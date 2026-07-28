"""Fail-closed D1 reconstruction for the UTR EditBench V2 datasets.

This module deliberately keeps four layers distinct:

* immutable input provenance;
* paper-clean, label-bearing normalized rows;
* canonical, label-bearing records used only by measured evaluation code; and
* a physically separate label-free candidate store.

Canonical edit scripts are deterministic constructions between measured
endpoints.  They are never represented as observed biological trajectories.
"""

from __future__ import annotations

import csv
import gzip
import hashlib
import json
import math
import os
import re
import statistics
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Iterator, Mapping, Sequence

from data.utr_benchmark_v2.edit_script import (
    apply_edit_script,
    canonicalize_edit_script,
)
from data.utr_benchmark_v2.records import (
    MEASURED_PAIR_TYPES,
    validate_canonical_record,
)


SCHEMA_VERSION = "utr_editbench_d1_v2"
RNA_ALPHABET = frozenset("ACGU")
PIPELINE_STAGES = (
    "download",
    "extract",
    "paper_clean",
    "canonical_clean",
    "build_source_candidate",
    "build_edit_scripts",
    "reproduce_labels",
    "audit_library_design",
    "audit_exposure",
)

CANONICAL_FIELDS = (
    "record_id",
    "dataset_id",
    "study_id",
    "assay_id",
    "context_id",
    "evidence_grade",
    "exposure_grade",
    "region",
    "organism",
    "cell_context",
    "reporter",
    "cargo",
    "endpoint",
    "endpoint_provenance",
    "timepoint",
    "source_id",
    "source_sequence",
    "candidate_sequence",
    "source_length",
    "candidate_length",
    "edit_script",
    "edit_types",
    "edit_positions",
    "reference_alleles",
    "alternate_alleles",
    "edit_count",
    "edit_distance",
    "source_value_raw",
    "candidate_value_raw",
    "delta_raw",
    "delta_normalized",
    "effect_standard_error",
    "replicate_count",
    "pair_type",
    "trajectory_observed",
    "trajectory_source",
    "trajectory_provenance",
    "coupling_type",
    "paper_split",
    "canonical_split",
    "source_group",
    "gene_group",
    "study_group",
    "context_group",
    "sequence_cluster",
    "sequence_provenance",
    "label_provenance",
    "download_manifest",
    "license",
    "quality_flags",
    "historical_exposure",
    "scaffold_group",
    "barcode_batch",
    "library_batch",
)

CANDIDATE_STORE_FIELDS = (
    "record_id",
    "candidate_id",
    "dataset_id",
    "study_id",
    "assay_id",
    "context_id",
    "evidence_grade",
    "exposure_grade",
    "region",
    "organism",
    "cell_context",
    "reporter",
    "cargo",
    "endpoint",
    "endpoint_provenance",
    "timepoint",
    "source_id",
    "source_sequence",
    "candidate_sequence",
    "source_length",
    "candidate_length",
    "edit_script",
    "edit_types",
    "edit_positions",
    "reference_alleles",
    "alternate_alleles",
    "edit_count",
    "edit_distance",
    "pair_type",
    "trajectory_observed",
    "trajectory_source",
    "trajectory_provenance",
    "coupling_type",
    "paper_split",
    "canonical_split",
    "source_group",
    "gene_group",
    "study_group",
    "context_group",
    "sequence_cluster",
    "sequence_provenance",
    "download_manifest",
    "license",
    "quality_flags",
    "historical_exposure",
    "intermediate_sequences",
    "library_design",
    "barcode_batch",
    "scaffold_group",
    "library_batch",
    "raw_source_sequence",
    "raw_candidate_sequence",
    "canonicalization_provenance",
)

CANDIDATE_STORE_FORBIDDEN_FIELDS = frozenset(
    {
        "source_value",
        "candidate_value",
        "source_value_raw",
        "candidate_value_raw",
        "delta",
        "delta_raw",
        "delta_normalized",
        "effect",
        "effect_size",
        "effect_standard_error",
        "label",
        "labels",
        "label_provenance",
        "reported_logfc",
        "score",
        "target",
    }
)
CANDIDATE_STORE_FORBIDDEN_KEY_TOKENS = frozenset(
    {
        "label",
        "labels",
        "source_value",
        "candidate_value",
        "value_raw",
        "delta",
        "effect",
        "effect_size",
        "score",
        "target",
        "outcome",
        "half_life",
        "halflife",
        "logfc",
        "mrl",
        "rl",
    }
)


ACTIVE_DATASET_POLICIES: dict[str, dict[str, Any]] = {
    "GSE114002": {
        "status": "eligible",
        "region": "five_utr",
        "evidence_grade": "E2",
        "exposure_grade": "previously_accessed",
        "historical_exposure": "previously_accessed_before_V2",
        "allowed_uses": [
            "source_paired_natural_SNV_grounding",
            "absolute_sequence_prior_separately_labelled",
        ],
        "forbidden_uses": [
            "absolute_sequence_as_intervention",
            "observed_trajectory_claim",
            "indel_efficacy_claim",
            "untouched_external_claim",
        ],
        "library_ascertainment": (
            "random_25_50nt_and_truncated_natural_5UTR_library; "
            "only source-anchored single-SNV rows are intervention records"
        ),
    },
    "GSE200304": {
        "status": "eligible",
        "region": "three_utr",
        "evidence_grade": "E2",
        "exposure_grade": "previously_accessed",
        "historical_exposure": "previously_accessed_before_V2",
        "allowed_uses": [
            "paired_3UTR_grounding",
            "multi_endpoint_control",
            "source_holdout",
        ],
        "forbidden_uses": [
            "indel_efficacy_claim",
            "observed_trajectory_claim",
            "raw_label_reproduction_claim",
            "untouched_external_claim",
        ],
        "library_ascertainment": (
            "patient mutation ascertainment; fixed 201-nt WT/mutant pairs"
        ),
    },
    "GSE246381": {
        "status": "eligible",
        "region": "five_utr",
        "evidence_grade": "E4",
        "exposure_grade": "E4",
        "historical_exposure": "historically_exposed_retrospective_E4",
        "allowed_uses": [
            "retrospective_stress_test",
            "diagnostics_without_selection",
        ],
        "forbidden_uses": [
            "training",
            "candidate_selection",
            "hyperparameter_selection",
            "untouched_external_claim",
            "sealed_external_claim",
            "primary_claim_alone",
        ],
        "library_ascertainment": "ascertained neurodevelopmental-disorder variants",
    },
    "GSE217518": {
        "status": "eligible",
        "region": "mixed_by_row",
        "evidence_grade": "E2",
        "exposure_grade": "previously_accessed",
        "historical_exposure": "previously_accessed_before_V2",
        "allowed_uses": [
            "paired_3UTR_half_life_grounding",
            "paired_5UTR_half_life_grounding",
            "measured_SUB_INS_DEL_endpoint_pairs",
            "source_holdout",
        ],
        "forbidden_uses": [
            "observed_trajectory_claim",
            "untouched_external_claim",
            "unfrozen_final_split_access",
        ],
        "library_ascertainment": (
            "disease-relevant UTR variants from official Figure4 tables; "
            "unique Ref/Mut endpoint pairs only"
        ),
    },
    "MPRAu_processed_ENCSR854RUF": {
        "status": "conditional",
        "region": "three_utr",
        "evidence_grade": "E2",
        "exposure_grade": "previously_accessed",
        "historical_exposure": "previously_accessed_before_V2",
        "allowed_uses": [
            "paired_3UTR_grounding_after_reference_freeze",
            "measured_5bp_deletion_subset_after_100_percent_roundtrip",
            "context_control_after_reference_freeze",
        ],
        "forbidden_uses": [
            "partial_reference_reconstruction",
            "insertion_efficacy_claim",
            "observed_trajectory_claim",
            "untouched_external_claim",
        ],
        "library_ascertainment": (
            "variant panel plus targeted non-overlapping 5-bp deletion tiling"
        ),
    },
}

BLOCKED_DATASET_POLICIES: dict[str, dict[str, Any]] = {
    "GSE145046": {
        "status": "blocked",
        "reason_code": "BLOCKED_SCAFFOLD_NOT_FROZEN",
        "read_final_labels": False,
        "role": "absolute_dense_landscape_only",
    },
    "GSE149487": {
        "status": "blocked",
        "reason_code": "BLOCKED_EXACT_PAIR_MAPPING_UNRECOVERED",
        "read_final_labels": False,
        "role": "blocked_canonical_pairs",
    },
    "ENCSR854RUF_raw62": {
        "status": "blocked",
        "reason_code": "OBSERVATIONAL_ONLY_NOT_INTERVENTION",
        "read_final_labels": False,
        "role": "observational_pretraining_candidate_only",
    },
    "GSE330741": {
        "status": "blocked",
        "reason_code": "METADATA_ONLY_FINAL_LABELS_UNOPENED",
        "read_final_labels": False,
        "role": "metadata_only_preaccess_freeze_candidate",
    },
    "GSE291719": {
        "status": "blocked",
        "reason_code": "METADATA_ONLY_FINAL_LABELS_UNOPENED",
        "read_final_labels": False,
        "role": "metadata_only_preaccess_freeze_candidate",
    },
    "GSE173083": {
        "status": "blocked",
        "reason_code": "EXCLUDED_OUTSIDE_UTR_ONLY_SCOPE",
        "read_final_labels": False,
        "role": "archived_context_only",
    },
    "GSE207584": {
        "status": "blocked",
        "reason_code": "EXCLUDED_OUTSIDE_UTR_ONLY_SCOPE",
        "read_final_labels": False,
        "role": "archived_context_only",
    },
}
D1_SCOPE_DATASETS = frozenset(
    set(ACTIVE_DATASET_POLICIES) | set(BLOCKED_DATASET_POLICIES)
)


def dataset_policy(dataset_id: str) -> dict[str, Any]:
    """Return a copy of the immutable D1 policy for ``dataset_id``."""
    if dataset_id in ACTIVE_DATASET_POLICIES:
        return dict(ACTIVE_DATASET_POLICIES[dataset_id])
    if dataset_id in BLOCKED_DATASET_POLICIES:
        return dict(BLOCKED_DATASET_POLICIES[dataset_id])
    return {
        "status": "blocked",
        "reason_code": "UNREGISTERED_DATASET_FAIL_CLOSED",
        "read_final_labels": False,
        "role": "unregistered",
    }


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _sha256_text(payload: str) -> str:
    return _sha256_bytes(payload.encode("utf-8"))


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(8 * 1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _normalize_name(value: Any) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def _normalized_row(row: Mapping[str, Any]) -> dict[str, Any]:
    return {_normalize_name(key): value for key, value in row.items()}


def _get(row: Mapping[str, Any], *names: str, default: Any = None) -> Any:
    normalized = _normalized_row(row)
    for name in names:
        key = _normalize_name(name)
        if key in normalized:
            value = normalized[key]
            if value is not None and str(value).strip() != "":
                return value
    return default


def _to_float(value: Any) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def _to_int(value: Any) -> int | None:
    number = _to_float(value)
    if number is None or not number.is_integer():
        return None
    return int(number)


def _normalize_sequence(value: Any) -> str:
    sequence = re.sub(r"\s+", "", str(value or "")).upper().replace("T", "U")
    if not sequence:
        raise ValueError("EMPTY_SEQUENCE")
    invalid = sorted(set(sequence) - RNA_ALPHABET)
    if invalid:
        raise ValueError("INVALID_SEQUENCE_ALPHABET:" + "".join(invalid))
    return sequence


def _median(values: Sequence[float]) -> float:
    if not values:
        raise ValueError("median requires values")
    return float(statistics.median(values))


def _row_identifier(row: Mapping[str, Any], row_index: int) -> str:
    value = _get(
        row,
        "row_id",
        "record_id",
        "seqid",
        "sequence_id",
        "variant_id",
        "id",
    )
    return str(value) if value is not None else f"row-{row_index:08d}"


def _row_fingerprint(row: Mapping[str, Any]) -> str:
    canonical = json.dumps(
        {str(key): str(value) for key, value in sorted(row.items())},
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    return _sha256_text(canonical)


def _reject(
    dataset_id: str,
    row: Mapping[str, Any],
    row_index: int,
    reason_code: str,
    detail: str = "",
) -> dict[str, Any]:
    return {
        "dataset_id": dataset_id,
        "row_id": _row_identifier(row, row_index),
        "row_fingerprint_sha256": _row_fingerprint(row),
        "reason_code": reason_code,
        "detail": detail,
    }


def _levenshtein_distance(source: str, candidate: str) -> int:
    previous = list(range(len(candidate) + 1))
    for source_index, source_base in enumerate(source, start=1):
        current = [source_index]
        for candidate_index, candidate_base in enumerate(candidate, start=1):
            current.append(
                min(
                    current[-1] + 1,
                    previous[candidate_index] + 1,
                    previous[candidate_index - 1] + (source_base != candidate_base),
                )
            )
        previous = current
    return previous[-1]


def _canonical_script(source: str, candidate: str) -> dict[str, Any]:
    result = canonicalize_edit_script(source, candidate)
    if not isinstance(result, Mapping):
        raise ValueError("EDIT_SCRIPT_API_MUST_RETURN_MAPPING")
    actions_raw = (
        result.get("actions") or result.get("edit_script") or result.get("script")
    )
    if not isinstance(actions_raw, Sequence) or isinstance(actions_raw, (str, bytes)):
        raise ValueError("EDIT_SCRIPT_RESULT_LACKS_ACTIONS")
    actions = [dict(action) for action in actions_raw]
    applied = apply_edit_script(source, actions)
    if applied != candidate:
        raise ValueError("EDIT_SCRIPT_ROUNDTRIP_FAILED")

    edit_actions = [
        action
        for action in actions
        if str(action.get("op") or action.get("action") or "").upper() != "STOP"
    ]
    edit_types = [
        str(action.get("op") or action.get("action") or "").upper()
        for action in edit_actions
    ]
    if any(action not in {"SUB", "INS", "DEL"} for action in edit_types):
        raise ValueError("UNSUPPORTED_EDIT_ACTION")

    positions: list[int] = []
    reference_alleles: list[str] = []
    alternate_alleles: list[str] = []
    for action in edit_actions:
        position = action.get("position", action.get("pos", action.get("index")))
        if isinstance(position, bool) or not isinstance(position, int):
            raise ValueError("EDIT_ACTION_LACKS_INTEGER_POSITION")
        positions.append(position)
        reference_alleles.append(
            str(
                action.get(
                    "reference",
                    action.get("ref", action.get("deleted", "")),
                )
                or ""
            ).upper()
        )
        alternate_alleles.append(
            str(
                action.get(
                    "alternate",
                    action.get("alt", action.get("inserted", "")),
                )
                or ""
            ).upper()
        )

    ambiguity_count = result.get(
        "equivalent_minimal_script_count",
        result.get("ambiguity_count", result.get("minimal_alignment_count", 1)),
    )
    try:
        ambiguity_count = int(ambiguity_count)
    except (TypeError, ValueError):
        raise ValueError("INVALID_EDIT_SCRIPT_AMBIGUITY_COUNT") from None
    if ambiguity_count < 1:
        raise ValueError("INVALID_EDIT_SCRIPT_AMBIGUITY_COUNT")

    minimal_edit_count = result.get(
        "minimal_edit_count",
        result.get("edit_distance", len(edit_actions)),
    )
    try:
        minimal_edit_count = int(minimal_edit_count)
    except (TypeError, ValueError):
        raise ValueError("INVALID_MINIMAL_EDIT_COUNT") from None

    return {
        "actions": actions,
        "edit_types": edit_types,
        "edit_positions": positions,
        "reference_alleles": reference_alleles,
        "alternate_alleles": alternate_alleles,
        "edit_count": len(edit_actions),
        "edit_distance": _levenshtein_distance(source, candidate),
        "minimal_edit_count": minimal_edit_count,
        "equivalent_minimal_script_count": ambiguity_count,
        "path_ambiguous": ambiguity_count > 1,
        "canonicalization": str(
            result.get("canonicalization")
            or result.get("tie_break")
            or "deterministic_minimal_script"
        ),
        "ambiguity_category": str(result.get("ambiguity_category") or "unspecified"),
        "count_scope": str(
            result.get("count_scope") or "minimum_cost_character_alignments"
        ),
    }


def _normalise_provenance(provenance: Mapping[str, Any]) -> dict[str, Any]:
    result = dict(provenance)
    if "input_files" in result:
        files: list[dict[str, Any]] = []
        for entry in result["input_files"]:
            path = Path(str(entry["path"]))
            item = dict(entry)
            if path.is_file():
                item.setdefault("bytes", path.stat().st_size)
                item.setdefault("sha256", _sha256_file(path))
            files.append(item)
        result["input_files"] = files
    elif result.get("input_path"):
        path = Path(str(result["input_path"]))
        if path.is_file():
            result.setdefault("input_bytes", path.stat().st_size)
            result.setdefault("input_sha256", _sha256_file(path))
    result.setdefault("download_manifest", "UNKNOWN")
    result.setdefault("license", "UNKNOWN")
    if "raw_artifact" not in result:
        if result.get("input_files"):
            result["raw_artifact"] = [
                {
                    key: item.get(key)
                    for key in ("path", "bytes", "sha256", "role")
                    if item.get(key) is not None
                }
                for item in result["input_files"]
            ]
        elif result.get("input_path"):
            result["raw_artifact"] = {
                key: result.get(key)
                for key in ("input_path", "input_bytes", "input_sha256")
                if result.get(key) is not None
            }
        else:
            result["raw_artifact"] = "not_opened_or_not_applicable"
    result.setdefault(
        "processed_artifact",
        "generated_by:data.utr_benchmark_v2.d1_builder:paper_clean",
    )
    raw_files: list[dict[str, Any]] = []
    if result.get("input_files"):
        raw_files = [
            dict(item) for item in result["input_files"] if isinstance(item, Mapping)
        ]
    elif result.get("input_path"):
        raw_files = [
            {
                "path": result.get("input_path"),
                "bytes": result.get("input_bytes"),
                "sha256": result.get("input_sha256"),
            }
        ]
    raw_files_complete = bool(raw_files) and all(
        Path(str(item.get("path", ""))).is_file()
        and isinstance(item.get("bytes"), int)
        and item["bytes"] >= 0
        and bool(re.fullmatch(r"[0-9a-f]{64}", str(item.get("sha256", ""))))
        for item in raw_files
    )
    download_manifest = result.get("download_manifest")
    if isinstance(download_manifest, Mapping):
        download_manifest_complete = bool(
            re.fullmatch(
                r"[0-9a-f]{64}",
                str(download_manifest.get("sha256", "")),
            )
        ) and bool(download_manifest.get("path"))
    else:
        manifest_path = Path(str(download_manifest or ""))
        download_manifest_complete = manifest_path.is_file()
        if download_manifest_complete:
            result["download_manifest"] = {
                "path": str(manifest_path),
                "bytes": manifest_path.stat().st_size,
                "sha256": _sha256_file(manifest_path),
            }
    license_text = str(result.get("license") or "").strip()
    license_complete = bool(license_text) and _normalize_name(license_text) not in {
        "unknown",
        "unresolved",
        "tbd",
        "missing",
    }
    result["provenance_audit"] = {
        "raw_files": raw_files,
        "raw_files_complete": raw_files_complete,
        "download_manifest_complete": download_manifest_complete,
        "license_complete": license_complete,
        "complete": (
            raw_files_complete and download_manifest_complete and license_complete
        ),
    }
    return result


def _stable_id(*parts: Any) -> str:
    payload = "\x1f".join(str(part) for part in parts)
    return _sha256_text(payload)[:24]


def _pair_record(
    *,
    dataset_id: str,
    row_id: str,
    source: str,
    candidate: str,
    endpoint: str,
    provenance: Mapping[str, Any],
    source_value: float | None,
    candidate_value: float | None,
    delta: float | None,
    delta_normalized: float | None,
    effect_standard_error: float | None,
    replicate_count: int | None,
    label_status: str,
    cell_context: str,
    context_id: str,
    assay_id: str,
    organism: str,
    reporter: str,
    cargo: str,
    timepoint: str | None,
    gene: str | None,
    source_group: str | None,
    sequence_cluster: str | None,
    library_design: str,
    barcode_batch: str | None,
    quality_flags: Sequence[str] = (),
    pair_type: str | None = None,
    region: str | None = None,
    raw_endpoint: str | None = None,
    endpoint_transformation: str = "identity",
    sequence_provenance: Mapping[str, Any] | None = None,
    raw_source_sequence: str | None = None,
    raw_candidate_sequence: str | None = None,
    canonicalization_provenance: Mapping[str, Any] | None = None,
    scaffold_group: str | None = None,
    library_batch: str | None = None,
) -> dict[str, Any]:
    policy = ACTIVE_DATASET_POLICIES[dataset_id]
    script = _canonical_script(source, candidate)
    source_hash = _sha256_text(source)
    candidate_hash = _sha256_text(candidate)
    source_id = f"{dataset_id}:source:{source_hash[:20]}"
    record_id = (
        f"{dataset_id}:record:"
        f"{_stable_id(row_id, source_hash, candidate_hash, endpoint, context_id)}"
    )
    if delta is None and source_value is not None and candidate_value is not None:
        delta = candidate_value - source_value
    if delta_normalized is None:
        delta_normalized = delta
    group = source_group or source_id
    actual_pair_type = pair_type or (
        "true_wt_mutant"
        if source_value is not None and candidate_value is not None
        else "retrospective_constructed_neighbor"
    )
    sequence_provenance_payload = dict(sequence_provenance or provenance)
    sequence_provenance_payload.setdefault(
        "raw_artifact", provenance.get("raw_artifact", "UNKNOWN")
    )
    sequence_provenance_payload.setdefault(
        "processed_artifact",
        provenance.get(
            "processed_artifact",
            "generated_by:data.utr_benchmark_v2.d1_builder:paper_clean",
        ),
    )
    label_provenance = {
        "status": label_status,
        "input_provenance": dict(provenance),
    }
    record: dict[str, Any] = {
        "record_id": record_id,
        "candidate_id": f"{dataset_id}:candidate:{candidate_hash[:20]}",
        "dataset_id": dataset_id,
        "study_id": dataset_id,
        "assay_id": assay_id,
        "context_id": context_id,
        "evidence_grade": policy["evidence_grade"],
        "exposure_grade": policy["exposure_grade"],
        "region": policy["region"],
        "organism": organism,
        "cell_context": cell_context,
        "reporter": reporter,
        "cargo": cargo,
        "endpoint": endpoint,
        "endpoint_provenance": {
            "raw_endpoint": raw_endpoint or endpoint,
            "transformation": endpoint_transformation,
        },
        "timepoint": timepoint,
        "source_id": source_id,
        "source_sequence": source,
        "candidate_sequence": candidate,
        "source_length": len(source),
        "candidate_length": len(candidate),
        "edit_script": script["actions"],
        "edit_types": script["edit_types"],
        "edit_positions": script["edit_positions"],
        "reference_alleles": script["reference_alleles"],
        "alternate_alleles": script["alternate_alleles"],
        "edit_count": script["edit_count"],
        "edit_distance": script["edit_distance"],
        "source_value_raw": source_value,
        "candidate_value_raw": candidate_value,
        "delta_raw": delta,
        "delta_normalized": delta_normalized,
        "effect_standard_error": effect_standard_error,
        "replicate_count": replicate_count,
        "pair_type": actual_pair_type,
        "trajectory_observed": False,
        "trajectory_source": "constructed",
        "trajectory_provenance": {
            "status": "canonical_minimum_edit_alignment_between_endpoints",
            "observed_biological_path": False,
        },
        "coupling_type": "constructed_alignment_coupling",
        "paper_split": (
            "retrospective_only" if dataset_id == "GSE246381" else "unassigned_D1"
        ),
        "canonical_split": (
            "retrospective_only" if dataset_id == "GSE246381" else "unassigned_B0"
        ),
        "source_group": group,
        "gene_group": gene or f"NOT_APPLICABLE:{dataset_id}:gene_mapping_unavailable",
        "study_group": dataset_id,
        "context_group": context_id,
        "sequence_cluster": sequence_cluster or group,
        "sequence_provenance": sequence_provenance_payload,
        "label_provenance": label_provenance,
        "download_manifest": provenance.get("download_manifest", "UNKNOWN"),
        "license": provenance.get("license", "UNKNOWN"),
        "quality_flags": sorted(set(quality_flags)),
        "historical_exposure": policy["historical_exposure"],
        "intermediate_sequences": [],
        "library_design": library_design,
        "barcode_batch": barcode_batch
        or f"NOT_APPLICABLE:{dataset_id}:barcode_batch_unavailable",
        "library_batch": library_batch
        or f"NOT_APPLICABLE:{dataset_id}:library_batch_unavailable",
        "scaffold_group": scaffold_group or f"{dataset_id}:reporter:{reporter}",
        "raw_source_sequence": raw_source_sequence,
        "raw_candidate_sequence": raw_candidate_sequence,
        "canonicalization_provenance": dict(canonicalization_provenance or {}),
        "edit_script_ambiguity": {
            "equivalent_minimal_script_count": script[
                "equivalent_minimal_script_count"
            ],
            "path_ambiguous": script["path_ambiguous"],
            "canonicalization": script["canonicalization"],
            "ambiguity_category": script["ambiguity_category"],
            "count_scope": script["count_scope"],
            "trajectory_observed": False,
        },
    }
    missing = [field for field in CANONICAL_FIELDS if field not in record]
    if missing:
        raise AssertionError(f"canonical record missing fields: {missing}")
    record["region"] = region or policy["region"]
    return validate_canonical_record(record)


def _absolute_record(
    *,
    dataset_id: str,
    row_id: str,
    candidate: str,
    endpoint: str,
    provenance: Mapping[str, Any],
    candidate_value: float,
    cell_context: str,
    assay_id: str,
    library_design: str,
) -> dict[str, Any]:
    policy = ACTIVE_DATASET_POLICIES[dataset_id]
    candidate_hash = _sha256_text(candidate)
    context_id = f"{dataset_id}:{cell_context}"
    record: dict[str, Any] = {
        "record_id": (
            f"{dataset_id}:absolute:"
            f"{_stable_id(row_id, candidate_hash, endpoint, context_id)}"
        ),
        "candidate_id": f"{dataset_id}:candidate:{candidate_hash[:20]}",
        "dataset_id": dataset_id,
        "study_id": dataset_id,
        "assay_id": assay_id,
        "context_id": context_id,
        "evidence_grade": policy["evidence_grade"],
        "exposure_grade": policy["exposure_grade"],
        "region": policy["region"],
        "organism": "human",
        "cell_context": cell_context,
        "reporter": "MPRA",
        "cargo": "eGFP_or_reporter",
        "endpoint": endpoint,
        "endpoint_provenance": {
            "raw_endpoint": endpoint,
            "transformation": "identity",
        },
        "timepoint": None,
        "source_id": None,
        "source_sequence": None,
        "candidate_sequence": candidate,
        "source_length": None,
        "candidate_length": len(candidate),
        "edit_script": None,
        "edit_types": [],
        "edit_positions": [],
        "reference_alleles": [],
        "alternate_alleles": [],
        "edit_count": 0,
        "edit_distance": None,
        "source_value_raw": None,
        "candidate_value_raw": candidate_value,
        "delta_raw": None,
        "delta_normalized": None,
        "effect_standard_error": None,
        "replicate_count": None,
        "pair_type": "absolute_property_only",
        "trajectory_observed": False,
        "trajectory_source": "latent",
        "trajectory_provenance": {
            "status": "not_applicable_absolute_property_record",
            "observed_biological_path": False,
        },
        "coupling_type": "property_conditioned_target_coupling",
        "paper_split": "absolute_prior_only",
        "canonical_split": "absolute_prior_only",
        "source_group": f"{dataset_id}:absolute_library",
        "gene_group": "NOT_APPLICABLE:GSE114002:absolute_library",
        "study_group": dataset_id,
        "context_group": context_id,
        "sequence_cluster": f"{dataset_id}:absolute:{candidate_hash[:20]}",
        "sequence_provenance": {
            **dict(provenance),
            "raw_artifact": provenance.get("raw_artifact", "UNKNOWN"),
            "processed_artifact": provenance.get(
                "processed_artifact",
                "generated_by:data.utr_benchmark_v2.d1_builder:paper_clean",
            ),
        },
        "label_provenance": {
            "status": "processed_or_raw_absolute_measurement",
            "input_provenance": dict(provenance),
        },
        "download_manifest": provenance.get("download_manifest", "UNKNOWN"),
        "license": provenance.get("license", "UNKNOWN"),
        "quality_flags": ["ABSOLUTE_SEQUENCE_NOT_INTERVENTION"],
        "historical_exposure": policy["historical_exposure"],
        "intermediate_sequences": [],
        "library_design": library_design,
        "barcode_batch": "NOT_APPLICABLE:GSE114002:absolute_library",
        "library_batch": f"GSE114002:{library_design}",
        "scaffold_group": f"GSE114002:absolute_library:{library_design}",
        "raw_source_sequence": None,
        "raw_candidate_sequence": candidate,
        "canonicalization_provenance": {
            "sequence_normalization": "uppercase;T_to_U",
            "source_constructed": False,
        },
        "edit_script_ambiguity": {
            "equivalent_minimal_script_count": 0,
            "path_ambiguous": False,
            "canonicalization": "not_applicable_absolute_sequence",
            "trajectory_observed": False,
        },
    }
    return validate_canonical_record(record)


def _candidate_record(record: Mapping[str, Any]) -> dict[str, Any]:
    candidate = {
        field: record[field] for field in CANDIDATE_STORE_FIELDS if field in record
    }
    forbidden = set(candidate) & CANDIDATE_STORE_FORBIDDEN_FIELDS
    if forbidden:
        raise AssertionError(
            "candidate store contains label-bearing fields: "
            + ",".join(sorted(forbidden))
        )
    nested_forbidden = candidate_store_label_paths(candidate)
    if nested_forbidden:
        raise AssertionError(
            "candidate store contains nested label-bearing keys: "
            + ",".join(nested_forbidden)
        )
    return candidate


def candidate_store_label_paths(
    payload: Any,
    *,
    prefix: str = "$",
) -> list[str]:
    """Return recursively nested paths whose keys can carry measurements.

    Candidate-store provenance remains audit-rich, but arbitrary config
    dictionaries cannot smuggle labels through nested provenance fields.
    """
    failures: list[str] = []
    if isinstance(payload, Mapping):
        for raw_key, value in payload.items():
            key = _normalize_name(raw_key)
            path = f"{prefix}.{raw_key}"
            if (
                key in CANDIDATE_STORE_FORBIDDEN_FIELDS
                or key in CANDIDATE_STORE_FORBIDDEN_KEY_TOKENS
                or key.endswith("_label")
                or key.startswith("label_")
                or key.endswith("_value")
                or key.startswith("delta_")
                or key.startswith("effect_")
            ):
                failures.append(path)
            failures.extend(candidate_store_label_paths(value, prefix=path))
    elif isinstance(payload, (list, tuple)):
        for index, value in enumerate(payload):
            failures.extend(
                candidate_store_label_paths(value, prefix=f"{prefix}[{index}]")
            )
    return sorted(set(failures))


def _paper_clean_pair(record: Mapping[str, Any], row_id: str) -> dict[str, Any]:
    paper = {
        "row_id": row_id,
        "record_id": record["record_id"],
        "dataset_id": record["dataset_id"],
        "source_sequence": record["source_sequence"],
        "candidate_sequence": record["candidate_sequence"],
        "endpoint": record["endpoint"],
        "source_value_raw": record["source_value_raw"],
        "candidate_value_raw": record["candidate_value_raw"],
        "delta_raw": record["delta_raw"],
        "label_reproduction_status": record["label_provenance"]["status"],
        "pair_type": record["pair_type"],
    }
    for field in (
        "raw_source_sequence",
        "raw_candidate_sequence",
        "canonicalization_provenance",
    ):
        if record.get(field) not in (None, {}):
            paper[field] = record[field]
    return paper


def _base_result(
    dataset_id: str,
    provenance: Mapping[str, Any],
    fixture_mode: bool,
) -> dict[str, Any]:
    policy = dataset_policy(dataset_id)
    return {
        "schema_version": SCHEMA_VERSION,
        "generated_at_utc": _utc_now(),
        "dataset_id": dataset_id,
        "fixture_mode": fixture_mode,
        "policy": policy,
        "status": "pending",
        "paper_eligible": False,
        "allowed_uses": list(policy.get("allowed_uses", [])),
        "forbidden_uses": list(policy.get("forbidden_uses", [])),
        "input_provenance": dict(provenance),
        "paper_clean_records": [],
        "label_records": [],
        "candidate_records": [],
        "rejected_records": [],
        "auxiliary_records": [],
        "raw_row_lineage_records": [],
        "accounting": {
            "total_input_rows": 0,
            "accepted_intervention_rows": 0,
            "accepted_absolute_rows": 0,
            "accepted_input_rows": 0,
            "auxiliary_source_anchor_rows": 0,
            "rejected_rows": 0,
            "accounted_rows": 0,
        },
        "library_ascertainment": policy.get("library_ascertainment", "UNKNOWN"),
        "label_reproduction": {"status": "not_started"},
    }


def _finalize_result(result: dict[str, Any]) -> dict[str, Any]:
    accounting = result["accounting"]
    accounting["accepted_intervention_rows"] = sum(
        record["pair_type"] not in {"absolute_property_only", "unlabeled_pretraining"}
        for record in result["label_records"]
    )
    accounting["accepted_absolute_rows"] = sum(
        record["pair_type"] in {"absolute_property_only", "unlabeled_pretraining"}
        for record in result["label_records"]
    )
    if not accounting.get("accepted_input_rows"):
        accounting["accepted_input_rows"] = len(result["label_records"])
    accounting["auxiliary_source_anchor_rows"] = len(result["auxiliary_records"])
    accounting["rejected_rows"] = len(result["rejected_records"])
    accounting["accounted_rows"] = (
        accounting["accepted_input_rows"]
        + accounting["auxiliary_source_anchor_rows"]
        + accounting["rejected_rows"]
    )
    if accounting["accounted_rows"] != accounting["total_input_rows"]:
        raise AssertionError(
            "D1 accounting mismatch: "
            f"{accounting['accounted_rows']} != {accounting['total_input_rows']}"
        )
    result["candidate_records"] = [
        _candidate_record(record) for record in result["label_records"]
    ]
    roundtrip_records = [
        record
        for record in result["label_records"]
        if record["pair_type"]
        not in {"absolute_property_only", "unlabeled_pretraining"}
    ]
    roundtrip_passed = sum(
        apply_edit_script(record["source_sequence"], record["edit_script"])
        == record["candidate_sequence"]
        for record in roundtrip_records
    )
    result["roundtrip_audit"] = {
        "intervention_records": len(roundtrip_records),
        "roundtrip_passed": roundtrip_passed,
        "fraction": (
            roundtrip_passed / len(roundtrip_records) if roundtrip_records else None
        ),
    }
    ambiguity_counts = [
        record["edit_script_ambiguity"]["equivalent_minimal_script_count"]
        for record in roundtrip_records
    ]
    result["edit_script_ambiguity"] = {
        "records": len(ambiguity_counts),
        "ambiguous_records": sum(count > 1 for count in ambiguity_counts),
        "max_equivalent_minimal_script_count": max(ambiguity_counts, default=0),
        "constructed_paths_marked_observed": sum(
            bool(record["trajectory_observed"]) for record in roundtrip_records
        ),
    }
    measured_records = [
        record
        for record in roundtrip_records
        if record["pair_type"] in MEASURED_PAIR_TYPES
    ]
    action_counts = {"SUB": 0, "INS": 0, "DEL": 0}
    action_instance_counts = {"SUB": 0, "INS": 0, "DEL": 0}
    edit_count_distribution: dict[str, int] = {}
    action_combination_distribution: dict[str, int] = {}
    pair_type_counts: dict[str, int] = {}
    for record in measured_records:
        for action in set(record["edit_types"]):
            action_counts[action] += 1
        for action in record["edit_types"]:
            action_instance_counts[action] += 1
        edit_count_key = str(record["edit_count"])
        edit_count_distribution[edit_count_key] = (
            edit_count_distribution.get(edit_count_key, 0) + 1
        )
        action_key = "+".join(sorted(set(record["edit_types"])))
        action_combination_distribution[action_key] = (
            action_combination_distribution.get(action_key, 0) + 1
        )
        pair_type = str(record["pair_type"])
        pair_type_counts[pair_type] = pair_type_counts.get(pair_type, 0) + 1
    indel_records = [
        record
        for record in measured_records
        if {"INS", "DEL"} & set(record["edit_types"])
    ]
    missing_or_nonfinite_label_rejections = sum(
        record["reason_code"]
        in {
            "MISSING_OR_NONFINITE_LABEL",
            "MISSING_PAIRED_PROCESSED_LABEL",
        }
        for record in result["rejected_records"]
    )
    result["measured_action_coverage"] = {
        "measured_endpoint_pair_records": len(measured_records),
        "record_counts_by_action_presence": action_counts,
        "canonical_action_instance_counts": action_instance_counts,
        "edit_count_distribution": dict(sorted(edit_count_distribution.items())),
        "action_combination_distribution": dict(
            sorted(action_combination_distribution.items())
        ),
        "pair_type_counts": dict(sorted(pair_type_counts.items())),
        "indel_endpoint_pairs": len(indel_records),
        "indel_endpoint_pairs_with_both_finite_values": sum(
            record["source_value_raw"] is not None
            and math.isfinite(record["source_value_raw"])
            and record["candidate_value_raw"] is not None
            and math.isfinite(record["candidate_value_raw"])
            for record in indel_records
        ),
        "missing_or_nonfinite_label_rejections": (
            missing_or_nonfinite_label_rejections
        ),
        "observed_trajectory_records": 0,
        "interpretation": (
            "action presence is derived from measured source/candidate "
            "endpoints; canonical action order is constructed and was not "
            "observed experimentally"
        ),
    }
    if result["status"] == "pending":
        result["status"] = "accepted_fixture" if result["fixture_mode"] else "accepted"
    result["paper_eligible"] = (
        result["status"] == "accepted"
        and bool(result["label_records"])
        and bool(result["input_provenance"].get("provenance_audit", {}).get("complete"))
        and result["roundtrip_audit"]["roundtrip_passed"]
        == result["roundtrip_audit"]["intervention_records"]
    )
    return result


def _build_gse114002(
    rows: Sequence[Mapping[str, Any]],
    provenance: Mapping[str, Any],
    fixture_mode: bool,
) -> dict[str, Any]:
    result = _base_result("GSE114002", provenance, fixture_mode)
    result["accounting"]["total_input_rows"] = len(rows)
    prepared: list[dict[str, Any]] = []
    anchors: dict[str, list[float]] = {}

    for index, row in enumerate(rows):
        row_id = _row_identifier(row, index)
        library = str(
            _get(row, "library", "library_kind", "library_type", default="")
        ).strip()
        raw_candidate = _get(
            row, "utr", "sequence", "candidate_sequence", "alt_sequence"
        )
        raw_source = _get(
            row, "mother", "source_sequence", "wt_sequence", "ref_sequence"
        )
        value = _to_float(
            _get(row, "rl", "mrl", "mean_ribosome_loading", "candidate_value")
        )
        try:
            candidate = _normalize_sequence(raw_candidate)
            source = _normalize_sequence(raw_source) if raw_source is not None else None
        except ValueError as exc:
            result["rejected_records"].append(
                _reject("GSE114002", row, index, str(exc).split(":")[0], str(exc))
            )
            continue
        if value is None:
            result["rejected_records"].append(
                _reject("GSE114002", row, index, "MISSING_OR_NONFINITE_LABEL")
            )
            continue
        item = {
            "index": index,
            "row": row,
            "row_id": row_id,
            "library": library,
            "source": source,
            "candidate": candidate,
            "value": value,
        }
        if source is not None and candidate == source:
            anchors.setdefault(source, []).append(value)
            result["auxiliary_records"].append(
                {
                    "row_id": row_id,
                    "role": "measured_source_anchor",
                    "source_id": f"GSE114002:source:{_sha256_text(source)[:20]}",
                    "source_sequence": source,
                    "value_raw": value,
                    "endpoint": "mean_ribosome_loading",
                }
            )
        else:
            prepared.append(item)

    for item in prepared:
        row = item["row"]
        index = item["index"]
        source = item["source"]
        candidate = item["candidate"]
        row_id = item["row_id"]
        library = item["library"]
        value = item["value"]
        if source is None:
            if not re.search(
                r"random|vary|trunc|absolute|designed", library, re.IGNORECASE
            ):
                result["rejected_records"].append(
                    _reject(
                        "GSE114002",
                        row,
                        index,
                        "UNCLASSIFIED_ABSOLUTE_LIBRARY",
                        f"library={library!r}",
                    )
                )
                continue
            record = _absolute_record(
                dataset_id="GSE114002",
                row_id=row_id,
                candidate=candidate,
                endpoint="mean_ribosome_loading",
                provenance=provenance,
                candidate_value=value,
                cell_context=str(
                    _get(row, "cell_context", "cell_type", default="HEK293T")
                ),
                assay_id=str(
                    _get(
                        row,
                        "assay_id",
                        "assay",
                        default="GSE114002_MPRA_mean_ribosome_loading",
                    )
                ),
                library_design=library or "absolute_library_unclassified",
            )
            result["label_records"].append(record)
            result["paper_clean_records"].append(_paper_clean_pair(record, row_id))
            continue

        if source not in anchors:
            result["rejected_records"].append(
                _reject(
                    "GSE114002",
                    row,
                    index,
                    "MISSING_SOURCE_ANCHOR",
                    "no source==candidate measured row for this mother sequence",
                )
            )
            continue
        try:
            record = _pair_record(
                dataset_id="GSE114002",
                row_id=row_id,
                source=source,
                candidate=candidate,
                endpoint="mean_ribosome_loading",
                provenance=provenance,
                source_value=_median(anchors[source]),
                candidate_value=value,
                delta=None,
                delta_normalized=None,
                effect_standard_error=_to_float(
                    _get(row, "effect_standard_error", "standard_error", "se")
                ),
                replicate_count=_to_int(_get(row, "replicate_count", "n_replicates")),
                label_status="PROVIDED_LABEL_ONLY",
                cell_context=str(
                    _get(row, "cell_context", "cell_type", default="HEK293T")
                ),
                context_id=str(
                    _get(
                        row,
                        "context_id",
                        default="GSE114002:HEK293T:unmodified_mRNA",
                    )
                ),
                assay_id=str(
                    _get(
                        row,
                        "assay_id",
                        "assay",
                        default="GSE114002_MPRA_mean_ribosome_loading",
                    )
                ),
                organism="human",
                reporter="MPRA",
                cargo=str(_get(row, "cargo", "cds", default="eGFP")),
                timepoint=None,
                gene=(str(_get(row, "gene", "symbol", default="")).strip() or None),
                source_group=f"GSE114002:mother:{_sha256_text(source)[:20]}",
                sequence_cluster=None,
                library_design=library or "source_anchored_snv",
                barcode_batch=str(_get(row, "barcode_batch", "batch", default=""))
                or None,
            )
        except ValueError as exc:
            result["rejected_records"].append(
                _reject(
                    "GSE114002",
                    row,
                    index,
                    str(exc).split(":")[0],
                    str(exc),
                )
            )
            continue
        if record["edit_count"] != 1 or record["edit_types"] != ["SUB"]:
            result["rejected_records"].append(
                _reject(
                    "GSE114002",
                    row,
                    index,
                    "UNSUPPORTED_NON_SNV_INTERVENTION",
                    f"edit_types={record['edit_types']}; edit_count={record['edit_count']}",
                )
            )
            continue
        result["label_records"].append(record)
        result["paper_clean_records"].append(_paper_clean_pair(record, row_id))

    result["label_reproduction"] = {
        "status": "PROVIDED_LABEL_ONLY",
        "source_anchor_aggregation": "median",
        "raw_reproduction_claim_allowed": False,
        "note": (
            "Deposited rl values are consumed as provided labels; raw-bin "
            "reconstruction was not performed by this D1 builder."
        ),
    }
    return _finalize_result(result)


def _pair_sequences(
    row: Mapping[str, Any],
) -> tuple[str, str]:
    source = _normalize_sequence(
        _get(
            row,
            "source_sequence",
            "wt_sequence",
            "wildtype_sequence",
            "reference_sequence",
            "ref_sequence",
            "refsequence",
            "wt",
            "reference",
        )
    )
    candidate = _normalize_sequence(
        _get(
            row,
            "candidate_sequence",
            "mutant_sequence",
            "variant_sequence",
            "alternate_sequence",
            "alt_sequence",
            "altsequence",
            "mutant",
            "alt",
        )
    )
    return source, candidate


GSE200304_PAIR_PATTERN = re.compile(r"_(WT|Mutant)$")


def extract_gse200304_exact_join(
    construct_rows: Sequence[Mapping[str, Any]],
    label_rows: Sequence[Mapping[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Join the official GSE200304 tables on merged_id == Barcode exactly."""
    labels: dict[str, float] = {}
    duplicate_barcodes: set[str] = set()
    invalid_label_rows = 0
    label_row_audit: list[dict[str, Any]] = []
    for row_index, row in enumerate(label_rows):
        barcode = str(_get(row, "Barcode", default="")).strip()
        value = _to_float(_get(row, "Freq"))
        fingerprint = _row_fingerprint(row)
        lineage = {
            "schema_version": "gse200304_raw_row_lineage_v1",
            "dataset_id": "GSE200304",
            "raw_table_role": "processed_label_table",
            "raw_row_index": row_index,
            "raw_row_key": f"processed_label_table:{row_index}",
            "raw_row_id": barcode or None,
            "raw_row_fingerprint_sha256": fingerprint,
            "lineage_id": _sha256_text(
                f"GSE200304\x1fprocessed_label_table\x1f{row_index}\x1f{fingerprint}"
            ),
            "disposition": "PENDING",
            "normalized_target": None,
        }
        if not barcode or value is None:
            invalid_label_rows += 1
            lineage["disposition"] = "REJECTED_INVALID_LABEL_ROW"
            label_row_audit.append(lineage)
            continue
        if barcode in labels:
            duplicate_barcodes.add(barcode)
        labels[barcode] = value
        label_row_audit.append(lineage)
    if duplicate_barcodes:
        raise ValueError(
            "GSE200304_DUPLICATE_LABEL_BARCODES:"
            + ",".join(sorted(duplicate_barcodes)[:5])
        )

    constructs_by_id: dict[str, Mapping[str, Any]] = {}
    construct_row_indices: dict[str, int] = {}
    construct_row_fingerprints: dict[str, str] = {}
    duplicate_construct_ids: set[str] = set()
    for row_index, row in enumerate(construct_rows):
        merged_id = str(_get(row, "merged_id", default="")).strip()
        if not merged_id:
            raise ValueError("GSE200304_CONSTRUCT_MISSING_MERGED_ID")
        if merged_id in constructs_by_id:
            duplicate_construct_ids.add(merged_id)
        constructs_by_id[merged_id] = row
        construct_row_indices[merged_id] = row_index
        construct_row_fingerprints[merged_id] = _row_fingerprint(row)
    if duplicate_construct_ids:
        raise ValueError(
            "GSE200304_DUPLICATE_CONSTRUCT_MERGED_IDS:"
            + ",".join(sorted(duplicate_construct_ids)[:5])
        )

    grouped: dict[str, dict[str, list[tuple[str, Mapping[str, Any]]]]] = {}
    controls: list[tuple[str, Mapping[str, Any]]] = []
    for merged_id, row in constructs_by_id.items():
        match = GSE200304_PAIR_PATTERN.search(merged_id)
        if match is None:
            controls.append((merged_id, row))
            continue
        allele = match.group(1)
        base = merged_id[: match.start()]
        grouped.setdefault(base, {"WT": [], "Mutant": []})[allele].append(
            (merged_id, row)
        )

    normalized: list[dict[str, Any]] = []
    normalized_target_by_construct: dict[str, dict[str, Any]] = {}
    construct_disposition: dict[str, str] = {}
    malformed_pair_constructs = 0
    pair_201nt_count = 0
    pair_hamming_distribution: dict[str, int] = {}
    for base in sorted(grouped):
        endpoints = grouped[base]
        if len(endpoints["WT"]) != 1 or len(endpoints["Mutant"]) != 1:
            for allele in ("WT", "Mutant"):
                for merged_id, row in endpoints[allele]:
                    normalized_target_by_construct[merged_id] = {
                        "target_kind": "rejected_row",
                        "row_id": merged_id,
                        "endpoint_role": allele,
                    }
                    construct_disposition[merged_id] = "REJECTED_UNPAIRED_ENDPOINT"
                    normalized.append(
                        {
                            "row_id": merged_id,
                            "_d1_rejection_reason": "UNPAIRED_ENDPOINT",
                            "_d1_rejection_detail": (
                                f"base={base}; WT={len(endpoints['WT'])}; "
                                f"Mutant={len(endpoints['Mutant'])}"
                            ),
                            "_raw_row_fingerprint": _row_fingerprint(row),
                        }
                    )
                    malformed_pair_constructs += 1
            continue
        wt_id, wt_row = endpoints["WT"][0]
        mutant_id, mutant_row = endpoints["Mutant"][0]
        normalized_target_by_construct[wt_id] = {
            "target_kind": "normalized_pair_input",
            "row_id": base,
            "endpoint_role": "source",
        }
        normalized_target_by_construct[mutant_id] = {
            "target_kind": "normalized_pair_input",
            "row_id": base,
            "endpoint_role": "candidate",
        }
        construct_disposition[wt_id] = "NORMALIZED_SOURCE_ENDPOINT"
        construct_disposition[mutant_id] = "NORMALIZED_CANDIDATE_ENDPOINT"
        wt_sequence = _get(wt_row, "201bp")
        mutant_sequence = _get(mutant_row, "201bp")
        wt_normalized = re.sub(r"\s+", "", str(wt_sequence or "")).upper()
        mutant_normalized = re.sub(r"\s+", "", str(mutant_sequence or "")).upper()
        if len(wt_normalized) == 201 and len(mutant_normalized) == 201:
            pair_201nt_count += 1
            hamming = sum(
                left != right for left, right in zip(wt_normalized, mutant_normalized)
            )
            key = str(hamming)
        else:
            key = "length_mismatch"
        pair_hamming_distribution[key] = pair_hamming_distribution.get(key, 0) + 1
        normalized.append(
            {
                "row_id": base,
                "source_sequence": wt_sequence,
                "candidate_sequence": mutant_sequence,
                "endpoint": "deposited_log2_cpm_small_seq_on_plasmid",
                "source_value": labels.get(wt_id),
                "candidate_value": labels.get(mutant_id),
                "source_group": f"GSE200304:{base}",
                "sequence_cluster": f"GSE200304:{base}",
                "source_merged_id": wt_id,
                "candidate_merged_id": mutant_id,
                "source_construct_type": _get(wt_row, "Type"),
                "candidate_construct_type": _get(mutant_row, "Type"),
                "label_join_rule": "construct.merged_id == labels.Barcode",
                "library_batch": "GSE200304:GSM6030637_Twist_Oligo_Order",
                "barcode_batch": "GSE200304:GSM6030637_merged_id",
            }
        )
    for merged_id, row in sorted(controls):
        normalized_target_by_construct[merged_id] = {
            "target_kind": "rejected_row",
            "row_id": merged_id,
            "endpoint_role": "control",
        }
        construct_disposition[merged_id] = "REJECTED_CONTROL_NOT_WT_MUTANT_PAIR"
        normalized.append(
            {
                "row_id": merged_id,
                "_d1_rejection_reason": "CONTROL_NOT_WT_MUTANT_PAIR",
                "_d1_rejection_detail": str(_get(row, "Type", default="control")),
                "_raw_row_fingerprint": _row_fingerprint(row),
            }
        )

    label_joined_to_construct = sum(barcode in constructs_by_id for barcode in labels)
    construct_lineage: list[dict[str, Any]] = []
    for merged_id, row in constructs_by_id.items():
        row_index = construct_row_indices[merged_id]
        fingerprint = construct_row_fingerprints[merged_id]
        construct_lineage.append(
            {
                "schema_version": "gse200304_raw_row_lineage_v1",
                "dataset_id": "GSE200304",
                "raw_table_role": "construct_table",
                "raw_row_index": row_index,
                "raw_row_key": f"construct_table:{row_index}",
                "raw_row_id": merged_id,
                "raw_row_fingerprint_sha256": fingerprint,
                "lineage_id": _sha256_text(
                    f"GSE200304\x1fconstruct_table\x1f{row_index}\x1f{fingerprint}"
                ),
                "disposition": construct_disposition[merged_id],
                "normalized_target": normalized_target_by_construct[merged_id],
            }
        )
    for lineage in label_row_audit:
        if lineage["disposition"] != "PENDING":
            continue
        barcode = str(lineage["raw_row_id"])
        target = normalized_target_by_construct.get(barcode)
        if target is None:
            lineage["disposition"] = "REJECTED_UNJOINED_LABEL_BARCODE"
            lineage["normalized_target"] = None
        else:
            lineage["disposition"] = "JOINED_TO_CONSTRUCT"
            lineage["normalized_target"] = dict(target)
    raw_row_lineage = sorted(
        construct_lineage + label_row_audit,
        key=lambda row: (row["raw_table_role"], row["raw_row_index"]),
    )
    disposition_counts: dict[str, int] = {}
    for lineage in raw_row_lineage:
        disposition = str(lineage["disposition"])
        disposition_counts[disposition] = disposition_counts.get(disposition, 0) + 1
    raw_row_keys = [str(row["raw_row_key"]) for row in raw_row_lineage]
    lineage_ids = [str(row["lineage_id"]) for row in raw_row_lineage]
    raw_fingerprints = [
        str(row["raw_row_fingerprint_sha256"]) for row in raw_row_lineage
    ]
    lineage_summary = {
        "schema_version": "gse200304_raw_row_lineage_summary_v1",
        "row_counts_by_table": {
            "construct_table": len(construct_rows),
            "processed_label_table": len(label_rows),
        },
        "total_raw_rows": len(raw_row_lineage),
        "unique_raw_row_keys": len(set(raw_row_keys)),
        "duplicate_raw_row_keys": len(raw_row_keys) - len(set(raw_row_keys)),
        "unique_lineage_ids": len(set(lineage_ids)),
        "duplicate_lineage_ids": len(lineage_ids) - len(set(lineage_ids)),
        "unique_raw_row_fingerprints": len(set(raw_fingerprints)),
        "duplicate_raw_row_fingerprints": (
            len(raw_fingerprints) - len(set(raw_fingerprints))
        ),
        "rows_with_normalized_target": sum(
            row["normalized_target"] is not None for row in raw_row_lineage
        ),
        "rows_without_normalized_target": sum(
            row["normalized_target"] is None for row in raw_row_lineage
        ),
        "disposition_counts": dict(sorted(disposition_counts.items())),
    }
    audit = {
        "join_rule": "construct.merged_id == labels.Barcode",
        "pair_rule": "remove terminal _WT or _Mutant from merged_id",
        "construct_rows": len(construct_rows),
        "construct_unique_merged_ids": len(constructs_by_id),
        "label_rows": len(label_rows),
        "label_unique_barcodes": len(labels),
        "invalid_label_rows": invalid_label_rows,
        "label_barcodes_joined_to_construct": label_joined_to_construct,
        "unjoined_label_barcodes": len(labels) - label_joined_to_construct,
        "sequence_pair_groups": sum(
            len(endpoints["WT"]) == 1 and len(endpoints["Mutant"]) == 1
            for endpoints in grouped.values()
        ),
        "control_constructs": len(controls),
        "labeled_control_constructs": sum(
            merged_id in labels for merged_id, _ in controls
        ),
        "malformed_pair_constructs": malformed_pair_constructs,
        "pair_201nt_count": pair_201nt_count,
        "pair_hamming_distribution": dict(sorted(pair_hamming_distribution.items())),
        "raw_to_normalized_accounting": {
            "construct_rows": len(construct_rows),
            "paired_construct_rows": 2
            * sum(
                len(endpoints["WT"]) == 1 and len(endpoints["Mutant"]) == 1
                for endpoints in grouped.values()
            ),
            "control_construct_rows": len(controls),
            "malformed_pair_construct_rows": malformed_pair_constructs,
        },
        "raw_row_lineage_summary": lineage_summary,
        "raw_row_lineage_records": raw_row_lineage,
    }
    return normalized, audit


def _build_gse200304(
    rows: Sequence[Mapping[str, Any]],
    provenance: Mapping[str, Any],
    fixture_mode: bool,
) -> dict[str, Any]:
    dataset_id = "GSE200304"
    result = _base_result(dataset_id, provenance, fixture_mode)
    result["accounting"]["total_input_rows"] = len(rows)
    label_coverage = {
        "both_labeled": 0,
        "source_only": 0,
        "candidate_only": 0,
        "neither_labeled": 0,
    }
    sequence_pair_signatures: set[tuple[str, str]] = set()
    for index, row in enumerate(rows):
        row_id = _row_identifier(row, index)
        if row.get("_d1_rejection_reason"):
            result["rejected_records"].append(
                _reject(
                    dataset_id,
                    row,
                    index,
                    str(row["_d1_rejection_reason"]),
                    str(row.get("_d1_rejection_detail", "")),
                )
            )
            continue
        try:
            source, candidate = _pair_sequences(row)
        except ValueError as exc:
            result["rejected_records"].append(
                _reject(dataset_id, row, index, str(exc).split(":")[0], str(exc))
            )
            continue
        if len(source) != 201 or len(candidate) != 201:
            result["rejected_records"].append(
                _reject(
                    dataset_id,
                    row,
                    index,
                    "UNEXPECTED_PAIR_LENGTH",
                    f"source={len(source)} candidate={len(candidate)} expected=201",
                )
            )
            continue
        endpoint = str(
            _get(
                row,
                "endpoint",
                "phenotype",
                "assay_endpoint",
                default="",
            )
        ).strip()
        if not endpoint:
            result["rejected_records"].append(
                _reject(dataset_id, row, index, "MISSING_ENDPOINT")
            )
            continue
        source_value = _to_float(
            _get(row, "source_value", "source_value_raw", "wt_value", "ref_value")
        )
        candidate_value = _to_float(
            _get(
                row,
                "candidate_value",
                "candidate_value_raw",
                "mutant_value",
                "variant_value",
                "alt_value",
            )
        )
        delta = _to_float(_get(row, "delta", "delta_raw", "effect_size"))
        both_labeled = source_value is not None and candidate_value is not None
        if both_labeled:
            coverage_role = "both_labeled"
        elif source_value is not None:
            coverage_role = "source_only"
        elif candidate_value is not None:
            coverage_role = "candidate_only"
        else:
            coverage_role = "neither_labeled"
        try:
            record = _pair_record(
                dataset_id=dataset_id,
                row_id=row_id,
                source=source,
                candidate=candidate,
                endpoint=endpoint,
                provenance=provenance,
                source_value=source_value,
                candidate_value=candidate_value,
                delta=delta if both_labeled else None,
                delta_normalized=(
                    _to_float(_get(row, "delta_normalized", "normalized_effect"))
                    if both_labeled
                    else None
                ),
                effect_standard_error=_to_float(
                    _get(row, "effect_standard_error", "standard_error", "se")
                ),
                replicate_count=_to_int(_get(row, "replicate_count", "n_replicates")),
                label_status=(
                    "PROVIDED_LABEL_ONLY_BOTH_ENDPOINTS"
                    if both_labeled
                    else f"SEQUENCE_MAPPING_{coverage_role.upper()}"
                ),
                cell_context=str(
                    _get(row, "cell_context", "cell_type", default="prostate_MPRA")
                ),
                context_id=str(
                    _get(
                        row,
                        "context_id",
                        default="GSE200304:prostate_MPRA",
                    )
                ),
                assay_id=str(
                    _get(row, "assay_id", "assay", default=f"GSE200304:{endpoint}")
                ),
                organism="human",
                reporter="3UTR_MPRA",
                cargo=str(_get(row, "cargo", "reporter", default="reporter")),
                timepoint=str(_get(row, "timepoint", default="")) or None,
                gene=(
                    str(
                        _get(
                            row,
                            "gene",
                            "symbol",
                            "gene_name",
                            default="",
                        )
                    ).strip()
                    or None
                ),
                source_group=str(
                    _get(row, "source_group", "gene", "symbol", default="")
                )
                or None,
                sequence_cluster=str(_get(row, "sequence_cluster", default="")) or None,
                library_design="patient_mutation_201nt_WT_mutant_pairs",
                barcode_batch=str(
                    _get(
                        row,
                        "barcode_batch",
                        "batch",
                        default="GSE200304:GSM6030637_merged_id",
                    )
                )
                or None,
                quality_flags=(
                    ["PROVIDED_LABEL_ONLY_NOT_RAW_REPRODUCED"]
                    if both_labeled
                    else [
                        "SEQUENCE_PAIR_MAPPING_ONLY_NOT_CLOSED_MEASURED",
                        coverage_role.upper(),
                    ]
                ),
                pair_type=(
                    "true_wt_mutant"
                    if both_labeled
                    else "retrospective_constructed_neighbor"
                ),
                raw_endpoint="Freq",
                endpoint_transformation="provided_label_identity",
                library_batch=str(
                    _get(
                        row,
                        "library_batch",
                        default="GSE200304:GSM6030637_Twist_Oligo_Order",
                    )
                ),
                scaffold_group="GSE200304:201bp_3UTR_MPRA_scaffold",
            )
        except ValueError as exc:
            result["rejected_records"].append(
                _reject(dataset_id, row, index, str(exc).split(":")[0], str(exc))
            )
            continue
        if record["edit_count"] != 1 or record["edit_types"] != ["SUB"]:
            result["rejected_records"].append(
                _reject(
                    dataset_id,
                    row,
                    index,
                    "UNSUPPORTED_NON_SINGLE_SUBSTITUTION",
                    f"edit_types={record['edit_types']}; edit_count={record['edit_count']}",
                )
            )
            continue
        if not both_labeled:
            record["paper_split"] = "sequence_mapping_only"
            record["canonical_split"] = "sequence_mapping_only"
            record = validate_canonical_record(record)
        result["label_records"].append(record)
        result["paper_clean_records"].append(_paper_clean_pair(record, row_id))
        label_coverage[coverage_role] += 1
        sequence_pair_signatures.add((source, candidate))

    accepted = len(sequence_pair_signatures)
    result["paper_count_reconciliation"] = {
        "paper_reported_pairs": 6892,
        "production_expected_sequence_pairs": 6885,
        "production_expected_both_labeled_pairs": 6120,
        "production_expected_source_only_pairs": 192,
        "production_expected_candidate_only_pairs": 225,
        "production_expected_neither_labeled_pairs": 348,
        (
            "fixture_observed_pairs"
            if fixture_mode
            else "production_observed_sequence_pairs"
        ): accepted,
        "observed_label_coverage": label_coverage,
        "known_discrepancy": 7,
        "status": (
            "fixture_only_not_a_production_count_gate"
            if fixture_mode
            else (
                "sequence_and_label_coverage_reconciled"
                if accepted == 6885
                and label_coverage
                == {
                    "both_labeled": 6120,
                    "source_only": 192,
                    "candidate_only": 225,
                    "neither_labeled": 348,
                }
                else "UNEXPECTED_PRODUCTION_PAIR_COUNT"
            )
        ),
    }
    result["label_reproduction"] = {
        "status": "PROVIDED_LABEL_ONLY",
        "raw_reproduction_claim_allowed": False,
        "closed_measured_eligible_pairs": label_coverage["both_labeled"],
        "sequence_mapping_only_pairs": (
            label_coverage["source_only"]
            + label_coverage["candidate_only"]
            + label_coverage["neither_labeled"]
        ),
    }
    return _finalize_result(result)


def _build_gse246381(
    rows: Sequence[Mapping[str, Any]],
    provenance: Mapping[str, Any],
    fixture_mode: bool,
) -> dict[str, Any]:
    dataset_id = "GSE246381"
    result = _base_result(dataset_id, provenance, fixture_mode)
    result["accounting"]["total_input_rows"] = len(rows)
    for index, row in enumerate(rows):
        row_id = _row_identifier(row, index)
        try:
            source, candidate = _pair_sequences(row)
        except ValueError as exc:
            result["rejected_records"].append(
                _reject(dataset_id, row, index, str(exc).split(":")[0], str(exc))
            )
            continue
        source_value = _to_float(
            _get(row, "source_value", "source_value_raw", "ref_value")
        )
        candidate_value = _to_float(
            _get(
                row,
                "candidate_value",
                "candidate_value_raw",
                "alt_value",
            )
        )
        delta = _to_float(_get(row, "delta", "delta_raw", "effect_size"))
        reported_effect = _to_float(
            _get(row, "reported_effect", "reported_logfc", "logfc")
        )
        label_loaded = any(
            value is not None
            for value in (
                source_value,
                candidate_value,
                delta,
                reported_effect,
            )
        )
        endpoint = str(_get(row, "endpoint", default="sequence_pair_only")).strip()
        seq_id = str(_get(row, "seqid", "sequence_id", default=row_id))
        family_match = re.search(r"(?:^|;)Family=([^;]+)", seq_id)
        family = (
            family_match.group(1)
            if family_match
            else str(
                _get(
                    row,
                    "family",
                    "source_group",
                    default=f"not_reported:{_sha256_text(source)[:12]}",
                )
            )
        )
        try:
            record = _pair_record(
                dataset_id=dataset_id,
                row_id=row_id,
                source=source,
                candidate=candidate,
                endpoint=endpoint,
                provenance=provenance,
                source_value=source_value,
                candidate_value=candidate_value,
                delta=delta,
                delta_normalized=_to_float(
                    _get(row, "delta_normalized", "normalized_effect")
                ),
                effect_standard_error=_to_float(
                    _get(row, "effect_standard_error", "standard_error", "se")
                ),
                replicate_count=_to_int(_get(row, "replicate_count", "n_replicates")),
                label_status=(
                    "historical_processed_label_available_retrospective_only"
                    if label_loaded
                    else "sequence_pair_only_label_not_loaded"
                ),
                cell_context=str(
                    _get(row, "cell_context", "cell_type", default="unspecified")
                ),
                context_id=str(
                    _get(
                        row,
                        "context_id",
                        default="GSE246381:retrospective_unspecified_context",
                    )
                ),
                assay_id=str(
                    _get(row, "assay_id", "assay", default=f"GSE246381:{endpoint}")
                ),
                organism=str(_get(row, "organism", default="human")),
                reporter="5UTR_MPRA",
                cargo=str(_get(row, "cargo", "reporter", default="tdTomato")),
                timepoint=str(_get(row, "timepoint", default="")) or None,
                gene=(
                    str(
                        _get(
                            row,
                            "gene",
                            "symbol",
                            "gene_name",
                            default="",
                        )
                    ).strip()
                    or None
                ),
                source_group=f"GSE246381:family:{family}",
                sequence_cluster=f"GSE246381:family:{family}",
                library_design="ascertained_NDD_variants",
                barcode_batch=str(_get(row, "barcode_batch", "batch", default=""))
                or None,
                quality_flags=[
                    "HISTORICALLY_EXPOSED_RETROSPECTIVE_E4",
                    "NO_TRAINING_OR_SELECTION",
                    "NOT_UNTOUCHED_OR_SEALED",
                ],
                raw_source_sequence=source,
                raw_candidate_sequence=candidate,
                canonicalization_provenance={
                    "normalization": "uppercase_strip_whitespace_T_to_U",
                    "source_field": "RefSequence_or_explicit_alias",
                    "candidate_field": "AltSequence_or_explicit_alias",
                },
            )
        except ValueError as exc:
            result["rejected_records"].append(
                _reject(dataset_id, row, index, str(exc).split(":")[0], str(exc))
            )
            continue
        if record["edit_count"] != 1 or record["edit_types"] != ["SUB"]:
            result["rejected_records"].append(
                _reject(
                    dataset_id,
                    row,
                    index,
                    "UNSUPPORTED_NON_SINGLE_SUBSTITUTION",
                    f"edit_types={record['edit_types']}; edit_count={record['edit_count']}",
                )
            )
            continue
        if reported_effect is not None:
            record["label_provenance"]["reported_effect_raw"] = reported_effect
        record = validate_canonical_record(record)
        result["label_records"].append(record)
        result["paper_clean_records"].append(_paper_clean_pair(record, row_id))
    result["label_reproduction"] = {
        "status": (
            "retrospective_labels_separated_when_present;"
            "sequence_pairs_remain_usable_without_loading_labels"
        ),
        "training_or_selection_allowed": False,
    }
    return _finalize_result(result)


GSE217518_U3_SUFFIX_DNA = "GCGGCCGCGCAATAACTAGC"
GSE217518_U5_PREFIX_DNA = "CGCTAGGGATCCTCTAGTCA"
GSE217518_U5_SUFFIX_DNA = "ACCGGTCGCCACC"
GSE217518_ENDPOINT_PATTERN = re.compile(r"_(Ref|Mut)_([0-9]+)$")


def _gse217518_region(value: Any) -> str:
    token = _normalize_name(value)
    if token in {"u3", "3utr", "three_utr", "utr3"}:
        return "three_utr"
    if token in {"u5", "5utr", "five_utr", "utr5"}:
        return "five_utr"
    raise ValueError("UNRESOLVED_UTR_REGION")


def _trim_gse217518_oligo(
    raw_sequence: Any,
    region: str,
) -> tuple[str, str, dict[str, Any]]:
    raw_rna = _normalize_sequence(raw_sequence)
    raw_dna = raw_rna.replace("U", "T")
    if region == "three_utr":
        if not raw_dna.endswith(GSE217518_U3_SUFFIX_DNA):
            raise ValueError("FLANK_MISMATCH")
        insert_dna = raw_dna[: -len(GSE217518_U3_SUFFIX_DNA)]
        audit = {
            "raw_alphabet": "RNA_ACGU",
            "boundary_matching_alphabet": "DNA_ACGT_after_U_to_T",
            "prefix_trimmed": "",
            "suffix_trimmed": GSE217518_U3_SUFFIX_DNA,
            "suffix_length": len(GSE217518_U3_SUFFIX_DNA),
            "rule": "trim_observed_fixed_U3_suffix_only",
            "paper_nominal_insert_length": 115,
            "observed_boundary_note": (
                "no uniform U3 forward-primer prefix is present in the "
                "official CSV; forcing an unobserved prefix is forbidden"
            ),
        }
    else:
        if not raw_dna.startswith(GSE217518_U5_PREFIX_DNA) or not raw_dna.endswith(
            GSE217518_U5_SUFFIX_DNA
        ):
            raise ValueError("FLANK_MISMATCH")
        insert_dna = raw_dna[
            len(GSE217518_U5_PREFIX_DNA) : -len(GSE217518_U5_SUFFIX_DNA)
        ]
        audit = {
            "raw_alphabet": "RNA_ACGU",
            "boundary_matching_alphabet": "DNA_ACGT_after_U_to_T",
            "prefix_trimmed": GSE217518_U5_PREFIX_DNA,
            "prefix_length": len(GSE217518_U5_PREFIX_DNA),
            "suffix_trimmed": GSE217518_U5_SUFFIX_DNA,
            "suffix_length": len(GSE217518_U5_SUFFIX_DNA),
            "rule": "trim_observed_fixed_U5_prefix20_and_suffix13",
            "paper_nominal_insert_length": 115,
            "observed_boundary_note": (
                "the official CSV carries a uniform 13-nt U5 suffix, not a "
                "forced 20-nt reverse-primer sequence"
            ),
        }
    canonical_rna = _normalize_sequence(insert_dna)
    audit["raw_length"] = len(raw_rna)
    audit["canonical_insert_length"] = len(canonical_rna)
    audit["canonical_alphabet"] = "RNA_ACGU_after_T_to_U"
    return raw_rna, canonical_rna, audit


def _gse217518_gene(seq_name: str) -> str | None:
    match = re.search(r"\(([^()]+)\)", seq_name)
    return match.group(1) if match else None


def _build_gse217518(
    rows: Sequence[Mapping[str, Any]],
    provenance: Mapping[str, Any],
    fixture_mode: bool,
) -> dict[str, Any]:
    dataset_id = "GSE217518"
    result = _base_result(dataset_id, provenance, fixture_mode)
    result["accounting"]["total_input_rows"] = len(rows)
    groups: dict[str, dict[str, list[dict[str, Any]]]] = {}
    trimmed_lengths: dict[str, list[int]] = {
        "three_utr": [],
        "five_utr": [],
    }
    raw_region_counts = {"three_utr": 0, "five_utr": 0}

    for index, row in enumerate(rows):
        seq_name = str(_get(row, "seqName", "seq_name", "id", default=""))
        match = GSE217518_ENDPOINT_PATTERN.search(seq_name)
        if not match:
            result["rejected_records"].append(
                _reject(
                    dataset_id,
                    row,
                    index,
                    "UNPARSEABLE_ENDPOINT_ID",
                    "expected terminal _(Ref|Mut)_[0-9]+",
                )
            )
            continue
        allele = match.group(1)
        base_group = seq_name[: match.start()]
        try:
            region = _gse217518_region(
                _get(row, "region", "utr_region", "table_region")
            )
            raw_rna, canonical, trim_audit = _trim_gse217518_oligo(
                _get(row, "sequence"),
                region,
            )
        except ValueError as exc:
            result["rejected_records"].append(
                _reject(dataset_id, row, index, str(exc), str(exc))
            )
            continue
        half_life = _to_float(_get(row, "halfLife", "half_life"))
        if half_life is None:
            result["rejected_records"].append(
                _reject(dataset_id, row, index, "MISSING_OR_NONFINITE_LABEL")
            )
            continue
        raw_region_counts[region] += 1
        trimmed_lengths[region].append(len(canonical))
        groups.setdefault(base_group, {"Ref": [], "Mut": []})[allele].append(
            {
                "index": index,
                "row": row,
                "seq_name": seq_name,
                "base_group": base_group,
                "region": region,
                "raw_sequence": raw_rna,
                "canonical_sequence": canonical,
                "trim_audit": trim_audit,
                "half_life": half_life,
            }
        )

    accepted_pairs_by_region = {"three_utr": 0, "five_utr": 0}
    unpaired_group_counts = {"three_utr": 0, "five_utr": 0, "mixed": 0}
    for base_group in sorted(groups):
        endpoints = groups[base_group]
        if len(endpoints["Ref"]) != 1 or len(endpoints["Mut"]) != 1:
            members = endpoints["Ref"] + endpoints["Mut"]
            regions = {member["region"] for member in members}
            region_key = next(iter(regions)) if len(regions) == 1 else "mixed"
            unpaired_group_counts[region_key] += 1
            for member in members:
                result["rejected_records"].append(
                    _reject(
                        dataset_id,
                        member["row"],
                        member["index"],
                        "UNPAIRED_ENDPOINT",
                        (
                            f"base_group has Ref={len(endpoints['Ref'])}, "
                            f"Mut={len(endpoints['Mut'])}; exactly one of each "
                            "is required"
                        ),
                    )
                )
            continue
        reference = endpoints["Ref"][0]
        mutant = endpoints["Mut"][0]
        if reference["region"] != mutant["region"]:
            for member in (reference, mutant):
                result["rejected_records"].append(
                    _reject(
                        dataset_id,
                        member["row"],
                        member["index"],
                        "PAIR_REGION_MISMATCH",
                    )
                )
            continue
        source = reference["canonical_sequence"]
        candidate = mutant["canonical_sequence"]
        mismatch_count = (
            sum(left != right for left, right in zip(source, candidate))
            if len(source) == len(candidate)
            else None
        )
        pair_type = (
            "measured_indel_pair"
            if len(source) != len(candidate)
            else (
                "true_wt_mutant"
                if mismatch_count == 1
                else "measured_multi_edit_family"
            )
        )
        region = reference["region"]
        table_name = "SHdiNT_U3" if region == "three_utr" else "SHdiNT_U5"
        quality_flags = [
            "OFFICIAL_FIGURE4_PROCESSED_HALF_LIFE",
            "RAW_RNA_NORMALIZED_VIA_U_TO_T_BOUNDARY_MATCH_THEN_T_TO_U",
            "CONSTRUCTED_CANONICAL_PATH_NOT_OBSERVED",
        ]
        if region == "three_utr":
            quality_flags.extend(
                [
                    "U3_OBSERVED_FIXED_SUFFIX20_TRIMMED",
                    "PAPER_115BP_VS_OBSERVED_U3_LENGTH_DISTRIBUTION_DISCLOSED",
                ]
            )
        else:
            quality_flags.extend(
                [
                    "U5_OBSERVED_FIXED_PREFIX20_SUFFIX13_TRIMMED",
                    "PAPER_115BP_BOUNDARY_AUDITED",
                ]
            )
        sequence_provenance = {
            **dict(provenance),
            "raw_artifact": provenance.get("raw_artifact", "UNKNOWN"),
            "processed_artifact": {
                "official_table": table_name,
                "reference_seqName": reference["seq_name"],
                "mutant_seqName": mutant["seq_name"],
            },
            "raw_sequence_preserved": True,
            "canonicalization": {
                "reference": reference["trim_audit"],
                "mutant": mutant["trim_audit"],
            },
        }
        try:
            record = _pair_record(
                dataset_id=dataset_id,
                row_id=base_group,
                source=source,
                candidate=candidate,
                endpoint="half_life",
                provenance=provenance,
                source_value=reference["half_life"],
                candidate_value=mutant["half_life"],
                delta=None,
                delta_normalized=None,
                effect_standard_error=None,
                replicate_count=None,
                label_status=(
                    "official_processed_Figure4_halfLife_not_raw_read_reproduced"
                ),
                cell_context="SH-SY5Y",
                context_id=f"GSE217518:{table_name}:SH-SY5Y",
                assay_id=f"GSE217518:{table_name}:halfLife",
                organism="human",
                reporter="SHdiNT_reporter",
                cargo="SHdiNT_reporter",
                timepoint=None,
                gene=_gse217518_gene(base_group),
                source_group=f"GSE217518:{base_group}",
                sequence_cluster=f"GSE217518:{base_group}",
                library_design="disease_relevant_UTR_variant_SHdiNT",
                barcode_batch=f"NOT_APPLICABLE:GSE217518:{table_name}:no_barcode_field",
                quality_flags=quality_flags,
                pair_type=pair_type,
                region=region,
                raw_endpoint="halfLife",
                endpoint_transformation="identity",
                sequence_provenance=sequence_provenance,
                raw_source_sequence=reference["raw_sequence"],
                raw_candidate_sequence=mutant["raw_sequence"],
                canonicalization_provenance={
                    "reference": reference["trim_audit"],
                    "mutant": mutant["trim_audit"],
                    "base_group_rule": "remove_terminal_(Ref|Mut)_[0-9]+",
                },
                scaffold_group=f"GSE217518:{table_name}:SHdiNT_scaffold",
                library_batch=f"GSE217518:{table_name}",
            )
        except ValueError as exc:
            for member in (reference, mutant):
                result["rejected_records"].append(
                    _reject(
                        dataset_id,
                        member["row"],
                        member["index"],
                        str(exc).split(":")[0],
                        str(exc),
                    )
                )
            continue
        if record["edit_count"] == 0:
            for member in (reference, mutant):
                result["rejected_records"].append(
                    _reject(
                        dataset_id,
                        member["row"],
                        member["index"],
                        "NOOP_REF_MUT_PAIR",
                    )
                )
            continue
        result["label_records"].append(record)
        paper_clean = _paper_clean_pair(record, base_group)
        paper_clean.update(
            {
                "raw_source_sequence": reference["raw_sequence"],
                "raw_candidate_sequence": mutant["raw_sequence"],
                "source_seqName": reference["seq_name"],
                "candidate_seqName": mutant["seq_name"],
                "canonicalization_provenance": record["canonicalization_provenance"],
            }
        )
        result["paper_clean_records"].append(paper_clean)
        result["accounting"]["accepted_input_rows"] += 2
        accepted_pairs_by_region[region] += 1

    def mode_or_none(values: Sequence[int]) -> int | None:
        return statistics.multimode(values)[0] if values else None

    result["paper_count_reconciliation"] = {
        "paper_reported_total_sequences": 12472,
        "paper_reported_pairs": 6555,
        "official_csv_rows": raw_region_counts,
        "official_csv_unique_pairs": accepted_pairs_by_region,
        "known_production_snapshot": {
            "three_utr_rows": 3275,
            "three_utr_unique_pairs": 1124,
            "five_utr_rows": 4601,
            "five_utr_unique_pairs": 1756,
        },
        "unpaired_group_counts": unpaired_group_counts,
        "canonical_insert_length_mode": {
            region: mode_or_none(lengths) for region, lengths in trimmed_lengths.items()
        },
        "status": (
            "fixture_only_not_a_production_count_gate"
            if fixture_mode
            else (
                "official_csv_counts_reconciled_and_paper_difference_disclosed"
                if raw_region_counts == {"three_utr": 3275, "five_utr": 4601}
                and accepted_pairs_by_region == {"three_utr": 1124, "five_utr": 1756}
                else "UNEXPECTED_PRODUCTION_COUNTS"
            )
        ),
    }
    result["label_reproduction"] = {
        "status": "official_processed_Figure4_halfLife_not_raw_read_reproduced",
        "raw_read_reproduction_claim_allowed": False,
    }
    return _finalize_result(result)


def _mprau_reference_is_complete(reference_audit: Mapping[str, Any] | None) -> bool:
    if not reference_audit:
        return False
    status = str(reference_audit.get("status", "")).lower()
    coverage = _to_float(reference_audit.get("reference_coverage"))
    roundtrip = _to_float(reference_audit.get("roundtrip_fraction"))
    reference_sha = str(reference_audit.get("reference_sha256") or "")
    return (
        status in {"frozen", "verified_frozen", "passed_frozen"}
        and coverage == 1.0
        and roundtrip == 1.0
        and bool(re.fullmatch(r"[0-9a-f]{64}", reference_sha))
    )


def _blocked_result(
    dataset_id: str,
    provenance: Mapping[str, Any],
    fixture_mode: bool,
    reason_code: str,
    *,
    reference_audit: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    result = _base_result(dataset_id, provenance, fixture_mode)
    result.update(
        {
            "status": "blocked",
            "reason_code": reason_code,
            "paper_eligible": False,
            "reference_audit": dict(reference_audit or {}),
            "read_final_labels": False,
        }
    )
    result["label_reproduction"] = {
        "status": "not_attempted_fail_closed",
        "reason_code": reason_code,
    }
    return _finalize_result(result)


def _build_mprau(
    rows: Sequence[Mapping[str, Any]],
    provenance: Mapping[str, Any],
    fixture_mode: bool,
    reference_audit: Mapping[str, Any] | None,
) -> dict[str, Any]:
    dataset_id = "MPRAu_processed_ENCSR854RUF"
    if not _mprau_reference_is_complete(reference_audit):
        return _blocked_result(
            dataset_id,
            provenance,
            fixture_mode,
            "BLOCKED_MPRAU_REFERENCE_NOT_FROZEN_100_PERCENT",
            reference_audit=reference_audit,
        )

    result = _base_result(dataset_id, provenance, fixture_mode)
    result["accounting"]["total_input_rows"] = len(rows)
    for index, row in enumerate(rows):
        row_id = _row_identifier(row, index)
        try:
            source, candidate = _pair_sequences(row)
        except ValueError as exc:
            result["rejected_records"].append(
                _reject(dataset_id, row, index, str(exc).split(":")[0], str(exc))
            )
            continue
        endpoint = str(_get(row, "endpoint", default="allele_specific_rna_abundance"))
        source_value = _to_float(
            _get(row, "source_value", "source_value_raw", "reference_value")
        )
        candidate_value = _to_float(
            _get(row, "candidate_value", "candidate_value_raw", "alternate_value")
        )
        delta = _to_float(_get(row, "delta", "delta_raw", "effect_size"))
        if source_value is None and candidate_value is None and delta is None:
            result["rejected_records"].append(
                _reject(dataset_id, row, index, "MISSING_PROCESSED_LABEL")
            )
            continue
        try:
            record = _pair_record(
                dataset_id=dataset_id,
                row_id=row_id,
                source=source,
                candidate=candidate,
                endpoint=endpoint,
                provenance=provenance,
                source_value=source_value,
                candidate_value=candidate_value,
                delta=delta,
                delta_normalized=_to_float(
                    _get(row, "delta_normalized", "normalized_effect")
                ),
                effect_standard_error=_to_float(
                    _get(row, "effect_standard_error", "standard_error", "se")
                ),
                replicate_count=_to_int(_get(row, "replicate_count", "n_replicates")),
                label_status="processed_labels_provided_not_raw_reproduced",
                cell_context=str(
                    _get(row, "cell_context", "cell_line", default="unspecified")
                ),
                context_id=str(
                    _get(
                        row,
                        "context_id",
                        default="MPRAu:unspecified_cell_context",
                    )
                ),
                assay_id=str(_get(row, "assay_id", "assay", default="MPRAu")),
                organism="human",
                reporter="3UTR_MPRAu",
                cargo="reporter",
                timepoint=None,
                gene=str(_get(row, "gene", "symbol", default="UNKNOWN")),
                source_group=str(_get(row, "source_group", "variant_group", default=""))
                or None,
                sequence_cluster=None,
                library_design=(
                    "variant_panel_plus_targeted_nonoverlapping_5bp_deletion_tiling"
                ),
                barcode_batch=str(_get(row, "barcode_batch", "batch", default=""))
                or None,
                quality_flags=["REFERENCE_RECONSTRUCTION_FROZEN_100_PERCENT"],
            )
        except ValueError as exc:
            result["rejected_records"].append(
                _reject(dataset_id, row, index, str(exc).split(":")[0], str(exc))
            )
            continue
        valid_sub = record["edit_count"] == 1 and record["edit_types"] == ["SUB"]
        valid_del = (
            record["edit_count"] == 1
            and record["edit_types"] == ["DEL"]
            and len(record["reference_alleles"][0]) == 5
            and record["alternate_alleles"][0] == ""
        )
        if not (valid_sub or valid_del):
            result["rejected_records"].append(
                _reject(
                    dataset_id,
                    row,
                    index,
                    "UNSUPPORTED_MPRAU_ACTION",
                    "only single SUB or documented 5-bp DEL is admissible",
                )
            )
            continue
        result["label_records"].append(record)
        result["paper_clean_records"].append(_paper_clean_pair(record, row_id))

    # MPRAu is all-or-nothing.  A partial accepted subset could manufacture a
    # deletion claim from incomplete reference reconstruction.
    if result["rejected_records"]:
        blocked = _blocked_result(
            dataset_id,
            provenance,
            fixture_mode,
            "BLOCKED_MPRAU_NOT_ALL_ROWS_ROUNDTRIP",
            reference_audit=reference_audit,
        )
        blocked["accounting"]["total_input_rows"] = len(rows)
        blocked["accounting"]["rejected_rows"] = len(rows)
        blocked["accounting"]["accounted_rows"] = len(rows)
        blocked["rejected_records"] = [
            {
                "dataset_id": dataset_id,
                "row_id": "all_rows",
                "row_fingerprint_sha256": _sha256_text(
                    json.dumps(
                        [
                            record["row_fingerprint_sha256"]
                            for record in result["rejected_records"]
                        ]
                    )
                ),
                "reason_code": "BLOCKED_MPRAU_NOT_ALL_ROWS_ROUNDTRIP",
                "detail": (
                    f"{len(result['rejected_records'])}/{len(rows)} input rows "
                    "failed canonical reconstruction; no partial subset admitted"
                ),
            }
        ]
        blocked["accounting"]["rejected_rows"] = 1
        blocked["accounting"]["accounted_rows"] = 1
        blocked["accounting"]["total_input_rows"] = 1
        return blocked
    result["reference_audit"] = dict(reference_audit or {})
    result["label_reproduction"] = {
        "status": "processed_labels_provided_not_raw_reproduced",
        "reference_reconstruction": "frozen_100_percent_roundtrip",
    }
    return _finalize_result(result)


def build_dataset_rows(
    dataset_id: str,
    rows: Iterable[Mapping[str, Any]],
    *,
    provenance: Mapping[str, Any],
    fixture_mode: bool = False,
    reference_audit: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    """Build one D1 dataset from already extracted rows.

    Blocked datasets return before iterating ``rows``.  This is intentional:
    metadata-only external candidates cannot have final labels opened merely
    because a caller accidentally supplied a row iterator.
    """
    normalized_provenance = _normalise_provenance(provenance)
    if dataset_id in BLOCKED_DATASET_POLICIES:
        policy = BLOCKED_DATASET_POLICIES[dataset_id]
        return _blocked_result(
            dataset_id,
            normalized_provenance,
            fixture_mode,
            str(policy["reason_code"]),
        )
    if dataset_id not in ACTIVE_DATASET_POLICIES:
        return _blocked_result(
            dataset_id,
            normalized_provenance,
            fixture_mode,
            "UNREGISTERED_DATASET_FAIL_CLOSED",
        )
    if (
        dataset_id == "MPRAu_processed_ENCSR854RUF"
        and not _mprau_reference_is_complete(reference_audit)
    ):
        # Return before iterating rows.  Processed MPRAu labels cannot be
        # opened while the exact reference/round-trip gate is incomplete.
        return _blocked_result(
            dataset_id,
            normalized_provenance,
            fixture_mode,
            "BLOCKED_MPRAU_REFERENCE_NOT_FROZEN_100_PERCENT",
            reference_audit=reference_audit,
        )
    materialized_rows = [dict(row) for row in rows]
    if dataset_id == "GSE114002":
        return _build_gse114002(materialized_rows, normalized_provenance, fixture_mode)
    if dataset_id == "GSE200304":
        return _build_gse200304(materialized_rows, normalized_provenance, fixture_mode)
    if dataset_id == "GSE246381":
        return _build_gse246381(materialized_rows, normalized_provenance, fixture_mode)
    if dataset_id == "GSE217518":
        return _build_gse217518(materialized_rows, normalized_provenance, fixture_mode)
    if dataset_id == "MPRAu_processed_ENCSR854RUF":
        return _build_mprau(
            materialized_rows,
            normalized_provenance,
            fixture_mode,
            reference_audit,
        )
    raise AssertionError(f"unhandled active dataset: {dataset_id}")


def _json_line(record: Mapping[str, Any]) -> str:
    return (
        json.dumps(
            record,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        )
        + "\n"
    )


def _install_exclusive(path: Path, temporary: Path) -> None:
    """Atomically install a new file and refuse every overwrite race."""
    try:
        os.link(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _write_jsonl_atomic(path: Path, records: Sequence[Mapping[str, Any]]) -> None:
    if path.exists():
        raise FileExistsError(f"refusing to overwrite D1 artifact: {path}")
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    with temporary.open("x", encoding="utf-8", newline="") as handle:
        for record in records:
            handle.write(_json_line(record))
    _install_exclusive(path, temporary)


def _write_json_atomic(path: Path, payload: Mapping[str, Any]) -> None:
    if path.exists():
        raise FileExistsError(f"refusing to overwrite D1 artifact: {path}")
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    with temporary.open("x", encoding="utf-8", newline="") as handle:
        handle.write(
            json.dumps(payload, indent=2, ensure_ascii=False, sort_keys=True) + "\n"
        )
    _install_exclusive(path, temporary)


def _output_entry(dataset_root: Path, path: Path, count: int) -> dict[str, Any]:
    return {
        "path": path.relative_to(dataset_root).as_posix(),
        "bytes": path.stat().st_size,
        "sha256": _sha256_file(path),
        "records": count,
    }


def _identity_binding(
    records: Sequence[Mapping[str, Any]],
    identity_field: str,
) -> dict[str, Any]:
    identities = [str(record.get(identity_field, "")) for record in records]
    return {
        "identity_field": identity_field,
        "records": len(records),
        "identities_unique": len(identities) == len(set(identities)),
        "identities_sha256": _sha256_text(
            "\n".join(identities) + ("\n" if identities else "")
        ),
    }


def write_dataset_result(
    result: Mapping[str, Any],
    stage_d1_root: Path,
) -> Path:
    """Write one immutable dataset snapshot below a fresh D1 stage root."""
    dataset_id = str(result["dataset_id"])
    dataset_root = stage_d1_root / "datasets" / dataset_id
    if dataset_root.exists() and any(dataset_root.iterdir()):
        raise FileExistsError(
            f"refusing to overwrite existing D1 dataset result: {dataset_root}"
        )
    dataset_root.mkdir(parents=True, exist_ok=True)

    paths = {
        "paper_clean": dataset_root / "paper_clean" / "records.jsonl",
        "label_store": dataset_root / "canonical" / "records_with_labels.jsonl",
        "candidate_store": dataset_root / "candidate_store" / "candidates.jsonl",
        "rejected": dataset_root / "rejected" / "records.jsonl",
        "auxiliary": dataset_root / "paper_clean" / "auxiliary_records.jsonl",
        "raw_row_lineage": dataset_root / "audit" / "raw_row_lineage.jsonl",
        "accounting": dataset_root / "audit" / "accounting.json",
    }
    _write_jsonl_atomic(paths["paper_clean"], result["paper_clean_records"])
    _write_jsonl_atomic(paths["label_store"], result["label_records"])
    _write_jsonl_atomic(paths["candidate_store"], result["candidate_records"])
    _write_jsonl_atomic(paths["rejected"], result["rejected_records"])
    _write_jsonl_atomic(paths["auxiliary"], result["auxiliary_records"])
    _write_jsonl_atomic(
        paths["raw_row_lineage"], result.get("raw_row_lineage_records", [])
    )
    _write_json_atomic(paths["accounting"], result["accounting"])

    manifest: dict[str, Any] = {
        key: value
        for key, value in result.items()
        if key
        not in {
            "paper_clean_records",
            "label_records",
            "candidate_records",
            "rejected_records",
            "auxiliary_records",
            "raw_row_lineage_records",
        }
    }
    manifest["outputs"] = {
        "paper_clean": _output_entry(
            dataset_root,
            paths["paper_clean"],
            len(result["paper_clean_records"]),
        ),
        "label_store": _output_entry(
            dataset_root, paths["label_store"], len(result["label_records"])
        ),
        "candidate_store": _output_entry(
            dataset_root,
            paths["candidate_store"],
            len(result["candidate_records"]),
        ),
        "rejected": _output_entry(
            dataset_root, paths["rejected"], len(result["rejected_records"])
        ),
        "auxiliary": _output_entry(
            dataset_root,
            paths["auxiliary"],
            len(result["auxiliary_records"]),
        ),
        "raw_row_lineage": _output_entry(
            dataset_root,
            paths["raw_row_lineage"],
            len(result.get("raw_row_lineage_records", [])),
        ),
        "accounting": _output_entry(dataset_root, paths["accounting"], 1),
    }
    manifest["content_bindings"] = {
        "paper_clean": _identity_binding(result["paper_clean_records"], "record_id"),
        "canonical": _identity_binding(result["label_records"], "record_id"),
        "candidate_store": _identity_binding(result["candidate_records"], "record_id"),
        "rejected": _identity_binding(result["rejected_records"], "row_id"),
        "auxiliary": _identity_binding(result["auxiliary_records"], "row_id"),
        "raw_row_lineage": _identity_binding(
            result.get("raw_row_lineage_records", []), "lineage_id"
        ),
    }
    manifest["physical_separation"] = {
        "candidate_store_contains_labels": False,
        "candidate_store_required_to_open_label_store": False,
        "candidate_store_path": manifest["outputs"]["candidate_store"]["path"],
        "label_store_path": manifest["outputs"]["label_store"]["path"],
        "paths_are_distinct": (
            manifest["outputs"]["candidate_store"]["path"]
            != manifest["outputs"]["label_store"]["path"]
        ),
    }
    _write_json_atomic(dataset_root / "manifest.json", manifest)
    return dataset_root


def iter_table_rows(
    path: Path,
    *,
    file_format: str | None = None,
    delimiter: str | None = None,
    sheet_name: str | None = None,
    defaults: Mapping[str, Any] | None = None,
) -> Iterator[dict[str, Any]]:
    """Stream CSV/TSV/JSONL/XLSX rows without guessing scientific semantics."""
    suffixes = "".join(path.suffixes).lower()
    fmt = (file_format or "").lower()
    if not fmt:
        if suffixes.endswith(".jsonl"):
            fmt = "jsonl"
        elif suffixes.endswith((".xlsx", ".xlsm")):
            fmt = "xlsx"
        elif suffixes.endswith((".tsv", ".tsv.gz")):
            fmt = "tsv"
        elif suffixes.endswith((".csv", ".csv.gz")):
            fmt = "csv"
        else:
            raise ValueError(f"cannot infer input format for {path}")
    defaults = dict(defaults or {})

    if fmt == "jsonl":
        with path.open(encoding="utf-8") as handle:
            for line_number, line in enumerate(handle, start=1):
                if not line.strip():
                    continue
                value = json.loads(line)
                if not isinstance(value, Mapping):
                    raise ValueError(
                        f"{path}:{line_number}: JSONL row is not an object"
                    )
                yield {**defaults, **dict(value)}
        return

    if fmt in {"csv", "tsv"}:
        opener = gzip.open if suffixes.endswith(".gz") else open
        actual_delimiter = delimiter or ("\t" if fmt == "tsv" else ",")
        with opener(path, "rt", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle, delimiter=actual_delimiter)
            if reader.fieldnames is None:
                raise ValueError(f"{path}: missing header")
            for row in reader:
                yield {**defaults, **dict(row)}
        return

    if fmt == "xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("XLSX_INPUT_REQUIRES_OPENPYXL") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if not sheet_name:
                raise ValueError("XLSX_INPUT_REQUIRES_EXPLICIT_SHEET_NAME")
            worksheet = workbook[sheet_name]
            rows = worksheet.iter_rows(values_only=True)
            try:
                header = [str(value or "") for value in next(rows)]
            except StopIteration:
                return
            for values in rows:
                yield {
                    **defaults,
                    **dict(zip(header, values)),
                }
        finally:
            workbook.close()
        return
    raise ValueError(f"unsupported input format: {fmt}")


def build_dataset_from_config(
    dataset_config: Mapping[str, Any],
    *,
    fixture_mode: bool = False,
) -> dict[str, Any]:
    """Build one dataset from a frozen, explicit input configuration."""
    dataset_id = str(dataset_config["dataset_id"])
    policy = dataset_policy(dataset_id)

    # Blocked/metadata-only datasets must not even stat or open supplied final
    # label paths.
    if policy["status"] == "blocked":
        return build_dataset_rows(
            dataset_id,
            (),
            provenance={
                "input_files": [],
                "download_manifest": dataset_config.get("download_manifest", "UNKNOWN"),
                "license": dataset_config.get("license", "UNKNOWN"),
                "input_access": "not_opened_due_to_fail_closed_dataset_policy",
            },
            fixture_mode=fixture_mode,
        )
    if (
        dataset_id == "MPRAu_processed_ENCSR854RUF"
        and not _mprau_reference_is_complete(dataset_config.get("reference_audit"))
    ):
        return build_dataset_rows(
            dataset_id,
            (),
            provenance={
                "input_files": [],
                "download_manifest": dataset_config.get("download_manifest", "UNKNOWN"),
                "license": dataset_config.get("license", "UNKNOWN"),
                "input_access": ("not_opened_due_to_fail_closed_dataset_policy"),
            },
            fixture_mode=fixture_mode,
            reference_audit=dataset_config.get("reference_audit"),
        )

    input_files = dataset_config.get("input_files", [])
    rows: list[dict[str, Any]] = []
    rows_by_role: dict[str, list[dict[str, Any]]] = {}
    provenance_files: list[dict[str, Any]] = []
    for input_spec in input_files:
        path = Path(str(input_spec["path"]))
        if not path.is_file():
            raise FileNotFoundError(path)
        provenance_files.append(
            {
                "path": str(path),
                "bytes": path.stat().st_size,
                "sha256": _sha256_file(path),
                "role": input_spec.get("role", "input"),
                "format": input_spec.get("format"),
                "delimiter": input_spec.get("delimiter"),
                "sheet_name": input_spec.get("sheet_name"),
                "defaults": dict(input_spec.get("defaults") or {}),
            }
        )
        loaded_rows = list(
            iter_table_rows(
                path,
                file_format=input_spec.get("format"),
                delimiter=input_spec.get("delimiter"),
                sheet_name=input_spec.get("sheet_name"),
                defaults=input_spec.get("defaults"),
            )
        )
        rows.extend(loaded_rows)
        role = _normalize_name(input_spec.get("role", "input"))
        rows_by_role.setdefault(role, []).extend(loaded_rows)
    provenance = {
        "input_files": provenance_files,
        "download_manifest": dataset_config.get("download_manifest", "UNKNOWN"),
        "license": dataset_config.get("license", "UNKNOWN"),
        "fixture": fixture_mode,
    }
    extraction_audit: dict[str, Any] | None = None
    if dataset_id == "GSE200304":
        construct_roles = {
            "construct",
            "construct_table",
            "constructs",
            "sequence_construct_table",
        }
        label_roles = {
            "label",
            "label_table",
            "labels",
            "processed_label_table",
            "processed_labels",
        }
        construct_rows = [
            row
            for role, role_rows in rows_by_role.items()
            if role in construct_roles
            for row in role_rows
        ]
        label_rows = [
            row
            for role, role_rows in rows_by_role.items()
            if role in label_roles
            for row in role_rows
        ]
        if construct_rows and label_rows:
            rows, extraction_audit = extract_gse200304_exact_join(
                construct_rows,
                label_rows,
            )
        elif not fixture_mode:
            raise ValueError(
                "GSE200304_REQUIRES_EXACT_TWO_TABLE_JOIN_WITH_ROLES_"
                "construct_table_AND_processed_label_table"
            )

    result = build_dataset_rows(
        dataset_id,
        rows,
        provenance=provenance,
        fixture_mode=fixture_mode,
        reference_audit=dataset_config.get("reference_audit"),
    )
    if extraction_audit is not None:
        lineage_records = extraction_audit.pop("raw_row_lineage_records", [])
        result["raw_row_lineage_records"] = lineage_records
        result["extraction_audit"] = extraction_audit
        exact_extraction = (
            extraction_audit["construct_rows"] == 13836
            and extraction_audit["construct_unique_merged_ids"] == 13836
            and extraction_audit["label_unique_barcodes"] == 12704
            and extraction_audit["label_barcodes_joined_to_construct"] == 12704
            and extraction_audit["unjoined_label_barcodes"] == 0
            and extraction_audit["sequence_pair_groups"] == 6885
            and extraction_audit["pair_201nt_count"] == 6885
            and extraction_audit["pair_hamming_distribution"] == {"1": 6885}
            and extraction_audit["control_constructs"] == 66
            and extraction_audit["labeled_control_constructs"] == 47
            and extraction_audit["malformed_pair_constructs"] == 0
        )
        result["extraction_audit"]["production_gate_exact"] = exact_extraction
        if not fixture_mode:
            result["paper_eligible"] = (
                bool(result["paper_eligible"])
                and exact_extraction
                and result["paper_count_reconciliation"]["status"]
                == "sequence_and_label_coverage_reconciled"
            )
    return result


def pipeline_stage_descriptor(
    dataset_id: str,
    stage: str,
) -> dict[str, Any]:
    """Return the audit contract behind a dataset-local stage entrypoint."""
    if stage not in PIPELINE_STAGES:
        raise ValueError(f"unsupported D1 pipeline stage: {stage}")
    policy = dataset_policy(dataset_id)
    return {
        "schema_version": "d1_pipeline_entrypoint_v2",
        "dataset_id": dataset_id,
        "stage": stage,
        "dataset_status": policy["status"],
        "implementation": "data.utr_benchmark_v2.d1_builder",
        "atomic_full_build_entrypoint": ("scripts/data/build_d1_utr_benchmark.py"),
        "writes_production_data": False,
        "scientific_gate_claimed": False,
        "note": (
            "This dataset-local entrypoint exposes the contract stage for "
            "audit. The full builder executes all stages atomically so a "
            "partial stage cannot be mistaken for a frozen D1 snapshot."
        ),
    }


def pipeline_stage_main(dataset_id: str, stage: str) -> int:
    print(
        json.dumps(
            pipeline_stage_descriptor(dataset_id, stage),
            ensure_ascii=False,
            sort_keys=True,
        )
    )
    return 0
