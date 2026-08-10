from __future__ import annotations

import errno
import gzip
import hashlib
import importlib.util
import json
import math
import shutil
import statistics
import subprocess
from pathlib import Path
from types import SimpleNamespace
from typing import Any

import pytest
from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[2]
MODULE_PATH = ROOT / "scripts" / "route_a_v3" / "qualify_gse149487_plumage.py"
V4_PATH = ROOT / "scripts" / "route_a_v3" / "reconstruct_gse149487_plumage.py"
TEST_PATH = Path(__file__).resolve()
ASSET_MANIFEST_PATH = ROOT / "configs" / "route_a_v3_gse149487_asset_manifest_v2.json"
PROTOCOL_PATH = ROOT / "configs" / "route_a_v3_gse149487_a1_qualification.json"
CANONICAL_SCHEMA_PATH = ROOT / "schemas" / "route_a_v3" / "canonical_intervention_record.schema.json"
A0_REGISTRY_PATH = ROOT / "docs" / "execution" / "route_a_v3_data_role_registry.yaml"

SPEC = importlib.util.spec_from_file_location("qualify_gse149487_plumage", MODULE_PATH)
assert SPEC and SPEC.loader
QUAL = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(QUAL)

V4_SPEC = importlib.util.spec_from_file_location("plumage_v4_for_fixtures", V4_PATH)
assert V4_SPEC and V4_SPEC.loader
V4 = importlib.util.module_from_spec(V4_SPEC)
V4_SPEC.loader.exec_module(V4)


SOURCE_SEQUENCE = "AACCGG"
CANDIDATE_SEQUENCE = "AATCGG"
MUTANT_DESCRIPTION = "GENE1_C_T_chr1_100_105"
WT_DESCRIPTION = "GENE1_WT_chr1_100_105"
SOURCE_KEYS = tuple(f"SOURCE_MEMBER_SECRET_{index:02d}" for index in range(20))
CANDIDATE_KEYS = tuple(f"CANDIDATE_MEMBER_SECRET_{index:02d}" for index in range(20))
ALL_KEYS = SOURCE_KEYS + CANDIDATE_KEYS


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def _write_moesm8(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = V4.MOESM8_SEQUENCE_SHEET
    sheet.append(list(V4.MOESM8_6A_RAW_HEADER))
    sheet.append(
        [
            len(CANDIDATE_SEQUENCE),
            "GENE1",
            "GENE1_chr1_100_C_T_UTR5",
            CANDIDATE_SEQUENCE,
        ]
    )
    sheet.append(
        [
            len(SOURCE_SEQUENCE),
            "GENE1",
            "GENE1_chr1_100_WT_UTR5",
            SOURCE_SEQUENCE,
        ]
    )
    workbook.save(path)


def _write_mapping_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = V4.LIM6C_SHEET
    sheet.append(list(V4.LIM6C_RAW_HEADER))
    for key in CANDIDATE_KEYS:
        sheet.append([MUTANT_DESCRIPTION, key, *([None] * 15)])
    for key in SOURCE_KEYS:
        sheet.append([WT_DESCRIPTION, key, *([None] * 15)])
    workbook.save(path)


def _write_reference_workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "REFERENCE_ONLY"
    workbook.active.append(["REFERENCE_ONLY"])
    workbook.save(path)


def _raw_count_for(
    assay: str,
    key: str,
    replicate: int,
    *,
    context: str = "PC3",
) -> int:
    candidate = key in CANDIDATE_KEYS
    assay_index = {"DNA": 1, "TOTALRNA": 2, "POLYSOME": 3}[assay]
    context_offset = {"PC3": 0, "293T": 10_000}[context]
    slot_sentinel = context_offset + assay_index * 100 + replicate
    biological_signal = {
        "DNA": 100,
        "TOTALRNA": 300 if candidate else 100,
        "POLYSOME": 500 if candidate else 100,
    }[assay]
    return biological_signal * 100 + slot_sentinel


def _write_raw_table(
    path: Path,
    *,
    assay: str,
    replicate: int,
    context: str = "PC3",
    keys: tuple[str, ...] = ALL_KEYS,
) -> None:
    with gzip.open(path, "wt", encoding="utf-8", newline="") as handle:
        handle.write("opaque_member\tcount\n")
        for key in keys:
            handle.write(
                f"{key}\t{_raw_count_for(assay, key, replicate, context=context)}\n"
            )


def _fixture_repo_and_data(
    tmp_path: Path,
    *,
    extra_raw_keys_by_context: dict[str, tuple[str, ...]] | None = None,
) -> tuple[Path, Path, Path, Path]:
    repo = tmp_path / "repo"
    data = tmp_path / "public_data"
    for relative in (
        "configs",
        "docs/goals",
        "docs/execution",
        "schemas/route_a_v3",
        "scripts/route_a_v3",
        "tests/route_a_v3",
    ):
        (repo / relative).mkdir(parents=True, exist_ok=True)
    data.mkdir()

    contract = repo / "docs/goals/MRNA_XEDITFLOW_ROUTE_A_V3.md"
    contract.write_text("fixture contract\n", encoding="utf-8")
    shutil.copy2(ROOT / "configs/route_a_v3_a1_qualification.json", repo / "configs/route_a_v3_a1_qualification.json")
    shutil.copy2(A0_REGISTRY_PATH, repo / "docs/execution/route_a_v3_data_role_registry.yaml")
    shutil.copy2(CANONICAL_SCHEMA_PATH, repo / "schemas/route_a_v3/canonical_intervention_record.schema.json")
    shutil.copy2(V4_PATH, repo / "scripts/route_a_v3/reconstruct_gse149487_plumage.py")
    shutil.copy2(MODULE_PATH, repo / "scripts/route_a_v3/qualify_gse149487_plumage.py")
    shutil.copy2(TEST_PATH, repo / "tests/route_a_v3/test_qualify_gse149487_plumage.py")

    moesm3 = data / "41467_2021_24445_MOESM3_ESM.xlsx"
    moesm8 = data / "41467_2021_24445_MOESM8_ESM.xlsx"
    mapping = data / "Lim_et_al_Supp_Tbl_6c_293T.xlsx"
    _write_reference_workbook(moesm3)
    _write_moesm8(moesm8)
    _write_mapping_workbook(mapping)

    raw_entries: list[dict[str, Any]] = []
    for context in QUAL.CONTEXTS:
        for assay in QUAL.ASSAYS:
            for replicate in QUAL.REPLICATES:
                filename = f"GSM_FIX_{context}_{assay}_rep{replicate}.txt.gz"
                path = data / filename
                extra_keys = (extra_raw_keys_by_context or {}).get(context, ())
                _write_raw_table(
                    path,
                    assay=assay,
                    replicate=replicate,
                    context=context,
                    keys=ALL_KEYS + extra_keys,
                )
                raw_entries.append(
                    {
                        "name": filename,
                        "bytes": path.stat().st_size,
                        "expected_bytes": path.stat().st_size,
                        "sha256": _sha256(path),
                        "url": f"https://example.invalid/{filename}",
                        "downloaded": True,
                    }
                )
    p0_manifest = data / "manifest.json"
    _write_json(p0_manifest, {"files": raw_entries})

    asset_manifest = json.loads(ASSET_MANIFEST_PATH.read_text(encoding="utf-8"))
    asset_manifest["source_manifest"]["sha256"] = _sha256(p0_manifest)
    supplement_paths = {
        "GSE149487_MOESM3": moesm3,
        "GSE149487_MOESM8": moesm8,
        "GSE149487_LIM6C_293T": mapping,
    }
    for asset in asset_manifest["assets"]:
        if asset["asset_id"] in supplement_paths:
            path = supplement_paths[asset["asset_id"]]
            asset["bytes"] = path.stat().st_size
            asset["sha256"] = _sha256(path)
    fixture_asset_manifest = repo / "configs/route_a_v3_gse149487_asset_manifest_v2.json"
    _write_json(fixture_asset_manifest, asset_manifest)

    protocol = json.loads(PROTOCOL_PATH.read_text(encoding="utf-8"))
    authority = protocol["authority"]
    authority["contract_sha256"] = _sha256(contract)
    authority["a1_qualification_sha256"] = _sha256(repo / authority["a1_qualification_path"])
    authority["data_role_registry_sha256"] = _sha256(repo / authority["data_role_registry_path"])
    authority["canonical_schema_sha256"] = _sha256(repo / authority["canonical_schema_path"])
    authority["asset_manifest_sha256"] = _sha256(fixture_asset_manifest)
    authority["v4_helper_sha256"] = _sha256(repo / authority["v4_helper_path"])
    authority["qualifier_sha256"] = _sha256(repo / authority["qualifier_path"])
    authority["focused_test_sha256"] = _sha256(repo / authority["focused_test_path"])
    authority["implementation_commit"] = "UNKNOWN_NOT_ASSERTED"
    fixture_protocol = repo / "configs/route_a_v3_gse149487_a1_qualification.json"
    _write_json(fixture_protocol, protocol)
    return repo, data, fixture_protocol, fixture_asset_manifest


def _minimal_bundle_payloads() -> dict[str, bytes]:
    return {
        name: b"{}\n"
        for name in QUAL.ALWAYS_OUTPUT_FILES
        if name != "SHA256SUMS"
    }


def _publication_run_metadata(suffix: str = "FALLBACK") -> dict[str, str]:
    return {
        "run_id": f"A1_{suffix}",
        "execution_id": f"GSE149487_{suffix}",
        "recorded_at": "2026-08-10T12:07:00+08:00",
    }


def _git(repo: Path, *args: str) -> str:
    return subprocess.run(
        ["git", "-C", str(repo), *args],
        check=True,
        capture_output=True,
        text=True,
    ).stdout.strip()


def _commit(repo: Path, message: str) -> str:
    _git(repo, "add", "-A")
    _git(
        repo,
        "-c",
        "user.name=Route A Fixture",
        "-c",
        "user.email=route-a-fixture@example.invalid",
        "commit",
        "-m",
        message,
    )
    return _git(repo, "rev-parse", "HEAD")


def _git_binding_fixture(tmp_path: Path) -> dict[str, Any]:
    repo = tmp_path / "git_binding_repo"
    repo.mkdir()
    _git(repo, "init")
    (repo / "README.md").write_text("accepted a0\n", encoding="utf-8")
    accepted = _commit(repo, "accepted a0")

    contract = repo / "docs/goals/contract.md"
    registry = repo / "docs/execution/registry.yaml"
    contract.parent.mkdir(parents=True)
    registry.parent.mkdir(parents=True)
    contract.write_text("active contract\n", encoding="utf-8")
    registry.write_text("active registry\n", encoding="utf-8")
    active = _commit(repo, "active authority")

    qualifier = repo / "scripts/qualifier.py"
    focused_test = repo / "tests/test_qualifier.py"
    qualifier.parent.mkdir(parents=True)
    focused_test.parent.mkdir(parents=True)
    qualifier.write_text("print('qualifier')\n", encoding="utf-8")
    focused_test.write_text("def test_fixture(): pass\n", encoding="utf-8")
    implementation = _commit(repo, "implementation")
    return {
        "repo": repo,
        "accepted": accepted,
        "active": active,
        "implementation": implementation,
        "authority_files": (
            ("docs/goals/contract.md", _sha256(contract)),
            ("docs/execution/registry.yaml", _sha256(registry)),
        ),
        "implementation_files": (
            ("scripts/qualifier.py", _sha256(qualifier)),
            ("tests/test_qualifier.py", _sha256(focused_test)),
        ),
        "qualifier": qualifier,
    }


def test_authority_protocol_and_additive_21_asset_manifest_are_closed() -> None:
    asset_manifest = QUAL._validate_asset_manifest(
        json.loads(ASSET_MANIFEST_PATH.read_text(encoding="utf-8"))
    )
    protocol = QUAL._validate_protocol(
        json.loads(PROTOCOL_PATH.read_text(encoding="utf-8"))
    )
    raw_assets = [asset for asset in asset_manifest["assets"] if asset["asset_kind"] == "GEO_RAW_COUNT"]
    supplement_assets = [
        asset for asset in asset_manifest["assets"] if asset["asset_kind"] == "SUPPLEMENT_WORKBOOK"
    ]
    assert len(asset_manifest["assets"]) == 21
    assert len(raw_assets) == 18
    assert len(supplement_assets) == 3
    assert {
        (asset["context"], asset["assay"], asset["biological_replicate"])
        for asset in raw_assets
    } == {
        (context, assay, replicate)
        for context in QUAL.CONTEXTS
        for assay in QUAL.ASSAYS
        for replicate in QUAL.REPLICATES
    }
    assert protocol["authority"]["asset_manifest_sha256"] == _sha256(ASSET_MANIFEST_PATH)
    assert protocol["authority"]["initial_contract_sha256"] == "d1c031aecdec710495f6861b380785cccd64663ac4bd97b4f479d6fdf372ea07"
    assert protocol["authority"]["contract_sha256"] == "3ba224de6277edd67387913cf1c83a5e1344e0ad44ef196db07d0772b45c4d79"
    assert protocol["authority"]["active_authority_commit"] == "d078060c81114687db5068902a5aad5d9bedbee6"
    assert protocol["authority"]["active_amendment_decision_ids"] == ["V3-DEC-017"]
    assert protocol["authority"]["v4_helper_sha256"] == _sha256(V4_PATH)
    assert protocol["authority"]["qualifier_sha256"] == _sha256(MODULE_PATH)
    assert protocol["authority"]["focused_test_sha256"] == _sha256(TEST_PATH)
    mapping_asset = next(
        asset
        for asset in asset_manifest["assets"]
        if asset["asset_id"] == "GSE149487_LIM6C_293T"
    )
    assert mapping_asset["source_commit"] == "d613b541d192d6c502a1ef8849c27e801a7fbfb9"
    assert mapping_asset["source_blob_git_sha1"] == (
        "2d4bae738074a1d1bffbafc5ec39da1dff679807"
    )
    assert mapping_asset["source_commit_author_date"] == "2024-06-17T19:25:39Z"
    assert mapping_asset["source_commit"] in mapping_asset["source_uri"]
    assert "/v1.0.0/" not in mapping_asset["source_uri"]
    assert protocol["mapping"]["outcome_blind_mapping_evidence_status"] == "UNKNOWN_NOT_ASSERTED"
    assert protocol["mapping"]["raw_key_reconciliation"][
        "prefrozen_excluded_key_sha256_by_class"
    ] == {"EXCLUDED_CONTROL": [], "OUT_OF_SCOPE": []}
    assert protocol["license_and_redistribution"]["audit_status"] == "UNKNOWN_NOT_ASSERTED"
    assert protocol["foundation_exposure"]["audit_status"] == "UNKNOWN_NOT_ASSERTED"
    assert protocol["canonical_v3"]["materialize_only_when_every_qualification_gate_passes"] is True
    assert protocol["scope"]["authority_update_allowed_by_qualifier"] is False
    assert protocol["input_contract"][
        "preopen_lstat_dev_ino_size_mtime_acceptance_authority"
    ] is False
    assert protocol["input_contract"]["verified_input_acceptance_authority"] == (
        "O_NOFOLLOW_FSTAT_BEFORE_AFTER_SHA256_AND_BYTES"
    )
    assert protocol["input_contract"][
        "same_descriptor_identity_and_size_must_remain_stable"
    ] is True
    implementation_commit = protocol["authority"]["implementation_commit"]
    implementation_blocker = "IMPLEMENTATION_COMMIT_UNKNOWN_NOT_ASSERTED"
    if implementation_commit == "UNKNOWN_NOT_ASSERTED":
        assert implementation_blocker in protocol["known_external_evidence_blockers"]
    else:
        assert QUAL.COMMIT_RE.fullmatch(implementation_commit)
        assert implementation_blocker not in protocol["known_external_evidence_blockers"]
    assert protocol["output_contract"]["primary_publication_mode"] == (
        QUAL.PRIMARY_PUBLICATION_MODE
    )
    assert protocol["output_contract"]["atomic_no_replace_kernel_primitive_primary"] is True
    assert protocol["output_contract"][
        "primary_commit_marker_written_last_in_staging_before_atomic_rename"
    ] is True
    assert protocol["output_contract"][
        "single_output_failure_transaction_claim_required"
    ] is True
    assert protocol["output_contract"][
        "transaction_claim_loser_failure_record_allowed"
    ] is False
    assert protocol["output_contract"]["directory_fsync_unsupported_errno_set"] == [
        "ENOSYS",
        "EINVAL",
        "ENOTSUP",
        "EOPNOTSUPP",
    ]
    assert protocol["output_contract"][
        "transaction_claim_parent_directory_fsync_unsupported_action"
    ] == "CONTINUE_WITH_FIXED_CAPABILITY_WARNING"
    assert protocol["output_contract"][
        "staging_directory_fsync_unsupported_action"
    ] == "CONTINUE_WITH_FIXED_CAPABILITY_WARNING"
    assert protocol["output_contract"]["staging_directory_fsync_other_error_action"] == (
        "FAIL_CLOSED_BEFORE_PUBLICATION"
    )
    assert protocol["output_contract"][
        "failure_record_post_commit_directory_fsync_error_action"
    ] == "RETURN_FAILURE_WITH_DURABILITY_WARNING"
    assert protocol["output_contract"]["atomic_no_replace_unsupported_errno_fallback"] == [
        "ENOSYS",
        "EINVAL",
        "ENOTSUP",
        "EOPNOTSUPP",
    ]
    assert protocol["output_contract"]["fallback_publication_mode"] == (
        QUAL.FALLBACK_PUBLICATION_MODE
    )
    assert protocol["output_contract"]["fallback_atomic_mkdir_loser_status"] == (
        "CONTENDED_NO_FAILURE_RECORD"
    )
    assert protocol["output_contract"]["fallback_bundle_file_write_mode"] == (
        "O_EXCL_AND_FILE_FSYNC"
    )
    assert protocol["output_contract"]["fallback_required_terminal_metadata_files"] == [
        QUAL.PUBLICATION_COMMIT_FILENAME
    ]
    assert protocol["output_contract"][
        "required_terminal_metadata_files_all_publication_modes"
    ] == [QUAL.PUBLICATION_COMMIT_FILENAME]
    assert protocol["output_contract"][
        "commit_marker_required_for_primary_and_fallback"
    ] is True
    assert protocol["output_contract"]["unmarked_output_directory_acceptance"] == (
        "REJECT_AS_PARTIAL_NOT_COMMITTED"
    )
    assert protocol["output_contract"]["post_commit_marker_validation_retry_count"] == 1
    assert protocol["output_contract"][
        "post_commit_marker_validation_failure_status"
    ] == "COMMITTED_NOT_ACCEPTED"
    assert protocol["output_contract"][
        "committed_not_accepted_failure_record_allowed"
    ] is False
    assert protocol["output_contract"][
        "committed_not_accepted_canonical_accepted"
    ] is False
    assert protocol["output_contract"]["success_stdout_includes_publication_mode"] is True
    assert protocol["output_contract"]["fallback_commit_marker_bindings"] == [
        "run_id",
        "execution_id",
        "SHA256SUMS_SHA256",
        "bundle_file_count",
        "bundle_filename_set_sha256",
        "final_output_directory_name_sha256",
        "final_output_target_sha256",
        "publication_mode",
        "committed_true",
    ]
    assert protocol["output_contract"][
        "fallback_commit_marker_validation_required_before_published_return"
    ] is True
    assert protocol["output_contract"][
        "fallback_directory_and_parent_fsync_after_commit_marker"
    ] is True
    assert protocol["output_contract"][
        "fallback_partial_directory_canonical_accepted"
    ] is False


def test_verified_inputs_ignore_stale_preopen_metadata_but_bind_fd_hash_and_bytes(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "stale_preopen_metadata.json"
    payload = b'{"stable":"verified descriptor bytes"}\n'
    source.write_bytes(payload)
    expected_sha256 = hashlib.sha256(payload).hexdigest()
    real_require_regular_file = QUAL._require_regular_file

    def stale_metadata_after_real_path_guards(
        path: Path,
        *,
        label: str,
        suffix: str | None = None,
    ) -> SimpleNamespace:
        info = real_require_regular_file(path, label=label, suffix=suffix)
        return SimpleNamespace(
            st_mode=info.st_mode,
            st_dev=info.st_dev + 101,
            st_ino=info.st_ino + 103,
            st_size=info.st_size + 107,
            st_mtime_ns=info.st_mtime_ns + 109,
        )

    monkeypatch.setattr(
        QUAL,
        "_require_regular_file",
        stale_metadata_after_real_path_guards,
    )
    observed, read_provenance = QUAL._read_verified_bytes(
        source,
        expected_sha256,
        expected_bytes=len(payload),
        label="stale pre-open JSON",
        suffix=".json",
    )
    hash_provenance = QUAL._verify_file_hash(
        source,
        expected_sha256,
        expected_bytes=len(payload),
        label="stale pre-open hash-only input",
        suffix=".json",
    )
    snapshot = tmp_path / "verified_snapshot.json"
    snapshot_provenance = QUAL._snapshot_verified_file(
        source,
        snapshot,
        expected_sha256,
        expected_bytes=len(payload),
        label="stale pre-open snapshot input",
        suffix=".json",
    )

    assert observed == payload
    assert snapshot.read_bytes() == payload
    assert read_provenance["sha256"] == expected_sha256
    assert hash_provenance == {
        "sha256": expected_sha256,
        "bytes": len(payload),
        "filename": source.name,
    }
    assert snapshot_provenance["sha256"] == expected_sha256
    assert snapshot.stat().st_mode & 0o777 == 0o400


def test_verified_input_open_rejects_symlink_and_nonregular_paths(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    payload = b"verified bytes\n"
    expected_sha256 = hashlib.sha256(payload).hexdigest()
    source = tmp_path / "regular.txt"
    source.write_bytes(payload)
    symlink = tmp_path / "linked.txt"
    symlink.symlink_to(source)
    with pytest.raises(QUAL.ScopeViolation, match="symlink component"):
        QUAL._read_verified_bytes(
            symlink,
            expected_sha256,
            label="symlink verified input",
        )

    nonregular = tmp_path / "directory.txt"
    nonregular.mkdir()
    with pytest.raises(QUAL.QualificationError, match="regular file"):
        QUAL._verify_file_hash(
            nonregular,
            expected_sha256,
            label="nonregular verified input",
        )

    def bypass_only_preopen_metadata_guard(
        _: Path,
        *,
        label: str,
        suffix: str | None = None,
    ) -> SimpleNamespace:
        del label, suffix
        return SimpleNamespace()

    monkeypatch.setattr(
        QUAL,
        "_require_regular_file",
        bypass_only_preopen_metadata_guard,
    )
    with pytest.raises(QUAL.QualificationError, match="without following symlinks"):
        QUAL._read_verified_bytes(
            symlink,
            expected_sha256,
            label="O_NOFOLLOW symlink verified input",
        )
    with pytest.raises(QUAL.QualificationError, match="opened descriptor is not a regular"):
        QUAL._verify_file_hash(
            nonregular,
            expected_sha256,
            label="fstat nonregular verified input",
        )


def test_verified_input_rejects_true_same_descriptor_mutation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "same_fd_mutation.bin"
    payload = b"same descriptor original bytes\n"
    source.write_bytes(payload)
    original_inode = source.stat().st_ino
    expected_sha256 = hashlib.sha256(payload).hexdigest()
    real_read = QUAL.os.read
    mutation_injected = False

    def read_then_mutate_same_inode(descriptor: int, size: int) -> bytes:
        nonlocal mutation_injected
        block = real_read(descriptor, size)
        if block and not mutation_injected:
            mutation_injected = True
            with source.open("ab", buffering=0) as handle:
                handle.write(b"same-fd-mutation\n")
                QUAL.os.fsync(handle.fileno())
        return block

    monkeypatch.setattr(QUAL.os, "read", read_then_mutate_same_inode)
    with pytest.raises(
        QUAL.QualificationError,
        match="changed while its verified bytes were captured",
    ):
        QUAL._read_verified_bytes(
            source,
            expected_sha256,
            expected_bytes=len(payload),
            label="mutated same descriptor input",
        )
    assert mutation_injected is True
    assert source.stat().st_ino == original_inode


def test_benjamini_hochberg_is_monotone_in_sorted_p_value_order() -> None:
    pvalues = [0.04, 0.001, 0.03, 0.2]
    adjusted = QUAL._benjamini_hochberg(pvalues)
    ordered = sorted(zip(pvalues, adjusted))
    assert all(left[1] <= right[1] for left, right in zip(ordered, ordered[1:]))
    assert adjusted == pytest.approx([0.05333333333333334, 0.004, 0.05333333333333334, 0.2])


def test_git_binding_verifies_authority_chain_blobs_and_clean_implementation(
    tmp_path: Path,
) -> None:
    fixture = _git_binding_fixture(tmp_path)
    result = QUAL._verify_git_binding(
        fixture["repo"],
        fixture["implementation"],
        accepted_a0_base_commit=fixture["accepted"],
        active_authority_commit=fixture["active"],
        authority_files=fixture["authority_files"],
        implementation_files=fixture["implementation_files"],
    )
    assert result["status"] == "PASS"
    assert result["accepted_a0_is_ancestor_of_active_authority"] is True
    assert result["active_authority_is_ancestor_of_implementation"] is True
    assert result["active_authority_file_hashes_match"] is True
    assert result["implementation_file_hashes_match"] is True
    assert result["worktree_clean"] is True


def test_git_binding_rejects_nonancestor_authority(tmp_path: Path) -> None:
    fixture = _git_binding_fixture(tmp_path)
    repo = fixture["repo"]
    _git(repo, "checkout", "--detach", fixture["accepted"])
    (repo / "sibling.txt").write_text("sibling authority\n", encoding="utf-8")
    sibling = _commit(repo, "sibling authority")
    _git(repo, "checkout", "--detach", fixture["implementation"])
    with pytest.raises(QUAL.QualificationError, match="active authority to implementation"):
        QUAL._verify_git_binding(
            repo,
            fixture["implementation"],
            accepted_a0_base_commit=fixture["accepted"],
            active_authority_commit=sibling,
            authority_files=fixture["authority_files"],
            implementation_files=fixture["implementation_files"],
        )


def test_git_binding_rejects_authority_and_implementation_blob_mismatch(
    tmp_path: Path,
) -> None:
    fixture = _git_binding_fixture(tmp_path)
    wrong_authority = (
        (fixture["authority_files"][0][0], "0" * 64),
        fixture["authority_files"][1],
    )
    with pytest.raises(QUAL.QualificationError, match="active authority file hash"):
        QUAL._verify_git_binding(
            fixture["repo"],
            fixture["implementation"],
            accepted_a0_base_commit=fixture["accepted"],
            active_authority_commit=fixture["active"],
            authority_files=wrong_authority,
            implementation_files=fixture["implementation_files"],
        )
    wrong_implementation = (
        (fixture["implementation_files"][0][0], "f" * 64),
        fixture["implementation_files"][1],
    )
    with pytest.raises(QUAL.QualificationError, match="implementation file hash"):
        QUAL._verify_git_binding(
            fixture["repo"],
            fixture["implementation"],
            accepted_a0_base_commit=fixture["accepted"],
            active_authority_commit=fixture["active"],
            authority_files=fixture["authority_files"],
            implementation_files=wrong_implementation,
        )


def test_git_binding_rejects_dirty_worktree(tmp_path: Path) -> None:
    fixture = _git_binding_fixture(tmp_path)
    fixture["qualifier"].write_text("print('dirty')\n", encoding="utf-8")
    with pytest.raises(QUAL.QualificationError, match="worktree is not clean"):
        QUAL._verify_git_binding(
            fixture["repo"],
            fixture["implementation"],
            accepted_a0_base_commit=fixture["accepted"],
            active_authority_commit=fixture["active"],
            authority_files=fixture["authority_files"],
            implementation_files=fixture["implementation_files"],
        )


def test_missing_component_is_explicitly_excluded_and_never_zero() -> None:
    ratios, reasons = QUAL._eligible_log2_ratios(
        members={"present", "missing"},
        numerator_counts={"present": 10},
        denominator_counts={"present": 5, "missing": 7},
        numerator_library_sum=100,
        denominator_library_sum=100,
        minimum_cpm=0.5,
    )
    assert ratios == pytest.approx([1.0])
    assert reasons == {"NUMERATOR_KEY_MISSING_NOT_ZERO": 1}


def test_paper_native_and_route_a_companion_are_separate() -> None:
    protocol = json.loads(PROTOCOL_PATH.read_text(encoding="utf-8"))
    pair = {
        "pair_id": "PAIR_1",
        "biological_source_group_id": "GROUP_1",
        "source_members": set(SOURCE_KEYS),
        "candidate_members": set(CANDIDATE_KEYS),
    }
    counts: dict[tuple[str, str, int], dict[str, int]] = {}
    library_sums: dict[tuple[str, str, int], int] = {}
    for context in QUAL.CONTEXTS:
        for assay in QUAL.ASSAYS:
            for replicate in QUAL.REPLICATES:
                slot = (context, assay, replicate)
                counts[slot] = {
                    key: _raw_count_for(
                        assay,
                        key,
                        replicate,
                        context=context,
                    )
                    for key in ALL_KEYS
                }
                library_sums[slot] = sum(counts[slot].values())
    paper, replicate_rows, effects, rejections = QUAL._build_paper_and_companion_results(
        pairs=[pair],
        counts=counts,
        library_sums=library_sums,
        protocol=protocol,
    )
    assert len(replicate_rows) == 12
    assert len(effects) == 4
    assert len(set(library_sums.values())) == 18
    assert all(effect["biological_replicate_count"] == 3 for effect in effects)
    assert all(effect["standard_error"] is not None for effect in effects)
    assert all(
        effect["record_type"] == "ROUTE_A_COMPANION_AGGREGATE_EFFECT"
        and effect["paper_native_row"] is False
        and len(effect["biological_replicate_deltas"]) == 3
        and effect["candidate_minus_source_delta"]
        == pytest.approx(statistics.fmean(effect["biological_replicate_deltas"]))
        and effect["standard_error"]
        == pytest.approx(
            statistics.stdev(effect["biological_replicate_deltas"]) / math.sqrt(3.0)
        )
        for effect in effects
    )
    effects_by_slot = {
        (effect["context"], effect["endpoint_id"]): effect for effect in effects
    }
    assert effects_by_slot[("PC3", "transcript_log2_totalrna_over_dna")][
        "biological_replicate_deltas"
    ] != effects_by_slot[("293T", "transcript_log2_totalrna_over_dna")][
        "biological_replicate_deltas"
    ]
    assert paper["computed_pair_context_endpoint_count"] == 4
    assert paper["paper_method_reproduced"] is False
    assert paper["method_evidence"]["method_source_status"] == "UNKNOWN_NOT_ASSERTED"
    assert all(row["significance_used_for_membership"] is False for row in paper["rows"])
    assert all(row["technical_members_are_independent_n"] is False for row in paper["rows"])
    assert not rejections


def test_power_simulation_uses_prefrozen_group_unit_and_is_deterministic() -> None:
    protocol = json.loads(PROTOCOL_PATH.read_text(encoding="utf-8"))
    first = QUAL._run_group_power_simulation(200, protocol)
    second = QUAL._run_group_power_simulation(200, protocol)
    assert first == second
    assert first["analysis_unit"] == "BIOLOGICAL_SOURCE_GROUP"
    assert first["post_dedup_group_count"] == 200
    assert first["simulation_trials"] == 1000
    assert first["bootstrap_resamples"] == 2000
    assert first["seed"] == 20260810
    assert first["status"] in {"PASS", "FAIL_CURRENT_PROTOCOL"}


def test_strict_outcome_blind_mapping_reuses_v4_pair_semantics(tmp_path: Path) -> None:
    moesm8 = tmp_path / "41467_2021_24445_MOESM8_ESM.xlsx"
    mapping = tmp_path / "Lim_et_al_Supp_Tbl_6c_293T.xlsx"
    _write_moesm8(moesm8)
    _write_mapping_workbook(mapping)
    helper = QUAL._load_v4_helpers(V4_PATH)
    sequence_asset = {
        "sha256": _sha256(moesm8),
        "workbook_contract": {
            "sequence_sheet": V4.MOESM8_SEQUENCE_SHEET,
            "sequence_header_sha256": V4.MOESM8_6A_RAW_HEADER_SHA256,
        },
    }
    mapping_asset = {
        "sha256": _sha256(mapping),
        "workbook_contract": {
            "mapping_sheet": V4.LIM6C_SHEET,
            "mapping_header_sha256": V4.LIM6C_RAW_HEADER_SHA256,
            "description_column_index": 0,
            "member_key_column_index": 1,
        },
    }
    entries, sequence_audit, sequence_counts = QUAL._load_sequence_universe(
        path=moesm8,
        effective_asset=sequence_asset,
        helper=helper,
    )
    pairs, mapping_audit, pair_audit, reasons, members = QUAL._load_outcome_blind_member_mapping(
        path=mapping,
        effective_asset=mapping_asset,
        helper=helper,
        sequence_entries=entries,
        mapping_evidence_status="PASS",
    )
    assert len(pairs) == 1
    assert pairs[0]["source_sequence"] == SOURCE_SEQUENCE
    assert pairs[0]["candidate_sequence"] == CANDIDATE_SEQUENCE
    assert pairs[0]["sequence_index_0_based"] == 2
    assert len(members) == 40
    assert not reasons
    assert sequence_counts == {
        "INCLUDED_STRICT_SNV_MUTANT_COORDINATE": 1,
        "REFERENCE_ONLY_STRICT_WT_COORDINATE": 1,
    }
    safe_output = json.dumps([*sequence_audit, *mapping_audit, *pair_audit], sort_keys=True)
    for key in ALL_KEYS:
        assert key not in safe_output
    assert SOURCE_SEQUENCE not in safe_output
    assert CANDIDATE_SEQUENCE not in safe_output


def test_raw_grid_allows_cross_context_missingness_without_imputation(tmp_path: Path) -> None:
    assets: list[dict[str, Any]] = []
    paths: dict[str, Path] = {}
    pc3_keys = ("key_a", "key_b")
    t293_keys = ("key_b", "key_c")
    for context, keys in (("PC3", pc3_keys), ("293T", t293_keys)):
        for assay in QUAL.ASSAYS:
            for replicate in QUAL.REPLICATES:
                asset_id = f"{context}_{assay}_{replicate}"
                path = tmp_path / f"{asset_id}.txt.gz"
                _write_raw_table(
                    path,
                    assay=assay,
                    replicate=replicate,
                    context=context,
                    keys=keys,
                )
                assets.append(
                    {
                        "asset_id": asset_id,
                        "asset_kind": "GEO_RAW_COUNT",
                        "context": context,
                        "assay": assay,
                        "biological_replicate": replicate,
                        "sha256": _sha256(path),
                    }
                )
                paths[asset_id] = path
    reconciliation_contract = json.loads(PROTOCOL_PATH.read_text(encoding="utf-8"))[
        "mapping"
    ]["raw_key_reconciliation"]
    _, _, audits, cross, reconciliation = QUAL._load_geo_count_grid(
        resolved_assets=assets,
        asset_paths=paths,
        mapped_members={"key_a", "key_b", "key_c"},
        reconciliation_contract=reconciliation_contract,
    )
    assert len(audits) == 18
    assert cross == {
        "pc3_unique_key_count": 2,
        "293t_unique_key_count": 2,
        "intersection_key_count": 1,
        "pc3_only_key_count": 1,
        "293t_only_key_count": 1,
        "union_key_count": 3,
        "missing_is_zero": False,
        "cross_context_missingness_allowed": True,
        "raw_keys_emitted": False,
    }
    assert reconciliation["status"] == "PASS"
    assert reconciliation["unclassified_raw_key_count"] == 0
    assert reconciliation["mapping_key_absent_from_both_contexts_count"] == 0

    _, _, _, _, absent_mapping = QUAL._load_geo_count_grid(
        resolved_assets=assets,
        asset_paths=paths,
        mapped_members={"key_a", "key_b", "key_c", "key_absent_from_both"},
        reconciliation_contract=reconciliation_contract,
    )
    assert absent_mapping["status"] == "BLOCKED_OUTCOME_BLIND_KEY_RECONCILIATION"
    assert absent_mapping["mapping_key_absent_from_both_contexts_count"] == 1

    exclusion_contract = json.loads(json.dumps(reconciliation_contract))
    exclusion_contract["prefrozen_excluded_key_sha256_by_class"]["OUT_OF_SCOPE"] = [
        QUAL._raw_key_classification_sha256("key_a"),
        QUAL._raw_key_classification_sha256("key_c"),
    ]
    _, _, _, _, explicitly_excluded = QUAL._load_geo_count_grid(
        resolved_assets=assets,
        asset_paths=paths,
        mapped_members={"key_b"},
        reconciliation_contract=exclusion_contract,
    )
    assert explicitly_excluded["status"] == "PASS"
    assert explicitly_excluded["unclassified_raw_key_count"] == 0
    assert explicitly_excluded["observed_prefrozen_exclusion_hash_count"] == 2


def test_canonical_v3_builder_materializes_only_qualified_records() -> None:
    protocol = json.loads(PROTOCOL_PATH.read_text(encoding="utf-8"))
    protocol["license_and_redistribution"].update(
        {
            "audit_status": "PASS",
            "license_id": "CC_BY_4_0_FIXTURE",
            "license_name": "CC BY 4.0",
            "license_uri": "https://creativecommons.org/licenses/by/4.0/",
            "verified_at": "2026-08-10T00:00:00+08:00",
        }
    )
    protocol["foundation_exposure"].update(
        {
            "audit_status": "PASS",
            "checkpoint_id": "FIXTURE_CHECKPOINT",
            "checkpoint_sha256": "a" * 64,
            "audit_id": "FIXTURE_EXPOSURE_AUDIT",
        }
    )
    pair = {
        "pair_id": "PAIR_FIXTURE",
        "biological_source_group_id": "GROUP_FIXTURE",
        "source_id": "SOURCE_FIXTURE",
        "candidate_id": "CANDIDATE_FIXTURE",
        "gene_group_id": "GENE_FIXTURE",
        "locus_id": "LOCUS_FIXTURE",
        "source_sequence": SOURCE_SEQUENCE,
        "candidate_sequence": CANDIDATE_SEQUENCE,
        "source_sequence_sha256": hashlib.sha256(SOURCE_SEQUENCE.encode()).hexdigest(),
        "candidate_sequence_sha256": hashlib.sha256(CANDIDATE_SEQUENCE.encode()).hexdigest(),
        "sequence_index_0_based": 2,
        "ref": "C",
        "alt": "T",
    }
    effects = [
        {
            "pair_id": pair["pair_id"],
            "context": context,
            "endpoint_id": endpoint["endpoint_id"],
            "candidate_minus_source_delta": 1.5,
            "standard_error": 0.2,
            "eligible": True,
        }
        for context in QUAL.CONTEXTS
        for endpoint in QUAL.ENDPOINTS
    ]
    license_exposure = {
        "canonical_license_status": "VERIFIED_PUBLIC",
        "redistribution_allowed": True,
    }
    schema = json.loads(CANONICAL_SCHEMA_PATH.read_text(encoding="utf-8"))
    records = QUAL._build_canonical_records(
        closed_pairs=[pair],
        effects=effects,
        effective_manifest_sha256="b" * 64,
        raw_bundle_sha256="c" * 64,
        protocol_sha256="d" * 64,
        script_sha256="e" * 64,
        protocol=protocol,
        schema=schema,
        license_exposure=license_exposure,
    )
    assert len(records) == 4
    assert all(record["data_role"] == "ORDINARY_A1_INTERVENTION" for record in records)
    assert all(record["eligibility"]["status"] == "QUALIFIED" for record in records)
    assert all(record["split"]["partition"] == "DEVELOPMENT" for record in records)
    assert all(record["replicate"]["replicate_count"] == 3 for record in records)
    with pytest.raises(
        QUAL.QualificationError,
        match="requires verified public redistribution",
    ):
        QUAL._build_canonical_records(
            closed_pairs=[pair],
            effects=effects,
            effective_manifest_sha256="b" * 64,
            raw_bundle_sha256="c" * 64,
            protocol_sha256="d" * 64,
            script_sha256="e" * 64,
            protocol=protocol,
            schema=schema,
            license_exposure={
                "canonical_license_status": "VERIFIED_NONREDISTRIBUTABLE",
                "redistribution_allowed": False,
            },
        )


def test_nonredistributable_is_the_only_blocker_for_full_sequence_canonical() -> None:
    protocol = json.loads(PROTOCOL_PATH.read_text(encoding="utf-8"))
    protocol["mapping"]["outcome_blind_mapping_evidence_status"] = "PASS"
    blockers = QUAL._qualification_blockers(
        protocol=protocol,
        git_binding={"status": "PASS"},
        sequence_class_counts={"REFERENCE_ONLY_STRICT_WT_COORDINATE": 1},
        mapping_reasons={},
        pair_count=1,
        closed_pair_count=1,
        paper_report={
            "method_evidence": {
                "method_source_status": "PASS",
                "multiple_testing_family_status": "PASS",
                "published_result_crosscheck_status": "PASS",
            },
            "paper_method_reproduced": True,
        },
        leakage_report={"leakage_audit_status": "PASS"},
        power_report={"status": "PASS"},
        license_exposure={
            "license_audit_status": "PASS",
            "canonical_sequence_materialization_allowed": False,
            "foundation_exposure_audit_status": "PASS",
        },
        raw_key_reconciliation={
            "unclassified_raw_key_count": 0,
            "mapped_and_explicit_exclusion_overlap_count": 0,
            "mapping_key_absent_from_both_contexts_count": 0,
        },
    )
    assert blockers == ["CANONICAL_SEQUENCE_REDISTRIBUTION_NOT_ALLOWED"]


def test_full_synthetic_run_uses_verified_snapshots_and_publishes_zero_canonical(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    repo, data, protocol, asset_manifest = _fixture_repo_and_data(tmp_path)
    output = tmp_path / "blocked_bundle"
    raw_path = next(data.glob("*.txt.gz"))
    real_require_regular_file = QUAL._require_regular_file
    p0_stale_preopen_metadata_seen = False

    def stale_only_for_p0_manifest_after_path_guards(
        path: Path,
        *,
        label: str,
        suffix: str | None = None,
    ) -> Any:
        nonlocal p0_stale_preopen_metadata_seen
        info = real_require_regular_file(path, label=label, suffix=suffix)
        if label != "P0 source manifest":
            return info
        p0_stale_preopen_metadata_seen = True
        return SimpleNamespace(
            st_mode=info.st_mode,
            st_dev=info.st_dev + 211,
            st_ino=info.st_ino + 223,
            st_size=info.st_size + 227,
            st_mtime_ns=info.st_mtime_ns + 229,
        )

    def replace_original_paths_after_snapshot() -> None:
        protocol.write_bytes(b"{}\n")
        asset_manifest.write_bytes(b"{}\n")
        (data / "manifest.json").write_bytes(b"{}\n")
        (repo / "schemas/route_a_v3/canonical_intervention_record.schema.json").write_bytes(
            b"{}\n"
        )
        (repo / "scripts/route_a_v3/reconstruct_gse149487_plumage.py").write_bytes(
            b"raise RuntimeError('original helper path was reopened')\n"
        )
        (data / "41467_2021_24445_MOESM8_ESM.xlsx").write_bytes(b"not an xlsx")
        (data / "Lim_et_al_Supp_Tbl_6c_293T.xlsx").write_bytes(b"not an xlsx")
        raw_path.write_bytes(b"not a gzip")

    monkeypatch.setattr(
        QUAL,
        "_POST_VERIFIED_INPUT_SNAPSHOT_HOOK",
        replace_original_paths_after_snapshot,
    )
    monkeypatch.setattr(
        QUAL,
        "_require_regular_file",
        stale_only_for_p0_manifest_after_path_guards,
    )
    expected_protocol_sha256 = _sha256(protocol)
    report = QUAL.qualify_gse149487_plumage(
        repo_root=repo,
        data_root=data,
        protocol_path=protocol,
        asset_manifest_path=asset_manifest,
        expected_protocol_sha256=expected_protocol_sha256,
        output_directory=output,
        run_id="A1_FIXTURE_RUN",
        execution_id="GSE149487_FIXTURE_EXECUTION",
        recorded_at="2026-08-10T12:00:00+08:00",
    )
    assert report["qualification_status"] == "BLOCKED_PENDING_PUBLIC_EVIDENCE"
    assert report["qualified"] is False
    assert report["canonical_record_count"] == 0
    assert report["candidate_canonical_record_count_before_gate"] == 4
    assert p0_stale_preopen_metadata_seen is True
    assert report["publication"]["status"] == "PUBLISHED_DURABLE"
    assert report["publication"]["publication_mode"] == QUAL.PRIMARY_PUBLICATION_MODE
    assert report["publication"]["terminal_commit_marker_validated"] is True
    assert report["publication"]["terminal_commit_marker_postcommit_revalidated"] is True
    assert report["authority"]["input_snapshot_lineage"] == {
        "small_json_mode": QUAL.VERIFIED_JSON_BYTES_MODE,
        "file_parser_mode": QUAL.PRIVATE_READ_ONLY_SNAPSHOT_MODE,
        "verified_json_documents": [
            "qualification_protocol",
            "asset_manifest",
            "canonical_schema",
            "p0_source_manifest",
        ],
        "private_read_only_asset_snapshot_count": 21,
        "v4_helper_loaded_from_private_read_only_snapshot": True,
        "scientific_parser_original_path_reopen_after_verification": False,
        "preopen_lstat_dev_ino_size_mtime_acceptance_authority": False,
        "verified_input_acceptance_authority": (
            "O_NOFOLLOW_FSTAT_BEFORE_AFTER_SHA256_AND_BYTES"
        ),
        "same_descriptor_identity_and_size_must_remain_stable": True,
        "snapshot_lifetime": "QUALIFICATION_PAYLOAD_BUILD_ONLY",
    }
    assert not (output / "canonical_intervention_records.jsonl").exists()
    assert {
        path.name for path in output.iterdir()
    } == set(QUAL.ALWAYS_OUTPUT_FILES) | {QUAL.PUBLICATION_COMMIT_FILENAME}
    assert "OUTCOME_BLIND_LONG_READ_MAPPING_PROVENANCE_UNKNOWN_NOT_ASSERTED" in report["blockers"]
    assert "PAPER_NATIVE_METHOD_SOURCE_UNKNOWN_NOT_ASSERTED" in report["blockers"]
    assert "LICENSE_AND_REDISTRIBUTION_UNKNOWN_NOT_ASSERTED" in report["blockers"]
    assert "CHECKPOINT_SPECIFIC_FOUNDATION_EXPOSURE_UNKNOWN_NOT_ASSERTED" in report["blockers"]
    assert report["raw_grid"]["outcome_blind_key_reconciliation"]["status"] == "PASS"
    companion_rows = [
        json.loads(line)
        for line in (output / "replicate_effect_summaries.jsonl")
        .read_text(encoding="utf-8")
        .splitlines()
    ]
    assert len(companion_rows) == 4
    assert all(
        row["record_type"] == "ROUTE_A_COMPANION_AGGREGATE_EFFECT"
        and row["paper_native_row"] is False
        and len(row["biological_replicate_deltas"]) == 3
        and row["candidate_minus_source_delta"]
        == pytest.approx(statistics.fmean(row["biological_replicate_deltas"]))
        for row in companion_rows
    )
    all_output = b"".join(path.read_bytes() for path in output.iterdir())
    for key in ALL_KEYS:
        assert key.encode() not in all_output
    assert SOURCE_SEQUENCE.encode() not in all_output
    assert CANDIDATE_SEQUENCE.encode() not in all_output
    sums = {
        line.split("  ", 1)[1]: line.split("  ", 1)[0]
        for line in (output / "SHA256SUMS").read_text(encoding="ascii").splitlines()
    }
    assert sums == {
        path.name: _sha256(path)
        for path in output.iterdir()
        if path.name not in {"SHA256SUMS", QUAL.PUBLICATION_COMMIT_FILENAME}
    }
    with pytest.raises(QUAL.QualificationError, match="overwrite"):
        QUAL.qualify_gse149487_plumage(
            repo_root=repo,
            data_root=data,
            protocol_path=protocol,
            asset_manifest_path=asset_manifest,
            expected_protocol_sha256="a" * 64,
            output_directory=output,
            run_id="A1_FIXTURE_RUN",
            execution_id="GSE149487_FIXTURE_EXECUTION_2",
            recorded_at="2026-08-10T12:01:00+08:00",
        )


def test_atomic_no_replace_publication_does_not_replace_racing_empty_directory(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "racing_output"
    original_rename = QUAL._rename_directory_noreplace
    competing_inode: dict[str, int] = {}

    def create_competing_target_then_rename(source: Path, destination: Path) -> None:
        destination.mkdir()
        competing_inode["value"] = destination.stat().st_ino
        original_rename(source, destination)

    monkeypatch.setattr(
        QUAL,
        "_rename_directory_noreplace",
        create_competing_target_then_rename,
    )
    with pytest.raises(QUAL.QualificationError, match="no-replace publication"):
        QUAL._publish_bundle(
            output,
            _minimal_bundle_payloads(),
            run_metadata=_publication_run_metadata("PRIMARY_RACE"),
        )
    assert output.is_dir()
    assert output.stat().st_ino == competing_inode["value"]
    assert list(output.iterdir()) == []


def test_primary_publication_marker_is_last_hash_bound_and_required_offline(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "primary_marker_bundle"
    run_metadata = _publication_run_metadata("PRIMARY_MARKER")
    staging_write_order: list[str] = []
    real_write_exclusive = QUAL._write_exclusive

    def record_staging_writes(path: Path, payload: bytes) -> None:
        if path.parent.name.startswith(f".{output.name}.partial-staging-"):
            staging_write_order.append(path.name)
        real_write_exclusive(path, payload)

    monkeypatch.setattr(QUAL, "_write_exclusive", record_staging_writes)
    publication = QUAL._publish_bundle(
        output,
        _minimal_bundle_payloads(),
        run_metadata=run_metadata,
    )

    assert publication["publication_mode"] == QUAL.PRIMARY_PUBLICATION_MODE
    assert publication["terminal_commit_marker_validated"] is True
    assert publication["terminal_commit_marker_postcommit_revalidated"] is True
    assert staging_write_order[-1] == QUAL.PUBLICATION_COMMIT_FILENAME
    assert staging_write_order.count(QUAL.PUBLICATION_COMMIT_FILENAME) == 1
    assert set(staging_write_order[:-1]) == set(QUAL.ALWAYS_OUTPUT_FILES)
    marker = QUAL._validate_publication_commit(
        output,
        expected_run_metadata=run_metadata,
        expected_publication_mode=QUAL.PRIMARY_PUBLICATION_MODE,
    )
    assert marker["publication_mode"] == QUAL.PRIMARY_PUBLICATION_MODE
    assert marker["sha256sums_sha256"] == _sha256(output / "SHA256SUMS")
    (output / QUAL.PUBLICATION_COMMIT_FILENAME).unlink()
    with pytest.raises(QUAL.QualificationError, match="publication commit marker"):
        QUAL._validate_publication_commit(
            output,
            expected_run_metadata=run_metadata,
            expected_publication_mode=QUAL.PRIMARY_PUBLICATION_MODE,
        )


def test_primary_pre_rename_staging_marker_is_never_consumer_accepted(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "primary_pre_rename_crash_target"
    run_metadata = _publication_run_metadata("PRIMARY_PRE_RENAME_CRASH")
    captured_staging: list[Path] = []

    def stop_after_staging_marker(source: Path, _: Path) -> list[str]:
        captured_staging.append(source)
        assert (source / QUAL.PUBLICATION_COMMIT_FILENAME).is_file()
        raise QUAL.QualificationError("injected stop before primary rename")

    monkeypatch.setattr(QUAL, "_rename_directory_noreplace", stop_after_staging_marker)
    with pytest.raises(QUAL.QualificationError, match="stop before primary rename"):
        QUAL._publish_bundle(
            output,
            _minimal_bundle_payloads(),
            run_metadata=run_metadata,
        )
    assert not output.exists()
    assert len(captured_staging) == 1
    assert captured_staging[0].is_dir()
    with pytest.raises(QUAL.QualificationError, match="directory-name hash mismatch"):
        QUAL._validate_publication_commit(
            captured_staging[0],
            expected_run_metadata=run_metadata,
            expected_publication_mode=QUAL.PRIMARY_PUBLICATION_MODE,
        )


def test_primary_post_commit_marker_loss_is_committed_not_accepted(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "primary_marker_lost_after_commit"
    failure = tmp_path / "PRIMARY_MARKER_LOST_FAILURE.json"
    real_rename = QUAL._rename_directory_noreplace

    def synthetic_payload_builder(**_: Any) -> tuple[dict[str, bytes], dict[str, Any]]:
        return _minimal_bundle_payloads(), {
            "qualification_status": "BLOCKED_PENDING_PUBLIC_EVIDENCE",
            "qualified": False,
            "canonical_record_count": 0,
        }

    def rename_then_remove_marker(source: Path, destination: Path) -> list[str]:
        warnings = real_rename(source, destination)
        (destination / QUAL.PUBLICATION_COMMIT_FILENAME).unlink()
        return warnings

    monkeypatch.setattr(QUAL, "_build_qualification_payloads", synthetic_payload_builder)
    monkeypatch.setattr(QUAL, "_rename_directory_noreplace", rename_then_remove_marker)
    result = QUAL.execute_qualification(
        repo_root=tmp_path / "repo",
        data_root=tmp_path / "data",
        protocol_path=tmp_path / "protocol.json",
        asset_manifest_path=tmp_path / "assets.json",
        expected_protocol_sha256="a" * 64,
        output_directory=output,
        failure_record_path=failure,
        run_id="A1_PRIMARY_MARKER_LOST",
        execution_id="GSE149487_PRIMARY_MARKER_LOST",
        recorded_at="2026-08-10T12:17:00+08:00",
    )

    assert result["kind"] == "COMMITTED_NOT_ACCEPTED"
    committed = result["committed_not_accepted"]
    assert committed["status"] == "COMMITTED_NOT_ACCEPTED"
    assert committed["publication_mode"] == QUAL.PRIMARY_PUBLICATION_MODE
    assert committed["directory_committed"] is True
    assert committed["publication_accepted"] is False
    assert committed["terminal_commit_marker_validated"] is False
    assert committed["qualified"] is False
    assert committed["canonical_accepted"] is False
    assert committed["canonical_record_count"] == 0
    assert committed["failure_record_materialized"] is False
    assert output.is_dir()
    assert not (output / QUAL.PUBLICATION_COMMIT_FILENAME).exists()
    assert not failure.exists()
    assert not QUAL._transaction_claim_path(output).exists()


def test_primary_transient_final_marker_read_failure_retries_then_publishes_warning(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "primary_marker_transient_read"
    failure = tmp_path / "PRIMARY_MARKER_TRANSIENT_FAILURE.json"
    real_validate = QUAL._validate_publication_commit
    injected_count = 0

    def synthetic_payload_builder(**_: Any) -> tuple[dict[str, bytes], dict[str, Any]]:
        return _minimal_bundle_payloads(), {
            "qualification_status": "BLOCKED_PENDING_PUBLIC_EVIDENCE",
            "qualified": False,
            "canonical_record_count": 0,
        }

    def fail_first_final_validation(
        directory: Path,
        **kwargs: Any,
    ) -> dict[str, Any]:
        nonlocal injected_count
        if directory == output and injected_count == 0:
            injected_count += 1
            raise QUAL.QualificationError("injected transient final marker read failure")
        return real_validate(directory, **kwargs)

    monkeypatch.setattr(QUAL, "_build_qualification_payloads", synthetic_payload_builder)
    monkeypatch.setattr(QUAL, "_validate_publication_commit", fail_first_final_validation)
    result = QUAL.execute_qualification(
        repo_root=tmp_path / "repo",
        data_root=tmp_path / "data",
        protocol_path=tmp_path / "protocol.json",
        asset_manifest_path=tmp_path / "assets.json",
        expected_protocol_sha256="a" * 64,
        output_directory=output,
        failure_record_path=failure,
        run_id="A1_PRIMARY_MARKER_TRANSIENT",
        execution_id="GSE149487_PRIMARY_MARKER_TRANSIENT",
        recorded_at="2026-08-10T12:18:00+08:00",
    )

    assert injected_count == 1
    assert result["kind"] == "BUNDLE"
    publication = result["report"]["publication"]
    assert publication["status"] == "PUBLISHED_WITH_POST_COMMIT_DURABILITY_WARNING"
    assert publication["terminal_commit_marker_validated"] is True
    assert publication["terminal_commit_marker_postcommit_revalidated"] is True
    assert "POST_COMMIT_TERMINAL_MARKER_REVALIDATION_RETRY_REQUIRED" in publication[
        "durability_warning_codes"
    ]
    assert QUAL._validate_publication_commit(
        output,
        expected_run_metadata={
            "run_id": "A1_PRIMARY_MARKER_TRANSIENT",
            "execution_id": "GSE149487_PRIMARY_MARKER_TRANSIENT",
            "recorded_at": "2026-08-10T12:18:00+08:00",
        },
        expected_publication_mode=QUAL.PRIMARY_PUBLICATION_MODE,
    )["committed"] is True
    assert not failure.exists()


def test_unsupported_atomic_rename_uses_validated_terminal_marker_written_last(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "fallback_bundle"
    run_metadata = _publication_run_metadata("FALLBACK_SUCCESS")
    unsupported_errno = getattr(errno, "EOPNOTSUPP", errno.ENOTSUP)
    write_order: list[str] = []
    real_write_exclusive = QUAL._write_exclusive

    def unsupported_atomic_rename(_: Path, __: Path) -> None:
        raise QUAL.AtomicNoReplaceUnsupported(unsupported_errno)

    def record_final_directory_writes(path: Path, payload: bytes) -> None:
        if path.parent == output:
            write_order.append(path.name)
        real_write_exclusive(path, payload)

    monkeypatch.setattr(QUAL, "_rename_directory_noreplace", unsupported_atomic_rename)
    monkeypatch.setattr(QUAL, "_write_exclusive", record_final_directory_writes)
    publication = QUAL._publish_bundle(
        output,
        _minimal_bundle_payloads(),
        run_metadata=run_metadata,
    )

    assert publication["status"] == "PUBLISHED_DURABLE"
    assert publication["publication_mode"] == QUAL.FALLBACK_PUBLICATION_MODE
    assert publication["fallback_trigger_errno"] == unsupported_errno
    assert publication["terminal_commit_marker_validated"] is True
    assert publication["source_staging_preserved"] is True
    assert write_order[-1] == QUAL.PUBLICATION_COMMIT_FILENAME
    assert write_order.count(QUAL.PUBLICATION_COMMIT_FILENAME) == 1
    assert set(write_order[:-1]) == set(QUAL.ALWAYS_OUTPUT_FILES)
    assert {path.name for path in output.iterdir()} == (
        set(QUAL.ALWAYS_OUTPUT_FILES) | {QUAL.PUBLICATION_COMMIT_FILENAME}
    )
    marker = QUAL._validate_publication_commit(
        output,
        expected_run_metadata=run_metadata,
    )
    assert marker["committed"] is True
    assert marker["publication_mode"] == QUAL.FALLBACK_PUBLICATION_MODE
    assert marker["bundle_file_count_excluding_commit_marker"] == len(
        QUAL.ALWAYS_OUTPUT_FILES
    )
    assert marker["sha256sums_sha256"] == _sha256(output / "SHA256SUMS")
    preserved_primary_staging = [
        path
        for path in tmp_path.iterdir()
        if path.name.startswith(f".{output.name}.partial-staging-")
    ]
    assert len(preserved_primary_staging) == 1
    with pytest.raises(QUAL.QualificationError, match="directory-name hash mismatch"):
        QUAL._validate_publication_commit(
            preserved_primary_staging[0],
            expected_run_metadata=run_metadata,
            expected_publication_mode=QUAL.PRIMARY_PUBLICATION_MODE,
        )


def test_fallback_mid_write_is_partial_without_marker_or_failure_record(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "fallback_partial"
    failure = tmp_path / "FALLBACK_PARTIAL_FAILURE.json"
    real_write_exclusive = QUAL._write_exclusive

    def synthetic_payload_builder(**_: Any) -> tuple[dict[str, bytes], dict[str, Any]]:
        return _minimal_bundle_payloads(), {
            "qualification_status": "BLOCKED_PENDING_PUBLIC_EVIDENCE",
            "qualified": False,
            "canonical_record_count": 0,
        }

    def unsupported_atomic_rename(_: Path, __: Path) -> None:
        raise QUAL.AtomicNoReplaceUnsupported(errno.EINVAL)

    def fail_during_final_directory_write(path: Path, payload: bytes) -> None:
        if path.parent == output and path.name == "GROUP_POWER_AUDIT.json":
            raise OSError("injected fallback member write failure")
        real_write_exclusive(path, payload)

    monkeypatch.setattr(QUAL, "_build_qualification_payloads", synthetic_payload_builder)
    monkeypatch.setattr(QUAL, "_rename_directory_noreplace", unsupported_atomic_rename)
    monkeypatch.setattr(QUAL, "_write_exclusive", fail_during_final_directory_write)
    result = QUAL.execute_qualification(
        repo_root=tmp_path / "repo",
        data_root=tmp_path / "data",
        protocol_path=tmp_path / "protocol.json",
        asset_manifest_path=tmp_path / "assets.json",
        expected_protocol_sha256="a" * 64,
        output_directory=output,
        failure_record_path=failure,
        run_id="A1_FALLBACK_PARTIAL",
        execution_id="GSE149487_FALLBACK_PARTIAL",
        recorded_at="2026-08-10T12:08:00+08:00",
    )

    assert result["kind"] == "PARTIAL_NOT_COMMITTED"
    assert result["partial"]["status"] == "PARTIAL_NOT_COMMITTED"
    assert result["partial"]["publication_mode"] == QUAL.FALLBACK_PUBLICATION_MODE
    assert result["partial"]["terminal_commit_marker_present"] is False
    assert result["partial"]["terminal_commit_marker_validated"] is False
    assert result["partial"]["canonical_accepted"] is False
    assert result["partial"]["canonical_record_count"] == 0
    assert result["partial"]["failure_record_materialized"] is False
    assert result["partial"]["retry_requires_new_run_id"] is True
    assert output.is_dir()
    assert any(output.iterdir())
    assert not (output / QUAL.PUBLICATION_COMMIT_FILENAME).exists()
    assert not failure.exists()
    assert not QUAL._transaction_claim_path(output).exists()


def test_fallback_concurrent_final_mkdir_loser_is_contended_without_failure(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "fallback_contended"
    failure = tmp_path / "FALLBACK_CONTENDED_FAILURE.json"

    def synthetic_payload_builder(**_: Any) -> tuple[dict[str, bytes], dict[str, Any]]:
        return _minimal_bundle_payloads(), {
            "qualification_status": "BLOCKED_PENDING_PUBLIC_EVIDENCE",
            "qualified": False,
            "canonical_record_count": 0,
        }

    def competitor_then_unsupported(_: Path, destination: Path) -> None:
        destination.mkdir()
        (destination / "COMPETING_OWNER").write_bytes(b"winner\n")
        raise QUAL.AtomicNoReplaceUnsupported(errno.ENOSYS)

    monkeypatch.setattr(QUAL, "_build_qualification_payloads", synthetic_payload_builder)
    monkeypatch.setattr(QUAL, "_rename_directory_noreplace", competitor_then_unsupported)
    result = QUAL.execute_qualification(
        repo_root=tmp_path / "repo",
        data_root=tmp_path / "data",
        protocol_path=tmp_path / "protocol.json",
        asset_manifest_path=tmp_path / "assets.json",
        expected_protocol_sha256="a" * 64,
        output_directory=output,
        failure_record_path=failure,
        run_id="A1_FALLBACK_CONTENDED",
        execution_id="GSE149487_FALLBACK_CONTENDED",
        recorded_at="2026-08-10T12:09:00+08:00",
    )

    assert result["kind"] == "CONTENDED"
    assert result["contention"]["status"] == "OUTPUT_PUBLICATION_CONTENDED"
    assert result["contention"]["failure_record_materialized"] is False
    assert (output / "COMPETING_OWNER").read_bytes() == b"winner\n"
    assert not (output / QUAL.PUBLICATION_COMMIT_FILENAME).exists()
    assert not failure.exists()


def test_fallback_commit_marker_hash_mismatch_is_rejected(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "fallback_hash_mismatch"
    run_metadata = _publication_run_metadata("FALLBACK_HASH_MISMATCH")

    def unsupported_atomic_rename(_: Path, __: Path) -> None:
        raise QUAL.AtomicNoReplaceUnsupported(errno.ENOTSUP)

    monkeypatch.setattr(QUAL, "_rename_directory_noreplace", unsupported_atomic_rename)
    publication = QUAL._publish_bundle(
        output,
        _minimal_bundle_payloads(),
        run_metadata=run_metadata,
    )
    assert publication["terminal_commit_marker_validated"] is True
    sums_path = output / "SHA256SUMS"
    sums_path.write_bytes(sums_path.read_bytes() + b"tampered\n")
    with pytest.raises(QUAL.QualificationError, match="SHA256SUMS hash mismatch"):
        QUAL._validate_publication_commit(
            output,
            expected_run_metadata=run_metadata,
        )


def test_fallback_post_commit_fsync_failure_is_published_warning(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "fallback_post_commit_warning"
    failure = tmp_path / "FALLBACK_POST_COMMIT_FAILURE.json"
    real_fsync_directory = QUAL._fsync_directory

    def synthetic_payload_builder(**_: Any) -> tuple[dict[str, bytes], dict[str, Any]]:
        return _minimal_bundle_payloads(), {
            "qualification_status": "BLOCKED_PENDING_PUBLIC_EVIDENCE",
            "qualified": False,
            "canonical_record_count": 0,
        }

    def unsupported_atomic_rename(_: Path, __: Path) -> None:
        raise QUAL.AtomicNoReplaceUnsupported(errno.EINVAL)

    def fail_output_fsync_only_after_marker(path: Path) -> None:
        if path == output and (output / QUAL.PUBLICATION_COMMIT_FILENAME).exists():
            raise OSError("injected fallback post-commit fsync failure")
        real_fsync_directory(path)

    monkeypatch.setattr(QUAL, "_build_qualification_payloads", synthetic_payload_builder)
    monkeypatch.setattr(QUAL, "_rename_directory_noreplace", unsupported_atomic_rename)
    monkeypatch.setattr(QUAL, "_fsync_directory", fail_output_fsync_only_after_marker)
    result = QUAL.execute_qualification(
        repo_root=tmp_path / "repo",
        data_root=tmp_path / "data",
        protocol_path=tmp_path / "protocol.json",
        asset_manifest_path=tmp_path / "assets.json",
        expected_protocol_sha256="a" * 64,
        output_directory=output,
        failure_record_path=failure,
        run_id="A1_FALLBACK_POST_COMMIT",
        execution_id="GSE149487_FALLBACK_POST_COMMIT",
        recorded_at="2026-08-10T12:10:00+08:00",
    )

    assert result["kind"] == "BUNDLE"
    publication = result["report"]["publication"]
    assert publication["status"] == "PUBLISHED_WITH_POST_COMMIT_DURABILITY_WARNING"
    assert publication["publication_mode"] == QUAL.FALLBACK_PUBLICATION_MODE
    assert publication["terminal_commit_marker_validated"] is True
    assert "POST_COMMIT_OUTPUT_DIRECTORY_FSYNC_FAILED" in publication[
        "durability_warning_codes"
    ]
    assert QUAL._validate_publication_commit(
        output,
        expected_run_metadata={
            "run_id": "A1_FALLBACK_POST_COMMIT",
            "execution_id": "GSE149487_FALLBACK_POST_COMMIT",
            "recorded_at": "2026-08-10T12:10:00+08:00",
        },
    )["committed"] is True
    assert not failure.exists()


def test_fallback_rejects_nonapproved_errno_without_creating_final_directory(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "fallback_nonapproved_errno"

    def incorrectly_classified_atomic_error(_: Path, __: Path) -> None:
        raise QUAL.AtomicNoReplaceUnsupported(errno.EACCES)

    monkeypatch.setattr(
        QUAL,
        "_rename_directory_noreplace",
        incorrectly_classified_atomic_error,
    )
    with pytest.raises(QUAL.QualificationError, match="non-approved errno"):
        QUAL._publish_bundle(
            output,
            _minimal_bundle_payloads(),
            run_metadata=_publication_run_metadata("FALLBACK_NONAPPROVED"),
        )
    assert not output.exists()


def test_unsupported_staging_directory_fsync_still_reaches_marker_fallback(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "fallback_after_staging_fsync_unsupported"
    run_metadata = _publication_run_metadata("STAGING_FSYNC_UNSUPPORTED")
    real_fsync_directory = QUAL._fsync_directory

    def unsupported_staging_directory_fsync(path: Path) -> None:
        if path.name.startswith(f".{output.name}.partial-staging-"):
            raise OSError(errno.EINVAL, "staging directory fsync unsupported")
        real_fsync_directory(path)

    def unsupported_atomic_rename(_: Path, __: Path) -> None:
        raise QUAL.AtomicNoReplaceUnsupported(errno.ENOTSUP)

    monkeypatch.setattr(QUAL, "_fsync_directory", unsupported_staging_directory_fsync)
    monkeypatch.setattr(QUAL, "_rename_directory_noreplace", unsupported_atomic_rename)
    publication = QUAL._publish_bundle(
        output,
        _minimal_bundle_payloads(),
        run_metadata=run_metadata,
    )

    assert publication["publication_mode"] == QUAL.FALLBACK_PUBLICATION_MODE
    assert publication["terminal_commit_marker_validated"] is True
    assert publication["status"] == "PUBLISHED_WITH_POST_COMMIT_DURABILITY_WARNING"
    assert "PRECOMMIT_STAGING_DIRECTORY_FSYNC_UNSUPPORTED" in publication[
        "durability_warning_codes"
    ]
    assert QUAL._validate_publication_commit(
        output,
        expected_run_metadata=run_metadata,
    )["committed"] is True


def test_nonapproved_staging_directory_fsync_error_fails_before_publication(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "staging_fsync_hard_failure"
    rename_called = False
    real_fsync_directory = QUAL._fsync_directory

    def hard_staging_directory_fsync_failure(path: Path) -> None:
        if path.name.startswith(f".{output.name}.partial-staging-"):
            raise OSError(errno.EIO, "staging directory fsync hard failure")
        real_fsync_directory(path)

    def must_not_attempt_rename(_: Path, __: Path) -> list[str]:
        nonlocal rename_called
        rename_called = True
        return []

    monkeypatch.setattr(QUAL, "_fsync_directory", hard_staging_directory_fsync_failure)
    monkeypatch.setattr(QUAL, "_rename_directory_noreplace", must_not_attempt_rename)
    with pytest.raises(QUAL.QualificationError, match="before atomic commit"):
        QUAL._publish_bundle(
            output,
            _minimal_bundle_payloads(),
            run_metadata=_publication_run_metadata("STAGING_FSYNC_HARD_FAILURE"),
        )
    assert rename_called is False
    assert not output.exists()


def test_post_commit_fsync_failure_returns_published_warning_without_failure_record(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "published_bundle"
    failure = tmp_path / "FAILURE.json"
    real_fsync_directory = QUAL._fsync_directory

    def synthetic_payload_builder(**_: Any) -> tuple[dict[str, bytes], dict[str, Any]]:
        return _minimal_bundle_payloads(), {
            "qualification_status": "BLOCKED_PENDING_PUBLIC_EVIDENCE",
            "qualified": False,
            "canonical_record_count": 0,
        }

    def fail_only_after_commit(path: Path) -> None:
        if path == output.parent and output.exists():
            raise OSError("injected post-commit durability failure")
        real_fsync_directory(path)

    monkeypatch.setattr(QUAL, "_build_qualification_payloads", synthetic_payload_builder)
    monkeypatch.setattr(QUAL, "_fsync_directory", fail_only_after_commit)
    result = QUAL.execute_qualification(
        repo_root=tmp_path / "repo",
        data_root=tmp_path / "data",
        protocol_path=tmp_path / "protocol.json",
        asset_manifest_path=tmp_path / "assets.json",
        expected_protocol_sha256="a" * 64,
        output_directory=output,
        failure_record_path=failure,
        run_id="A1_POST_COMMIT_WARNING",
        execution_id="GSE149487_POST_COMMIT_WARNING",
        recorded_at="2026-08-10T12:03:00+08:00",
    )
    assert result["kind"] == "BUNDLE"
    assert result["report"]["publication"]["status"] == (
        "PUBLISHED_WITH_POST_COMMIT_DURABILITY_WARNING"
    )
    assert result["report"]["publication"]["directory_committed"] is True
    assert "POST_COMMIT_PARENT_DIRECTORY_FSYNC_FAILED" in result["report"][
        "publication"
    ]["durability_warning_codes"]
    assert output.is_dir()
    assert not failure.exists()


def test_primary_post_commit_parent_descriptor_close_failure_is_published_warning(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "primary_close_warning_bundle"
    failure = tmp_path / "PRIMARY_CLOSE_WARNING_FAILURE.json"
    real_open = QUAL.os.open
    real_close = QUAL.os.close
    publication_parent_descriptors: set[int] = set()

    def synthetic_payload_builder(**_: Any) -> tuple[dict[str, bytes], dict[str, Any]]:
        return _minimal_bundle_payloads(), {
            "qualification_status": "BLOCKED_PENDING_PUBLIC_EVIDENCE",
            "qualified": False,
            "canonical_record_count": 0,
        }

    def track_publication_parent_open(
        path: Path,
        flags: int,
        *args: Any,
        **kwargs: Any,
    ) -> int:
        descriptor = real_open(path, flags, *args, **kwargs)
        if (
            Path(path) == output.parent
            and flags & getattr(QUAL.os, "O_DIRECTORY", 0)
        ):
            publication_parent_descriptors.add(descriptor)
        return descriptor

    def close_then_report_error(descriptor: int) -> None:
        real_close(descriptor)
        if descriptor in publication_parent_descriptors:
            publication_parent_descriptors.remove(descriptor)
            raise OSError("injected post-commit parent descriptor close failure")

    monkeypatch.setattr(QUAL, "_build_qualification_payloads", synthetic_payload_builder)
    monkeypatch.setattr(QUAL.os, "open", track_publication_parent_open)
    monkeypatch.setattr(QUAL.os, "close", close_then_report_error)
    result = QUAL.execute_qualification(
        repo_root=tmp_path / "repo",
        data_root=tmp_path / "data",
        protocol_path=tmp_path / "protocol.json",
        asset_manifest_path=tmp_path / "assets.json",
        expected_protocol_sha256="a" * 64,
        output_directory=output,
        failure_record_path=failure,
        run_id="A1_PRIMARY_CLOSE_WARNING",
        execution_id="GSE149487_PRIMARY_CLOSE_WARNING",
        recorded_at="2026-08-10T12:11:00+08:00",
    )

    assert result["kind"] == "BUNDLE"
    publication = result["report"]["publication"]
    assert publication["publication_mode"] == QUAL.PRIMARY_PUBLICATION_MODE
    assert publication["status"] == "PUBLISHED_WITH_POST_COMMIT_DURABILITY_WARNING"
    assert "POST_COMMIT_PARENT_DIRECTORY_DESCRIPTOR_CLOSE_FAILED" in publication[
        "durability_warning_codes"
    ]
    assert output.is_dir()
    assert not failure.exists()


def test_unsupported_claim_parent_fsync_is_warning_and_does_not_block_publication(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "claim_fsync_unsupported_bundle"
    failure = tmp_path / "CLAIM_FSYNC_UNSUPPORTED_FAILURE.json"
    claim = QUAL._transaction_claim_path(output)
    real_fsync_directory = QUAL._fsync_directory
    injected = False

    def synthetic_payload_builder(**_: Any) -> tuple[dict[str, bytes], dict[str, Any]]:
        return _minimal_bundle_payloads(), {
            "qualification_status": "BLOCKED_PENDING_PUBLIC_EVIDENCE",
            "qualified": False,
            "canonical_record_count": 0,
        }

    def unsupported_only_for_claim_acquisition(path: Path) -> None:
        nonlocal injected
        if not injected and path == output.parent and claim.exists():
            injected = True
            raise OSError(errno.EOPNOTSUPP, "claim parent fsync unsupported")
        real_fsync_directory(path)

    monkeypatch.setattr(QUAL, "_build_qualification_payloads", synthetic_payload_builder)
    monkeypatch.setattr(QUAL, "_fsync_directory", unsupported_only_for_claim_acquisition)
    result = QUAL.execute_qualification(
        repo_root=tmp_path / "repo",
        data_root=tmp_path / "data",
        protocol_path=tmp_path / "protocol.json",
        asset_manifest_path=tmp_path / "assets.json",
        expected_protocol_sha256="a" * 64,
        output_directory=output,
        failure_record_path=failure,
        run_id="A1_CLAIM_FSYNC_UNSUPPORTED",
        execution_id="GSE149487_CLAIM_FSYNC_UNSUPPORTED",
        recorded_at="2026-08-10T12:12:00+08:00",
    )

    assert injected is True
    assert result["kind"] == "BUNDLE"
    assert "TRANSACTION_CLAIM_PARENT_DIRECTORY_FSYNC_UNSUPPORTED" in result[
        "report"
    ]["publication"]["durability_warning_codes"]
    assert result["report"]["publication"]["status"] == (
        "PUBLISHED_WITH_POST_COMMIT_DURABILITY_WARNING"
    )
    assert output.is_dir()
    assert not failure.exists()
    assert not claim.exists()


def test_nonapproved_claim_parent_fsync_error_fails_closed_before_build(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "claim_fsync_hard_failure_bundle"
    failure = tmp_path / "CLAIM_FSYNC_HARD_FAILURE.json"
    claim = QUAL._transaction_claim_path(output)
    build_called = False
    real_fsync_directory = QUAL._fsync_directory

    def must_not_build(**_: Any) -> tuple[dict[str, bytes], dict[str, Any]]:
        nonlocal build_called
        build_called = True
        return _minimal_bundle_payloads(), {}

    def hard_failure_only_for_claim_acquisition(path: Path) -> None:
        if path == output.parent and claim.exists():
            raise OSError(errno.EIO, "claim parent fsync hard failure")
        real_fsync_directory(path)

    monkeypatch.setattr(QUAL, "_build_qualification_payloads", must_not_build)
    monkeypatch.setattr(QUAL, "_fsync_directory", hard_failure_only_for_claim_acquisition)
    with pytest.raises(QUAL.QualificationError, match="durably acquired"):
        QUAL.execute_qualification(
            repo_root=tmp_path / "repo",
            data_root=tmp_path / "data",
            protocol_path=tmp_path / "protocol.json",
            asset_manifest_path=tmp_path / "assets.json",
            expected_protocol_sha256="a" * 64,
            output_directory=output,
            failure_record_path=failure,
            run_id="A1_CLAIM_FSYNC_HARD_FAILURE",
            execution_id="GSE149487_CLAIM_FSYNC_HARD_FAILURE",
            recorded_at="2026-08-10T12:13:00+08:00",
        )
    assert build_called is False
    assert not output.exists()
    assert not failure.exists()
    assert not claim.exists()


def test_failure_record_parent_fsync_error_returns_failure_truth_with_warning(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "failure_truth_output"
    failure = tmp_path / "FAILURE_TRUTH.json"
    real_fsync_directory = QUAL._fsync_directory

    def fail_qualification(**_: Any) -> dict[str, Any]:
        raise QUAL.QualificationError("injected pre-publication qualification failure")

    def fail_directory_fsync_after_failure_commit(path: Path) -> None:
        if path == failure.parent and failure.exists():
            raise OSError(errno.EINVAL, "failure parent fsync unsupported")
        real_fsync_directory(path)

    monkeypatch.setattr(QUAL, "qualify_gse149487_plumage", fail_qualification)
    monkeypatch.setattr(QUAL, "_fsync_directory", fail_directory_fsync_after_failure_commit)
    result = QUAL.execute_qualification(
        repo_root=tmp_path / "repo",
        data_root=tmp_path / "data",
        protocol_path=tmp_path / "protocol.json",
        asset_manifest_path=tmp_path / "assets.json",
        expected_protocol_sha256="a" * 64,
        output_directory=output,
        failure_record_path=failure,
        run_id="A1_FAILURE_TRUTH",
        execution_id="GSE149487_FAILURE_TRUTH",
        recorded_at="2026-08-10T12:14:00+08:00",
    )

    assert result["kind"] == "FAILURE"
    assert result["failure"]["status"] == (
        "FAIL_CLOSED_BEFORE_SUCCESS_BUNDLE_PUBLICATION"
    )
    assert result["failure"]["failure_record_durability_warning_codes"] == [
        "FAILURE_RECORD_PARENT_DIRECTORY_FSYNC_FAILED"
    ]
    assert failure.is_file()
    assert json.loads(failure.read_text(encoding="utf-8"))["status"] == (
        "FAIL_CLOSED_BEFORE_SUCCESS_BUNDLE_PUBLICATION"
    )
    assert not output.exists()


def test_failure_record_post_fsync_descriptor_close_error_preserves_failure_truth(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "failure_close_truth_output"
    failure = tmp_path / "FAILURE_CLOSE_TRUTH.json"
    real_open = QUAL.os.open
    real_close = QUAL.os.close
    failure_descriptors: set[int] = set()

    def fail_qualification(**_: Any) -> dict[str, Any]:
        raise QUAL.QualificationError("injected pre-publication qualification failure")

    def track_failure_open(
        path: Path,
        flags: int,
        *args: Any,
        **kwargs: Any,
    ) -> int:
        descriptor = real_open(path, flags, *args, **kwargs)
        if Path(path) == failure and flags & QUAL.os.O_EXCL:
            failure_descriptors.add(descriptor)
        return descriptor

    def close_failure_after_real_close(descriptor: int) -> None:
        real_close(descriptor)
        if descriptor in failure_descriptors:
            failure_descriptors.remove(descriptor)
            raise OSError("injected failure-record descriptor close error")

    monkeypatch.setattr(QUAL, "qualify_gse149487_plumage", fail_qualification)
    monkeypatch.setattr(QUAL.os, "open", track_failure_open)
    monkeypatch.setattr(QUAL.os, "close", close_failure_after_real_close)
    result = QUAL.execute_qualification(
        repo_root=tmp_path / "repo",
        data_root=tmp_path / "data",
        protocol_path=tmp_path / "protocol.json",
        asset_manifest_path=tmp_path / "assets.json",
        expected_protocol_sha256="a" * 64,
        output_directory=output,
        failure_record_path=failure,
        run_id="A1_FAILURE_CLOSE_TRUTH",
        execution_id="GSE149487_FAILURE_CLOSE_TRUTH",
        recorded_at="2026-08-10T12:20:00+08:00",
    )

    assert result["kind"] == "FAILURE"
    assert result["failure"]["canonical_record_count"] == 0
    assert result["failure"]["failure_record_durability_warning_codes"] == [
        "FAILURE_RECORD_DESCRIPTOR_CLOSE_FAILED"
    ]
    assert failure.is_file()
    assert json.loads(failure.read_text(encoding="utf-8"))["status"] == (
        "FAIL_CLOSED_BEFORE_SUCCESS_BUNDLE_PUBLICATION"
    )
    assert not output.exists()


def test_main_failure_stdout_reports_persisted_truth_warning_counts(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
) -> None:
    def synthetic_failure_result(**_: Any) -> dict[str, Any]:
        return {
            "kind": "FAILURE",
            "failure": {
                "status": "FAIL_CLOSED_BEFORE_SUCCESS_BUNDLE_PUBLICATION",
                "failure_type": "QualificationError",
                "failure_record_durability_warning_codes": [
                    "FAILURE_RECORD_PARENT_DIRECTORY_FSYNC_FAILED"
                ],
                "transaction_claim_warning_codes": [
                    "TRANSACTION_CLAIM_CLEANUP_DIRECTORY_FSYNC_FAILED"
                ],
            },
        }

    monkeypatch.setattr(QUAL, "execute_qualification", synthetic_failure_result)
    exit_code = QUAL.main(
        [
            "--repo-root",
            str(tmp_path / "repo"),
            "--data-root",
            str(tmp_path / "data"),
            "--protocol",
            str(tmp_path / "protocol.json"),
            "--asset-manifest",
            str(tmp_path / "assets.json"),
            "--protocol-sha256",
            "a" * 64,
            "--output-dir",
            str(tmp_path / "output"),
            "--failure-record",
            str(tmp_path / "FAILURE.json"),
            "--run-id",
            "A1_FAILURE_STDOUT",
            "--execution-id",
            "GSE149487_FAILURE_STDOUT",
            "--recorded-at",
            "2026-08-10T12:15:00+08:00",
        ]
    )
    stdout = json.loads(capsys.readouterr().out)
    assert exit_code == 2
    assert stdout["status"] == "FAIL_CLOSED_BEFORE_SUCCESS_BUNDLE_PUBLICATION"
    assert stdout["failure_record_durability_warning_count"] == 1
    assert stdout["transaction_claim_warning_count"] == 1
    assert stdout["durability_warning_count"] == 2
    assert stdout["canonical_record_count"] == 0


def test_main_bundle_stdout_includes_publication_mode(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
) -> None:
    def synthetic_bundle_result(**_: Any) -> dict[str, Any]:
        return {
            "kind": "BUNDLE",
            "report": {
                "qualification_status": "BLOCKED_PENDING_PUBLIC_EVIDENCE",
                "qualified": False,
                "canonical_record_count": 0,
                "qualified_independent_ordinary_study_count": 0,
                "publication": {
                    "publication_mode": QUAL.PRIMARY_PUBLICATION_MODE,
                    "status": "PUBLISHED_DURABLE",
                    "durability_warning_codes": [],
                },
            },
        }

    monkeypatch.setattr(QUAL, "execute_qualification", synthetic_bundle_result)
    exit_code = QUAL.main(
        [
            "--repo-root",
            str(tmp_path / "repo"),
            "--data-root",
            str(tmp_path / "data"),
            "--protocol",
            str(tmp_path / "protocol.json"),
            "--asset-manifest",
            str(tmp_path / "assets.json"),
            "--protocol-sha256",
            "a" * 64,
            "--output-dir",
            str(tmp_path / "output"),
            "--failure-record",
            str(tmp_path / "FAILURE.json"),
            "--run-id",
            "A1_BUNDLE_STDOUT",
            "--execution-id",
            "GSE149487_BUNDLE_STDOUT",
            "--recorded-at",
            "2026-08-10T12:16:00+08:00",
        ]
    )
    stdout = json.loads(capsys.readouterr().out)
    assert exit_code == 0
    assert stdout["publication_mode"] == QUAL.PRIMARY_PUBLICATION_MODE
    assert stdout["publication_status"] == "PUBLISHED_DURABLE"
    assert stdout["canonical_record_count"] == 0


def test_main_committed_not_accepted_is_nonzero_and_never_accepts_canonical(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
) -> None:
    def synthetic_committed_not_accepted(**_: Any) -> dict[str, Any]:
        return {
            "kind": "COMMITTED_NOT_ACCEPTED",
            "committed_not_accepted": {
                "status": "COMMITTED_NOT_ACCEPTED",
                "publication_mode": QUAL.PRIMARY_PUBLICATION_MODE,
                "durability_warning_codes": [],
                "transaction_claim_warning_codes": [],
            },
        }

    monkeypatch.setattr(
        QUAL,
        "execute_qualification",
        synthetic_committed_not_accepted,
    )
    exit_code = QUAL.main(
        [
            "--repo-root",
            str(tmp_path / "repo"),
            "--data-root",
            str(tmp_path / "data"),
            "--protocol",
            str(tmp_path / "protocol.json"),
            "--asset-manifest",
            str(tmp_path / "assets.json"),
            "--protocol-sha256",
            "a" * 64,
            "--output-dir",
            str(tmp_path / "output"),
            "--failure-record",
            str(tmp_path / "FAILURE.json"),
            "--run-id",
            "A1_COMMITTED_NOT_ACCEPTED_STDOUT",
            "--execution-id",
            "GSE149487_COMMITTED_NOT_ACCEPTED_STDOUT",
            "--recorded-at",
            "2026-08-10T12:19:00+08:00",
        ]
    )
    stdout = json.loads(capsys.readouterr().out)
    assert exit_code == 5
    assert stdout["status"] == "COMMITTED_NOT_ACCEPTED"
    assert stdout["directory_committed"] is True
    assert stdout["publication_accepted"] is False
    assert stdout["qualified"] is False
    assert stdout["canonical_accepted"] is False
    assert stdout["canonical_record_count"] == 0
    assert stdout["failure_record_materialized"] is False


def test_execute_claim_loser_does_not_build_or_publish_failure(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "contended_bundle"
    failure = tmp_path / "CONTENDED_FAILURE.json"
    claim = QUAL._transaction_claim_path(output)
    claim.write_text("owned by winner\n", encoding="utf-8")

    def must_not_build(**_: Any) -> tuple[dict[str, bytes], dict[str, Any]]:
        raise AssertionError("claim loser reached payload construction")

    monkeypatch.setattr(QUAL, "_build_qualification_payloads", must_not_build)
    result = QUAL.execute_qualification(
        repo_root=tmp_path / "repo",
        data_root=tmp_path / "data",
        protocol_path=tmp_path / "protocol.json",
        asset_manifest_path=tmp_path / "assets.json",
        expected_protocol_sha256="a" * 64,
        output_directory=output,
        failure_record_path=failure,
        run_id="A1_TRANSACTION_LOSER",
        execution_id="GSE149487_TRANSACTION_LOSER",
        recorded_at="2026-08-10T12:04:00+08:00",
    )
    assert result["kind"] == "CONTENDED"
    assert result["contention"]["status"] == "TRANSACTION_CLAIM_CONTENDED"
    assert result["contention"]["failure_record_materialized"] is False
    assert not output.exists()
    assert not failure.exists()


def test_execute_second_attempt_after_success_never_creates_failure_double_truth(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "single_truth_bundle"
    failure = tmp_path / "SINGLE_TRUTH_FAILURE.json"

    def synthetic_payload_builder(**_: Any) -> tuple[dict[str, bytes], dict[str, Any]]:
        return _minimal_bundle_payloads(), {
            "qualification_status": "BLOCKED_PENDING_PUBLIC_EVIDENCE",
            "qualified": False,
            "canonical_record_count": 0,
        }

    monkeypatch.setattr(QUAL, "_build_qualification_payloads", synthetic_payload_builder)
    common = {
        "repo_root": tmp_path / "repo",
        "data_root": tmp_path / "data",
        "protocol_path": tmp_path / "protocol.json",
        "asset_manifest_path": tmp_path / "assets.json",
        "expected_protocol_sha256": "a" * 64,
        "output_directory": output,
        "failure_record_path": failure,
        "run_id": "A1_SINGLE_TRUTH",
        "recorded_at": "2026-08-10T12:05:00+08:00",
    }
    winner = QUAL.execute_qualification(
        **common,
        execution_id="GSE149487_SINGLE_TRUTH_WINNER",
    )
    loser = QUAL.execute_qualification(
        **common,
        execution_id="GSE149487_SINGLE_TRUTH_LOSER",
    )
    assert winner["kind"] == "BUNDLE"
    assert loser["kind"] == "CONTENDED"
    assert loser["contention"]["status"] == "OUTPUT_ALREADY_COMMITTED_OR_OCCUPIED"
    assert output.is_dir()
    assert not failure.exists()
    assert not QUAL._transaction_claim_path(output).exists()


def test_execute_requires_failure_record_to_be_output_sibling(tmp_path: Path) -> None:
    output = tmp_path / "bundle"
    outside_parent = tmp_path / "elsewhere"
    outside_parent.mkdir()
    with pytest.raises(QUAL.QualificationError, match="exclusive sibling"):
        QUAL.execute_qualification(
            repo_root=tmp_path / "repo",
            data_root=tmp_path / "data",
            protocol_path=tmp_path / "protocol.json",
            asset_manifest_path=tmp_path / "assets.json",
            expected_protocol_sha256="a" * 64,
            output_directory=output,
            failure_record_path=outside_parent / "FAILURE.json",
            run_id="A1_NON_SIBLING_FAILURE",
            execution_id="GSE149487_NON_SIBLING_FAILURE",
            recorded_at="2026-08-10T12:06:00+08:00",
        )
    assert not output.exists()
    assert not (outside_parent / "FAILURE.json").exists()


def test_unclassified_raw_key_blocks_canonical_without_identifier_leakage(
    tmp_path: Path,
) -> None:
    secret_extra_key = "UNCLASSIFIED_RAW_SECRET_SENTINEL"
    repo, data, protocol, asset_manifest = _fixture_repo_and_data(
        tmp_path,
        extra_raw_keys_by_context={"PC3": (secret_extra_key,)},
    )
    output = tmp_path / "raw_key_blocked_bundle"
    report = QUAL.qualify_gse149487_plumage(
        repo_root=repo,
        data_root=data,
        protocol_path=protocol,
        asset_manifest_path=asset_manifest,
        expected_protocol_sha256=_sha256(protocol),
        output_directory=output,
        run_id="A1_RAW_KEY_RECONCILIATION",
        execution_id="GSE149487_RAW_KEY_RECONCILIATION",
        recorded_at="2026-08-10T12:02:00+08:00",
    )
    assert report["canonical_record_count"] == 0
    assert "RAW_KEY_UNCLASSIFIED_OUTCOME_BLIND_RECONCILIATION_NOT_ZERO" in report[
        "blockers"
    ]
    reconciliation = report["raw_grid"]["outcome_blind_key_reconciliation"]
    assert reconciliation["status"] == "BLOCKED_OUTCOME_BLIND_KEY_RECONCILIATION"
    assert reconciliation["unclassified_raw_key_count"] == 1
    assert reconciliation["per_context"]["PC3"]["classification_counts"][
        "UNCLASSIFIED"
    ] == 1
    assert reconciliation["per_context"]["293T"]["classification_counts"][
        "UNCLASSIFIED"
    ] == 0
    assert not (output / "canonical_intervention_records.jsonl").exists()
    assert secret_extra_key.encode() not in b"".join(
        path.read_bytes() for path in output.iterdir()
    )


def test_execution_error_publishes_exclusive_failure_record(tmp_path: Path) -> None:
    failure = tmp_path / "FAILURE.json"
    output = tmp_path / "never_created"
    result = QUAL.execute_qualification(
        repo_root=tmp_path / "restricted_repo",
        data_root=tmp_path / "data",
        protocol_path=tmp_path / "protocol.json",
        asset_manifest_path=tmp_path / "assets.json",
        expected_protocol_sha256="a" * 64,
        output_directory=output,
        failure_record_path=failure,
        run_id="A1_FIXTURE_FAILURE",
        execution_id="GSE149487_FIXTURE_FAILURE",
        recorded_at="2026-08-10T12:00:00+08:00",
    )
    assert result["kind"] == "FAILURE"
    payload = json.loads(failure.read_text(encoding="utf-8"))
    assert payload["status"] == "FAIL_CLOSED_BEFORE_SUCCESS_BUNDLE_PUBLICATION"
    assert payload["canonical_record_count"] == 0
    assert payload["qualified"] is False
    assert not output.exists()
    with pytest.raises(QUAL.QualificationError, match="overwrite"):
        QUAL.execute_qualification(
            repo_root=tmp_path / "restricted_repo",
            data_root=tmp_path / "data",
            protocol_path=tmp_path / "protocol.json",
            asset_manifest_path=tmp_path / "assets.json",
            expected_protocol_sha256="a" * 64,
            output_directory=output,
            failure_record_path=failure,
            run_id="A1_FIXTURE_FAILURE",
            execution_id="GSE149487_FIXTURE_FAILURE_2",
            recorded_at="2026-08-10T12:01:00+08:00",
        )
