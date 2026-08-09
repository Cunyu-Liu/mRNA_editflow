from __future__ import annotations

import hashlib
import importlib.util
import json
import math
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[2]
MODULE_PATH = (
    ROOT / "scripts" / "route_a_v3" / "reconstruct_gse149487_plumage.py"
)
SPEC = importlib.util.spec_from_file_location(
    "reconstruct_gse149487_plumage", MODULE_PATH
)
assert SPEC and SPEC.loader
RECON = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(RECON)


SOURCE_SEQUENCE = "AACCGG"
CANDIDATE_SEQUENCE = "AATCGG"
MUTANT_DESCRIPTION = "GENE1_C_T_chr1_100_105"
WT_DESCRIPTION = "GENE1_WT_chr1_100_105"
RAW_MUTANT_BARCODE_IDS = tuple(
    f"RAW_MUTANT_BARCODE_SECRET_{index:02d}" for index in range(1, 23)
)
RAW_WT_BARCODE_IDS = tuple(
    f"RAW_WT_BARCODE_SECRET_{index:02d}" for index in range(1, 23)
)
RAW_BARCODE_IDS = RAW_MUTANT_BARCODE_IDS + RAW_WT_BARCODE_IDS
AUTHORITY_PROTOCOL = ROOT / "configs" / "route_a_v3_plumage_reconstruction.json"


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _lim6c_block_values(
    *,
    description: str,
    barcode: str,
    transcript_ratios: tuple[float, float, float],
    te_ratios: tuple[float, float, float],
) -> tuple[list[Any], list[Any]]:
    dna = (0.0, 0.0, 0.0)
    total = tuple(
        dna_value + ratio
        for dna_value, ratio in zip(dna, transcript_ratios)
    )
    polysome = tuple(
        total_value + ratio
        for total_value, ratio in zip(total, te_ratios)
    )
    return (
        [
            description,
            barcode,
            total[0],
            dna[0],
            total[1],
            dna[1],
            total[2],
            dna[2],
        ],
        [
            description,
            barcode,
            total[0],
            polysome[0],
            total[1],
            polysome[1],
            total[2],
            polysome[2],
        ],
    )


def _append_6c_row(
    sheet: Any,
    *,
    description: str,
    barcode: str,
    transcript_ratios: tuple[float, float, float],
    te_ratios: tuple[float, float, float],
) -> None:
    transcript_values, te_values = _lim6c_block_values(
        description=description,
        barcode=barcode,
        transcript_ratios=transcript_ratios,
        te_ratios=te_ratios,
    )
    sheet.append([*transcript_values, None, *te_values])


def _write_moesm8(
    path: Path,
    *,
    source_sequence: str = SOURCE_SEQUENCE,
    candidate_sequence: str = CANDIDATE_SEQUENCE,
    ref: str = "C",
    alt: str = "T",
) -> None:
    workbook = Workbook()
    sequence_sheet = workbook.active
    sequence_sheet.title = RECON.MOESM8_SEQUENCE_SHEET
    sequence_sheet.append(list(RECON.MOESM8_6A_RAW_HEADER))
    sequence_sheet.append(
        [
            len(candidate_sequence),
            "GENE1",
            f"GENE1_chr1_100_{ref}_{alt}_UTR5",
            candidate_sequence,
        ]
    )
    sequence_sheet.append(
        [
            len(source_sequence),
            "GENE1",
            "GENE1_chr1_100_WT_UTR5",
            source_sequence,
        ]
    )
    for sheet_name in RECON.MOESM8_SHEETS[1:]:
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(["REFERENCE_ONLY_NOT_READ"])
    workbook.save(path)


def _write_lim6c(path: Path, *, include_wt: bool = True) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = RECON.LIM6C_SHEET
    sheet.append(list(RECON.LIM6C_RAW_HEADER))
    logical_rows: list[tuple[list[Any], list[Any]]] = []
    for index, barcode in enumerate(RAW_MUTANT_BARCODE_IDS):
        lower_half = index < 11
        logical_rows.append(
            _lim6c_block_values(
                description=MUTANT_DESCRIPTION,
                barcode=barcode,
                transcript_ratios=(3.0, 5.0, 8.0)
                if lower_half
                else (5.0, 7.0, 10.0),
                te_ratios=(2.0, 3.0, 4.0)
                if lower_half
                else (4.0, 5.0, 6.0),
            )
        )
    if include_wt:
        for index, barcode in enumerate(RAW_WT_BARCODE_IDS):
            lower_half = index < 11
            logical_rows.append(
                _lim6c_block_values(
                    description=WT_DESCRIPTION,
                    barcode=barcode,
                    transcript_ratios=(1.0, 2.0, 3.0)
                    if lower_half
                    else (3.0, 4.0, 5.0),
                    te_ratios=(1.0, 1.0, 1.0)
                    if lower_half
                    else (3.0, 3.0, 3.0),
                )
            )
    # The two physical blocks are deliberately shuffled relative to one another.
    # A valid implementation must join only by (description, barcode).
    right_rows = [right for _, right in reversed(logical_rows)]
    for (left, _), right in zip(logical_rows, right_rows):
        sheet.append([*left, None, *right])
    workbook.save(path)


def _write_protocol(path: Path, *, moesm8_sha256: str, lim6c_sha256: str) -> None:
    protocol = json.loads(AUTHORITY_PROTOCOL.read_text(encoding="utf-8"))
    protocol["inputs"]["moesm8"]["sha256"] = moesm8_sha256
    protocol["inputs"]["lim6c_293t"]["sha256"] = lim6c_sha256
    path.write_text(
        json.dumps(protocol, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _fixture_inputs(
    tmp_path: Path,
    *,
    include_wt: bool = True,
    source_sequence: str = SOURCE_SEQUENCE,
    candidate_sequence: str = CANDIDATE_SEQUENCE,
    ref: str = "C",
    alt: str = "T",
) -> tuple[Path, Path]:
    moesm8 = tmp_path / "41467_2021_24445_MOESM8_ESM.xlsx"
    lim6c = tmp_path / "Lim_et_al_Supp_Tbl_6c_293T.xlsx"
    _write_moesm8(
        moesm8,
        source_sequence=source_sequence,
        candidate_sequence=candidate_sequence,
        ref=ref,
        alt=alt,
    )
    _write_lim6c(lim6c, include_wt=include_wt)
    _write_protocol(
        tmp_path / "route_a_v3_plumage_reconstruction.json",
        moesm8_sha256=_sha256(moesm8),
        lim6c_sha256=_sha256(lim6c),
    )
    return moesm8, lim6c


def _run(moesm8: Path, lim6c: Path, output: Path) -> dict[str, Any]:
    protocol = moesm8.parent / "route_a_v3_plumage_reconstruction.json"
    _write_protocol(
        protocol,
        moesm8_sha256=_sha256(moesm8),
        lim6c_sha256=_sha256(lim6c),
    )
    return RECON.reconstruct_gse149487_plumage(
        moesm8_path=moesm8,
        lim6c_path=lim6c,
        protocol_path=protocol,
        expected_moesm8_sha256=_sha256(moesm8),
        expected_lim6c_sha256=_sha256(lim6c),
        expected_protocol_sha256=_sha256(protocol),
        output_directory=output,
    )


def _protocol_path(workbook: Path) -> Path:
    return workbook.parent / "route_a_v3_plumage_reconstruction.json"


def _jsonl(path: Path) -> list[dict[str, Any]]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines()]


def _rewrite_cell(path: Path, *, row: int, column: int, value: Any) -> None:
    workbook = load_workbook(path)
    try:
        workbook[RECON.LIM6C_SHEET].cell(row=row, column=column).value = value
        workbook.save(path)
    finally:
        workbook.close()


def test_exact_formula_three_replicates_se_and_closed_method(tmp_path: Path) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path)
    output = tmp_path / "reconstruction"

    report = _run(moesm8, lim6c, output)
    records = _jsonl(output / "development_companion_effect_records.jsonl")
    replicate_rows = _jsonl(output / "replicate_effect_summaries.jsonl")
    barcode_audit = _jsonl(output / "BARCODE_RATIO_AUDIT.jsonl")
    description_audit = _jsonl(output / "DESCRIPTION_CLASSIFICATION_AUDIT.jsonl")
    moesm8_6a_audit = _jsonl(output / "MOESM8_6A_CLASSIFICATION_AUDIT.jsonl")

    assert len(records) == 2
    assert not (output / "canonical_intervention_records.jsonl").exists()
    assert all(
        row["schema_id"] == "route_a_v3.development_companion_effect_record"
        and row["record_type"] == "DEVELOPMENT_COMPANION_EFFECT_SUMMARY"
        and row["canonical_intervention_record_v3_materialized"] is False
        for row in records
    )
    transcript = next(
        row
        for row in records
        if row["endpoint_id"] == "transcript_log2_totalrna_over_dna"
    )
    assert transcript["biological_replicate_deltas"] == pytest.approx([2.0, 3.0, 5.0])
    assert transcript["effect_delta_mutant_minus_wt"] == pytest.approx(10.0 / 3.0)
    assert transcript["standard_error"] == pytest.approx(math.sqrt(7.0) / 3.0)
    assert transcript["effective_n"] == 3
    assert transcript["biological_replicate_count"] == 3
    assert transcript["effective_n_unit"] == "BIOLOGICAL_REPLICATE"
    assert transcript["barcode_is_independent_n"] is False
    assert transcript["source_sequence"] == SOURCE_SEQUENCE
    assert transcript["candidate_sequence"] == CANDIDATE_SEQUENCE
    assert transcript["edit"] == {
        "alt": "T",
        "ref": "C",
        "sequence_index_0_based": 2,
        "type": "SNV",
    }
    assert transcript["group_binding"]["gene"] == "GENE1"
    assert transcript["group_binding"]["design_family"] == RECON.DESIGN_FAMILY
    assert transcript["group_binding"]["wt_construct_id"] == transcript["wt_construct_id"]
    assert transcript["study_group_id"] == "PLUMAGE_LIM_2021"
    assert transcript["independent_study_id"] == "PLUMAGE_LIM_2021"
    assert transcript["independent_study_count"] == 1
    assert transcript["context"] == "293T"
    assert transcript["provenance"]["lim6c_293t"][
        "endpoint_source_block_id"
    ] == RECON.TRANSCRIPT_BLOCK_ID
    assert transcript["provenance"]["lim6c_293t"]["row_position_join_used"] is False
    translation_efficiency = next(
        row
        for row in records
        if row["endpoint_id"] == "te_log2_polysome_over_totalrna"
    )
    assert translation_efficiency["biological_replicate_deltas"] == pytest.approx(
        [1.0, 2.0, 3.0]
    )
    assert translation_efficiency["effect_delta_mutant_minus_wt"] == pytest.approx(
        2.0
    )
    assert translation_efficiency["standard_error"] == pytest.approx(
        1.0 / math.sqrt(3.0)
    )
    assert translation_efficiency["provenance"]["lim6c_293t"][
        "endpoint_source_block_id"
    ] == RECON.TRANSLATION_EFFICIENCY_BLOCK_ID

    transcript_rep1 = next(
        row
        for row in replicate_rows
        if row["endpoint_id"] == "transcript_log2_totalrna_over_dna"
        and row["replicate_id"] == "rep1"
    )
    assert transcript_rep1["mutant_construct_barcode_ratio_median"] == pytest.approx(4.0)
    assert transcript_rep1["wt_construct_barcode_ratio_median"] == pytest.approx(2.0)
    assert transcript_rep1["replicate_delta_mutant_minus_wt"] == pytest.approx(2.0)
    assert transcript_rep1["mutant_technical_barcode_count_retained"] == 22

    assert len(barcode_audit) == 264  # 44 barcodes x 3 reps x 2 endpoints
    audit_mutant_rep1 = next(
        row
        for row in barcode_audit
        if row["arm"] == "MUTANT"
        and row["endpoint_id"] == "transcript_log2_totalrna_over_dna"
        and row["biological_replicate_id"] == "rep1"
        and row["barcode_log2_ratio"] == pytest.approx(3.0)
    )
    assert audit_mutant_rep1["eligible"] is True
    assert audit_mutant_rep1["exclusion_reason"] is None
    assert audit_mutant_rep1["numerator_log2_cpm"] == pytest.approx(3.0)
    assert audit_mutant_rep1["denominator_log2_cpm"] == pytest.approx(0.0)
    assert "numerator_cpm" not in audit_mutant_rep1
    assert "denominator_cpm" not in audit_mutant_rep1
    assert audit_mutant_rep1["original_cpm_minimum_inclusive"] == 0.5
    assert audit_mutant_rep1["published_log2_cpm_minimum_inclusive"] == -1.0
    assert len(audit_mutant_rep1["barcode_id_sha256"]) == 64
    assert audit_mutant_rep1["source_asset_id"] == "GSE149487_LIM6C_293T"
    assert audit_mutant_rep1["sheet_name"] == "Sheet1"
    assert audit_mutant_rep1["source_block_id"] == RECON.TRANSCRIPT_BLOCK_ID
    assert audit_mutant_rep1["physical_row_number"] == 2
    assert len(audit_mutant_rep1["row_locator_sha256"]) == 64
    assert audit_mutant_rep1["raw_barcode_id_emitted"] is False
    assert "barcode" not in audit_mutant_rep1

    mutant_01_hash = RECON._salted_barcode_sha256(
        RAW_MUTANT_BARCODE_IDS[0],
        lim6c_sha256=_sha256(lim6c),
    )
    te_mutant_rep1 = next(
        row
        for row in barcode_audit
        if row["barcode_id_sha256"] == mutant_01_hash
        and row["endpoint_id"] == "te_log2_polysome_over_totalrna"
        and row["biological_replicate_id"] == "rep1"
    )
    assert te_mutant_rep1["source_block_id"] == (
        RECON.TRANSLATION_EFFICIENCY_BLOCK_ID
    )
    assert te_mutant_rep1["physical_row_number"] == 45
    assert te_mutant_rep1["key_present_in_endpoint_block"] is True

    assert len(description_audit) == 2
    assert all(row["union_key_count"] == 22 for row in description_audit)
    assert all(row["transcript_block_key_count"] == 22 for row in description_audit)
    assert all(
        row["translation_efficiency_block_key_count"] == 22
        for row in description_audit
    )
    assert {row["classification_category"] for row in description_audit} == {
        "included",
        "reference_only",
    }
    assert {
        row["classification"] for row in description_audit
    } == {
        "INCLUDED_STRICT_SNV_MUTANT",
        "REFERENCE_ONLY_STRICT_WT",
    }
    assert len(moesm8_6a_audit) == 2
    assert {row["classification"] for row in moesm8_6a_audit} == {
        "INCLUDED_STRICT_SNV_MUTANT_COORDINATE",
        "REFERENCE_ONLY_STRICT_WT_COORDINATE",
    }
    assert all(
        row["raw_coordinate_emitted"] is False
        and row["raw_sequence_emitted"] is False
        and "coordinate" not in row
        and "sequence" not in row
        for row in moesm8_6a_audit
    )

    assert report["paper_method_reproduced"] is False
    assert report["companion_summary_only"] is True
    assert report["paper_method"] == {
        "barcode_distribution_test": "TWO_SIDED_MANN_WHITNEY_MUTANT_VS_WT",
        "multiple_testing_rule": "FDR_LT_0.1",
        "reproduced": False,
    }
    assert "PC3_AND_18_GEO_RAW_COUNT_TABLE_JOIN_NOT_INCLUDED" in report["blockers"]
    assert "CANONICAL_INTERVENTION_RECORD_V3_NOT_MATERIALIZED" in report["blockers"]
    assert "UNADJUDICATED_DESCRIPTION_CLASSES_EXCLUDED" in report["blockers"]
    assert "UNADJUDICATED_6A_COORDINATE_CLASSES_EXCLUDED" in report["blockers"]
    assert report["workbook_structure"]["moesm8_6a_header_sha256"] == (
        RECON.MOESM8_6A_RAW_HEADER_SHA256
    )
    assert report["workbook_structure"]["lim6c_header_sha256"] == (
        RECON.LIM6C_RAW_HEADER_SHA256
    )
    block_structure = report["workbook_structure"]["lim6c_block_semantics"]
    assert block_structure["physical_data_rows"] == 44
    assert block_structure["transcript_present_key_rows"] == 44
    assert block_structure["translation_efficiency_present_key_rows"] == 44
    assert block_structure["translation_efficiency_blank_tail_rows"] == 0
    assert block_structure["hashed_key_intersection"] == 44
    assert block_structure["hashed_key_transcript_only"] == 0
    assert block_structure["hashed_key_translation_efficiency_only"] == 0
    assert block_structure["key_aligned_totalrna_exact_equal"] == 44
    assert block_structure["key_aligned_totalrna_any_difference"] == 0
    assert block_structure["rowwise_key_equal_count_descriptive_only"] == 0
    assert block_structure["row_position_join_used"] is False
    assert report["lim6c_block_semantics"]["endpoint_uses_only_its_own_block"] is True
    assert report["lim6c_block_semantics"]["missing_key_is_zero"] is False
    assert report["summary"]["barcode_ratio_audit_row_count"] == 264
    assert report["study_group_id"] == "PLUMAGE_LIM_2021"
    assert report["independent_study_id"] == "PLUMAGE_LIM_2021"
    assert report["independent_study_count"] == 1
    assert report["endpoint_or_context_increases_independent_study_count"] is False
    assert report["summary"]["canonical_record_count"] == 0
    assert report["canonical_intervention_record_v3_materialized"] is False
    assert report["summary"]["development_companion_effect_record_count"] == 2
    assert report["summary"]["moesm8_6a_physical_data_row_denominator"] == 2
    assert report["summary"]["moesm8_6a_classification_complete"] is True
    assert report["companion_method"]["analysis_method"] == (
        "ROUTE_A_COMPANION_NOT_PAPER_TEST"
    )


def test_published_log2_cpm_negative_values_and_minus_one_boundary_are_valid(
    tmp_path: Path,
) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path)
    # The source workbook stores published log2(CPM per barcode), not CPM.
    # Both -0.5 and the exact -1 floor are valid component values.
    _rewrite_cell(lim6c, row=2, column=4, value=-0.5)
    _rewrite_cell(lim6c, row=3, column=4, value=-1.0)
    output = tmp_path / "published-log2-cpm-boundary"

    report = _run(moesm8, lim6c, output)
    rows = _jsonl(output / "replicate_effect_summaries.jsonl")
    rep1 = next(
        row
        for row in rows
        if row["endpoint_id"] == "transcript_log2_totalrna_over_dna"
        and row["replicate_id"] == "rep1"
    )
    assert rep1["mutant_technical_barcode_count_retained"] == 22
    audit_rows = _jsonl(output / "BARCODE_RATIO_AUDIT.jsonl")
    minus_half = next(
        row
        for row in audit_rows
        if row["endpoint_id"] == "transcript_log2_totalrna_over_dna"
        and row["biological_replicate_id"] == "rep1"
        and row["physical_row_number"] == 2
    )
    assert minus_half["numerator_log2_cpm"] == pytest.approx(3.0)
    assert minus_half["denominator_log2_cpm"] == pytest.approx(-0.5)
    assert minus_half["barcode_log2_ratio"] == pytest.approx(3.5)
    assert minus_half["eligible"] is True
    assert minus_half["exclusion_reason"] is None

    minus_one = next(
        row
        for row in audit_rows
        if row["endpoint_id"] == "transcript_log2_totalrna_over_dna"
        and row["biological_replicate_id"] == "rep1"
        and row["physical_row_number"] == 3
    )
    assert minus_one["numerator_log2_cpm"] == pytest.approx(3.0)
    assert minus_one["denominator_log2_cpm"] == pytest.approx(-1.0)
    assert minus_one["barcode_log2_ratio"] == pytest.approx(4.0)
    assert minus_one["eligible"] is True
    assert minus_one["exclusion_reason"] is None
    assert report["companion_method"]["input_value_scale"] == (
        "PUBLISHED_LOG2_CPM_PER_BARCODE"
    )
    assert report["companion_method"][
        "published_log2_cpm_minimum_inclusive"
    ] == -1.0
    assert report["companion_method"]["original_cpm_minimum_inclusive"] == 0.5
    assert report["companion_method"]["clipping_allowed"] is False
    assert report["companion_method"]["missing_is_zero"] is False


def test_prefrozen_primary_floor_rejects_arm_with_only_nineteen_barcodes(
    tmp_path: Path,
) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path)
    for row_number in (2, 3, 4):
        _rewrite_cell(lim6c, row=row_number, column=4, value=None)
    output = tmp_path / "below-primary-floor"

    report = _run(moesm8, lim6c, output)
    records = _jsonl(output / "development_companion_effect_records.jsonl")
    transcript = next(
        row
        for row in records
        if row["endpoint_id"] == "transcript_log2_totalrna_over_dna"
    )
    assert transcript["effect_delta_mutant_minus_wt"] is None
    assert transcript["standard_error"] is None
    assert transcript["effective_n"] is None
    assert transcript["primary_technical_coverage_floor_per_arm_replicate"] == 20
    assert transcript["primary_technical_coverage_status"] == "REJECTED"
    summaries = _jsonl(output / "replicate_effect_summaries.jsonl")
    rep1 = next(
        row
        for row in summaries
        if row["endpoint_id"] == "transcript_log2_totalrna_over_dna"
        and row["replicate_id"] == "rep1"
    )
    assert rep1["mutant_technical_barcode_count_retained"] == 19
    assert rep1["replicate_rejection_reason"] == (
        "MUTANT_ARM_TECHNICAL_COVERAGE_BELOW_PREFROZEN_20"
    )
    assert report["coverage_threshold_prefrozen"] is True
    assert report["companion_method"][
        "primary_min_eligible_barcodes_per_arm_endpoint_replicate"
    ] == 20
    assert report["companion_method"]["secondary_audit_only_barcode_floors"] == [
        10,
        50,
    ]


def test_endpoint_is_rejected_when_effective_biological_n_is_not_three(
    tmp_path: Path,
) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path)
    # Remove every mutant transcript-ratio technical unit from rep3 only.
    for row_number in range(2, 24):
        _rewrite_cell(lim6c, row=row_number, column=8, value=None)
    output = tmp_path / "missing-replicate"

    report = _run(moesm8, lim6c, output)
    records = _jsonl(output / "development_companion_effect_records.jsonl")
    rows = _jsonl(output / "replicate_effect_summaries.jsonl")

    assert {row["endpoint_id"] for row in records} == {
        "transcript_log2_totalrna_over_dna",
        "te_log2_polysome_over_totalrna",
    }
    rejected_record = next(
        row
        for row in records
        if row["endpoint_id"] == "transcript_log2_totalrna_over_dna"
    )
    assert rejected_record["effect_delta_mutant_minus_wt"] is None
    assert rejected_record["standard_error"] is None
    assert rejected_record["effective_n"] is None
    assert rejected_record["primary_technical_coverage_status"] == "REJECTED"
    assert report["summary"]["development_companion_null_effect_record_count"] == 1
    assert report["summary"][
        "development_companion_nonnull_effect_record_count"
    ] == 1
    assert report["summary"]["endpoint_rejection_reason_counts"] == {
        "BIOLOGICAL_REPLICATE_COUNT_NOT_THREE": 1
    }
    rejected = next(
        row
        for row in rows
        if row["endpoint_id"] == "transcript_log2_totalrna_over_dna"
        and row["replicate_id"] == "rep3"
    )
    assert rejected["replicate_status"] == "REJECTED"
    assert rejected["endpoint_effective_n"] is None
    assert rejected["complete_biological_replicate_count_observed"] == 2
    assert rejected["required_biological_replicate_count"] == 3
    assert rejected["replicate_rejection_reason"] == (
        "NO_QUALIFYING_MUTANT_TECHNICAL_BARCODES"
    )


def test_missing_wt_is_rejected_with_aggregate_reason(tmp_path: Path) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path, include_wt=False)
    output = tmp_path / "missing-wt"

    report = _run(moesm8, lim6c, output)

    assert (
        output / "development_companion_effect_records.jsonl"
    ).read_bytes() == b""
    assert (output / "replicate_effect_summaries.jsonl").read_bytes() == b""
    assert (output / "BARCODE_RATIO_AUDIT.jsonl").read_bytes() == b""
    assert report["summary"]["sequence_pair_rejection_reason_counts"] == {
        "MISSING_WT_LIM6C_DESCRIPTION": 1
    }
    assert report["summary"]["endpoint_rejection_reason_counts"] == {
        "MISSING_WT_LIM6C_DESCRIPTION": 2
    }
    assert report["summary"]["canonical_record_count"] == 0
    assert report["summary"]["development_companion_effect_record_count"] == 0


def test_all_unique_descriptions_are_classified_without_join_shrinking_denominator(
    tmp_path: Path,
) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path)
    unadjudicated_description = "GENE1_COMPLEX_HAPLOTYPE_chr1_100_105"
    unadjudicated_barcode = "RAW_UNADJUDICATED_BARCODE_SECRET"
    workbook = load_workbook(lim6c)
    try:
        _append_6c_row(
            workbook[RECON.LIM6C_SHEET],
            description=unadjudicated_description,
            barcode=unadjudicated_barcode,
            transcript_ratios=(1.0, 1.0, 1.0),
            te_ratios=(1.0, 1.0, 1.0),
        )
        workbook.save(lim6c)
    finally:
        workbook.close()

    output = tmp_path / "description-universe"
    report = _run(moesm8, lim6c, output)
    rows = _jsonl(output / "DESCRIPTION_CLASSIFICATION_AUDIT.jsonl")

    assert len(rows) == 3
    assert {row["description"] for row in rows} == {
        MUTANT_DESCRIPTION,
        WT_DESCRIPTION,
        unadjudicated_description,
    }
    rejected = next(row for row in rows if row["description"] == unadjudicated_description)
    assert rejected["classification"] == "REJECTED_UNADJUDICATED_DESCRIPTION"
    assert rejected["classification_category"] == "rejected_with_reason"
    assert rejected["classification_reason"] == (
        "UNADJUDICATED_DESCRIPTION_CLASS_EXCLUDED"
    )
    assert report["summary"]["description_universe_denominator"] == 3
    assert report["summary"]["description_classification_complete"] is True
    assert report["summary"]["description_classification_counts"] == {
        "INCLUDED_STRICT_SNV_MUTANT": 1,
        "REFERENCE_ONLY_STRICT_WT": 1,
        "REJECTED_UNADJUDICATED_DESCRIPTION": 1,
    }
    assert sum(report["summary"]["description_classification_counts"].values()) == 3
    assert "UNADJUDICATED_DESCRIPTION_CLASSES_EXCLUDED" in report["blockers"]
    assert unadjudicated_barcode.encode("utf-8") not in b"".join(
        (output / name).read_bytes() for name in RECON.OUTPUT_FILES
    )


def test_nonstrict_and_blank_6a_rows_are_rejected_audited_and_do_not_abort(
    tmp_path: Path,
) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path)
    raw_nonstrict_coordinate = "NONSTRICT_COORDINATE_DO_NOT_EMIT"
    raw_nonstrict_sequence = "NONSTRICT_SEQUENCE_DO_NOT_EMIT"
    raw_nonstrict_gene = "UNVALIDATED_GENE_DO_NOT_EMIT"
    workbook = load_workbook(moesm8)
    try:
        sheet = workbook[RECON.MOESM8_SEQUENCE_SHEET]
        sheet.insert_rows(3)  # Preserve a fully blank physical row inside the table.
        sheet.append(
            [
                "NOT_A_POSITIVE_INTEGER",
                raw_nonstrict_gene,
                raw_nonstrict_coordinate,
                raw_nonstrict_sequence,
            ]
        )
        workbook.save(moesm8)
    finally:
        workbook.close()

    output = tmp_path / "moesm8-6a-universe"
    report = _run(moesm8, lim6c, output)
    rows = _jsonl(output / "MOESM8_6A_CLASSIFICATION_AUDIT.jsonl")

    assert len(rows) == 4
    assert {row["physical_row_number"] for row in rows} == {2, 3, 4, 5}
    assert {row["classification"] for row in rows} == set(
        RECON.MOESM8_6A_ALLOWED_CLASSES
    )
    blank = next(row for row in rows if row["physical_row_number"] == 3)
    assert blank["classification"] == "REJECTED_EMPTY_ROW"
    assert blank["classification_reason"] == "EMPTY_PHYSICAL_DATA_ROW"
    assert blank["coordinate_value_type"] == "EMPTY"
    assert blank["coordinate_signature"] == "EMPTY_PHYSICAL_DATA_ROW"

    rejected = next(row for row in rows if row["physical_row_number"] == 5)
    assert rejected["classification"] == "REJECTED_UNADJUDICATED_COORDINATE"
    assert rejected["classification_reason"] == (
        "COORDINATE_NOT_STRICT_MUTANT_OR_WT_REGEX"
    )
    assert rejected["coordinate_value_type"] == "STRING"
    assert rejected["coordinate_signature"] == (
        "NONSTRICT_STRING_FIELDS_5_EXACT_UTR5_SUFFIX_NO"
    )
    expected_coordinate_hash = hashlib.sha256(
        (
            f"{_sha256(moesm8)}\0{RECON.MOESM8_SEQUENCE_SHEET}\0"
            "physical_row=5\0coordinate_type=STRING\0"
            f"{raw_nonstrict_coordinate}"
        ).encode("utf-8")
    ).hexdigest()
    assert rejected["coordinate_sha256"] == expected_coordinate_hash
    assert len(rejected["row_locator_sha256"]) == 64
    assert all(
        row["raw_coordinate_emitted"] is False
        and row["raw_sequence_emitted"] is False
        and "coordinate" not in row
        and "sequence" not in row
        for row in rows
    )

    counts = report["summary"]["moesm8_6a_classification_counts"]
    assert counts == {
        "INCLUDED_STRICT_SNV_MUTANT_COORDINATE": 1,
        "REFERENCE_ONLY_STRICT_WT_COORDINATE": 1,
        "REJECTED_UNADJUDICATED_COORDINATE": 1,
        "REJECTED_EMPTY_ROW": 1,
    }
    assert report["summary"]["moesm8_6a_physical_data_row_denominator"] == 4
    assert report["workbook_structure"]["moesm8_6a_row_count"] == 4
    assert sum(counts.values()) == 4
    assert report["summary"]["development_companion_effect_record_count"] == 2
    assert report["summary"]["canonical_record_count"] == 0
    assert report["canonical_intervention_record_v3_materialized"] is False
    assert report["qualification"] is False
    assert "UNADJUDICATED_6A_COORDINATE_CLASSES_EXCLUDED" in report["blockers"]

    all_output = b"".join(
        (output / name).read_bytes() for name in RECON.OUTPUT_FILES
    )
    assert raw_nonstrict_coordinate.encode("utf-8") not in all_output
    assert raw_nonstrict_sequence.encode("utf-8") not in all_output
    assert raw_nonstrict_gene.encode("utf-8") not in all_output


def test_candidate_must_be_exactly_one_snv(tmp_path: Path) -> None:
    moesm8, lim6c = _fixture_inputs(
        tmp_path,
        candidate_sequence="TATCGG",  # source differs at indices 0 and 2
    )
    output = tmp_path / "two-edits"

    report = _run(moesm8, lim6c, output)

    assert report["summary"]["canonical_record_count"] == 0
    assert report["summary"]["development_companion_effect_record_count"] == 0
    assert report["summary"]["sequence_pair_rejection_reason_counts"] == {
        "NOT_EXACTLY_ONE_SNV": 1
    }


def test_sequence_snv_must_agree_with_description_ref_alt(tmp_path: Path) -> None:
    moesm8, lim6c = _fixture_inputs(
        tmp_path,
        source_sequence="AAACGG",
        candidate_sequence="AATCGG",
        ref="C",
        alt="T",
    )
    output = tmp_path / "allele-mismatch"

    report = _run(moesm8, lim6c, output)

    assert report["summary"]["sequence_pair_rejection_reason_counts"] == {
        "SNV_SEQUENCE_REF_ALT_MISMATCH": 1
    }
    assert report["summary"]["canonical_record_count"] == 0
    assert report["summary"]["development_companion_effect_record_count"] == 0


def test_missing_required_column_fails_exact_header_validation(tmp_path: Path) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path)
    workbook = load_workbook(lim6c)
    try:
        workbook[RECON.LIM6C_SHEET].cell(row=1, column=17).value = None
        workbook.save(lim6c)
    finally:
        workbook.close()

    with pytest.raises(RECON.ReconstructionError, match="exact raw header mismatch"):
        _run(moesm8, lim6c, tmp_path / "bad-header")


def test_totalrna_must_match_by_key_across_shuffled_blocks(tmp_path: Path) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path)
    _rewrite_cell(lim6c, row=2, column=12, value=999.0)

    with pytest.raises(
        RECON.ReconstructionError, match="key-aligned TotalRNA mismatch"
    ):
        _run(moesm8, lim6c, tmp_path / "key-aligned-mismatch")


def test_right_blank_tail_and_left_right_only_keys_stay_null_not_zero(
    tmp_path: Path,
) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path)
    left_only_barcode = "RAW_LEFT_ONLY_BARCODE_SECRET"
    right_only_barcode = "RAW_RIGHT_ONLY_BARCODE_SECRET"
    left_values, _ = _lim6c_block_values(
        description=MUTANT_DESCRIPTION,
        barcode=left_only_barcode,
        transcript_ratios=(5.0, 7.0, 10.0),
        te_ratios=(4.0, 5.0, 6.0),
    )
    workbook = load_workbook(lim6c)
    try:
        sheet = workbook[RECON.LIM6C_SHEET]
        # Replace one valid WT key in J:Q only. The replaced key remains
        # transcript-only and the new key is TE-only for the included WT arm.
        sheet.cell(row=2, column=11).value = right_only_barcode
        sheet.append(
            [*left_values, None, *([None] * 8)]
        )
        workbook.save(lim6c)
    finally:
        workbook.close()

    output = tmp_path / "right-blank-tail"
    report = _run(moesm8, lim6c, output)
    structure = report["workbook_structure"]["lim6c_block_semantics"]
    assert structure["physical_data_rows"] == 45
    assert structure["transcript_present_key_rows"] == 45
    assert structure["translation_efficiency_present_key_rows"] == 44
    assert structure["translation_efficiency_blank_tail_rows"] == 1
    assert structure["hashed_key_intersection"] == 43
    assert structure["hashed_key_transcript_only"] == 2
    assert structure["hashed_key_translation_efficiency_only"] == 1

    left_only_hash = RECON._salted_barcode_sha256(
        left_only_barcode,
        lim6c_sha256=_sha256(lim6c),
    )
    audit_rows = _jsonl(output / "BARCODE_RATIO_AUDIT.jsonl")
    te_missing = [
        row
        for row in audit_rows
        if row["barcode_id_sha256"] == left_only_hash
        and row["endpoint_id"] == "te_log2_polysome_over_totalrna"
    ]
    assert len(te_missing) == 3
    assert all(
        row["key_present_in_endpoint_block"] is False
        and row["numerator_log2_cpm"] is None
        and row["denominator_log2_cpm"] is None
        and row["barcode_log2_ratio"] is None
        and row["exclusion_reason"] == "KEY_MISSING_FROM_ENDPOINT_BLOCK_NOT_ZERO"
        and row["physical_row_number"] is None
        and row["row_locator_sha256"] is None
        for row in te_missing
    )
    right_only_hash = RECON._salted_barcode_sha256(
        right_only_barcode,
        lim6c_sha256=_sha256(lim6c),
    )
    transcript_missing = [
        row
        for row in audit_rows
        if row["barcode_id_sha256"] == right_only_hash
        and row["endpoint_id"] == "transcript_log2_totalrna_over_dna"
    ]
    assert len(transcript_missing) == 3
    assert all(
        row["source_block_id"] == RECON.TRANSCRIPT_BLOCK_ID
        and row["key_present_in_endpoint_block"] is False
        and row["numerator_log2_cpm"] is None
        and row["denominator_log2_cpm"] is None
        and row["barcode_log2_ratio"] is None
        and row["exclusion_reason"] == "KEY_MISSING_FROM_ENDPOINT_BLOCK_NOT_ZERO"
        and row["physical_row_number"] is None
        and row["row_locator_sha256"] is None
        for row in transcript_missing
    )
    te_present = [
        row
        for row in audit_rows
        if row["barcode_id_sha256"] == right_only_hash
        and row["endpoint_id"] == "te_log2_polysome_over_totalrna"
    ]
    assert len(te_present) == 3
    assert all(
        row["source_block_id"] == RECON.TRANSLATION_EFFICIENCY_BLOCK_ID
        and row["key_present_in_endpoint_block"] is True
        and row["physical_row_number"] == 2
        and row["row_locator_sha256"] is not None
        for row in te_present
    )
    all_output = b"".join(
        (output / name).read_bytes() for name in RECON.OUTPUT_FILES
    )
    assert left_only_barcode.encode("utf-8") not in all_output
    assert right_only_barcode.encode("utf-8") not in all_output


def test_endpoint_arm_missing_after_outer_inventory_is_rejected_not_copied(
    tmp_path: Path,
) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path)
    te_only_description = "TE_ONLY_UNADJUDICATED_DESCRIPTION"
    workbook = load_workbook(lim6c)
    try:
        sheet = workbook[RECON.LIM6C_SHEET]
        # In the reversed right block, physical rows 24:45 contain all mutant
        # keys. Move those keys into a distinct TE-only description without
        # changing their published log2CPM values.
        for physical_row in range(24, 46):
            sheet.cell(row=physical_row, column=10).value = te_only_description
        workbook.save(lim6c)
    finally:
        workbook.close()

    output = tmp_path / "endpoint-arm-missing"
    report = _run(moesm8, lim6c, output)
    structure = report["workbook_structure"]["lim6c_block_semantics"]
    assert structure["hashed_key_intersection"] == 22
    assert structure["hashed_key_transcript_only"] == 22
    assert structure["hashed_key_translation_efficiency_only"] == 22

    records = _jsonl(output / "development_companion_effect_records.jsonl")
    te_record = next(
        row
        for row in records
        if row["endpoint_id"] == "te_log2_polysome_over_totalrna"
    )
    assert te_record["effect_delta_mutant_minus_wt"] is None
    assert te_record["standard_error"] is None
    assert te_record["effective_n"] is None

    summaries = _jsonl(output / "replicate_effect_summaries.jsonl")
    te_summaries = [
        row
        for row in summaries
        if row["endpoint_id"] == "te_log2_polysome_over_totalrna"
    ]
    assert len(te_summaries) == 3
    assert all(
        row["replicate_rejection_reason"]
        == "MUTANT_ARM_MISSING_FROM_ENDPOINT_BLOCK"
        and row["mutant_technical_barcode_count_total"] == 0
        and row["mutant_endpoint_block_missing_key_count"] == 22
        for row in te_summaries
    )
    missing_audit = [
        row
        for row in _jsonl(output / "BARCODE_RATIO_AUDIT.jsonl")
        if row["arm"] == "MUTANT"
        and row["endpoint_id"] == "te_log2_polysome_over_totalrna"
    ]
    assert len(missing_audit) == 66
    assert all(
        row["key_present_in_endpoint_block"] is False
        and row["numerator_log2_cpm"] is None
        and row["denominator_log2_cpm"] is None
        and row["barcode_log2_ratio"] is None
        and row["exclusion_reason"] == "KEY_MISSING_FROM_ENDPOINT_BLOCK_NOT_ZERO"
        for row in missing_audit
    )
    description_rows = _jsonl(output / "DESCRIPTION_CLASSIFICATION_AUDIT.jsonl")
    te_only = next(
        row for row in description_rows if row["description"] == te_only_description
    )
    assert te_only["transcript_block_key_count"] == 0
    assert te_only["translation_efficiency_block_key_count"] == 22
    assert te_only["classification"] == "REJECTED_UNADJUDICATED_DESCRIPTION"


def test_duplicate_key_within_either_endpoint_block_fails_closed(
    tmp_path: Path,
) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path)
    workbook = load_workbook(lim6c)
    try:
        sheet = workbook[RECON.LIM6C_SHEET]
        sheet.cell(row=3, column=10).value = sheet.cell(row=2, column=10).value
        sheet.cell(row=3, column=11).value = sheet.cell(row=2, column=11).value
        workbook.save(lim6c)
    finally:
        workbook.close()

    with pytest.raises(RECON.ReconstructionError, match="duplicate description/barcode"):
        _run(moesm8, lim6c, tmp_path / "duplicate-key")


def test_published_log2_cpm_below_minus_one_fails_closed(tmp_path: Path) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path)
    _rewrite_cell(lim6c, row=2, column=12, value=-1.000001)

    with pytest.raises(
        RECON.ReconstructionError,
        match="log2CPM below the inclusive -1 minimum",
    ):
        _run(moesm8, lim6c, tmp_path / "below-log2-cpm-floor")


@pytest.mark.parametrize(
    ("invalid_value", "error_pattern"),
    [
        (True, "contains non-numeric log2CPM"),
        ("0", "contains non-numeric log2CPM"),
        ("=1+1", "contains a forbidden formula cell"),
    ],
)
def test_present_log2_cpm_cell_rejects_bool_string_and_formula(
    tmp_path: Path,
    invalid_value: Any,
    error_pattern: str,
) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path)
    _rewrite_cell(lim6c, row=2, column=12, value=invalid_value)

    with pytest.raises(RECON.ReconstructionError, match=error_pattern):
        _run(moesm8, lim6c, tmp_path / "invalid-log2-cpm-cell")


def test_input_hash_drift_fails_before_any_workbook_read(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path)

    def forbidden_read(*args: Any, **kwargs: Any) -> Any:
        raise AssertionError("workbook parser must not run before both hashes pass")

    monkeypatch.setattr(RECON, "load_workbook", forbidden_read)
    monkeypatch.setattr(RECON.pd, "read_excel", forbidden_read)
    with pytest.raises(RECON.ReconstructionError, match="MOESM8 workbook SHA-256 mismatch"):
        RECON.reconstruct_gse149487_plumage(
            moesm8_path=moesm8,
            lim6c_path=lim6c,
            protocol_path=_protocol_path(moesm8),
            expected_moesm8_sha256="0" * 64,
            expected_lim6c_sha256=_sha256(lim6c),
            expected_protocol_sha256=_sha256(_protocol_path(moesm8)),
            output_directory=tmp_path / "hash-drift",
        )


def test_protocol_hash_bound_semantics_fail_before_workbook_read(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path)
    protocol_path = _protocol_path(moesm8)
    protocol = json.loads(protocol_path.read_text(encoding="utf-8"))
    protocol["route_a_companion_summary"][
        "minimum_eligible_barcodes_per_arm_per_endpoint_per_replicate"
    ] = 19
    protocol_path.write_text(json.dumps(protocol) + "\n", encoding="utf-8")

    def forbidden_workbook_read(*args: Any, **kwargs: Any) -> Any:
        raise AssertionError("protocol semantics must fail before workbook parsing")

    monkeypatch.setattr(RECON, "load_workbook", forbidden_workbook_read)
    monkeypatch.setattr(RECON.pd, "read_excel", forbidden_workbook_read)
    with pytest.raises(RECON.ReconstructionError, match="protocol field mismatch"):
        RECON.reconstruct_gse149487_plumage(
            moesm8_path=moesm8,
            lim6c_path=lim6c,
            protocol_path=protocol_path,
            expected_moesm8_sha256=_sha256(moesm8),
            expected_lim6c_sha256=_sha256(lim6c),
            expected_protocol_sha256=_sha256(protocol_path),
            output_directory=tmp_path / "protocol-drift",
        )


def test_protocol_published_log2_cpm_scale_is_strictly_bound_before_workbook_read(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path)
    protocol_path = _protocol_path(moesm8)
    protocol = json.loads(protocol_path.read_text(encoding="utf-8"))
    protocol["paper_faithful_measurement_transform"][
        "published_log2_cpm_minimum_inclusive"
    ] = 0.0
    protocol_path.write_text(json.dumps(protocol) + "\n", encoding="utf-8")

    def forbidden_workbook_read(*args: Any, **kwargs: Any) -> Any:
        raise AssertionError("log2CPM scale drift must fail before workbook parsing")

    monkeypatch.setattr(RECON, "load_workbook", forbidden_workbook_read)
    with pytest.raises(RECON.ReconstructionError, match="protocol field mismatch"):
        RECON.reconstruct_gse149487_plumage(
            moesm8_path=moesm8,
            lim6c_path=lim6c,
            protocol_path=protocol_path,
            expected_moesm8_sha256=_sha256(moesm8),
            expected_lim6c_sha256=_sha256(lim6c),
            expected_protocol_sha256=_sha256(protocol_path),
            output_directory=tmp_path / "protocol-log2-cpm-scale-drift",
        )


def test_protocol_companion_and_canonical_boundary_is_strictly_bound(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path)
    protocol_path = _protocol_path(moesm8)
    protocol = json.loads(protocol_path.read_text(encoding="utf-8"))
    protocol["companion_lineage"][
        "canonical_intervention_record_v3_materialized"
    ] = True
    protocol_path.write_text(json.dumps(protocol) + "\n", encoding="utf-8")

    def forbidden_workbook_read(*args: Any, **kwargs: Any) -> Any:
        raise AssertionError("lineage boundary must fail before workbook parsing")

    monkeypatch.setattr(RECON, "load_workbook", forbidden_workbook_read)
    monkeypatch.setattr(RECON.pd, "read_excel", forbidden_workbook_read)
    with pytest.raises(RECON.ReconstructionError, match="protocol field mismatch"):
        RECON.reconstruct_gse149487_plumage(
            moesm8_path=moesm8,
            lim6c_path=lim6c,
            protocol_path=protocol_path,
            expected_moesm8_sha256=_sha256(moesm8),
            expected_lim6c_sha256=_sha256(lim6c),
            expected_protocol_sha256=_sha256(protocol_path),
            output_directory=tmp_path / "protocol-lineage-drift",
        )


def test_protocol_moesm8_6a_universe_is_strictly_bound_before_workbook_read(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path)
    protocol_path = _protocol_path(moesm8)
    protocol = json.loads(protocol_path.read_text(encoding="utf-8"))
    protocol["moesm8_6a_universe"][
        "successful_join_may_shrink_physical_row_denominator"
    ] = True
    protocol_path.write_text(json.dumps(protocol) + "\n", encoding="utf-8")

    def forbidden_workbook_read(*args: Any, **kwargs: Any) -> Any:
        raise AssertionError("6a universe drift must fail before workbook parsing")

    monkeypatch.setattr(RECON, "load_workbook", forbidden_workbook_read)
    monkeypatch.setattr(RECON.pd, "read_excel", forbidden_workbook_read)
    with pytest.raises(RECON.ReconstructionError, match="protocol field mismatch"):
        RECON.reconstruct_gse149487_plumage(
            moesm8_path=moesm8,
            lim6c_path=lim6c,
            protocol_path=protocol_path,
            expected_moesm8_sha256=_sha256(moesm8),
            expected_lim6c_sha256=_sha256(lim6c),
            expected_protocol_sha256=_sha256(protocol_path),
            output_directory=tmp_path / "protocol-6a-universe-drift",
        )


def test_protocol_lim6c_independent_block_semantics_are_strictly_bound(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path)
    protocol_path = _protocol_path(moesm8)
    protocol = json.loads(protocol_path.read_text(encoding="utf-8"))
    protocol["lim6c_block_semantics"]["row_position_may_join_blocks"] = True
    protocol_path.write_text(json.dumps(protocol) + "\n", encoding="utf-8")

    def forbidden_workbook_read(*args: Any, **kwargs: Any) -> Any:
        raise AssertionError("Lim 6c block semantics must fail before workbook parsing")

    monkeypatch.setattr(RECON, "load_workbook", forbidden_workbook_read)
    monkeypatch.setattr(RECON.pd, "read_excel", forbidden_workbook_read)
    with pytest.raises(RECON.ReconstructionError, match="protocol field mismatch"):
        RECON.reconstruct_gse149487_plumage(
            moesm8_path=moesm8,
            lim6c_path=lim6c,
            protocol_path=protocol_path,
            expected_moesm8_sha256=_sha256(moesm8),
            expected_lim6c_sha256=_sha256(lim6c),
            expected_protocol_sha256=_sha256(protocol_path),
            output_directory=tmp_path / "protocol-lim6c-block-drift",
        )


def test_workbook_basenames_fail_before_payload_read(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path)

    def forbidden_hash(*args: Any, **kwargs: Any) -> Any:
        raise AssertionError("basename mismatch must fail before payload hashing")

    monkeypatch.setattr(RECON, "_sha256_and_size", forbidden_hash)
    renamed_moesm8 = tmp_path / "renamed-moesm8.xlsx"
    moesm8.rename(renamed_moesm8)
    with pytest.raises(RECON.ReconstructionError, match="MOESM8 workbook basename mismatch"):
        RECON.reconstruct_gse149487_plumage(
            moesm8_path=renamed_moesm8,
            lim6c_path=lim6c,
            protocol_path=_protocol_path(lim6c),
            expected_moesm8_sha256="0" * 64,
            expected_lim6c_sha256="0" * 64,
            expected_protocol_sha256="0" * 64,
            output_directory=tmp_path / "bad-moesm8-basename",
        )

    renamed_moesm8.rename(moesm8)
    renamed_lim6c = tmp_path / "renamed-lim6c.xlsx"
    lim6c.rename(renamed_lim6c)
    with pytest.raises(RECON.ReconstructionError, match="Lim 6c workbook basename mismatch"):
        RECON.reconstruct_gse149487_plumage(
            moesm8_path=moesm8,
            lim6c_path=renamed_lim6c,
            protocol_path=_protocol_path(moesm8),
            expected_moesm8_sha256="0" * 64,
            expected_lim6c_sha256="0" * 64,
            expected_protocol_sha256="0" * 64,
            output_directory=tmp_path / "bad-lim6c-basename",
        )


def test_case_insensitive_forbidden_path_fails_before_payload_read(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _, lim6c = _fixture_inputs(tmp_path)

    def forbidden_hash(*args: Any, **kwargs: Any) -> Any:
        raise AssertionError("forbidden path must fail before payload hashing")

    monkeypatch.setattr(RECON, "_sha256_and_size", forbidden_hash)
    with pytest.raises(RECON.ScopeViolation, match="rejected before read"):
        RECON.reconstruct_gse149487_plumage(
            moesm8_path=tmp_path / "SeAlEd" / "input.xlsx",
            lim6c_path=lim6c,
            protocol_path=_protocol_path(lim6c),
            expected_moesm8_sha256="0" * 64,
            expected_lim6c_sha256="0" * 64,
            expected_protocol_sha256=_sha256(_protocol_path(lim6c)),
            output_directory=tmp_path / "safe-output",
        )


def test_input_and_output_symlinks_are_rejected(tmp_path: Path) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path)
    moesm8_link = tmp_path / "moesm8-link.xlsx"
    moesm8_link.symlink_to(moesm8)

    with pytest.raises(RECON.ReconstructionError, match="symlink"):
        RECON.reconstruct_gse149487_plumage(
            moesm8_path=moesm8_link,
            lim6c_path=lim6c,
            protocol_path=_protocol_path(moesm8),
            expected_moesm8_sha256=_sha256(moesm8),
            expected_lim6c_sha256=_sha256(lim6c),
            expected_protocol_sha256=_sha256(_protocol_path(moesm8)),
            output_directory=tmp_path / "input-link-out",
        )

    real_output_parent = tmp_path / "real-output-parent"
    real_output_parent.mkdir()
    output_parent_link = tmp_path / "output-parent-link"
    output_parent_link.symlink_to(real_output_parent, target_is_directory=True)
    with pytest.raises(RECON.ReconstructionError, match="symlink"):
        RECON.reconstruct_gse149487_plumage(
            moesm8_path=moesm8,
            lim6c_path=lim6c,
            protocol_path=_protocol_path(moesm8),
            expected_moesm8_sha256=_sha256(moesm8),
            expected_lim6c_sha256=_sha256(lim6c),
            expected_protocol_sha256=_sha256(_protocol_path(moesm8)),
            output_directory=output_parent_link / "out",
        )


def test_existing_output_is_preserved_without_workbook_parse(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path)
    output = tmp_path / "existing"
    output.mkdir()
    marker = output / "preserve.txt"
    marker.write_bytes(b"preserve-me")

    def forbidden_read(*args: Any, **kwargs: Any) -> Any:
        raise AssertionError("existing output must fail before workbook parse")

    monkeypatch.setattr(RECON, "load_workbook", forbidden_read)
    with pytest.raises(RECON.ReconstructionError, match="refusing to overwrite"):
        _run(moesm8, lim6c, output)
    assert marker.read_bytes() == b"preserve-me"
    assert not (tmp_path / ".existing.publish-claim").exists()
    assert not list(tmp_path.glob(".existing.partial-staging-*"))


def test_staging_write_failure_never_creates_final_and_retains_partial_evidence(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path)
    output = tmp_path / "injected-write-failure"
    original_write = RECON._write_exclusive

    def injected_failure(path: Path, payload: bytes) -> None:
        if path.name == "replicate_effect_summaries.jsonl":
            raise OSError("injected staging write failure")
        original_write(path, payload)

    monkeypatch.setattr(RECON, "_write_exclusive", injected_failure)
    with pytest.raises(OSError, match="injected staging write failure"):
        _run(moesm8, lim6c, output)

    assert not output.exists()
    claim = tmp_path / ".injected-write-failure.publish-claim"
    partials = list(tmp_path.glob(".injected-write-failure.partial-staging-*"))
    assert claim.is_file()
    assert len(partials) == 1
    assert partials[0].is_dir()
    assert (
        partials[0] / "development_companion_effect_records.jsonl"
    ).is_file()
    assert not (partials[0] / "RECONSTRUCTION_REPORT.json").exists()
    claim_payload = json.loads(claim.read_text(encoding="utf-8"))
    assert claim_payload["status"] == "PARTIAL_OUTPUT_STAGING_ACTIVE_NOT_FINAL"
    assert claim_payload["staging_directory_name"] == partials[0].name


def test_development_companion_schema_does_not_impersonate_canonical_v3(
    tmp_path: Path,
) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path)
    output = tmp_path / "schema-boundary"

    report = _run(moesm8, lim6c, output)
    companion_rows = _jsonl(
        output / "development_companion_effect_records.jsonl"
    )

    assert companion_rows
    assert "canonical_intervention_records.jsonl" not in RECON.OUTPUT_FILES
    assert not (output / "canonical_intervention_records.jsonl").exists()
    assert all(
        row["schema_id"] != "CanonicalInterventionRecordV3"
        and row["record_type"] == "DEVELOPMENT_COMPANION_EFFECT_SUMMARY"
        and row["canonical_intervention_record_v3_materialized"] is False
        for row in companion_rows
    )
    assert report["canonical_intervention_record_v3_materialized"] is False
    assert report["summary"]["canonical_record_count"] == 0
    assert report["summary"]["development_companion_effect_record_count"] == 2
    assert "CANONICAL_INTERVENTION_RECORD_V3_NOT_MATERIALIZED" in report["blockers"]


def test_deterministic_outputs_checksums_and_no_raw_barcode_ids(tmp_path: Path) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path)
    first = tmp_path / "first"
    second = tmp_path / "second"

    first_report = _run(moesm8, lim6c, first)
    second_report = _run(moesm8, lim6c, second)

    assert first_report == second_report
    assert not list(tmp_path.glob(".*.publish-claim"))
    assert not list(tmp_path.glob(".*.partial-staging-*"))
    for name in RECON.OUTPUT_FILES:
        assert (first / name).read_bytes() == (second / name).read_bytes()

    all_output = b"".join((first / name).read_bytes() for name in RECON.OUTPUT_FILES)
    for raw_barcode in RAW_BARCODE_IDS:
        assert raw_barcode.encode("utf-8") not in all_output

    expected_sums = {
        name: _sha256(first / name)
        for name in RECON.OUTPUT_FILES
        if name != "SHA256SUMS"
    }
    observed_sums = {
        line.split("  ", 1)[1]: line.split("  ", 1)[0]
        for line in (first / "SHA256SUMS").read_text(encoding="ascii").splitlines()
    }
    assert observed_sums == expected_sums


def test_closed_qualification_tokens_and_aggregate_only_stdout(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    moesm8, lim6c = _fixture_inputs(tmp_path)
    output = tmp_path / "cli"

    result = RECON.main(
        [
            "--moesm8",
            str(moesm8),
            "--moesm8-sha256",
            _sha256(moesm8),
            "--lim6c",
            str(lim6c),
            "--lim6c-sha256",
            _sha256(lim6c),
            "--protocol",
            str(_protocol_path(moesm8)),
            "--protocol-sha256",
            _sha256(_protocol_path(moesm8)),
            "--output-dir",
            str(output),
        ]
    )

    assert result == 0
    stdout = capsys.readouterr().out
    summary = json.loads(stdout)
    assert summary == {
        "canonical_intervention_record_v3_materialized": False,
        "canonical_record_count": 0,
        "companion_summary_only": True,
        "dataset_id": "GSE149487",
        "development_companion_effect_record_count": 2,
        "qualified": False,
        "reconstruction_status": "DEVELOPMENT_RECONSTRUCTED_NOT_QUALIFIED",
        "rejected_endpoint_pair_count": 0,
    }
    assert SOURCE_SEQUENCE not in stdout
    assert MUTANT_DESCRIPTION not in stdout
    for raw_barcode in RAW_BARCODE_IDS:
        assert raw_barcode not in stdout

    report = json.loads((output / "RECONSTRUCTION_REPORT.json").read_text())
    for key in (
        "qualification",
        "qualified",
        "training",
        "training_allowed",
        "model_selection",
        "model_selection_allowed",
        "paper_method_reproduced",
        "whole_study_context_closed",
    ):
        assert report[key] is False
    assert report["coverage_threshold_prefrozen"] is True
    assert report["coverage_threshold_freeze_timing"] == (
        "PREFROZEN_BEFORE_REAL_RECONSTRUCTION_RESULTS"
    )
    assert report["protocol"]["strict_runtime_constant_binding"] is True
    assert report["canonical_intervention_record_v3_materialized"] is False
    assert report["summary"]["canonical_record_count"] == 0
    assert report["summary"]["development_companion_effect_record_count"] == 2
    assert report["input_workbooks"]["moesm8"] == {
        "actual_filename": RECON.EXPECTED_MOESM8_FILENAME,
        "bytes": report["input_workbooks"]["moesm8"]["bytes"],
        "expected_filename": RECON.EXPECTED_MOESM8_FILENAME,
        "filename_match": True,
        "sequence_sheet": RECON.MOESM8_SEQUENCE_SHEET,
        "sha256": _sha256(moesm8),
        "source_asset_id": RECON.MOESM8_SOURCE_ASSET_ID,
    }
    assert report["input_workbooks"]["lim6c_293t"]["actual_filename"] == (
        RECON.EXPECTED_LIM6C_FILENAME
    )
    assert report["input_workbooks"]["lim6c_293t"]["expected_filename"] == (
        RECON.EXPECTED_LIM6C_FILENAME
    )
    assert report["input_workbooks"]["lim6c_293t"]["filename_match"] is True
    assert report["license"] == "UNKNOWN_BLOCKED"
    assert report["eligibility"] == "PENDING_BLOCKED"
    assert report["exposure"] == "DEVELOPMENT_ONLY"
    assert report["split"] == "EXCLUDED"
    assert report["split_execution"] == "NOT_RUN"
    assert report["scope"]["not_included"] == [
        "PC3",
        "18_GEO_RAW_COUNT_TABLES",
        "MOESM3",
        "PUBLISHED_6D_6E_LFC_DESCRIPTIVE_COMPARISON",
    ]
