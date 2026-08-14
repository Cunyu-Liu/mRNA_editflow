from __future__ import annotations

import copy
import hashlib
import importlib.util
import inspect
import json
import sys
from pathlib import Path

import openpyxl
import pytest


ROOT = Path(__file__).resolve().parents[2]
CONFIG_PATH = ROOT / (
    "configs/route_a_v3_encsr854ruf_dec027_dataset_specific_a1_preflight_v1.json"
)
MODULE_PATH = ROOT / (
    "scripts/route_a_v3/preflight_encsr854ruf_dec027_dataset_specific_a1.py"
)
SPEC = importlib.util.spec_from_file_location("encsr854ruf_dec027_preflight", MODULE_PATH)
assert SPEC is not None and SPEC.loader is not None
PREFLIGHT = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = PREFLIGHT
SPEC.loader.exec_module(PREFLIGHT)


def _disk_protocol() -> dict:
    return PREFLIGHT.load_protocol(CONFIG_PATH)


def _protocol() -> dict:
    """Return the clean I2 fixture from either a legal disk I2 or disk B2."""

    protocol = PREFLIGHT._normalise_own_binding(_disk_protocol())
    PREFLIGHT.validate_protocol(protocol)
    return protocol


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _bind_own(protocol: dict) -> dict:
    value = copy.deepcopy(protocol)
    own = value["implementation_binding"]["own_preflight_group"]
    own.update(
        {
            "status": PREFLIGHT.BOUND,
            "implementation_commit": "3" * 40,
            "implementation_script_sha256": "c" * 64,
            "implementation_test_sha256": "d" * 64,
        }
    )
    PREFLIGHT.validate_protocol(value)
    return value


def _sequence(first: str) -> str:
    return first + "A" * 132


def _make_assets(
    tmp_path: Path,
    protocol: dict,
    *,
    omit_alt_member: bool = False,
    omit_alt_fasta: bool = False,
) -> tuple[Path, Path, dict]:
    author = tmp_path / "author"
    data = author / "data"
    scripts = author / "scripts"
    data.mkdir(parents=True)
    scripts.mkdir(parents=True)

    reporter_ids = (
        "SECRET_REPORTER_REF_A",
        "SECRET_REPORTER_ALT_A",
        "SECRET_REPORTER_REF_B",
        "SECRET_REPORTER_ALT_B",
    )
    sequences = (
        _sequence("A"),
        _sequence("C"),
        _sequence("G"),
        _sequence("T"),
    )
    fasta_entries = [
        f">{reporter_ids[0]}/SECRET_UNUSED_ALIAS\n{sequences[0]}\n",
        f">{reporter_ids[1]}\n{sequences[1]}\n",
        f">{reporter_ids[2]}\n{sequences[2]}\n",
    ]
    if not omit_alt_fasta:
        fasta_entries.append(f">{reporter_ids[3]}\n{sequences[3]}\n")
    fasta_path = data / "GWASrewritepos_CMS_alignment_file.fasta"
    fasta_path.write_text("".join(fasta_entries), encoding="utf-8")

    array_path = data / "GWASrewritepos_CMS_arrayassign"
    array_rows = ["oligo_name\ttag\tdup\tref_name\tCMS_array\tGWAS_array"]
    for index, reporter_id in enumerate(reporter_ids):
        array_rows.append(f"PRIVATE_{index}\tref\t0\t{reporter_id}\t1\t1")
    array_path.write_text("\n".join(array_rows) + "\n", encoding="utf-8")
    r_path = scripts / "generate_processed_DESeq2_data.R"
    r_path.write_text("# synthetic public-method fixture\n", encoding="utf-8")

    workbook_path = tmp_path / "publisher.xlsx"
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    workbook.create_sheet("README").append(["aggregate synthetic fixture"])

    result = workbook.create_sheet("Variant MPRAu Results")
    result_headers = ["mpra_variant_id"]
    for context in PREFLIGHT.CONTEXTS:
        result_headers.extend(
            [f"log2FoldChange_Skew_{context}", f"lfcSE_Skew_{context}"]
        )
    result.append(result_headers)
    result.append(["SECRET_PAIR_A"] + [1.0, 0.1] * 6)
    result.append(["SECRET_PAIR_B", "NA", 0.2] + [2.0, 0.2] * 5)

    oligo = workbook.create_sheet("Oligo Variant Info")
    oligo.append(
        [
            "mpra_variant_id",
            "tag",
            "oligo_id",
            "ref_allele",
            "alt_allele",
            "other_var_in_oligo_window",
        ]
    )
    oligo.append(["SECRET_PAIR_A", "ref", reporter_ids[0], "A", "C", "NA"])
    oligo.append(["SECRET_PAIR_A", "alt", reporter_ids[1], "A", "C", "NA"])
    oligo.append(["SECRET_PAIR_B", "ref", reporter_ids[2], "G", "T", "index_error"])
    if not omit_alt_member:
        oligo.append(["SECRET_PAIR_B", "alt", reporter_ids[3], "G", "T", "index_error"])

    for sheet_name in (
        "Raw Counts HEK CMS",
        "Raw Counts HEK GWAS",
        "Raw Counts HEK Remaining",
    ):
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(["oligo_id", "aggregate_R1"])
        for reporter_id in reporter_ids:
            sheet.append([reporter_id, 10])
    workbook.save(workbook_path)
    workbook.close()

    bound = _bind_own(protocol)
    publisher = bound["ordinary_public_asset_contract"]["publisher_processed_asset"]
    publisher["byte_count"] = workbook_path.stat().st_size
    publisher["sha256"] = _sha256(workbook_path)
    required = bound["ordinary_public_asset_contract"]["author_repository"]["required_assets"]
    for relative, path in (
        ("data/GWASrewritepos_CMS_alignment_file.fasta", fasta_path),
        ("data/GWASrewritepos_CMS_arrayassign", array_path),
        ("scripts/generate_processed_DESeq2_data.R", r_path),
    ):
        required[relative]["byte_count"] = path.stat().st_size
        required[relative]["sha256"] = _sha256(path)
    PREFLIGHT.validate_protocol(bound)
    return workbook_path, author, bound


def test_candidate_freezes_authority_runtime_and_scientific_disposition() -> None:
    protocol = _protocol()
    baseline = protocol["fresh_baseline"]
    assert baseline["latest_settled_runtime_event_id"] == "A1-EVT-059"
    assert baseline["production_head"] == PREFLIGHT.ENCSR_I1_COMMIT
    authority = protocol["implementation_binding"]["authority_group"]
    assert authority["authority_commit"] == PREFLIGHT.AUTHORITY_COMMIT
    assert tuple(authority["authority_exact_changed_paths"]) == PREFLIGHT.AUTHORITY_EXACT12
    runtime = protocol["implementation_binding"]["runtime_group"]
    assert runtime["i1_commit"] == PREFLIGHT.RUNTIME_I1_COMMIT
    assert runtime["i2_commit"] == PREFLIGHT.RUNTIME_I2_COMMIT
    assert runtime["b2_commit"] == PREFLIGHT.RUNTIME_B2_COMMIT
    assert tuple(runtime["i1_exact_changed_paths"]) == PREFLIGHT.RUNTIME_EXACT3
    predecessor = protocol["implementation_binding"]["gse217518_predecessor_group"]
    encsr_i1 = protocol["implementation_binding"]["encsr854ruf_i1_group"]
    own = protocol["implementation_binding"]["own_preflight_group"]
    assert predecessor["status"] == PREFLIGHT.BOUND
    assert predecessor["i1_commit"] == PREFLIGHT.GSE217_I1_COMMIT
    assert predecessor["i2_commit"] == PREFLIGHT.GSE217_I2_COMMIT
    assert predecessor["b2_commit"] == PREFLIGHT.GSE217_B2_COMMIT
    assert predecessor["i3_commit"] == PREFLIGHT.GSE217_I3_COMMIT
    assert predecessor["b3_commit"] == PREFLIGHT.GSE217_B3_COMMIT
    assert encsr_i1 == {
        "status": PREFLIGHT.BOUND,
        "i1_expected_parent": PREFLIGHT.GSE217_B3_COMMIT,
        "i1_commit": PREFLIGHT.ENCSR_I1_COMMIT,
        "i1_exact_changed_paths": list(PREFLIGHT.EXACT3),
        "i1_blob_sha256_by_path": PREFLIGHT.ENCSR_I1_BLOBS,
    }
    assert own["status"] == PREFLIGHT.UNKNOWN
    assert tuple(own["implementation_exact_changed_paths"]) == PREFLIGHT.EXACT3
    assert protocol["public_research_snapshot"]["normalized_gate_counts"] == {
        "pass": 3,
        "partial_or_conditional": 3,
        "fail": 1,
        "unknown_not_asserted": 4,
        "total": 11,
    }
    assert protocol["scientific_state"]["contribution_delta"] == {
        "ordinary": 0,
        "a1": 0,
        "true_a2": 0,
        "canonical_records": 0,
    }


def test_legal_disk_i_and_disk_b_protocols_are_both_accepted(tmp_path: Path) -> None:
    disk_i = _protocol()
    i_path = tmp_path / "i.json"
    i_path.write_text(json.dumps(disk_i), encoding="utf-8")
    loaded_i = PREFLIGHT.load_protocol(i_path)
    assert loaded_i["implementation_binding"]["own_preflight_group"]["status"] == PREFLIGHT.UNKNOWN

    disk_b = _bind_own(_protocol())
    b_path = tmp_path / "b.json"
    b_path.write_text(json.dumps(disk_b), encoding="utf-8")
    loaded_b = PREFLIGHT.load_protocol(b_path)
    assert loaded_b["implementation_binding"]["own_preflight_group"]["status"] == PREFLIGHT.BOUND


def test_checked_in_disk_i2_or_b2_normalises_to_clean_i2() -> None:
    disk = _disk_protocol()
    disk_status = disk["implementation_binding"]["own_preflight_group"]["status"]
    assert disk_status in {PREFLIGHT.UNKNOWN, PREFLIGHT.BOUND}
    clean = _protocol()
    assert clean["implementation_binding"]["own_preflight_group"]["status"] == PREFLIGHT.UNKNOWN
    assert PREFLIGHT._normalise_own_binding(disk) == clean


def test_inactive_candidate_stops_before_git_asset_or_output_io(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    protocol = _protocol()
    config_path = tmp_path / "candidate.json"
    config_path.write_text(json.dumps(protocol), encoding="utf-8")
    calls = {"git": 0, "assets": 0, "output": 0}

    def bomb_git(*_args: object, **_kwargs: object) -> dict:
        calls["git"] += 1
        raise AssertionError("Git must not be inspected")

    def bomb_assets(*_args: object, **_kwargs: object) -> dict:
        calls["assets"] += 1
        raise AssertionError("assets must not be inspected")

    def bomb_output(*_args: object, **_kwargs: object) -> Path:
        calls["output"] += 1
        raise AssertionError("output must not be inspected")

    monkeypatch.setattr(PREFLIGHT, "_audit_repository", bomb_git)
    monkeypatch.setattr(PREFLIGHT, "_audit_prepared_assets", bomb_assets)
    monkeypatch.setattr(PREFLIGHT, "_write_exactly_one", bomb_output)
    output_dir = tmp_path / "must-not-exist"
    with pytest.raises(PREFLIGHT.ActivationBlocked, match="UNKNOWN_NOT_ASSERTED"):
        PREFLIGHT.execute(
            config_path,
            tmp_path / "missing-repository",
            tmp_path / "missing-publisher.xlsx",
            tmp_path / "missing-author",
            output_dir,
        )
    assert calls == {"git": 0, "assets": 0, "output": 0}
    assert not output_dir.exists()


def test_predecessor_history_drift_or_partial_own_binding_is_rejected() -> None:
    predecessor = _protocol()
    predecessor["implementation_binding"]["gse217518_predecessor_group"][
        "i2_commit"
    ] = "1" * 40
    with pytest.raises(PREFLIGHT.ProtocolError, match="i2_commit differs"):
        PREFLIGHT.validate_protocol(predecessor)

    own = _protocol()
    own["implementation_binding"]["own_preflight_group"]["implementation_commit"] = "3" * 40
    with pytest.raises(PREFLIGHT.ProtocolError, match="partially bound"):
        PREFLIGHT.validate_protocol(own)


def test_repository_auditor_freezes_full_chain_and_rejects_stale_copy(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    protocol = _bind_own(_protocol())
    repo = tmp_path / "repo"
    protocol["repository_authority"]["production_repo_root"] = str(repo)
    config_path = repo / PREFLIGHT.CONFIG_REPO_PATH
    script_path = repo / PREFLIGHT.SCRIPT_REPO_PATH
    test_path = repo / PREFLIGHT.TEST_REPO_PATH
    for path in (config_path, script_path, test_path):
        path.parent.mkdir(parents=True, exist_ok=True)
    config_path.write_text(json.dumps(protocol), encoding="utf-8")
    script_path.write_bytes(b"stale copied producer")
    test_path.write_bytes(b"bound focused test")
    head = "4" * 40

    def fake_run_git(_repo: Path, *arguments: str) -> str:
        if arguments in (("rev-parse", "HEAD"), ("rev-parse", "@{upstream}")):
            return head
        if arguments == ("rev-parse", "--abbrev-ref", "HEAD"):
            return PREFLIGHT.PRODUCTION_BRANCH
        if arguments == ("rev-parse", "--abbrev-ref", "@{upstream}"):
            return PREFLIGHT.PRODUCTION_UPSTREAM
        if arguments == ("status", "--porcelain=v1", "--untracked-files=all"):
            return ""
        raise AssertionError(arguments)

    verified: list[dict] = []

    def fake_verify(_repo: Path, **kwargs: object) -> None:
        verified.append(dict(kwargs))

    gse_i_protocol = {
        "bindings": {
            "implementation": {
                "status": PREFLIGHT.UNKNOWN,
                "implementation_commit": PREFLIGHT.UNKNOWN,
                "implementation_script_sha256": PREFLIGHT.UNKNOWN,
                "implementation_test_sha256": PREFLIGHT.UNKNOWN,
                "fixed": "preserved",
            }
        }
    }
    gse_b_protocol = copy.deepcopy(gse_i_protocol)
    gse_b_protocol["bindings"]["implementation"].update(
        {
            "status": PREFLIGHT.BOUND,
            "implementation_commit": "1" * 40,
            "implementation_script_sha256": "a" * 64,
            "implementation_test_sha256": "a" * 64,
        }
    )
    own_i_protocol = PREFLIGHT._normalise_own_binding(protocol)

    def fake_blob(_repo: Path, commit: str, path: str) -> bytes:
        own = protocol["implementation_binding"]["own_preflight_group"]
        if commit in (PREFLIGHT.GSE217_I2_COMMIT, PREFLIGHT.GSE217_I3_COMMIT) and path == PREFLIGHT.GSE217_CONFIG_PATH:
            return json.dumps(gse_i_protocol).encode("utf-8")
        if commit in (PREFLIGHT.GSE217_B2_COMMIT, PREFLIGHT.GSE217_B3_COMMIT) and path == PREFLIGHT.GSE217_CONFIG_PATH:
            return json.dumps(gse_b_protocol).encode("utf-8")
        if commit == own["implementation_commit"] and path == PREFLIGHT.CONFIG_REPO_PATH:
            return json.dumps(own_i_protocol).encode("utf-8")
        if commit == head and path == PREFLIGHT.CONFIG_REPO_PATH:
            return config_path.read_bytes()
        if commit == own["implementation_commit"] and path == PREFLIGHT.SCRIPT_REPO_PATH:
            return b"bound producer"
        if commit == own["implementation_commit"] and path == PREFLIGHT.TEST_REPO_PATH:
            return b"bound focused test"
        raise AssertionError((commit, path))

    monkeypatch.setattr(PREFLIGHT, "_run_git", fake_run_git)
    monkeypatch.setattr(PREFLIGHT, "_live_origin_head", lambda *_: head)
    monkeypatch.setattr(PREFLIGHT, "_verify_commit", fake_verify)
    monkeypatch.setattr(PREFLIGHT, "_git_blob", fake_blob)
    monkeypatch.setattr(PREFLIGHT, "__file__", str(script_path))
    with pytest.raises(PREFLIGHT.RepositoryError, match="differs from ENCSR854RUF I2"):
        PREFLIGHT._audit_repository(protocol, config_path, repo)
    assert [item["label"] for item in verified] == [
        "DEC027 authority A",
        "DEC027 runtime I1",
        "DEC027 runtime I2",
        "DEC027 runtime B2",
        "GSE217518 I1",
        "GSE217518 I2",
        "GSE217518 B2",
        "GSE217518 I3",
        "GSE217518 B3",
        "ENCSR854RUF I1",
        "ENCSR854RUF I2",
        "ENCSR854RUF B2",
    ]
    assert verified[4]["expected_parent"] == PREFLIGHT.RUNTIME_B2_COMMIT
    assert verified[9]["expected_parent"] == PREFLIGHT.GSE217_B3_COMMIT
    assert verified[10]["expected_parent"] == PREFLIGHT.ENCSR_I1_COMMIT
    assert verified[11]["expected_parent"] == "3" * 40


def test_repository_failure_precedes_asset_and_output_io(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    protocol = _bind_own(_protocol())
    config_path = tmp_path / "bound.json"
    config_path.write_text(json.dumps(protocol), encoding="utf-8")
    calls = {"repository": 0, "assets": 0, "output": 0}

    def fail_repository(*_args: object, **_kwargs: object) -> dict:
        calls["repository"] += 1
        raise PREFLIGHT.RepositoryError("stale producer")

    def bomb_assets(*_args: object, **_kwargs: object) -> dict:
        calls["assets"] += 1
        raise AssertionError("asset inspection must follow repository closure")

    def bomb_output(*_args: object, **_kwargs: object) -> Path:
        calls["output"] += 1
        raise AssertionError("output inspection must follow repository closure")

    monkeypatch.setattr(PREFLIGHT, "_audit_repository", fail_repository)
    monkeypatch.setattr(PREFLIGHT, "_audit_prepared_assets", bomb_assets)
    monkeypatch.setattr(PREFLIGHT, "_write_exactly_one", bomb_output)
    with pytest.raises(PREFLIGHT.RepositoryError, match="stale producer"):
        PREFLIGHT.execute(
            config_path,
            tmp_path / "repo",
            tmp_path / "publisher.xlsx",
            tmp_path / "author",
            tmp_path / "output",
        )
    assert calls == {"repository": 1, "assets": 0, "output": 0}


def test_synthetic_assets_emit_only_aggregate_geometry(tmp_path: Path) -> None:
    workbook, author, bound = _make_assets(tmp_path, _protocol())
    geometry = PREFLIGHT._audit_prepared_assets(bound, workbook, author)
    record = PREFLIGHT.build_aggregate_record(
        bound,
        geometry,
        {
            "status": "SYNTHETIC_TEST_BINDING_ONLY",
            "head": "4" * 40,
            "implementation_commit": "3" * 40,
            "binding_commit": "4" * 40,
        },
    )
    assert geometry["published_pair_count"] == 2
    assert geometry["published_reporter_count"] == 4
    assert geometry["author_fasta_header_count"] == 4
    assert geometry["author_fasta_expanded_alias_token_count"] == 5
    assert geometry["source_candidate_crosswalk_missing_count"] == 0
    assert geometry["declared_allele_length_to_sequence_replay_mismatch_count"] == 0
    assert geometry["index_error_affected_pair_count"] == 1
    assert geometry[
        "finite_effect_and_lfcse_all_six_context_pair_count_after_index_error_exclusion"
    ] == 1
    assert record["status"] == "TERMINAL_AGGREGATE_PREFLIGHT_STOP_NOT_QUALIFIED"
    assert record["qualification_or_credit_change"] is False
    assert record["contribution_delta"] == {
        "ordinary": 0,
        "a1": 0,
        "true_a2": 0,
        "canonical_records": 0,
    }
    serialized = json.dumps(record, sort_keys=True)
    for poison in (
        "SECRET_REPORTER",
        "SECRET_PAIR",
        "SECRET_UNUSED_ALIAS",
        _sequence("A"),
        _sequence("C"),
    ):
        assert poison not in serialized


@pytest.mark.parametrize(
    ("omit_alt_member", "omit_alt_fasta", "message"),
    (
        (True, False, "exactly one ref and one alt"),
        (False, True, "crosswalk is incomplete"),
    ),
)
def test_malformed_pair_or_unmapped_reporter_fails_closed(
    tmp_path: Path,
    omit_alt_member: bool,
    omit_alt_fasta: bool,
    message: str,
) -> None:
    workbook, author, bound = _make_assets(
        tmp_path,
        _protocol(),
        omit_alt_member=omit_alt_member,
        omit_alt_fasta=omit_alt_fasta,
    )
    with pytest.raises(PREFLIGHT.AssetAuditError, match=message):
        PREFLIGHT._audit_prepared_assets(bound, workbook, author)


def test_observed_geometry_mismatch_downgrades_pass() -> None:
    protocol = _protocol()
    geometry = copy.deepcopy(protocol["public_research_snapshot"]["aggregate_geometry"])
    geometry.update(
        {
            "result_missing_pair_count": 0,
            "unexpected_result_pair_count": 0,
            "endpoint_context_count": 6,
            "declared_allele_length_to_sequence_replay_mismatch_count": 1,
        }
    )
    gates = PREFLIGHT.evaluate_gate_statuses(protocol, geometry)
    by_id = {gate["gate_id"]: gate for gate in gates}
    construct = by_id["FULL_REPORTER_AND_THREE_UTR_CONTEXT_CLOSED"]
    assert construct["normalized_status"] == "FAIL"
    assert construct["reason_code"] == "OBSERVED_REPORTER_INSERT_OR_EDIT_REPLAY_DIFFERS"


def _frozen_record() -> dict:
    protocol = _protocol()
    geometry = copy.deepcopy(protocol["public_research_snapshot"]["aggregate_geometry"])
    geometry.update(
        {
            "result_missing_pair_count": 0,
            "unexpected_result_pair_count": 0,
            "endpoint_context_count": 6,
        }
    )
    return PREFLIGHT.build_aggregate_record(
        protocol,
        geometry,
        {
            "status": "SYNTHETIC_TEST_BINDING_ONLY",
            "head": "4" * 40,
            "implementation_commit": "3" * 40,
            "binding_commit": "4" * 40,
        },
    )


def test_atomic_fixed_report_is_idempotent_and_never_replaces(tmp_path: Path) -> None:
    report = _frozen_record()
    output = tmp_path / "aggregate"
    first = PREFLIGHT._write_exactly_one(report, output)
    original = first.read_bytes()
    assert PREFLIGHT._write_exactly_one(report, output) == first
    different = copy.deepcopy(report)
    different["status"] = "DIFFERENT_REPORT_FOR_NO_REPLACE_TEST"
    with pytest.raises(PREFLIGHT.PublicationError, match="replacement refused"):
        PREFLIGHT._write_exactly_one(different, output)
    assert first.read_bytes() == original
    assert [path.name for path in output.iterdir()] == [PREFLIGHT.REPORT_FILENAME]


def test_atomic_failure_removes_temp_final_and_new_directory(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    output = tmp_path / "aggregate"
    real_fsync_directory = PREFLIGHT._fsync_directory
    calls = {"count": 0}

    def fail_after_link(path: Path) -> None:
        calls["count"] += 1
        if calls["count"] == 2:
            raise OSError("injected directory fsync failure")
        real_fsync_directory(path)

    monkeypatch.setattr(PREFLIGHT, "_fsync_directory", fail_after_link)
    with pytest.raises(PREFLIGHT.PublicationError, match="atomically publish"):
        PREFLIGHT._write_exactly_one(_frozen_record(), output)
    assert not output.exists()


def test_cli_has_only_fixed_production_path_and_no_public_analysis_bypass() -> None:
    parser_source = inspect.getsource(PREFLIGHT._parser)
    main_source = inspect.getsource(PREFLIGHT.main)
    execute_source = inspect.getsource(PREFLIGHT.execute)
    assert "static-review" not in parser_source
    assert "public_review_summary" not in main_source
    assert "_audit_prepared_assets" in execute_source
    assert "build_aggregate_record" in execute_source
    assert "_write_exactly_one" in execute_source
