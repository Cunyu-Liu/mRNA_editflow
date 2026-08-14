from __future__ import annotations

import copy
import csv
import gzip
import hashlib
import importlib.util
import json
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[2]
PROTOCOL_PATH = (
    ROOT
    / "configs"
    / "route_a_v3_gse207584_moesm7_aggregate_endpoint_universe_preflight_v1.json"
)
MODULE_PATH = (
    ROOT
    / "scripts"
    / "route_a_v3"
    / "preflight_gse207584_moesm7_aggregate_endpoint_universe.py"
)
SPEC = importlib.util.spec_from_file_location("gse207584_moesm7_preflight", MODULE_PATH)
assert SPEC and SPEC.loader
PREFLIGHT = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = PREFLIGHT
SPEC.loader.exec_module(PREFLIGHT)


def _protocol() -> dict[str, object]:
    value = json.loads(PROTOCOL_PATH.read_text(encoding="utf-8"))
    PREFLIGHT._validate_protocol(value)
    return value


def _i_protocol() -> dict[str, object]:
    value = json.loads(PROTOCOL_PATH.read_text(encoding="utf-8"))
    binding = value["implementation_binding"]
    binding["production_repo_root"] = str(PREFLIGHT.PRODUCTION_REPO_ROOT)
    binding["initial_implementation_commit"] = PREFLIGHT.INITIAL_IMPLEMENTATION_COMMIT
    binding["initial_implementation_frozen_blobs"] = [
        {
            "path": path,
            "sha256": PREFLIGHT.INITIAL_IMPLEMENTATION_FROZEN_BLOBS[path],
        }
        for path in PREFLIGHT.EXPECTED_EXACT3
    ]
    for field in PREFLIGHT.UNKNOWN_BINDING_SCALARS:
        binding[field] = PREFLIGHT.UNKNOWN
    PREFLIGHT._validate_protocol(value)
    return value


def _bound_protocol() -> dict[str, object]:
    value = _i_protocol()
    binding = value["implementation_binding"]
    binding["status"] = PREFLIGHT.BOUND
    binding["implementation_commit"] = "2" * 40
    binding["implementation_script_sha256"] = "3" * 64
    binding["implementation_test_sha256"] = "4" * 64
    PREFLIGHT._validate_protocol(value)
    return value


def _fixture_binding(*args: object) -> dict[str, str]:
    return {
        "status": "TEST_FIXTURE_BOUND_WITHOUT_GIT",
        "current_predecessor_commit": PREFLIGHT.CURRENT_PREDECESSOR_COMMIT,
        "implementation_commit": "2" * 40,
        "binding_commit": "8" * 40,
    }


def _write_protocol(path: Path, protocol: object) -> tuple[Path, bytes]:
    payload = (json.dumps(protocol, indent=2) + "\n").encode("utf-8")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(payload)
    return path, payload


def _family_sizes(protocol: dict[str, object]) -> list[int]:
    histogram = protocol["prefrozen_geometry"]["family_size_histogram"]
    return [
        size
        for size, family_count in sorted(
            (int(size), count) for size, count in histogram.items()
        )
        for _ in range(family_count)
    ]


def _synthetic_assets(
    tmp_path: Path,
    *,
    defect: str | None = None,
) -> tuple[Path, Path]:
    if defect not in {None, "MISSING_PAIR", "SIGNATURE_MISMATCH"}:
        raise AssertionError("unsupported synthetic defect")
    protocol = _protocol()
    workbook_path = tmp_path / "41598_2022_15526_MOESM7_ESM.xlsx"
    geo_path = tmp_path / "GSE207584_Zebrafish-library-perfect.csv.gz"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Supplemental-Table-6-library-pe"
    sheet.append(list(PREFLIGHT.PUBLISHER_HEADER))
    family_assets: list[
        tuple[str, list[str], list[tuple[str, tuple[float, ...]]]]
    ] = []
    for family_index, size in enumerate(_family_sizes(protocol)):
        family = f"FAMILY_MEMBER_POISON_{family_index:03d}"
        candidates = [
            f"CANDIDATE_MEMBER_POISON_{family_index:03d}_{index:02d}"
            for index in range(size)
        ]
        endpoints: list[tuple[str, tuple[float, ...]]] = []
        for endpoint_index in range(size):
            group = f"GROUP_MEMBER_POISON_{endpoint_index % 7}"
            base = family_index * 1000.0 + endpoint_index * 20.0
            values = tuple(base + offset + 0.125 for offset in range(9))
            endpoints.append((group, values))
            sheet.append(
                [
                    family,
                    group,
                    *values,
                    -0.01 * (endpoint_index + 1),
                    0.01 * (endpoint_index + 1),
                ]
            )
        family_assets.append((family, candidates, endpoints))
    workbook.save(workbook_path)

    omitted = False
    mutated = False
    with gzip.open(geo_path, "wt", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerow(PREFLIGHT.GEO_HEADER)
        for family, candidates, endpoints in reversed(family_assets):
            for candidate in reversed(candidates):
                for group, source_values in reversed(endpoints):
                    if defect == "MISSING_PAIR" and not omitted:
                        omitted = True
                        continue
                    values = list(source_values)
                    if defect == "SIGNATURE_MISMATCH" and not mutated:
                        values[0] += 1.0
                        mutated = True
                    elif not mutated:
                        # Publisher XLSX and GEO CSV may differ below the frozen
                        # nine-decimal signature resolution.  This must not make
                        # row order relevant or fabricate a different endpoint.
                        values[0] += 4e-11
                        mutated = True
                    writer.writerow([candidate, family, group, *values])
    return workbook_path, geo_path


def _execute(tmp_path: Path, *, defect: str | None = None) -> dict[str, object]:
    publisher, geo = _synthetic_assets(tmp_path, defect=defect)
    output = tmp_path / PREFLIGHT.REPORT_FILENAME
    return PREFLIGHT.execute(
        PROTOCOL_PATH,
        publisher,
        geo,
        output,
        binding_auditor=_fixture_binding,
    )


def test_disk_protocol_freezes_dec023_scope_geometry_rights_and_zero_claims() -> None:
    protocol = _protocol()
    binding = protocol["implementation_binding"]
    authority = protocol["authority_assessment"]
    geometry = protocol["prefrozen_geometry"]
    rights = protocol["rights_policy"]
    claims = protocol["claim_boundary"]

    assert protocol["protocol_status"] == PREFLIGHT.PROTOCOL_STATUS
    assert binding["current_predecessor_commit"] == (
        "0a6586814460b211cc730c463390e68f64aaa4f1"
    )
    assert binding["initial_implementation_commit"] == (
        "374ea6166c74c898751c7a3d4d6951664ca1d524"
    )
    assert binding["initial_implementation_frozen_blobs"] == [
        {
            "path": path,
            "sha256": PREFLIGHT.INITIAL_IMPLEMENTATION_FROZEN_BLOBS[path],
        }
        for path in PREFLIGHT.EXPECTED_EXACT3
    ]
    assert tuple(binding["implementation_commit_exact_changed_paths"]) == (
        PREFLIGHT.EXPECTED_EXACT3
    )
    assert binding["binding_commit_exact_changed_paths"] == [
        PREFLIGHT.CONFIG_PATH
    ]
    values = [binding[field] for field in PREFLIGHT.UNKNOWN_BINDING_SCALARS]
    if binding["status"] == PREFLIGHT.UNKNOWN:
        assert values == [PREFLIGHT.UNKNOWN] * 4
    else:
        assert binding["status"] == PREFLIGHT.BOUND
        assert PREFLIGHT.HEX40_RE.fullmatch(binding["implementation_commit"])
        assert hashlib.sha256(MODULE_PATH.read_bytes()).hexdigest() == (
            binding["implementation_script_sha256"]
        )
        assert hashlib.sha256(Path(__file__).read_bytes()).hexdigest() == (
            binding["implementation_test_sha256"]
        )
    assert authority[
        "current_authority_sufficient_for_this_candidate_and_aggregate_preflight"
    ] is True
    assert authority["new_row_level_authority_granted"] is False
    assert authority["sequence_field_classes_read_count"] == 0
    assert geometry["publisher_endpoint_row_count"] == 955
    assert geometry["geo_body_row_count"] == 10227
    assert geometry["row_order_pairing_permitted"] is False
    assert geometry["sequence_pairing_permitted"] is False
    assert rights["private_academic_analysis_allowed"] is True
    assert rights["aggregate_derived_reporting_allowed"] is True
    assert rights["member_payload_redistribution_allowed"] is False
    assert all(claims[key] is False for key in PREFLIGHT.FALSE_CLAIM_KEYS)
    assert claims["current_credit_delta"] == {
        "ordinary": 0,
        "A1": 0,
        "true_A2": 0,
    }


def test_i_to_b_normalization_changes_exactly_four_scalars() -> None:
    i_protocol = _i_protocol()
    b_protocol = _bound_protocol()

    assert [
        i_protocol["implementation_binding"][field]
        for field in PREFLIGHT.UNKNOWN_BINDING_SCALARS
    ] == [PREFLIGHT.UNKNOWN] * 4
    assert PREFLIGHT._semantic_diff_paths(i_protocol, b_protocol) == set(
        PREFLIGHT.UNKNOWN_TO_BOUND_PATHS
    )
    assert PREFLIGHT._normalise_binding_to_i(b_protocol) == i_protocol


def test_unknown_or_partial_binding_stops_before_asset_and_output_io(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    calls = {"asset": 0, "output": 0}

    def forbidden_aggregate(*args: object) -> dict[str, object]:
        calls["asset"] += 1
        raise AssertionError("asset readers must not run")

    def forbidden_publish(*args: object) -> str:
        calls["output"] += 1
        raise AssertionError("output publisher must not run")

    monkeypatch.setattr(PREFLIGHT, "aggregate", forbidden_aggregate)
    monkeypatch.setattr(PREFLIGHT, "_publish_no_replace", forbidden_publish)
    output = tmp_path / PREFLIGHT.REPORT_FILENAME
    unknown_protocol = _i_protocol()
    unknown_path, _ = _write_protocol(
        tmp_path / "unknown" / PREFLIGHT.PROTOCOL_BASENAME,
        unknown_protocol,
    )
    with pytest.raises(PREFLIGHT.BindingNotFrozen, match="not BOUND"):
        PREFLIGHT.execute(
            unknown_path,
            tmp_path / "missing-publisher.xlsx",
            tmp_path / "missing-geo.csv.gz",
            output,
        )
    assert calls == {"asset": 0, "output": 0}
    assert not output.exists()

    partial = copy.deepcopy(unknown_protocol)
    partial["implementation_binding"]["implementation_commit"] = "2" * 40
    partial_path, _ = _write_protocol(
        tmp_path / "partial" / PREFLIGHT.PROTOCOL_BASENAME,
        partial,
    )
    with pytest.raises(PREFLIGHT.ProtocolError, match="partially known"):
        PREFLIGHT.execute(
            partial_path,
            tmp_path / "missing-publisher.xlsx",
            tmp_path / "missing-geo.csv.gz",
            output,
        )
    assert calls == {"asset": 0, "output": 0}
    assert not output.exists()


def _fake_bound_repository(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> tuple[dict[str, object], Path, bytes, Path]:
    repo_root = (tmp_path / "repo").resolve()
    repo_root.mkdir()
    monkeypatch.setattr(PREFLIGHT, "PRODUCTION_REPO_ROOT", repo_root)

    initial_protocol = json.loads(PROTOCOL_PATH.read_text(encoding="utf-8"))
    initial_protocol["implementation_binding"] = {
        "historical_lifecycle": "FROZEN_INITIAL_I1_TEST_FIXTURE"
    }
    initial_config_blob = (json.dumps(initial_protocol, indent=2) + "\n").encode(
        "utf-8"
    )
    initial_script_blob = b"GSE207584 endpoint-universe initial implementation I1\n"
    initial_test_blob = b"GSE207584 endpoint-universe initial focused test I1\n"
    frozen_initial_blobs = {
        PREFLIGHT.CONFIG_PATH: hashlib.sha256(initial_config_blob).hexdigest(),
        PREFLIGHT.SCRIPT_PATH: hashlib.sha256(initial_script_blob).hexdigest(),
        PREFLIGHT.TEST_PATH: hashlib.sha256(initial_test_blob).hexdigest(),
    }
    monkeypatch.setattr(
        PREFLIGHT,
        "INITIAL_IMPLEMENTATION_FROZEN_BLOBS",
        frozen_initial_blobs,
    )

    script_blob = b"GSE207584 endpoint-universe repair implementation I2\n"
    test_blob = b"GSE207584 endpoint-universe repair focused test I2\n"
    protocol = _bound_protocol()
    binding = protocol["implementation_binding"]
    binding["implementation_script_sha256"] = hashlib.sha256(script_blob).hexdigest()
    binding["implementation_test_sha256"] = hashlib.sha256(test_blob).hexdigest()
    PREFLIGHT._validate_protocol(protocol)
    i_protocol = PREFLIGHT._normalise_binding_to_i(protocol)

    protocol_path, protocol_payload = _write_protocol(
        repo_root / PREFLIGHT.CONFIG_PATH,
        protocol,
    )
    script_path = repo_root / PREFLIGHT.SCRIPT_PATH
    script_path.parent.mkdir(parents=True, exist_ok=True)
    script_path.write_bytes(script_blob)
    test_path = repo_root / PREFLIGHT.TEST_PATH
    test_path.parent.mkdir(parents=True, exist_ok=True)
    test_path.write_bytes(test_blob)
    monkeypatch.setattr(PREFLIGHT, "EXECUTING_SCRIPT_PATH", script_path.resolve())

    predecessor = PREFLIGHT.CURRENT_PREDECESSOR_COMMIT
    initial_implementation = PREFLIGHT.INITIAL_IMPLEMENTATION_COMMIT
    implementation = str(binding["implementation_commit"])
    head = "8" * 40
    branch = PREFLIGHT.PRODUCTION_BRANCH
    remote = PREFLIGHT.REMOTE_NAME
    git_text = {
        ("rev-parse", "--abbrev-ref", "HEAD"): branch,
        ("status", "--porcelain=v1", "--untracked-files=all"): "",
        ("rev-parse", "HEAD"): head,
        (
            "rev-parse",
            "--abbrev-ref",
            "--symbolic-full-name",
            "@{u}",
        ): f"{remote}/{branch}",
        ("rev-parse", "@{u}"): head,
        ("rev-parse", f"refs/remotes/{remote}/{branch}"): head,
        (
            "ls-remote",
            "--heads",
            remote,
            f"refs/heads/{branch}",
        ): f"{head}\trefs/heads/{branch}",
        (
            "rev-list",
            "--parents",
            "-n",
            "1",
            initial_implementation,
        ): f"{initial_implementation} {predecessor}",
        (
            "rev-list",
            "--parents",
            "-n",
            "1",
            implementation,
        ): f"{implementation} {initial_implementation}",
        ("rev-list", "--parents", "-n", "1", head): f"{head} {implementation}",
        (
            "diff-tree",
            "--no-commit-id",
            "--name-only",
            "-r",
            initial_implementation,
        ): "\n".join(PREFLIGHT.EXPECTED_EXACT3),
        (
            "diff-tree",
            "--no-commit-id",
            "--name-only",
            "-r",
            implementation,
        ): "\n".join(PREFLIGHT.EXPECTED_EXACT3),
        (
            "diff-tree",
            "--no-commit-id",
            "--name-only",
            "-r",
            head,
        ): PREFLIGHT.CONFIG_PATH,
    }
    i_payload = (json.dumps(i_protocol, indent=2) + "\n").encode("utf-8")
    git_blobs = {
        (initial_implementation, PREFLIGHT.CONFIG_PATH): initial_config_blob,
        (initial_implementation, PREFLIGHT.SCRIPT_PATH): initial_script_blob,
        (initial_implementation, PREFLIGHT.TEST_PATH): initial_test_blob,
        (implementation, PREFLIGHT.CONFIG_PATH): i_payload,
        (implementation, PREFLIGHT.SCRIPT_PATH): script_blob,
        (implementation, PREFLIGHT.TEST_PATH): test_blob,
        (head, PREFLIGHT.CONFIG_PATH): protocol_payload,
        (head, PREFLIGHT.SCRIPT_PATH): script_blob,
        (head, PREFLIGHT.TEST_PATH): test_blob,
    }

    def fake_git_text(root: Path, *arguments: str) -> str:
        assert root == repo_root
        return git_text[arguments]

    def fake_git_blob(root: Path, commit: str, relative_path: str) -> bytes:
        assert root == repo_root
        return git_blobs[(commit, relative_path)]

    monkeypatch.setattr(PREFLIGHT, "_run_git_text", fake_git_text)
    monkeypatch.setattr(PREFLIGHT, "_git_blob", fake_git_blob)
    return protocol, protocol_path, protocol_payload, repo_root


def test_default_binding_auditor_closes_current_refs_exact3_i_and_config_only_b(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    protocol, protocol_path, protocol_payload, repo_root = _fake_bound_repository(
        tmp_path,
        monkeypatch,
    )

    result = PREFLIGHT._default_binding_auditor(
        protocol,
        protocol_path,
        protocol_payload,
        repo_root,
    )

    assert result == {
        "status": "BOUND_FROZEN_I1_REPAIR_EXACT3_I2_CONFIG_ONLY_B2_VERIFIED",
        "current_predecessor_commit": PREFLIGHT.CURRENT_PREDECESSOR_COMMIT,
        "initial_implementation_commit": PREFLIGHT.INITIAL_IMPLEMENTATION_COMMIT,
        "implementation_commit": "2" * 40,
        "binding_commit": "8" * 40,
        "upstream_head": "8" * 40,
        "tracking_head": "8" * 40,
        "live_origin_head": "8" * 40,
    }


def test_stale_executing_copy_is_rejected_before_asset_or_output_io(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    protocol, protocol_path, protocol_payload, repo_root = _fake_bound_repository(
        tmp_path,
        monkeypatch,
    )
    stale_copy = tmp_path / "copied_preflight.py"
    stale_copy.write_bytes((repo_root / PREFLIGHT.SCRIPT_PATH).read_bytes())
    monkeypatch.setattr(PREFLIGHT, "EXECUTING_SCRIPT_PATH", stale_copy.resolve())

    with pytest.raises(PREFLIGHT.ProtocolError, match="stale or copied"):
        PREFLIGHT._default_binding_auditor(
            protocol,
            protocol_path,
            protocol_payload,
            repo_root,
        )


def test_complete_cartesian_geometry_closes_only_endpoint_universe(tmp_path: Path) -> None:
    report = _execute(tmp_path)
    comparison = report["aggregate_geometry"]["publisher_geo_comparison"]

    assert report["status"] == PREFLIGHT.STATUS_CLOSED
    assert all(
        report["gates"][gate_id]["status"] == PREFLIGHT.PASS
        for gate_id in PREFLIGHT.GATE_IDS[:5]
    )
    assert report["gates"][PREFLIGHT.GATE_IDS[5]]["status"] == PREFLIGHT.UNKNOWN
    assert comparison == {
        "family_union_count": 100,
        "family_endpoint_candidate_count_match_count": 100,
        "family_endpoint_signature_set_match_count": 100,
        "family_square_row_count_match_count": 100,
        "family_complete_cartesian_count": 100,
        "publisher_endpoint_signatures_not_in_geo_count": 0,
        "geo_endpoint_signatures_not_in_publisher_count": 0,
        "expected_geo_rows_from_publisher_family_squares": 10227,
        "numeric_signature_decimal_places": 9,
        "row_order_pairing_used": False,
        "sequence_pairing_used": False,
        "source_to_candidate_mapping_identifiable": False,
    }
    assert report["claim_boundary"]["source_to_candidate_mapping_established"] is False
    assert report["claim_boundary"]["true_a2_status_allowed_or_changed"] is False

    rendered = (tmp_path / PREFLIGHT.REPORT_FILENAME).read_text(encoding="utf-8")
    assert "MEMBER_POISON" not in rendered
    assert report["internal_access_attestation"] == {
        "ordinary_public_assets_read_count": 2,
        "sequence_asset_read_count": 0,
        "derived_effect_value_cell_read_count": 0,
        "private_or_restricted_asset_read_count": 0,
        "sealed_asset_contact_count": 0,
        "member_identifier_sequence_or_row_payload_output_count": 0,
        "row_or_member_material_persisted_count": 0,
        "split_assignment_output_count": 0,
        "training_run_count": 0,
        "gpu_run_count": 0,
        "model_selection_count": 0,
    }


def test_missing_geo_pair_fails_coverage_and_cartesian_gates(tmp_path: Path) -> None:
    report = _execute(tmp_path, defect="MISSING_PAIR")

    assert report["status"] == PREFLIGHT.STATUS_STOP
    assert report["gates"][PREFLIGHT.GATE_IDS[0]]["status"] == PREFLIGHT.PASS
    assert report["gates"][PREFLIGHT.GATE_IDS[1]]["status"] == PREFLIGHT.FAIL
    assert report["gates"][PREFLIGHT.GATE_IDS[2]]["status"] == PREFLIGHT.FAIL
    assert report["gates"][PREFLIGHT.GATE_IDS[5]]["status"] == PREFLIGHT.UNKNOWN


def test_value_signature_mismatch_does_not_pass_as_geometry_pairing(tmp_path: Path) -> None:
    report = _execute(tmp_path, defect="SIGNATURE_MISMATCH")

    assert report["status"] == PREFLIGHT.STATUS_STOP
    assert report["gates"][PREFLIGHT.GATE_IDS[0]]["status"] == PREFLIGHT.PASS
    assert report["gates"][PREFLIGHT.GATE_IDS[1]]["status"] == PREFLIGHT.PASS
    assert report["gates"][PREFLIGHT.GATE_IDS[2]]["status"] == PREFLIGHT.FAIL
    comparison = report["aggregate_geometry"]["publisher_geo_comparison"]
    # The untouched candidates still cover the original publisher endpoint;
    # the one altered candidate contributes one extra, non-publisher signature.
    assert comparison["publisher_endpoint_signatures_not_in_geo_count"] == 0
    assert comparison["geo_endpoint_signatures_not_in_publisher_count"] == 1


def test_publisher_header_drift_stops_before_output(tmp_path: Path) -> None:
    publisher, geo = _synthetic_assets(tmp_path)
    workbook = load_workbook(publisher)
    workbook.active["A1"] = "candidate_id"
    workbook.save(publisher)
    output = tmp_path / PREFLIGHT.REPORT_FILENAME

    with pytest.raises(PREFLIGHT.AssetError, match="publisher endpoint header differs"):
        PREFLIGHT.execute(
            PROTOCOL_PATH,
            publisher,
            geo,
            output,
            binding_auditor=_fixture_binding,
        )
    assert not output.exists()


def test_existing_different_aggregate_report_is_never_overwritten(tmp_path: Path) -> None:
    publisher, geo = _synthetic_assets(tmp_path)
    output = tmp_path / PREFLIGHT.REPORT_FILENAME
    output.write_text("owner material\n", encoding="utf-8")

    with pytest.raises(PREFLIGHT.OutputError, match="different payload"):
        PREFLIGHT.execute(
            PROTOCOL_PATH,
            publisher,
            geo,
            output,
            binding_auditor=_fixture_binding,
        )
    assert output.read_text(encoding="utf-8") == "owner material\n"


def test_existing_exact_aggregate_report_is_accepted_idempotently(
    tmp_path: Path,
) -> None:
    output = tmp_path / PREFLIGHT.REPORT_FILENAME
    payload = PREFLIGHT._json_bytes({"aggregate": True})
    output.write_bytes(payload)

    status = PREFLIGHT._publish_no_replace(output, payload)

    assert status == "EXISTING_EXACT_PAYLOAD_ACCEPTED"
    assert output.read_bytes() == payload
    assert list(tmp_path.glob(f".{PREFLIGHT.REPORT_FILENAME}.*.tmp")) == []


def test_partial_publication_failure_leaves_no_final_or_temp(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output = tmp_path / "output" / PREFLIGHT.REPORT_FILENAME

    def fail_directory_fsync(path: Path) -> None:
        raise OSError("injected directory fsync failure")

    monkeypatch.setattr(PREFLIGHT, "_fsync_directory", fail_directory_fsync)
    with pytest.raises(PREFLIGHT.OutputError, match="cannot publish"):
        PREFLIGHT._publish_no_replace(
            output,
            PREFLIGHT._json_bytes({"aggregate": True}),
        )

    assert not output.exists()
    assert list(output.parent.iterdir()) == []
