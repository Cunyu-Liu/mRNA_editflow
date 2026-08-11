from __future__ import annotations

import copy
import csv
import hashlib
import importlib.util
import io
import inspect
import json
import os
import re
import sys
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook


STAGING_ROOT = Path(__file__).resolve().parents[2]
SCRIPT = (
    STAGING_ROOT
    / "scripts/route_a_v3/produce_gse200304_dec019_canonical_row_lineage_gate.py"
)
CONFIG = (
    STAGING_ROOT
    / "configs/route_a_v3_gse200304_dec019_canonical_row_lineage_gate_v1.json"
)
V3_ROOT = Path(os.environ.get("G200_V3_CONSUMER_ROOT", STAGING_ROOT)).resolve()
V3_IMPLEMENTATION_COMMIT = "86d16c181fc9deaf83597da9c1523e4fea9c7493"
V3_BOUND_CONFIG_SHA256 = "8c88eb6c708fa309ff0c87a0f64fce1bb205a0212b35a85ad3fb3505e8d7613b"


def _load_module(path: Path, name: str):
    spec = importlib.util.spec_from_file_location(name, path)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module


producer = _load_module(SCRIPT, "g200_lineage_gate_producer_under_test")


def v3_bound_config_payload() -> bytes:
    path = (
        V3_ROOT
        / "configs/route_a_v3_gse200304_dec019_reported_endpoint_a1_activation_v3.json"
    )
    payload = path.read_bytes()
    config = json.loads(payload)
    binding = config["implementation_binding"]
    assert binding["status"] == "BOUND"
    assert binding["implementation_commit"] == V3_IMPLEMENTATION_COMMIT
    assert binding["implementation_script_sha256"] == (
        "90e840b721e5d07d4437d429d5b42f5a91fc262e560b3b331095db65dbb18fa6"
    )
    assert binding["implementation_test_sha256"] == (
        "ca0d5221748aaecc10b31edb691f8244a0fe2b94cf67ae9a8f493ac8d3f75ca5"
    )
    assert hashlib.sha256(payload).hexdigest() == V3_BOUND_CONFIG_SHA256
    return payload


def load_config() -> dict:
    return json.loads(CONFIG.read_text(encoding="utf-8"))


def unknown_config() -> dict:
    config = load_config()
    binding = config["implementation_binding"]
    binding.update(
        {
            "status": "UNKNOWN_NOT_ASSERTED",
            "implementation_commit": "UNKNOWN_NOT_ASSERTED",
            "implementation_script_sha256": "UNKNOWN_NOT_ASSERTED",
            "implementation_test_sha256": "UNKNOWN_NOT_ASSERTED",
        }
    )
    producer.validate_static_config(config)
    return config


def bound_config() -> dict:
    config = unknown_config()
    binding = config["implementation_binding"]
    binding.update(
        {
            "status": "BOUND",
            "implementation_commit": "1" * 40,
            "implementation_script_sha256": "2" * 64,
            "implementation_test_sha256": "3" * 64,
        }
    )
    producer.validate_static_config(config)
    return config


def fake_production_summary(config: dict) -> producer.AuditSummary:
    members = {
        item["asset_id"]: item for item in config["source_contract"]["members"]
    }
    return producer.AuditSummary(
        processed_pair_count=6772,
        canonical_record_count=6547,
        s2_only_pair_count=113,
        s3_only_pair_count=0,
        locator_merkle_root_sha256="4" * 64,
        table_s2_sha256=members["PMC10540565_TABLE_S2"]["sha256"],
        table_s3_sha256=members["PMC10540565_TABLE_S3"]["sha256"],
    )


S2_HEADER = ["ID", "Type", "201bp", "5' End", "3'End", "Full_Oligo"]
PRIMARY_HEADER = [
    "barcode",
    "Gene",
    "Comparison",
    "xtail_log2FC_TE",
    "xtail_pvalue",
    "xtail_FDR",
    "Translation_Sig",
]
CONTROL_HEADER = [
    "barcode",
    "TE_1",
    "TE_2",
    "TE_3",
    "TE_4",
    "TE_5",
    "TE_6",
    "Txn_1",
    "Txn_2",
    "Txn_3",
    "Txn_4",
    "Txn_5",
    "Txn_6",
]
PAIR_1 = "GENE:1_A-C"
PAIR_2 = "GENE:2_G-T"
PAIR_3 = "GENE:3_C-A"


def make_s2(*, bad_header: bool = False) -> bytes:
    stream = io.StringIO(newline="")
    writer = csv.writer(stream, lineterminator="\r\n")
    header = list(S2_HEADER)
    if bad_header:
        header[0] = "wrong"
    writer.writerow(header)
    opaque = "ACGT" * 55
    rows = [
        [PAIR_1, "WT", opaque, "opaque", "opaque", "opaque"],
        [PAIR_1, "Mutant", opaque[:-1] + "A", "opaque", "opaque", "opaque"],
        [PAIR_2, "WT", opaque, "quoted,opaque", "opaque", "opaque"],
        [PAIR_2, "Mutant", opaque[:-1] + "C", "quoted,opaque", "opaque", "opaque"],
        [PAIR_2, "WT", opaque, "quoted,opaque", "opaque", "opaque"],
        [PAIR_2, "Mutant", opaque[:-1] + "C", "quoted,opaque", "opaque", "opaque"],
        [PAIR_3, "WT", opaque, "opaque\ncontinued", "opaque", "opaque"],
        [PAIR_3, "Mutant", opaque[:-1] + "G", "opaque\ncontinued", "opaque", "opaque"],
        ["CTRL", "Control", "opaque", "opaque", "opaque", "opaque"],
    ]
    writer.writerows(rows)
    return stream.getvalue().encode("utf-8")


def make_s3(
    *,
    second_key: str = PAIR_2,
    primary_mixed: bool = False,
    bad_header: bool = False,
    significance_label: str = "Significant",
    official_row_geometry: bool = False,
) -> bytes:
    workbook = Workbook()
    primary = workbook.active
    primary.title = "S2A_Polysome_MPRA_Mut_Stats"
    header = list(PRIMARY_HEADER)
    if bad_header:
        header[0] = "wrong"
    primary.append(header)
    primary.append([PAIR_1, "forbidden_gene_one", "HighPoly:RNA", 1.1, 0.2, 0.3, significance_label])
    primary.append([PAIR_1, "forbidden_gene_one", "TotalPoly:RNA", 1.2, 0.2, 0.3, significance_label])
    primary.append([second_key, "forbidden_gene_two", "HighPoly:RNA", 2.1, 0.4, 0.5, "Not Significant"])
    if primary_mixed:
        primary.append([second_key, "forbidden_gene_two", "TotalPoly:RNA", "NA", 0.4, "NA", "Not Significant"])
    else:
        primary.append([second_key, "forbidden_gene_two", "TotalPoly:RNA", "NA", "NA", "NA", "Not Significant"])
    control = workbook.create_sheet("S2B_Poly_MPRA_Control_Stats")
    control.append(CONTROL_HEADER)
    control.append(["forbidden_control"] + list(range(1, 13)))
    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()
    payload = stream.getvalue()
    return add_official_row_geometry(payload) if official_row_geometry else payload


def add_official_row_geometry(payload: bytes) -> bytes:
    source = io.BytesIO(payload)
    target = io.BytesIO()
    with zipfile.ZipFile(source, "r") as incoming, zipfile.ZipFile(target, "w") as outgoing:
        for info in incoming.infolist():
            member = incoming.read(info.filename)
            if info.filename in {
                "xl/worksheets/sheet1.xml",
                "xl/worksheets/sheet2.xml",
            }:
                width = 7 if info.filename.endswith("sheet1.xml") else 13
                member = member.replace(
                    b'<worksheet xmlns="http://schemas.openxmlformats.org/'
                    b'spreadsheetml/2006/main">',
                    b'<worksheet xmlns="http://schemas.openxmlformats.org/'
                    b'spreadsheetml/2006/main" xmlns:x14ac="'
                    + producer.X14AC_NAMESPACE_URI
                    + b'">',
                    1,
                )
                member = re.sub(
                    rb'<row r="([1-9][0-9]*)">',
                    rb'<row r="\1" spans="1:'
                    + str(width).encode("ascii")
                    + rb'" x14ac:dyDescent="0.25">',
                    member,
                )
            outgoing.writestr(info, member)
    return target.getvalue()


def tiny_locator_contract() -> dict:
    contract = copy.deepcopy(load_config()["locator_contract"])
    contract.update(
        {
            "processed_pair_count": 2,
            "canonical_record_count": 1,
            "primary_na_pair_count": 1,
        }
    )
    contract["table_s2"] = {
        "exact_header": S2_HEADER,
        "canonical_compact_header_json_sha256": producer._canonical_header_sha256(S2_HEADER),
        "raw_row_count": 9,
        "unique_content_row_count": 7,
        "duplicate_extra_row_count": 2,
        "duplicated_content_group_count": 2,
        "duplicated_content_multiplicity": 2,
        "duplicated_pair_count": 1,
        "raw_type_counts": {"Control": 1, "Mutant": 4, "WT": 4},
        "deduplicated_type_counts": {"Control": 1, "Mutant": 3, "WT": 3},
        "raw_id_row_multiplicity_counts": {"1": 1, "2": 2, "4": 1},
        "deduplicated_pair_count": 3,
        "deduplicated_control_count": 1,
        "pair_id_grammar": producer.PAIR_ID_RE.pattern,
        "pair_rule": "EXACTLY_ONE_WT_AND_ONE_MUTANT_AFTER_EXACT_FULL_ROW_DEDUPLICATION",
    }
    contract["table_s3"] = {
        "exact_sheet_names": [
            "S2A_Polysome_MPRA_Mut_Stats",
            "S2B_Poly_MPRA_Control_Stats",
        ],
        "primary_sheet_name": "S2A_Polysome_MPRA_Mut_Stats",
        "primary_exact_header": PRIMARY_HEADER,
        "primary_header_sha256": producer._canonical_header_sha256(PRIMARY_HEADER),
        "primary_data_row_count": 4,
        "pair_key_count": 2,
        "comparison_row_counts": {"HighPoly:RNA": 2, "TotalPoly:RNA": 2},
        "comparison_count_per_pair": 2,
        "finite_statistic_rows": {"HighPoly:RNA": 2, "TotalPoly:RNA": 1},
        "na_statistic_rows": {"HighPoly:RNA": 0, "TotalPoly:RNA": 1},
        "both_comparisons_finite_pair_count": 1,
        "primary_only_finite_pair_count": 0,
        "secondary_only_finite_pair_count": 1,
        "neither_comparison_finite_pair_count": 0,
        "statistics_missing_token": "NA",
        "statistics_must_be_all_numeric_present_or_all_exact_na": True,
        "control_sheet_name": "S2B_Poly_MPRA_Control_Stats",
        "control_exact_header": CONTROL_HEADER,
        "control_header_sha256": producer._canonical_header_sha256(CONTROL_HEADER),
        "control_data_row_count": 1,
        "control_data_access_policy": "DIMENSION_AND_HEADER_ONLY_DO_NOT_READ_DATA_CELLS",
    }
    contract["join"] = {
        "table_s2_pair_count": 3,
        "table_s3_pair_count": 2,
        "joined_pair_count": 2,
        "table_s2_absent_from_table_s3_count": 1,
        "table_s3_not_in_table_s2_count": 0,
        "s3_pair_set_must_be_subset_of_s2": True,
        "each_s3_pair_must_join_exactly_one_s2_pair": True,
    }
    return contract


def source_contract_for(directory: Path, s2: bytes, s3: bytes) -> dict:
    config = load_config()
    source = copy.deepcopy(config["source_contract"])
    source["data_root"] = os.fspath(directory)
    contents = {
        "ASSET_ACQUISITION_MANIFEST.json": b"manifest",
        "NCBI_PRJNA824033_RUNINFO.csv": b"runinfo",
        "NIHMS1928233-supplement-3.csv": s2,
        "NIHMS1928233-supplement-4.xlsx": s3,
        "PUBLICATION_COMMIT.json": b"marker",
        "SHA256SUMS": b"checksums",
        "slschuster_3UTRMutationalMPRA-v1.2.zip": b"opaque",
    }
    by_name = {item["relative_path"]: item for item in source["members"]}
    for name, payload in contents.items():
        item = by_name[name]
        item["bytes"] = len(payload)
        item["sha256"] = hashlib.sha256(payload).hexdigest()
    return source, contents


def write_source_directory(directory: Path, contents: dict[str, bytes]) -> None:
    directory.mkdir()
    for name, payload in contents.items():
        (directory / name).write_bytes(payload)


def predecessor_fixture(root: Path) -> tuple[dict, dict, dict[str, bytes]]:
    config = load_config()
    predecessor = copy.deepcopy(
        config["authority_inputs"]["required_predecessor_authority"]
    )
    predecessor["trusted_absolute_bundle_path"] = os.fspath(root)
    predecessor["terminal_marker_final_output_target_sha256"] = hashlib.sha256(
        os.fspath(root).encode("utf-8")
    ).hexdigest()
    payloads = {
        "INPUT_INTEGRITY_AUDIT.json": producer.json_bytes({"aggregate_only": True}),
        "PUBLISHED_ENDPOINT_AUDIT.json": producer.json_bytes({"aggregate_only": True}),
        "QUALIFICATION_REPORT.json": producer.json_bytes({"qualified": False}),
    }
    payloads["SHA256SUMS"] = "".join(
        f"{hashlib.sha256(payloads[name]).hexdigest()}  {name}\n"
        for name in sorted(payloads)
    ).encode("ascii")
    declared = sorted(payloads)
    payloads["PUBLICATION_COMMIT.json"] = producer.json_bytes(
        {
            "record_type": "GSE200304_PUBLISHED_ENDPOINT_A1_PUBLICATION_COMMIT",
            "bundle_member_names": declared,
            "bundle_member_count": 4,
            "sha256sums_sha256": hashlib.sha256(
                payloads["SHA256SUMS"]
            ).hexdigest(),
            "final_output_target_sha256": predecessor[
                "terminal_marker_final_output_target_sha256"
            ],
            "terminal_publication_operation": "FSYNCED_STAGED_HARDLINK_NO_REPLACE",
            "committed": True,
            "terminal_marker_written_last": True,
        }
    )
    predecessor["members"] = [
        {
            "name": name,
            "bytes": len(payloads[name]),
            "sha256": hashlib.sha256(payloads[name]).hexdigest(),
        }
        for name in (
            "INPUT_INTEGRITY_AUDIT.json",
            "PUBLISHED_ENDPOINT_AUDIT.json",
            "QUALIFICATION_REPORT.json",
            "SHA256SUMS",
            "PUBLICATION_COMMIT.json",
        )
    ]
    lineage = predecessor["runtime_lineage_authority"]
    runtime_config = {
        "event_id": lineage["event_id"],
        "protocol_id": lineage["protocol_id"],
        "contract_id": producer.CONTRACT_ID,
        "phase_id": producer.PHASE_ID,
        "dataset_id": producer.DATASET_ID,
        "implementation_binding": {
            "status": "BOUND",
            "implementation_commit": lineage["implementation_commit"],
            "implementation_script_sha256": lineage["implementation_script_sha256"],
            "implementation_test_sha256": lineage["implementation_test_sha256"],
            "compiled_core_sha256": lineage["compiled_core_sha256"],
        },
        "runtime": {
            "artifact_root": os.fspath(root),
            "artifact_members": [
                {**member, "artifact_type": f"TEST_{index}"}
                for index, member in enumerate(predecessor["members"])
            ],
        },
        "artifact_truth": {
            "publication_state": "COMMITTED_ACCEPTED",
            "terminal_record_type": (
                "GSE200304_PUBLISHED_ENDPOINT_A1_PUBLICATION_COMMIT"
            ),
            "terminal_marker_written_last": True,
            "terminal_publication_operation": (
                "FSYNCED_STAGED_HARDLINK_NO_REPLACE"
            ),
            "terminal_declared_member_names": declared,
        },
    }
    config["authority_inputs"]["required_predecessor_authority"] = predecessor
    return config, runtime_config, payloads


def test_frozen_config_supports_exact_i_or_b_lifecycle_state():
    config = load_config()
    producer.validate_static_config(config)
    assert producer.config_core_sha256(config) == producer.FROZEN_CONFIG_CORE_SHA256
    binding = config["implementation_binding"]
    assert binding["status"] in {"UNKNOWN_NOT_ASSERTED", "BOUND"}
    if binding["status"] == "UNKNOWN_NOT_ASSERTED":
        assert all(
            binding[key] == "UNKNOWN_NOT_ASSERTED"
            for key in (
                "implementation_commit",
                "implementation_script_sha256",
                "implementation_test_sha256",
            )
        )
    else:
        producer.validate_implementation_binding(config)
        assert binding["implementation_commit"] != "UNKNOWN_NOT_ASSERTED"
        assert binding["implementation_script_sha256"] == hashlib.sha256(
            SCRIPT.read_bytes()
        ).hexdigest()
        assert binding["implementation_test_sha256"] == hashlib.sha256(
            Path(__file__).read_bytes()
        ).hexdigest()
    assert config["locator_contract"]["join"] == {
        "table_s2_pair_count": 6885,
        "table_s3_pair_count": 6772,
        "joined_pair_count": 6772,
        "table_s2_absent_from_table_s3_count": 113,
        "table_s3_not_in_table_s2_count": 0,
        "s3_pair_set_must_be_subset_of_s2": True,
        "each_s3_pair_must_join_exactly_one_s2_pair": True,
    }
    assert config["locator_contract"]["canonical_record_count"] == 6547
    assert config["authority_inputs"]["dec019_gse200304_v3"]["config_sha256"] == (
        V3_BOUND_CONFIG_SHA256
    )
    assert config["repository_authority"]["implementation_base_commit"] == (
        "de35ce44d7744b89c8b52291343d9f1d6ea674a0"
    )
    assert len(config["source_contract"]["members"]) == 7
    assert config["repository_authority"]["authority_file_git_mode"] == "100644"
    assert config["repository_authority"][
        "worktree_authority_files_must_be_regular_single_link"
    ] is True


def test_unknown_binding_stops_before_any_source_or_output_callback():
    calls: list[str] = []

    def source_reader(_config):
        calls.append("source")
        raise AssertionError("must not be called")

    def output_factory(_config):
        calls.append("output")
        raise AssertionError("must not be called")

    with pytest.raises(producer.BindingError, match="stopped before source/output"):
        producer.execute(
            unknown_config(),
            production=False,
            source_reader=source_reader,
            output_directory_factory=output_factory,
        )
    assert calls == []


def test_production_execution_rejects_all_injectable_components_before_authority(
    monkeypatch,
):
    calls: list[str] = []

    def forbidden(*_args, **_kwargs):
        calls.append("called")
        raise AssertionError("production callback must not be called")

    monkeypatch.setattr(producer, "validate_production_authority", forbidden)
    with pytest.raises(producer.ScopeViolation, match="forbids injectable"):
        producer.execute(
            bound_config(),
            production=True,
            source_reader=forbidden,
        )
    assert calls == []


def test_i_to_b_lifecycle_is_exact_four_scalars_only():
    before = unknown_config()
    after = bound_config()
    producer.validate_i_to_b_config_pair(
        before,
        after,
        implementation_commit="1" * 40,
        script_sha256="2" * 64,
        test_sha256="3" * 64,
    )
    drifted = copy.deepcopy(after)
    drifted["locator_contract"]["canonical_record_count"] = 6546
    with pytest.raises(producer.BindingError):
        producer.validate_i_to_b_config_pair(
            before,
            drifted,
            implementation_commit="1" * 40,
            script_sha256="2" * 64,
            test_sha256="3" * 64,
        )


def test_gate_record_is_current_v3_validator_compatible_and_exact():
    config = bound_config()
    record = producer.build_gate_record(config, fake_production_summary(config))
    v3 = _load_module(
        V3_ROOT / "scripts/route_a_v3/adjudicate_gse200304_dec019_reported_endpoint_a1_v3.py",
        "g200_v3_compatibility_validator",
    )
    v3_config_payload = v3_bound_config_payload()
    v3_config = json.loads(v3_config_payload)
    slot = next(
        item
        for item in v3_config["evidence_contract"]["slots"]
        if item["slot_id"] == producer.GATE_ID
    )
    accepted = v3._validate_gate_record(producer.json_bytes(record), slot, v3_config)
    assert accepted["facts"] == {
        "deterministic_row_locator_frozen": True,
        "table_s2_hash_bound": True,
        "table_s3_hash_bound": True,
        "s2_s3_join_rule_frozen": True,
        "multi_asset_lineage_closed": True,
        "canonical_record_count": 6547,
        "processed_pair_count": 6772,
        "raw_replay_role": producer.RAW_REPLAY_ROLE,
        "raw_replay_status": "NOT_RUN",
        "independent_raw_reproduction_claimed": False,
        "locator_lineage_commitment_algorithm": (
            producer.LOCATOR_LINEAGE_COMMITMENT_ALGORITHM
        ),
        "locator_lineage_merkle_root_sha256": "4" * 64,
    }
    producer._validate_v3_compatibility(
        v3_config_payload,
        config,
    )
    v3_script_payload = (
        V3_ROOT
        / "scripts/route_a_v3/adjudicate_gse200304_dec019_reported_endpoint_a1_v3.py"
    ).read_bytes()
    consumed = producer._consume_gate_with_v3_assets(
        producer.json_bytes(record),
        config,
        v3_config_payload=v3_config_payload,
        v3_script_payload=v3_script_payload,
    )
    assert consumed == record
    rejected = copy.deepcopy(record)
    rejected["privacy"]["contains_row_identifier"] = True
    with pytest.raises(producer.GateProducerError, match="consumer rejected"):
        producer._consume_gate_with_v3_assets(
            producer.json_bytes(rejected),
            config,
            v3_config_payload=v3_config_payload,
            v3_script_payload=v3_script_payload,
        )


def test_s2_selective_parser_captures_only_header_then_id_type():
    rows = list(producer.iter_rfc4180_selective(make_s2()))
    assert rows[0].captured == tuple(S2_HEADER)
    assert all(len(row.captured) == 2 for row in rows[1:])
    assert all("ACGTACGTACGTACGTACGT" not in value for row in rows for value in row.captured)
    contract = tiny_locator_contract()
    audit = producer.audit_table_s2(make_s2(), contract["table_s2"], contract)
    assert len(audit.pair_keys) == 3
    assert all(isinstance(key, bytes) and len(key) == 32 for key in audit.pair_keys)


def test_domain_hash_length_prefixes_domain_and_every_payload_field():
    assert producer._domain_hash(b"a", (b"bc",)) != producer._domain_hash(
        b"ab", (b"c",)
    )
    assert producer._domain_hash(b"domain-a", (b"left", b"right")) != (
        producer._domain_hash(b"domain-b", (b"left", b"right"))
    )
    assert producer._domain_hash(b"domain", (b"left", b"right")) != (
        producer._domain_hash(b"domain", (b"leftright",))
    )


def test_selective_s2_s3_audit_passes_without_significance_membership():
    contract = tiny_locator_contract()
    first = producer.audit_tables(make_s2(), make_s3(significance_label="Significant"), contract)
    second = producer.audit_tables(
        make_s2(), make_s3(significance_label="arbitrary_non_gate_label"), contract
    )
    assert first.processed_pair_count == second.processed_pair_count == 2
    assert first.canonical_record_count == second.canonical_record_count == 1
    assert first.s2_only_pair_count == second.s2_only_pair_count == 1
    assert first.s3_only_pair_count == second.s3_only_pair_count == 0
    # The internal lineage root binds the whole immutable source hash, so a
    # label-byte change changes provenance without changing membership.
    assert first.locator_merkle_root_sha256 != second.locator_merkle_root_sha256


def test_s3_poison_values_never_reach_selected_element_or_control_full_reader(
    monkeypatch,
):
    original_selected = producer._selected_element
    original_member = producer._xlsx_read_member
    selected_payloads: list[bytes] = []
    prefix_calls: list[str] = []

    def guarded_selected(span, payload, *, label):
        raw = payload[span.start : span.end]
        for poison in (
            b"forbidden_gene_one",
            b"forbidden_gene_two",
            b"Significant",
            b"Not Significant",
            b"<v>1.1</v>",
            b"<v>1.2</v>",
            b"<v>2.1</v>",
        ):
            assert poison not in raw
        selected_payloads.append(raw)
        return original_selected(span, payload, label=label)

    def guarded_member(archive, name, **kwargs):
        assert name != "xl/worksheets/sheet2.xml"
        return original_member(archive, name, **kwargs)

    original_prefix = producer._xlsx_read_through_first_row

    def guarded_prefix(archive, name, **kwargs):
        prefix_calls.append(name)
        payload = original_prefix(archive, name, **kwargs)
        assert b"forbidden_control" not in payload
        return payload

    monkeypatch.setattr(producer, "_selected_element", guarded_selected)
    monkeypatch.setattr(producer, "_xlsx_read_member", guarded_member)
    monkeypatch.setattr(producer, "_xlsx_read_through_first_row", guarded_prefix)
    producer.audit_tables(
        make_s2(),
        make_s3(official_row_geometry=True),
        tiny_locator_contract(),
    )
    assert selected_payloads
    assert prefix_calls == ["xl/worksheets/sheet2.xml"]

    numeric_cell = b'<c r="D2" t="n"><v>POISON_NUMERIC_LEXEME</v></c>'
    [span] = list(
        producer._xml_element_spans(numeric_cell, "c", label="numeric poison")
    )
    token = producer._statistic_token(
        span,
        numeric_cell,
        label="numeric poison",
        locator_contract=tiny_locator_contract(),
    )
    assert token.kind == "NUMERIC_PRESENT"
    assert ".text" not in inspect.getsource(producer._statistic_token)
    assert "ET." not in inspect.getsource(producer._extract_primary_tokens)
    assert "ET." not in inspect.getsource(producer._closed_worksheet_row_number)


def _opening_row_tag(payload: bytes):
    return next(
        tag
        for tag in producer._iter_xml_tags(payload, label="structural row fixture")
        if tag.local_name == "row" and tag.kind in {"start", "empty"}
    )


def test_official_row_attribute_geometry_is_accepted_without_value_materialization():
    worksheet = (
        b'<worksheet xmlns="http://schemas.openxmlformats.org/'
        b'spreadsheetml/2006/main" xmlns:x14ac="'
        + producer.X14AC_NAMESPACE_URI
        + b'"><sheetData><row x14ac:dyDescent="0.25" r="1" '
        b'spans="1:7"></row></sheetData></worksheet>'
    )
    assert producer._worksheet_has_official_x14ac_namespace(
        worksheet,
        label="structural worksheet fixture",
    ) is True
    assert producer._closed_worksheet_row_number(
        _opening_row_tag(worksheet),
        label="structural row fixture",
        official_x14ac_namespace=True,
        expected_span="1:7",
    ) == 1
    assert ".text" not in inspect.getsource(producer._closed_worksheet_row_number)
    assert "ET." not in inspect.getsource(producer._lexical_tag_attributes)


@pytest.mark.parametrize(
    ("raw_tag", "official_namespace"),
    [
        (b'<row r="1" spans="1:7" x14ac:dyDescent="0.25" extra="x"/>', True),
        (b'<row r="1" r="2"/>', False),
        (b'<row r=\"1\" illegal?="x"/>', False),
        (b'<row r="1" spans="1:8" x14ac:dyDescent="0.25"/>', True),
        (b'<row r="1" spans="1:7" x14ac:dyDescent="1.25"/>', True),
        (b'<row r="1" spans="1:7" x14ac:dyDescent="0.25"/>', False),
        (b'<row r="1"/>', True),
        (b'<evil:row r="1"/>', False),
    ],
)
def test_row_attribute_geometry_rejects_extra_duplicate_illegal_or_unbound_forms(
    raw_tag,
    official_namespace,
):
    with pytest.raises(producer.TableAuditError):
        producer._closed_worksheet_row_number(
            _opening_row_tag(raw_tag),
            label="structural row fixture",
            official_x14ac_namespace=official_namespace,
            expected_span="1:7",
        )


def test_wrong_official_row_namespace_binding_fails_closed():
    worksheet = (
        b'<worksheet xmlns="http://schemas.openxmlformats.org/'
        b'spreadsheetml/2006/main" xmlns:x14ac="urn:not-the-official-binding">'
        b'<sheetData><row r="1" spans="1:7" '
        b'x14ac:dyDescent="0.25"/></sheetData></worksheet>'
    )
    with pytest.raises(producer.TableAuditError, match="namespace binding differs"):
        producer._worksheet_has_official_x14ac_namespace(
            worksheet,
            label="structural worksheet fixture",
        )


def test_wrong_default_worksheet_namespace_binding_fails_closed():
    worksheet = (
        b'<worksheet xmlns="urn:not-spreadsheetml"><sheetData>'
        b'<row r="1"/></sheetData></worksheet>'
    )
    with pytest.raises(producer.TableAuditError, match="default namespace binding differs"):
        producer._worksheet_has_official_x14ac_namespace(
            worksheet,
            label="structural worksheet fixture",
        )


@pytest.mark.parametrize(
    "declaration",
    [
        b'xmlns="urn:rebound-default"',
        b'xmlns:x14ac="urn:rebound-x14ac"',
    ],
)
def test_descendant_default_or_x14ac_namespace_rebinding_fails_closed(declaration):
    worksheet = (
        b'<worksheet xmlns="'
        + producer.SPREADSHEETML_NAMESPACE_URI
        + b'" xmlns:x14ac="'
        + producer.X14AC_NAMESPACE_URI
        + b'"><sheetData '
        + declaration
        + b'><row r="1" spans="1:7" x14ac:dyDescent="0.25"/>'
        b'</sheetData></worksheet>'
    )
    with pytest.raises(
        producer.TableAuditError,
        match="descendant worksheet namespace declaration is forbidden",
    ):
        producer._worksheet_has_official_x14ac_namespace(
            worksheet,
            label="structural worksheet fixture",
        )


def test_primary_xml_never_slices_forbidden_or_numeric_cell_payloads():
    poisons = (
        b"forbidden_gene_one",
        b"forbidden_gene_two",
        b"Significant",
        b"Not Significant",
        b"<v>1.1</v>",
        b"<v>1.2</v>",
        b"<v>2.1</v>",
    )

    class SliceGuard(bytes):
        def __getitem__(self, key):
            result = super().__getitem__(key)
            if isinstance(key, slice) and isinstance(result, bytes):
                assert all(poison not in result for poison in poisons)
            return result

    with zipfile.ZipFile(io.BytesIO(make_s3()), "r") as archive:
        primary = SliceGuard(archive.read("xl/worksheets/sheet1.xml"))
    headers, rows, wanted = producer._extract_primary_tokens(
        primary,
        tiny_locator_contract(),
    )
    assert len(headers) == 7
    assert len(rows) == 4
    assert isinstance(wanted, dict)


@pytest.mark.parametrize("which", ["s2", "s3"])
def test_header_mismatch_fails_closed(which):
    s2 = make_s2(bad_header=which == "s2")
    s3 = make_s3(bad_header=which == "s3")
    with pytest.raises(producer.TableAuditError, match="header"):
        producer.audit_tables(s2, s3, tiny_locator_contract())


def test_join_membership_mismatch_fails_closed():
    with pytest.raises(producer.TableAuditError, match="subset|join"):
        producer.audit_tables(
            make_s2(), make_s3(second_key="OTHER:9_A-G"), tiny_locator_contract()
        )


def test_count_and_finite_state_mismatch_fail_closed():
    count_contract = tiny_locator_contract()
    count_contract["table_s2"]["raw_row_count"] = 8
    with pytest.raises(producer.TableAuditError, match="raw row count"):
        producer.audit_tables(make_s2(), make_s3(), count_contract)
    with pytest.raises(producer.TableAuditError, match="mixed numeric/NA"):
        producer.audit_tables(make_s2(), make_s3(primary_mixed=True), tiny_locator_contract())


def test_duplicate_locator_digest_is_rejected():
    def collision(_key, _comparison, _contract):
        return b"x" * 32

    contract = tiny_locator_contract()
    contract["canonical_record_count"] = 2
    contract["primary_na_pair_count"] = 0
    contract["table_s3"]["finite_statistic_rows"]["TotalPoly:RNA"] = 2
    contract["table_s3"]["na_statistic_rows"]["TotalPoly:RNA"] = 0
    contract["table_s3"]["both_comparisons_finite_pair_count"] = 2
    contract["table_s3"]["secondary_only_finite_pair_count"] = 0
    # Make both TotalPoly rows finite.
    workbook = make_s3()
    book = __import__("openpyxl").load_workbook(io.BytesIO(workbook))
    sheet = book["S2A_Polysome_MPRA_Mut_Stats"]
    for column in range(4, 7):
        sheet.cell(row=5, column=column).value = 0.5
    output = io.BytesIO()
    book.save(output)
    book.close()
    with pytest.raises(producer.TableAuditError, match="duplicate canonical locator"):
        producer.audit_tables(
            make_s2(), output.getvalue(), contract, locator_digest=collision
        )


def test_source_closure_hashes_all_seven_but_materializes_only_s2_s3(tmp_path):
    s2, s3 = make_s2(), make_s3()
    source, contents = source_contract_for(tmp_path / "source", s2, s3)
    write_source_directory(tmp_path / "source", contents)
    config = load_config()
    config["source_contract"] = source
    phases: list[str] = []
    observed = producer.read_source_inputs(config, fault=phases.append)
    assert observed == {
        "PMC10540565_TABLE_S2": s2,
        "PMC10540565_TABLE_S3": s3,
    }
    assert {
        phase.removeprefix("after_read:")
        for phase in phases
        if phase.startswith("after_read:")
    } == {item["asset_id"] for item in source["members"]}

    (tmp_path / "source" / "NIHMS1928233-supplement-3.csv").write_bytes(s2 + b"x")
    with pytest.raises(producer.InputIntegrityError, match="byte count"):
        producer.read_source_inputs(config)


def test_source_same_size_hash_mismatch_fails_closed(tmp_path):
    s2, s3 = make_s2(), make_s3()
    source, contents = source_contract_for(tmp_path / "source", s2, s3)
    write_source_directory(tmp_path / "source", contents)
    config = load_config()
    config["source_contract"] = source
    changed = bytearray(s2)
    changed[-1] ^= 1
    (tmp_path / "source" / "NIHMS1928233-supplement-3.csv").write_bytes(changed)
    with pytest.raises(producer.InputIntegrityError, match="SHA differs"):
        producer.read_source_inputs(config)


def test_non_scientific_source_member_hash_mismatch_fails_closed(tmp_path):
    s2, s3 = make_s2(), make_s3()
    source, contents = source_contract_for(tmp_path / "source", s2, s3)
    write_source_directory(tmp_path / "source", contents)
    config = load_config()
    config["source_contract"] = source
    path = tmp_path / "source" / "ASSET_ACQUISITION_MANIFEST.json"
    changed = bytearray(path.read_bytes())
    changed[0] ^= 1
    path.write_bytes(changed)
    with pytest.raises(producer.InputIntegrityError, match="SHA differs"):
        producer.read_source_inputs(config)


@pytest.mark.parametrize("violation", ["mode", "hardlink", "symlink"])
def test_worktree_authority_read_requires_mode_nofollow_and_single_link(
    tmp_path,
    violation,
):
    repo = tmp_path / "repo"
    target = repo / "scripts" / "producer.py"
    target.parent.mkdir(parents=True)
    target.write_bytes(b"authority")
    target.chmod(0o644)
    assert producer._read_worktree_authority_file(
        repo, "scripts/producer.py"
    ) == b"authority"
    if violation == "mode":
        target.chmod(0o600)
    elif violation == "hardlink":
        os.link(target, tmp_path / "other-link")
    else:
        target.unlink()
        target.symlink_to(tmp_path / "outside")
        (tmp_path / "outside").write_bytes(b"authority")
    with pytest.raises(producer.BindingError):
        producer._read_worktree_authority_file(repo, "scripts/producer.py")


def test_source_extra_member_and_hardlink_fail_closed(tmp_path):
    s2, s3 = make_s2(), make_s3()
    source, contents = source_contract_for(tmp_path / "source", s2, s3)
    write_source_directory(tmp_path / "source", contents)
    config = load_config()
    config["source_contract"] = source
    (tmp_path / "source" / "extra").write_bytes(b"x")
    with pytest.raises(producer.InputIntegrityError, match="seven-member"):
        producer.read_source_inputs(config)
    (tmp_path / "source" / "extra").unlink()
    os.link(
        tmp_path / "source" / "NIHMS1928233-supplement-3.csv",
        tmp_path / "source" / "hardlink",
    )
    (tmp_path / "source" / "hardlink").unlink()
    # nlink returns to one after the test link is removed; create a hard link
    # outside the source bundle so exact membership remains seven.
    os.link(
        tmp_path / "source" / "NIHMS1928233-supplement-3.csv",
        tmp_path / "outside_link",
    )
    with pytest.raises(producer.InputIntegrityError, match="single-link"):
        producer.read_source_inputs(config)


def test_source_file_rename_replacement_is_detected(tmp_path):
    s2, s3 = make_s2(), make_s3()
    source, contents = source_contract_for(tmp_path / "source", s2, s3)
    write_source_directory(tmp_path / "source", contents)
    config = load_config()
    config["source_contract"] = source

    def replace(phase):
        if phase == "after_open:PMC10540565_TABLE_S2":
            target = tmp_path / "source" / "NIHMS1928233-supplement-3.csv"
            target.rename(tmp_path / "s2_original")
            target.write_bytes(s2)

    with pytest.raises(producer.InputIntegrityError, match="replaced"):
        producer.read_source_inputs(config, fault=replace)


def test_predecessor_five_member_bundle_and_evt037_config_replay(tmp_path):
    root = tmp_path / "predecessor"
    root.mkdir()
    config, runtime_config, payloads = predecessor_fixture(root)
    for name, payload in payloads.items():
        (root / name).write_bytes(payload)
    producer._validate_runtime_lineage_config(runtime_config, config)
    predecessor = config["authority_inputs"]["required_predecessor_authority"]
    observed = producer._read_exact_predecessor_bundle(
        root,
        predecessor,
        forbidden_tokens=config["source_contract"]["forbidden_path_tokens"],
    )
    producer._validate_predecessor_payloads(
        observed,
        predecessor,
        runtime_config,
    )
    assert observed == payloads

    drifted_runtime = copy.deepcopy(runtime_config)
    drifted_runtime["runtime"]["artifact_root"] = os.fspath(tmp_path / "other")
    with pytest.raises(producer.BindingError, match="artifact root"):
        producer._validate_runtime_lineage_config(drifted_runtime, config)


def test_predecessor_hash_and_exact_member_closure_fail_closed(tmp_path):
    root = tmp_path / "predecessor"
    root.mkdir()
    config, _runtime_config, payloads = predecessor_fixture(root)
    for name, payload in payloads.items():
        (root / name).write_bytes(payload)
    predecessor = config["authority_inputs"]["required_predecessor_authority"]
    target = root / "QUALIFICATION_REPORT.json"
    changed = bytearray(target.read_bytes())
    changed[0] ^= 1
    target.write_bytes(changed)
    with pytest.raises(producer.InputIntegrityError, match="SHA differs"):
        producer._read_exact_predecessor_bundle(
            root,
            predecessor,
            forbidden_tokens=config["source_contract"]["forbidden_path_tokens"],
        )
    target.write_bytes(payloads["QUALIFICATION_REPORT.json"])
    (root / "unexpected").write_bytes(b"x")
    with pytest.raises(producer.InputIntegrityError, match="five-member"):
        producer._read_exact_predecessor_bundle(
            root,
            predecessor,
            forbidden_tokens=config["source_contract"]["forbidden_path_tokens"],
        )


def test_single_file_publisher_create_idempotent_and_no_overwrite(tmp_path):
    config = bound_config()
    record = producer.build_gate_record(config, fake_production_summary(config))
    payload = producer.json_bytes(record)
    assert producer.publish_single_gate(payload, tmp_path) == "CREATED_EXCLUSIVE"
    target = tmp_path / producer.OUTPUT_BASENAME
    assert target.read_bytes() == payload
    assert target.stat().st_nlink == 1
    assert producer.publish_single_gate(payload, tmp_path) == "EXISTING_EXACT_IDEMPOTENT"
    target.write_bytes(b"different")
    with pytest.raises(producer.AmbiguousPublicationError, match="differs"):
        producer.publish_single_gate(payload, tmp_path)


def test_publisher_prelink_fault_is_confirmed_absent(tmp_path):
    payload = producer.json_bytes(
        producer.build_gate_record(bound_config(), fake_production_summary(bound_config()))
    )

    def fail(observed):
        if observed == "after_temp_fsync":
            raise RuntimeError("injected")

    with pytest.raises(producer.PublicationError, match="CONFIRMED_ABSENT"):
        producer.publish_single_gate(payload, tmp_path, fault=fail)
    assert not (tmp_path / producer.OUTPUT_BASENAME).exists()
    assert not list(tmp_path.glob(f".{producer.OUTPUT_BASENAME}.tmp.*"))


@pytest.mark.parametrize(
    "phase",
    ["after_final_link", "after_temp_unlink", "before_final_accept"],
)
def test_publisher_postlink_fault_rolls_forward_to_committed_exact(tmp_path, phase):
    payload = producer.json_bytes(
        producer.build_gate_record(bound_config(), fake_production_summary(bound_config()))
    )

    def fail(observed):
        if observed == phase:
            raise RuntimeError("injected")

    assert producer.publish_single_gate(payload, tmp_path, fault=fail) == (
        "COMMITTED_EXACT_AFTER_RECOVERY"
    )
    target = tmp_path / producer.OUTPUT_BASENAME
    assert target.read_bytes() == payload
    assert target.stat().st_nlink == 1
    assert not list(tmp_path.glob(f".{producer.OUTPUT_BASENAME}.tmp.*"))


def test_publisher_recovers_from_one_directory_fsync_failure_after_link(
    tmp_path,
    monkeypatch,
):
    payload = producer.json_bytes(
        producer.build_gate_record(bound_config(), fake_production_summary(bound_config()))
    )
    real_fsync = producer.os.fsync
    calls = 0

    def fail_first_postlink_directory_fsync(descriptor):
        nonlocal calls
        calls += 1
        if calls == 2:
            raise OSError("injected post-link directory fsync failure")
        return real_fsync(descriptor)

    monkeypatch.setattr(producer.os, "fsync", fail_first_postlink_directory_fsync)
    assert producer.publish_single_gate(payload, tmp_path) == (
        "COMMITTED_EXACT_AFTER_RECOVERY"
    )
    target = tmp_path / producer.OUTPUT_BASENAME
    assert target.read_bytes() == payload
    assert target.stat().st_nlink == 1


def test_publisher_persistent_postlink_fsync_failure_is_explicit_unverified(
    tmp_path,
    monkeypatch,
):
    payload = producer.json_bytes(
        producer.build_gate_record(bound_config(), fake_production_summary(bound_config()))
    )
    real_fsync = producer.os.fsync
    calls = 0

    def fail_all_postlink_directory_fsyncs(descriptor):
        nonlocal calls
        calls += 1
        if calls >= 2:
            raise OSError("injected persistent directory fsync failure")
        return real_fsync(descriptor)

    monkeypatch.setattr(producer.os, "fsync", fail_all_postlink_directory_fsyncs)
    with pytest.raises(producer.PublicationStateUnverifiedError, match="UNVERIFIED"):
        producer.publish_single_gate(payload, tmp_path)
    target = tmp_path / producer.OUTPUT_BASENAME
    assert target.read_bytes() == payload
    assert target.stat().st_nlink == 1


def test_publisher_stale_temp_and_multilink_existing_are_ambiguous(tmp_path):
    payload = producer.json_bytes(
        producer.build_gate_record(bound_config(), fake_production_summary(bound_config()))
    )
    stale = tmp_path / f".{producer.OUTPUT_BASENAME}.tmp.stale"
    stale.write_bytes(payload)
    with pytest.raises(producer.AmbiguousPublicationError, match="stale"):
        producer.publish_single_gate(payload, tmp_path)
    stale.unlink()
    target = tmp_path / producer.OUTPUT_BASENAME
    target.write_bytes(payload)
    os.link(target, tmp_path.parent / f"{tmp_path.name}-second-link")
    with pytest.raises(producer.AmbiguousPublicationError, match="single-link"):
        producer.publish_single_gate(payload, tmp_path)


def test_publisher_temp_path_replacement_is_unverified_and_not_unlinked(tmp_path):
    payload = producer.json_bytes(
        producer.build_gate_record(bound_config(), fake_production_summary(bound_config()))
    )
    replacement: Path | None = None

    def replace(phase):
        nonlocal replacement
        if phase == "after_temp_fsync":
            [staged] = list(tmp_path.glob(f".{producer.OUTPUT_BASENAME}.tmp.*"))
            staged.rename(tmp_path / "original-staged-inode")
            staged.write_bytes(b"replacement")
            replacement = staged

    with pytest.raises(producer.PublicationStateUnverifiedError, match="UNVERIFIED"):
        producer.publish_single_gate(payload, tmp_path, fault=replace)
    assert replacement is not None and replacement.read_bytes() == b"replacement"
    assert (tmp_path / "original-staged-inode").read_bytes() == payload


def test_output_directory_rename_replacement_is_detected(tmp_path):
    output = tmp_path / "output"
    output.mkdir()
    payload = producer.json_bytes(
        producer.build_gate_record(bound_config(), fake_production_summary(bound_config()))
    )

    def replace(phase):
        if phase == "after_temp_fsync":
            output.rename(tmp_path / "moved")
            output.mkdir()

    with pytest.raises(producer.PublicationStateUnverifiedError, match="UNVERIFIED"):
        producer.publish_single_gate(payload, output, fault=replace)
    assert not (output / producer.OUTPUT_BASENAME).exists()
    # The retained descriptor now names a renamed-out directory.  Its owned
    # temp is deliberately preserved because cleanup there would target a
    # non-canonical parent.
    assert list((tmp_path / "moved").glob(f".{producer.OUTPUT_BASENAME}.tmp.*"))


@pytest.mark.parametrize("success_path", ["created", "existing", "recovery"])
def test_publisher_late_parent_replacement_after_retained_membership_is_unverified(
    tmp_path,
    monkeypatch,
    success_path,
):
    output = tmp_path / "output"
    output.mkdir()
    payload = producer.json_bytes(
        producer.build_gate_record(bound_config(), fake_production_summary(bound_config()))
    )
    if success_path == "existing":
        assert producer.publish_single_gate(payload, output) == "CREATED_EXCLUSIVE"

    original_membership = producer._assert_gate_directory_membership
    moved = tmp_path / "renamed-out-output"
    swapped = False

    def replace_after_retained_membership(directory_fd, *, expected):
        nonlocal swapped
        original_membership(directory_fd, expected=expected)
        if expected == {producer.OUTPUT_BASENAME} and not swapped:
            swapped = True
            output.rename(moved)
            output.mkdir()
            (output / producer.OUTPUT_BASENAME).write_bytes(payload)

    def fault(phase):
        if success_path == "recovery" and phase == "after_final_link":
            raise RuntimeError("force committed recovery path")

    monkeypatch.setattr(
        producer,
        "_assert_gate_directory_membership",
        replace_after_retained_membership,
    )
    with pytest.raises(producer.PublicationStateUnverifiedError, match="UNVERIFIED"):
        producer.publish_single_gate(
            payload,
            output,
            fault=fault if success_path == "recovery" else None,
        )
    assert swapped is True
    assert (moved / producer.OUTPUT_BASENAME).read_bytes() == payload
    assert (output / producer.OUTPUT_BASENAME).read_bytes() == payload


def test_gate_privacy_rejects_forbidden_key_and_never_emits_row_values():
    config = bound_config()
    record = producer.build_gate_record(config, fake_production_summary(config))
    payload = producer.json_bytes(record).decode("ascii")
    for forbidden_value in (
        PAIR_1,
        PAIR_2,
        "forbidden_gene_one",
        "Significant",
        "ACGT" * 55,
    ):
        assert forbidden_value not in payload
    with pytest.raises(producer.GateProducerError, match="forbidden output key"):
        producer._assert_no_forbidden_output(
            {"barcode": "not-persistable"},
            {"barcode"},
            label="test",
        )


def test_valid_all_dna_alphabet_merkle_digest_is_not_misclassified_as_sequence():
    config = bound_config()
    summary = fake_production_summary(config)
    summary = producer.AuditSummary(
        **{
            **summary.__dict__,
            "locator_merkle_root_sha256": "a" * 64,
        }
    )
    record = producer.build_gate_record(config, summary)
    producer.validate_gate_record(record, config)
    assert record["facts"]["locator_lineage_merkle_root_sha256"] == "a" * 64


def test_inspector_recomputes_source_evidence_and_exact_compares_bytes(
    tmp_path,
    monkeypatch,
):
    config = bound_config()
    summary = fake_production_summary(config)
    expected = producer.build_gate_record(config, summary)
    payload = producer.json_bytes(expected)
    producer.publish_single_gate(payload, tmp_path)
    calls: list[str] = []

    def authority(_config):
        calls.append("authority")
        return {"mode": "TEST_AUTHORITY"}

    def audit(_config):
        calls.append("source_and_evidence")
        return expected, summary

    def predecessor(_config):
        calls.append("predecessor")
        return {"member_count": 5}

    def consumer(observed_payload, _config):
        calls.append("consumer")
        assert observed_payload == payload
        return expected

    monkeypatch.setattr(producer, "validate_production_authority", authority)
    monkeypatch.setattr(producer, "replay_predecessor_authority", predecessor)
    monkeypatch.setattr(producer, "_audit_expected_gate", audit)
    monkeypatch.setattr(producer, "consume_gate_with_current_v3", consumer)
    monkeypatch.setattr(producer, "_configured_output_directory", lambda _config: tmp_path)
    result = producer.inspect_production_gate(config)
    assert result["mode"] == "INSPECT_RECOMPUTED_SOURCE_AND_EXACT_GATE"
    assert calls == ["authority", "predecessor", "source_and_evidence", "consumer"]

    monkeypatch.setattr(
        producer,
        "_read_and_validate_gate_at",
        lambda *_args, **_kwargs: (expected, payload + b" "),
    )
    with pytest.raises(producer.AmbiguousPublicationError, match="reconstructed"):
        producer.inspect_production_gate(config)


def test_inspector_late_parent_replacement_after_retained_membership_is_unverified(
    tmp_path,
    monkeypatch,
):
    output = tmp_path / "output"
    output.mkdir()
    config = bound_config()
    summary = fake_production_summary(config)
    expected = producer.build_gate_record(config, summary)
    payload = producer.json_bytes(expected)
    assert producer.publish_single_gate(payload, output) == "CREATED_EXCLUSIVE"
    calls: list[str] = []

    monkeypatch.setattr(
        producer,
        "validate_production_authority",
        lambda _config: calls.append("authority") or {"mode": "TEST_AUTHORITY"},
    )
    monkeypatch.setattr(
        producer,
        "replay_predecessor_authority",
        lambda _config: calls.append("predecessor") or {"member_count": 5},
    )
    monkeypatch.setattr(
        producer,
        "_audit_expected_gate",
        lambda _config: (calls.append("source_and_evidence") or (expected, summary)),
    )
    monkeypatch.setattr(
        producer,
        "consume_gate_with_current_v3",
        lambda *_args: calls.append("consumer") or expected,
    )
    monkeypatch.setattr(
        producer,
        "_configured_output_directory",
        lambda _config: output,
    )

    original_membership = producer._assert_gate_directory_membership
    retained_checks = 0
    moved = tmp_path / "renamed-out-inspection-output"

    def replace_after_final_retained_membership(directory_fd, *, expected):
        nonlocal retained_checks
        original_membership(directory_fd, expected=expected)
        if expected != {producer.OUTPUT_BASENAME}:
            return
        retained_checks += 1
        if retained_checks == 2:
            output.rename(moved)
            output.mkdir()
            (output / producer.OUTPUT_BASENAME).write_bytes(payload)

    monkeypatch.setattr(
        producer,
        "_assert_gate_directory_membership",
        replace_after_final_retained_membership,
    )
    with pytest.raises(producer.PublicationStateUnverifiedError, match="UNVERIFIED"):
        producer.inspect_production_gate(config)
    assert retained_checks == 2
    assert calls == ["authority", "predecessor", "source_and_evidence"]
    assert (moved / producer.OUTPUT_BASENAME).read_bytes() == payload
    assert (output / producer.OUTPUT_BASENAME).read_bytes() == payload


def test_public_inspector_rejects_nontrusted_path(tmp_path):
    with pytest.raises(producer.ScopeViolation, match="trusted output slot"):
        producer.read_and_validate_committed_gate(
            tmp_path / producer.OUTPUT_BASENAME,
            bound_config(),
        )


def test_script_has_no_openpyxl_or_significance_value_access():
    source = SCRIPT.read_text(encoding="utf-8")
    assert "from openpyxl" not in source
    assert "import openpyxl" not in source
    assert "significance =" not in source
    assert "row[6]" not in source
    assert "NUMERIC_XML_RE" not in source
    assert "ET.iterparse" not in source
