"""P1-04: Unified data loader for cross-fitted predictor ensemble.

Supports loading and merging labels from:
    - Sample 2019 MPRA (GSE114002): 5'UTR random 50-mer + designed + variable length
    - Cao 2021 5'UTR TE: GENCODE v17 5'UTR + Ribo-seq TE features (HEK293T/PC3/Muscle)
    - Saluki half-life: human + mouse transcriptome-wide half-life PCA-1
    - CodonBERT stability: mRNA_Stability.csv + CoV_Vaccine_Degradation.csv

All datasets use the same deterministic SHA-256 split contract:
    - Hash key: uppercase sequence (Sample 2019, CodonBERT) or uppercase gene name (Saluki)
    - Bucket: <0.8 train, <0.9 val, else test
    - Test split is FROZEN — no predictor in P1-04 trains on it
"""
from __future__ import annotations

import csv
import gzip
import hashlib
import io
import json
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterator, List, Optional, Tuple

import numpy as np

_REPO_ROOT = Path(__file__).resolve().parents[2]  # mrna_editflow/ (parents[0]=predictors/, [1]=models/, [2]=project root)
DATA_ROOT = _REPO_ROOT / "data" / "raw"


# ---------------------------------------------------------------------------
# Split contract (must match P1-02A/P1-03 manifest scripts)
# ---------------------------------------------------------------------------

def split_bucket_sequence(seq: str) -> str:
    """Deterministic 80/10/10 split via SHA-256 of uppercased sequence."""
    h = hashlib.sha256(seq.upper().encode()).hexdigest()
    v = int(h[:8], 16) / 0xFFFFFFFF
    if v < 0.8:
        return "train"
    elif v < 0.9:
        return "val"
    return "test"


def split_bucket_gene(gene: str) -> str:
    """Deterministic 80/10/10 split via SHA-256 of uppercased gene name."""
    h = hashlib.sha256(gene.upper().encode()).hexdigest()
    v = int(h[:8], 16) / 0xFFFFFFFF
    if v < 0.8:
        return "train"
    elif v < 0.9:
        return "val"
    return "test"


# ---------------------------------------------------------------------------
# Data records
# ---------------------------------------------------------------------------

@dataclass
class PredictorRecord:
    """Unified record for predictor training/eval.

    Attributes:
        sequence: nucleotide sequence (5'UTR, CDS, or full mRNA)
        label: regression target (RL, TE, half-life, stability, degradation rate)
        dataset: source dataset name
        split: train / val / test
        cell_type: cell type if known (e.g., "HEK293T", "HeLa", "Muscle")
        chemistry: RNA chemistry if known (e.g., "unmodified", "pseudouridine")
        gene: gene name if known (Saluki, Cao 2021)
        metadata: extra fields (replicate, library_kind, etc.)
    """
    sequence: str
    label: float
    dataset: str
    split: str
    cell_type: Optional[str] = None
    chemistry: Optional[str] = None
    gene: Optional[str] = None
    metadata: Dict = None


# ---------------------------------------------------------------------------
# Per-dataset loaders
# ---------------------------------------------------------------------------

def load_sample2019(
    gsm_filter: Optional[List[str]] = None,
    split_filter: Optional[List[str]] = None,
) -> Iterator[Dict]:
    """Load Sample 2019 MPRA records.

    Yields dicts: {sequence, label, dataset, split, cell_type, chemistry, replicate, library_kind, gsm}

    Args:
        gsm_filter: If provided, only yield records from these GSMs
        split_filter: If provided, only yield records in these splits
    """
    root = DATA_ROOT / "sample2019_mpra"
    manifest_path = root / "manifest.json"
    if not manifest_path.exists():
        raise FileNotFoundError(f"Sample 2019 manifest not found: {manifest_path}")

    with open(manifest_path) as f:
        manifest = json.load(f)

    for entry in manifest["samples"]:
        if "error" in entry:
            continue
        gsm = entry["gsm"]
        if gsm_filter and gsm not in gsm_filter:
            continue
        filename = entry["filename"]
        path = root / filename
        if not path.exists():
            continue

        # Read CSV.gz
        with gzip.open(path, "rt", newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader)
            norm = [h.strip().lower() for h in header]
            # Locate UTR and RL columns
            utr_col = None
            rl_col = None
            for i, name in enumerate(norm):
                if name in ("utr", "seq", "sequence"):
                    utr_col = i
                elif name in ("rl", "mrl", "ribosome_load"):
                    rl_col = i
            if utr_col is None or rl_col is None:
                continue

            for row in reader:
                if len(row) <= max(utr_col, rl_col):
                    continue
                try:
                    seq = row[utr_col].strip().upper().replace("T", "U")
                    label = float(row[rl_col])
                except (ValueError, IndexError):
                    continue
                if not seq or len(seq) < 5:
                    continue
                split = split_bucket_sequence(seq)
                if split_filter and split not in split_filter:
                    continue
                yield {
                    "sequence": seq,
                    "label": label,
                    "dataset": "sample2019_mpra",
                    "split": split,
                    "cell_type": entry.get("cell_type", "HEK293T"),
                    "chemistry": entry.get("rna_chemistry"),
                    "replicate": entry.get("replicate"),
                    "library_kind": entry.get("library_kind"),
                    "gsm": gsm,
                }


def load_saluki_halflife(
    species: Optional[str] = None,
    split_filter: Optional[List[str]] = None,
) -> Iterator[Dict]:
    """Load Saluki half-life records from .npz files.

    Args:
        species: "human" or "mouse" or None (both)
        split_filter: list of split names to keep
    """
    root = DATA_ROOT / "saluki_halflife"
    manifest_path = root / "manifest.json"
    if not manifest_path.exists():
        raise FileNotFoundError(f"Saluki manifest not found: {manifest_path}")

    with open(manifest_path) as f:
        manifest = json.load(f)

    target_files = []
    for entry in manifest["files"]:
        if "error" in entry:
            continue
        fname = entry["filename"]
        if species == "human" and "human" not in fname:
            continue
        if species == "mouse" and "mouse" not in fname:
            continue
        target_files.append(entry)

    for entry in target_files:
        path = root / entry["filename"]
        if not path.exists():
            continue
        data = np.load(path, allow_pickle=True)
        X = data["X"]  # (N, L, C)
        y = data["y"]  # (N,)
        genes = data["genes"]  # (N,)

        # Decode one-hot to sequence string (one-hot in first 4 channels)
        # X is int8 of shape (N, 12288, 6); channels 0-3 = ACGT, 4 = CDS, 5 = splice
        alphabet = "ACGU"
        N, L, C = X.shape
        # Pre-compute argmax across first 4 channels
        onehot = X[:, :, :4].astype(np.int8)
        idx = onehot.argmax(axis=2)  # (N, L)

        for i in range(N):
            gene = str(genes[i])
            split = split_bucket_gene(gene)
            if split_filter and split not in split_filter:
                continue
            # Build sequence from one-hot
            # Only positions with any onehot set are valid
            mask = onehot[i].sum(axis=1) > 0
            if mask.sum() == 0:
                continue
            seq_indices = idx[i][mask]
            seq = "".join(alphabet[j] for j in seq_indices)
            yield {
                "sequence": seq,
                "label": float(y[i]),
                "dataset": "saluki_halflife",
                "split": split,
                "cell_type": entry.get("cell_types", ["unknown"])[0],
                "chemistry": "unmodified",
                "gene": gene,
                "species": entry.get("species"),
                "length": len(seq),
            }


def load_codonbert_stability(
    kind_filter: Optional[str] = None,
    split_filter: Optional[List[str]] = None,
) -> Iterator[Dict]:
    """Load CodonBERT stability records.

    Args:
        kind_filter: "mRNA_Stability" or "CoV_Vaccine_Degradation"
        split_filter: list of split names to keep
    """
    root = DATA_ROOT / "codonbert_stability"
    manifest_path = root / "manifest.json"
    if not manifest_path.exists():
        raise FileNotFoundError(f"CodonBERT manifest not found: {manifest_path}")

    with open(manifest_path) as f:
        manifest = json.load(f)

    for entry in manifest["files"]:
        if "error" in entry:
            continue
        fname = entry["filename"]
        if kind_filter and kind_filter not in fname:
            continue
        path = root / fname
        if not path.exists():
            continue

        with open(path, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                seq = (row.get("Sequence") or "").strip().upper().replace("T", "U")
                try:
                    label = float(row.get("Value"))
                except (TypeError, ValueError):
                    continue
                if not seq or len(seq) < 5:
                    continue
                split = split_bucket_sequence(seq)
                if split_filter and split not in split_filter:
                    continue
                yield {
                    "sequence": seq,
                    "label": label,
                    "dataset": "codonbert_stability",
                    "split": split,
                    "cell_type": None,
                    "chemistry": "unmodified",
                    "kind": entry.get("kind"),
                    "source_dataset": row.get("Dataset"),
                    "source_split": row.get("Split"),
                }


def load_lepplek2022_persistseq(
    label_kind: str = "half_life",
    split_filter: Optional[List[str]] = None,
) -> Iterator[Dict]:
    """Load Leppek 2022 PERSIST-Seq records from GSE173083 Table S1.

    Yields dicts: {sequence, label, dataset, split, cell_type, chemistry, label_kind, ...}

    Args:
        label_kind: which label to use as regression target.
            Options: "mrl", "half_life", "protein_output",
                     "in_cell_stability", "in_solution_stability"
        split_filter: list of split names to keep
    """
    root = DATA_ROOT / "lepplek2022_persistseq"
    manifest_path = root / "manifest.json"
    if not manifest_path.exists():
        raise FileNotFoundError(f"Lepkek 2022 manifest not found: {manifest_path}")

    with open(manifest_path) as f:
        manifest = json.load(f)

    column_mapping = manifest.get("column_mapping", {})
    seq_col = column_mapping.get("sequence")
    label_col_map = {
        "mrl": column_mapping.get("mrl"),
        "half_life": column_mapping.get("half_life"),
        "protein_output": column_mapping.get("protein_output"),
        "in_cell_stability": column_mapping.get("in_cell_stability"),
        "in_solution_stability": column_mapping.get("in_solution_stability"),
    }
    if label_kind not in label_col_map:
        raise ValueError(
            f"Unknown label_kind={label_kind!r}; "
            f"options={list(label_col_map)}"
        )
    label_col = label_col_map[label_kind]
    if not label_col:
        raise ValueError(
            f"Label {label_kind!r} has no mapped column in manifest"
        )

    # Find the xlsx file (Table S1)
    s1_path = None
    for entry in manifest.get("files", []):
        if entry.get("kind") == "persist_seq_main":
            s1_path = root / entry["filename"]
            break
    if s1_path is None or not s1_path.exists():
        raise FileNotFoundError(
            "GSE173083 Table S1 (persist_seq_main) not found in manifest"
        )

    # Parse xlsx
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "openpyxl is required to parse Lepkek 2022 Table S1. "
            "Run: pip install openpyxl"
        ) from e

    wb = load_workbook(s1_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = list(next(rows_iter))
    cols = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header)]

    # Find column indices
    seq_idx = cols.index(seq_col) if seq_col in cols else None
    label_idx = cols.index(label_col) if label_col in cols else None
    if seq_idx is None or label_idx is None:
        wb.close()
        raise ValueError(
            f"Required columns not found: seq={seq_col!r} label={label_col!r}"
        )

    # Optional metadata columns
    meta_cols = {
        "human_name": "Human-readable name",
        "group": "Group",
        "type_5utr": "Type 5 UTR",
        "type_cds": "Type CDS",
        "type_3utr": "Type 3 UTR",
        "cai": "CAI",
        "gc_content_cds": "GC content (CDS)",
        "rna_length": "RNA length",
    }
    meta_idx = {k: cols.index(v) if v in cols else None
                for k, v in meta_cols.items()}

    for row in rows_iter:
        if row is None or all(v is None for v in row):
            continue
        if seq_idx >= len(row) or label_idx >= len(row):
            continue
        raw_seq = row[seq_idx]
        raw_label = row[label_idx]
        if raw_seq is None or raw_label is None:
            continue
        seq = str(raw_seq).strip().upper().replace("T", "U")
        # Filter to ACGU
        seq = "".join(c for c in seq if c in "ACGU")
        if len(seq) < 20:
            continue
        try:
            label = float(raw_label)
        except (ValueError, TypeError):
            continue
        if not np.isfinite(label):
            continue

        split = split_bucket_sequence(seq)
        if split_filter and split not in split_filter:
            continue

        metadata = {}
        for k, idx in meta_idx.items():
            if idx is not None and idx < len(row) and row[idx] is not None:
                metadata[k] = row[idx]

        yield {
            "sequence": seq,
            "label": label,
            "dataset": "lepplek2022_persistseq",
            "split": split,
            "cell_type": "HEK293T",
            "chemistry": "unmodified",
            "label_kind": label_kind,
            "source_gse": "GSE173083",
            "source_table": "S1",
            "metadata": metadata,
        }
    wb.close()


# ---------------------------------------------------------------------------
# Unified loader
# ---------------------------------------------------------------------------

def load_all(
    datasets: Optional[List[str]] = None,
    split_filter: Optional[List[str]] = None,
) -> Iterator[Dict]:
    """Load records from all P1-02A/P1-03 datasets.

    Args:
        datasets: list of dataset names to include; None = all
            Options: "sample2019_mpra", "saluki_halflife",
                     "codonbert_stability", "lepplek2022_persistseq"
        split_filter: list of split names to keep
    """
    if datasets is None:
        datasets = [
            "sample2019_mpra", "saluki_halflife", "codonbert_stability",
            "lepplek2022_persistseq",
        ]

    if "sample2019_mpra" in datasets:
        yield from load_sample2019(split_filter=split_filter)
    if "saluki_halflife" in datasets:
        yield from load_saluki_halflife(split_filter=split_filter)
    if "codonbert_stability" in datasets:
        yield from load_codonbert_stability(split_filter=split_filter)
    if "lepplek2022_persistseq" in datasets:
        yield from load_lepplek2022_persistseq(split_filter=split_filter)


def load_split_stats() -> Dict[str, Dict[str, int]]:
    """Load split counts per dataset from manifests (no parsing required).

    Returns:
        {dataset_name: {split: count, ...}, ...}
    """
    stats: Dict[str, Dict[str, int]] = {}
    # Sample 2019
    s2019_manifest = DATA_ROOT / "sample2019_mpra" / "manifest.json"
    if s2019_manifest.exists():
        with open(s2019_manifest) as f:
            m = json.load(f)
        counts = {"train": 0, "val": 0, "test": 0}
        for entry in m["samples"]:
            if "error" in entry:
                continue
            for k in counts:
                counts[k] += entry.get("split_stats", {}).get(k, 0)
        stats["sample2019_mpra"] = counts
    # Saluki
    saluki_manifest = DATA_ROOT / "saluki_halflife" / "manifest.json"
    if saluki_manifest.exists():
        with open(saluki_manifest) as f:
            m = json.load(f)
        counts = {"train": 0, "val": 0, "test": 0}
        for entry in m["files"]:
            for k in counts:
                counts[k] += entry.get("split_stats", {}).get(k, 0)
        stats["saluki_halflife"] = counts
    # CodonBERT
    cb_manifest = DATA_ROOT / "codonbert_stability" / "manifest.json"
    if cb_manifest.exists():
        with open(cb_manifest) as f:
            m = json.load(f)
        counts = {"train": 0, "val": 0, "test": 0}
        for entry in m["files"]:
            for k in counts:
                counts[k] += entry.get("split_stats", {}).get(k, 0)
        stats["codonbert_stability"] = counts
    # Leppek 2022 PERSIST-Seq
    lp_manifest = DATA_ROOT / "lepplek2022_persistseq" / "manifest.json"
    if lp_manifest.exists():
        with open(lp_manifest) as f:
            m = json.load(f)
        stats["lepplek2022_persistseq"] = m.get("split_stats", {"train": 0, "val": 0, "test": 0})
    return stats


# ---------------------------------------------------------------------------
# Batching utilities
# ---------------------------------------------------------------------------

def pad_sequences(sequences: List[str], max_len: Optional[int] = None,
                  pad_char: str = "N") -> Tuple[List[str], int]:
    """Pad sequences to common length. Returns (padded_seqs, max_len)."""
    if max_len is None:
        max_len = max(len(s) for s in sequences) if sequences else 0
    padded = [s + pad_char * (max_len - len(s)) for s in sequences]
    return padded, max_len


def onehot_encode(sequences: List[str], max_len: int,
                  alphabet: str = "ACGU") -> np.ndarray:
    """Encode sequences as one-hot (N, L, 4) int8 array."""
    n = len(sequences)
    arr = np.zeros((n, max_len, len(alphabet)), dtype=np.int8)
    char_to_idx = {c: i for i, c in enumerate(alphabet)}
    for i, seq in enumerate(sequences):
        for j, ch in enumerate(seq[:max_len]):
            if ch in char_to_idx:
                arr[i, j, char_to_idx[ch]] = 1
    return arr


__all__ = [
    "PredictorRecord",
    "split_bucket_sequence",
    "split_bucket_gene",
    "load_sample2019",
    "load_saluki_halflife",
    "load_codonbert_stability",
    "load_all",
    "load_split_stats",
    "pad_sequences",
    "onehot_encode",
]
