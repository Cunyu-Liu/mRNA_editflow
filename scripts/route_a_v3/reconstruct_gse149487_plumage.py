#!/usr/bin/env python3
"""Deterministically reconstruct a development-only GSE149487 companion table.

The two accepted inputs are the ordinary-public MOESM8 supplementary workbook
and the separately materialized Lim Table 6c HEK293T workbook.  This program is
deliberately fail closed: all caller-supplied paths and both expected SHA-256
digests are checked before either workbook is opened, workbook structure is
exactly validated, symlinks and overwrites are rejected, and no raw barcode ID
is emitted.

The generated effect is *not* the paper's two-sided Mann-Whitney/FDR analysis.
The Lim 6c component columns contain published log2(CPM per barcode), so each
barcode log2 ratio is computed by component subtraction after applying the
inclusive -1 log2CPM floor equivalent to the paper's 0.5 CPM floor.  This is a
Route A companion summary: barcode-level log2 ratios are reduced to a construct
median within each biological replicate, mutant-minus-WT replicate deltas are
averaged across exactly three biological replicates, and the SE is the sample
SD of those three deltas divided by sqrt(3).  Barcodes remain technical units
and never become the effective N.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import re
import secrets
import stat
import statistics
import tempfile
import unicodedata
from collections import Counter
from numbers import Real
from pathlib import Path
from typing import Any, BinaryIO, Callable, Mapping, Sequence

import pandas as pd
from openpyxl import load_workbook


CONTRACT_ID = "mrna_xeditflow_route_a_v3"
DATASET_ID = "GSE149487"
DATASET_ALIAS = "PLUMAGE"
STUDY_GROUP_ID = "PLUMAGE_LIM_2021"
INDEPENDENT_STUDY_ID = STUDY_GROUP_ID
SCHEMA_VERSION = "1.0.0"
RECONSTRUCTION_STATUS = "DEVELOPMENT_RECONSTRUCTED_NOT_QUALIFIED"
PROTOCOL_ID = "ROUTE_A_V3_GSE149487_PLUMAGE_PARTIAL_RECONSTRUCTION_V1"
PROTOCOL_STATUS = "PREFROZEN_BEFORE_REAL_RECONSTRUCTION_RESULTS"
DESIGN_FAMILY = "PLUMAGE_FULL_LENGTH_5UTR_SNV"
CONTEXT = "293T"
REGION = "5UTR"
MOESM8_SOURCE_ASSET_ID = "GSE149487_MOESM8"
LIM6C_SOURCE_ASSET_ID = "GSE149487_LIM6C_293T"
EXPECTED_MOESM8_FILENAME = "41467_2021_24445_MOESM8_ESM.xlsx"
EXPECTED_LIM6C_FILENAME = "Lim_et_al_Supp_Tbl_6c_293T.xlsx"
PUBLIC_LIM6C_SHA256 = (
    "f991b39b85efce98fe752887a49052043c495ae81b6dfc5dfeded09013b591df"
)
PRIMARY_MIN_ELIGIBLE_BARCODES_PER_ARM_REPLICATE = 20
SECONDARY_AUDIT_ONLY_BARCODE_FLOORS: tuple[int, ...] = (10, 50)
ORIGINAL_CPM_MINIMUM_INCLUSIVE = 0.5
PUBLISHED_LOG2_CPM_MINIMUM_INCLUSIVE = -1.0
PROTOCOL_UNRESOLVED_BLOCKERS: tuple[str, ...] = (
    "PC3_AND_18_GEO_RAW_COUNT_TABLE_JOIN_NOT_INCLUDED",
    "SUPPLEMENTS_NOT_LISTED_IN_CURRENT_P0_MANIFEST",
    "LICENSE_AND_REDISTRIBUTION_NOT_CLOSED",
    "CHECKPOINT_SPECIFIC_EXPOSURE_NOT_CLOSED",
    "GROUP_LEAKAGE_AND_SPLIT_NOT_FROZEN",
    "PAPER_NATIVE_MANN_WHITNEY_AND_MULTIPLE_TESTING_NOT_REPRODUCED",
    "CANONICAL_INTERVENTION_RECORD_V3_NOT_MATERIALIZED",
    "UNADJUDICATED_DESCRIPTION_CLASSES_EXCLUDED",
    "UNADJUDICATED_6A_COORDINATE_CLASSES_EXCLUDED",
)

FORBIDDEN_PATH_TOKENS: tuple[str, ...] = (
    "sealed",
    "restricted",
    "gse246381",
)
SHA256_RE = re.compile(r"^[0-9a-f]{64}$")

MOESM8_SHEETS: tuple[str, ...] = (
    "6a 5' UTR sequences",
    "6b short-read counts",
    "6c - histogram raw data",
    "6d transcript FDR<0.1",
    "6e TE FDR<0.1",
)
MOESM8_SEQUENCE_SHEET = MOESM8_SHEETS[0]
MOESM8_6A_COLUMNS: tuple[str, ...] = (
    "5' UTR length (bp)",
    "Gene name",
    "5' UTR genomic coordinate",
    "sequence of 5' UTR",
)
MOESM8_6A_RAW_HEADER: tuple[str, ...] = MOESM8_6A_COLUMNS
MOESM8_6A_RAW_HEADER_SHA256 = (
    "c88ccfd6e7d4d5910e75092119fa35198264f0bc5de4c513604504ccc56ee042"
)

LIM6C_SHEET = "Sheet1"
LIM6C_COLUMNS: tuple[str, ...] = (
    "description",
    "barcode",
    "293T_TotalRNA_rep1",
    "293T_DNA_rep1",
    "293T_TotalRNA_rep2",
    "293T_DNA_rep2",
    "293T_TotalRNA_rep3",
    "293T_DNA_rep3",
    "Unnamed: 8",
    "description.1",
    "barcode.1",
    "293T_TotalRNA_rep1.1",
    "293T_polysome_rep1",
    "293T_TotalRNA_rep2.1",
    "293T_polysome_rep2",
    "293T_TotalRNA_rep3.1",
    "293T_polysome_rep3",
)
LIM6C_RAW_HEADER: tuple[str, ...] = (
    "description",
    "barcode",
    "293T_TotalRNA_rep1",
    "293T_DNA_rep1",
    "293T_TotalRNA_rep2",
    "293T_DNA_rep2",
    "293T_TotalRNA_rep3",
    "293T_DNA_rep3",
    "",
    "description",
    "barcode",
    "293T_TotalRNA_rep1",
    "293T_polysome_rep1",
    "293T_TotalRNA_rep2",
    "293T_polysome_rep2",
    "293T_TotalRNA_rep3",
    "293T_polysome_rep3",
)
LIM6C_RAW_HEADER_SHA256 = (
    "9cd3d648f639e02c0191659d8108f0a1e9794afc196566eb5dd8470e1400f203"
)
TRANSCRIPT_BLOCK_ID = "TRANSCRIPT_A_H"
TRANSLATION_EFFICIENCY_BLOCK_ID = "TRANSLATION_EFFICIENCY_J_Q"
TRANSCRIPT_BLOCK_COLUMNS: tuple[str, ...] = LIM6C_COLUMNS[:8]
TRANSLATION_EFFICIENCY_BLOCK_COLUMNS: tuple[str, ...] = (
    "description",
    "barcode",
    "293T_TotalRNA_rep1",
    "293T_polysome_rep1",
    "293T_TotalRNA_rep2",
    "293T_polysome_rep2",
    "293T_TotalRNA_rep3",
    "293T_polysome_rep3",
)
LIM6C_KEY_COLUMNS: tuple[str, str] = ("description", "barcode")
PUBLIC_LIM6C_BLOCK_INVENTORY: Mapping[str, int] = {
    "physical_data_rows": 179791,
    "transcript_present_key_rows": 179791,
    "translation_efficiency_present_key_rows": 152489,
    "translation_efficiency_blank_tail_rows": 27302,
    "hashed_key_intersection": 151182,
    "hashed_key_transcript_only": 28609,
    "hashed_key_translation_efficiency_only": 1307,
    "key_aligned_totalrna_exact_equal": 151182,
    "key_aligned_totalrna_any_difference": 0,
    "rowwise_key_equal_count_descriptive_only": 43,
}

DESCRIPTION_MUTANT_RE = re.compile(
    r"^(?P<gene>[^_]+)_(?P<ref>[ACGT])_(?P<alt>[ACGT])_"
    r"(?P<chrom>chr[^_]+)_(?P<start>[0-9]+)_(?P<end>[0-9]+)$"
)
DESCRIPTION_WT_RE = re.compile(
    r"^(?P<gene>[^_]+)_WT_(?P<chrom>chr[^_]+)_"
    r"(?P<start>[0-9]+)_(?P<end>[0-9]+)$"
)
DESCRIPTION_ALLOWED_CLASSES: tuple[str, ...] = (
    "INCLUDED_STRICT_SNV_MUTANT",
    "REFERENCE_ONLY_STRICT_WT",
    "REJECTED_UNADJUDICATED_DESCRIPTION",
)
COORDINATE_MUTANT_RE = re.compile(
    r"^(?P<gene>[^_]+)_(?P<chrom>chr[^_]+)_(?P<pos>[0-9]+)_"
    r"(?P<ref>[ACGT])_(?P<alt>[ACGT])_UTR5$"
)
COORDINATE_WT_RE = re.compile(
    r"^(?P<gene>[^_]+)_(?P<chrom>chr[^_]+)_(?P<pos>[0-9]+)_WT_UTR5$"
)
MOESM8_6A_ALLOWED_CLASSES: tuple[str, ...] = (
    "INCLUDED_STRICT_SNV_MUTANT_COORDINATE",
    "REFERENCE_ONLY_STRICT_WT_COORDINATE",
    "REJECTED_UNADJUDICATED_COORDINATE",
    "REJECTED_EMPTY_ROW",
)

ENDPOINTS: tuple[Mapping[str, str], ...] = (
    {
        "endpoint_id": "transcript_log2_totalrna_over_dna",
        "block_id": TRANSCRIPT_BLOCK_ID,
        "excel_columns": "A:H",
        "numerator_template": "293T_TotalRNA_rep{replicate}",
        "denominator_template": "293T_DNA_rep{replicate}",
        "definition": "TotalRNA_log2CPM - DNA_log2CPM",
    },
    {
        "endpoint_id": "te_log2_polysome_over_totalrna",
        "block_id": TRANSLATION_EFFICIENCY_BLOCK_ID,
        "excel_columns": "J:Q",
        "numerator_template": "293T_polysome_rep{replicate}",
        "denominator_template": "293T_TotalRNA_rep{replicate}",
        "definition": "polysome_log2CPM - TotalRNA_log2CPM",
    },
)

OUTPUT_FILES: tuple[str, ...] = (
    "development_companion_effect_records.jsonl",
    "replicate_effect_summaries.jsonl",
    "BARCODE_RATIO_AUDIT.jsonl",
    "DESCRIPTION_CLASSIFICATION_AUDIT.jsonl",
    "MOESM8_6A_CLASSIFICATION_AUDIT.jsonl",
    "RECONSTRUCTION_REPORT.json",
    "SHA256SUMS",
)


class ReconstructionError(RuntimeError):
    """Raised when the reconstruction cannot proceed without ambiguity."""


class ScopeViolation(ReconstructionError):
    """Raised before payload reads when a path enters a forbidden scope."""


def _reject_forbidden_path(path: Path | str, *, label: str) -> None:
    text = os.fspath(path).casefold()
    matches = sorted(token for token in FORBIDDEN_PATH_TOKENS if token in text)
    if matches:
        raise ScopeViolation(
            f"{label} rejected before read; forbidden path token(s): "
            + ",".join(matches)
        )


def _absolute_without_resolving(path: Path) -> Path:
    return Path(os.path.abspath(os.path.expanduser(os.fspath(path))))


def _reject_symlink_components(path: Path, *, label: str) -> None:
    """Reject a symlink at the leaf or in any existing path component."""

    if not path.is_absolute():
        raise ReconstructionError(f"internal error: {label} path is not absolute")
    current = Path(path.anchor)
    for part in path.parts[1:]:
        current = current / part
        try:
            info = current.lstat()
        except FileNotFoundError:
            break
        if stat.S_ISLNK(info.st_mode):
            raise ReconstructionError(f"{label} must not contain a symlink component")


def _require_regular_xlsx(path: Path, *, label: str) -> os.stat_result:
    _reject_forbidden_path(path, label=label)
    _reject_symlink_components(path, label=label)
    try:
        info = path.lstat()
    except FileNotFoundError as exc:
        raise ReconstructionError(f"{label} is missing") from exc
    if not stat.S_ISREG(info.st_mode):
        raise ReconstructionError(f"{label} must be a regular file")
    if path.suffix.casefold() != ".xlsx":
        raise ReconstructionError(f"{label} must be an .xlsx workbook")
    return info


def _require_regular_json(path: Path, *, label: str) -> os.stat_result:
    _reject_forbidden_path(path, label=label)
    _reject_symlink_components(path, label=label)
    try:
        info = path.lstat()
    except FileNotFoundError as exc:
        raise ReconstructionError(f"{label} is missing") from exc
    if not stat.S_ISREG(info.st_mode):
        raise ReconstructionError(f"{label} must be a regular file")
    if path.suffix.casefold() != ".json":
        raise ReconstructionError(f"{label} must be a .json document")
    return info


def _require_output_target(path: Path) -> None:
    _reject_forbidden_path(path, label="output directory")
    _reject_symlink_components(path, label="output directory")
    if path.exists() or path.is_symlink():
        raise ReconstructionError("refusing to overwrite existing output directory")
    parent = path.parent
    _reject_symlink_components(parent, label="output parent directory")
    try:
        info = parent.lstat()
    except FileNotFoundError as exc:
        raise ReconstructionError("output parent directory is missing") from exc
    if not stat.S_ISDIR(info.st_mode):
        raise ReconstructionError("output parent must be a directory")


def _prepare_paths_before_read(
    moesm8_path: Path,
    lim6c_path: Path,
    protocol_path: Path,
    output_directory: Path,
) -> tuple[
    Path,
    Path,
    Path,
    Path,
    os.stat_result,
    os.stat_result,
    os.stat_result,
]:
    raw = (
        (Path(moesm8_path), "MOESM8 workbook"),
        (Path(lim6c_path), "Lim 6c workbook"),
        (Path(protocol_path), "reconstruction protocol"),
        (Path(output_directory), "output directory"),
    )

    # Scope-check all caller strings before any filesystem operation.
    for path, label in raw:
        _reject_forbidden_path(path, label=label)

    moesm8, lim6c, protocol, output = tuple(
        _absolute_without_resolving(path) for path, _ in raw
    )
    for path, (_, label) in zip((moesm8, lim6c, protocol, output), raw):
        _reject_forbidden_path(path, label=label)

    moesm8_info = _require_regular_xlsx(moesm8, label="MOESM8 workbook")
    lim6c_info = _require_regular_xlsx(lim6c, label="Lim 6c workbook")
    if moesm8.name != EXPECTED_MOESM8_FILENAME:
        raise ReconstructionError(
            "MOESM8 workbook basename mismatch; "
            f"expected={EXPECTED_MOESM8_FILENAME!r}, actual={moesm8.name!r}"
        )
    if lim6c.name != EXPECTED_LIM6C_FILENAME:
        raise ReconstructionError(
            "Lim 6c workbook basename mismatch; "
            f"expected={EXPECTED_LIM6C_FILENAME!r}, actual={lim6c.name!r}"
        )
    protocol_info = _require_regular_json(
        protocol, label="reconstruction protocol"
    )
    if (moesm8_info.st_dev, moesm8_info.st_ino) == (
        lim6c_info.st_dev,
        lim6c_info.st_ino,
    ):
        raise ReconstructionError("the two workbook inputs must be distinct files")
    if (protocol_info.st_dev, protocol_info.st_ino) in {
        (moesm8_info.st_dev, moesm8_info.st_ino),
        (lim6c_info.st_dev, lim6c_info.st_ino),
    }:
        raise ReconstructionError("protocol and workbook inputs must be distinct files")
    _require_output_target(output)
    return (
        moesm8,
        lim6c,
        protocol,
        output,
        moesm8_info,
        lim6c_info,
        protocol_info,
    )


def _normalize_expected_sha256(value: str, *, label: str) -> str:
    if not isinstance(value, str):
        raise ReconstructionError(f"{label} expected SHA-256 must be a string")
    normalized = value.casefold()
    if not SHA256_RE.fullmatch(normalized):
        raise ReconstructionError(f"{label} expected SHA-256 must be 64 hex digits")
    return normalized


def _sha256_and_size(path: Path) -> tuple[str, int]:
    digest = hashlib.sha256()
    size = 0
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1 << 20), b""):
            digest.update(block)
            size += len(block)
    return digest.hexdigest(), size


def _identity(info: os.stat_result) -> tuple[int, int, int, int]:
    return (info.st_dev, info.st_ino, info.st_size, info.st_mtime_ns)


def _verify_input_hash(
    path: Path,
    expected_sha256: str,
    initial_info: os.stat_result,
    *,
    label: str,
) -> dict[str, Any]:
    observed, observed_size = _sha256_and_size(path)
    after = _require_regular_xlsx(path, label=label)
    if _identity(initial_info) != _identity(after) or observed_size != after.st_size:
        raise ReconstructionError(f"{label} changed while its SHA-256 was computed")
    if observed != expected_sha256:
        raise ReconstructionError(f"{label} SHA-256 mismatch")
    return {
        "sha256": observed,
        "bytes": observed_size,
        "identity": _identity(after),
        "filename": path.name,
    }


def _verify_protocol_hash(
    path: Path,
    expected_sha256: str,
    initial_info: os.stat_result,
) -> dict[str, Any]:
    observed, observed_size = _sha256_and_size(path)
    after = _require_regular_json(path, label="reconstruction protocol")
    if _identity(initial_info) != _identity(after) or observed_size != after.st_size:
        raise ReconstructionError(
            "reconstruction protocol changed while its SHA-256 was computed"
        )
    if observed != expected_sha256:
        raise ReconstructionError("reconstruction protocol SHA-256 mismatch")
    return {
        "sha256": observed,
        "bytes": observed_size,
        "identity": _identity(after),
        "filename": path.name,
    }


def _protocol_value(document: Mapping[str, Any], path: str) -> Any:
    value: Any = document
    for component in path.split("."):
        if not isinstance(value, Mapping) or component not in value:
            raise ReconstructionError(f"protocol is missing required field: {path}")
        value = value[component]
    return value


def _load_and_validate_protocol(
    path: Path,
    *,
    expected_moesm8_sha256: str,
    expected_lim6c_sha256: str,
) -> dict[str, Any]:
    try:
        document = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ReconstructionError("reconstruction protocol is not valid UTF-8 JSON") from exc
    if not isinstance(document, dict):
        raise ReconstructionError("reconstruction protocol root must be an object")

    expected_values: tuple[tuple[str, Any], ...] = (
        ("contract_id", CONTRACT_ID),
        ("schema_version", "3.0.0"),
        ("protocol_id", PROTOCOL_ID),
        ("protocol_status", PROTOCOL_STATUS),
        ("dataset_id", DATASET_ID),
        ("study_group_id", "PLUMAGE_LIM_2021"),
        ("independent_study_count", 1),
        ("scope.ordinary_public_data_only", True),
        ("scope.included_contexts", [CONTEXT]),
        ("scope.included_endpoints", ["TRANSCRIPT", "TRANSLATION_EFFICIENCY"]),
        ("scope.included_biological_replicates", [1, 2, 3]),
        ("scope.pc3_included", False),
        ("scope.eighteen_geo_raw_count_tables_joined", False),
        ("scope.published_significant_only_tables_used_for_membership", False),
        ("scope.training_allowed", False),
        ("scope.model_selection_allowed", False),
        ("scope.qualification_allowed", False),
        ("inputs.moesm8.filename", EXPECTED_MOESM8_FILENAME),
        ("inputs.moesm8.sha256", expected_moesm8_sha256),
        ("inputs.moesm8.sheet", MOESM8_SEQUENCE_SHEET),
        ("inputs.moesm8.header_sha256", MOESM8_6A_RAW_HEADER_SHA256),
        ("inputs.lim6c_293t.filename", EXPECTED_LIM6C_FILENAME),
        ("inputs.lim6c_293t.sha256", expected_lim6c_sha256),
        ("inputs.lim6c_293t.sheet", LIM6C_SHEET),
        ("inputs.lim6c_293t.header_sha256", LIM6C_RAW_HEADER_SHA256),
        (
            "inputs.header_canonicalization.version",
            "XLSX_DENSE_HEADER_NFC_LF_COMPACT_JSON_V1",
        ),
        ("inputs.header_canonicalization.empty_physical_header_cell", ""),
        ("inputs.header_canonicalization.unicode_normalization", "NFC"),
        (
            "inputs.header_canonicalization.newline_normalization",
            "CRLF_OR_CR_TO_LF",
        ),
        ("inputs.header_canonicalization.json_ensure_ascii", False),
        ("inputs.header_canonicalization.json_separators", [",", ":"]),
        (
            "paper_faithful_measurement_transform.scope",
            "WITHIN_SAMPLE_BARCODE_MEASUREMENT_ONLY",
        ),
        (
            "paper_faithful_measurement_transform.input_value_scale",
            "PUBLISHED_LOG2_CPM_PER_BARCODE",
        ),
        (
            "paper_faithful_measurement_transform.published_source_axis",
            "SUPPLEMENTARY_FIGURE_8B_LOG2_CPM_PER_BARCODE",
        ),
        (
            "paper_faithful_measurement_transform.original_cpm_minimum_inclusive",
            ORIGINAL_CPM_MINIMUM_INCLUSIVE,
        ),
        (
            "paper_faithful_measurement_transform.published_log2_cpm_minimum_inclusive",
            PUBLISHED_LOG2_CPM_MINIMUM_INCLUSIVE,
        ),
        ("paper_faithful_measurement_transform.pseudocount", None),
        (
            "paper_faithful_measurement_transform.transcript_ratio",
            "TOTALRNA_LOG2_CPM_MINUS_DNA_LOG2_CPM",
        ),
        (
            "paper_faithful_measurement_transform.translation_efficiency_ratio",
            "POLYSOME_LOG2_CPM_MINUS_TOTALRNA_LOG2_CPM",
        ),
        (
            "paper_faithful_measurement_transform.both_ratio_components_must_meet_threshold",
            True,
        ),
        (
            "paper_faithful_measurement_transform.component_filter",
            "BOTH_PUBLISHED_LOG2_CPM_COMPONENTS_GTE_NEGATIVE_ONE",
        ),
        (
            "paper_faithful_measurement_transform.values_below_published_log2_cpm_minimum_are_allowed",
            False,
        ),
        ("paper_faithful_measurement_transform.clipping_allowed", False),
        ("paper_faithful_measurement_transform.missing_is_zero", False),
        ("route_a_companion_summary.classification", "ROUTE_A_COMPANION_NOT_PAPER_TEST"),
        ("route_a_companion_summary.barcode_role", "TECHNICAL_MEASUREMENT_UNIT_ONLY"),
        ("route_a_companion_summary.biological_replicate_role", "INFERENCE_UNIT"),
        (
            "route_a_companion_summary.within_arm_replicate_aggregation",
            "MEDIAN_OF_ELIGIBLE_BARCODE_LOG2_RATIOS",
        ),
        (
            "route_a_companion_summary.within_replicate_effect",
            "MUTANT_MEDIAN_MINUS_WT_MEDIAN",
        ),
        (
            "route_a_companion_summary.across_replicate_effect",
            "EQUAL_WEIGHT_MEAN_OF_THREE_REPLICATE_EFFECTS",
        ),
        (
            "route_a_companion_summary.standard_error",
            "SAMPLE_SD_OF_THREE_REPLICATE_EFFECTS_DIVIDED_BY_SQRT_3",
        ),
        ("route_a_companion_summary.required_biological_replicates", 3),
        (
            "route_a_companion_summary.minimum_eligible_barcodes_per_arm_per_endpoint_per_replicate",
            PRIMARY_MIN_ELIGIBLE_BARCODES_PER_ARM_REPLICATE,
        ),
        (
            "route_a_companion_summary.audit_only_sensitivity_floors",
            list(SECONDARY_AUDIT_ONLY_BARCODE_FLOORS),
        ),
        (
            "route_a_companion_summary.coverage_floor_purpose",
            "TECHNICAL_MEDIAN_STABILITY_NOT_EFFECTIVE_N_OR_POWER",
        ),
        ("route_a_companion_summary.sensitivity_may_change_primary_eligibility", False),
        ("route_a_companion_summary.barcode_count_may_weight_replicates", False),
        ("route_a_companion_summary.barcode_count_may_increase_effective_n", False),
        ("route_a_companion_summary.barcode_count_may_define_standard_error", False),
        ("companion_lineage.barcode_long_table_required", True),
        ("companion_lineage.raw_barcode_identifier_output_allowed", False),
        ("companion_lineage.stable_barcode_sha256_required", True),
        ("companion_lineage.eligible_and_excluded_observations_both_retained", True),
        ("companion_lineage.exact_exclusion_reason_required", True),
        ("companion_lineage.raw_row_locator_required", True),
        ("companion_lineage.replicate_summary_required", True),
        ("companion_lineage.development_companion_effect_summary_required", True),
        ("companion_lineage.canonical_intervention_record_v3_materialized", False),
        ("companion_lineage.all_outputs_sha256_bound", True),
        (
            "description_universe.included_mutant_description_regex",
            DESCRIPTION_MUTANT_RE.pattern,
        ),
        (
            "description_universe.reference_wt_description_regex",
            DESCRIPTION_WT_RE.pattern,
        ),
        (
            "description_universe.allowed_classes",
            list(DESCRIPTION_ALLOWED_CLASSES),
        ),
        ("description_universe.all_unique_descriptions_must_be_classified", True),
        (
            "description_universe.unadjudicated_class_action",
            "REJECT_WITH_REASON_AND_RETAIN_BLOCKER",
        ),
        (
            "description_universe.unadjudicated_blocker",
            "UNADJUDICATED_DESCRIPTION_CLASSES_EXCLUDED",
        ),
        (
            "description_universe.successful_join_may_shrink_unique_description_denominator",
            False,
        ),
        (
            "moesm8_6a_universe.allowed_classes",
            list(MOESM8_6A_ALLOWED_CLASSES),
        ),
        (
            "moesm8_6a_universe.all_physical_data_rows_must_be_classified",
            True,
        ),
        (
            "moesm8_6a_universe.unadjudicated_coordinate_action",
            "REJECT_WITH_ROW_LOCATOR_HASH_AND_RETAIN_BLOCKER",
        ),
        (
            "moesm8_6a_universe.unadjudicated_blocker",
            "UNADJUDICATED_6A_COORDINATE_CLASSES_EXCLUDED",
        ),
        (
            "moesm8_6a_universe.successful_join_may_shrink_physical_row_denominator",
            False,
        ),
        ("lim6c_block_semantics.physical_data_rows", 179791),
        (
            "lim6c_block_semantics.transcript_block.excel_columns",
            "A:H",
        ),
        (
            "lim6c_block_semantics.transcript_block.key_columns",
            list(LIM6C_KEY_COLUMNS),
        ),
        (
            "lim6c_block_semantics.transcript_block.present_key_rows",
            179791,
        ),
        (
            "lim6c_block_semantics.translation_efficiency_block.excel_columns",
            "J:Q",
        ),
        (
            "lim6c_block_semantics.translation_efficiency_block.key_columns",
            list(LIM6C_KEY_COLUMNS),
        ),
        (
            "lim6c_block_semantics.translation_efficiency_block.present_key_rows",
            152489,
        ),
        (
            "lim6c_block_semantics.translation_efficiency_block.blank_tail_rows",
            27302,
        ),
        ("lim6c_block_semantics.join_key", list(LIM6C_KEY_COLUMNS)),
        ("lim6c_block_semantics.blocks_are_row_aligned", False),
        ("lim6c_block_semantics.row_position_may_join_blocks", False),
        ("lim6c_block_semantics.outer_join_for_inventory_only", True),
        ("lim6c_block_semantics.endpoint_uses_only_its_own_block", True),
        (
            "lim6c_block_semantics.duplicate_key_rows_allowed_within_block",
            False,
        ),
        ("lim6c_block_semantics.hashed_key_intersection", 151182),
        ("lim6c_block_semantics.hashed_key_transcript_only", 28609),
        (
            "lim6c_block_semantics.hashed_key_translation_efficiency_only",
            1307,
        ),
        (
            "lim6c_block_semantics.key_aligned_totalrna_all_three_replicates_exact_equal",
            151182,
        ),
        (
            "lim6c_block_semantics.key_aligned_totalrna_any_difference",
            0,
        ),
        (
            "lim6c_block_semantics.missing_key_in_endpoint_block_is_zero",
            False,
        ),
        ("canonical_boundaries.evidence_status", "BLOCKED_PENDING_PUBLIC_EVIDENCE"),
        ("canonical_boundaries.claim_status", "NOT_ESTABLISHED"),
        ("canonical_boundaries.data_role", "EXCLUDED"),
        ("canonical_boundaries.split_partition", "EXCLUDED"),
        ("canonical_boundaries.license_status", "UNKNOWN_BLOCKED"),
        ("canonical_boundaries.checkpoint_exposure_status", "AUDIT_PENDING"),
        ("canonical_boundaries.beneficial_direction", "HIGHER_IS_BETTER"),
        (
            "canonical_boundaries.endpoint_or_context_increases_independent_study_count",
            False,
        ),
        (
            "canonical_boundaries.engineering_reconstruction_establishes_study_qualification",
            False,
        ),
        ("model_results_may_change_this_protocol", False),
    )
    for field_path, expected in expected_values:
        observed = _protocol_value(document, field_path)
        if observed != expected or type(observed) is not type(expected):
            raise ReconstructionError(
                f"protocol field mismatch: {field_path}"
            )
    if document.get("unresolved_blockers") != list(PROTOCOL_UNRESOLVED_BLOCKERS):
        raise ReconstructionError("protocol unresolved_blockers mismatch")
    return document


def _require_exact_columns(
    frame: pd.DataFrame,
    expected: Sequence[str],
    *,
    label: str,
) -> None:
    actual = tuple(str(column) for column in frame.columns)
    if actual != tuple(expected):
        missing = sorted(set(expected) - set(actual))
        unexpected = sorted(set(actual) - set(expected))
        raise ReconstructionError(
            f"{label} exact header mismatch; missing={missing}, unexpected={unexpected}"
        )


def _canonical_header_value(value: Any) -> str:
    if value is None:
        return ""
    if not isinstance(value, str):
        raise ReconstructionError("workbook raw header contains a non-string value")
    normalized_lines = value.replace("\r\n", "\n").replace("\r", "\n")
    return unicodedata.normalize("NFC", normalized_lines)


def _canonical_header(columns: Sequence[Any]) -> tuple[str, ...]:
    return tuple(_canonical_header_value(value) for value in columns)


def _header_sha256(columns: Sequence[Any]) -> str:
    canonical = _canonical_header(columns)
    payload = json.dumps(
        list(canonical), ensure_ascii=False, separators=(",", ":")
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def _load_exact_workbook_tables(
    moesm8_path: Path,
    lim6c_path: Path,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    try:
        moesm8_workbook = load_workbook(
            moesm8_path, read_only=True, data_only=True
        )
        try:
            if tuple(moesm8_workbook.sheetnames) != MOESM8_SHEETS:
                raise ReconstructionError("MOESM8 exact sheet set/order mismatch")
            moesm8_sheet = moesm8_workbook[MOESM8_SEQUENCE_SHEET]
            moesm8_raw_header_physical = tuple(
                cell.value
                for cell in next(
                    moesm8_sheet.iter_rows(
                        min_row=1,
                        max_row=1,
                        min_col=1,
                        max_col=moesm8_sheet.max_column,
                    )
                )
            )
            moesm8_raw_header = _canonical_header(moesm8_raw_header_physical)
            moesm8_data_rows = (
                [
                    tuple(cell.value for cell in row)
                    for row in moesm8_sheet.iter_rows(
                        min_row=2,
                        max_row=moesm8_sheet.max_row,
                        min_col=1,
                        max_col=len(MOESM8_6A_COLUMNS),
                    )
                ]
                if moesm8_sheet.max_row >= 2
                else []
            )
        finally:
            moesm8_workbook.close()

        # Keep formulas visible so they can be rejected rather than silently
        # consuming an implementation-dependent cached value.
        lim6c_workbook = load_workbook(
            lim6c_path, read_only=True, data_only=False
        )
        try:
            if tuple(lim6c_workbook.sheetnames) != (LIM6C_SHEET,):
                raise ReconstructionError("Lim 6c exact sheet set/order mismatch")
            lim6c_sheet = lim6c_workbook[LIM6C_SHEET]
            lim6c_raw_header_physical = tuple(
                cell.value
                for cell in next(
                    lim6c_sheet.iter_rows(
                        min_row=1,
                        max_row=1,
                        min_col=1,
                        max_col=lim6c_sheet.max_column,
                    )
                )
            )
            lim6c_raw_header = _canonical_header(lim6c_raw_header_physical)
            lim6c_physical_rows = [
                (physical_row_number, tuple(row_values))
                for physical_row_number, row_values in enumerate(
                    lim6c_sheet.iter_rows(
                        min_row=2,
                        max_row=lim6c_sheet.max_row,
                        min_col=1,
                        max_col=len(LIM6C_RAW_HEADER),
                        values_only=True,
                    ),
                    start=2,
                )
            ]
        finally:
            lim6c_workbook.close()

        if moesm8_raw_header != MOESM8_6A_RAW_HEADER:
            raise ReconstructionError("MOESM8 6a exact raw header mismatch")
        if lim6c_raw_header != LIM6C_RAW_HEADER:
            raise ReconstructionError("Lim 6c exact raw header mismatch")
        moesm8_header_sha256 = _header_sha256(moesm8_raw_header)
        lim6c_header_sha256 = _header_sha256(lim6c_raw_header)
        if moesm8_header_sha256 != MOESM8_6A_RAW_HEADER_SHA256:
            raise ReconstructionError("MOESM8 6a ordered-header SHA-256 mismatch")
        if lim6c_header_sha256 != LIM6C_RAW_HEADER_SHA256:
            raise ReconstructionError("Lim 6c ordered-header SHA-256 mismatch")

        # Preserve every physical row by frozen Excel position.  Lim 6c's A:H
        # and J:Q blocks are intentionally parsed independently; no pandas
        # duplicate-header mangling or row-position join is allowed.
        table_6a = pd.DataFrame(
            moesm8_data_rows,
            columns=list(MOESM8_6A_COLUMNS),
        )
        (
            transcript_block,
            translation_efficiency_block,
            lim6c_block_structure,
        ) = _parse_and_validate_lim6c_physical_blocks(
            lim6c_physical_rows
        )
    except ReconstructionError:
        raise
    except Exception as exc:
        raise ReconstructionError("workbook parse failed closed") from exc

    _require_exact_columns(table_6a, MOESM8_6A_COLUMNS, label="MOESM8 6a")
    _require_exact_columns(
        transcript_block,
        (*TRANSCRIPT_BLOCK_COLUMNS, "_physical_row_number", "_block_id"),
        label="Lim 6c transcript block",
    )
    _require_exact_columns(
        translation_efficiency_block,
        (
            *TRANSLATION_EFFICIENCY_BLOCK_COLUMNS,
            "_physical_row_number",
            "_block_id",
        ),
        label="Lim 6c translation-efficiency block",
    )
    if table_6a.empty:
        raise ReconstructionError("MOESM8 6a must contain at least one data row")
    if transcript_block.empty:
        raise ReconstructionError(
            "Lim 6c transcript block must contain at least one data row"
        )

    structure = {
        "moesm8_sheet_count": len(MOESM8_SHEETS),
        "moesm8_6a_row_count": int(len(table_6a)),
        "moesm8_6a_column_count": len(MOESM8_6A_COLUMNS),
        "moesm8_6a_header_sha256": moesm8_header_sha256,
        "lim6c_sheet_count": 1,
        "lim6c_row_count": int(len(lim6c_physical_rows)),
        "lim6c_column_count": len(LIM6C_COLUMNS),
        "lim6c_header_sha256": lim6c_header_sha256,
        "ordered_header_hash_algorithm": (
            "sha256(compact_json_utf8(dense_A_to_max_header_vector;"
            "empty='';unicode=NFC;newlines=LF))"
        ),
        "ordered_header_hash_version": "XLSX_DENSE_HEADER_NFC_LF_COMPACT_JSON_V1",
        "lim6c_block_semantics": lim6c_block_structure,
    }
    return table_6a, transcript_block, translation_efficiency_block, structure


def _normalize_sequence(value: Any, *, label: str) -> str:
    if not isinstance(value, str):
        raise ReconstructionError(f"{label} must be a sequence string")
    sequence = value.strip().upper().replace("U", "T")
    if not sequence or any(base not in "ACGT" for base in sequence):
        raise ReconstructionError(f"{label} must contain only A/C/G/T/U")
    return sequence


def _exact_positive_integer(value: Any, *, label: str) -> int:
    if isinstance(value, bool):
        raise ReconstructionError(f"{label} must be a positive integer")
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ReconstructionError(f"{label} must be a positive integer") from exc
    if not math.isfinite(number) or number <= 0 or not number.is_integer():
        raise ReconstructionError(f"{label} must be a positive integer")
    return int(number)


def _is_blank_cell(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def _coordinate_value_type(value: Any) -> str:
    if _is_blank_cell(value):
        return "EMPTY"
    if isinstance(value, str):
        return "STRING"
    if isinstance(value, bool):
        return "BOOLEAN"
    if isinstance(value, int):
        return "INTEGER"
    if isinstance(value, float):
        return "FLOAT"
    return type(value).__name__.upper()


def _coordinate_hash_material(value: Any) -> bytes:
    if _is_blank_cell(value):
        return b""
    if isinstance(value, str):
        return value.encode("utf-8")
    if isinstance(value, float) and not math.isfinite(value):
        return repr(value).encode("ascii")
    if isinstance(value, (bool, int, float)):
        return json.dumps(
            value,
            allow_nan=False,
            ensure_ascii=True,
            separators=(",", ":"),
        ).encode("ascii")
    if hasattr(value, "isoformat"):
        return str(value.isoformat()).encode("utf-8")
    return str(value).encode("utf-8")


def _moesm8_6a_coordinate_sha256(
    value: Any,
    *,
    physical_row_number: int,
    moesm8_sha256: str,
) -> str:
    prefix = (
        f"{moesm8_sha256}\0{MOESM8_SEQUENCE_SHEET}\0"
        f"physical_row={physical_row_number}\0"
        f"coordinate_type={_coordinate_value_type(value)}\0"
    ).encode("utf-8")
    return hashlib.sha256(prefix + _coordinate_hash_material(value)).hexdigest()


def _moesm8_6a_row_locator_sha256(
    physical_row_number: int,
    *,
    moesm8_sha256: str,
) -> str:
    locator = (
        f"{moesm8_sha256}\0{MOESM8_SEQUENCE_SHEET}\0"
        f"physical_row={physical_row_number}"
    ).encode("utf-8")
    return hashlib.sha256(locator).hexdigest()


def _unparsed_coordinate_signature(value: Any) -> str:
    value_type = _coordinate_value_type(value)
    if value_type == "EMPTY":
        return "EMPTY_COORDINATE_CELL"
    if not isinstance(value, str):
        return f"NONSTRICT_{value_type}_COORDINATE"
    field_count = value.count("_") + 1
    utr5_suffix = "YES" if value.endswith("_UTR5") else "NO"
    return (
        f"NONSTRICT_STRING_FIELDS_{field_count}_"
        f"EXACT_UTR5_SUFFIX_{utr5_suffix}"
    )


def _parse_6a_table(
    table: pd.DataFrame,
    *,
    moesm8_sha256: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    entries: list[dict[str, Any]] = []
    classification_rows: list[dict[str, Any]] = []
    for row_number, row in enumerate(table.itertuples(index=False, name=None), start=2):
        length_value, gene_value, coordinate_value, sequence_value = row
        row_is_empty = all(_is_blank_cell(value) for value in row)
        coordinate = coordinate_value if isinstance(coordinate_value, str) else None
        mutant_match = (
            COORDINATE_MUTANT_RE.fullmatch(coordinate)
            if coordinate is not None
            else None
        )
        wt_match = (
            COORDINATE_WT_RE.fullmatch(coordinate)
            if coordinate is not None
            else None
        )

        if row_is_empty:
            classification = "REJECTED_EMPTY_ROW"
            signature = "EMPTY_PHYSICAL_DATA_ROW"
            reason = "EMPTY_PHYSICAL_DATA_ROW"
        elif mutant_match is not None:
            classification = "INCLUDED_STRICT_SNV_MUTANT_COORDINATE"
            signature = "STRICT_SNV_MUTANT_COORDINATE_REGEX_FULLMATCH"
            reason = "STRICT_SNV_MUTANT_COORDINATE_INCLUDED"
        elif wt_match is not None:
            classification = "REFERENCE_ONLY_STRICT_WT_COORDINATE"
            signature = "STRICT_WT_COORDINATE_REGEX_FULLMATCH"
            reason = "STRICT_WT_COORDINATE_REFERENCE_ONLY"
        else:
            classification = "REJECTED_UNADJUDICATED_COORDINATE"
            signature = _unparsed_coordinate_signature(coordinate_value)
            reason = (
                "EMPTY_COORDINATE_IN_NONEMPTY_ROW"
                if _is_blank_cell(coordinate_value)
                else "COORDINATE_NOT_STRICT_MUTANT_OR_WT_REGEX"
                if isinstance(coordinate_value, str)
                else "NON_STRING_COORDINATE"
            )

        classification_rows.append(
            {
                "schema_id": "route_a_v3.moesm8_6a_classification_audit",
                "schema_version": SCHEMA_VERSION,
                "dataset_id": DATASET_ID,
                "study_group_id": STUDY_GROUP_ID,
                "independent_study_id": INDEPENDENT_STUDY_ID,
                "independent_study_count": 1,
                "source_asset_id": MOESM8_SOURCE_ASSET_ID,
                "sheet_name": MOESM8_SEQUENCE_SHEET,
                "physical_row_number": row_number,
                "row_locator_sha256": _moesm8_6a_row_locator_sha256(
                    row_number,
                    moesm8_sha256=moesm8_sha256,
                ),
                "coordinate_sha256": _moesm8_6a_coordinate_sha256(
                    coordinate_value,
                    physical_row_number=row_number,
                    moesm8_sha256=moesm8_sha256,
                ),
                "coordinate_value_type": _coordinate_value_type(coordinate_value),
                "coordinate_signature": signature,
                "classification": classification,
                "classification_reason": reason,
                "raw_coordinate_emitted": False,
                "raw_sequence_emitted": False,
            }
        )

        if mutant_match is None and wt_match is None:
            continue

        if (
            not isinstance(gene_value, str)
            or not gene_value
            or gene_value.strip() != gene_value
        ):
            raise ReconstructionError("MOESM8 6a contains an invalid Gene name")
        match = mutant_match or wt_match
        assert match is not None
        if gene_value != match.group("gene"):
            raise ReconstructionError("MOESM8 6a Gene name/coordinate mismatch")

        sequence = _normalize_sequence(
            sequence_value, label=f"MOESM8 6a sequence row {row_number}"
        )
        declared_length = _exact_positive_integer(
            length_value, label="MOESM8 6a declared sequence length"
        )
        if declared_length != len(sequence):
            raise ReconstructionError("MOESM8 6a declared/full sequence length mismatch")

        entry: dict[str, Any] = {
            "type": "mutant" if mutant_match is not None else "wt",
            "gene": match.group("gene"),
            "chrom": match.group("chrom"),
            "position": int(match.group("pos")),
            "sequence": sequence,
            "coordinate": coordinate,
            "physical_row_number": row_number,
        }
        if mutant_match is not None:
            ref = mutant_match.group("ref")
            alt = mutant_match.group("alt")
            if ref == alt:
                raise ReconstructionError("MOESM8 6a mutant coordinate has ref == alt")
            entry.update({"ref": ref, "alt": alt})
        entries.append(entry)
    return entries, classification_rows


def _strict_text_column(
    table: pd.DataFrame,
    column: str,
    *,
    label: str = "Lim 6c",
) -> None:
    for value in table[column].tolist():
        if not isinstance(value, str) or not value or value.strip() != value:
            raise ReconstructionError(f"{label} {column} contains an invalid value")


def _numeric_log2_cpm_column(
    table: pd.DataFrame,
    column: str,
    *,
    label: str = "Lim 6c",
) -> pd.Series:
    original = table[column]
    invalid_formula = original.map(
        lambda value: isinstance(value, str) and value.startswith("=")
    )
    if bool(invalid_formula.any()):
        raise ReconstructionError(
            f"{label} {column} contains a forbidden formula cell"
        )
    invalid_physical_type = original.map(
        lambda value: not _is_blank_cell(value)
        and (isinstance(value, bool) or not isinstance(value, Real))
    )
    if bool(invalid_physical_type.any()):
        raise ReconstructionError(f"{label} {column} contains non-numeric log2CPM")
    converted = pd.to_numeric(original, errors="coerce")
    invalid_nonblank = original.notna() & converted.isna()
    if bool(invalid_nonblank.any()):
        raise ReconstructionError(f"{label} {column} contains non-numeric log2CPM")
    if bool(((converted == float("inf")) | (converted == float("-inf"))).any()):
        raise ReconstructionError(f"{label} {column} contains non-finite log2CPM")
    if bool(
        (converted.dropna() < PUBLISHED_LOG2_CPM_MINIMUM_INCLUSIVE).any()
    ):
        raise ReconstructionError(
            f"{label} {column} contains log2CPM below the inclusive -1 minimum"
        )
    return converted.astype(float)


def _lim6c_key_sha256(description: str, barcode: str) -> str:
    payload = json.dumps(
        [description, barcode],
        ensure_ascii=False,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def _validate_lim6c_endpoint_block(
    frame: pd.DataFrame,
    *,
    block_id: str,
    value_columns: Sequence[str],
) -> pd.DataFrame:
    normalized = frame.copy()
    label = f"Lim 6c {block_id}"
    for column in LIM6C_KEY_COLUMNS:
        _strict_text_column(normalized, column, label=label)
    if bool(normalized.duplicated(subset=list(LIM6C_KEY_COLUMNS)).any()):
        raise ReconstructionError(
            f"{label} has duplicate description/barcode keys"
        )
    for column in value_columns:
        normalized[column] = _numeric_log2_cpm_column(
            normalized,
            column,
            label=label,
        )
    return normalized


def _optional_numeric_exact_equal(left: Any, right: Any) -> bool:
    left_missing = bool(pd.isna(left))
    right_missing = bool(pd.isna(right))
    return (left_missing and right_missing) or (
        not left_missing and not right_missing and left == right
    )


def _parse_and_validate_lim6c_physical_blocks(
    physical_rows: Sequence[tuple[int, Sequence[Any]]],
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    transcript_rows: list[tuple[Any, ...]] = []
    translation_efficiency_rows: list[tuple[Any, ...]] = []
    right_blank_tail_started = False
    right_blank_tail_rows = 0

    for physical_row_number, raw_row in physical_rows:
        values = tuple(raw_row)
        if len(values) != len(LIM6C_RAW_HEADER):
            raise ReconstructionError("Lim 6c physical row width mismatch")
        if not _is_blank_cell(values[8]):
            raise ReconstructionError("Lim 6c separator column I must be blank")

        transcript_values = values[:8]
        translation_efficiency_values = values[9:17]
        transcript_rows.append(
            (
                *transcript_values,
                physical_row_number,
                TRANSCRIPT_BLOCK_ID,
            )
        )

        right_is_blank = all(
            _is_blank_cell(value) for value in translation_efficiency_values
        )
        if right_is_blank:
            right_blank_tail_started = True
            right_blank_tail_rows += 1
            continue
        if right_blank_tail_started:
            raise ReconstructionError(
                "Lim 6c blank translation-efficiency rows must form a contiguous tail"
            )
        translation_efficiency_rows.append(
            (
                *translation_efficiency_values,
                physical_row_number,
                TRANSLATION_EFFICIENCY_BLOCK_ID,
            )
        )

    transcript_block = pd.DataFrame(
        transcript_rows,
        columns=[
            *TRANSCRIPT_BLOCK_COLUMNS,
            "_physical_row_number",
            "_block_id",
        ],
    )
    translation_efficiency_block = pd.DataFrame(
        translation_efficiency_rows,
        columns=[
            *TRANSLATION_EFFICIENCY_BLOCK_COLUMNS,
            "_physical_row_number",
            "_block_id",
        ],
    )
    transcript_block = _validate_lim6c_endpoint_block(
        transcript_block,
        block_id=TRANSCRIPT_BLOCK_ID,
        value_columns=TRANSCRIPT_BLOCK_COLUMNS[2:],
    )
    translation_efficiency_block = _validate_lim6c_endpoint_block(
        translation_efficiency_block,
        block_id=TRANSLATION_EFFICIENCY_BLOCK_ID,
        value_columns=TRANSLATION_EFFICIENCY_BLOCK_COLUMNS[2:],
    )

    transcript_hash_to_key: dict[str, tuple[str, str]] = {}
    for description, barcode in transcript_block.loc[
        :, list(LIM6C_KEY_COLUMNS)
    ].itertuples(index=False, name=None):
        key = (str(description), str(barcode))
        key_hash = _lim6c_key_sha256(*key)
        prior = transcript_hash_to_key.setdefault(key_hash, key)
        if prior != key:
            raise ReconstructionError("Lim 6c hashed join-key collision")

    translation_hash_to_key: dict[str, tuple[str, str]] = {}
    for description, barcode in translation_efficiency_block.loc[
        :, list(LIM6C_KEY_COLUMNS)
    ].itertuples(index=False, name=None):
        key = (str(description), str(barcode))
        key_hash = _lim6c_key_sha256(*key)
        prior = translation_hash_to_key.setdefault(key_hash, key)
        if prior != key:
            raise ReconstructionError("Lim 6c hashed join-key collision")

    transcript_hashes = set(transcript_hash_to_key)
    translation_hashes = set(translation_hash_to_key)
    intersection_hashes = transcript_hashes & translation_hashes

    totalrna_columns = [
        "293T_TotalRNA_rep1",
        "293T_TotalRNA_rep2",
        "293T_TotalRNA_rep3",
    ]
    transcript_totalrna = {
        _lim6c_key_sha256(str(description), str(barcode)): tuple(values)
        for description, barcode, *values in transcript_block.loc[
            :, [*LIM6C_KEY_COLUMNS, *totalrna_columns]
        ].itertuples(index=False, name=None)
    }
    translation_totalrna = {
        _lim6c_key_sha256(str(description), str(barcode)): tuple(values)
        for description, barcode, *values in translation_efficiency_block.loc[
            :, [*LIM6C_KEY_COLUMNS, *totalrna_columns]
        ].itertuples(index=False, name=None)
    }
    totalrna_mismatch_count = sum(
        not all(
            _optional_numeric_exact_equal(left_value, right_value)
            for left_value, right_value in zip(
                transcript_totalrna[key_hash],
                translation_totalrna[key_hash],
            )
        )
        for key_hash in intersection_hashes
    )
    if totalrna_mismatch_count:
        raise ReconstructionError(
            "Lim 6c key-aligned TotalRNA mismatch across endpoint blocks; "
            f"keys={totalrna_mismatch_count}"
        )

    transcript_key_by_physical_row = {
        int(row_number): (str(description), str(barcode))
        for description, barcode, row_number in transcript_block.loc[
            :, [*LIM6C_KEY_COLUMNS, "_physical_row_number"]
        ].itertuples(index=False, name=None)
    }
    translation_key_by_physical_row = {
        int(row_number): (str(description), str(barcode))
        for description, barcode, row_number in translation_efficiency_block.loc[
            :, [*LIM6C_KEY_COLUMNS, "_physical_row_number"]
        ].itertuples(index=False, name=None)
    }
    rowwise_key_equal_count = sum(
        transcript_key_by_physical_row[row_number]
        == translation_key_by_physical_row[row_number]
        for row_number in set(transcript_key_by_physical_row)
        & set(translation_key_by_physical_row)
    )

    structure = {
        "physical_data_rows": len(physical_rows),
        "transcript_present_key_rows": len(transcript_block),
        "translation_efficiency_present_key_rows": len(
            translation_efficiency_block
        ),
        "translation_efficiency_blank_tail_rows": right_blank_tail_rows,
        "transcript_duplicate_key_rows": 0,
        "translation_efficiency_duplicate_key_rows": 0,
        "hashed_key_intersection": len(intersection_hashes),
        "hashed_key_transcript_only": len(
            transcript_hashes - translation_hashes
        ),
        "hashed_key_translation_efficiency_only": len(
            translation_hashes - transcript_hashes
        ),
        "hashed_key_union": len(transcript_hashes | translation_hashes),
        "key_aligned_totalrna_exact_equal": len(intersection_hashes),
        "key_aligned_totalrna_any_difference": 0,
        "rowwise_key_equal_count_descriptive_only": rowwise_key_equal_count,
        "blocks_are_row_aligned": False,
        "row_position_join_used": False,
        "outer_join_used_for_inventory_only": True,
        "missing_key_is_zero": False,
        "join_key": list(LIM6C_KEY_COLUMNS),
        "join_key_hash_scheme": (
            "sha256(compact_json_utf8([description,barcode]))"
        ),
    }
    return transcript_block, translation_efficiency_block, structure


def _enforce_public_lim6c_inventory(structure: Mapping[str, Any]) -> None:
    for field, expected in PUBLIC_LIM6C_BLOCK_INVENTORY.items():
        if structure.get(field) != expected:
            raise ReconstructionError(
                f"public Lim 6c block inventory mismatch: {field}"
            )


def _parse_description(value: str) -> dict[str, Any] | None:
    mutant_match = DESCRIPTION_MUTANT_RE.fullmatch(value)
    if mutant_match is not None:
        start = int(mutant_match.group("start"))
        end = int(mutant_match.group("end"))
        if start > end or mutant_match.group("ref") == mutant_match.group("alt"):
            return None
        return {
            "type": "mutant",
            "gene": mutant_match.group("gene"),
            "ref": mutant_match.group("ref"),
            "alt": mutant_match.group("alt"),
            "chrom": mutant_match.group("chrom"),
            "start": start,
            "end": end,
            "description": value,
        }
    wt_match = DESCRIPTION_WT_RE.fullmatch(value)
    if wt_match is not None:
        start = int(wt_match.group("start"))
        end = int(wt_match.group("end"))
        if start > end:
            return None
        return {
            "type": "wt",
            "gene": wt_match.group("gene"),
            "chrom": wt_match.group("chrom"),
            "start": start,
            "end": end,
            "description": value,
        }
    return None


def _matching_sequence_pair(
    description: Mapping[str, Any],
    entries_6a: Sequence[Mapping[str, Any]],
) -> tuple[dict[str, Any] | None, str | None]:
    mutant_matches = [
        dict(entry)
        for entry in entries_6a
        if entry["type"] == "mutant"
        and entry["gene"] == description["gene"]
        and entry["chrom"] == description["chrom"]
        and entry["ref"] == description["ref"]
        and entry["alt"] == description["alt"]
        and description["start"] <= entry["position"] <= description["end"]
    ]
    if not mutant_matches:
        return None, "MISSING_MUTANT_6A_FULL_LENGTH_SEQUENCE"
    if len(mutant_matches) != 1:
        return None, "AMBIGUOUS_MUTANT_6A_FULL_LENGTH_SEQUENCE"
    mutant = mutant_matches[0]

    wt_matches = [
        dict(entry)
        for entry in entries_6a
        if entry["type"] == "wt"
        and entry["gene"] == description["gene"]
        and entry["chrom"] == description["chrom"]
        and entry["position"] == mutant["position"]
    ]
    if not wt_matches:
        return None, "MISSING_WT_6A_FULL_LENGTH_SEQUENCE"
    if len(wt_matches) != 1:
        return None, "AMBIGUOUS_WT_6A_FULL_LENGTH_SEQUENCE"
    wt = wt_matches[0]

    source = wt["sequence"]
    candidate = mutant["sequence"]
    if len(source) != len(candidate):
        return None, "NOT_EXACTLY_ONE_SNV"
    differences = [
        index
        for index, (source_base, candidate_base) in enumerate(zip(source, candidate))
        if source_base != candidate_base
    ]
    if len(differences) != 1:
        return None, "NOT_EXACTLY_ONE_SNV"
    difference_index = differences[0]
    if (
        source[difference_index] != description["ref"]
        or candidate[difference_index] != description["alt"]
    ):
        return None, "SNV_SEQUENCE_REF_ALT_MISMATCH"

    return {
        "source_sequence": source,
        "candidate_sequence": candidate,
        "sequence_index_0_based": difference_index,
        "mutant_6a_coordinate": mutant["coordinate"],
        "wt_6a_coordinate": wt["coordinate"],
        "variant_position": mutant["position"],
    }, None


def _technical_barcode_observations(
    rows: pd.DataFrame,
    *,
    numerator_column: str,
    denominator_column: str,
) -> list[dict[str, Any]]:
    observations: list[dict[str, Any]] = []
    for description, barcode, physical_row, block_id, raw_numerator, raw_denominator in zip(
        rows["description"].tolist(),
        rows["barcode"].tolist(),
        rows["_physical_row_number"].tolist(),
        rows["_block_id"].tolist(),
        rows[numerator_column].tolist(),
        rows[denominator_column].tolist(),
    ):
        numerator = None if pd.isna(raw_numerator) else float(raw_numerator)
        denominator = None if pd.isna(raw_denominator) else float(raw_denominator)
        if numerator is None and denominator is None:
            reason = "BOTH_COMPONENT_LOG2_CPMS_MISSING"
        elif numerator is None:
            reason = "NUMERATOR_LOG2_CPM_MISSING"
        elif denominator is None:
            reason = "DENOMINATOR_LOG2_CPM_MISSING"
        elif (
            numerator < PUBLISHED_LOG2_CPM_MINIMUM_INCLUSIVE
            and denominator < PUBLISHED_LOG2_CPM_MINIMUM_INCLUSIVE
        ):
            reason = "BOTH_COMPONENT_LOG2_CPMS_BELOW_NEGATIVE_ONE"
        elif numerator < PUBLISHED_LOG2_CPM_MINIMUM_INCLUSIVE:
            reason = "NUMERATOR_LOG2_CPM_BELOW_NEGATIVE_ONE"
        elif denominator < PUBLISHED_LOG2_CPM_MINIMUM_INCLUSIVE:
            reason = "DENOMINATOR_LOG2_CPM_BELOW_NEGATIVE_ONE"
        else:
            reason = None
        ratio = (
            numerator - denominator
            if reason is None and numerator is not None and denominator is not None
            else None
        )
        observations.append(
            {
                "description": description,
                "barcode": barcode,
                "physical_row_number": int(physical_row),
                "source_block_id": str(block_id),
                "key_present_in_endpoint_block": True,
                "numerator_log2_cpm": numerator,
                "denominator_log2_cpm": denominator,
                "eligible": reason is None,
                "exclusion_reason": reason,
                "ratio": ratio,
            }
        )
    return observations


def _technical_barcode_ratios(
    rows: pd.DataFrame,
    *,
    numerator_column: str,
    denominator_column: str,
) -> list[float]:
    """Return eligible ratios only; retained for focused formula testing."""

    return [
        float(observation["ratio"])
        for observation in _technical_barcode_observations(
            rows,
            numerator_column=numerator_column,
            denominator_column=denominator_column,
        )
        if observation["eligible"]
    ]


def _stable_digest(prefix: str, payload: Mapping[str, Any], *, length: int = 20) -> str:
    encoded = json.dumps(
        payload, ensure_ascii=True, separators=(",", ":"), sort_keys=True
    ).encode("utf-8")
    return prefix + hashlib.sha256(encoded).hexdigest()[:length]


def _construct_ids(
    *,
    mutant_description: str,
    wt_description: str,
    sequence_pair: Mapping[str, Any],
) -> tuple[str, str]:
    wt_construct_id = _stable_digest(
        "GSE149487_WT_",
        {
            "description": wt_description,
            "sequence": sequence_pair["source_sequence"],
        },
    )
    mutant_construct_id = _stable_digest(
        "GSE149487_MUT_",
        {
            "description": mutant_description,
            "sequence": sequence_pair["candidate_sequence"],
        },
    )
    return wt_construct_id, mutant_construct_id


def _salted_barcode_sha256(
    barcode: str,
    *,
    lim6c_sha256: str,
) -> str:
    payload = f"{DATASET_ID}\0{lim6c_sha256}\0{barcode}".encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def _row_locator_sha256(
    physical_row_number: int,
    *,
    lim6c_sha256: str,
    block_id: str,
) -> str:
    locator = (
        f"{lim6c_sha256}\0{LIM6C_SHEET}\0{block_id}\0"
        f"physical_row={physical_row_number}"
    ).encode("utf-8")
    return hashlib.sha256(locator).hexdigest()


def _description_sha256(description: str, *, lim6c_sha256: str) -> str:
    payload = f"{lim6c_sha256}\0{description}".encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def _closed_record_fields() -> dict[str, Any]:
    return {
        "evidence": "BLOCKED_PENDING_PUBLIC_EVIDENCE",
        "claim": "NOT_ESTABLISHED",
        "role": "ORDINARY_DEVELOPMENT",
        "data_role": "EXCLUDED",
        "split": "EXCLUDED",
        "split_execution": "NOT_RUN",
        "exposure": "DEVELOPMENT_ONLY",
        "checkpoint_exposure_status": "AUDIT_PENDING",
        "sequence_exposed": True,
        "label_exposed": True,
        "legacy_use": True,
        "license": "UNKNOWN_BLOCKED",
        "eligibility": "PENDING_BLOCKED",
        "qualification": False,
        "qualified": False,
        "training": False,
        "model_selection": False,
    }


def _make_replicate_summaries(
    *,
    pair_id: str,
    endpoint: Mapping[str, str],
    mutant_rows: pd.DataFrame,
    wt_rows: pd.DataFrame,
    mutant_description: str,
    wt_description: str,
    mutant_union_barcodes: set[str],
    wt_union_barcodes: set[str],
    mutant_construct_id: str,
    wt_construct_id: str,
    moesm8_sha256: str,
    lim6c_sha256: str,
    protocol_sha256: str,
    script_sha256: str,
    barcode_audit_sink: Callable[[Mapping[str, Any]], None],
) -> tuple[list[dict[str, Any]], list[float]]:
    pending: list[dict[str, Any]] = []
    deltas: list[float] = []
    for replicate in (1, 2, 3):
        numerator = endpoint["numerator_template"].format(replicate=replicate)
        denominator = endpoint["denominator_template"].format(replicate=replicate)
        mutant_present_observations = _technical_barcode_observations(
            mutant_rows,
            numerator_column=numerator,
            denominator_column=denominator,
        )
        wt_present_observations = _technical_barcode_observations(
            wt_rows,
            numerator_column=numerator,
            denominator_column=denominator,
        )
        def retain_union_with_explicit_missing(
            *,
            description: str,
            present: Sequence[Mapping[str, Any]],
            union_barcodes: set[str],
        ) -> list[dict[str, Any]]:
            present_by_barcode = {
                str(observation["barcode"]): dict(observation)
                for observation in present
            }
            if len(present_by_barcode) != len(present):
                raise ReconstructionError(
                    "internal error: duplicate endpoint-block barcode within description"
                )
            complete: list[dict[str, Any]] = []
            for barcode in sorted(union_barcodes):
                observation = present_by_barcode.get(barcode)
                if observation is not None:
                    complete.append(observation)
                    continue
                complete.append(
                    {
                        "description": description,
                        "barcode": barcode,
                        "physical_row_number": None,
                        "source_block_id": endpoint["block_id"],
                        "key_present_in_endpoint_block": False,
                        "numerator_log2_cpm": None,
                        "denominator_log2_cpm": None,
                        "eligible": False,
                        "exclusion_reason": (
                            "KEY_MISSING_FROM_ENDPOINT_BLOCK_NOT_ZERO"
                        ),
                        "ratio": None,
                    }
                )
            return complete

        mutant_observations = retain_union_with_explicit_missing(
            description=mutant_description,
            present=mutant_present_observations,
            union_barcodes=mutant_union_barcodes,
        )
        wt_observations = retain_union_with_explicit_missing(
            description=wt_description,
            present=wt_present_observations,
            union_barcodes=wt_union_barcodes,
        )
        mutant_ratios = [
            float(observation["ratio"])
            for observation in mutant_observations
            if observation["eligible"]
        ]
        wt_ratios = [
            float(observation["ratio"])
            for observation in wt_observations
            if observation["eligible"]
        ]

        for arm, construct_id, observations in (
            ("MUTANT", mutant_construct_id, mutant_observations),
            ("WT", wt_construct_id, wt_observations),
        ):
            for observation in observations:
                barcode = str(observation["barcode"])
                description = str(observation["description"])
                physical_row_raw = observation["physical_row_number"]
                physical_row = (
                    int(physical_row_raw)
                    if physical_row_raw is not None
                    else None
                )
                source_block_id = str(observation["source_block_id"])
                if source_block_id != endpoint["block_id"]:
                    raise ReconstructionError(
                        "internal error: endpoint received rows from the wrong Lim 6c block"
                    )
                barcode_audit_sink(
                    {
                        "schema_version": SCHEMA_VERSION,
                        "dataset_id": DATASET_ID,
                        "study_group_id": STUDY_GROUP_ID,
                        "independent_study_id": INDEPENDENT_STUDY_ID,
                        "independent_study_count": 1,
                        "pair_id": pair_id,
                        "construct_id": construct_id,
                        "arm": arm,
                        "cell_context": CONTEXT,
                        "biological_replicate_id": f"rep{replicate}",
                        "biological_replicate_index": replicate,
                        "endpoint_id": endpoint["endpoint_id"],
                        "endpoint_definition": endpoint["definition"],
                        "numerator_log2_cpm_field": numerator,
                        "denominator_log2_cpm_field": denominator,
                        "numerator_log2_cpm": observation[
                            "numerator_log2_cpm"
                        ],
                        "denominator_log2_cpm": observation[
                            "denominator_log2_cpm"
                        ],
                        "original_cpm_minimum_inclusive": (
                            ORIGINAL_CPM_MINIMUM_INCLUSIVE
                        ),
                        "published_log2_cpm_minimum_inclusive": (
                            PUBLISHED_LOG2_CPM_MINIMUM_INCLUSIVE
                        ),
                        "component_filter": (
                            "BOTH_PUBLISHED_LOG2_CPM_COMPONENTS_GTE_NEGATIVE_ONE"
                        ),
                        "primary_min_eligible_barcodes_per_arm_replicate": (
                            PRIMARY_MIN_ELIGIBLE_BARCODES_PER_ARM_REPLICATE
                        ),
                        "eligible": observation["eligible"],
                        "exclusion_reason": observation["exclusion_reason"],
                        "barcode_log2_ratio": observation["ratio"],
                        "barcode_id_sha256": _salted_barcode_sha256(
                            barcode, lim6c_sha256=lim6c_sha256
                        ),
                        "block_join_key_sha256": _lim6c_key_sha256(
                            description,
                            barcode,
                        ),
                        "source_asset_id": LIM6C_SOURCE_ASSET_ID,
                        "sheet_name": LIM6C_SHEET,
                        "source_block_id": source_block_id,
                        "source_excel_columns": endpoint["excel_columns"],
                        "key_present_in_endpoint_block": observation[
                            "key_present_in_endpoint_block"
                        ],
                        "physical_row_number": physical_row,
                        "row_locator_sha256": (
                            _row_locator_sha256(
                                physical_row,
                                lim6c_sha256=lim6c_sha256,
                                block_id=source_block_id,
                            )
                            if physical_row is not None
                            else None
                        ),
                        "raw_barcode_id_emitted": False,
                        "technical_unit": "BARCODE",
                        "barcode_is_independent_n": False,
                        "transform_scope": "PAPER_FAITHFUL_RATIO_INPUT_ONLY",
                        "paper_test_reproduced": False,
                        "paper_test_membership_status": (
                            "KEY_ABSENT_FROM_ENDPOINT_BLOCK_NOT_TESTED"
                            if not observation["key_present_in_endpoint_block"]
                            else "ELIGIBLE_RATIO_ONLY_NOT_TESTED"
                            if observation["eligible"]
                            else "EXCLUDED_BY_DUAL_COMPONENT_LOG2_CPM_FILTER"
                        ),
                        "provenance": {
                            "moesm8_sha256": moesm8_sha256,
                            "lim6c_293t_sha256": lim6c_sha256,
                            "lim6c_sheet": LIM6C_SHEET,
                            "lim6c_block_id": source_block_id,
                            "protocol_sha256": protocol_sha256,
                            "script_sha256": script_sha256,
                        },
                    }
                )
        mutant_present_count = len(mutant_present_observations)
        wt_present_count = len(wt_present_observations)
        reason: str | None = None
        if mutant_present_count == 0 and wt_present_count == 0:
            reason = "BOTH_ARMS_MISSING_FROM_ENDPOINT_BLOCK"
        elif mutant_present_count == 0:
            reason = "MUTANT_ARM_MISSING_FROM_ENDPOINT_BLOCK"
        elif wt_present_count == 0:
            reason = "WT_ARM_MISSING_FROM_ENDPOINT_BLOCK"
        elif not mutant_ratios and not wt_ratios:
            reason = "NO_QUALIFYING_MUTANT_OR_WT_TECHNICAL_BARCODES"
        elif not mutant_ratios:
            reason = "NO_QUALIFYING_MUTANT_TECHNICAL_BARCODES"
        elif not wt_ratios:
            reason = "NO_QUALIFYING_WT_TECHNICAL_BARCODES"
        elif (
            len(mutant_ratios) < PRIMARY_MIN_ELIGIBLE_BARCODES_PER_ARM_REPLICATE
            and len(wt_ratios) < PRIMARY_MIN_ELIGIBLE_BARCODES_PER_ARM_REPLICATE
        ):
            reason = "BOTH_ARMS_TECHNICAL_COVERAGE_BELOW_PREFROZEN_20"
        elif len(mutant_ratios) < PRIMARY_MIN_ELIGIBLE_BARCODES_PER_ARM_REPLICATE:
            reason = "MUTANT_ARM_TECHNICAL_COVERAGE_BELOW_PREFROZEN_20"
        elif len(wt_ratios) < PRIMARY_MIN_ELIGIBLE_BARCODES_PER_ARM_REPLICATE:
            reason = "WT_ARM_TECHNICAL_COVERAGE_BELOW_PREFROZEN_20"

        mutant_median = (
            float(statistics.median(mutant_ratios)) if mutant_ratios else None
        )
        wt_median = float(statistics.median(wt_ratios)) if wt_ratios else None
        delta = (
            mutant_median - wt_median
            if reason is None
            and mutant_median is not None
            and wt_median is not None
            else None
        )
        if delta is not None:
            deltas.append(float(delta))
        pending.append(
            {
                "schema_version": SCHEMA_VERSION,
                "dataset_id": DATASET_ID,
                "study_group_id": STUDY_GROUP_ID,
                "independent_study_id": INDEPENDENT_STUDY_ID,
                "independent_study_count": 1,
                "pair_id": pair_id,
                "endpoint_id": endpoint["endpoint_id"],
                "endpoint_definition": endpoint["definition"],
                "replicate_id": f"rep{replicate}",
                "biological_replicate_index": replicate,
                "mutant_construct_barcode_ratio_median": mutant_median,
                "wt_construct_barcode_ratio_median": wt_median,
                "replicate_delta_mutant_minus_wt": delta,
                "mutant_technical_barcode_count_total": mutant_present_count,
                "mutant_union_barcode_key_count": len(mutant_union_barcodes),
                "mutant_endpoint_block_missing_key_count": (
                    len(mutant_union_barcodes) - mutant_present_count
                ),
                "mutant_technical_barcode_count_retained": len(mutant_ratios),
                "wt_technical_barcode_count_total": wt_present_count,
                "wt_union_barcode_key_count": len(wt_union_barcodes),
                "wt_endpoint_block_missing_key_count": (
                    len(wt_union_barcodes) - wt_present_count
                ),
                "wt_technical_barcode_count_retained": len(wt_ratios),
                "primary_min_eligible_barcodes_per_arm_replicate": (
                    PRIMARY_MIN_ELIGIBLE_BARCODES_PER_ARM_REPLICATE
                ),
                "replicate_status": "ACCEPTED" if reason is None else "REJECTED",
                "replicate_rejection_reason": reason,
                "technical_unit": "BARCODE",
                "barcode_is_independent_n": False,
                "analysis_method": "ROUTE_A_COMPANION_NOT_PAPER_TEST",
                "paper_test_reproduced": False,
                "protocol_sha256": protocol_sha256,
            }
        )
    endpoint_accepted = len(deltas) == 3
    for summary in pending:
        summary["endpoint_aggregate_status"] = (
            "COMPANION_SUMMARY_ACCEPTED_NOT_QUALIFIED"
            if endpoint_accepted
            else "REJECTED_BIOLOGICAL_REPLICATE_COUNT_NOT_THREE"
        )
        summary["required_biological_replicate_count"] = 3
        summary["complete_biological_replicate_count_observed"] = len(deltas)
        summary["endpoint_effective_n"] = 3 if endpoint_accepted else None
        summary["effective_n_unit"] = "BIOLOGICAL_REPLICATE"
    return pending, deltas


def _development_companion_effect_record(
    *,
    description: Mapping[str, Any],
    wt_description: str,
    sequence_pair: Mapping[str, Any],
    endpoint: Mapping[str, str],
    deltas: Sequence[float],
    replicate_summaries: Sequence[Mapping[str, Any]],
    pair_id: str,
    moesm8_provenance: Mapping[str, Any],
    lim6c_provenance: Mapping[str, Any],
    protocol_provenance: Mapping[str, Any],
    script_sha256: str,
) -> dict[str, Any]:
    primary_coverage_pass = len(deltas) == 3
    delta = float(statistics.fmean(deltas)) if primary_coverage_pass else None
    standard_error = (
        float(statistics.stdev(deltas) / math.sqrt(3.0))
        if primary_coverage_pass
        else None
    )

    wt_construct_id, mutant_construct_id = _construct_ids(
        mutant_description=str(description["description"]),
        wt_description=wt_description,
        sequence_pair=sequence_pair,
    )
    group_binding = {
        "wt_construct_id": wt_construct_id,
        "gene": description["gene"],
        "locus": {
            "chromosome": description["chrom"],
            "description_start": description["start"],
            "description_end": description["end"],
            "variant_position_6a": sequence_pair["variant_position"],
        },
        "design_family": DESIGN_FAMILY,
    }
    group_id = _stable_digest("GSE149487_GROUP_", group_binding)
    record_id = pair_id + "::" + endpoint["endpoint_id"]

    record: dict[str, Any] = {
        "schema_id": "route_a_v3.development_companion_effect_record",
        "schema_version": SCHEMA_VERSION,
        "record_type": "DEVELOPMENT_COMPANION_EFFECT_SUMMARY",
        "canonical_intervention_record_v3_materialized": False,
        "record_id": record_id,
        "pair_id": pair_id,
        "dataset_id": DATASET_ID,
        "dataset_alias": DATASET_ALIAS,
        "study_group_id": STUDY_GROUP_ID,
        "independent_study_id": INDEPENDENT_STUDY_ID,
        "independent_study_count": 1,
        "context": CONTEXT,
        "region": REGION,
        "design_family": DESIGN_FAMILY,
        "group_id": group_id,
        "group_binding": group_binding,
        "wt_construct_id": wt_construct_id,
        "mutant_construct_id": mutant_construct_id,
        "gene": description["gene"],
        "chromosome": description["chrom"],
        "description_start": description["start"],
        "description_end": description["end"],
        "source_sequence": sequence_pair["source_sequence"],
        "candidate_sequence": sequence_pair["candidate_sequence"],
        "edit": {
            "type": "SNV",
            "sequence_index_0_based": sequence_pair["sequence_index_0_based"],
            "ref": description["ref"],
            "alt": description["alt"],
        },
        "endpoint_id": endpoint["endpoint_id"],
        "endpoint_definition": endpoint["definition"],
        "direction_convention": "HIGHER_IS_BETTER",
        "direction_interpretation": (
            "Positive delta means mutant median ratio exceeds WT median ratio; "
            "this convention is not a biological-benefit or qualification claim."
        ),
        "effect_delta_mutant_minus_wt": delta,
        "standard_error": standard_error,
        "biological_replicate_deltas": [
            summary["replicate_delta_mutant_minus_wt"]
            for summary in replicate_summaries
        ],
        "biological_replicate_count": 3,
        "complete_effect_biological_replicate_count": len(deltas),
        "effective_n": 3 if primary_coverage_pass else None,
        "effective_n_unit": "BIOLOGICAL_REPLICATE",
        "technical_unit": "BARCODE",
        "barcode_is_independent_n": False,
        "primary_technical_coverage_floor_per_arm_replicate": (
            PRIMARY_MIN_ELIGIBLE_BARCODES_PER_ARM_REPLICATE
        ),
        "primary_technical_coverage_status": (
            "PASS" if primary_coverage_pass else "REJECTED"
        ),
        "companion_effect_status": (
            "ESTIMATED_DEVELOPMENT_ONLY"
            if primary_coverage_pass
            else "NULL_REJECTED_INSUFFICIENT_PRIMARY_TECHNICAL_COVERAGE"
        ),
        "endpoint_rejection_reason": (
            None
            if primary_coverage_pass
            else "BIOLOGICAL_REPLICATE_COUNT_NOT_THREE_AFTER_PRIMARY_COVERAGE"
        ),
        "analysis_method": "ROUTE_A_COMPANION_NOT_PAPER_TEST",
        "paper_inferential_test_reproduced": False,
        "inferential_test": "NOT_RUN",
        "provenance": {
            "moesm8": {
                "sha256": moesm8_provenance["sha256"],
                "sheet": MOESM8_SEQUENCE_SHEET,
                "mutant_coordinate_locator": sequence_pair["mutant_6a_coordinate"],
                "wt_coordinate_locator": sequence_pair["wt_6a_coordinate"],
            },
            "lim6c_293t": {
                "sha256": lim6c_provenance["sha256"],
                "sheet": LIM6C_SHEET,
                "endpoint_source_block_id": endpoint["block_id"],
                "endpoint_source_excel_columns": endpoint["excel_columns"],
                "block_join_key": list(LIM6C_KEY_COLUMNS),
                "row_position_join_used": False,
                "mutant_description_locator": description["description"],
                "wt_description_locator": wt_description,
            },
            "script": {
                "name": Path(__file__).name,
                "sha256": script_sha256,
            },
            "protocol": {
                "protocol_id": PROTOCOL_ID,
                "sha256": protocol_provenance["sha256"],
            },
        },
    }
    record.update(_closed_record_fields())
    return record


def _json_bytes(payload: Any, *, pretty: bool = False) -> bytes:
    kwargs: dict[str, Any] = {
        "allow_nan": False,
        "ensure_ascii": True,
        "sort_keys": True,
    }
    if pretty:
        kwargs["indent"] = 2
    else:
        kwargs["separators"] = (",", ":")
    return (json.dumps(payload, **kwargs) + "\n").encode("utf-8")


def _jsonl_bytes(rows: Sequence[Mapping[str, Any]]) -> bytes:
    return b"".join(_json_bytes(row) for row in rows)


def _write_exclusive(path: Path, payload: bytes) -> None:
    try:
        descriptor = os.open(path, os.O_WRONLY | os.O_CREAT | os.O_EXCL, 0o644)
    except FileExistsError as exc:
        raise ReconstructionError("refusing to overwrite output artifact") from exc
    with os.fdopen(descriptor, "wb") as handle:
        handle.write(payload)
        handle.flush()
        os.fsync(handle.fileno())


def _write_exclusive_stream(path: Path, source: BinaryIO) -> str:
    try:
        descriptor = os.open(path, os.O_WRONLY | os.O_CREAT | os.O_EXCL, 0o644)
    except FileExistsError as exc:
        raise ReconstructionError("refusing to overwrite output artifact") from exc
    digest = hashlib.sha256()
    source.seek(0)
    with os.fdopen(descriptor, "wb") as handle:
        for block in iter(lambda: source.read(1 << 20), b""):
            handle.write(block)
            digest.update(block)
        handle.flush()
        os.fsync(handle.fileno())
    return digest.hexdigest()


def _fsync_directory(path: Path) -> None:
    descriptor = os.open(path, os.O_RDONLY)
    try:
        os.fsync(descriptor)
    finally:
        os.close(descriptor)


def _create_output_staging(
    output: Path,
) -> tuple[Path, Path, tuple[int, int]]:
    """Claim an output name and create a unique hidden sibling staging dir.

    The fixed-name O_EXCL claim coordinates cooperating invocations.  Python's
    standard library does not expose a portable no-replace directory rename,
    so publication also rechecks the final target immediately before rename.
    A failure intentionally leaves the claim/staging evidence for audit.
    """

    _require_output_target(output)
    token = secrets.token_hex(16)
    staging = output.parent / f".{output.name}.partial-staging-{token}"
    claim = output.parent / f".{output.name}.publish-claim"
    try:
        descriptor = os.open(
            claim,
            os.O_WRONLY | os.O_CREAT | os.O_EXCL,
            0o600,
        )
    except FileExistsError as exc:
        raise ReconstructionError(
            "output publication claim already exists; inspect partial evidence"
        ) from exc

    claim_identity: tuple[int, int]
    with os.fdopen(descriptor, "wb") as handle:
        claim_info = os.fstat(handle.fileno())
        claim_identity = (claim_info.st_dev, claim_info.st_ino)
        handle.write(
            _json_bytes(
                {
                    "status": "PARTIAL_OUTPUT_STAGING_ACTIVE_NOT_FINAL",
                    "target_directory_name": output.name,
                    "staging_directory_name": staging.name,
                }
            )
        )
        handle.flush()
        os.fsync(handle.fileno())

    try:
        staging.mkdir(mode=0o700, exist_ok=False)
    except FileExistsError as exc:
        raise ReconstructionError(
            "unique partial staging directory unexpectedly already exists"
        ) from exc
    _fsync_directory(output.parent)
    return staging, claim, claim_identity


def _publish_staged_output(
    *,
    staging: Path,
    output: Path,
    claim: Path,
    claim_identity: tuple[int, int],
) -> None:
    """Publish one fully written sibling directory and never replace a target."""

    _fsync_directory(staging)
    _require_output_target(output)
    try:
        current_claim = claim.lstat()
    except FileNotFoundError as exc:
        raise ReconstructionError(
            "output publication claim disappeared; partial staging retained"
        ) from exc
    if not stat.S_ISREG(current_claim.st_mode) or (
        current_claim.st_dev,
        current_claim.st_ino,
    ) != claim_identity:
        raise ReconstructionError(
            "output publication claim changed; partial staging retained"
        )

    try:
        os.rename(staging, output)
    except OSError as exc:
        raise ReconstructionError(
            "atomic output publication failed; partial staging retained"
        ) from exc

    # Once rename succeeds the final directory is complete.  Durability and
    # claim cleanup are best-effort so a post-publication housekeeping failure
    # cannot be misreported as a failed reconstruction with a final-looking dir.
    try:
        _fsync_directory(output.parent)
    except OSError:
        pass
    try:
        current_claim = claim.lstat()
        if stat.S_ISREG(current_claim.st_mode) and (
            current_claim.st_dev,
            current_claim.st_ino,
        ) == claim_identity:
            claim.unlink()
    except OSError:
        pass
    try:
        _fsync_directory(output.parent)
    except OSError:
        pass


def _build_artifacts(
    *,
    table_6a: pd.DataFrame,
    transcript_block: pd.DataFrame,
    translation_efficiency_block: pd.DataFrame,
    workbook_structure: Mapping[str, Any],
    moesm8_provenance: Mapping[str, Any],
    lim6c_provenance: Mapping[str, Any],
    protocol_provenance: Mapping[str, Any],
    script_sha256: str,
    barcode_audit_sink: Callable[[Mapping[str, Any]], None],
) -> tuple[
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[dict[str, Any]],
    dict[str, Any],
]:
    entries_6a, moesm8_6a_classification_rows = _parse_6a_table(
        table_6a,
        moesm8_sha256=str(moesm8_provenance["sha256"]),
    )
    if len(moesm8_6a_classification_rows) != len(table_6a):
        raise ReconstructionError(
            "internal error: MOESM8 6a physical-row classification is incomplete"
        )
    moesm8_6a_classification_counts = Counter(
        row["classification"] for row in moesm8_6a_classification_rows
    )
    if (
        set(moesm8_6a_classification_counts) - set(MOESM8_6A_ALLOWED_CLASSES)
        or sum(moesm8_6a_classification_counts.values()) != len(table_6a)
    ):
        raise ReconstructionError(
            "internal error: MOESM8 6a classification denominator mismatch"
        )
    moesm8_6a_rejection_reasons = Counter(
        str(row["classification_reason"])
        for row in moesm8_6a_classification_rows
        if str(row["classification"]).startswith("REJECTED_")
    )
    transcript_groups = {
        str(description): group.reset_index(drop=True)
        for description, group in transcript_block.groupby(
            "description", sort=True, dropna=False
        )
    }
    translation_efficiency_groups = {
        str(description): group.reset_index(drop=True)
        for description, group in translation_efficiency_block.groupby(
            "description", sort=True, dropna=False
        )
    }
    description_universe = sorted(
        set(transcript_groups) | set(translation_efficiency_groups)
    )
    parsed = {
        description: _parse_description(description)
        for description in description_universe
    }
    mutant_descriptions = sorted(
        description
        for description, fields in parsed.items()
        if fields is not None and fields["type"] == "mutant"
    )
    wt_descriptions = {
        description
        for description, fields in parsed.items()
        if fields is not None and fields["type"] == "wt"
    }
    unsupported_description_count = sum(fields is None for fields in parsed.values())

    development_companion_effect_records: list[dict[str, Any]] = []
    replicate_summaries: list[dict[str, Any]] = []
    sequence_pair_rejections: Counter[str] = Counter()
    endpoint_rejections: Counter[str] = Counter()
    replicate_rejections: Counter[str] = Counter()
    used_wt_descriptions: set[str] = set()
    description_decisions: dict[str, dict[str, Any]] = {}
    sequence_valid_pair_count = 0
    barcode_audit_row_count = 0

    def emit_barcode_audit(row: Mapping[str, Any]) -> None:
        nonlocal barcode_audit_row_count
        barcode_audit_sink(row)
        barcode_audit_row_count += 1

    for mutant_description in mutant_descriptions:
        fields = parsed[mutant_description]
        assert fields is not None
        wt_description = (
            f"{fields['gene']}_WT_{fields['chrom']}_"
            f"{fields['start']}_{fields['end']}"
        )
        if wt_description not in parsed:
            reason = "MISSING_WT_LIM6C_DESCRIPTION"
            description_decisions[mutant_description] = {
                "classification": "REJECTED_UNADJUDICATED_DESCRIPTION",
                "classification_category": "rejected_with_reason",
                "classification_reason": reason,
                "pair_id": None,
            }
            sequence_pair_rejections[reason] += 1
            endpoint_rejections[reason] += len(ENDPOINTS)
            continue

        sequence_pair, sequence_reason = _matching_sequence_pair(fields, entries_6a)
        if sequence_pair is None:
            assert sequence_reason is not None
            description_decisions[mutant_description] = {
                "classification": "REJECTED_UNADJUDICATED_DESCRIPTION",
                "classification_category": "rejected_with_reason",
                "classification_reason": sequence_reason,
                "pair_id": None,
            }
            sequence_pair_rejections[sequence_reason] += 1
            endpoint_rejections[sequence_reason] += len(ENDPOINTS)
            continue

        sequence_valid_pair_count += 1
        used_wt_descriptions.add(wt_description)
        pair_id = _stable_digest(
            "GSE149487_PAIR_",
            {
                "mutant_description": mutant_description,
                "wt_description": wt_description,
                "source_sequence": sequence_pair["source_sequence"],
                "candidate_sequence": sequence_pair["candidate_sequence"],
            },
        )
        description_decisions[mutant_description] = {
            "classification": "INCLUDED_STRICT_SNV_MUTANT",
            "classification_category": "included",
            "classification_reason": (
                "STRICT_SNV_PAIR_INCLUDED_FOR_DEVELOPMENT_COMPANION"
            ),
            "pair_id": pair_id,
        }

        def union_barcodes(description: str) -> set[str]:
            barcodes: set[str] = set()
            for description_groups in (
                transcript_groups,
                translation_efficiency_groups,
            ):
                rows = description_groups.get(description)
                if rows is not None:
                    barcodes.update(str(value) for value in rows["barcode"].tolist())
            return barcodes

        mutant_union_barcodes = union_barcodes(mutant_description)
        wt_union_barcodes = union_barcodes(wt_description)
        wt_construct_id, mutant_construct_id = _construct_ids(
            mutant_description=mutant_description,
            wt_description=wt_description,
            sequence_pair=sequence_pair,
        )
        for endpoint in ENDPOINTS:
            if endpoint["block_id"] == TRANSCRIPT_BLOCK_ID:
                endpoint_groups = transcript_groups
                empty_endpoint_rows = transcript_block.iloc[0:0]
            elif endpoint["block_id"] == TRANSLATION_EFFICIENCY_BLOCK_ID:
                endpoint_groups = translation_efficiency_groups
                empty_endpoint_rows = translation_efficiency_block.iloc[0:0]
            else:
                raise ReconstructionError("internal error: unknown endpoint block")
            mutant_rows = endpoint_groups.get(
                mutant_description,
                empty_endpoint_rows,
            )
            wt_rows = endpoint_groups.get(wt_description, empty_endpoint_rows)
            summaries, deltas = _make_replicate_summaries(
                pair_id=pair_id,
                endpoint=endpoint,
                mutant_rows=mutant_rows,
                wt_rows=wt_rows,
                mutant_description=mutant_description,
                wt_description=wt_description,
                mutant_union_barcodes=mutant_union_barcodes,
                wt_union_barcodes=wt_union_barcodes,
                mutant_construct_id=mutant_construct_id,
                wt_construct_id=wt_construct_id,
                moesm8_sha256=str(moesm8_provenance["sha256"]),
                lim6c_sha256=str(lim6c_provenance["sha256"]),
                protocol_sha256=str(protocol_provenance["sha256"]),
                script_sha256=script_sha256,
                barcode_audit_sink=emit_barcode_audit,
            )
            replicate_summaries.extend(summaries)
            for summary in summaries:
                reason = summary["replicate_rejection_reason"]
                if reason is not None:
                    replicate_rejections[str(reason)] += 1
            if len(deltas) != 3:
                endpoint_rejections["BIOLOGICAL_REPLICATE_COUNT_NOT_THREE"] += 1
            development_companion_effect_records.append(
                _development_companion_effect_record(
                    description=fields,
                    wt_description=wt_description,
                    sequence_pair=sequence_pair,
                    endpoint=endpoint,
                    deltas=deltas,
                    replicate_summaries=summaries,
                    pair_id=pair_id,
                    moesm8_provenance=moesm8_provenance,
                    lim6c_provenance=lim6c_provenance,
                    protocol_provenance=protocol_provenance,
                    script_sha256=script_sha256,
                )
            )

    description_classification_rows: list[dict[str, Any]] = []
    for description in description_universe:
        fields = parsed[description]
        if fields is None:
            decision = {
                "classification": "REJECTED_UNADJUDICATED_DESCRIPTION",
                "classification_category": "rejected_with_reason",
                "classification_reason": (
                    "UNADJUDICATED_DESCRIPTION_CLASS_EXCLUDED"
                ),
                "pair_id": None,
            }
        elif fields["type"] == "wt":
            decision = {
                "classification": "REFERENCE_ONLY_STRICT_WT",
                "classification_category": "reference_only",
                "classification_reason": (
                    "WT_REFERENCE_FOR_INCLUDED_MUTANT"
                    if description in used_wt_descriptions
                    else "WT_REFERENCE_NOT_USED_BY_INCLUDED_MUTANT"
                ),
                "pair_id": None,
            }
        else:
            decision = description_decisions.get(description)
            if decision is None:
                raise ReconstructionError(
                    "internal error: strict mutant description was not classified"
                )

        transcript_description_rows = transcript_groups.get(description)
        translation_description_rows = translation_efficiency_groups.get(description)
        candidate_sources: list[tuple[int, str, str]] = []
        if transcript_description_rows is not None:
            candidate_sources.append(
                (
                    int(transcript_description_rows["_physical_row_number"].min()),
                    TRANSCRIPT_BLOCK_ID,
                    "A:H",
                )
            )
        if translation_description_rows is not None:
            candidate_sources.append(
                (
                    int(translation_description_rows["_physical_row_number"].min()),
                    TRANSLATION_EFFICIENCY_BLOCK_ID,
                    "J:Q",
                )
            )
        if not candidate_sources:
            raise ReconstructionError(
                "internal error: description has no Lim 6c source block"
            )
        first_physical_row, first_source_block_id, first_source_excel_columns = min(
            candidate_sources,
            key=lambda item: (
                item[0],
                0 if item[1] == TRANSCRIPT_BLOCK_ID else 1,
            ),
        )
        transcript_barcodes = (
            set(str(value) for value in transcript_description_rows["barcode"].tolist())
            if transcript_description_rows is not None
            else set()
        )
        translation_barcodes = (
            set(str(value) for value in translation_description_rows["barcode"].tolist())
            if translation_description_rows is not None
            else set()
        )
        description_classification_rows.append(
            {
                "schema_id": "route_a_v3.description_classification_audit",
                "schema_version": SCHEMA_VERSION,
                "dataset_id": DATASET_ID,
                "study_group_id": STUDY_GROUP_ID,
                "independent_study_id": INDEPENDENT_STUDY_ID,
                "independent_study_count": 1,
                "context": CONTEXT,
                "source_asset_id": LIM6C_SOURCE_ASSET_ID,
                "sheet_name": LIM6C_SHEET,
                "description": description,
                "description_sha256": _description_sha256(
                    description,
                    lim6c_sha256=str(lim6c_provenance["sha256"]),
                ),
                "first_physical_row_number": first_physical_row,
                "first_source_block_id": first_source_block_id,
                "first_source_excel_columns": first_source_excel_columns,
                "first_row_locator_sha256": _row_locator_sha256(
                    first_physical_row,
                    lim6c_sha256=str(lim6c_provenance["sha256"]),
                    block_id=first_source_block_id,
                ),
                "transcript_block_key_count": len(transcript_barcodes),
                "translation_efficiency_block_key_count": len(
                    translation_barcodes
                ),
                "union_key_count": len(transcript_barcodes | translation_barcodes),
                "classification": decision["classification"],
                "classification_category": decision["classification_category"],
                "classification_reason": decision["classification_reason"],
                "pair_id": decision["pair_id"],
                "raw_barcode_id_emitted": False,
            }
        )

    if len(description_classification_rows) != len(description_universe):
        raise ReconstructionError(
            "internal error: description-universe classification is incomplete"
        )
    description_classification_counts = Counter(
        row["classification"] for row in description_classification_rows
    )
    if sum(description_classification_counts.values()) != len(description_universe):
        raise ReconstructionError(
            "internal error: description classification denominator mismatch"
        )
    description_rejection_reasons = Counter(
        str(row["classification_reason"])
        for row in description_classification_rows
        if row["classification_category"] == "rejected_with_reason"
    )
    description_classification_category_counts = Counter(
        row["classification_category"] for row in description_classification_rows
    )

    development_companion_effect_records.sort(key=lambda row: row["record_id"])
    replicate_summaries.sort(
        key=lambda row: (
            row["pair_id"],
            row["endpoint_id"],
            row["biological_replicate_index"],
        )
    )
    endpoint_slot_count = len(mutant_descriptions) * len(ENDPOINTS)
    development_companion_nonnull_effect_record_count = sum(
        record["effect_delta_mutant_minus_wt"] is not None
        for record in development_companion_effect_records
    )
    rejected_endpoint_slot_count = (
        endpoint_slot_count - development_companion_nonnull_effect_record_count
    )
    if rejected_endpoint_slot_count != sum(endpoint_rejections.values()):
        raise ReconstructionError("internal error: endpoint rejection accounting mismatch")

    report: dict[str, Any] = {
        "contract_id": CONTRACT_ID,
        "schema_version": SCHEMA_VERSION,
        "dataset_id": DATASET_ID,
        "dataset_alias": DATASET_ALIAS,
        "study_group_id": STUDY_GROUP_ID,
        "independent_study_id": INDEPENDENT_STUDY_ID,
        "independent_study_count": 1,
        "endpoint_or_context_increases_independent_study_count": False,
        "reconstruction_status": RECONSTRUCTION_STATUS,
        "canonical_intervention_record_v3_materialized": False,
        "development_companion_effect_summary_materialized": True,
        "qualification": False,
        "qualified": False,
        "training": False,
        "training_allowed": False,
        "model_selection": False,
        "model_selection_allowed": False,
        "paper_method_reproduced": False,
        "companion_summary_only": True,
        "whole_study_context_closed": False,
        "coverage_threshold_prefrozen": True,
        "coverage_threshold_freeze_timing": (
            "PREFROZEN_BEFORE_REAL_RECONSTRUCTION_RESULTS"
        ),
        "output_publication": {
            "bundle_written_in_hidden_sibling_staging_directory": True,
            "fixed_name_o_excl_publication_claim": True,
            "final_directory_created_only_by_sibling_directory_rename": True,
            "existing_final_target_rejected": True,
            "partial_staging_evidence_retained_on_failure": True,
            "portable_stdlib_directory_rename_noreplace_available": False,
            "residual_uncooperative_destination_creation_race": True,
        },
        "evidence": "BLOCKED_PENDING_PUBLIC_EVIDENCE",
        "claim": "NOT_ESTABLISHED",
        "role": "ORDINARY_DEVELOPMENT",
        "data_role": "EXCLUDED",
        "split": "EXCLUDED",
        "split_execution": "NOT_RUN",
        "exposure": "DEVELOPMENT_ONLY",
        "checkpoint_exposure_status": "AUDIT_PENDING",
        "sequence_exposed": True,
        "label_exposed": True,
        "legacy_use": True,
        "license": "UNKNOWN_BLOCKED",
        "eligibility": "PENDING_BLOCKED",
        "scope": {
            "included": [
                "MOESM8_6A_FULL_LENGTH_SEQUENCES",
                "LIM6C_293T_PUBLISHED_LOG2_CPM_PER_BARCODE",
                "HASHED_BARCODE_RATIO_AUDIT",
                "DESCRIPTION_UNIVERSE_CLASSIFICATION_AUDIT",
                "MOESM8_6A_PHYSICAL_ROW_CLASSIFICATION_AUDIT",
            ],
            "not_included": [
                "PC3",
                "18_GEO_RAW_COUNT_TABLES",
                "MOESM3",
                "PUBLISHED_6D_6E_LFC_DESCRIPTIVE_COMPARISON",
            ],
        },
        "blockers": list(PROTOCOL_UNRESOLVED_BLOCKERS),
        "protocol": {
            "protocol_id": PROTOCOL_ID,
            "protocol_status": PROTOCOL_STATUS,
            "filename": protocol_provenance["filename"],
            "sha256": protocol_provenance["sha256"],
            "bytes": protocol_provenance["bytes"],
            "strict_runtime_constant_binding": True,
        },
        "paper_method": {
            "barcode_distribution_test": "TWO_SIDED_MANN_WHITNEY_MUTANT_VS_WT",
            "multiple_testing_rule": "FDR_LT_0.1",
            "reproduced": False,
        },
        "companion_method": {
            "analysis_method": "ROUTE_A_COMPANION_NOT_PAPER_TEST",
            "input_value_scale": "PUBLISHED_LOG2_CPM_PER_BARCODE",
            "original_cpm_minimum_inclusive": ORIGINAL_CPM_MINIMUM_INCLUSIVE,
            "published_log2_cpm_minimum_inclusive": (
                PUBLISHED_LOG2_CPM_MINIMUM_INCLUSIVE
            ),
            "component_filter": (
                "BOTH_PUBLISHED_LOG2_CPM_COMPONENTS_GTE_NEGATIVE_ONE"
            ),
            "transcript_ratio_transform": (
                "TOTALRNA_LOG2_CPM_MINUS_DNA_LOG2_CPM"
            ),
            "translation_efficiency_ratio_transform": (
                "POLYSOME_LOG2_CPM_MINUS_TOTALRNA_LOG2_CPM"
            ),
            "clipping_allowed": False,
            "missing_is_zero": False,
            "primary_min_eligible_barcodes_per_arm_endpoint_replicate": (
                PRIMARY_MIN_ELIGIBLE_BARCODES_PER_ARM_REPLICATE
            ),
            "primary_technical_coverage_freeze_timing": (
                "PREFROZEN_BEFORE_REAL_RECONSTRUCTION_RESULTS"
            ),
            "secondary_audit_only_barcode_floors": list(
                SECONDARY_AUDIT_ONLY_BARCODE_FLOORS
            ),
            "secondary_floors_can_change_primary_eligibility": False,
            "within_construct_endpoint_replicate_reducer": "MEDIAN",
            "replicate_delta": "MUTANT_MEDIAN_MINUS_WT_MEDIAN",
            "effect": "ARITHMETIC_MEAN_OF_THREE_REPLICATE_DELTAS",
            "standard_error": "SAMPLE_SD_OF_THREE_DELTAS_DIVIDED_BY_SQRT_3",
            "effective_n_unit": "BIOLOGICAL_REPLICATE",
            "required_effective_n": 3,
            "barcodes_are_technical_units": True,
            "barcodes_are_independent_n": False,
        },
        "paper_faithful_transform_scope": {
            "included": [
                "PUBLISHED_SAMPLE_LOG2_CPM_VALUES",
                "DUAL_COMPONENT_LOG2_CPM_GTE_NEGATIVE_ONE_FILTER",
                "PER_BARCODE_LOG2_RATIO_BY_COMPONENT_SUBTRACTION",
            ],
            "not_included": [
                "MANN_WHITNEY_TEST",
                "FDR_ADJUSTMENT",
                "PUBLISHED_SIGNIFICANT_MEMBERSHIP_FILTER",
            ],
            "barcode_ratio_audit_complete_for_reconstructed_pairs": True,
            "raw_barcode_ids_emitted": False,
        },
        "lim6c_block_semantics": {
            **dict(workbook_structure["lim6c_block_semantics"]),
            "transcript_endpoint_source_block_id": TRANSCRIPT_BLOCK_ID,
            "translation_efficiency_endpoint_source_block_id": (
                TRANSLATION_EFFICIENCY_BLOCK_ID
            ),
            "endpoint_uses_only_its_own_block": True,
            "union_description_universe_used": True,
            "outer_key_union_used_for_inventory_and_explicit_missing_only": True,
            "missing_endpoint_key_emission": (
                "NULL_WITH_KEY_MISSING_FROM_ENDPOINT_BLOCK_NOT_ZERO"
            ),
            "row_position_join_used": False,
        },
        "barcode_ratio_audit": {
            "filename": "BARCODE_RATIO_AUDIT.jsonl",
            "input_value_scale": "PUBLISHED_LOG2_CPM_PER_BARCODE",
            "component_value_fields": [
                "numerator_log2_cpm",
                "denominator_log2_cpm",
            ],
            "ratio_calculation": (
                "NUMERATOR_LOG2_CPM_MINUS_DENOMINATOR_LOG2_CPM"
            ),
            "complete_for_sequence_valid_mutant_wt_pairs": True,
            "eligible_and_excluded_observations_retained": True,
            "raw_barcode_ids_emitted": False,
            "barcode_id_sha256_scheme": (
                "sha256(utf8(dataset_id + NUL + lim6c_sha256 + NUL + raw_barcode))"
            ),
            "row_locator_sha256_scheme": (
                "sha256(utf8(lim6c_sha256 + NUL + sheet + NUL + block_id + NUL + "
                "'physical_row=' + one_based_physical_row))"
            ),
            "raw_row_locator_fields_emitted": [
                "source_asset_id",
                "sheet_name",
                "source_block_id",
                "physical_row_number",
            ],
            "missing_endpoint_key_has_null_physical_row_and_locator": True,
            "raw_row_locator_is_ordinary_public_non_sensitive": True,
            "barcode_counts_may_weight_replicates": False,
            "barcode_counts_may_increase_effective_n": False,
            "barcode_counts_may_define_standard_error": False,
        },
        "description_classification_audit": {
            "filename": "DESCRIPTION_CLASSIFICATION_AUDIT.jsonl",
            "source_asset_id": LIM6C_SOURCE_ASSET_ID,
            "sheet_name": LIM6C_SHEET,
            "universe": "UNION_OF_ALL_UNIQUE_DESCRIPTIONS_ACROSS_BOTH_LIM6C_BLOCKS",
            "classification_values": list(DESCRIPTION_ALLOWED_CLASSES),
            "classification_category_values": [
                "included",
                "reference_only",
                "rejected_with_reason",
            ],
            "complete": True,
            "raw_barcode_ids_emitted": False,
        },
        "moesm8_6a_classification_audit": {
            "filename": "MOESM8_6A_CLASSIFICATION_AUDIT.jsonl",
            "source_asset_id": MOESM8_SOURCE_ASSET_ID,
            "sheet_name": MOESM8_SEQUENCE_SHEET,
            "universe": "ALL_PHYSICAL_MOESM8_6A_DATA_ROWS",
            "classification_values": list(MOESM8_6A_ALLOWED_CLASSES),
            "complete": True,
            "successful_join_may_shrink_physical_row_denominator": False,
            "raw_coordinates_emitted": False,
            "raw_sequences_emitted": False,
            "coordinate_sha256_scheme": (
                "sha256(utf8(moesm8_sha256 + NUL + sheet + NUL + "
                "'physical_row=' + row + NUL + 'coordinate_type=' + type + NUL) "
                "+ canonical_coordinate_value_bytes)"
            ),
            "coordinate_value_canonicalization": (
                "EMPTY_TO_ZERO_BYTES;STRING_TO_EXACT_UTF8;BOOL_INT_FINITE_FLOAT_"
                "TO_COMPACT_JSON;NONFINITE_FLOAT_TO_ASCII_REPR;"
                "ISOFORMAT_WHEN_AVAILABLE;OTHERWISE_UTF8_STRING"
            ),
            "row_locator_sha256_scheme": (
                "sha256(utf8(moesm8_sha256 + NUL + sheet + NUL + "
                "'physical_row=' + one_based_physical_row))"
            ),
        },
        "input_workbooks": {
            "moesm8": {
                "source_asset_id": MOESM8_SOURCE_ASSET_ID,
                "expected_filename": EXPECTED_MOESM8_FILENAME,
                "actual_filename": moesm8_provenance["filename"],
                "filename_match": True,
                "sha256": moesm8_provenance["sha256"],
                "bytes": moesm8_provenance["bytes"],
                "sequence_sheet": MOESM8_SEQUENCE_SHEET,
            },
            "lim6c_293t": {
                "source_asset_id": LIM6C_SOURCE_ASSET_ID,
                "expected_filename": EXPECTED_LIM6C_FILENAME,
                "actual_filename": lim6c_provenance["filename"],
                "filename_match": True,
                "sha256": lim6c_provenance["sha256"],
                "bytes": lim6c_provenance["bytes"],
                "sheet": LIM6C_SHEET,
            },
        },
        "workbook_structure": dict(workbook_structure),
        "script": {
            "filename": Path(__file__).name,
            "sha256": script_sha256,
        },
        "summary": {
            "moesm8_6a_physical_data_row_denominator": len(table_6a),
            "moesm8_6a_classification_complete": True,
            "moesm8_6a_classification_counts": {
                classification: int(
                    moesm8_6a_classification_counts.get(classification, 0)
                )
                for classification in MOESM8_6A_ALLOWED_CLASSES
            },
            "moesm8_6a_rejection_reason_counts": dict(
                sorted(moesm8_6a_rejection_reasons.items())
            ),
            "lim6c_unique_description_count": len(description_universe),
            "description_universe_denominator": len(description_universe),
            "lim6c_transcript_block_unique_description_count": len(
                transcript_groups
            ),
            "lim6c_translation_efficiency_block_unique_description_count": len(
                translation_efficiency_groups
            ),
            "lim6c_transcript_present_key_rows": len(transcript_block),
            "lim6c_translation_efficiency_present_key_rows": len(
                translation_efficiency_block
            ),
            "lim6c_hashed_key_intersection": workbook_structure[
                "lim6c_block_semantics"
            ]["hashed_key_intersection"],
            "lim6c_hashed_key_transcript_only": workbook_structure[
                "lim6c_block_semantics"
            ]["hashed_key_transcript_only"],
            "lim6c_hashed_key_translation_efficiency_only": workbook_structure[
                "lim6c_block_semantics"
            ]["hashed_key_translation_efficiency_only"],
            "lim6c_key_aligned_totalrna_exact_equal": workbook_structure[
                "lim6c_block_semantics"
            ]["key_aligned_totalrna_exact_equal"],
            "lim6c_key_aligned_totalrna_any_difference": workbook_structure[
                "lim6c_block_semantics"
            ]["key_aligned_totalrna_any_difference"],
            "lim6c_missing_endpoint_keys_are_zero": False,
            "lim6c_row_position_join_used": False,
            "description_classification_complete": True,
            "description_classification_counts": {
                classification: int(
                    description_classification_counts.get(classification, 0)
                )
                for classification in DESCRIPTION_ALLOWED_CLASSES
            },
            "description_classification_category_counts": {
                category: int(
                    description_classification_category_counts.get(category, 0)
                )
                for category in (
                    "included",
                    "reference_only",
                    "rejected_with_reason",
                )
            },
            "description_rejection_reason_counts": dict(
                sorted(description_rejection_reasons.items())
            ),
            "strict_mutant_description_count": len(mutant_descriptions),
            "strict_wt_description_count": len(wt_descriptions),
            "unsupported_unique_description_count": unsupported_description_count,
            "sequence_valid_pair_count": sequence_valid_pair_count,
            "sequence_rejected_pair_count": sum(sequence_pair_rejections.values()),
            "unused_strict_wt_description_count": len(
                wt_descriptions - used_wt_descriptions
            ),
            "endpoint_pair_slot_count": endpoint_slot_count,
            "canonical_record_count": 0,
            "canonical_intervention_record_v3_materialized": False,
            "development_companion_effect_record_count": len(
                development_companion_effect_records
            ),
            "development_companion_nonnull_effect_record_count": (
                development_companion_nonnull_effect_record_count
            ),
            "development_companion_null_effect_record_count": (
                len(development_companion_effect_records)
                - development_companion_nonnull_effect_record_count
            ),
            "rejected_endpoint_pair_count": rejected_endpoint_slot_count,
            "replicate_summary_count": len(replicate_summaries),
            "barcode_ratio_audit_row_count": barcode_audit_row_count,
            "sequence_pair_rejection_reason_counts": dict(
                sorted(sequence_pair_rejections.items())
            ),
            "endpoint_rejection_reason_counts": dict(
                sorted(endpoint_rejections.items())
            ),
            "replicate_rejection_reason_counts": dict(
                sorted(replicate_rejections.items())
            ),
        },
        "outputs": list(OUTPUT_FILES),
    }
    return (
        development_companion_effect_records,
        replicate_summaries,
        description_classification_rows,
        moesm8_6a_classification_rows,
        report,
    )


def reconstruct_gse149487_plumage(
    *,
    moesm8_path: Path,
    lim6c_path: Path,
    protocol_path: Path,
    expected_moesm8_sha256: str,
    expected_lim6c_sha256: str,
    expected_protocol_sha256: str,
    output_directory: Path,
) -> dict[str, Any]:
    """Build deterministic V3 development artifacts in a new directory."""

    expected_moesm8 = _normalize_expected_sha256(
        expected_moesm8_sha256, label="MOESM8"
    )
    expected_lim6c = _normalize_expected_sha256(
        expected_lim6c_sha256, label="Lim 6c"
    )
    expected_protocol = _normalize_expected_sha256(
        expected_protocol_sha256, label="reconstruction protocol"
    )
    (
        moesm8,
        lim6c,
        protocol,
        output,
        moesm8_info,
        lim6c_info,
        protocol_info,
    ) = _prepare_paths_before_read(
        Path(moesm8_path),
        Path(lim6c_path),
        Path(protocol_path),
        Path(output_directory),
    )

    # Both expected hashes are verified before the first workbook parser call.
    moesm8_provenance = _verify_input_hash(
        moesm8,
        expected_moesm8,
        moesm8_info,
        label="MOESM8 workbook",
    )
    lim6c_provenance = _verify_input_hash(
        lim6c,
        expected_lim6c,
        lim6c_info,
        label="Lim 6c workbook",
    )
    protocol_provenance = _verify_protocol_hash(
        protocol,
        expected_protocol,
        protocol_info,
    )
    _load_and_validate_protocol(
        protocol,
        expected_moesm8_sha256=expected_moesm8,
        expected_lim6c_sha256=expected_lim6c,
    )

    (
        table_6a,
        transcript_block,
        translation_efficiency_block,
        workbook_structure,
    ) = _load_exact_workbook_tables(moesm8, lim6c)
    public_lim6c_inventory_gate_applied = expected_lim6c == PUBLIC_LIM6C_SHA256
    if public_lim6c_inventory_gate_applied:
        _enforce_public_lim6c_inventory(
            workbook_structure["lim6c_block_semantics"]
        )
    workbook_structure["lim6c_block_semantics"][
        "public_inventory_gate_applied"
    ] = public_lim6c_inventory_gate_applied

    # Detect any workbook replacement or mutation between verification and parse.
    for path, expected, provenance, label in (
        (moesm8, expected_moesm8, moesm8_provenance, "MOESM8 workbook"),
        (lim6c, expected_lim6c, lim6c_provenance, "Lim 6c workbook"),
    ):
        info = _require_regular_xlsx(path, label=label)
        observed, observed_size = _sha256_and_size(path)
        if (
            observed != expected
            or observed_size != provenance["bytes"]
            or _identity(info) != tuple(provenance["identity"])
        ):
            raise ReconstructionError(f"{label} changed after pre-read verification")
    protocol_after = _require_regular_json(
        protocol, label="reconstruction protocol"
    )
    protocol_observed, protocol_size = _sha256_and_size(protocol)
    if (
        protocol_observed != expected_protocol
        or protocol_size != protocol_provenance["bytes"]
        or _identity(protocol_after) != tuple(protocol_provenance["identity"])
    ):
        raise ReconstructionError(
            "reconstruction protocol changed after pre-read verification"
        )

    script_sha256, _ = _sha256_and_size(Path(__file__).resolve())
    # A spooled stream keeps the full per-barcode audit reproducible without
    # retaining roughly six copies of the large Lim 6c table in Python objects.
    with tempfile.SpooledTemporaryFile(max_size=32 << 20, mode="w+b") as audit_stream:
        (
            development_companion_effect_records,
            replicate_summaries,
            description_classification_rows,
            moesm8_6a_classification_rows,
            report,
        ) = _build_artifacts(
            table_6a=table_6a,
            transcript_block=transcript_block,
            translation_efficiency_block=translation_efficiency_block,
            workbook_structure=workbook_structure,
            moesm8_provenance=moesm8_provenance,
            lim6c_provenance=lim6c_provenance,
            protocol_provenance=protocol_provenance,
            script_sha256=script_sha256,
            barcode_audit_sink=lambda row: audit_stream.write(_json_bytes(row)),
        )

        payloads = {
            "development_companion_effect_records.jsonl": _jsonl_bytes(
                development_companion_effect_records
            ),
            "replicate_effect_summaries.jsonl": _jsonl_bytes(replicate_summaries),
            "DESCRIPTION_CLASSIFICATION_AUDIT.jsonl": _jsonl_bytes(
                description_classification_rows
            ),
            "MOESM8_6A_CLASSIFICATION_AUDIT.jsonl": _jsonl_bytes(
                moesm8_6a_classification_rows
            ),
            "RECONSTRUCTION_REPORT.json": _json_bytes(report, pretty=True),
        }

        # Build the complete bundle in a hidden sibling and publish only after
        # every file, fsync, and checksum succeeds.  Failures intentionally keep
        # the hidden partial evidence and never create the final directory.
        staging, claim, claim_identity = _create_output_staging(output)

        artifact_hashes: dict[str, str] = {}
        for name in (
            "development_companion_effect_records.jsonl",
            "replicate_effect_summaries.jsonl",
            "DESCRIPTION_CLASSIFICATION_AUDIT.jsonl",
            "MOESM8_6A_CLASSIFICATION_AUDIT.jsonl",
        ):
            _write_exclusive(staging / name, payloads[name])
            artifact_hashes[name] = hashlib.sha256(payloads[name]).hexdigest()

        audit_name = "BARCODE_RATIO_AUDIT.jsonl"
        artifact_hashes[audit_name] = _write_exclusive_stream(
            staging / audit_name, audit_stream
        )

        report_name = "RECONSTRUCTION_REPORT.json"
        _write_exclusive(staging / report_name, payloads[report_name])
        artifact_hashes[report_name] = hashlib.sha256(
            payloads[report_name]
        ).hexdigest()

        checksum_payload = "".join(
            f"{artifact_hashes[name]}  {name}\n"
            for name in sorted(artifact_hashes)
        ).encode("ascii")
        _write_exclusive(staging / "SHA256SUMS", checksum_payload)
        _publish_staged_output(
            staging=staging,
            output=output,
            claim=claim,
            claim_identity=claim_identity,
        )
    return report


def reconstruct(**kwargs: Any) -> dict[str, Any]:
    """Short import-friendly alias for :func:`reconstruct_gse149487_plumage`."""

    return reconstruct_gse149487_plumage(**kwargs)


def _parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--moesm8-workbook", "--moesm8", required=True, type=Path, dest="moesm8"
    )
    parser.add_argument(
        "--moesm8-sha256",
        "--expected-moesm8-sha256",
        required=True,
        dest="moesm8_sha256",
    )
    parser.add_argument(
        "--lim6c-workbook", "--lim6c", required=True, type=Path, dest="lim6c"
    )
    parser.add_argument(
        "--lim6c-sha256",
        "--expected-lim6c-sha256",
        required=True,
        dest="lim6c_sha256",
    )
    parser.add_argument("--protocol", required=True, type=Path)
    parser.add_argument("--protocol-sha256", required=True)
    parser.add_argument(
        "--output-directory", "--output-dir", required=True, type=Path, dest="output"
    )
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = _parse_args(argv)
    report = reconstruct_gse149487_plumage(
        moesm8_path=args.moesm8,
        lim6c_path=args.lim6c,
        protocol_path=args.protocol,
        expected_moesm8_sha256=args.moesm8_sha256,
        expected_lim6c_sha256=args.lim6c_sha256,
        expected_protocol_sha256=args.protocol_sha256,
        output_directory=args.output,
    )
    summary = report["summary"]
    print(
        json.dumps(
            {
                "canonical_record_count": 0,
                "canonical_intervention_record_v3_materialized": False,
                "development_companion_effect_record_count": summary[
                    "development_companion_effect_record_count"
                ],
                "companion_summary_only": True,
                "dataset_id": DATASET_ID,
                "qualified": False,
                "reconstruction_status": RECONSTRUCTION_STATUS,
                "rejected_endpoint_pair_count": summary[
                    "rejected_endpoint_pair_count"
                ],
            },
            sort_keys=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
