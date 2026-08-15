#!/usr/bin/env python3
"""Convert final public GSE186455 allelic effects to Route 2 SUB records."""

from __future__ import annotations

import argparse
import csv
import gzip
import io
import json
import math
import os
import shutil
import tarfile
import tempfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
CONFIG_PATH = REPO_ROOT / "configs/route_a_v3_route2_gse186455_converter_v1.json"
CANONICAL_SCHEMA_VERSION = "mrna_editflow_route2_canonical.v1"
BASES = set("ACGT")


class ConversionError(RuntimeError):
    pass


def _require(condition: bool, message: str) -> None:
    if not condition:
        raise ConversionError(message)


def _number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def load_config(path: Path = CONFIG_PATH) -> dict[str, Any]:
    config = json.loads(path.read_text(encoding="utf-8"))
    validate_config(config)
    return config


def validate_config(config: Mapping[str, Any]) -> None:
    _require(config["schema_version"] == "route_a_v3_route2_gse186455_converter.v1", "unexpected schema version")
    study = config["study"]
    _require(study["study_unit_id"] == "GSE186455", "unexpected study")
    _require(study["pool_assignment"] == "DEVELOPMENT", "study left Development")
    _require(study["qualification_class"] == "DEVELOPMENT_RELAXED", "qualification class changed")
    _require(study["study_role"] == "A1_LIBRARY_INDEPENDENCE_REFERENCE_DEVELOPMENT", "study role changed")
    _require(study["biological_context_ids"] == ["N2A_TRANSCRIPT_ABUNDANCE", "VGLUT_TRANSCRIPT_ABUNDANCE"], "context set changed")
    _require(study["conversion_scope"] == "FINAL_PUBLISHED_REF_VS_ALT_LMM_STRICT_SUB_ONLY", "conversion scope changed")
    expected = config["input"]
    _require(expected["expected_tar_member_count"] == 13, "tar member count changed")
    _require(expected["expected_unique_element_count"] == 1624, "element count changed")
    _require(expected["expected_source_group_count"] == 653, "source group count changed")
    _require(expected["expected_complete_ref_alt_pair_count"] == 649, "pair count changed")
    _require(expected["expected_action_counts"] == {"LENGTH_CHANGE": 77, "SUB": 572}, "action counts changed")
    _require(expected["expected_author_bad_alt_count"] == 13, "author exclusion count changed")
    _require(expected["expected_author_bad_alt_sub_count"] == 1, "author SUB exclusion count changed")
    _require(expected["expected_workbook_element_count_each"] == 649, "workbook universe changed")
    _require(expected["expected_finite_effect_count"] == {"N2A_TRANSCRIPT_ABUNDANCE": 520, "VGLUT_TRANSCRIPT_ABUNDANCE": 305}, "published effect coverage changed")
    _require(expected["expected_context_record_counts"] == {"N2A_TRANSCRIPT_ABUNDANCE": 471, "VGLUT_TRANSCRIPT_ABUNDANCE": 281}, "context output count changed")
    _require(expected["expected_canonical_record_count"] == 752, "canonical count changed")
    _require(expected["expected_distinct_sequence_pair_count"] == 461, "sequence-pair count changed")
    _require(expected["expected_exact_sequence_pair_context_duplicate_count"] == 18, "exact duplicate count changed")
    _require(config["action_policy"] == {"allowed_candidate_action": "SUB", "ins_supported": False, "del_supported": False}, "action policy changed")
    output = config["output"]
    _require(output["overwrite_allowed"] is False and output["public_redistribution_allowed"] is False, "output policy changed")
    _require(output["directory"].startswith("/mnt/cunyuliu/mrna_xeditflow_routea_v3/route2/"), "output leaves Route 2 root")
    development = config["development_policy"]
    _require(development["training_eligible"] is True and development["model_selection_eligible"] is True, "Development use disabled")
    _require(development["confirmatory_evaluation_eligible"] is False, "confirmatory use enabled")
    _require(development["source_endpoint_value_representation"] is None and development["candidate_endpoint_value_representation"] is None, "absolute endpoints fabricated")
    _require(development["reported_standard_error_status"] == "PUBLISHED_LMM_EFFECT_SE_NOT_REPORTED", "SE status changed")
    _require(development["biological_standard_error_representation"] is None, "standard error fabricated")
    _require(development["exact_sequence_pair_group_binding_required"] is True, "exact duplicate grouping disabled")
    _require(development["near_duplicate_split_status"] == "NOT_RUN", "split overstated")
    credit = config["credit_policy"]
    _require(credit["measured_candidate"] is True and credit["generated_candidate"] is False, "candidate role changed")
    _require(not any(credit["qualified_credit_delta"].values()), "conversion increases qualified credit")
    _require(credit["qualified_counts_after_conversion"] == {"ordinary": 1, "a1": 1, "true_a2": 0, "canonical_records": 6547}, "qualified facts changed")
    _require(config["scientific_claim_status"] == "NOT_ESTABLISHED", "scientific claim overstated")


def _load_sequence_groups(path: Path) -> tuple[dict[str, dict[str, str]], Counter[str]]:
    sequences: dict[str, str] = {}
    stats: Counter[str] = Counter()
    with tarfile.open(path) as archive:
        for member in archive.getmembers():
            if not member.isfile() or not member.name.endswith(".tab.gz"):
                continue
            stats["tar_member_count"] += 1
            extracted = archive.extractfile(member)
            _require(extracted is not None, f"cannot read tar member: {member.name}")
            with gzip.GzipFile(fileobj=extracted) as compressed:
                reader = csv.DictReader(io.TextIOWrapper(compressed, encoding="utf-8"), delimiter="\t")
                _require(reader.fieldnames is not None and {"sequence", "seqName"}.issubset(reader.fieldnames), f"unexpected columns: {member.name}")
                for row in reader:
                    stats["tar_row_count"] += 1
                    name = str(row["seqName"])
                    sequence = str(row["sequence"]).upper()
                    _require(sequence and not (set(sequence) - BASES), f"invalid sequence: {name}")
                    previous = sequences.setdefault(name, sequence)
                    _require(previous == sequence, f"sequence conflict: {name}")
    stats["unique_element_count"] = len(sequences)

    groups: dict[str, dict[str, str]] = defaultdict(dict)
    for name, sequence in sequences.items():
        pair_id, separator, allele = name.rpartition("_")
        _require(bool(separator) and allele in {"ref", "alt", "shuf"}, f"unparsed element name: {name}")
        previous = groups[pair_id].setdefault(allele, sequence)
        _require(previous == sequence, f"allele conflict: {name}")
    stats["source_group_count"] = len(groups)
    for alleles in groups.values():
        if "ref" not in alleles or "alt" not in alleles:
            stats["incomplete_ref_alt_pair_count"] += 1
            continue
        stats["complete_ref_alt_pair_count"] += 1
        source = alleles["ref"]
        candidate = alleles["alt"]
        if len(source) != len(candidate):
            stats["action_LENGTH_CHANGE"] += 1
        else:
            distance = sum(left != right for left, right in zip(source, candidate))
            stats["action_SUB" if distance == 1 else "action_OTHER"] += 1
    return dict(groups), stats


def _read_published_effects(path: Path, config: Mapping[str, Any]) -> tuple[dict[str, dict[str, dict[str, float]]], Counter[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    contexts = config["published_effect_sheets"]
    _require(set(workbook.sheetnames) == set(config["input"]["expected_workbook_sheets"]), "publisher workbook sheet set changed")
    effects: dict[str, dict[str, dict[str, float]]] = {}
    stats: Counter[str] = Counter()
    for context, sheet_name in contexts.items():
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        header = list(next(rows))
        required = {"Element", "logFC", "pval", "fdr", "bonferroni"}
        _require(required.issubset(header), f"published effect columns absent: {sheet_name}")
        context_values: dict[str, dict[str, float]] = {}
        elements: set[str] = set()
        for values in rows:
            row = dict(zip(header, values))
            if row.get("Element") is None:
                continue
            pair_id = str(row["Element"])
            _require(pair_id not in elements, f"duplicate publisher row: {context}:{pair_id}")
            elements.add(pair_id)
            parsed = {key: _number(row.get(key)) for key in ("logFC", "pval", "fdr", "bonferroni")}
            if all(value is not None for value in parsed.values()):
                _require(all(0.0 <= parsed[key] <= 1.0 for key in ("pval", "fdr", "bonferroni")), f"invalid probability: {context}:{pair_id}")
                context_values[pair_id] = {key: float(value) for key, value in parsed.items()}
                stats[f"finite_effect_{context}"] += 1
            else:
                stats[f"not_tested_{context}"] += 1
        stats[f"workbook_element_{context}"] = len(elements)
        effects[context] = context_values
    return effects, stats


def _eligible_units(config: Mapping[str, Any], groups: Mapping[str, dict[str, str]], effects: Mapping[str, dict[str, dict[str, float]]]) -> tuple[list[dict[str, Any]], Counter[str]]:
    bad_alt = set(config["author_bad_alt_pair_ids"])
    _require(bad_alt.issubset(groups), "author bad-alt set is outside sequence universe")
    stats: Counter[str] = Counter(author_bad_alt_declared_count=len(bad_alt))
    units: list[dict[str, Any]] = []
    for pair_id in sorted(groups):
        alleles = groups[pair_id]
        if "ref" not in alleles or "alt" not in alleles:
            stats["reject_incomplete_ref_alt_pair"] += 1
            continue
        source = alleles["ref"]
        candidate = alleles["alt"]
        if len(source) != len(candidate):
            stats["reject_length_change_pair"] += 1
            continue
        changes = [index for index, (left, right) in enumerate(zip(source, candidate)) if left != right]
        if len(changes) != 1:
            stats["reject_non_single_sub_pair"] += 1
            continue
        if pair_id in bad_alt:
            stats["reject_author_bad_alt_sub_pair"] += 1
            continue
        position = changes[0]
        for context in config["study"]["biological_context_ids"]:
            endpoint = effects[context].get(pair_id)
            if endpoint is None:
                stats[f"reject_not_tested_{context}"] += 1
                continue
            units.append({
                "pair_id": pair_id,
                "source_sequence": source,
                "candidate_sequence": candidate,
                "position_zero_based": position,
                "biological_context_id": context,
                **endpoint,
            })
            stats[f"eligible_{context}"] += 1
    return units, stats


def _canonical_records(config: Mapping[str, Any], units: list[dict[str, Any]]) -> list[dict[str, Any]]:
    source_sequences = sorted({unit["source_sequence"] for unit in units})
    source_groups = {sequence: f"GSE186455_SOURCE_GROUP_{index:04d}" for index, sequence in enumerate(source_sequences, start=1)}
    sequence_pairs = sorted({(unit["source_sequence"], unit["candidate_sequence"]) for unit in units})
    pair_groups = {pair: f"GSE186455_EXACT_PAIR_GROUP_{index:04d}" for index, pair in enumerate(sequence_pairs, start=1)}
    development = config["development_policy"]
    credit = config["credit_policy"]
    replication = config["published_replication"]
    records: list[dict[str, Any]] = []
    for unit in sorted(units, key=lambda value: (value["pair_id"], value["biological_context_id"])):
        context = unit["biological_context_id"]
        position = unit["position_zero_based"]
        source = unit["source_sequence"]
        candidate = unit["candidate_sequence"]
        source_group = source_groups[source]
        records.append({
            "schema_version": CANONICAL_SCHEMA_VERSION,
            "canonical_record_id": f"GSE186455:{unit['pair_id']}:context:{context}",
            "study_unit_id": "GSE186455",
            "pool_assignment": config["study"]["pool_assignment"],
            "qualification_class": config["study"]["qualification_class"],
            "study_role": config["study"]["study_role"],
            "conversion_scope": config["study"]["conversion_scope"],
            "region": "3UTR",
            "biological_context_id": context,
            "assay_id": config["study"]["assay_id"],
            "endpoint_id": config["study"]["endpoint_id"],
            "endpoint_direction": config["study"]["endpoint_direction"],
            "source_id": source_group,
            "source_sequence": source,
            "source_endpoint_value": development["source_endpoint_value_representation"],
            "candidate_id": f"GSE186455:{unit['pair_id']}:candidate",
            "candidate_sequence": candidate,
            "candidate_endpoint_value": development["candidate_endpoint_value_representation"],
            "edit_operations": [{"type": "SUB", "position_zero_based": position, "ref": source[position], "alt": candidate[position]}],
            "direction_normalized_delta": unit["logFC"],
            "biological_standard_error": development["biological_standard_error_representation"],
            "standard_error_status": development["reported_standard_error_status"],
            "published_p_value": unit["pval"],
            "published_fdr": unit["fdr"],
            "published_bonferroni": unit["bonferroni"],
            "published_biological_replicate_count": replication[context]["published_biological_replicate_count"],
            "analyzed_biological_replicate_count": replication[context]["analyzed_biological_replicate_count"],
            "effective_n": None,
            "group_id": source_group,
            "exact_sequence_pair_group_id": pair_groups[(source, candidate)],
            "study_group_id": "GSE186455_LAGUNAS_2023_ASD_3UTR_MPRA_LIBRARY",
            "pair_id": unit["pair_id"],
            "gene_symbol": unit["pair_id"].rsplit("_", 1)[0].split(".", 1)[0],
            "measured_candidate": credit["measured_candidate"],
            "generated_candidate": credit["generated_candidate"],
            "training_eligible": development["training_eligible"],
            "model_selection_eligible": development["model_selection_eligible"],
            "confirmatory_evaluation_eligible": development["confirmatory_evaluation_eligible"],
            "historical_exposure_status": development["historical_exposure_status"],
            "near_duplicate_split_status": development["near_duplicate_split_status"],
        })
    return records


def _write_json(path: Path, value: Mapping[str, Any]) -> None:
    path.write_text(json.dumps(value, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def execute(config: Mapping[str, Any], tar_path: Path, workbook_path: Path, output_dir: Path) -> dict[str, Any]:
    _require(not output_dir.exists(), f"output directory already exists: {output_dir}")
    _require(tar_path.is_file(), f"sequence archive absent: {tar_path}")
    _require(workbook_path.is_file(), f"publisher workbook absent: {workbook_path}")
    expected = config["input"]
    groups, sequence_stats = _load_sequence_groups(tar_path)
    _require(sequence_stats["tar_member_count"] == expected["expected_tar_member_count"], "tar member count differs")
    _require(sequence_stats["unique_element_count"] == expected["expected_unique_element_count"], "unique element count differs")
    _require(sequence_stats["source_group_count"] == expected["expected_source_group_count"], "source group count differs")
    _require(sequence_stats["complete_ref_alt_pair_count"] == expected["expected_complete_ref_alt_pair_count"], "complete pair count differs")
    _require(sequence_stats["incomplete_ref_alt_pair_count"] == expected["expected_incomplete_ref_alt_pair_count"], "incomplete pair count differs")
    _require({"LENGTH_CHANGE": sequence_stats["action_LENGTH_CHANGE"], "SUB": sequence_stats["action_SUB"]} == expected["expected_action_counts"], "action counts differ")
    _require(sequence_stats["action_OTHER"] == 0, "unexpected equal-length multi-edit pair")

    effects, workbook_stats = _read_published_effects(workbook_path, config)
    for context in config["study"]["biological_context_ids"]:
        _require(workbook_stats[f"workbook_element_{context}"] == expected["expected_workbook_element_count_each"], f"workbook element count differs: {context}")
        _require(workbook_stats[f"finite_effect_{context}"] == expected["expected_finite_effect_count"][context], f"finite effect count differs: {context}")
        _require(set(effects[context]).issubset(groups), f"publisher effect universe is outside sequence universe: {context}")

    units, reject_stats = _eligible_units(config, groups, effects)
    _require(reject_stats["author_bad_alt_declared_count"] == expected["expected_author_bad_alt_count"], "bad-alt set count differs")
    _require(reject_stats["reject_author_bad_alt_sub_pair"] == expected["expected_author_bad_alt_sub_count"], "bad-alt SUB count differs")
    for context, count in expected["expected_context_record_counts"].items():
        _require(reject_stats[f"eligible_{context}"] == count, f"eligible context count differs: {context}")
    _require(len(units) == expected["expected_canonical_record_count"], "canonical unit count differs")
    distinct_pairs = {(unit["source_sequence"], unit["candidate_sequence"]) for unit in units}
    _require(len(distinct_pairs) == expected["expected_distinct_sequence_pair_count"], "distinct sequence-pair count differs")
    distinct_units = {(unit["source_sequence"], unit["candidate_sequence"], unit["biological_context_id"]) for unit in units}
    duplicate_count = len(units) - len(distinct_units)
    _require(duplicate_count == expected["expected_exact_sequence_pair_context_duplicate_count"], "exact sequence-pair context duplicate count differs")
    records = _canonical_records(config, units)
    _require(len(records) == len(units), "record materialization count differs")
    _require(len({record["canonical_record_id"] for record in records}) == len(records), "canonical record IDs are not unique")

    output = config["output"]
    output_dir.parent.mkdir(parents=True, exist_ok=True)
    temporary = Path(tempfile.mkdtemp(prefix=f".{output_dir.name}.", dir=output_dir.parent))
    try:
        canonical_path = temporary / output["canonical_filename"]
        with canonical_path.open("w", encoding="utf-8") as handle:
            for record in records:
                handle.write(json.dumps(record, sort_keys=True) + "\n")
        summary = {
            "status": "PASS_GSE186455_FINAL_PUBLISHED_LMM_STRICT_SUB_DEVELOPMENT",
            "study_unit_id": "GSE186455",
            "canonical_record_count": len(records),
            "context_record_counts": {context: reject_stats[f"eligible_{context}"] for context in config["study"]["biological_context_ids"]},
            "distinct_published_pair_id_count": len({unit["pair_id"] for unit in units}),
            "distinct_source_sequence_count": len({unit["source_sequence"] for unit in units}),
            "distinct_source_candidate_sequence_pair_count": len(distinct_pairs),
            "exact_sequence_pair_context_duplicate_count": duplicate_count,
            "qualified_credit_delta": config["credit_policy"]["qualified_credit_delta"],
            "qualified_counts_after_conversion": config["credit_policy"]["qualified_counts_after_conversion"],
            "scientific_claim_status": config["scientific_claim_status"],
            "limitations": config["limitations"],
        }
        rejects = {
            "sequence_archive": dict(sorted(sequence_stats.items())),
            "publisher_workbook": dict(sorted(workbook_stats.items())),
            "eligibility": dict(sorted(reject_stats.items())),
            "missing_is_zero": False,
            "non_sub_actions_retained": 0,
            "standard_errors_fabricated": 0,
        }
        _write_json(temporary / output["conversion_summary_filename"], summary)
        _write_json(temporary / output["reject_summary_filename"], rejects)
        os.replace(temporary, output_dir)
    except Exception:
        shutil.rmtree(temporary, ignore_errors=True)
        raise
    return summary


def main() -> int:
    parser = argparse.ArgumentParser(description="Convert GSE186455 final published allelic LMM effects")
    parser.add_argument("--config", type=Path, default=CONFIG_PATH)
    parser.add_argument("--sequence-tar", type=Path)
    parser.add_argument("--publisher-workbook", type=Path)
    parser.add_argument("--output-dir", type=Path)
    args = parser.parse_args()
    config = load_config(args.config)
    inputs = config["input"]
    summary = execute(
        config,
        args.sequence_tar or Path(inputs["sequence_tar_path"]),
        args.publisher_workbook or Path(inputs["publisher_workbook_path"]),
        args.output_dir or Path(config["output"]["directory"]),
    )
    print(json.dumps(summary, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
