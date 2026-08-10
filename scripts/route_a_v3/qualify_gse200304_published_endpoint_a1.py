#!/usr/bin/env python3
"""Audit the official GSE200304 published endpoint, while staying blocked.

Only the two frozen PMC Table S2/S3 assets are accepted.  The program emits
aggregate audits and an immutable ``BLOCKED_NOT_QUALIFIED`` disposition; it
has no raw-replay, canonical-materialization, training, model-selection, or
row-level output path.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import math
import os
import re
import secrets
import stat
import subprocess
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any, Callable, Iterable, Mapping, Sequence

from openpyxl import load_workbook


CONTRACT_ID = "mrna_xeditflow_route_a_v3"
SCHEMA_VERSION = "3.0.0"
PROTOCOL_ID = "ROUTE_A_V3_GSE200304_PUBLISHED_ENDPOINT_A1_V1"
PROTOCOL_STATUS = "GO_AGGREGATE_AUDIT_ONLY"
ACTIVATION_STATUS = "NO_GO_ACTIVATION"
QUALIFICATION_STATUS = "BLOCKED_NOT_QUALIFIED"
DATASET_ID = "GSE200304"
PROTOCOL_BASENAME = "route_a_v3_gse200304_published_endpoint_a1.json"
OUTPUT_ID = "ROUTE_A_V3_GSE200304_PUBLISHED_ENDPOINT_A1_BLOCKED_BUNDLE_V1"
SUCCESS_OUTCOME = "ENGINEERING_SUCCESS_IMMUTABLY_BLOCKED"
FAILURE_OUTCOME = "FAIL_CLOSED"
PUBLICATION_MARKER = "PUBLICATION_COMMIT.json"
SHA256SUMS_FILENAME = "SHA256SUMS"
SUCCESS_JSON_FILES = (
    "INPUT_INTEGRITY_AUDIT.json",
    "PUBLISHED_ENDPOINT_AUDIT.json",
    "QUALIFICATION_REPORT.json",
)
FAILURE_JSON_FILES = ("FAILURE_REPORT.json",)

SHA256_RE = re.compile(r"[0-9a-f]{64}")
COMMIT_RE = re.compile(r"[0-9a-f]{40}")
LONG_NUCLEOTIDE_RE = re.compile(r"[ACGTN]{20,}", re.IGNORECASE)
PAIR_ID_RE = re.compile(r"^[^:]+:[1-9][0-9]*_[ACGT]-[ACGT]$")
BASE_COMPLEMENT = {"A": "T", "T": "A", "C": "G", "G": "C"}

EXPECTED_DATA_ROOT = Path(
    "/mnt/cunyuliu/mrna_xeditflow_routea_v3/data/A1/GSE200304/"
    "GSE200304_PUBLIC_ASSETS_20260810T143731P0800"
)
FORBIDDEN_PATH_TOKENS = (
    "gse246381",
    "access_log",
    "sealed",
    "restricted",
    "fastq",
    "raw_replay",
    "zenodo",
    ".sam",
    ".bam",
    ".cram",
)

EXPECTED_AUTHORITY = {
    "contract_path": "docs/goals/MRNA_XEDITFLOW_ROUTE_A_V3.md",
    "contract_sha256": "cbac4c3dcba8f1b8df95d8edad52d19e3c126d1c865d0cc423537c754cc90982",
    "data_role_registry_path": "docs/execution/route_a_v3_data_role_registry.yaml",
    "data_role_registry_sha256": "746439ef5d88d8167176d19e9c675746fdc78984a66f6f123f77f6ec49523030",
    "decision_log_path": "docs/execution/route_a_v3_decision_log.yaml",
    "decision_log_sha256": "a5b041fab24d9a4309603a085fa3fcab936d69a899285bfa752689a2ee5fd4fd",
    "active_authority_commit": "d328bf04c394d4960ac11058e079c063e09280af",
    "staging_parent_head": "4eac98feb35da3ad83891a3a966dc069cbd19e99",
}

IMPLEMENTATION_BINDING_UNKNOWN = {
    "status": "UNKNOWN_NOT_ASSERTED",
    "repository_root_rule": "PROTOCOL_PATH_PARENT_PARENT",
    "implementation_commit": "UNKNOWN_NOT_ASSERTED",
    "qualifier_path": "scripts/route_a_v3/qualify_gse200304_published_endpoint_a1.py",
    "qualifier_blob_sha256": "UNKNOWN_NOT_ASSERTED",
    "test_path": "tests/route_a_v3/test_qualify_gse200304_published_endpoint_a1.py",
    "test_blob_sha256": "UNKNOWN_NOT_ASSERTED",
    "post_implementation_allowed_changed_paths": [
        "configs/route_a_v3_gse200304_published_endpoint_a1.json"
    ],
    "binding_commit_must_be_direct_child": True,
    "current_head_must_strictly_descend": True,
    "active_authority_must_be_ancestor": True,
    "clean_worktree_required": True,
    "running_script_must_match_qualifier_blob": True,
}

EXPECTED_DECISION_BOUNDARY = {
    "immutable_for_this_protocol": True,
    "qualified": False,
    "ordinary_study_contribution": 0,
    "a1_intervention_study_contribution": 0,
    "true_a2_dense_study_contribution": 0,
    "canonical_record_count": 0,
    "canonical_materialization_allowed": False,
    "training_allowed": False,
    "model_selection_allowed": False,
    "next_phase_authorized": False,
    "owner_policy_decision_made": False,
    "checkpoint_specific_use_cleared": False,
}

EXPECTED_SCOPE = {
    "ordinary_public_data_only": True,
    "region": "3UTR",
    "maximum_independent_study_count": 1,
    "published_endpoint_evidence_role": "AGGREGATE_AUDIT_ONLY_PENDING_OWNER_POLICY",
    "official_table_asset_count": 2,
    "raw_fastq_allowed": False,
    "sam_bam_cram_allowed": False,
    "raw_replay_allowed": False,
    "zenodo_code_execution_allowed": False,
    "network_access_allowed": False,
    "input_writes_allowed": False,
    "significance_filtering_allowed_for_membership": False,
    "row_level_output_allowed": False,
    "sequence_output_allowed": False,
    "barcode_output_allowed": False,
    "annotation_label_output_allowed": False,
    "row_identifier_output_allowed": False,
}

EXPECTED_ASSETS: tuple[dict[str, Any], ...] = (
    {
        "asset_id": "SOURCE_BUNDLE_MANIFEST",
        "role": "SOURCE_BUNDLE_CLOSURE_METADATA",
        "relative_path": "ASSET_ACQUISITION_MANIFEST.json",
        "format": "JSON",
        "bytes": 6426,
        "sha256": "8318990d9e3b6a0e6265bf9d1e8bc20f56f0ecfd994e83d279e733258642100c",
        "parsed_for_endpoint_statistics": False,
    },
    {
        "asset_id": "NCBI_PRJNA824033_RUNINFO",
        "role": "SOURCE_BUNDLE_CLOSURE_PROVENANCE_ONLY",
        "relative_path": "NCBI_PRJNA824033_RUNINFO.csv",
        "format": "CSV_OPAQUE_NOT_PARSED",
        "bytes": 12042,
        "sha256": "34bcedafebc41ee9ccd79483f331b62f2443df31d12691abc0a961a7201848f4",
        "parsed_for_endpoint_statistics": False,
    },
    {
        "asset_id": "PMC10540565_TABLE_S2",
        "role": "OFFICIAL_PUBLISHED_DESIGN_PAIR_GEOMETRY_ONLY",
        "relative_path": "NIHMS1928233-supplement-3.csv",
        "source_uri": (
            "https://pmc-oa-opendata.s3.amazonaws.com/PMC10540565.1/"
            "NIHMS1928233-supplement-3.csv"
        ),
        "source_authority": "NCBI_PMC_OA_OPEN_DATA",
        "format": "CSV_UTF8",
        "bytes": 7323186,
        "sha256": "812f3c983cb7c4f473200741ffd6d73bcab911c9e354934542e018e7b0cf8a6d",
        "parsed_for_endpoint_statistics": True,
    },
    {
        "asset_id": "PMC10540565_TABLE_S3",
        "role": "OFFICIAL_PUBLISHED_XTAIL_ENDPOINT_AGGREGATES_ONLY",
        "relative_path": "NIHMS1928233-supplement-4.xlsx",
        "source_uri": (
            "https://pmc-oa-opendata.s3.amazonaws.com/PMC10540565.1/"
            "NIHMS1928233-supplement-4.xlsx"
        ),
        "source_authority": "NCBI_PMC_OA_OPEN_DATA",
        "format": "XLSX_STATIC_VALUES",
        "bytes": 864791,
        "sha256": "ec2aab60fcb0be87f2bcc1b1a5a1f786b23bb429edc9851a4034a3e8983dfa08",
        "parsed_for_endpoint_statistics": True,
    },
    {
        "asset_id": "SOURCE_BUNDLE_PUBLICATION_COMMIT",
        "role": "SOURCE_BUNDLE_TERMINAL_COMMIT_METADATA",
        "relative_path": "PUBLICATION_COMMIT.json",
        "format": "JSON",
        "bytes": 1095,
        "sha256": "4742508195f28bf8c7ab1f7cb8bb0b68c32304f31b19c8f8979d098fa75786a5",
        "parsed_for_endpoint_statistics": False,
    },
    {
        "asset_id": "SOURCE_BUNDLE_SHA256SUMS",
        "role": "SOURCE_BUNDLE_CHECKSUM_METADATA",
        "relative_path": "SHA256SUMS",
        "format": "SHA256SUMS",
        "bytes": 491,
        "sha256": "20da85cd34f0574829392b5de1d7c48cc9782219847f56ccc07dffd579d79f15",
        "parsed_for_endpoint_statistics": False,
    },
    {
        "asset_id": "ZENODO_8007705_V1_2",
        "role": "SOURCE_BUNDLE_CLOSURE_OPAQUE_NEVER_EXECUTED",
        "relative_path": "slschuster_3UTRMutationalMPRA-v1.2.zip",
        "format": "ZIP_OPAQUE_HASH_ONLY_NOT_PARSED",
        "bytes": 46209,
        "sha256": "1c1b1979c1d5bd7fefa54e80a59f982228d0f1498eb0cff2883b753ee5eb0ae4",
        "parsed_for_endpoint_statistics": False,
    },
)

EXPECTED_PARSED_ASSET_IDS = ["PMC10540565_TABLE_S2", "PMC10540565_TABLE_S3"]
EXPECTED_SOURCE_CLOSURE = {
    "exact_member_names": [
        "ASSET_ACQUISITION_MANIFEST.json",
        "NCBI_PRJNA824033_RUNINFO.csv",
        "NIHMS1928233-supplement-3.csv",
        "NIHMS1928233-supplement-4.xlsx",
        "PUBLICATION_COMMIT.json",
        "SHA256SUMS",
        "slschuster_3UTRMutationalMPRA-v1.2.zip",
    ],
    "source_manifest_asset_count": 4,
    "source_checksum_entry_count": 5,
    "source_terminal_marker_committed": True,
    "source_terminal_marker_written_last": True,
    "opaque_zenodo_archive_container_parsed": False,
    "opaque_zenodo_code_executed": False,
}

EXPECTED_S2 = {
    "asset_id": "PMC10540565_TABLE_S2",
    "exact_header": ["ID", "Type", "201bp", "5' End", "3'End", "Full_Oligo"],
    "canonical_compact_header_json_sha256": "8342cace9a613739f9e1e112b1f7fe446def48589e03f45feb6418b2ce8d3412",
    "raw_row_count": 13850,
    "unique_content_row_count": 13836,
    "duplicate_extra_row_count": 14,
    "duplicated_content_group_count": 14,
    "duplicated_content_multiplicity": 2,
    "duplicated_pair_count": 7,
    "raw_type_counts": {"Control": 66, "Mutant": 6892, "WT": 6892},
    "deduplicated_type_counts": {"Control": 66, "Mutant": 6885, "WT": 6885},
    "raw_id_row_multiplicity_counts": {"1": 66, "2": 6878, "4": 7},
    "deduplicated_pair_count": 6885,
    "deduplicated_control_count": 66,
    "pair_rule": "EXACTLY_ONE_WT_AND_ONE_MUTANT_AFTER_EXACT_FULL_ROW_DEDUPLICATION",
    "duplicate_pair_rule": "EXACTLY_SEVEN_PAIR_KEYS_HAVE_TWO_IDENTICAL_COPIES_OF_EACH_ARM",
    "paired_window_length": 201,
    "central_zero_based_index": 100,
    "paired_alphabet": "ACGT",
    "paired_edit_rule": "EXACTLY_ONE_CENTRAL_SNV",
    "central_base_change_counts": {
        "A>C": 238,
        "A>G": 673,
        "A>T": 371,
        "C>A": 636,
        "C>G": 281,
        "C>T": 1261,
        "G>A": 1151,
        "G>C": 288,
        "G>T": 701,
        "T>A": 350,
        "T>C": 669,
        "T>G": 266,
    },
    "pair_id_grammar": "^[^:]+:[1-9][0-9]*_[ACGT]-[ACGT]$",
    "pair_id_grammar_match_count": 6885,
    "all_pair_orientation_counts": {
        "FORWARD": 3497,
        "REVERSE_COMPLEMENT": 3388,
        "UNRESOLVED": 0,
    },
    "distinct_candidate_count": 6885,
    "distinct_wt_201nt_proxy_count": 6882,
    "wt_201nt_proxy_pool_size_counts": {"1": 6879, "2": 3},
    "three_or_more_candidate_wt_201nt_proxy_pool_count": 0,
    "true_a2_dense_pool_count": 0,
}

EXPECTED_S3 = {
    "asset_id": "PMC10540565_TABLE_S3",
    "exact_sheet_names": [
        "S2A_Polysome_MPRA_Mut_Stats",
        "S2B_Poly_MPRA_Control_Stats",
    ],
    "primary_sheet": {
        "name": "S2A_Polysome_MPRA_Mut_Stats",
        "exact_header": [
            "barcode",
            "Gene",
            "Comparison",
            "xtail_log2FC_TE",
            "xtail_pvalue",
            "xtail_FDR",
            "Translation_Sig",
        ],
        "canonical_compact_header_json_sha256": "d204a821928cb76b2fbc29201d3bcd103e6f4d3fa9cc526bd669604d74ef2ea5",
        "data_row_count": 13544,
        "pair_key_count": 6772,
        "comparison_row_counts": {"HighPoly:RNA": 6772, "TotalPoly:RNA": 6772},
        "comparison_count_per_pair": 2,
        "distinct_gene_annotation_count": 1947,
        "comparison_gene_annotation_must_agree": True,
        "statistics_columns": ["xtail_log2FC_TE", "xtail_pvalue", "xtail_FDR"],
        "statistics_missing_token": "NA",
        "statistics_must_be_all_finite_or_all_na_per_row": True,
        "finite_statistic_rows": {"HighPoly:RNA": 6538, "TotalPoly:RNA": 6547},
        "na_statistic_rows": {"HighPoly:RNA": 234, "TotalPoly:RNA": 225},
        "both_comparisons_finite_pair_count": 6538,
        "primary_only_finite_pair_count": 9,
        "secondary_only_finite_pair_count": 0,
        "neither_comparison_finite_pair_count": 225,
        "significant_rows": {"HighPoly:RNA": 58, "TotalPoly:RNA": 174},
        "nonsignificant_rows": {"HighPoly:RNA": 6714, "TotalPoly:RNA": 6598},
        "cell_type_counts": {
            "key_annotation_comparison_string_each": 13544,
            "statistic_numeric_each": 13085,
            "statistic_na_string_each": 459,
            "translation_formula": 13544,
            "translation_cached_string": 13544,
        },
        "translation_formula_execution_allowed": False,
        "cached_translation_counts_role": "DESCRIPTIVE_ONLY_NOT_MEMBERSHIP_OR_GATE",
        "cached_translation_values_used_for_gate": False,
        "cached_translation_values_used_for_membership": False,
        "membership_rule": "ALL_PAIR_KEYS_BEFORE_TRANSLATION_SIGNIFICANCE",
    },
    "control_sheet": {
        "name": "S2B_Poly_MPRA_Control_Stats",
        "exact_header": [
            "barcode",
            "TE_1",
            "TE_2",
            "TE_3",
            "TE_4",
            "TE_5",
            "TE_6",
            "Txn_1",
            "Txn_2",
            "Txn_3",
            "Txn_4",
            "Txn_5",
            "Txn_6",
        ],
        "canonical_compact_header_json_sha256": "9008ea2fd8533da367e9dacad56f7089130574eea3b9829dc3fce2a76ad5d292",
        "data_row_count": 29,
        "finite_measurement_count": 343,
        "na_measurement_count": 5,
        "measurement_counts_role": "FROZEN_DESCRIPTIVE_ONLY_NOT_READ_OR_GATE",
        "data_access_policy": "HEADER_AND_DIMENSIONS_ONLY",
        "data_cells_must_not_be_read": True,
        "role": "OPAQUE_CONTROL_ONLY_EXCLUDED_FROM_MUTATION_ENDPOINT_AND_QUALIFICATION_COUNTS",
    },
    "join": {
        "table_s2_pair_count": 6885,
        "table_s3_pair_count": 6772,
        "joined_pair_count": 6772,
        "table_s2_absent_from_table_s3_count": 113,
        "table_s3_not_in_table_s2_count": 0,
        "joined_pair_orientation_counts": {
            "FORWARD": 3451,
            "REVERSE_COMPLEMENT": 3321,
            "UNRESOLVED": 0,
        },
    },
}

EXPECTED_ENDPOINT_BOUNDARY = {
    "primary_endpoint_id": "TOTAL_POLYSOME_TRANSLATION_EFFICIENCY",
    "primary_comparison_value": "TotalPoly:RNA",
    "primary_membership_pair_count": 6772,
    "primary_finite_effect_pair_count": 6547,
    "primary_absent_pair_count": 113,
    "primary_na_pair_count": 225,
    "primary_total_attrition_count": 338,
    "attrition_equation": "6885_MINUS_113_MINUS_225_EQUALS_6547",
    "primary_complete_distinct_wt_201nt_proxy_group_count": 6544,
    "primary_complete_wt_201nt_proxy_pool_size_counts": {"1": 6541, "2": 3},
    "wt_201nt_grouping_authority": False,
    "wt_201nt_grouping_proxy_only": True,
    "biological_source_group_authority_closed": False,
    "membership_selected_by_significance": False,
    "study_level_reported_biological_replicate_count": 6,
    "row_level_effective_replicate_count": None,
    "row_level_effective_replicate_count_status": "NOT_REPORTED",
    "standard_error": None,
    "standard_error_status": "NOT_REPORTED_IN_TABLE_S3_NOT_DERIVABLE_FROM_PUBLISHED_TABLE",
    "power_effective_n": None,
    "power_effective_n_status": "NOT_ESTABLISHED",
    "effect_geometry": "PAIRWISE_WT_VS_SINGLE_MUTANT_ONLY",
    "true_a2_dense_candidate_count": 0,
    "published_endpoint_is_not_raw_replay": True,
    "published_endpoint_is_not_canonical_materialization": True,
}

EXPECTED_PUBLICATION_CONTRACT = {
    "output_id": OUTPUT_ID,
    "success_json_files": list(SUCCESS_JSON_FILES),
    "failure_json_files": list(FAILURE_JSON_FILES),
    "checksum_file": SHA256SUMS_FILENAME,
    "terminal_marker": PUBLICATION_MARKER,
    "terminal_marker_written_last": True,
    "terminal_publication_operation": "FSYNCED_STAGED_HARDLINK_NO_REPLACE",
    "acceptance_critical_reads_after_terminal_visibility": False,
    "exclusive_no_overwrite": True,
    "aggregate_only": True,
    "raw_rows_or_row_keys_may_be_output": False,
}

BASE_BLOCKERS = (
    "OWNER_POLICY_FOR_PUBLISHED_ENDPOINT_USE_NOT_FROZEN",
    "CHECKPOINT_SPECIFIC_ENDPOINT_USE_NOT_CLEARED",
    "BIOLOGICAL_SOURCE_GROUP_AUTHORITY_NOT_CLOSED",
    "CURRENT_AUTHORITY_80S_BLOCKER_SCOPE_NOT_ROUTED_FOR_PUBLISHED_ENDPOINT_REUSE",
    "OUTCOME_BLIND_SPLIT_AND_LEAKAGE_POLICY_NOT_FROZEN",
    "POWER_AND_CONFIDENCE_INTERVAL_ADEQUACY_NOT_ESTABLISHED",
    "CANONICAL_REPORTED_ENDPOINT_SEMANTICS_NOT_ADJUDICATED",
    "ROW_LEVEL_REPLICATE_AND_STANDARD_ERROR_ADJUDICATION_NOT_CLOSED",
)
IMPLEMENTATION_BINDING_BLOCKER = "IMPLEMENTATION_BINDING_UNKNOWN_NOT_ASSERTED"

PROTOCOL_TOP_LEVEL_KEYS = {
    "contract_id",
    "schema_version",
    "protocol_id",
    "protocol_status",
    "activation_status",
    "qualification_status",
    "dataset_id",
    "authority",
    "implementation_binding",
    "decision_neutral_boundary",
    "scope",
    "input_contract",
    "table_contract",
    "endpoint_boundary",
    "publication_contract",
    "unresolved_blockers",
}

PROHIBITED_OUTPUT_KEYS = {
    "barcode",
    "barcodes",
    "gene",
    "genes",
    "label",
    "labels",
    "pair_id",
    "pair_ids",
    "row_id",
    "row_ids",
    "sequence",
    "sequences",
    "source_sequence",
    "candidate_sequence",
    "raw_rows",
}

_POST_VERIFIED_INPUT_SNAPSHOT_HOOK: Callable[[], None] | None = None
_PUBLICATION_FAULT_HOOK: Callable[[str], None] | None = None


class QualificationError(RuntimeError):
    code = "QUALIFICATION_FAILED"


class ScopeViolation(QualificationError):
    code = "SCOPE_REJECTED_BEFORE_READ"


class ProtocolError(QualificationError):
    code = "PROTOCOL_INVALID"


class InputIntegrityError(QualificationError):
    code = "INPUT_INTEGRITY_FAILED"


class TableAuditError(QualificationError):
    code = "TABLE_AUDIT_FAILED"


class PublicationError(QualificationError):
    code = "PUBLICATION_FAILED"


class PublicationContention(PublicationError):
    code = "OUTPUT_TARGET_EXISTS_NO_OVERWRITE"


class PartialPrecommitError(PublicationError):
    code = "PARTIAL_PRECOMMIT"


FileIdentity = tuple[int, int, int, int, int, int, int]
DirectoryIdentity = tuple[int, int, int]


@dataclass(frozen=True)
class DirectoryBinding:
    path: Path
    fd: int
    identity: DirectoryIdentity
    label: str


@dataclass(frozen=True)
class S2State:
    pair_keys: frozenset[str]
    wt_201nt_by_pair: Mapping[str, str]
    candidate_201nt_by_pair: Mapping[str, str]
    orientation_by_pair: Mapping[str, str]
    aggregate: Mapping[str, Any]


@dataclass(frozen=True)
class S3State:
    pair_keys: frozenset[str]
    finite_pair_keys_by_comparison: Mapping[str, frozenset[str]]
    aggregate: Mapping[str, Any]


def _sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _canonical_header_sha256(header: Sequence[str]) -> str:
    return _sha256_bytes(
        json.dumps(list(header), ensure_ascii=False, separators=(",", ":")).encode(
            "utf-8"
        )
    )


def _pretty_json_bytes(value: Mapping[str, Any]) -> bytes:
    try:
        return (
            json.dumps(
                value,
                ensure_ascii=False,
                sort_keys=True,
                indent=2,
                allow_nan=False,
            )
            + "\n"
        ).encode("utf-8")
    except (TypeError, ValueError) as exc:
        raise PublicationError("output document is not finite canonical JSON") from exc


def _strict_json_object(
    payload: bytes, *, error_type: type[QualificationError], label: str
) -> dict[str, Any]:
    def reject_duplicate_keys(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
        value: dict[str, Any] = {}
        for key, child in pairs:
            if key in value:
                raise ValueError("duplicate JSON key")
            value[key] = child
        return value

    def reject_nonfinite(token: str) -> Any:
        raise ValueError(f"non-finite JSON token {token}")

    try:
        value = json.loads(
            payload.decode("utf-8"),
            object_pairs_hook=reject_duplicate_keys,
            parse_constant=reject_nonfinite,
        )
    except (UnicodeDecodeError, json.JSONDecodeError, ValueError) as exc:
        raise error_type(f"{label} is not strict finite UTF-8 JSON") from exc
    if not isinstance(value, dict):
        raise error_type(f"{label} root is not an object")
    return value


def _expected_blockers(binding: Mapping[str, Any]) -> list[str]:
    blockers = list(BASE_BLOCKERS)
    if binding.get("status") == "UNKNOWN_NOT_ASSERTED":
        blockers.append(IMPLEMENTATION_BINDING_BLOCKER)
    return blockers


def _validate_binding_document(binding: Mapping[str, Any]) -> None:
    if set(binding) != set(IMPLEMENTATION_BINDING_UNKNOWN):
        raise ProtocolError("implementation-binding keys differ from frozen schema")
    for key in (
        "repository_root_rule",
        "qualifier_path",
        "test_path",
        "post_implementation_allowed_changed_paths",
        "binding_commit_must_be_direct_child",
        "current_head_must_strictly_descend",
        "active_authority_must_be_ancestor",
        "clean_worktree_required",
        "running_script_must_match_qualifier_blob",
    ):
        if binding[key] != IMPLEMENTATION_BINDING_UNKNOWN[key]:
            raise ProtocolError("implementation-binding fixed field drifted")
    status = binding.get("status")
    if status == "UNKNOWN_NOT_ASSERTED":
        if dict(binding) != IMPLEMENTATION_BINDING_UNKNOWN:
            raise ProtocolError("unknown implementation binding is not exact")
        return
    if status != "BOUND":
        raise ProtocolError("implementation-binding status is outside enum")
    if COMMIT_RE.fullmatch(str(binding.get("implementation_commit"))) is None:
        raise ProtocolError("bound implementation commit is invalid")
    for key in ("qualifier_blob_sha256", "test_blob_sha256"):
        if SHA256_RE.fullmatch(str(binding.get(key))) is None:
            raise ProtocolError("bound implementation blob hash is invalid")


def _validate_protocol(protocol: Mapping[str, Any]) -> None:
    if set(protocol) != PROTOCOL_TOP_LEVEL_KEYS:
        raise ProtocolError("protocol top-level keys differ from closed schema")
    expected_scalars = {
        "contract_id": CONTRACT_ID,
        "schema_version": SCHEMA_VERSION,
        "protocol_id": PROTOCOL_ID,
        "protocol_status": PROTOCOL_STATUS,
        "activation_status": ACTIVATION_STATUS,
        "qualification_status": QUALIFICATION_STATUS,
        "dataset_id": DATASET_ID,
    }
    for key, expected in expected_scalars.items():
        if protocol.get(key) != expected:
            raise ProtocolError(f"protocol scalar {key} drifted")
    if protocol.get("authority") != EXPECTED_AUTHORITY:
        raise ProtocolError("authority binding drifted")
    binding = protocol.get("implementation_binding")
    if not isinstance(binding, Mapping):
        raise ProtocolError("implementation binding is not an object")
    _validate_binding_document(binding)
    if protocol.get("decision_neutral_boundary") != EXPECTED_DECISION_BOUNDARY:
        raise ProtocolError("decision-neutral boundary drifted")
    if protocol.get("scope") != EXPECTED_SCOPE:
        raise ProtocolError("scope drifted")

    input_contract = protocol.get("input_contract")
    if not isinstance(input_contract, Mapping):
        raise ProtocolError("input contract is not an object")
    expected_input_keys = {
        "data_root",
        "forbidden_path_tokens",
        "same_descriptor_verified_snapshot_required",
        "root_to_leaf_o_nofollow_required",
        "single_link_regular_files_required",
        "all_input_hashes_and_sizes_required_before_parsing",
        "source_bundle_member_count",
        "parsed_scientific_asset_ids",
        "source_bundle_members",
        "source_bundle_closure",
    }
    if set(input_contract) != expected_input_keys:
        raise ProtocolError("input-contract keys drifted")
    if input_contract.get("data_root") != os.fspath(EXPECTED_DATA_ROOT):
        raise ProtocolError("input root drifted")
    if [str(value).casefold() for value in input_contract["forbidden_path_tokens"]] != list(
        FORBIDDEN_PATH_TOKENS
    ):
        raise ProtocolError("forbidden path token set drifted")
    for key in (
        "same_descriptor_verified_snapshot_required",
        "root_to_leaf_o_nofollow_required",
        "single_link_regular_files_required",
        "all_input_hashes_and_sizes_required_before_parsing",
    ):
        if input_contract.get(key) is not True:
            raise ProtocolError("required snapshot guard is disabled")
    if input_contract.get("source_bundle_member_count") != len(EXPECTED_ASSETS):
        raise ProtocolError("source-bundle member count drifted")
    if input_contract.get("parsed_scientific_asset_ids") != EXPECTED_PARSED_ASSET_IDS:
        raise ProtocolError("parsed scientific asset allowlist drifted")
    if input_contract.get("source_bundle_members") != list(EXPECTED_ASSETS):
        raise ProtocolError("source-bundle member binding drifted")
    if input_contract.get("source_bundle_closure") != EXPECTED_SOURCE_CLOSURE:
        raise ProtocolError("source-bundle closure contract drifted")
    for asset in EXPECTED_ASSETS:
        if SHA256_RE.fullmatch(str(asset["sha256"])) is None or not isinstance(
            asset["bytes"], int
        ):
            raise ProtocolError("official asset hash or size is invalid")
        _safe_relative_parts(str(asset["relative_path"]), label="official asset path")

    table_contract = protocol.get("table_contract")
    if table_contract != {"table_s2": EXPECTED_S2, "table_s3": EXPECTED_S3}:
        raise ProtocolError("table contract drifted")
    if protocol.get("endpoint_boundary") != EXPECTED_ENDPOINT_BOUNDARY:
        raise ProtocolError("endpoint boundary drifted")
    if protocol.get("publication_contract") != EXPECTED_PUBLICATION_CONTRACT:
        raise ProtocolError("publication contract drifted")
    if protocol.get("unresolved_blockers") != _expected_blockers(binding):
        raise ProtocolError("blocker list does not match implementation-binding state")


def _git_capture(repository_root: Path, arguments: Sequence[str]) -> bytes:
    try:
        completed = subprocess.run(
            ["git", "-C", os.fspath(repository_root), *arguments],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            check=False,
        )
    except OSError as exc:
        raise ProtocolError("git executable unavailable for binding") from exc
    if completed.returncode != 0:
        raise ProtocolError("git binding query failed closed")
    return completed.stdout


def _git_is_ancestor(repository_root: Path, ancestor: str, descendant: str) -> bool:
    try:
        completed = subprocess.run(
            [
                "git",
                "-C",
                os.fspath(repository_root),
                "merge-base",
                "--is-ancestor",
                ancestor,
                descendant,
            ],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            check=False,
        )
    except OSError as exc:
        raise ProtocolError("git executable unavailable for binding") from exc
    if completed.returncode not in {0, 1}:
        raise ProtocolError("git ancestry query failed closed")
    return completed.returncode == 0


def _validate_i_to_b_protocol_transition(
    implementation_protocol: Mapping[str, Any],
    binding_protocol: Mapping[str, Any],
    binding: Mapping[str, Any],
) -> None:
    if implementation_protocol.get("implementation_binding") != (
        IMPLEMENTATION_BINDING_UNKNOWN
    ):
        raise ProtocolError("implementation commit protocol is not exact UNKNOWN-I")
    if binding_protocol.get("implementation_binding") != dict(binding):
        raise ProtocolError("binding commit protocol does not contain the running binding")
    if implementation_protocol.get("unresolved_blockers") != [
        *BASE_BLOCKERS,
        IMPLEMENTATION_BINDING_BLOCKER,
    ]:
        raise ProtocolError("implementation commit blockers are not exact UNKNOWN-I")
    if binding_protocol.get("unresolved_blockers") != list(BASE_BLOCKERS):
        raise ProtocolError("binding commit scientific blockers drifted")
    implementation_core = dict(implementation_protocol)
    binding_core = dict(binding_protocol)
    implementation_core.pop("implementation_binding", None)
    binding_core.pop("implementation_binding", None)
    implementation_core.pop("unresolved_blockers", None)
    binding_core.pop("unresolved_blockers", None)
    if implementation_core != binding_core:
        raise ProtocolError(
            "config-only B changed protocol core outside binding-derived fields"
        )


def _verify_implementation_binding(
    binding: Mapping[str, Any],
    authority: Mapping[str, Any],
    repository_root: Path,
    *,
    running_script_path: Path | None = None,
) -> dict[str, Any]:
    """Verify UNKNOWN-I or the future direct config-only B commit."""

    _validate_binding_document(binding)
    if binding["status"] == "UNKNOWN_NOT_ASSERTED":
        return {
            "status": "UNKNOWN_NOT_ASSERTED",
            "verified": False,
            "implementation_commit": "UNKNOWN_NOT_ASSERTED",
            "binding_commit": "UNKNOWN_NOT_ASSERTED",
            "clean_worktree": "UNKNOWN_NOT_ASSERTED",
            "config_only_direct_child": "UNKNOWN_NOT_ASSERTED",
            "authority_blobs_match": "UNKNOWN_NOT_ASSERTED",
            "implementation_blobs_match": "UNKNOWN_NOT_ASSERTED",
            "running_script_matches_bound_blob": "UNKNOWN_NOT_ASSERTED",
        }

    root = _absolute_without_resolving(repository_root)
    current_head = _git_capture(root, ["rev-parse", "HEAD"]).decode("ascii").strip()
    if COMMIT_RE.fullmatch(current_head) is None:
        raise ProtocolError("current Git HEAD is invalid")
    if _git_capture(root, ["status", "--porcelain=v1"]):
        raise ProtocolError("bound execution requires a clean worktree")
    active = str(authority["active_authority_commit"])
    staging_parent = str(authority["staging_parent_head"])
    implementation = str(binding["implementation_commit"])
    if active == staging_parent or not _git_is_ancestor(root, active, staging_parent):
        raise ProtocolError("active authority is not a strict staging-parent ancestor")
    if staging_parent == implementation or not _git_is_ancestor(
        root, staging_parent, implementation
    ):
        raise ProtocolError("staging parent is not a strict implementation ancestor")
    parents = _git_capture(root, ["rev-list", "--parents", "-n", "1", current_head]).decode(
        "ascii"
    ).split()
    if len(parents) != 2 or parents[1] != implementation:
        raise ProtocolError("binding commit is not the direct implementation child")
    changed = _git_capture(
        root, ["diff", "--name-only", implementation, current_head, "--"]
    ).decode("utf-8").splitlines()
    if changed != list(binding["post_implementation_allowed_changed_paths"]):
        raise ProtocolError("binding commit changes are not exactly config-only")
    protocol_relative = str(binding["post_implementation_allowed_changed_paths"][0])
    implementation_protocol_blob = _git_capture(
        root, ["show", f"{implementation}:{protocol_relative}"]
    )
    binding_protocol_blob = _git_capture(
        root, ["show", f"{current_head}:{protocol_relative}"]
    )
    implementation_protocol = _strict_json_object(
        implementation_protocol_blob,
        error_type=ProtocolError,
        label="implementation-commit protocol",
    )
    binding_protocol = _strict_json_object(
        binding_protocol_blob,
        error_type=ProtocolError,
        label="binding-commit protocol",
    )
    _validate_i_to_b_protocol_transition(
        implementation_protocol, binding_protocol, binding
    )

    for path_key, hash_key in (
        ("contract_path", "contract_sha256"),
        ("data_role_registry_path", "data_role_registry_sha256"),
        ("decision_log_path", "decision_log_sha256"),
    ):
        blob = _git_capture(root, ["show", f"{active}:{authority[path_key]}"])
        if _sha256_bytes(blob) != authority[hash_key]:
            raise ProtocolError("active-authority blob differs from frozen hash")
    implementation_blobs: dict[str, bytes] = {}
    for path_key, hash_key in (
        ("qualifier_path", "qualifier_blob_sha256"),
        ("test_path", "test_blob_sha256"),
    ):
        blob = _git_capture(root, ["show", f"{implementation}:{binding[path_key]}"])
        if _sha256_bytes(blob) != binding[hash_key]:
            raise ProtocolError("implementation blob differs from frozen hash")
        implementation_blobs[path_key] = blob
    script_path = Path(__file__) if running_script_path is None else running_script_path
    running_bytes, _ = _read_path_verified_snapshot(
        script_path,
        label="running qualifier",
        expected_sha256=str(binding["qualifier_blob_sha256"]),
    )
    if running_bytes != implementation_blobs["qualifier_path"]:
        raise ProtocolError("running qualifier differs from implementation commit")
    return {
        "status": "PASS_BOUND_IMPLEMENTATION",
        "verified": True,
        "implementation_commit": implementation,
        "binding_commit": current_head,
        "clean_worktree": True,
        "config_only_direct_child": True,
        "authority_blobs_match": True,
        "implementation_blobs_match": True,
        "running_script_matches_bound_blob": True,
    }


def _reject_forbidden_path(path: Path | str, *, label: str) -> None:
    text = os.fspath(path).casefold()
    hits = sorted(token for token in FORBIDDEN_PATH_TOKENS if token in text)
    if hits:
        raise ScopeViolation(
            f"{label} rejected before read because a forbidden path token was present"
        )


def _absolute_without_resolving(path: Path | str) -> Path:
    candidate = Path(path).expanduser()
    if candidate.is_absolute():
        return Path(os.path.normpath(os.fspath(candidate)))
    return Path(os.path.abspath(os.fspath(candidate)))


def _safe_basename(name: str, *, label: str) -> str:
    if (
        not isinstance(name, str)
        or not name
        or Path(name).name != name
        or name in {".", ".."}
    ):
        raise ScopeViolation(f"{label} is not a safe basename")
    _reject_forbidden_path(name, label=label)
    return name


def _safe_relative_parts(value: str, *, label: str) -> tuple[str, ...]:
    if not isinstance(value, str) or not value:
        raise ScopeViolation(f"{label} is not a nonempty relative path")
    _reject_forbidden_path(value, label=label)
    pure = PurePosixPath(value)
    if pure.is_absolute() or not pure.parts or any(
        part in {"", ".", ".."} for part in pure.parts
    ):
        raise ScopeViolation(f"{label} is not a safe relative path")
    return tuple(_safe_basename(part, label=label) for part in pure.parts)


def _file_identity(info: os.stat_result) -> FileIdentity:
    return (
        info.st_dev,
        info.st_ino,
        info.st_mode,
        info.st_nlink,
        info.st_size,
        info.st_mtime_ns,
        info.st_ctime_ns,
    )


def _directory_identity(info: os.stat_result) -> DirectoryIdentity:
    return (info.st_dev, info.st_ino, info.st_mode)


def _open_directory_no_symlinks(path: Path | str, *, label: str) -> DirectoryBinding:
    absolute = _absolute_without_resolving(path)
    _reject_forbidden_path(absolute, label=label)
    nofollow = getattr(os, "O_NOFOLLOW", None)
    directory = getattr(os, "O_DIRECTORY", None)
    if nofollow is None or directory is None:
        raise ScopeViolation(f"{label} requires O_NOFOLLOW and O_DIRECTORY")
    flags = os.O_RDONLY | nofollow | directory | getattr(os, "O_CLOEXEC", 0)
    try:
        descriptor = os.open(absolute.anchor, flags)
    except OSError as exc:
        raise ScopeViolation(f"{label} filesystem root cannot be opened safely") from exc
    try:
        for component in absolute.parts[1:]:
            _safe_basename(component, label=label)
            try:
                next_descriptor = os.open(component, flags, dir_fd=descriptor)
            except OSError as exc:
                raise ScopeViolation(f"{label} contains a symlink or non-directory") from exc
            opened = os.fstat(next_descriptor)
            if not stat.S_ISDIR(opened.st_mode):
                os.close(next_descriptor)
                raise ScopeViolation(f"{label} component is not a directory")
            os.close(descriptor)
            descriptor = next_descriptor
        opened = os.fstat(descriptor)
        if not stat.S_ISDIR(opened.st_mode):
            raise ScopeViolation(f"{label} descriptor is not a directory")
        return DirectoryBinding(
            path=absolute,
            fd=descriptor,
            identity=_directory_identity(opened),
            label=label,
        )
    except Exception:
        os.close(descriptor)
        raise


def _assert_directory_binding(binding: DirectoryBinding) -> None:
    current = os.fstat(binding.fd)
    if not stat.S_ISDIR(current.st_mode) or _directory_identity(current) != binding.identity:
        raise ScopeViolation(f"{binding.label} descriptor identity changed")
    reopened = _open_directory_no_symlinks(
        binding.path, label=f"{binding.label} namespace rebind"
    )
    try:
        if reopened.identity != binding.identity:
            raise ScopeViolation(f"{binding.label} namespace binding changed")
    finally:
        os.close(reopened.fd)


def _open_relative_parent(root: DirectoryBinding, parts: Sequence[str], *, label: str) -> int:
    descriptor = os.dup(root.fd)
    flags = (
        os.O_RDONLY
        | getattr(os, "O_CLOEXEC", 0)
        | getattr(os, "O_DIRECTORY", 0)
        | getattr(os, "O_NOFOLLOW", 0)
    )
    try:
        for component in parts:
            try:
                next_descriptor = os.open(component, flags, dir_fd=descriptor)
            except OSError as exc:
                raise ScopeViolation(f"{label} parent cannot be opened safely") from exc
            opened = os.fstat(next_descriptor)
            if not stat.S_ISDIR(opened.st_mode):
                os.close(next_descriptor)
                raise ScopeViolation(f"{label} parent is not a directory")
            os.close(descriptor)
            descriptor = next_descriptor
        return descriptor
    except Exception:
        os.close(descriptor)
        raise


def _open_regular_leaf_at(
    parent_fd: int,
    leaf: str,
    *,
    label: str,
    require_single_link: bool,
) -> tuple[int, os.stat_result]:
    flags = (
        os.O_RDONLY
        | getattr(os, "O_CLOEXEC", 0)
        | getattr(os, "O_NOFOLLOW", 0)
        | getattr(os, "O_NONBLOCK", 0)
    )
    try:
        descriptor = os.open(leaf, flags, dir_fd=parent_fd)
    except OSError as exc:
        raise ScopeViolation(f"{label} cannot be opened safely") from exc
    try:
        opened = os.fstat(descriptor)
        if not stat.S_ISREG(opened.st_mode):
            raise ScopeViolation(f"{label} is not a regular file")
        if require_single_link and opened.st_nlink != 1:
            raise ScopeViolation(f"{label} is not a single-link snapshot")
    except Exception:
        os.close(descriptor)
        raise
    return descriptor, opened


def _assert_leaf_binding(
    parent_fd: int,
    leaf: str,
    identity: FileIdentity,
    *,
    label: str,
    require_single_link: bool,
) -> None:
    descriptor, reopened = _open_regular_leaf_at(
        parent_fd,
        leaf,
        label=label,
        require_single_link=require_single_link,
    )
    try:
        if _file_identity(reopened) != identity:
            raise ScopeViolation(f"{label} namespace binding changed")
    finally:
        os.close(descriptor)


def _read_relative_verified_snapshot(
    root: DirectoryBinding,
    relative_path: str,
    *,
    label: str,
    expected_sha256: str | None = None,
    expected_bytes: int | None = None,
) -> tuple[bytes, dict[str, Any], FileIdentity]:
    parts = _safe_relative_parts(relative_path, label=label)
    parent_fd = _open_relative_parent(root, parts[:-1], label=label)
    leaf = parts[-1]
    digest = hashlib.sha256()
    try:
        descriptor, opened = _open_regular_leaf_at(
            parent_fd, leaf, label=label, require_single_link=True
        )
        identity = _file_identity(opened)
        chunks: list[bytes] = []
        try:
            while True:
                block = os.read(descriptor, 1 << 20)
                if not block:
                    break
                digest.update(block)
                chunks.append(block)
            final = os.fstat(descriptor)
            if _file_identity(final) != identity:
                raise ScopeViolation(f"{label} changed during same-descriptor capture")
            payload = b"".join(chunks)
            if len(payload) != opened.st_size:
                raise ScopeViolation(f"{label} byte count changed during capture")
        finally:
            os.close(descriptor)
        _assert_leaf_binding(
            parent_fd,
            leaf,
            identity,
            label=f"{label} post-capture",
            require_single_link=True,
        )
    finally:
        os.close(parent_fd)
    observed_sha256 = digest.hexdigest()
    if expected_sha256 is not None and observed_sha256 != expected_sha256:
        raise InputIntegrityError(f"{label} SHA-256 mismatch")
    if expected_bytes is not None and len(payload) != expected_bytes:
        raise InputIntegrityError(f"{label} byte count mismatch")
    return payload, {"sha256": observed_sha256, "bytes": len(payload)}, identity


def _assert_relative_identity(
    root: DirectoryBinding,
    relative_path: str,
    identity: FileIdentity,
    *,
    label: str,
) -> None:
    parts = _safe_relative_parts(relative_path, label=label)
    parent_fd = _open_relative_parent(root, parts[:-1], label=label)
    try:
        _assert_leaf_binding(
            parent_fd,
            parts[-1],
            identity,
            label=label,
            require_single_link=True,
        )
    finally:
        os.close(parent_fd)


def _read_path_verified_snapshot(
    path: Path | str,
    *,
    label: str,
    expected_sha256: str | None = None,
) -> tuple[bytes, dict[str, Any]]:
    absolute = _absolute_without_resolving(path)
    parent = _open_directory_no_symlinks(absolute.parent, label=f"{label} parent")
    try:
        payload, provenance, identity = _read_relative_verified_snapshot(
            parent,
            absolute.name,
            label=label,
            expected_sha256=expected_sha256,
        )
        _assert_relative_identity(parent, absolute.name, identity, label=label)
        _assert_directory_binding(parent)
        return payload, provenance
    finally:
        os.close(parent.fd)


def _preflight_paths_before_read(
    protocol_path: Path | str,
    data_root: Path | str,
    output_directory: Path | str,
) -> tuple[Path, Path, Path]:
    raw = (
        (Path(protocol_path), "protocol path"),
        (Path(data_root), "ordinary public data root"),
        (Path(output_directory), "output path"),
    )
    # This loop deliberately precedes expanduser, stat, resolve, open and parse.
    for path, label in raw:
        _reject_forbidden_path(path, label=label)
    absolute = tuple(_absolute_without_resolving(path) for path, _ in raw)
    for path, (_, label) in zip(absolute, raw):
        _reject_forbidden_path(path, label=label)
    protocol, root, output = absolute
    if protocol.name != PROTOCOL_BASENAME:
        raise ScopeViolation("protocol basename is outside the frozen allowlist")
    if root != _absolute_without_resolving(EXPECTED_DATA_ROOT):
        raise ScopeViolation("data root differs from the frozen official-table root")
    try:
        output.relative_to(root)
    except ValueError:
        pass
    else:
        raise ScopeViolation("output path overlaps the input root")
    if output in {protocol, protocol.parent, root}:
        raise ScopeViolation("output path overlaps an authority or input path")
    return protocol, root, output


def _load_protocol(path: Path, expected_sha256: str) -> tuple[dict[str, Any], dict[str, Any]]:
    if SHA256_RE.fullmatch(expected_sha256) is None:
        raise ProtocolError("launch protocol SHA-256 is invalid")
    payload, provenance = _read_path_verified_snapshot(
        path, label="published-endpoint protocol", expected_sha256=expected_sha256
    )
    value = _strict_json_object(
        payload, error_type=ProtocolError, label="published-endpoint protocol"
    )
    _validate_protocol(value)
    return value, {
        **provenance,
        "launch_expected_sha256": expected_sha256,
    }


def _audit_source_bundle_closure(
    payloads: Mapping[str, bytes], input_contract: Mapping[str, Any]
) -> dict[str, Any]:
    if set(payloads) != {asset["asset_id"] for asset in EXPECTED_ASSETS}:
        raise InputIntegrityError("source-bundle captured member set differs")
    by_id = {asset["asset_id"]: asset for asset in EXPECTED_ASSETS}
    checksum_ids = {
        "SOURCE_BUNDLE_MANIFEST",
        "NCBI_PRJNA824033_RUNINFO",
        "PMC10540565_TABLE_S2",
        "PMC10540565_TABLE_S3",
        "ZENODO_8007705_V1_2",
    }
    expected_checksum_lines = [
        f"{by_id[asset_id]['sha256']}  {by_id[asset_id]['relative_path']}"
        for asset_id in sorted(
            checksum_ids, key=lambda value: str(by_id[value]["relative_path"])
        )
    ]
    try:
        checksum_lines = payloads["SOURCE_BUNDLE_SHA256SUMS"].decode("ascii").splitlines()
    except UnicodeDecodeError as exc:
        raise InputIntegrityError("source-bundle SHA256SUMS is not ASCII") from exc
    if checksum_lines != expected_checksum_lines:
        raise InputIntegrityError("source-bundle SHA256SUMS closure differs")

    source_manifest = _strict_json_object(
        payloads["SOURCE_BUNDLE_MANIFEST"],
        error_type=InputIntegrityError,
        label="source-bundle acquisition manifest",
    )
    source_marker = _strict_json_object(
        payloads["SOURCE_BUNDLE_PUBLICATION_COMMIT"],
        error_type=InputIntegrityError,
        label="source-bundle terminal marker",
    )
    if (
        source_manifest.get("record_type")
        != "ROUTE_A_V3_PUBLIC_ASSET_ACQUISITION_MANIFEST"
        or source_manifest.get("contract_id") != CONTRACT_ID
        or source_manifest.get("dataset_id") != DATASET_ID
        or source_manifest.get("status") != "ASSETS_ACQUIRED_NOT_QUALIFIED"
    ):
        raise InputIntegrityError("source acquisition manifest identity differs")
    source_assets = source_manifest.get("assets")
    if not isinstance(source_assets, list) or len(source_assets) != input_contract[
        "source_bundle_closure"
    ]["source_manifest_asset_count"]:
        raise InputIntegrityError("source acquisition manifest asset count differs")
    expected_manifest_assets = {
        by_id[asset_id]["relative_path"]: (
            by_id[asset_id]["bytes"],
            by_id[asset_id]["sha256"],
        )
        for asset_id in (
            "NCBI_PRJNA824033_RUNINFO",
            "PMC10540565_TABLE_S2",
            "PMC10540565_TABLE_S3",
            "ZENODO_8007705_V1_2",
        )
    }
    observed_manifest_assets: dict[str, tuple[int, str]] = {}
    for entry in source_assets:
        if not isinstance(entry, Mapping):
            raise InputIntegrityError("source acquisition asset entry is invalid")
        filename = entry.get("filename")
        observed_manifest_assets[str(filename)] = (
            entry.get("bytes"),
            entry.get("sha256"),
        )
    if observed_manifest_assets != expected_manifest_assets:
        raise InputIntegrityError("source acquisition manifest hashes or sizes differ")
    scientific = source_manifest.get("scientific_boundaries")
    if not isinstance(scientific, Mapping) or any(
        scientific.get(key) != expected
        for key, expected in {
            "ordinary_study_contribution": 0,
            "a1_study_contribution": 0,
            "true_a2_study_contribution": 0,
            "qualified": False,
            "training_started": False,
            "next_phase_authorized": False,
        }.items()
    ):
        raise InputIntegrityError("source acquisition scientific boundary differs")

    expected_marker_members = sorted(
        set(input_contract["source_bundle_closure"]["exact_member_names"])
        - {"PUBLICATION_COMMIT.json"}
    )
    marker_required = {
        "record_type": "ROUTE_A_V3_PUBLIC_ASSET_ACQUISITION_COMMIT",
        "contract_id": CONTRACT_ID,
        "dataset_id": DATASET_ID,
        "intended_final_path": os.fspath(EXPECTED_DATA_ROOT),
        "member_files": expected_marker_members,
        "member_file_count": 6,
        "sha256sums_sha256": by_id["SOURCE_BUNDLE_SHA256SUMS"]["sha256"],
        "commit_marker_written_last": True,
        "committed": True,
        "scientific_status": "ASSETS_ACQUIRED_NOT_QUALIFIED",
        "ordinary_study_contribution": 0,
        "a1_study_contribution": 0,
        "true_a2_study_contribution": 0,
        "training_started": False,
        "next_phase_authorized": False,
    }
    if any(source_marker.get(key) != expected for key, expected in marker_required.items()):
        raise InputIntegrityError("source terminal marker closure differs")
    return {
        "status": "PASS_EXACT_SEVEN_MEMBER_SOURCE_BUNDLE_CLOSURE",
        "exact_member_count": len(payloads),
        "checksum_entry_count": len(checksum_lines),
        "source_manifest_asset_count": len(source_assets),
        "terminal_marker_committed": True,
        "terminal_marker_written_last": True,
        "endpoint_statistics_parsed_asset_count": len(EXPECTED_PARSED_ASSET_IDS),
        "opaque_zenodo_member_byte_verified": True,
        "opaque_zenodo_archive_container_parsed": False,
        "opaque_zenodo_code_executed": False,
        "raw_reads_or_alignments_opened": False,
    }


def _audit_table_s2(payload: bytes, spec: Mapping[str, Any]) -> S2State:
    try:
        text = payload.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise TableAuditError("Table S2 is not UTF-8 CSV") from exc
    reader = csv.reader(io.StringIO(text, newline=""), strict=True)
    try:
        header = next(reader)
    except StopIteration as exc:
        raise TableAuditError("Table S2 is empty") from exc
    if header != spec["exact_header"]:
        raise TableAuditError("Table S2 header differs from frozen header")
    if _canonical_header_sha256(header) != spec["canonical_compact_header_json_sha256"]:
        raise TableAuditError("Table S2 header hash differs from frozen hash")
    rows: list[tuple[str, ...]] = []
    for row in reader:
        if len(row) != len(header) or any("\x00" in value for value in row):
            raise TableAuditError("Table S2 row geometry is invalid")
        if any(value == "" for value in row):
            raise TableAuditError("Table S2 contains an empty required field")
        rows.append(tuple(row))
    if len(rows) != spec["raw_row_count"]:
        raise TableAuditError("Table S2 raw row count differs from frozen count")

    content_counts = Counter(rows)
    unique_rows = list(content_counts)
    multiplicities = Counter(content_counts.values())
    duplicate_extra = sum(count - 1 for count in content_counts.values())
    if len(unique_rows) != spec["unique_content_row_count"]:
        raise TableAuditError("Table S2 unique-content count differs")
    if duplicate_extra != spec["duplicate_extra_row_count"]:
        raise TableAuditError("Table S2 duplicate-extra count differs")
    if sum(count > 1 for count in content_counts.values()) != spec[
        "duplicated_content_group_count"
    ]:
        raise TableAuditError("Table S2 duplicated-content group count differs")
    expected_multiplicities = {
        1: spec["unique_content_row_count"] - spec["duplicated_content_group_count"],
        spec["duplicated_content_multiplicity"]: spec["duplicated_content_group_count"],
    }
    if dict(multiplicities) != expected_multiplicities:
        raise TableAuditError("Table S2 duplicate multiplicities differ")

    raw_types = Counter(row[1] for row in rows)
    unique_types = Counter(row[1] for row in unique_rows)
    if dict(raw_types) != spec["raw_type_counts"]:
        raise TableAuditError("Table S2 raw type counts differ")
    if dict(unique_types) != spec["deduplicated_type_counts"]:
        raise TableAuditError("Table S2 deduplicated type counts differ")
    if set(raw_types) != {"WT", "Mutant", "Control"}:
        raise TableAuditError("Table S2 contains an unexpected row type")

    raw_by_key: dict[str, list[tuple[str, ...]]] = defaultdict(list)
    unique_by_key: dict[str, list[tuple[str, ...]]] = defaultdict(list)
    for row in rows:
        raw_by_key[row[0]].append(row)
    for row in unique_rows:
        unique_by_key[row[0]].append(row)
    raw_key_hist = Counter(len(group) for group in raw_by_key.values())
    if {str(key): value for key, value in sorted(raw_key_hist.items())} != spec[
        "raw_id_row_multiplicity_counts"
    ]:
        raise TableAuditError("Table S2 raw key multiplicity reconciliation differs")

    duplicated_pairs = 0
    for group in raw_by_key.values():
        if len(group) != 4:
            continue
        group_counts = Counter(group)
        if (
            len(group_counts) != 2
            or set(group_counts.values()) != {2}
            or Counter(row[1] for row in group_counts) != {"WT": 1, "Mutant": 1}
        ):
            raise TableAuditError("Table S2 duplicate-pair reconciliation is not exact")
        duplicated_pairs += 1
    if duplicated_pairs != spec["duplicated_pair_count"]:
        raise TableAuditError("Table S2 duplicated-pair count differs")

    pair_rows: dict[str, dict[str, tuple[str, ...]]] = {}
    control_count = 0
    for key, group in unique_by_key.items():
        types = Counter(row[1] for row in group)
        if types == {"Control": 1}:
            control_count += 1
            continue
        if types != {"WT": 1, "Mutant": 1} or len(group) != 2:
            raise TableAuditError("Table S2 pair is not exactly WT plus Mutant")
        pair_rows[key] = {row[1]: row for row in group}
    if len(pair_rows) != spec["deduplicated_pair_count"]:
        raise TableAuditError("Table S2 deduplicated pair count differs")
    if control_count != spec["deduplicated_control_count"]:
        raise TableAuditError("Table S2 control count differs")

    base_changes: Counter[str] = Counter()
    pair_orientations: Counter[str] = Counter()
    orientation_by_pair: dict[str, str] = {}
    grammar_match_count = 0
    candidate_201nt: set[str] = set()
    candidate_201nt_by_pair: dict[str, str] = {}
    wt_201nt_by_pair: dict[str, str] = {}
    proxy_pools: dict[str, set[str]] = defaultdict(set)
    expected_length = int(spec["paired_window_length"])
    central = int(spec["central_zero_based_index"])
    alphabet = set(str(spec["paired_alphabet"]))
    for key, arms in pair_rows.items():
        match = PAIR_ID_RE.fullmatch(key)
        if match is None:
            raise TableAuditError("Table S2 intervention key does not match exact grammar")
        grammar_match_count += 1
        wt = arms["WT"][2]
        mutant = arms["Mutant"][2]
        if len(wt) != expected_length or len(mutant) != expected_length:
            raise TableAuditError("Table S2 paired window length differs")
        if set(wt) - alphabet or set(mutant) - alphabet:
            raise TableAuditError("Table S2 paired window alphabet differs")
        differences = [index for index, bases in enumerate(zip(wt, mutant)) if bases[0] != bases[1]]
        if differences != [central]:
            raise TableAuditError("Table S2 pair is not exactly one central SNV")
        base_change = f"{wt[central]}>{mutant[central]}"
        base_changes[base_change] += 1
        ref_alt = key.rsplit("_", 1)[1]
        ref = ref_alt[0]
        alt = ref_alt[2]
        forward = wt[central] == ref and mutant[central] == alt
        reverse = (
            wt[central] == BASE_COMPLEMENT[ref]
            and mutant[central] == BASE_COMPLEMENT[alt]
        )
        if forward == reverse:
            raise TableAuditError("Table S2 pair orientation is unresolved or ambiguous")
        pair_orientation = "FORWARD" if forward else "REVERSE_COMPLEMENT"
        pair_orientations[pair_orientation] += 1
        orientation_by_pair[key] = pair_orientation
        candidate_201nt.add(mutant)
        candidate_201nt_by_pair[key] = mutant
        wt_201nt_by_pair[key] = wt
        proxy_pools[wt].add(mutant)
    if spec["pair_id_grammar"] != PAIR_ID_RE.pattern:
        raise TableAuditError("Table S2 pair-key grammar contract differs")
    if grammar_match_count != spec["pair_id_grammar_match_count"]:
        raise TableAuditError("Table S2 pair-key grammar match count differs")
    observed_pair_orientations = {
        "FORWARD": pair_orientations["FORWARD"],
        "REVERSE_COMPLEMENT": pair_orientations["REVERSE_COMPLEMENT"],
        "UNRESOLVED": 0,
    }
    if observed_pair_orientations != spec["all_pair_orientation_counts"]:
        raise TableAuditError("Table S2 pair orientation aggregate differs")
    if dict(sorted(base_changes.items())) != spec["central_base_change_counts"]:
        raise TableAuditError("Table S2 central base-change aggregate differs")
    if len(candidate_201nt) != spec["distinct_candidate_count"]:
        raise TableAuditError("Table S2 distinct-candidate count differs")
    if len(proxy_pools) != spec["distinct_wt_201nt_proxy_count"]:
        raise TableAuditError("Table S2 WT 201-nt proxy count differs")
    pool_hist = Counter(len(candidates) for candidates in proxy_pools.values())
    if {str(key): value for key, value in sorted(pool_hist.items())} != spec[
        "wt_201nt_proxy_pool_size_counts"
    ]:
        raise TableAuditError("Table S2 WT proxy pool-size aggregate differs")
    dense_count = sum(size >= 3 for size in map(len, proxy_pools.values()))
    if dense_count != spec["three_or_more_candidate_wt_201nt_proxy_pool_count"]:
        raise TableAuditError("Table S2 dense proxy-pool count differs")

    aggregate = {
        "status": "PASS_EXACT_DUPLICATE_PAIR_RECONCILIATION",
        "raw_row_count": len(rows),
        "unique_content_row_count": len(unique_rows),
        "duplicate_extra_row_count": duplicate_extra,
        "duplicated_content_group_count": sum(
            count > 1 for count in content_counts.values()
        ),
        "duplicated_pair_count": duplicated_pairs,
        "deduplicated_pair_count": len(pair_rows),
        "deduplicated_control_count": control_count,
        "exactly_one_central_snv_pair_count": sum(base_changes.values()),
        "central_zero_based_index": central,
        "central_base_change_counts": dict(sorted(base_changes.items())),
        "pair_id_grammar_match_count": grammar_match_count,
        "all_pair_orientation_counts": observed_pair_orientations,
        "distinct_candidate_count": len(candidate_201nt),
        "distinct_wt_201nt_proxy_count": len(proxy_pools),
        "wt_201nt_proxy_pool_size_counts": {
            str(key): value for key, value in sorted(pool_hist.items())
        },
        "wt_201nt_grouping_authority": False,
        "wt_201nt_grouping_proxy_only": True,
        "three_or_more_candidate_wt_201nt_proxy_pool_count": dense_count,
        "true_a2_dense_pool_count": 0,
        "header_sha256": spec["canonical_compact_header_json_sha256"],
    }
    return S2State(
        pair_keys=frozenset(pair_rows),
        wt_201nt_by_pair=wt_201nt_by_pair,
        candidate_201nt_by_pair=candidate_201nt_by_pair,
        orientation_by_pair=orientation_by_pair,
        aggregate=aggregate,
    )


def _is_finite_number(value: Any) -> bool:
    return (
        isinstance(value, (int, float))
        and not isinstance(value, bool)
        and math.isfinite(float(value))
    )


def _audit_table_s3(payload: bytes, spec: Mapping[str, Any]) -> S3State:
    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            names = [name.casefold() for name in archive.namelist()]
    except (OSError, zipfile.BadZipFile) as exc:
        raise TableAuditError("Table S3 is not a valid XLSX archive") from exc
    if any(
        name.endswith("vbaproject.bin") or name.startswith("xl/externallinks/")
        for name in names
    ):
        raise TableAuditError("Table S3 contains disallowed active or external content")
    try:
        workbook = load_workbook(
            io.BytesIO(payload), read_only=True, data_only=True, keep_links=False
        )
        formula_workbook = load_workbook(
            io.BytesIO(payload), read_only=True, data_only=False, keep_links=False
        )
    except Exception as exc:
        raise TableAuditError("Table S3 workbook cannot be parsed") from exc
    try:
        if workbook.sheetnames != spec["exact_sheet_names"]:
            raise TableAuditError("Table S3 sheet names differ from frozen list")
        if formula_workbook.sheetnames != spec["exact_sheet_names"]:
            raise TableAuditError("Table S3 formula-view sheet names differ")
        primary_spec = spec["primary_sheet"]
        cell_spec = primary_spec["cell_type_counts"]
        formula_primary = formula_workbook[primary_spec["name"]]
        formula_rows = formula_primary.iter_rows()
        try:
            formula_header = [cell.value for cell in next(formula_rows)]
        except StopIteration as exc:
            raise TableAuditError("Table S3 formula-view primary sheet is empty") from exc
        if formula_header != primary_spec["exact_header"]:
            raise TableAuditError("Table S3 formula-view primary header differs")
        formula_data_rows = 0
        leading_string_counts = [0, 0, 0]
        statistic_numeric_counts = [0, 0, 0]
        statistic_na_string_counts = [0, 0, 0]
        translation_formula_count = 0
        for cells in formula_rows:
            row = list(cells)
            if len(row) != len(formula_header):
                raise TableAuditError("Table S3 formula-view row width differs")
            formula_data_rows += 1
            for index in range(3):
                if row[index].data_type == "s" and isinstance(row[index].value, str):
                    leading_string_counts[index] += 1
                else:
                    raise TableAuditError("Table S3 leading cell type differs")
            for offset, index in enumerate(range(3, 6)):
                if row[index].data_type == "n" and _is_finite_number(row[index].value):
                    statistic_numeric_counts[offset] += 1
                elif row[index].data_type == "s" and row[index].value == "NA":
                    statistic_na_string_counts[offset] += 1
                else:
                    raise TableAuditError("Table S3 statistic cell type differs")
            if row[6].data_type != "f":
                raise TableAuditError("Table S3 translation cell is not a formula cell")
            translation_formula_count += 1
        if formula_data_rows != primary_spec["data_row_count"]:
            raise TableAuditError("Table S3 formula-view row count differs")
        if leading_string_counts != [
            cell_spec["key_annotation_comparison_string_each"]
        ] * 3:
            raise TableAuditError("Table S3 leading string-cell counts differ")
        if statistic_numeric_counts != [cell_spec["statistic_numeric_each"]] * 3:
            raise TableAuditError("Table S3 numeric-cell counts differ")
        if statistic_na_string_counts != [
            cell_spec["statistic_na_string_each"]
        ] * 3:
            raise TableAuditError("Table S3 NA string-cell counts differ")
        if translation_formula_count != cell_spec["translation_formula"]:
            raise TableAuditError("Table S3 translation formula count differs")

        primary = workbook[primary_spec["name"]]
        primary_rows = primary.iter_rows(values_only=True)
        try:
            primary_header = list(next(primary_rows))
        except StopIteration as exc:
            raise TableAuditError("Table S3 primary sheet is empty") from exc
        if primary_header != primary_spec["exact_header"]:
            raise TableAuditError("Table S3 primary header differs")
        if _canonical_header_sha256(primary_header) != primary_spec[
            "canonical_compact_header_json_sha256"
        ]:
            raise TableAuditError("Table S3 primary header hash differs")

        comparison_counts: Counter[str] = Counter()
        finite_counts: Counter[str] = Counter()
        na_counts: Counter[str] = Counter()
        significant_counts: Counter[str] = Counter()
        nonsignificant_counts: Counter[str] = Counter()
        other_cached_translation_counts: Counter[str] = Counter()
        cached_translation_string_count = 0
        per_pair_comparisons: dict[str, set[str]] = defaultdict(set)
        annotation_by_pair: dict[str, str] = {}
        distinct_annotations: set[str] = set()
        finite_pair_keys: dict[str, set[str]] = defaultdict(set)
        primary_row_count = 0
        comparisons = set(primary_spec["comparison_row_counts"])
        missing_token = primary_spec["statistics_missing_token"]
        for values in primary_rows:
            row = list(values)
            if len(row) != len(primary_header):
                raise TableAuditError("Table S3 primary row width differs")
            primary_row_count += 1
            key, annotation, comparison = row[0], row[1], row[2]
            if not isinstance(key, str) or not key:
                raise TableAuditError("Table S3 primary pair key is invalid")
            if not isinstance(annotation, str) or not annotation:
                raise TableAuditError("Table S3 primary annotation is invalid")
            if comparison not in comparisons:
                raise TableAuditError("Table S3 primary comparison is outside allowlist")
            if comparison in per_pair_comparisons[key]:
                raise TableAuditError("Table S3 duplicates a pair-comparison row")
            per_pair_comparisons[key].add(comparison)
            previous_annotation = annotation_by_pair.setdefault(key, annotation)
            if previous_annotation != annotation:
                raise TableAuditError("Table S3 annotation differs between comparisons")
            distinct_annotations.add(annotation)
            comparison_counts[comparison] += 1
            statistics = row[3:6]
            all_finite = all(_is_finite_number(value) for value in statistics)
            all_na = all(value == missing_token for value in statistics)
            if all_finite:
                effect, pvalue, fdr = (float(value) for value in statistics)
                del effect
                if not 0.0 <= pvalue <= 1.0 or not 0.0 <= fdr <= 1.0:
                    raise TableAuditError("Table S3 p-value or FDR is outside [0,1]")
                finite_counts[comparison] += 1
                finite_pair_keys[comparison].add(key)
            elif all_na:
                na_counts[comparison] += 1
            else:
                raise TableAuditError("Table S3 statistics are mixed finite/NA")
            significance = row[6]
            if not isinstance(significance, str):
                raise TableAuditError("Table S3 cached translation cell is not a string")
            cached_translation_string_count += 1
            if significance == "Significant":
                significant_counts[comparison] += 1
            elif significance == "Not Significant":
                nonsignificant_counts[comparison] += 1
            else:
                other_cached_translation_counts[comparison] += 1
        if primary_row_count != primary_spec["data_row_count"]:
            raise TableAuditError("Table S3 primary row count differs")
        if len(per_pair_comparisons) != primary_spec["pair_key_count"]:
            raise TableAuditError("Table S3 primary pair-key count differs")
        if any(value != comparisons for value in per_pair_comparisons.values()):
            raise TableAuditError("Table S3 pair membership is not comparison-complete")
        if dict(comparison_counts) != primary_spec["comparison_row_counts"]:
            raise TableAuditError("Table S3 comparison row counts differ")
        if dict(finite_counts) != primary_spec["finite_statistic_rows"]:
            raise TableAuditError("Table S3 finite statistic counts differ")
        if dict(na_counts) != primary_spec["na_statistic_rows"]:
            raise TableAuditError("Table S3 NA statistic counts differ")
        high_complete = finite_pair_keys["HighPoly:RNA"]
        primary_complete = finite_pair_keys["TotalPoly:RNA"]
        completeness_partition = {
            "both_comparisons_finite_pair_count": len(
                high_complete & primary_complete
            ),
            "primary_only_finite_pair_count": len(
                primary_complete - high_complete
            ),
            "secondary_only_finite_pair_count": len(
                high_complete - primary_complete
            ),
            "neither_comparison_finite_pair_count": len(
                set(per_pair_comparisons) - (high_complete | primary_complete)
            ),
        }
        if any(
            completeness_partition[key] != primary_spec[key]
            for key in completeness_partition
        ):
            raise TableAuditError("Table S3 comparison completeness partition differs")
        if cached_translation_string_count != cell_spec["translation_cached_string"]:
            raise TableAuditError("Table S3 cached translation string count differs")
        if len(distinct_annotations) != primary_spec["distinct_gene_annotation_count"]:
            raise TableAuditError("Table S3 annotation cardinality differs")

        control_spec = spec["control_sheet"]
        control = workbook[control_spec["name"]]
        formula_control = formula_workbook[control_spec["name"]]
        expected_control_rows = control_spec["data_row_count"] + 1
        expected_control_columns = len(control_spec["exact_header"])
        if (
            control.max_row != expected_control_rows
            or control.max_column != expected_control_columns
            or formula_control.max_row != expected_control_rows
            or formula_control.max_column != expected_control_columns
        ):
            raise TableAuditError("Table S3 opaque control dimensions differ")
        control_rows = control.iter_rows(
            min_row=1, max_row=1, max_col=expected_control_columns, values_only=True
        )
        try:
            control_header = list(next(control_rows))
        except StopIteration as exc:
            raise TableAuditError("Table S3 control sheet is empty") from exc
        if control_header != control_spec["exact_header"]:
            raise TableAuditError("Table S3 control header differs")
        if _canonical_header_sha256(control_header) != control_spec[
            "canonical_compact_header_json_sha256"
        ]:
            raise TableAuditError("Table S3 control header hash differs")
    finally:
        workbook.close()
        formula_workbook.close()

    aggregate = {
        "status": "PASS_FULL_MEMBERSHIP_PUBLISHED_ENDPOINT_AUDIT",
        "sheet_name_count": len(spec["exact_sheet_names"]),
        "primary_data_row_count": primary_row_count,
        "primary_pair_count": len(per_pair_comparisons),
        "comparisons_per_pair": len(comparisons),
        "finite_statistic_rows": dict(sorted(finite_counts.items())),
        "na_statistic_rows": dict(sorted(na_counts.items())),
        **completeness_partition,
        "significant_rows": dict(sorted(significant_counts.items())),
        "nonsignificant_rows": dict(sorted(nonsignificant_counts.items())),
        "other_cached_translation_rows": dict(
            sorted(other_cached_translation_counts.items())
        ),
        "cached_translation_counts_role": "DESCRIPTIVE_ONLY_NOT_MEMBERSHIP_OR_GATE",
        "cached_translation_values_used_for_gate": False,
        "cached_translation_values_used_for_membership": False,
        "translation_formula_count": translation_formula_count,
        "translation_formula_executed": False,
        "cached_translation_string_count": cached_translation_string_count,
        "membership_includes_significant_and_nonsignificant": True,
        "significance_used_for_membership": False,
        "distinct_annotation_count": len(distinct_annotations),
        "primary_header_sha256": primary_spec[
            "canonical_compact_header_json_sha256"
        ],
        "control_header_sha256": control_spec[
            "canonical_compact_header_json_sha256"
        ],
        "opaque_control_row_count": control_spec["data_row_count"],
        "opaque_control_column_count": len(control_spec["exact_header"]),
        "opaque_control_data_cell_read_count": 0,
        "opaque_control_data_access_policy": "HEADER_AND_DIMENSIONS_ONLY",
        "opaque_control_excluded_from_endpoint": True,
        "opaque_control_excluded_from_qualification_counts": True,
        "active_content_executed": False,
    }
    return S3State(
        pair_keys=frozenset(per_pair_comparisons),
        finite_pair_keys_by_comparison={
            comparison: frozenset(keys)
            for comparison, keys in finite_pair_keys.items()
        },
        aggregate=aggregate,
    )


def _audit_join_and_endpoint(
    s2: S2State,
    s3: S3State,
    s3_spec: Mapping[str, Any],
    endpoint_spec: Mapping[str, Any],
) -> dict[str, Any]:
    s2_only = s2.pair_keys - s3.pair_keys
    s3_only = s3.pair_keys - s2.pair_keys
    joined = s2.pair_keys & s3.pair_keys
    join_spec = s3_spec["join"]
    joined_orientation_counter = Counter(
        s2.orientation_by_pair[key] for key in joined
    )
    joined_orientation_counts = {
        "FORWARD": joined_orientation_counter["FORWARD"],
        "REVERSE_COMPLEMENT": joined_orientation_counter["REVERSE_COMPLEMENT"],
        "UNRESOLVED": 0,
    }
    observed_join = {
        "table_s2_pair_count": len(s2.pair_keys),
        "table_s3_pair_count": len(s3.pair_keys),
        "joined_pair_count": len(joined),
        "table_s2_absent_from_table_s3_count": len(s2_only),
        "table_s3_not_in_table_s2_count": len(s3_only),
        "joined_pair_orientation_counts": joined_orientation_counts,
    }
    if observed_join != join_spec:
        raise TableAuditError("Table S2/S3 exact join counts differ")
    primary = str(endpoint_spec["primary_comparison_value"])
    finite_primary = s3.finite_pair_keys_by_comparison.get(primary, frozenset())
    if not finite_primary <= joined:
        raise TableAuditError("primary finite membership is outside exact join")
    primary_proxy_pools: dict[str, set[str]] = defaultdict(set)
    for key in finite_primary:
        primary_proxy_pools[s2.wt_201nt_by_pair[key]].add(
            s2.candidate_201nt_by_pair[key]
        )
    primary_proxy_count = len(primary_proxy_pools)
    primary_proxy_pool_hist = Counter(
        len(candidates) for candidates in primary_proxy_pools.values()
    )
    primary_proxy_pool_counts = {
        str(size): count for size, count in sorted(primary_proxy_pool_hist.items())
    }
    observed_endpoint = {
        "primary_membership_pair_count": len(s3.pair_keys),
        "primary_finite_effect_pair_count": len(finite_primary),
        "primary_absent_pair_count": len(s2_only),
        "primary_na_pair_count": len(s3.pair_keys) - len(finite_primary),
        "primary_total_attrition_count": len(s2_only)
        + len(s3.pair_keys)
        - len(finite_primary),
        "primary_complete_distinct_wt_201nt_proxy_group_count": primary_proxy_count,
        "primary_complete_wt_201nt_proxy_pool_size_counts": primary_proxy_pool_counts,
    }
    for key, value in observed_endpoint.items():
        if value != endpoint_spec[key]:
            raise TableAuditError("primary endpoint attrition or proxy count differs")
    return {
        "status": "PASS_EXACT_JOIN_AND_PRIMARY_ATTRITION_AUDIT",
        **observed_join,
        **observed_endpoint,
        "attrition_equation_verified": True,
        "membership_selected_by_significance": False,
        "study_level_reported_biological_replicate_count": endpoint_spec[
            "study_level_reported_biological_replicate_count"
        ],
        "row_level_effective_replicate_count": None,
        "row_level_effective_replicate_count_status": endpoint_spec[
            "row_level_effective_replicate_count_status"
        ],
        "standard_error": None,
        "standard_error_status": endpoint_spec["standard_error_status"],
        "power_effective_n": None,
        "power_effective_n_status": endpoint_spec["power_effective_n_status"],
        "effect_geometry": endpoint_spec["effect_geometry"],
        "true_a2_dense_candidate_count": 0,
        "wt_201nt_grouping_authority": False,
        "wt_201nt_grouping_proxy_only": True,
        "biological_source_group_authority_closed": False,
    }


def _assert_aggregate_safe_document(value: Any) -> None:
    if isinstance(value, Mapping):
        for key, child in value.items():
            if str(key).casefold() in PROHIBITED_OUTPUT_KEYS:
                raise PublicationError("output contains a prohibited row-level field")
            _assert_aggregate_safe_document(child)
    elif isinstance(value, list):
        for child in value:
            _assert_aggregate_safe_document(child)
    elif isinstance(value, str) and LONG_NUCLEOTIDE_RE.search(value):
        raise PublicationError("output contains a nucleotide payload")


def _require_exact_keys(
    value: Any, expected: set[str], *, label: str
) -> Mapping[str, Any]:
    if not isinstance(value, Mapping) or set(value) != expected:
        raise PublicationError(f"{label} keys differ from the closed schema")
    return value


def _require_count_map(
    value: Any, allowed_keys: set[str], *, label: str
) -> Mapping[str, int]:
    if not isinstance(value, Mapping) or not set(value) <= allowed_keys:
        raise PublicationError(f"{label} is outside the closed aggregate schema")
    if any(
        not isinstance(count, int) or isinstance(count, bool) or count < 0
        for count in value.values()
    ):
        raise PublicationError(f"{label} contains a non-count value")
    return value


def _validate_s2_output(value: Any) -> None:
    expected = {
        "status": "PASS_EXACT_DUPLICATE_PAIR_RECONCILIATION",
        "raw_row_count": EXPECTED_S2["raw_row_count"],
        "unique_content_row_count": EXPECTED_S2["unique_content_row_count"],
        "duplicate_extra_row_count": EXPECTED_S2["duplicate_extra_row_count"],
        "duplicated_content_group_count": EXPECTED_S2[
            "duplicated_content_group_count"
        ],
        "duplicated_pair_count": EXPECTED_S2["duplicated_pair_count"],
        "deduplicated_pair_count": EXPECTED_S2["deduplicated_pair_count"],
        "deduplicated_control_count": EXPECTED_S2["deduplicated_control_count"],
        "exactly_one_central_snv_pair_count": EXPECTED_S2[
            "deduplicated_pair_count"
        ],
        "central_zero_based_index": EXPECTED_S2["central_zero_based_index"],
        "central_base_change_counts": EXPECTED_S2["central_base_change_counts"],
        "pair_id_grammar_match_count": EXPECTED_S2["pair_id_grammar_match_count"],
        "all_pair_orientation_counts": EXPECTED_S2[
            "all_pair_orientation_counts"
        ],
        "distinct_candidate_count": EXPECTED_S2["distinct_candidate_count"],
        "distinct_wt_201nt_proxy_count": EXPECTED_S2[
            "distinct_wt_201nt_proxy_count"
        ],
        "wt_201nt_proxy_pool_size_counts": EXPECTED_S2[
            "wt_201nt_proxy_pool_size_counts"
        ],
        "wt_201nt_grouping_authority": False,
        "wt_201nt_grouping_proxy_only": True,
        "three_or_more_candidate_wt_201nt_proxy_pool_count": EXPECTED_S2[
            "three_or_more_candidate_wt_201nt_proxy_pool_count"
        ],
        "true_a2_dense_pool_count": 0,
        "header_sha256": EXPECTED_S2["canonical_compact_header_json_sha256"],
    }
    if value != expected:
        raise PublicationError("Table S2 output differs from the closed aggregate")


def _validate_s3_output(value: Any) -> None:
    primary = EXPECTED_S3["primary_sheet"]
    control = EXPECTED_S3["control_sheet"]
    exact_keys = {
        "status",
        "sheet_name_count",
        "primary_data_row_count",
        "primary_pair_count",
        "comparisons_per_pair",
        "finite_statistic_rows",
        "na_statistic_rows",
        "both_comparisons_finite_pair_count",
        "primary_only_finite_pair_count",
        "secondary_only_finite_pair_count",
        "neither_comparison_finite_pair_count",
        "significant_rows",
        "nonsignificant_rows",
        "other_cached_translation_rows",
        "cached_translation_counts_role",
        "cached_translation_values_used_for_gate",
        "cached_translation_values_used_for_membership",
        "translation_formula_count",
        "translation_formula_executed",
        "cached_translation_string_count",
        "membership_includes_significant_and_nonsignificant",
        "significance_used_for_membership",
        "distinct_annotation_count",
        "primary_header_sha256",
        "control_header_sha256",
        "opaque_control_row_count",
        "opaque_control_column_count",
        "opaque_control_data_cell_read_count",
        "opaque_control_data_access_policy",
        "opaque_control_excluded_from_endpoint",
        "opaque_control_excluded_from_qualification_counts",
        "active_content_executed",
    }
    table = _require_exact_keys(value, exact_keys, label="Table S3 output")
    dynamic_keys = {
        "significant_rows",
        "nonsignificant_rows",
        "other_cached_translation_rows",
    }
    expected_fixed = {
        "status": "PASS_FULL_MEMBERSHIP_PUBLISHED_ENDPOINT_AUDIT",
        "sheet_name_count": len(EXPECTED_S3["exact_sheet_names"]),
        "primary_data_row_count": primary["data_row_count"],
        "primary_pair_count": primary["pair_key_count"],
        "comparisons_per_pair": primary["comparison_count_per_pair"],
        "finite_statistic_rows": primary["finite_statistic_rows"],
        "na_statistic_rows": primary["na_statistic_rows"],
        "both_comparisons_finite_pair_count": primary[
            "both_comparisons_finite_pair_count"
        ],
        "primary_only_finite_pair_count": primary[
            "primary_only_finite_pair_count"
        ],
        "secondary_only_finite_pair_count": primary[
            "secondary_only_finite_pair_count"
        ],
        "neither_comparison_finite_pair_count": primary[
            "neither_comparison_finite_pair_count"
        ],
        "cached_translation_counts_role": "DESCRIPTIVE_ONLY_NOT_MEMBERSHIP_OR_GATE",
        "cached_translation_values_used_for_gate": False,
        "cached_translation_values_used_for_membership": False,
        "translation_formula_count": primary["cell_type_counts"][
            "translation_formula"
        ],
        "translation_formula_executed": False,
        "cached_translation_string_count": primary["cell_type_counts"][
            "translation_cached_string"
        ],
        "membership_includes_significant_and_nonsignificant": True,
        "significance_used_for_membership": False,
        "distinct_annotation_count": primary["distinct_gene_annotation_count"],
        "primary_header_sha256": primary[
            "canonical_compact_header_json_sha256"
        ],
        "control_header_sha256": control[
            "canonical_compact_header_json_sha256"
        ],
        "opaque_control_row_count": control["data_row_count"],
        "opaque_control_column_count": len(control["exact_header"]),
        "opaque_control_data_cell_read_count": 0,
        "opaque_control_data_access_policy": "HEADER_AND_DIMENSIONS_ONLY",
        "opaque_control_excluded_from_endpoint": True,
        "opaque_control_excluded_from_qualification_counts": True,
        "active_content_executed": False,
    }
    if any(table[key] != expected for key, expected in expected_fixed.items()):
        raise PublicationError("Table S3 output fixed aggregate drifted")
    comparisons = set(primary["comparison_row_counts"])
    descriptive = {
        key: _require_count_map(table[key], comparisons, label=key)
        for key in dynamic_keys
    }
    for comparison, row_count in primary["comparison_row_counts"].items():
        if sum(values.get(comparison, 0) for values in descriptive.values()) != row_count:
            raise PublicationError("cached translation descriptive partition drifted")


def _validate_endpoint_output(value: Any) -> None:
    endpoint = EXPECTED_ENDPOINT_BOUNDARY
    expected = {
        "status": "PASS_EXACT_JOIN_AND_PRIMARY_ATTRITION_AUDIT",
        **EXPECTED_S3["join"],
        "primary_membership_pair_count": endpoint["primary_membership_pair_count"],
        "primary_finite_effect_pair_count": endpoint[
            "primary_finite_effect_pair_count"
        ],
        "primary_absent_pair_count": endpoint["primary_absent_pair_count"],
        "primary_na_pair_count": endpoint["primary_na_pair_count"],
        "primary_total_attrition_count": endpoint[
            "primary_total_attrition_count"
        ],
        "primary_complete_distinct_wt_201nt_proxy_group_count": endpoint[
            "primary_complete_distinct_wt_201nt_proxy_group_count"
        ],
        "primary_complete_wt_201nt_proxy_pool_size_counts": endpoint[
            "primary_complete_wt_201nt_proxy_pool_size_counts"
        ],
        "attrition_equation_verified": True,
        "membership_selected_by_significance": False,
        "study_level_reported_biological_replicate_count": endpoint[
            "study_level_reported_biological_replicate_count"
        ],
        "row_level_effective_replicate_count": None,
        "row_level_effective_replicate_count_status": endpoint[
            "row_level_effective_replicate_count_status"
        ],
        "standard_error": None,
        "standard_error_status": endpoint["standard_error_status"],
        "power_effective_n": None,
        "power_effective_n_status": endpoint["power_effective_n_status"],
        "effect_geometry": endpoint["effect_geometry"],
        "true_a2_dense_candidate_count": 0,
        "wt_201nt_grouping_authority": False,
        "wt_201nt_grouping_proxy_only": True,
        "biological_source_group_authority_closed": False,
    }
    if value != expected:
        raise PublicationError("endpoint output differs from the closed aggregate")


def _validate_success_payloads(payloads: Mapping[str, Mapping[str, Any]]) -> None:
    if set(payloads) != set(SUCCESS_JSON_FILES):
        raise PublicationError("success payload member set is not closed")
    report = _require_exact_keys(
        payloads["QUALIFICATION_REPORT.json"],
        {
            "contract_id",
            "protocol_id",
            "dataset_id",
            "execution_outcome",
            "protocol_status",
            "activation_status",
            "qualification_status",
            "scientific_claim_status",
            "engineering_audits_passed",
            "qualified",
            "ordinary_study_contribution",
            "a1_intervention_study_contribution",
            "true_a2_dense_study_contribution",
            "canonical_record_count",
            "canonical_materialization_allowed",
            "training_allowed",
            "model_selection_allowed",
            "next_phase_authorized",
            "owner_policy_decision_made",
            "checkpoint_specific_use_cleared",
            "biological_source_group_authority_closed",
            "current_authority_endpoint_scope_status",
            "implementation_binding",
            "unresolved_blockers",
            "raw_replay_status",
            "zenodo_code_execution_status",
            "aggregate_only",
        },
        label="success report",
    )
    required = {
        "execution_outcome": SUCCESS_OUTCOME,
        "qualification_status": QUALIFICATION_STATUS,
        "qualified": False,
        "ordinary_study_contribution": 0,
        "a1_intervention_study_contribution": 0,
        "true_a2_dense_study_contribution": 0,
        "canonical_record_count": 0,
        "canonical_materialization_allowed": False,
        "training_allowed": False,
        "model_selection_allowed": False,
        "next_phase_authorized": False,
        "owner_policy_decision_made": False,
        "checkpoint_specific_use_cleared": False,
        "biological_source_group_authority_closed": False,
        "current_authority_endpoint_scope_status": "REQUIRED_80S_ROLE_AUTHORITY_ABSENT_NOT_ROUTED_FOR_PUBLISHED_ENDPOINT_REUSE",
        "aggregate_only": True,
    }
    for key, expected in required.items():
        if report.get(key) != expected:
            raise PublicationError("success report decision-neutral invariant drifted")
    for key, expected in {
        "contract_id": CONTRACT_ID,
        "protocol_id": PROTOCOL_ID,
        "dataset_id": DATASET_ID,
        "protocol_status": PROTOCOL_STATUS,
        "activation_status": ACTIVATION_STATUS,
        "scientific_claim_status": "NOT_ESTABLISHED",
        "engineering_audits_passed": True,
        "raw_replay_status": "NOT_RUN_NOT_IN_SCOPE",
        "zenodo_code_execution_status": "NOT_RUN_NOT_IN_SCOPE",
    }.items():
        if report[key] != expected:
            raise PublicationError("success report fixed identity or scope drifted")
    binding = _require_exact_keys(
        report["implementation_binding"],
        {
            "status",
            "verified",
            "implementation_commit",
            "binding_commit",
            "clean_worktree",
            "config_only_direct_child",
            "authority_blobs_match",
            "implementation_blobs_match",
            "running_script_matches_bound_blob",
        },
        label="success binding audit",
    )
    if (
        binding["status"] != "PASS_BOUND_IMPLEMENTATION"
        or binding["verified"] is not True
        or COMMIT_RE.fullmatch(str(binding["implementation_commit"])) is None
        or COMMIT_RE.fullmatch(str(binding["binding_commit"])) is None
        or any(
            binding[key] is not True
            for key in (
                "clean_worktree",
                "config_only_direct_child",
                "authority_blobs_match",
                "implementation_blobs_match",
                "running_script_matches_bound_blob",
            )
        )
    ):
        raise PublicationError("success binding audit is not exact PASS_BOUND")
    if report.get("unresolved_blockers") != _expected_blockers(
        report["implementation_binding"]
    ):
        raise PublicationError("success report blockers drifted")
    integrity = _require_exact_keys(
        payloads["INPUT_INTEGRITY_AUDIT.json"],
        {
            "contract_id",
            "protocol_id",
            "dataset_id",
            "status",
            "protocol_snapshot",
            "asset_count",
            "official_table_asset_count",
            "assets",
            "source_bundle_closure",
            "all_hashes_and_sizes_verified_before_parsing",
            "same_descriptor_snapshot_verified",
            "root_to_leaf_o_nofollow_verified",
            "single_link_regular_files_verified",
            "network_accessed",
            "input_payload_writes",
            "raw_fastq_or_alignment_input_count",
            "external_code_executed",
            "opaque_zenodo_member_byte_verified",
            "opaque_zenodo_archive_container_parsed",
            "opaque_zenodo_code_executed",
            "aggregate_only",
        },
        label="input integrity output",
    )
    snapshot = _require_exact_keys(
        integrity["protocol_snapshot"],
        {"sha256", "bytes", "launch_expected_sha256"},
        label="protocol snapshot",
    )
    if (
        SHA256_RE.fullmatch(str(snapshot["sha256"])) is None
        or snapshot["sha256"] != snapshot["launch_expected_sha256"]
        or not isinstance(snapshot["bytes"], int)
        or isinstance(snapshot["bytes"], bool)
        or snapshot["bytes"] <= 0
    ):
        raise PublicationError("protocol snapshot provenance drifted")
    assets = integrity["assets"]
    if not isinstance(assets, list) or len(assets) != len(EXPECTED_ASSETS):
        raise PublicationError("input integrity asset list differs")
    for observed, expected_asset in zip(assets, EXPECTED_ASSETS):
        item = _require_exact_keys(
            observed,
            {"asset_id", "role", "relative_locator", "bytes", "sha256", "verified"},
            label="input integrity asset",
        )
        if item != {
            "asset_id": expected_asset["asset_id"],
            "role": expected_asset["role"],
            "relative_locator": expected_asset["relative_path"],
            "bytes": expected_asset["bytes"],
            "sha256": expected_asset["sha256"],
            "verified": True,
        }:
            raise PublicationError("input integrity asset provenance drifted")
    closure = _require_exact_keys(
        integrity["source_bundle_closure"],
        {
            "status",
            "exact_member_count",
            "checksum_entry_count",
            "source_manifest_asset_count",
            "terminal_marker_committed",
            "terminal_marker_written_last",
            "endpoint_statistics_parsed_asset_count",
            "opaque_zenodo_member_byte_verified",
            "opaque_zenodo_archive_container_parsed",
            "opaque_zenodo_code_executed",
            "raw_reads_or_alignments_opened",
        },
        label="source closure output",
    )
    if closure != {
        "status": "PASS_EXACT_SEVEN_MEMBER_SOURCE_BUNDLE_CLOSURE",
        "exact_member_count": len(EXPECTED_ASSETS),
        "checksum_entry_count": 5,
        "source_manifest_asset_count": 4,
        "terminal_marker_committed": True,
        "terminal_marker_written_last": True,
        "endpoint_statistics_parsed_asset_count": len(EXPECTED_PARSED_ASSET_IDS),
        "opaque_zenodo_member_byte_verified": True,
        "opaque_zenodo_archive_container_parsed": False,
        "opaque_zenodo_code_executed": False,
        "raw_reads_or_alignments_opened": False,
    }:
        raise PublicationError("source closure output drifted")
    for key, expected in {
        "contract_id": CONTRACT_ID,
        "protocol_id": PROTOCOL_ID,
        "dataset_id": DATASET_ID,
        "status": "PASS_EXACT_OFFICIAL_TABLE_INPUT_INTEGRITY",
        "asset_count": len(EXPECTED_ASSETS),
        "official_table_asset_count": len(EXPECTED_PARSED_ASSET_IDS),
        "all_hashes_and_sizes_verified_before_parsing": True,
        "same_descriptor_snapshot_verified": True,
        "root_to_leaf_o_nofollow_verified": True,
        "single_link_regular_files_verified": True,
        "network_accessed": False,
        "input_payload_writes": 0,
        "raw_fastq_or_alignment_input_count": 0,
        "external_code_executed": False,
        "opaque_zenodo_member_byte_verified": True,
        "opaque_zenodo_archive_container_parsed": False,
        "opaque_zenodo_code_executed": False,
        "aggregate_only": True,
    }.items():
        if integrity[key] != expected:
            raise PublicationError("input integrity fixed invariant drifted")
    published = _require_exact_keys(
        payloads["PUBLISHED_ENDPOINT_AUDIT.json"],
        {
            "contract_id",
            "protocol_id",
            "dataset_id",
            "status",
            "table_s2",
            "table_s3",
            "endpoint_boundary",
            "opaque_control_sheet_role",
            "published_endpoint_is_not_raw_replay",
            "published_endpoint_is_not_canonical_materialization",
            "aggregate_only",
        },
        label="published endpoint output",
    )
    for key, expected in {
        "contract_id": CONTRACT_ID,
        "protocol_id": PROTOCOL_ID,
        "dataset_id": DATASET_ID,
        "status": "PASS_MECHANICAL_AUDITS_REMAINS_BLOCKED",
        "opaque_control_sheet_role": "CONTROL_ONLY_EXCLUDED_FROM_MUTATION_ENDPOINT_AND_QUALIFICATION_COUNTS",
        "published_endpoint_is_not_raw_replay": True,
        "published_endpoint_is_not_canonical_materialization": True,
        "aggregate_only": True,
    }.items():
        if published[key] != expected:
            raise PublicationError("published endpoint fixed invariant drifted")
    _validate_s2_output(published["table_s2"])
    _validate_s3_output(published["table_s3"])
    _validate_endpoint_output(published["endpoint_boundary"])
    for payload in payloads.values():
        _assert_aggregate_safe_document(payload)


def _validate_failure_payloads(payloads: Mapping[str, Mapping[str, Any]]) -> None:
    if set(payloads) != set(FAILURE_JSON_FILES):
        raise PublicationError("failure payload member set is not closed")
    report = _require_exact_keys(
        payloads["FAILURE_REPORT.json"],
        {
            "contract_id",
            "protocol_id",
            "dataset_id",
            "execution_outcome",
            "failure_code",
            "status",
            "qualified",
            "ordinary_study_contribution",
            "a1_intervention_study_contribution",
            "true_a2_dense_study_contribution",
            "canonical_record_count",
            "training_allowed",
            "model_selection_allowed",
            "next_phase_authorized",
            "aggregate_only",
        },
        label="failure report",
    )
    for key, expected in {
        "execution_outcome": FAILURE_OUTCOME,
        "qualified": False,
        "ordinary_study_contribution": 0,
        "a1_intervention_study_contribution": 0,
        "true_a2_dense_study_contribution": 0,
        "canonical_record_count": 0,
        "training_allowed": False,
        "model_selection_allowed": False,
        "next_phase_authorized": False,
        "aggregate_only": True,
    }.items():
        if report.get(key) != expected:
            raise PublicationError("failure report invariant drifted")
    if (
        report["contract_id"] != CONTRACT_ID
        or report["protocol_id"] != PROTOCOL_ID
        or report["dataset_id"] != DATASET_ID
        or report["status"] != "FAIL_CLOSED_BEFORE_BLOCKED_SUCCESS_BUNDLE"
        or report["failure_code"]
        not in {
            ProtocolError.code,
            InputIntegrityError.code,
            TableAuditError.code,
            QualificationError.code,
        }
    ):
        raise PublicationError("failure report identity or code drifted")
    _assert_aggregate_safe_document(report)


def _validate_closed_payloads(
    payloads: Mapping[str, Mapping[str, Any]], *, outcome: str
) -> None:
    if outcome == SUCCESS_OUTCOME:
        _validate_success_payloads(payloads)
    elif outcome == FAILURE_OUTCOME:
        _validate_failure_payloads(payloads)
    else:
        raise PublicationError("publication outcome is outside the closed enum")


def _build_success_payloads(
    *,
    protocol: Mapping[str, Any],
    protocol_provenance: Mapping[str, Any],
    asset_provenance: Mapping[str, Mapping[str, Any]],
    source_bundle_closure: Mapping[str, Any],
    implementation_binding: Mapping[str, Any],
    s2: S2State,
    s3: S3State,
    endpoint_audit: Mapping[str, Any],
) -> dict[str, dict[str, Any]]:
    assets = []
    for asset in protocol["input_contract"]["source_bundle_members"]:
        observed = asset_provenance[asset["asset_id"]]
        assets.append(
            {
                "asset_id": asset["asset_id"],
                "role": asset["role"],
                "relative_locator": asset["relative_path"],
                "bytes": observed["bytes"],
                "sha256": observed["sha256"],
                "verified": True,
            }
        )
    integrity = {
        "contract_id": CONTRACT_ID,
        "protocol_id": PROTOCOL_ID,
        "dataset_id": DATASET_ID,
        "status": "PASS_EXACT_OFFICIAL_TABLE_INPUT_INTEGRITY",
        "protocol_snapshot": dict(protocol_provenance),
        "asset_count": len(assets),
        "official_table_asset_count": len(EXPECTED_PARSED_ASSET_IDS),
        "assets": assets,
        "source_bundle_closure": dict(source_bundle_closure),
        "all_hashes_and_sizes_verified_before_parsing": True,
        "same_descriptor_snapshot_verified": True,
        "root_to_leaf_o_nofollow_verified": True,
        "single_link_regular_files_verified": True,
        "network_accessed": False,
        "input_payload_writes": 0,
        "raw_fastq_or_alignment_input_count": 0,
        "external_code_executed": False,
        "opaque_zenodo_member_byte_verified": True,
        "opaque_zenodo_archive_container_parsed": False,
        "opaque_zenodo_code_executed": False,
        "aggregate_only": True,
    }
    endpoint = {
        "contract_id": CONTRACT_ID,
        "protocol_id": PROTOCOL_ID,
        "dataset_id": DATASET_ID,
        "status": "PASS_MECHANICAL_AUDITS_REMAINS_BLOCKED",
        "table_s2": dict(s2.aggregate),
        "table_s3": dict(s3.aggregate),
        "endpoint_boundary": dict(endpoint_audit),
        "opaque_control_sheet_role": (
            "CONTROL_ONLY_EXCLUDED_FROM_MUTATION_ENDPOINT_AND_QUALIFICATION_COUNTS"
        ),
        "published_endpoint_is_not_raw_replay": True,
        "published_endpoint_is_not_canonical_materialization": True,
        "aggregate_only": True,
    }
    report = {
        "contract_id": CONTRACT_ID,
        "protocol_id": PROTOCOL_ID,
        "dataset_id": DATASET_ID,
        "execution_outcome": SUCCESS_OUTCOME,
        "protocol_status": PROTOCOL_STATUS,
        "activation_status": ACTIVATION_STATUS,
        "qualification_status": QUALIFICATION_STATUS,
        "scientific_claim_status": "NOT_ESTABLISHED",
        "engineering_audits_passed": True,
        "qualified": False,
        "ordinary_study_contribution": 0,
        "a1_intervention_study_contribution": 0,
        "true_a2_dense_study_contribution": 0,
        "canonical_record_count": 0,
        "canonical_materialization_allowed": False,
        "training_allowed": False,
        "model_selection_allowed": False,
        "next_phase_authorized": False,
        "owner_policy_decision_made": False,
        "checkpoint_specific_use_cleared": False,
        "biological_source_group_authority_closed": False,
        "current_authority_endpoint_scope_status": "REQUIRED_80S_ROLE_AUTHORITY_ABSENT_NOT_ROUTED_FOR_PUBLISHED_ENDPOINT_REUSE",
        "implementation_binding": dict(implementation_binding),
        "unresolved_blockers": list(protocol["unresolved_blockers"]),
        "raw_replay_status": "NOT_RUN_NOT_IN_SCOPE",
        "zenodo_code_execution_status": "NOT_RUN_NOT_IN_SCOPE",
        "aggregate_only": True,
    }
    payloads = {
        "INPUT_INTEGRITY_AUDIT.json": integrity,
        "PUBLISHED_ENDPOINT_AUDIT.json": endpoint,
        "QUALIFICATION_REPORT.json": report,
    }
    _validate_success_payloads(payloads)
    return payloads


def _failure_payload(code: str) -> dict[str, dict[str, Any]]:
    allowed_codes = {
        ProtocolError.code,
        InputIntegrityError.code,
        TableAuditError.code,
        QualificationError.code,
    }
    if code not in allowed_codes:
        code = QualificationError.code
    payloads = {
        "FAILURE_REPORT.json": {
            "contract_id": CONTRACT_ID,
            "protocol_id": PROTOCOL_ID,
            "dataset_id": DATASET_ID,
            "execution_outcome": FAILURE_OUTCOME,
            "failure_code": code,
            "status": "FAIL_CLOSED_BEFORE_BLOCKED_SUCCESS_BUNDLE",
            "qualified": False,
            "ordinary_study_contribution": 0,
            "a1_intervention_study_contribution": 0,
            "true_a2_dense_study_contribution": 0,
            "canonical_record_count": 0,
            "training_allowed": False,
            "model_selection_allowed": False,
            "next_phase_authorized": False,
            "aggregate_only": True,
        }
    }
    _validate_failure_payloads(payloads)
    return payloads


def _safe_output_name(name: str) -> str:
    if name not in {
        *SUCCESS_JSON_FILES,
        *FAILURE_JSON_FILES,
        SHA256SUMS_FILENAME,
        PUBLICATION_MARKER,
    }:
        raise PublicationError("output filename is outside the closed allowlist")
    return name


def _publication_fault(phase: str) -> None:
    if _PUBLICATION_FAULT_HOOK is not None:
        _PUBLICATION_FAULT_HOOK(phase)


def _write_exclusive_at(
    directory_fd: int,
    name: str,
    payload: bytes,
    *,
    allow_staged_name: bool = False,
) -> None:
    if not allow_staged_name:
        _safe_output_name(name)
    else:
        _safe_basename(name, label="staged terminal marker")
    flags = (
        os.O_WRONLY
        | os.O_CREAT
        | os.O_EXCL
        | getattr(os, "O_CLOEXEC", 0)
        | getattr(os, "O_NOFOLLOW", 0)
    )
    try:
        descriptor = os.open(name, flags, 0o640, dir_fd=directory_fd)
    except OSError as exc:
        raise PublicationError("exclusive publication member cannot be created") from exc
    try:
        view = memoryview(payload)
        while view:
            written = os.write(descriptor, view)
            if written <= 0:
                raise PublicationError("exclusive publication write made no progress")
            view = view[written:]
        os.fsync(descriptor)
    except Exception as exc:
        if isinstance(exc, QualificationError):
            raise
        raise PublicationError("exclusive publication member write failed") from exc
    finally:
        os.close(descriptor)


def _read_member_snapshot_at(
    directory_fd: int,
    name: str,
    *,
    allow_staged_name: bool = False,
) -> tuple[bytes, FileIdentity]:
    if not allow_staged_name:
        _safe_output_name(name)
    descriptor, opened = _open_regular_leaf_at(
        directory_fd,
        name,
        label="publication member",
        require_single_link=False,
    )
    identity = _file_identity(opened)
    try:
        chunks: list[bytes] = []
        while True:
            block = os.read(descriptor, 1 << 20)
            if not block:
                break
            chunks.append(block)
        final = os.fstat(descriptor)
        if _file_identity(final) != identity:
            raise PublicationError("publication member changed during read")
        payload = b"".join(chunks)
        if len(payload) != opened.st_size:
            raise PublicationError("publication member byte count changed")
    finally:
        os.close(descriptor)
    _assert_leaf_binding(
        directory_fd,
        name,
        identity,
        label="publication member post-read",
        require_single_link=False,
    )
    return payload, identity


def _read_member_at(
    directory_fd: int,
    name: str,
    *,
    allow_staged_name: bool = False,
) -> bytes:
    payload, _ = _read_member_snapshot_at(
        directory_fd, name, allow_staged_name=allow_staged_name
    )
    return payload


def _publication_marker_document(
    *, outcome: str, member_payloads: Mapping[str, bytes], output_directory: Path
) -> dict[str, Any]:
    target = _absolute_without_resolving(output_directory)
    return {
        "schema_version": "1.0.0",
        "record_type": "GSE200304_PUBLISHED_ENDPOINT_A1_PUBLICATION_COMMIT",
        "contract_id": CONTRACT_ID,
        "protocol_id": PROTOCOL_ID,
        "output_id": OUTPUT_ID,
        "execution_outcome": outcome,
        "bundle_member_names": sorted(member_payloads),
        "bundle_member_count": len(member_payloads),
        "sha256sums_sha256": _sha256_bytes(member_payloads[SHA256SUMS_FILENAME]),
        "final_output_directory_name_sha256": _sha256_bytes(target.name.encode("utf-8")),
        "final_output_target_sha256": _sha256_bytes(os.fspath(target).encode("utf-8")),
        "terminal_publication_operation": "FSYNCED_STAGED_HARDLINK_NO_REPLACE",
        "committed": True,
        "terminal_marker_written_last": True,
    }


def _validate_checksum_payload(member_payloads: Mapping[str, bytes]) -> None:
    sums = member_payloads[SHA256SUMS_FILENAME]
    try:
        lines = sums.decode("ascii").splitlines()
    except UnicodeDecodeError as exc:
        raise PublicationError("SHA256SUMS is not ASCII") from exc
    names = sorted(set(member_payloads) - {SHA256SUMS_FILENAME})
    expected = [f"{_sha256_bytes(member_payloads[name])}  {name}" for name in names]
    if lines != expected:
        raise PublicationError("SHA256SUMS differs from closed member bytes")


def _precommit_validate(
    output_fd: int,
    parent_fd: int,
    staged_marker_name: str,
    *,
    member_payloads: Mapping[str, bytes],
    marker_payload: bytes,
) -> tuple[dict[str, FileIdentity], FileIdentity]:
    if set(os.listdir(output_fd)) != set(member_payloads):
        raise PublicationError("precommit output member set differs")
    member_identities: dict[str, FileIdentity] = {}
    for name, expected in member_payloads.items():
        observed, identity = _read_member_snapshot_at(output_fd, name)
        if observed != expected:
            raise PublicationError("precommit output bytes differ")
        member_identities[name] = identity
    observed_marker, marker_identity = _read_member_snapshot_at(
        parent_fd, staged_marker_name, allow_staged_name=True
    )
    if observed_marker != marker_payload:
        raise PublicationError("staged terminal-marker bytes differ")
    _validate_checksum_payload(member_payloads)
    return member_identities, marker_identity


def _assert_precommit_identities_at_link_boundary(
    output_fd: int,
    parent_fd: int,
    staged_marker_name: str,
    member_identities: Mapping[str, FileIdentity],
    marker_identity: FileIdentity,
) -> None:
    if set(os.listdir(output_fd)) != set(member_identities):
        raise PublicationError("publication member set changed before terminal link")
    for name, identity in member_identities.items():
        _assert_leaf_binding(
            output_fd,
            name,
            identity,
            label="publication member at terminal-link boundary",
            require_single_link=True,
        )
    _assert_leaf_binding(
        parent_fd,
        staged_marker_name,
        marker_identity,
        label="staged terminal marker at terminal-link boundary",
        require_single_link=True,
    )


def _open_output_directory_at(parent_fd: int, name: str) -> tuple[int, os.stat_result]:
    flags = (
        os.O_RDONLY
        | getattr(os, "O_CLOEXEC", 0)
        | getattr(os, "O_DIRECTORY", 0)
        | getattr(os, "O_NOFOLLOW", 0)
    )
    descriptor = os.open(name, flags, dir_fd=parent_fd)
    opened = os.fstat(descriptor)
    if not stat.S_ISDIR(opened.st_mode):
        os.close(descriptor)
        raise PublicationError("output namespace is not a directory")
    return descriptor, opened


def _publish_closed_bundle(
    output_directory: Path,
    payloads: Mapping[str, Mapping[str, Any]],
    *,
    outcome: str,
) -> dict[str, Any]:
    _validate_closed_payloads(payloads, outcome=outcome)
    encoded = {name: _pretty_json_bytes(payloads[name]) for name in sorted(payloads)}
    checksum = "".join(
        f"{_sha256_bytes(encoded[name])}  {name}\n" for name in sorted(encoded)
    ).encode("ascii")
    members = {**encoded, SHA256SUMS_FILENAME: checksum}
    target = _absolute_without_resolving(output_directory)
    parent = _open_directory_no_symlinks(target.parent, label="output parent")
    output_fd: int | None = None
    staged_name: str | None = None
    output_created = False
    committed = False
    precommit_validated = False
    postcommit_warnings: list[str] = []
    core_error: Exception | None = None
    try:
        _safe_basename(target.name, label="output directory basename")
        try:
            os.mkdir(target.name, 0o700, dir_fd=parent.fd)
            output_created = True
        except FileExistsError as exc:
            raise PublicationContention("exclusive output target already exists") from exc
        except OSError as exc:
            raise PublicationError("exclusive output directory cannot be created") from exc
        output_fd, _ = _open_output_directory_at(parent.fd, target.name)
        for name in sorted(members):
            _write_exclusive_at(output_fd, name, members[name])
        _publication_fault("precommit_output_fsync")
        os.fsync(output_fd)
        _publication_fault("precommit_parent_fsync")
        os.fsync(parent.fd)
        _assert_directory_binding(parent)

        marker_document = _publication_marker_document(
            outcome=outcome, member_payloads=members, output_directory=target
        )
        marker_payload = _pretty_json_bytes(marker_document)
        staged_name = (
            ".publication-"
            + _sha256_bytes(target.name.encode("utf-8"))[:12]
            + "-"
            + secrets.token_hex(8)
            + ".stage"
        )
        _write_exclusive_at(
            parent.fd, staged_name, marker_payload, allow_staged_name=True
        )
        _publication_fault("precommit_validation")
        # Fault injection and all mutable preparatory work precede the final
        # byte validation.  The returned identities are rechecked immediately
        # at the no-replace hardlink boundary.
        _publication_fault("terminal_hardlink")
        member_identities, marker_identity = _precommit_validate(
            output_fd,
            parent.fd,
            staged_name,
            member_payloads=members,
            marker_payload=marker_payload,
        )
        _assert_precommit_identities_at_link_boundary(
            output_fd,
            parent.fd,
            staged_name,
            member_identities,
            marker_identity,
        )
        precommit_validated = True
        try:
            os.link(
                staged_name,
                PUBLICATION_MARKER,
                src_dir_fd=parent.fd,
                dst_dir_fd=output_fd,
                follow_symlinks=False,
            )
        except FileExistsError as exc:
            raise PublicationContention("terminal marker already exists") from exc
        except OSError as exc:
            raise PublicationError("terminal hardlink publication failed") from exc
        # This assignment is the commit truth boundary.  No acceptance-critical
        # read or ordinary-failure transition is permitted below it.
        committed = True
    except Exception as exc:
        core_error = exc

    if committed:
        if staged_name is not None:
            try:
                _publication_fault("postcommit_unlink_stage")
                os.unlink(staged_name, dir_fd=parent.fd)
                staged_name = None
            except Exception:
                postcommit_warnings.append("STAGED_TERMINAL_LINK_CLEANUP_WARNING")
        for phase, descriptor, warning in (
            ("postcommit_output_fsync", output_fd, "POSTCOMMIT_OUTPUT_FSYNC_WARNING"),
            ("postcommit_parent_fsync", parent.fd, "POSTCOMMIT_PARENT_FSYNC_WARNING"),
        ):
            if descriptor is None:
                postcommit_warnings.append(warning)
                continue
            try:
                _publication_fault(phase)
                os.fsync(descriptor)
            except Exception:
                postcommit_warnings.append(warning)
    elif staged_name is not None:
        try:
            os.unlink(staged_name, dir_fd=parent.fd)
            staged_name = None
        except Exception:
            pass

    if output_fd is not None:
        try:
            os.close(output_fd)
        except OSError:
            if committed:
                postcommit_warnings.append("POSTCOMMIT_OUTPUT_CLOSE_WARNING")
            elif core_error is None:
                core_error = PublicationError("precommit output close failed")
    try:
        os.close(parent.fd)
    except OSError:
        if committed:
            postcommit_warnings.append("POSTCOMMIT_PARENT_CLOSE_WARNING")
        elif core_error is None:
            core_error = PublicationError("precommit parent close failed")

    if committed:
        return {
            "publication_state": (
                "COMMITTED_WITH_DURABILITY_WARNING"
                if postcommit_warnings
                else "COMMITTED_ACCEPTED"
            ),
            "execution_outcome": outcome,
            "committed": True,
            "accepted": precommit_validated,
            "postcommit_warning_codes": sorted(set(postcommit_warnings)),
            "qualified": False,
            "ordinary_study_contribution": 0,
            "a1_intervention_study_contribution": 0,
            "true_a2_dense_study_contribution": 0,
            "canonical_record_count": 0,
            "training_allowed": False,
            "model_selection_allowed": False,
            "next_phase_authorized": False,
            "terminal_marker": PUBLICATION_MARKER,
            "terminal_marker_written_last": True,
            "terminal_publication_operation": "FSYNCED_STAGED_HARDLINK_NO_REPLACE",
            "no_acceptance_critical_read_after_commit": True,
            "no_overwrite": True,
        }
    if isinstance(core_error, QualificationError):
        if output_created and not isinstance(core_error, PublicationContention):
            raise PartialPrecommitError("publication stopped before commit truth") from core_error
        raise core_error
    if core_error is not None:
        if output_created:
            raise PartialPrecommitError("publication stopped before commit truth") from core_error
        raise PublicationError("publication failed before output creation") from core_error
    raise PublicationError("publication ended without commit truth")


def validate_published_bundle(output_directory: Path) -> dict[str, Any]:
    """Independent consumer validation; never part of publisher commit truth."""

    target = _absolute_without_resolving(output_directory)
    parent = _open_directory_no_symlinks(target.parent, label="published bundle parent")
    output_fd: int | None = None
    try:
        output_fd, _ = _open_output_directory_at(parent.fd, target.name)
        marker_payload = _read_member_at(output_fd, PUBLICATION_MARKER)
        marker = _strict_json_object(
            marker_payload,
            error_type=PublicationError,
            label="terminal publication marker",
        )
        outcome = marker.get("execution_outcome") if isinstance(marker, Mapping) else None
        if outcome == SUCCESS_OUTCOME:
            json_names = set(SUCCESS_JSON_FILES)
        elif outcome == FAILURE_OUTCOME:
            json_names = set(FAILURE_JSON_FILES)
        else:
            raise PublicationError("terminal marker outcome is outside enum")
        expected_names = json_names | {SHA256SUMS_FILENAME, PUBLICATION_MARKER}
        if set(os.listdir(output_fd)) != expected_names:
            raise PublicationError("committed bundle member set differs")
        members = {
            name: _read_member_at(output_fd, name)
            for name in sorted(expected_names - {PUBLICATION_MARKER})
        }
        expected_marker = _publication_marker_document(
            outcome=outcome, member_payloads=members, output_directory=target
        )
        if marker != expected_marker:
            raise PublicationError("terminal marker does not bind exact bundle")
        _validate_checksum_payload(members)
        decoded: dict[str, Mapping[str, Any]] = {}
        for name in json_names:
            decoded[name] = _strict_json_object(
                members[name],
                error_type=PublicationError,
                label="committed publication JSON",
            )
        _validate_closed_payloads(decoded, outcome=outcome)
        return {
            "publication_state": "COMMITTED_ACCEPTED",
            "committed": True,
            "accepted": True,
            "execution_outcome": outcome,
        }
    except OSError as exc:
        raise PublicationError("published bundle consumer failed closed") from exc
    finally:
        if output_fd is not None:
            os.close(output_fd)
        os.close(parent.fd)


def _qualify_preflighted(
    *,
    protocol_path: Path,
    protocol_sha256: str,
    data_root: Path,
    output_directory: Path,
) -> dict[str, Any]:
    try:
        protocol, protocol_provenance = _load_protocol(
            protocol_path, protocol_sha256
        )
        binding_audit = _verify_implementation_binding(
            protocol["implementation_binding"],
            protocol["authority"],
            protocol_path.parents[1],
        )
        if binding_audit.get("verified") is not True:
            raise ProtocolError(
                "implementation binding must be BOUND before official-asset access"
            )
        root = _open_directory_no_symlinks(
            data_root, label="official GSE200304 published-table root"
        )
        payloads: dict[str, bytes] = {}
        provenance: dict[str, dict[str, Any]] = {}
        identities: dict[str, tuple[str, FileIdentity]] = {}
        try:
            if sorted(os.listdir(root.fd)) != sorted(
                protocol["input_contract"]["source_bundle_closure"][
                    "exact_member_names"
                ]
            ):
                raise InputIntegrityError(
                    "source-bundle directory is not the exact seven-member closure"
                )
            # Every frozen input is captured and verified before any table parse.
            for asset in protocol["input_contract"]["source_bundle_members"]:
                payload, observed, identity = _read_relative_verified_snapshot(
                    root,
                    asset["relative_path"],
                    label=f"official asset {asset['asset_id']}",
                    expected_sha256=asset["sha256"],
                    expected_bytes=asset["bytes"],
                )
                payloads[asset["asset_id"]] = payload
                provenance[asset["asset_id"]] = observed
                identities[asset["asset_id"]] = (asset["relative_path"], identity)
            if _POST_VERIFIED_INPUT_SNAPSHOT_HOOK is not None:
                _POST_VERIFIED_INPUT_SNAPSHOT_HOOK()
            _assert_directory_binding(root)
            if sorted(os.listdir(root.fd)) != sorted(
                protocol["input_contract"]["source_bundle_closure"][
                    "exact_member_names"
                ]
            ):
                raise InputIntegrityError(
                    "source-bundle member closure changed after snapshot capture"
                )
            for asset_id, (relative, identity) in identities.items():
                _assert_relative_identity(
                    root,
                    relative,
                    identity,
                    label=f"post-snapshot {asset_id}",
                )
        finally:
            os.close(root.fd)
    except QualificationError:
        raise
    except OSError as exc:
        raise InputIntegrityError("input I/O failed closed") from exc

    source_closure = _audit_source_bundle_closure(
        payloads, protocol["input_contract"]
    )
    table_contract = protocol["table_contract"]
    s2 = _audit_table_s2(
        payloads["PMC10540565_TABLE_S2"], table_contract["table_s2"]
    )
    s3 = _audit_table_s3(
        payloads["PMC10540565_TABLE_S3"], table_contract["table_s3"]
    )
    endpoint_audit = _audit_join_and_endpoint(
        s2,
        s3,
        table_contract["table_s3"],
        protocol["endpoint_boundary"],
    )
    output_payloads = _build_success_payloads(
        protocol=protocol,
        protocol_provenance=protocol_provenance,
        asset_provenance=provenance,
        source_bundle_closure=source_closure,
        implementation_binding=binding_audit,
        s2=s2,
        s3=s3,
        endpoint_audit=endpoint_audit,
    )
    return _publish_closed_bundle(
        output_directory, output_payloads, outcome=SUCCESS_OUTCOME
    )


def qualify_gse200304_published_endpoint_a1(
    *,
    protocol_path: Path,
    protocol_sha256: str,
    data_root: Path,
    output_directory: Path,
) -> dict[str, Any]:
    protocol, root, output = _preflight_paths_before_read(
        protocol_path, data_root, output_directory
    )
    return _qualify_preflighted(
        protocol_path=protocol,
        protocol_sha256=protocol_sha256,
        data_root=root,
        output_directory=output,
    )


def execute_qualification(
    *,
    protocol_path: Path,
    protocol_sha256: str,
    data_root: Path,
    output_directory: Path,
) -> dict[str, Any]:
    # Forbidden path rejection deliberately occurs outside the failure publisher.
    protocol, root, output = _preflight_paths_before_read(
        protocol_path, data_root, output_directory
    )
    try:
        return _qualify_preflighted(
            protocol_path=protocol,
            protocol_sha256=protocol_sha256,
            data_root=root,
            output_directory=output,
        )
    except (PublicationError, ProtocolError):
        raise
    except QualificationError as exc:
        return _publish_closed_bundle(
            output, _failure_payload(exc.code), outcome=FAILURE_OUTCOME
        )


def _parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--protocol", required=True, type=Path)
    parser.add_argument("--protocol-sha256", required=True)
    parser.add_argument("--data-root", required=True, type=Path)
    parser.add_argument("--output-directory", required=True, type=Path)
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = _parse_args(argv)
    try:
        result = execute_qualification(
            protocol_path=args.protocol,
            protocol_sha256=args.protocol_sha256,
            data_root=args.data_root,
            output_directory=args.output_directory,
        )
    except PartialPrecommitError:
        print(
            json.dumps(
                {
                    "publication_state": "PARTIAL_PRECOMMIT",
                    "committed": False,
                    "qualified": False,
                    "canonical_record_count": 0,
                    "training_allowed": False,
                    "next_phase_authorized": False,
                },
                sort_keys=True,
            )
        )
        return 3
    except QualificationError as exc:
        print(
            json.dumps(
                {
                    "publication_state": "FAIL_CLOSED",
                    "failure_code": exc.code,
                    "committed": False,
                    "qualified": False,
                    "canonical_record_count": 0,
                    "training_allowed": False,
                    "next_phase_authorized": False,
                },
                sort_keys=True,
            )
        )
        return 2
    print(json.dumps(result, sort_keys=True))
    return 0 if result["execution_outcome"] == SUCCESS_OUTCOME else 2


if __name__ == "__main__":
    raise SystemExit(main())
