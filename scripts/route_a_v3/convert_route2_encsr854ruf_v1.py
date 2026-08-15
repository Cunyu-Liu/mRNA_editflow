#!/usr/bin/env python3
"""Convert public ENCSR854RUF author FASTA pairs to six-context Route 2 records."""

from __future__ import annotations

import argparse
import json
import math
import os
import shutil
import tempfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
CONFIG_PATH = REPO_ROOT / "configs/route_a_v3_route2_encsr854ruf_converter_v1.json"
CANONICAL_SCHEMA_VERSION = "mrna_editflow_route2_canonical.v1"
BASES = set("ACGT")
COMPLEMENT = str.maketrans("ACGT", "TGCA")


class ConversionError(RuntimeError):
    pass


def _require(condition: bool, message: str) -> None:
    if not condition:
        raise ConversionError(message)


def _finite(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(value)


def load_config(path: Path = CONFIG_PATH) -> dict[str, Any]:
    config = json.loads(path.read_text(encoding="utf-8"))
    validate_config(config)
    return config


def validate_config(config: Mapping[str, Any]) -> None:
    _require(config["schema_version"] == "route_a_v3_route2_encsr854ruf_converter.v1", "unexpected schema version")
    study = config["study"]
    _require(study["study_unit_id"] == "ENCSR854RUF", "unexpected study")
    _require(study["pool_assignment"] == "DEVELOPMENT" and study["qualification_class"] == "DEVELOPMENT_RELAXED", "study pool changed")
    _require(study["study_role"] == "A1_EXPOSURE_POSITIVE_STRICT_SUB_DEVELOPMENT", "study role changed")
    _require(study["conversion_scope"] == "PUBLIC_AUTHOR_FASTA_FINITE_ALL_SIX_CONTEXT_SUB_ONLY", "scope changed")
    expected = config["input"]
    _require(expected["expected_published_pair_count"] == 15266, "published pair count changed")
    _require(expected["expected_complete_six_context_sub_pair_count"] == 13302, "eligible SUB pair count changed")
    _require(expected["expected_canonical_record_count"] == 79800, "canonical count changed")
    _require(expected["expected_sequence_length"] == 133, "sequence length changed")
    _require(expected["fasta_alias_expansion_rule"] == "SPLIT_HEADER_ON_SLASH", "alias rule changed")
    _require(expected["index_error_rule"] == "OTHER_VAR_IN_OLIGO_WINDOW_CONTAINS_CASE_INSENSITIVE_INDEX_ERROR_TOKEN", "index rule changed")
    _require(config["action_policy"] == {"allowed_candidate_action": "SUB", "ins_supported": False, "del_supported": False}, "action policy changed")
    output = config["output"]
    _require(output["overwrite_allowed"] is False and output["public_redistribution_allowed"] is False, "output rights changed")
    _require(output["directory"].startswith("/mnt/cunyuliu/mrna_xeditflow_routea_v3/route2/"), "output leaves Route 2 root")
    development = config["development_policy"]
    _require(development["training_eligible"] is True and development["model_selection_eligible"] is True, "Development disabled")
    _require(development["historical_exposure_status"] == "KNOWN_EXPOSED_DEVELOPMENT_ONLY", "exposure understated")
    _require(development["near_duplicate_split_status"] == "NOT_RUN", "split overstated")
    _require(development["source_endpoint_value_representation"] is None and development["candidate_endpoint_value_representation"] is None, "absolute endpoints fabricated")
    _require(development["missing_endpoint_is_zero"] is False, "missing endpoint changed to zero")
    credit = config["credit_policy"]
    _require(credit["measured_candidate"] is True and credit["generated_candidate"] is False, "candidate role changed")
    _require(not any(credit["qualified_credit_delta"].values()), "conversion increases qualified credit")
    _require(credit["qualified_counts_after_conversion"] == {"ordinary": 1, "a1": 1, "true_a2": 0, "canonical_records": 6547}, "qualified facts changed")
    _require(config["scientific_claim_status"] == "NOT_ESTABLISHED", "scientific claim overstated")


def _load_fasta_aliases(path: Path, expected_length: int) -> tuple[dict[str, str], Counter[str], int]:
    records: list[tuple[str, str]] = []
    header: str | None = None
    parts: list[str] = []
    with path.open(encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if line.startswith(">"):
                if header is not None:
                    records.append((header, "".join(parts).upper()))
                header = line[1:]
                parts = []
            elif line:
                parts.append(line)
    if header is not None:
        records.append((header, "".join(parts).upper()))
    alias_values: dict[str, set[str]] = defaultdict(set)
    stats: Counter[str] = Counter()
    for fasta_header, sequence in records:
        stats["header_count"] += 1
        valid = len(sequence) == expected_length and not (set(sequence) - BASES)
        stats["valid_record_count" if valid else "invalid_record_count"] += 1
        for alias in fasta_header.split("/"):
            stats["expanded_alias_token_count"] += 1
            if valid:
                alias_values[alias].add(sequence)
                stats["valid_alias_token_count"] += 1
    conflicts = sum(len(values) > 1 for values in alias_values.values())
    aliases = {alias: next(iter(values)) for alias, values in alias_values.items() if len(values) == 1}
    return aliases, stats, conflicts


def _rows(sheet) -> tuple[list[str], list[tuple[int, dict[str, Any]]]]:
    iterator = sheet.iter_rows(values_only=True)
    header = list(next(iterator))
    return header, [(number, dict(zip(header, values))) for number, values in enumerate(iterator, start=2)]


def _pair_geometry(workbook, aliases: Mapping[str, str]) -> tuple[dict[str, dict[str, Any]], set[str], Counter[str]]:
    header, rows = _rows(workbook["Oligo Variant Info"])
    required = {"mpra_variant_id", "tag", "oligo_id", "ref_allele", "alt_allele", "strand", "other_var_in_oligo_window"}
    _require(required.issubset(header), f"oligo columns absent: {sorted(required - set(header))}")
    grouped: dict[str, list[tuple[int, dict[str, Any]]]] = defaultdict(list)
    index_errors: set[str] = set()
    for row_number, row in rows:
        pair_id = str(row["mpra_variant_id"])
        grouped[pair_id].append((row_number, row))
        tokens = {part.strip().lower() for part in str(row.get("other_var_in_oligo_window") or "").split(",")}
        if "index_error" in tokens:
            index_errors.add(pair_id)
    stats: Counter[str] = Counter()
    pairs: dict[str, dict[str, Any]] = {}
    metadata_fields = ("variant_id", "chrom", "oligo_starts", "oligo_ends", "strand", "var_start", "var_end", "ref_allele", "alt_allele", "genes", "transcripts", "gene_symbols")
    for pair_id, members in grouped.items():
        if len(members) == 2:
            stats["pair_size_two_count"] += 1
        tags = Counter(str(row["tag"]).lower() for _, row in members)
        if len(members) != 2 or tags != Counter({"ref": 1, "alt": 1}):
            stats["invalid_pair_geometry_count"] += 1
            continue
        by_tag = {str(row["tag"]).lower(): (number, row) for number, row in members}
        ref_sequence = aliases.get(str(by_tag["ref"][1]["oligo_id"]))
        alt_sequence = aliases.get(str(by_tag["alt"][1]["oligo_id"]))
        if ref_sequence is None or alt_sequence is None:
            stats["crosswalk_missing_pair_count"] += 1
            continue
        if any(by_tag["ref"][1][field] != by_tag["alt"][1][field] for field in metadata_fields):
            stats["metadata_disagreement_pair_count"] += 1
            continue
        if ref_sequence == alt_sequence:
            stats["sequence_identical_pair_count"] += 1
            continue
        distance = sum(left != right for left, right in zip(ref_sequence, alt_sequence)) if len(ref_sequence) == len(alt_sequence) else None
        action = "SUB" if distance == 1 else "OTHER"
        stats[f"action_{action}"] += 1
        edit_position = None
        if action == "SUB":
            edit_position = next(index for index, (left, right) in enumerate(zip(ref_sequence, alt_sequence)) if left != right)
            declared_ref = str(by_tag["ref"][1]["ref_allele"]).upper()
            declared_alt = str(by_tag["ref"][1]["alt_allele"]).upper()
            if str(by_tag["ref"][1]["strand"]) == "-":
                declared_ref = declared_ref.translate(COMPLEMENT)[::-1]
                declared_alt = declared_alt.translate(COMPLEMENT)[::-1]
            if len(declared_ref) != 1 or len(declared_alt) != 1 or ref_sequence[edit_position] != declared_ref or alt_sequence[edit_position] != declared_alt:
                stats["sub_declared_allele_replay_mismatch_count"] += 1
        pairs[pair_id] = {
            "pair_id": pair_id,
            "source_sequence": ref_sequence,
            "candidate_sequence": alt_sequence,
            "action": action,
            "edit_position_zero_based": edit_position,
            "supplement_row_number": by_tag["ref"][0],
            "metadata": by_tag["ref"][1],
        }
    stats["published_pair_count"] = len(grouped)
    return pairs, index_errors, stats


def _eligible_pairs(config: Mapping[str, Any], workbook, pairs: Mapping[str, dict[str, Any]], index_errors: set[str]) -> tuple[list[dict[str, Any]], Counter[str], int]:
    header, rows = _rows(workbook["Variant MPRAu Results"])
    contexts = config["study"]["biological_context_ids"]
    required = {"mpra_variant_id"}
    for context in contexts:
        required.update({f"log2FoldChange_Skew_{context}", f"lfcSE_Skew_{context}"})
    _require(required.issubset(header), f"result columns absent: {sorted(required - set(header))}")
    results = {str(row["mpra_variant_id"]): row for _, row in rows}
    _require(set(results) == set(pairs), "result and pair universes differ")
    stats: Counter[str] = Counter(result_row_count=len(rows))
    negative_se_count = 0
    eligible = []
    for pair_id, pair in pairs.items():
        if pair_id in index_errors:
            stats["index_error_pair_count"] += 1
            continue
        result = results[pair_id]
        endpoints = {}
        complete = True
        for context in contexts:
            effect = result[f"log2FoldChange_Skew_{context}"]
            standard_error = result[f"lfcSE_Skew_{context}"]
            if not (_finite(effect) and _finite(standard_error)):
                complete = False
                break
            negative_se_count += standard_error < 0
            endpoints[context] = {"effect": float(effect), "standard_error": float(standard_error)}
        if not complete:
            stats["incomplete_six_context_effect_or_se_pair_count"] += 1
            continue
        stats[f"complete_six_context_{pair['action'].lower()}_pair_count"] += 1
        if pair["action"] == "SUB":
            value = dict(pair)
            value["endpoints"] = endpoints
            eligible.append(value)
    return eligible, stats, negative_se_count


def _resolve_units(eligible_pairs: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, int]]:
    units: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for pair in eligible_pairs:
        for context, endpoint in pair["endpoints"].items():
            value = dict(pair)
            value.update({"biological_context_id": context, **endpoint})
            units[(pair["source_sequence"], pair["candidate_sequence"], context)].append(value)
    resolved = []
    duplicate_units = conflict_units = conflict_rows = 0
    for key in sorted(units):
        values = units[key]
        distinct = {(round(value["effect"], 12), round(value["standard_error"], 12)) for value in values}
        if len(values) > 1 and len(distinct) > 1:
            conflict_units += 1
            conflict_rows += len(values)
            continue
        if len(values) > 1:
            duplicate_units += 1
        resolved.append(sorted(values, key=lambda value: (value["supplement_row_number"], value["pair_id"]))[0])
    return resolved, {
        "endpoint_slot_count_before_exact_dedup": sum(len(values) for values in units.values()),
        "exact_duplicate_unit_count": duplicate_units,
        "conflicting_exact_unit_count": conflict_units,
        "conflicting_endpoint_row_count": conflict_rows,
        "distinct_sequence_pair_count": len({(value["source_sequence"], value["candidate_sequence"]) for value in resolved}),
    }


def _canonical_records(config: Mapping[str, Any], resolved: list[dict[str, Any]]) -> list[dict[str, Any]]:
    source_groups = {source: f"ENCSR854RUF_SOURCE_GROUP_{index:05d}" for index, source in enumerate(sorted({value["source_sequence"] for value in resolved}), start=1)}
    development = config["development_policy"]
    credit = config["credit_policy"]
    replicate_counts = config["published_biological_replicate_counts"]
    records = []
    for value in sorted(resolved, key=lambda item: (item["pair_id"], item["biological_context_id"])):
        group_id = source_groups[value["source_sequence"]]
        position = value["edit_position_zero_based"]
        metadata = value["metadata"]
        records.append({
            "schema_version": CANONICAL_SCHEMA_VERSION,
            "canonical_record_id": f"ENCSR854RUF:{value['pair_id']}:context:{value['biological_context_id']}",
            "study_unit_id": "ENCSR854RUF",
            "pool_assignment": config["study"]["pool_assignment"],
            "qualification_class": config["study"]["qualification_class"],
            "study_role": config["study"]["study_role"],
            "conversion_scope": config["study"]["conversion_scope"],
            "region": "3UTR",
            "biological_context_id": value["biological_context_id"],
            "assay_id": config["study"]["assay_id"],
            "endpoint_id": config["study"]["endpoint_id"],
            "endpoint_direction": config["study"]["endpoint_direction"],
            "source_id": group_id,
            "source_sequence": value["source_sequence"],
            "source_endpoint_value": development["source_endpoint_value_representation"],
            "candidate_id": f"ENCSR854RUF:{value['pair_id']}:candidate",
            "candidate_sequence": value["candidate_sequence"],
            "candidate_endpoint_value": development["candidate_endpoint_value_representation"],
            "edit_operations": [{"type": "SUB", "position_zero_based": position, "ref": value["source_sequence"][position], "alt": value["candidate_sequence"][position]}],
            "direction_normalized_delta": value["effect"],
            "biological_standard_error": value["standard_error"],
            "standard_error_status": development["reported_standard_error_status"],
            "published_biological_replicate_count": replicate_counts[value["biological_context_id"]],
            "effective_n": None,
            "group_id": group_id,
            "study_group_id": "ENCSR854RUF_TEWHEY_2021",
            "pair_id": value["pair_id"],
            "variant_id": metadata["variant_id"],
            "gene_symbols": metadata["gene_symbols"],
            "strand": metadata["strand"],
            "supplement_row_number": value["supplement_row_number"],
            "measured_candidate": credit["measured_candidate"],
            "generated_candidate": credit["generated_candidate"],
            "training_eligible": development["training_eligible"],
            "model_selection_eligible": development["model_selection_eligible"],
            "historical_exposure_status": development["historical_exposure_status"],
            "near_duplicate_split_status": development["near_duplicate_split_status"],
        })
    return records


def _write_json(path: Path, value: Mapping[str, Any]) -> None:
    path.write_text(json.dumps(value, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def execute(config: Mapping[str, Any], workbook_path: Path, fasta_path: Path, output_dir: Path) -> dict[str, Any]:
    _require(not output_dir.exists(), f"output directory already exists: {output_dir}")
    _require(workbook_path.is_file(), f"publisher workbook absent: {workbook_path}")
    _require(fasta_path.is_file(), f"author FASTA absent: {fasta_path}")
    expected = config["input"]
    aliases, fasta_stats, alias_conflicts = _load_fasta_aliases(fasta_path, expected["expected_sequence_length"])
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    pairs, index_errors, pair_stats = _pair_geometry(workbook, aliases)
    eligible, eligibility_stats, negative_se_count = _eligible_pairs(config, workbook, pairs, index_errors)
    resolved, unit_stats = _resolve_units(eligible)
    records = _canonical_records(config, resolved)
    context_counts = Counter(record["biological_context_id"] for record in records)
    observed = {
        "workbook_sheet_count": len(workbook.sheetnames),
        "fasta_header_count": fasta_stats["header_count"],
        "fasta_valid_record_count": fasta_stats["valid_record_count"],
        "fasta_expanded_alias_token_count": fasta_stats["expanded_alias_token_count"],
        "fasta_valid_alias_token_count": fasta_stats["valid_alias_token_count"],
        "fasta_invalid_record_count": fasta_stats["invalid_record_count"],
        "fasta_alias_conflict_count": alias_conflicts,
        "published_pair_count": pair_stats["published_pair_count"],
        "pair_size_two_count": pair_stats["pair_size_two_count"],
        "crosswalk_missing_pair_count": pair_stats["crosswalk_missing_pair_count"],
        "metadata_disagreement_pair_count": pair_stats["metadata_disagreement_pair_count"],
        "sequence_identical_pair_count": pair_stats["sequence_identical_pair_count"],
        "action_counts": {"OTHER": pair_stats["action_OTHER"], "SUB": pair_stats["action_SUB"]},
        "sub_declared_allele_replay_mismatch_count": pair_stats["sub_declared_allele_replay_mismatch_count"],
        "result_row_count": eligibility_stats["result_row_count"],
        "index_error_pair_count": eligibility_stats["index_error_pair_count"],
        "incomplete_six_context_effect_or_se_pair_count": eligibility_stats["incomplete_six_context_effect_or_se_pair_count"],
        "complete_six_context_sub_pair_count": eligibility_stats["complete_six_context_sub_pair_count"],
        "complete_six_context_other_action_pair_count": eligibility_stats["complete_six_context_other_pair_count"],
        "negative_standard_error_count": negative_se_count,
        **unit_stats,
        "canonical_record_count": len(records),
        "context_record_counts": dict(sorted(context_counts.items())),
        "distinct_source_group_count": len({record["source_sequence"] for record in records}),
    }
    exact = (
        observed["workbook_sheet_count"] == expected["expected_workbook_sheet_count"]
        and observed["fasta_header_count"] == expected["expected_fasta_header_count"]
        and observed["fasta_valid_record_count"] == expected["expected_fasta_valid_record_count"]
        and observed["fasta_expanded_alias_token_count"] == expected["expected_fasta_expanded_alias_token_count"]
        and observed["fasta_valid_alias_token_count"] == expected["expected_fasta_valid_alias_token_count"]
        and observed["fasta_invalid_record_count"] == expected["expected_fasta_invalid_record_count"]
        and observed["fasta_alias_conflict_count"] == expected["expected_fasta_alias_conflict_count"]
        and observed["published_pair_count"] == expected["expected_published_pair_count"]
        and observed["pair_size_two_count"] == expected["expected_pair_size_two_count"]
        and observed["crosswalk_missing_pair_count"] == expected["expected_crosswalk_missing_pair_count"]
        and observed["metadata_disagreement_pair_count"] == expected["expected_metadata_disagreement_pair_count"]
        and observed["sequence_identical_pair_count"] == expected["expected_sequence_identical_pair_count"]
        and observed["action_counts"] == expected["expected_action_counts"]
        and observed["sub_declared_allele_replay_mismatch_count"] == expected["expected_sub_declared_allele_replay_mismatch_count"]
        and observed["result_row_count"] == expected["expected_result_row_count"]
        and observed["index_error_pair_count"] == expected["expected_index_error_pair_count"]
        and observed["incomplete_six_context_effect_or_se_pair_count"] == expected["expected_incomplete_six_context_effect_or_se_pair_count"]
        and observed["complete_six_context_sub_pair_count"] == expected["expected_complete_six_context_sub_pair_count"]
        and observed["complete_six_context_other_action_pair_count"] == expected["expected_complete_six_context_other_action_pair_count"]
        and observed["negative_standard_error_count"] == expected["expected_negative_standard_error_count"]
        and observed["endpoint_slot_count_before_exact_dedup"] == expected["expected_endpoint_slot_count_before_exact_dedup"]
        and observed["exact_duplicate_unit_count"] == expected["expected_exact_duplicate_unit_count"]
        and observed["conflicting_exact_unit_count"] == expected["expected_conflicting_exact_unit_count"]
        and observed["distinct_sequence_pair_count"] == expected["expected_distinct_sequence_pair_count"]
        and observed["canonical_record_count"] == expected["expected_canonical_record_count"]
        and set(observed["context_record_counts"].values()) == {expected["expected_context_record_count_each"]}
        and set(observed["context_record_counts"]) == set(config["study"]["biological_context_ids"])
        and observed["distinct_source_group_count"] == expected["expected_distinct_source_group_count"]
    )
    status = "CONVERTED_DEVELOPMENT_RELAXED_EXPOSURE_POSITIVE_SIX_CONTEXT_SUB_ONLY" if exact else "UNCONVERTIBLE_FOR_ROUTE2_V1_GEOMETRY_MISMATCH"
    output_dir.parent.mkdir(parents=True, exist_ok=True)
    temporary = Path(tempfile.mkdtemp(prefix=f".{output_dir.name}.", dir=output_dir.parent))
    try:
        canonical_path = temporary / config["output"]["canonical_filename"]
        with canonical_path.open("w", encoding="utf-8") as handle:
            if exact:
                for record in records:
                    handle.write(json.dumps(record, sort_keys=True) + "\n")
        summary = {
            "converter_id": config["converter_id"], "study_unit_id": "ENCSR854RUF",
            "pool_assignment": "DEVELOPMENT", "qualification_class": "DEVELOPMENT_RELAXED",
            "conversion_scope": config["study"]["conversion_scope"], "status": status, **observed,
            "measured_candidate_count": len(records) if exact else 0, "generated_candidate_count": 0,
            "qualified_credit_delta": config["credit_policy"]["qualified_credit_delta"],
            "qualified_counts_after_conversion": config["credit_policy"]["qualified_counts_after_conversion"],
            "public_redistribution_allowed": config["output"]["public_redistribution_allowed"],
            "historical_exposure_status": config["development_policy"]["historical_exposure_status"],
            "near_duplicate_split_status": config["development_policy"]["near_duplicate_split_status"],
            "limitations": config["limitations"], "scientific_claim_status": config["scientific_claim_status"],
        }
        reject_summary = {
            "converter_id": config["converter_id"], "study_unit_id": "ENCSR854RUF",
            "other_action_pair_count": observed["action_counts"]["OTHER"],
            "publisher_index_error_pair_count": observed["index_error_pair_count"],
            "incomplete_six_context_effect_or_se_pair_count": observed["incomplete_six_context_effect_or_se_pair_count"],
            "exact_duplicate_unit_count_collapsed": observed["exact_duplicate_unit_count"],
            "conflicting_exact_unit_count": observed["conflicting_exact_unit_count"],
            "reject_payload_in_summary": False,
        }
        _write_json(temporary / config["output"]["conversion_summary_filename"], summary)
        _write_json(temporary / config["output"]["reject_summary_filename"], reject_summary)
        os.rename(temporary, output_dir)
        return summary
    except Exception:
        shutil.rmtree(temporary, ignore_errors=True)
        raise


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", type=Path, default=CONFIG_PATH)
    parser.add_argument("--workbook", type=Path)
    parser.add_argument("--fasta", type=Path)
    parser.add_argument("--output-dir", type=Path)
    args = parser.parse_args()
    config = load_config(args.config)
    summary = execute(config, args.workbook or Path(config["input"]["publisher_workbook_path"]), args.fasta or Path(config["input"]["author_fasta_path"]), args.output_dir or Path(config["output"]["directory"]))
    print(json.dumps(summary, sort_keys=True))
    return 0 if summary["status"].startswith("CONVERTED_") else 2


if __name__ == "__main__":
    raise SystemExit(main())
