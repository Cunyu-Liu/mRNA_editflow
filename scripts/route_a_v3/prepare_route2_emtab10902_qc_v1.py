#!/usr/bin/env python3
"""Prepare and close the outcome-blind E-MTAB-10902 N-zip read-count QC gate."""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


EXPECTED_DESIGN_ROWS = 6266
EXPECTED_UNIQUE_SEQUENCES = 6260
PUBLISHER_REPORTED_PASSING_DESIGNS = 5679
MINIMUM_READS = 20
MINIMUM_PASSING_SAMPLES = 3
SAMPLE_IDS = (
    "ERR7337821",
    "ERR7337822",
    "ERR7337823",
    "ERR7337824",
    "ERR7337825",
    "ERR7337826",
)
BASES = set("ACGT")


class QcPreparationError(RuntimeError):
    pass


def _require(condition: bool, message: str) -> None:
    if not condition:
        raise QcPreparationError(message)


def load_design_rows(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    _require("Supp_Table_2a" in workbook.sheetnames, "Supplementary Table 2a is absent")
    rows = workbook["Supp_Table_2a"].iter_rows(values_only=True)
    next(rows)
    next(rows)
    header = tuple(next(rows))
    required = {
        "Source gene id",
        "Source gene name",
        "Source tile id",
        "Mutation type",
        "Mutation position",
        "size",
        "Sequence",
    }
    _require(required <= set(header), "Supplementary Table 2a schema changed")
    result = [dict(zip(header, row)) for row in rows if any(value is not None for value in row)]
    _require(len(result) == EXPECTED_DESIGN_ROWS, "Supplementary Table 2a design count changed")
    return result


def normalized_sequence(row: dict[str, Any], row_number: int) -> str:
    sequence = str(row["Sequence"]).upper()
    _require(len(sequence) == int(row["size"]), f"design length differs at row {row_number}")
    _require(len(sequence) in {85, 90, 100}, f"unsupported design length at row {row_number}")
    _require(not (set(sequence) - BASES), f"invalid design alphabet at row {row_number}")
    return sequence


def build_library(workbook_path: Path, output_dir: Path) -> dict[str, Any]:
    _require(not output_dir.exists(), f"output directory already exists: {output_dir}")
    rows = load_design_rows(workbook_path)
    by_sequence: dict[str, list[int]] = defaultdict(list)
    for row_number, row in enumerate(rows, start=1):
        by_sequence[normalized_sequence(row, row_number)].append(row_number)
    _require(len(by_sequence) == EXPECTED_UNIQUE_SEQUENCES, "unique design sequence count changed")

    output_dir.mkdir(parents=True)
    fasta_path = output_dir / "emtab10902_nzip_unique_designs.fa"
    map_path = output_dir / "design_row_to_sequence_id.jsonl"
    summary_path = output_dir / "preparation_summary.json"
    sequence_id_by_sequence: dict[str, str] = {}
    with fasta_path.open("w", encoding="ascii") as handle:
        for sequence_number, sequence in enumerate(sorted(by_sequence), start=1):
            sequence_id = f"NZSEQ{sequence_number:05d}"
            sequence_id_by_sequence[sequence] = sequence_id
            handle.write(f">{sequence_id}\n{sequence}\n")
    with map_path.open("w", encoding="utf-8") as handle:
        for row_number, row in enumerate(rows, start=1):
            sequence = normalized_sequence(row, row_number)
            payload = {
                "design_row_number": row_number,
                "sequence_id": sequence_id_by_sequence[sequence],
                "source_gene_id": str(row["Source gene id"]),
                "source_gene_name": str(row["Source gene name"]),
                "source_tile_id": str(row["Source tile id"]),
                "mutation_type": str(row["Mutation type"]),
                "mutation_position": str(row["Mutation position"]),
            }
            handle.write(json.dumps(payload, sort_keys=True) + "\n")

    duplicate_groups = [members for members in by_sequence.values() if len(members) > 1]
    summary = {
        "schema_version": "route_a_v3_route2_emtab10902_qc_preparation.v1",
        "status": "COUNTING_LIBRARY_PREPARED_OUTCOME_BLIND",
        "source_sheet": "Supp_Table_2a",
        "evaluation_outcome_sheet_read": False,
        "design_row_count": len(rows),
        "unique_sequence_count": len(by_sequence),
        "duplicate_sequence_group_count": len(duplicate_groups),
        "duplicate_design_row_count": sum(len(group) for group in duplicate_groups),
        "fasta_filename": fasta_path.name,
        "mapping_filename": map_path.name,
    }
    summary_path.write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return summary


def load_mapping(path: Path) -> list[dict[str, Any]]:
    rows = [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines()]
    _require(len(rows) == EXPECTED_DESIGN_ROWS, "design-to-sequence mapping count changed")
    _require({row["design_row_number"] for row in rows} == set(range(1, EXPECTED_DESIGN_ROWS + 1)), "design mapping row numbers changed")
    _require(len({row["sequence_id"] for row in rows}) == EXPECTED_UNIQUE_SEQUENCES, "design mapping sequence count changed")
    return rows


def load_read_counts(path: Path) -> dict[str, int]:
    result: dict[str, int] = {}
    with path.open(encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, start=1):
            fields = line.rstrip("\n").split("\t")
            _require(len(fields) >= 3, f"count row width changed in {path.name}:{line_number}")
            sequence_id, _umi_count, read_count = fields[:3]
            _require(sequence_id not in result, f"duplicate sequence id in {path.name}")
            try:
                result[sequence_id] = int(read_count)
            except ValueError as exc:
                raise QcPreparationError(f"noninteger read count in {path.name}:{line_number}") from exc
            _require(result[sequence_id] >= 0, f"negative read count in {path.name}:{line_number}")
    _require(len(result) == EXPECTED_UNIQUE_SEQUENCES, f"count member count changed in {path.name}")
    return result


def close_qc(mapping_path: Path, count_specs: list[str], output_dir: Path) -> dict[str, Any]:
    _require(not output_dir.exists(), f"output directory already exists: {output_dir}")
    parsed: dict[str, Path] = {}
    for spec in count_specs:
        sample_id, separator, raw_path = spec.partition("=")
        _require(separator == "=" and sample_id and raw_path, f"invalid count specification: {spec}")
        _require(sample_id not in parsed, f"duplicate count sample: {sample_id}")
        parsed[sample_id] = Path(raw_path)
    _require(set(parsed) == set(SAMPLE_IDS), "count sample set differs from the six primary WT secondary N-zip samples")

    mapping = load_mapping(mapping_path)
    counts_by_sample = {sample_id: load_read_counts(parsed[sample_id]) for sample_id in SAMPLE_IDS}
    expected_ids = {row["sequence_id"] for row in mapping}
    for sample_id, counts in counts_by_sample.items():
        _require(set(counts) == expected_ids, f"count sequence ids changed in {sample_id}")

    output_dir.mkdir(parents=True)
    membership_path = output_dir / "publisher_read_qc_membership.jsonl"
    passed_designs = 0
    passed_sequence_ids: set[str] = set()
    with membership_path.open("w", encoding="utf-8") as handle:
        for row in mapping:
            sequence_id = row["sequence_id"]
            read_counts = {sample_id: counts_by_sample[sample_id][sequence_id] for sample_id in SAMPLE_IDS}
            samples_at_least_20 = sum(value >= MINIMUM_READS for value in read_counts.values())
            passes = samples_at_least_20 >= MINIMUM_PASSING_SAMPLES
            passed_designs += int(passes)
            if passes:
                passed_sequence_ids.add(sequence_id)
            payload = {
                **row,
                "read_counts": read_counts,
                "samples_at_least_20_reads": samples_at_least_20,
                "passes_publisher_read_qc": passes,
            }
            handle.write(json.dumps(payload, sort_keys=True) + "\n")

    summary = {
        "schema_version": "route_a_v3_route2_emtab10902_qc_membership.v1",
        "status": "PUBLISHER_READ_QC_REPRODUCED" if passed_designs == PUBLISHER_REPORTED_PASSING_DESIGNS else "PUBLISHER_READ_QC_COUNT_MISMATCH",
        "evaluation_outcome_sheet_read": False,
        "sample_ids": list(SAMPLE_IDS),
        "minimum_reads_per_sample": MINIMUM_READS,
        "minimum_passing_sample_count": MINIMUM_PASSING_SAMPLES,
        "design_row_count": len(mapping),
        "unique_sequence_count": len(expected_ids),
        "passed_design_row_count": passed_designs,
        "passed_unique_sequence_count": len(passed_sequence_ids),
        "publisher_reported_passed_design_count": PUBLISHER_REPORTED_PASSING_DESIGNS,
        "publisher_reported_count_reproduced": passed_designs == PUBLISHER_REPORTED_PASSING_DESIGNS,
        "membership_filename": membership_path.name,
    }
    (output_dir / "qc_membership_summary.json").write_text(
        json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    return summary


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)
    build = subparsers.add_parser("build-library")
    build.add_argument("--workbook", type=Path, required=True)
    build.add_argument("--output-dir", type=Path, required=True)
    close = subparsers.add_parser("close-qc")
    close.add_argument("--mapping", type=Path, required=True)
    close.add_argument("--count", action="append", default=[])
    close.add_argument("--output-dir", type=Path, required=True)
    args = parser.parse_args()
    if args.command == "build-library":
        summary = build_library(args.workbook, args.output_dir)
    else:
        summary = close_qc(args.mapping, args.count, args.output_dir)
    print(json.dumps(summary, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
