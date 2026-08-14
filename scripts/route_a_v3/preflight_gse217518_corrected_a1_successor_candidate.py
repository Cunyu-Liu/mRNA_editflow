#!/usr/bin/env python3
"""Aggregate-only corrected GSE217518 A1 successor candidate.

This candidate separates the published core half-life QC (R2 > 0.5 and
MSE < 1) from outcome-dependent filters used only in downstream reference-
sequence motif analyses.  It audits the exact eleven DEC027 gate IDs without
serialising member identifiers, sequences, row effects, row standard errors,
or split assignments.

DEC027 authority A, the settled A1-EVT-059 authority-runtime I/B predecessor,
and the first GSE217518 implementation commit I1 are frozen and byte-bound.  The
replacement I2 candidate's own four-scalar implementation group remains grouped
UNKNOWN, so the production entry point fails before Git, official-asset, or
output I/O.  Once B2 separately binds I2, production verifies the entire
A -> runtime I/B -> I1 -> I2/B2 Git chain, the executing I2 script bytes, and a
clean HEAD equal to upstream and the live origin.  Pure in-memory evaluation
changes no qualification, credit, canonical, training, GPU, model, or A7 state.
"""

from __future__ import annotations

import argparse
import csv
import gzip
import hashlib
import json
import math
import os
import re
import subprocess
import tempfile
from collections import Counter, defaultdict
from copy import deepcopy
from pathlib import Path
from typing import Any, Mapping, Sequence

import openpyxl


STAGING_ROOT = Path(__file__).resolve().parents[2]
CONFIG_PATH = (
    STAGING_ROOT
    / "configs/route_a_v3_gse217518_corrected_a1_successor_candidate_v1.json"
)

SCHEMA_VERSION = "route_a_v3_gse217518_corrected_a1_successor_candidate.v1"
PROTOCOL_ID = "GSE217518_CORRECTED_A1_SUCCESSOR_CANDIDATE_V1"
OBSERVATION_SCHEMA = "route_a_v3_gse217518_corrected_a1_public_aggregate_observation.v1"
REPORT_FILENAME = "GSE217518_CORRECTED_A1_SUCCESSOR_AGGREGATE_PREFLIGHT_V1.json"
UNKNOWN = "UNKNOWN_NOT_ASSERTED"
BOUND = "BOUND"
PASS = "PASS"
BLOCKED = "BLOCKED"
NOT_RUN = "NOT_RUN"

PRODUCTION_REPO_ROOT = Path(
    "/home/cunyuliu/mrna_editflow_goal/worktrees/routea_v3_a1_20260810"
)
PRODUCTION_BRANCH = "routea-v3-a1-20260810"
PRODUCTION_UPSTREAM = f"origin/{PRODUCTION_BRANCH}"
AUTHORITY_COMMIT = "3e0ad158a0b45b2f26ed82da3afe60667c712cd6"
AUTHORITY_PARENT = "b1ca33d852bad111ff31b4f60493d8c43c63d1a3"
AUTHORITY_EXACT12: tuple[str, ...] = (
    "configs/route_a_v3.yaml",
    "configs/route_a_v3_a1_qualification.json",
    "docs/contracts/amendments/mrna_xeditflow_route_a_v3_dec027.yaml",
    "docs/contracts/supersession_mrna_xeditflow_v1_1_to_route_a_v3.yaml",
    "docs/execution/route_a_v3_a1_interim.yaml",
    "docs/execution/route_a_v3_a6_interim.yaml",
    "docs/execution/route_a_v3_data_role_registry.yaml",
    "docs/execution/route_a_v3_decision_log.yaml",
    "docs/execution/route_a_v3_registry_manifest.json",
    "docs/execution/route_a_v3_task_registry.yaml",
    "scripts/route_a_v3/validate_a0_bundle.py",
    "tests/route_a_v3/test_a0_integrity_guards.py",
)
AUTHORITY_BLOBS: dict[str, str] = {
    "configs/route_a_v3.yaml": (
        "c5ec7d236443b506c09fd3f09e149ce5d082daff618887989af6e59472727a27"
    ),
    "configs/route_a_v3_a1_qualification.json": (
        "261339c38f4b8bbd48bf8f63f6a588be57af9f6229119e84bf661d7ee8f855db"
    ),
    "docs/contracts/amendments/mrna_xeditflow_route_a_v3_dec027.yaml": (
        "2a27c296539e8e665873363778d91cc223f56f933a815e9509b10a7267f6b5c4"
    ),
    "docs/contracts/supersession_mrna_xeditflow_v1_1_to_route_a_v3.yaml": (
        "1c4b6e29c09eb24798207138047b68909d7bda8bacc3a2eab8e17a7ca789b44b"
    ),
    "docs/execution/route_a_v3_a1_interim.yaml": (
        "fb50929ae2dfa0bdd1c50b003fc43c0e13b89baade203e8384c8b6ea3eba7b1e"
    ),
    "docs/execution/route_a_v3_a6_interim.yaml": (
        "1d44bcfe8669a55dc42f619ed43178f0637e20a4297ca996ecab3f7165612769"
    ),
    "docs/execution/route_a_v3_data_role_registry.yaml": (
        "80217a8114286f84960237819ac5b2d5828afbc23af118576541cc7cee64ae4e"
    ),
    "docs/execution/route_a_v3_decision_log.yaml": (
        "e0dc2a7fb186c5c8d00c1c5604602b1b2f87b26241191d4da55a405a02387e05"
    ),
    "docs/execution/route_a_v3_registry_manifest.json": (
        "73a39a566aa0310a80cc83f4eb17ddb95cabc87e5b070ef6484c05178ba32b75"
    ),
    "docs/execution/route_a_v3_task_registry.yaml": (
        "a64d0b8bb5eb466b06daa46ed109bd19901ee775910bc5cc9221c39ead63a4bc"
    ),
    "scripts/route_a_v3/validate_a0_bundle.py": (
        "81d1a8dc49375f53a1edd5c3f41625e734eacc31a4c328c8137340a041d77e65"
    ),
    "tests/route_a_v3/test_a0_integrity_guards.py": (
        "106f847e957a40fffec4c1b57f8f572325ef96e8a3dc5729db68499e320f380b"
    ),
}
RUNTIME_CONFIG_PATH = "configs/route_a_v3_dec027_authority_runtime_sync_v1.json"
RUNTIME_SCRIPT_PATH = "scripts/route_a_v3/dec027_authority_runtime_sync.py"
RUNTIME_TEST_PATH = "tests/route_a_v3/test_dec027_authority_runtime_sync.py"
RUNTIME_EXACT3: tuple[str, ...] = (
    RUNTIME_CONFIG_PATH,
    RUNTIME_SCRIPT_PATH,
    RUNTIME_TEST_PATH,
)
RUNTIME_I_COMMIT = "5d66e8dc83eb9966f7698ac0fc677f1b06af8ea6"
RUNTIME_B_COMMIT = "e60956cf59cbddc0406c5d116fb9714906db36e1"
RUNTIME_I_BLOBS: dict[str, str] = {
    RUNTIME_CONFIG_PATH: (
        "3d5af87e7512568ed663b211c24a8586eeb9f03936a397cf2d2ddaeb2a21f57b"
    ),
    RUNTIME_SCRIPT_PATH: (
        "44dcda8897e747cfe363668ddc23d8dd9c53a7f3ffab692a1bb4e7cf738973ca"
    ),
    RUNTIME_TEST_PATH: (
        "ff250d4f011d8526e9a4a7bf13049f1f47346faa1c7ea512cbf447a6fb59ba4a"
    ),
}
RUNTIME_B_BLOBS: dict[str, str] = {
    RUNTIME_CONFIG_PATH: (
        "e5c1f96ec57b220fd36ff4677deb37d6dc0be06e02f21af3837e17a51e91e5ee"
    ),
    RUNTIME_SCRIPT_PATH: RUNTIME_I_BLOBS[RUNTIME_SCRIPT_PATH],
    RUNTIME_TEST_PATH: RUNTIME_I_BLOBS[RUNTIME_TEST_PATH],
}
CONFIG_REPO_PATH = (
    "configs/route_a_v3_gse217518_corrected_a1_successor_candidate_v1.json"
)
SCRIPT_REPO_PATH = (
    "scripts/route_a_v3/preflight_gse217518_corrected_a1_successor_candidate.py"
)
TEST_REPO_PATH = (
    "tests/route_a_v3/test_preflight_gse217518_corrected_a1_successor_candidate.py"
)
EXACT3: tuple[str, ...] = (CONFIG_REPO_PATH, SCRIPT_REPO_PATH, TEST_REPO_PATH)
I1_COMMIT = "17a35f0f88cc988b938aaf25d94a8b32f0cacfc8"
I1_BLOBS: dict[str, str] = {
    CONFIG_REPO_PATH: (
        "0aa3324d3cfdfd50837ea32a4d1efef754fe70abdab9805f373401f21a1ccb41"
    ),
    SCRIPT_REPO_PATH: (
        "6ca04bdc464ac30f1c3b83830b74c6621816bd25308e345f39e2c5ee94f21b4c"
    ),
    TEST_REPO_PATH: (
        "b08209856fb852991c1b795864304fcda62a4f63419197c068ec6d1f0fd34691"
    ),
}
OWN_BINDING_FIELDS: tuple[str, ...] = (
    "status",
    "implementation_commit",
    "implementation_script_sha256",
    "implementation_test_sha256",
)

GATE_IDS: tuple[str, ...] = (
    "PUBLIC_SOURCE_ASSET_IDENTITY_AND_PRIMARY_ROUTE_CLOSED",
    "SOURCE_REFERENCE_TO_CANDIDATE_CROSSWALK_CLOSED",
    "FULL_115BP_CONSTRUCT_REPORTER_AND_REGION_CONTEXT_CLOSED",
    "ENDPOINT_DIRECTION_SCALE_TRANSFORM_AND_SEMANTICS_CLOSED",
    "THREE_INDEPENDENT_BIOLOGICAL_EXPERIMENTS_AND_VALID_STANDARD_ERROR_CLOSED",
    "MISSING_OUTLIER_QC_AND_SELECTION_CLOSED",
    "LICENSE_AND_REUSE_RIGHTS_CLOSED",
    "HISTORICAL_ANALYTIC_OR_CHECKPOINT_EXPOSURE_CLOSED",
    "OUTCOME_BLIND_SOURCE_GROUP_NEAR_DUPLICATE_SPLIT_AND_ZERO_LEAKAGE_READINESS_CLOSED",
    "POST_DEDUP_INDEPENDENT_SOURCE_GROUP_EFFECTIVE_N_CLOSED",
    "PREFROZEN_SOURCE_GROUP_POWER_AND_FULL_CI_WIDTH_CLOSED",
)

ROOT_KEYS = {
    "schema_version",
    "protocol_id",
    "contract_id",
    "phase_id",
    "decision_id",
    "document_status",
    "dataset_id",
    "baseline",
    "repository_authority",
    "bindings",
    "production_activation_rule",
    "scope",
    "official_public_sources",
    "official_asset_contract",
    "required_gate_ids_exactly",
    "prefrozen_information_thresholds",
    "article_and_project_facts",
    "current_aggregate_observation",
    "output_contract",
    "terminal_state",
}

FORBIDDEN_OUTPUT_KEYS = {
    "member_id",
    "sequence",
    "source_sequence",
    "candidate_sequence",
    "row_effect",
    "row_standard_error",
    "split_assignment",
}


class CandidateContractError(RuntimeError):
    """Static candidate protocol is incomplete or inconsistent."""


class BindingNotFrozen(RuntimeError):
    """Production execution was attempted before all bindings were frozen."""


class PublicAssetError(RuntimeError):
    """Official public inputs do not meet the frozen asset schema."""


class OutputError(RuntimeError):
    """Aggregate report publication cannot satisfy the frozen contract."""


def _reject_duplicate_keys(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in pairs:
        if key in result:
            raise CandidateContractError(f"duplicate JSON key: {key}")
        result[key] = value
    return result


def _reject_nonfinite(token: str) -> None:
    raise CandidateContractError(f"non-finite JSON constant: {token}")


def load_json_object(path: Path, *, label: str = "JSON") -> dict[str, Any]:
    try:
        payload = path.read_bytes()
    except OSError as exc:
        raise CandidateContractError(f"cannot read {label}: {path}") from exc
    try:
        value = json.loads(
            payload.decode("utf-8"),
            object_pairs_hook=_reject_duplicate_keys,
            parse_constant=_reject_nonfinite,
        )
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise CandidateContractError(f"invalid strict UTF-8 {label}: {path}") from exc
    if not isinstance(value, dict):
        raise CandidateContractError(f"{label} root must be an object")
    return value


def _expect_keys(value: Mapping[str, Any], expected: set[str], label: str) -> None:
    observed = set(value)
    if observed != expected:
        raise CandidateContractError(
            f"{label} key closure differs; "
            f"missing={sorted(expected - observed)}, extra={sorted(observed - expected)}"
        )


def _is_hex(value: Any, length: int) -> bool:
    return (
        isinstance(value, str)
        and len(value) == length
        and all(char in "0123456789abcdef" for char in value)
    )


def _validate_sha_map(value: Any, paths: Sequence[str], *, label: str) -> None:
    if not isinstance(value, Mapping) or set(value) != set(paths):
        raise CandidateContractError(f"{label} path closure differs")
    if any(not _is_hex(digest, 64) for digest in value.values()):
        raise CandidateContractError(f"{label} contains an invalid SHA-256")


def _validate_authority_binding(binding: Mapping[str, Any]) -> None:
    expected = {
        "status": BOUND,
        "authority_commit": AUTHORITY_COMMIT,
        "authority_expected_parent": AUTHORITY_PARENT,
        "authority_exact_changed_paths": list(AUTHORITY_EXACT12),
        "authority_blob_sha256_by_path": AUTHORITY_BLOBS,
    }
    if binding != expected:
        raise CandidateContractError("DEC027 authority A binding differs")


def _runtime_binding_mode(binding: Mapping[str, Any]) -> str:
    expected_fields = [
        "status",
        "runtime_event_id",
        "implementation_commit",
        "implementation_blob_sha256_by_path",
        "binding_commit",
        "binding_expected_parent",
        "binding_blob_sha256_by_path",
    ]
    if binding.get("unknown_to_bound_fields") != expected_fields:
        raise CandidateContractError("runtime grouped binding field list differs")
    if binding.get("implementation_expected_parent") != AUTHORITY_COMMIT:
        raise CandidateContractError("runtime implementation parent differs")
    if binding.get("implementation_exact_changed_paths") != list(RUNTIME_EXACT3):
        raise CandidateContractError("runtime implementation exact3 differs")
    if binding.get("binding_exact_changed_paths") != [RUNTIME_CONFIG_PATH]:
        raise CandidateContractError("runtime binding must be config-only")
    dynamic = tuple(binding.get(field) for field in expected_fields)
    if binding.get("status") == UNKNOWN:
        if dynamic != (UNKNOWN,) * len(expected_fields):
            raise CandidateContractError(
                "runtime binding must remain grouped UNKNOWN, not partially bound"
            )
        return UNKNOWN
    if binding.get("status") != BOUND:
        raise CandidateContractError("runtime binding status is invalid")
    if binding.get("runtime_event_id") != "A1-EVT-059":
        raise CandidateContractError("runtime event must be A1-EVT-059")
    implementation = binding.get("implementation_commit")
    binding_commit = binding.get("binding_commit")
    if not _is_hex(implementation, 40) or not _is_hex(binding_commit, 40):
        raise CandidateContractError(
            "runtime implementation or binding commit is invalid"
        )
    if implementation != RUNTIME_I_COMMIT or binding_commit != RUNTIME_B_COMMIT:
        raise CandidateContractError("settled A1-EVT-059 runtime I/B identity differs")
    if binding.get("binding_expected_parent") != implementation:
        raise CandidateContractError("runtime binding parent differs")
    _validate_sha_map(
        binding.get("implementation_blob_sha256_by_path"),
        RUNTIME_EXACT3,
        label="runtime I blobs",
    )
    _validate_sha_map(
        binding.get("binding_blob_sha256_by_path"),
        RUNTIME_EXACT3,
        label="runtime B blobs",
    )
    if binding.get("implementation_blob_sha256_by_path") != RUNTIME_I_BLOBS:
        raise CandidateContractError("settled A1-EVT-059 runtime I blobs differ")
    if binding.get("binding_blob_sha256_by_path") != RUNTIME_B_BLOBS:
        raise CandidateContractError("settled A1-EVT-059 runtime B blobs differ")
    return BOUND


def _implementation_binding_mode(
    binding: Mapping[str, Any], *, runtime_mode: str
) -> str:
    frozen_i1 = binding.get("frozen_i1_predecessor")
    if frozen_i1 != {
        "status": BOUND,
        "implementation_commit": I1_COMMIT,
        "implementation_expected_parent": RUNTIME_B_COMMIT,
        "implementation_exact_changed_paths": list(EXACT3),
        "implementation_blob_sha256_by_path": I1_BLOBS,
    }:
        raise CandidateContractError("frozen GSE217518 I1 predecessor differs")
    if binding.get("implementation_expected_parent") != I1_COMMIT:
        raise CandidateContractError("dynamic GSE217518 I2 parent differs")
    if binding.get("unknown_to_bound_fields") != list(OWN_BINDING_FIELDS):
        raise CandidateContractError("implementation four-scalar group differs")
    if binding.get("implementation_script_path") != SCRIPT_REPO_PATH:
        raise CandidateContractError("implementation script path differs")
    if binding.get("implementation_test_path") != TEST_REPO_PATH:
        raise CandidateContractError("implementation test path differs")
    if binding.get("implementation_exact_changed_paths") != list(EXACT3):
        raise CandidateContractError("implementation exact3 differs")
    if binding.get("binding_exact_changed_paths") != [CONFIG_REPO_PATH]:
        raise CandidateContractError("implementation B must be config-only")
    dynamic = tuple(binding.get(field) for field in OWN_BINDING_FIELDS)
    if binding.get("status") == UNKNOWN:
        if dynamic != (UNKNOWN,) * len(OWN_BINDING_FIELDS):
            raise CandidateContractError(
                "implementation binding must remain grouped UNKNOWN, not partially bound"
            )
        return UNKNOWN
    if binding.get("status") != BOUND:
        raise CandidateContractError("implementation binding status is invalid")
    if runtime_mode != BOUND:
        raise CandidateContractError(
            "BOUND implementation requires the DEC027 runtime predecessor BOUND"
        )
    if not _is_hex(binding.get("implementation_commit"), 40):
        raise CandidateContractError("implementation commit is invalid")
    for field in (
        "implementation_script_sha256",
        "implementation_test_sha256",
    ):
        if not _is_hex(binding.get(field), 64):
            raise CandidateContractError(f"{field} is invalid")
    return BOUND


def validate_protocol(config: Mapping[str, Any]) -> None:
    _expect_keys(config, ROOT_KEYS, "protocol")
    expected_scalars = {
        "schema_version": SCHEMA_VERSION,
        "protocol_id": PROTOCOL_ID,
        "contract_id": "mrna_xeditflow_route_a_v3",
        "phase_id": "A1",
        "decision_id": "V3-DEC-027",
        "document_status": "DRAFT_CANDIDATE_NOT_ACTIVE_PROTOCOL",
        "dataset_id": "GSE217518",
    }
    for key, expected in expected_scalars.items():
        if config.get(key) != expected:
            raise CandidateContractError(f"protocol {key} differs")

    if config.get("baseline") != {
        "remote_branch": PRODUCTION_BRANCH,
        "dec027_authority_head": AUTHORITY_COMMIT,
        "dec027_authority_parent": AUTHORITY_PARENT,
        "fresh_verified_clean_at_authority_freeze": True,
        "pre_dec027_projection_event": "A1-EVT-058",
        "dec027_runtime_event_expected_when_bound": "A1-EVT-059",
    }:
        raise CandidateContractError("DEC027 baseline differs")

    if config.get("required_gate_ids_exactly") != list(GATE_IDS):
        raise CandidateContractError("the DEC027 exact eleven gate IDs differ")

    repository = config.get("repository_authority")
    if repository != {
        "production_repo_root": str(PRODUCTION_REPO_ROOT),
        "branch": PRODUCTION_BRANCH,
        "upstream_ref": PRODUCTION_UPSTREAM,
        "live_origin_head_required": True,
        "clean_worktree_and_index_required": True,
    }:
        raise CandidateContractError("production repository authority differs")

    bindings = config.get("bindings")
    if not isinstance(bindings, Mapping) or set(bindings) != {
        "authority",
        "runtime",
        "implementation",
    }:
        raise CandidateContractError("binding group closure differs")
    for name in ("authority", "runtime", "implementation"):
        if not isinstance(bindings[name], Mapping):
            raise CandidateContractError(f"{name} binding must be an object")
    _expect_keys(
        bindings["runtime"],
        {
            "status",
            "runtime_event_id",
            "implementation_commit",
            "implementation_expected_parent",
            "implementation_exact_changed_paths",
            "implementation_blob_sha256_by_path",
            "binding_commit",
            "binding_expected_parent",
            "binding_exact_changed_paths",
            "binding_blob_sha256_by_path",
            "unknown_to_bound_fields",
        },
        "runtime binding",
    )
    _expect_keys(
        bindings["implementation"],
        {
            "frozen_i1_predecessor",
            "status",
            "implementation_commit",
            "implementation_expected_parent",
            "implementation_script_path",
            "implementation_script_sha256",
            "implementation_test_path",
            "implementation_test_sha256",
            "implementation_exact_changed_paths",
            "binding_exact_changed_paths",
            "unknown_to_bound_fields",
        },
        "implementation binding",
    )
    _validate_authority_binding(bindings["authority"])
    runtime_mode = _runtime_binding_mode(bindings["runtime"])
    _implementation_binding_mode(bindings["implementation"], runtime_mode=runtime_mode)

    activation = config.get("production_activation_rule")
    if not isinstance(activation, Mapping):
        raise CandidateContractError("production activation rule must be an object")
    _expect_keys(
        activation,
        {
            "all_three_binding_groups_must_be_bound",
            "required_commit_chain",
            "clean_head_equals_upstream_equals_live_origin_required",
            "direct_parent_changed_path_and_blob_audit_required",
            "executing_script_and_focused_test_must_match_implementation_i",
            "binding_commit_may_change_only_the_four_own_binding_scalars",
            "fail_before_official_asset_read",
            "fail_before_output_directory_creation",
            "candidate_evaluation_is_not_production_execution",
        },
        "production activation rule",
    )
    if any(
        activation.get(key) is not True
        for key in (
            "all_three_binding_groups_must_be_bound",
            "clean_head_equals_upstream_equals_live_origin_required",
            "direct_parent_changed_path_and_blob_audit_required",
            "executing_script_and_focused_test_must_match_implementation_i",
            "binding_commit_may_change_only_the_four_own_binding_scalars",
            "fail_before_official_asset_read",
            "fail_before_output_directory_creation",
            "candidate_evaluation_is_not_production_execution",
        )
    ):
        raise CandidateContractError("production fail-before-input/output rule differs")
    if activation.get("required_commit_chain") != (
        "DEC027_A_TO_RUNTIME_I_TO_RUNTIME_B_TO_GSE217518_I1_"
        "TO_GSE217518_I2_TO_GSE217518_B2"
    ):
        raise CandidateContractError("production commit chain differs")

    scope = config.get("scope")
    if not isinstance(scope, Mapping):
        raise CandidateContractError("scope must be an object")
    for key in (
        "ordinary_public_only",
        "aggregate_output_only",
    ):
        if scope.get(key) is not True:
            raise CandidateContractError(f"scope.{key} must be true")
    for key in (
        "member_identifier_output_allowed",
        "sequence_output_allowed",
        "row_effect_output_allowed",
        "row_standard_error_output_allowed",
        "split_assignment_output_allowed",
        "persistent_member_level_intermediate_allowed",
        "private_or_sealed_access_allowed",
        "qualification_allowed",
        "credit_change_allowed",
        "canonical_materialization_allowed",
        "training_allowed",
        "gpu_work_allowed",
        "model_selection_allowed",
        "a7_allowed",
    ):
        if scope.get(key) is not False:
            raise CandidateContractError(f"scope.{key} must be false")

    thresholds = config.get("prefrozen_information_thresholds")
    expected_thresholds = {
        "analysis_unit": "INDEPENDENT_SOURCE_GROUP",
        "alternative_spearman_rho": 0.25,
        "alpha_two_sided": 0.05,
        "target_power_minimum": 0.8,
        "confidence_level": 0.95,
        "maximum_full_ci_width": 0.3,
        "required_effective_n_reference": 156,
        "formal_power_execution_allowed_by_candidate": False,
    }
    if thresholds != expected_thresholds:
        raise CandidateContractError("prefrozen information thresholds differ")

    assets = config.get("official_asset_contract")
    if not isinstance(assets, Mapping):
        raise CandidateContractError("official asset contract must be an object")
    processed = assets.get("processed_assets")
    if not isinstance(processed, list) or len(processed) != 4:
        raise CandidateContractError("exactly four processed assets are required")
    expected_geometry = {
        ("HEK293T", "3UTR"),
        ("HEK293T", "5UTR"),
        ("SH-SY5Y", "3UTR"),
        ("SH-SY5Y", "5UTR"),
    }
    if {
        (item.get("cell_line"), item.get("region"))
        for item in processed
        if isinstance(item, Mapping)
    } != expected_geometry:
        raise CandidateContractError("cell/region asset geometry differs")
    if assets.get("replicate_count_per_cell_region") != 3:
        raise CandidateContractError("replicate count differs")
    if assets.get("measurement_columns_per_asset") != 9:
        raise CandidateContractError("measurement column count differs")

    observation = config.get("current_aggregate_observation")
    if not isinstance(observation, Mapping):
        raise CandidateContractError("current aggregate observation must be an object")
    validate_observation(observation)

    output = config.get("output_contract")
    if (
        not isinstance(output, Mapping)
        or output.get("report_filename") != REPORT_FILENAME
    ):
        raise CandidateContractError("output contract differs")
    _expect_keys(
        output,
        {
            "report_filename",
            "final_file_count",
            "same_directory_temporary_file_required",
            "file_fsync_before_publish_required",
            "atomic_no_replace_hard_link_required",
            "directory_fsync_after_publish_required",
            "identical_existing_report_is_idempotent",
            "different_existing_report_action",
            "member_payload_allowed",
            "sequence_payload_allowed",
            "row_effect_or_standard_error_payload_allowed",
            "split_assignment_payload_allowed",
        },
        "output contract",
    )
    if output.get("final_file_count") != 1:
        raise CandidateContractError("output file count must be one")
    for key in (
        "same_directory_temporary_file_required",
        "file_fsync_before_publish_required",
        "atomic_no_replace_hard_link_required",
        "directory_fsync_after_publish_required",
        "identical_existing_report_is_idempotent",
    ):
        if output.get(key) is not True:
            raise CandidateContractError(f"output_contract.{key} must be true")
    if output.get("different_existing_report_action") != "REJECT_WITHOUT_REPLACEMENT":
        raise CandidateContractError("different existing report action differs")
    for key in (
        "member_payload_allowed",
        "sequence_payload_allowed",
        "row_effect_or_standard_error_payload_allowed",
        "split_assignment_payload_allowed",
    ):
        if output.get(key) is not False:
            raise CandidateContractError(f"output_contract.{key} must be false")

    terminal = config.get("terminal_state")
    if not isinstance(terminal, Mapping):
        raise CandidateContractError("terminal state must be an object")
    if terminal.get("qualified") is not False:
        raise CandidateContractError("candidate may not qualify the dataset")
    if terminal.get("contribution") != {
        "ordinary": 0,
        "a1": 0,
        "true_a2": 0,
        "canonical_records": 0,
    }:
        raise CandidateContractError("candidate contribution must be zero")
    for key in (
        "training_allowed",
        "gpu_work_allowed",
        "model_selection_allowed",
        "a7_allowed",
        "next_phase_authorized",
    ):
        if terminal.get(key) is not False:
            raise CandidateContractError(f"terminal_state.{key} must be false")


def load_protocol(path: Path = CONFIG_PATH) -> dict[str, Any]:
    config = load_json_object(path, label="candidate protocol")
    validate_protocol(config)
    return config


def _require_bool(value: Any, label: str) -> bool:
    if not isinstance(value, bool):
        raise CandidateContractError(f"{label} must be boolean")
    return value


def _require_nonnegative_int(value: Any, label: str) -> int:
    if isinstance(value, bool) or not isinstance(value, int) or value < 0:
        raise CandidateContractError(f"{label} must be a nonnegative integer")
    return value


def validate_observation(observation: Mapping[str, Any]) -> None:
    if observation.get("schema_version") != OBSERVATION_SCHEMA:
        raise CandidateContractError("aggregate observation schema differs")
    for key in (
        "asset_geometry",
        "crosswalk",
        "construct_context",
        "endpoint",
        "replicate_and_se",
        "missing_qc_selection",
        "rights",
        "exposure",
        "split",
        "effective_n",
        "power",
    ):
        if not isinstance(observation.get(key), Mapping):
            raise CandidateContractError(f"observation.{key} must be an object")
    _require_nonnegative_int(
        observation.get("official_processed_asset_count"),
        "official_processed_asset_count",
    )
    _require_nonnegative_int(
        observation["asset_geometry"].get("supplement_rows"), "supplement_rows"
    )
    _require_bool(
        observation["crosswalk"].get("raw_role_tokens_closed"),
        "crosswalk.raw_role_tokens_closed",
    )
    _require_bool(
        observation["power"].get("formal_power_run"), "power.formal_power_run"
    )


def _gate(
    gate_id: str,
    status: str,
    reason_code: str,
    aggregate_evidence: Mapping[str, Any],
) -> dict[str, Any]:
    if gate_id not in GATE_IDS:
        raise CandidateContractError(f"unexpected gate id: {gate_id}")
    if status not in {PASS, BLOCKED, NOT_RUN}:
        raise CandidateContractError(f"unsupported gate status: {status}")
    return {
        "gate_id": gate_id,
        "status": status,
        "reason_code": reason_code,
        "fact_class": "CONFIRMED_FACT",
        "aggregate_evidence": dict(aggregate_evidence),
    }


def evaluate_observation(
    config: Mapping[str, Any], observation: Mapping[str, Any]
) -> dict[str, Any]:
    """Evaluate one aggregate observation; never read or write a member asset."""

    validate_protocol(config)
    validate_observation(observation)
    geometry = observation["asset_geometry"]
    crosswalk = observation["crosswalk"]
    construct = observation["construct_context"]
    endpoint = observation["endpoint"]
    replicate = observation["replicate_and_se"]
    missing = observation["missing_qc_selection"]
    rights = observation["rights"]
    exposure = observation["exposure"]
    split = observation["split"]
    effective_n = observation["effective_n"]
    power = observation["power"]

    gates: list[dict[str, Any]] = []

    asset_pass = (
        observation.get("all_primary_routes_resolved") is True
        and observation.get("official_processed_asset_count") == 4
        and observation.get("official_supplement_count") == 1
    )
    gates.append(
        _gate(
            GATE_IDS[0],
            PASS if asset_pass else BLOCKED,
            (
                "OFFICIAL_GEO_ELIFE_AND_AUTHOR_ARCHIVE_PRIMARY_ROUTES_CLOSED"
                if asset_pass
                else "PRIMARY_PUBLIC_ASSET_ROUTE_NOT_CLOSED"
            ),
            {
                "official_public_source_count": observation.get(
                    "official_public_source_count"
                ),
                "official_processed_asset_count": observation.get(
                    "official_processed_asset_count"
                ),
                "official_supplement_count": observation.get(
                    "official_supplement_count"
                ),
            },
        )
    )

    crosswalk_pass = (
        crosswalk.get("raw_role_tokens_closed") is True
        and crosswalk.get(
            "shared_reference_and_supplement_to_raw_exact_crosswalk_closed"
        )
        is True
        and geometry.get("allele_singleton_group_count") == 0
    )
    gates.append(
        _gate(
            GATE_IDS[1],
            PASS if crosswalk_pass else BLOCKED,
            (
                "EXACT_SOURCE_REFERENCE_CANDIDATE_CROSSWALK_CLOSED"
                if crosswalk_pass
                else "RAW_ROLE_TOKENS_PRESENT_BUT_SHARED_REFERENCE_AND_SUPPLEMENT_RAW_CROSSWALK_NOT_CLOSED"
            ),
            {
                "all_raw_rows_role_labeled": geometry.get(
                    "all_raw_rows_have_exactly_one_ref_or_mut_role_token"
                ),
                "syntactic_one_ref_one_mut_group_count": geometry.get(
                    "syntactic_one_ref_one_mut_group_count"
                ),
                "allele_singleton_group_count": geometry.get(
                    "allele_singleton_group_count"
                ),
                "groups_with_role_multiplicity_count": geometry.get(
                    "groups_with_role_multiplicity_count"
                ),
            },
        )
    )

    construct_pass = (
        construct.get("publisher_115bp_rule_and_reporter_context_present") is True
        and construct.get("exact_boundary_aware_construct_replay_closed_for_all_rows")
        is True
    )
    gates.append(
        _gate(
            GATE_IDS[2],
            PASS if construct_pass else BLOCKED,
            (
                "EXACT_115BP_BOUNDARY_AWARE_CONSTRUCT_AND_REPORTER_CONTEXT_REPLAY_CLOSED"
                if construct_pass
                else "PUBLISHED_115BP_AND_REPORTER_RULE_PRESENT_BUT_EXACT_ALL_ROW_CONSTRUCT_REPLAY_NOT_CLOSED"
            ),
            {
                "publisher_construct_rule_present": construct.get(
                    "publisher_115bp_rule_and_reporter_context_present"
                ),
                "versioned_transcript_rows": geometry.get("versioned_transcript_rows"),
                "supplement_rows": geometry.get("supplement_rows"),
                "exact_all_row_replay_closed": construct.get(
                    "exact_boundary_aware_construct_replay_closed_for_all_rows"
                ),
            },
        )
    )

    endpoint_pass = (
        endpoint.get("identity_direction_scale_transform_and_pair_semantics_closed")
        is True
    )
    gates.append(
        _gate(
            GATE_IDS[3],
            PASS if endpoint_pass else BLOCKED,
            (
                "PUBLISHER_HALF_LIFE_DIRECTION_NORMALIZATION_TRANSFORM_AND_PAIR_MODEL_CLOSED"
                if endpoint_pass
                else "ENDPOINT_DIRECTION_SCALE_TRANSFORM_OR_PAIR_SEMANTICS_NOT_CLOSED"
            ),
            {
                "endpoint_identity": config["article_and_project_facts"][
                    "endpoint_identity"
                ],
                "endpoint_unit": config["article_and_project_facts"]["endpoint_unit"],
                "endpoint_direction": config["article_and_project_facts"][
                    "endpoint_direction"
                ],
                "complete_endpoint_rows_either": geometry.get(
                    "complete_endpoint_rows_either"
                ),
            },
        )
    )

    replicate_pass = (
        replicate.get("publisher_independent_experiment_count") == 3
        and replicate.get("raw_three_by_three_geometry_closed") is True
        and replicate.get("valid_row_level_standard_error_replay_closed") is True
    )
    gates.append(
        _gate(
            GATE_IDS[4],
            PASS if replicate_pass else BLOCKED,
            (
                "THREE_INDEPENDENT_EXPERIMENTS_AND_VALID_STANDARD_ERROR_REPLAY_CLOSED"
                if replicate_pass
                else "THREE_INDEPENDENT_EXPERIMENTS_PUBLISHED_BUT_VALID_STANDARD_ERROR_REPLAY_NOT_CLOSED"
            ),
            {
                "publisher_independent_experiment_count": replicate.get(
                    "publisher_independent_experiment_count"
                ),
                "raw_three_by_three_geometry_closed": replicate.get(
                    "raw_three_by_three_geometry_closed"
                ),
                "valid_standard_error_replay_closed": replicate.get(
                    "valid_row_level_standard_error_replay_closed"
                ),
            },
        )
    )

    qc_pass = (
        missing.get("core_r2_mse_rule_executable") is True
        and missing.get("missing_endpoint_is_zero") is False
        and missing.get("downstream_ref_only_quantile_filter_applies_to_core_a1_effect")
        is False
    )
    gates.append(
        _gate(
            GATE_IDS[5],
            PASS if qc_pass else BLOCKED,
            (
                "CORE_R2_MSE_QC_AND_MISSING_SELECTION_CLOSED_DOWNSTREAM_MOTIF_FILTER_NOT_APPLICABLE"
                if qc_pass
                else "CORE_QC_MISSINGNESS_OR_SELECTION_SEMANTICS_NOT_CLOSED"
            ),
            {
                "core_qc_r_squared_strictly_greater_than": config[
                    "article_and_project_facts"
                ]["core_qc_r_squared_strictly_greater_than"],
                "core_qc_mse_strictly_less_than": config["article_and_project_facts"][
                    "core_qc_mse_strictly_less_than"
                ],
                "complete_endpoint_rows_sh": geometry.get("complete_endpoint_rows_sh"),
                "complete_endpoint_rows_hek": geometry.get(
                    "complete_endpoint_rows_hek"
                ),
                "missing_endpoint_rows_both": geometry.get(
                    "missing_endpoint_rows_both"
                ),
                "missing_endpoint_is_zero": missing.get("missing_endpoint_is_zero"),
                "downstream_ref_only_filter_applies_to_core_a1": missing.get(
                    "downstream_ref_only_quantile_filter_applies_to_core_a1_effect"
                ),
            },
        )
    )

    rights_pass = (
        rights.get("article_cc_by_4_0") is True
        and rights.get("geo_data_availability_explicit") is True
        and rights.get("existing_data_scope_rights_audit_pass") is True
    )
    gates.append(
        _gate(
            GATE_IDS[6],
            PASS if rights_pass else BLOCKED,
            (
                "ELIFE_CC_BY_GEO_AVAILABILITY_AND_EXISTING_DATA_SCOPE_RIGHTS_AUDIT_CLOSED"
                if rights_pass
                else "DATA_SCOPE_LICENSE_OR_REUSE_RIGHTS_NOT_CLOSED"
            ),
            {
                "article_cc_by_4_0": rights.get("article_cc_by_4_0"),
                "geo_data_availability_explicit": rights.get(
                    "geo_data_availability_explicit"
                ),
                "existing_data_scope_rights_audit_pass": rights.get(
                    "existing_data_scope_rights_audit_pass"
                ),
            },
        )
    )

    exposure_pass = (
        exposure.get("internal_historical_analytic_record_count") == 0
        and exposure.get("internal_historical_checkpoint_training_row_count") == 0
        and exposure.get("successor_role_adjudication_closed") is True
    )
    gates.append(
        _gate(
            GATE_IDS[7],
            PASS if exposure_pass else BLOCKED,
            (
                "ZERO_HISTORICAL_ANALYTIC_OR_CHECKPOINT_EXPOSURE_AND_ROLE_CLOSED"
                if exposure_pass
                else "INTERNAL_HISTORICAL_ANALYTIC_AND_CHECKPOINT_EXPOSURE_PRESENT_WITHOUT_SUCCESSOR_ROLE_ADJUDICATION"
            ),
            {
                "external_foundation_sequence_exposure_count": exposure.get(
                    "external_foundation_sequence_exposure_count"
                ),
                "internal_historical_analytic_record_count": exposure.get(
                    "internal_historical_analytic_record_count"
                ),
                "internal_historical_checkpoint_training_row_count": exposure.get(
                    "internal_historical_checkpoint_training_row_count"
                ),
                "successor_role_adjudication_closed": exposure.get(
                    "successor_role_adjudication_closed"
                ),
            },
        )
    )

    split_pass = (
        split.get("outcome_blind_source_group_split_readiness_closed") is True
        and split.get("near_duplicate_zero_leakage_audit_run") is True
        and split.get("split_assignment_execution_count") == 0
    )
    gates.append(
        _gate(
            GATE_IDS[8],
            PASS if split_pass else NOT_RUN,
            (
                "OUTCOME_BLIND_SOURCE_GROUP_AND_NEAR_DUPLICATE_ZERO_LEAKAGE_READINESS_CLOSED"
                if split_pass
                else "SOURCE_GROUP_PREREQUISITE_NOT_CLOSED_AND_SPLIT_LEAKAGE_READINESS_NOT_RUN"
            ),
            {
                "outcome_blind_readiness_closed": split.get(
                    "outcome_blind_source_group_split_readiness_closed"
                ),
                "near_duplicate_zero_leakage_audit_run": split.get(
                    "near_duplicate_zero_leakage_audit_run"
                ),
                "split_assignment_execution_count": split.get(
                    "split_assignment_execution_count"
                ),
            },
        )
    )

    n_pass = (
        effective_n.get("audit_run") is True
        and isinstance(
            effective_n.get("post_dedup_independent_source_group_count"), int
        )
        and not isinstance(
            effective_n.get("post_dedup_independent_source_group_count"), bool
        )
        and effective_n.get("post_dedup_independent_source_group_count") >= 0
    )
    gates.append(
        _gate(
            GATE_IDS[9],
            PASS if n_pass else NOT_RUN,
            (
                "POST_DEDUP_INDEPENDENT_SOURCE_GROUP_EFFECTIVE_N_CLOSED"
                if n_pass
                else "POST_DEDUP_INDEPENDENT_SOURCE_GROUP_AUDIT_NOT_RUN"
            ),
            {
                "nominal_syntactic_pair_count_not_effective_n": geometry.get(
                    "syntactic_one_ref_one_mut_group_count"
                ),
                "post_dedup_independent_source_group_count": effective_n.get(
                    "post_dedup_independent_source_group_count"
                ),
                "audit_run": effective_n.get("audit_run"),
            },
        )
    )

    power_pass = (
        power.get("formal_power_run") is True
        and power.get("full_ci_width_run") is True
        and n_pass
    )
    gates.append(
        _gate(
            GATE_IDS[10],
            PASS if power_pass else NOT_RUN,
            (
                "PREFROZEN_SOURCE_GROUP_POWER_AND_FULL_CI_WIDTH_CLOSED"
                if power_pass
                else "FORMAL_POWER_NOT_AUTHORIZED_OR_RUN_AND_POST_DEDUP_N_NOT_CLOSED"
            ),
            {
                "analysis_unit": config["prefrozen_information_thresholds"][
                    "analysis_unit"
                ],
                "required_effective_n_reference": config[
                    "prefrozen_information_thresholds"
                ]["required_effective_n_reference"],
                "target_power_minimum": config["prefrozen_information_thresholds"][
                    "target_power_minimum"
                ],
                "maximum_full_ci_width": config["prefrozen_information_thresholds"][
                    "maximum_full_ci_width"
                ],
                "formal_power_run": power.get("formal_power_run"),
                "full_ci_width_run": power.get("full_ci_width_run"),
            },
        )
    )

    if [item["gate_id"] for item in gates] != list(GATE_IDS):
        raise CandidateContractError("evaluated gate order differs")
    counts = Counter(item["status"] for item in gates)
    all_pass = counts[PASS] == len(GATE_IDS)
    terminal = dict(config["terminal_state"])
    report = {
        "schema_version": "route_a_v3_gse217518_corrected_a1_successor_aggregate_preflight.v1",
        "protocol_id": PROTOCOL_ID,
        "record_type": "AGGREGATE_CORRECTED_A1_QUALIFICATION_PREFLIGHT_ONLY",
        "dataset_id": "GSE217518",
        "candidate_protocol_status": config["document_status"],
        "production_binding_status": {
            name: config["bindings"][name]["status"]
            for name in ("authority", "runtime", "implementation")
        },
        "result_status": (
            "ALL_ELEVEN_PREFLIGHT_GATES_PASS_PROMOTION_REQUEST_ONLY"
            if all_pass
            else "STOP_CORRECTED_PREFLIGHT_GATES_NOT_CLOSED"
        ),
        "all_required_gates_pass": all_pass,
        "gate_counts": {
            PASS: counts[PASS],
            BLOCKED: counts[BLOCKED],
            NOT_RUN: counts[NOT_RUN],
            "TOTAL": len(gates),
        },
        "gates": gates,
        "aggregate_geometry": {
            "supplement_rows": geometry.get("supplement_rows"),
            "supplement_region_counts": geometry.get("supplement_region_counts"),
            "versioned_transcript_rows": geometry.get("versioned_transcript_rows"),
            "unique_versioned_transcripts": geometry.get(
                "unique_versioned_transcripts"
            ),
            "complete_endpoint_rows_sh": geometry.get("complete_endpoint_rows_sh"),
            "complete_endpoint_rows_hek": geometry.get("complete_endpoint_rows_hek"),
            "complete_endpoint_rows_both": geometry.get("complete_endpoint_rows_both"),
            "complete_endpoint_rows_either": geometry.get(
                "complete_endpoint_rows_either"
            ),
            "syntactic_one_ref_one_mut_group_count": geometry.get(
                "syntactic_one_ref_one_mut_group_count"
            ),
            "allele_singleton_group_count": geometry.get(
                "allele_singleton_group_count"
            ),
            "post_dedup_independent_source_group_count": effective_n.get(
                "post_dedup_independent_source_group_count"
            ),
        },
        "corrected_qc_disposition": {
            "core_half_life_qc": "R_SQUARED_GT_0_5_AND_MSE_LT_1",
            "downstream_reference_only_quantile_filter": (
                "NOT_APPLICABLE_TO_CORE_A1_PAIR_EFFECT"
            ),
            "blanket_author_qc_blocker_retained": False,
        },
        "worth_formal_binding": True,
        "formal_binding_recommendation": (
            "WORTH_FORMAL_BINDING_FOR_ONE_BOUNDED_AGGREGATE_CORRECTED_PREFLIGHT_RUN"
        ),
        "terminal_state": terminal,
        "scope_attestation": {
            "ordinary_public_only": True,
            "aggregate_output_only": True,
            "member_identifier_output_count": 0,
            "sequence_output_count": 0,
            "row_effect_output_count": 0,
            "row_standard_error_output_count": 0,
            "split_assignment_output_count": 0,
            "qualification_run_count": 0,
            "canonical_materialization_count": 0,
            "training_run_count": 0,
            "gpu_run_count": 0,
            "model_selection_run_count": 0,
            "private_or_sealed_input_count": 0,
        },
    }
    _assert_aggregate_only(report)
    return report


def _assert_aggregate_only(value: Any, *, path: str = "report") -> None:
    if isinstance(value, Mapping):
        for key, item in value.items():
            if str(key).lower() in FORBIDDEN_OUTPUT_KEYS:
                raise CandidateContractError(f"forbidden output key at {path}.{key}")
            _assert_aggregate_only(item, path=f"{path}.{key}")
    elif isinstance(value, list):
        for index, item in enumerate(value):
            _assert_aggregate_only(item, path=f"{path}[{index}]")
    elif isinstance(value, float) and not math.isfinite(value):
        raise CandidateContractError(f"non-finite output at {path}")


def _raw_asset_geometry(path: Path) -> tuple[dict[str, Any], set[str]]:
    try:
        handle = gzip.open(path, "rt", encoding="utf-8-sig", newline="")
    except OSError as exc:
        raise PublicAssetError(f"cannot open official processed asset: {path}") from exc
    keys: list[str] = []
    measurement_missing = 0
    measurement_nonfinite = 0
    with handle:
        reader = csv.reader(handle)
        try:
            header = next(reader)
        except StopIteration as exc:
            raise PublicAssetError(f"empty processed asset: {path}") from exc
        if len(header) != 10:
            raise PublicAssetError(f"processed asset column count differs: {path}")
        for row in reader:
            if len(row) != 10:
                raise PublicAssetError(f"processed asset row width differs: {path}")
            key = row[0]
            tokens = re.split(r"([_|:\-/])", key)
            hits = [
                (index, token.casefold())
                for index, token in enumerate(tokens)
                if token.casefold() in {"ref", "mut"}
            ]
            if len(hits) != 1:
                raise PublicAssetError(
                    f"processed asset row does not have one allele role: {path}"
                )
            keys.append(key)
            for raw in row[1:]:
                if raw == "":
                    measurement_missing += 1
                    continue
                try:
                    numeric = float(raw)
                except ValueError as exc:
                    raise PublicAssetError(
                        f"processed asset measurement is nonnumeric: {path}"
                    ) from exc
                if not math.isfinite(numeric):
                    measurement_nonfinite += 1

    if len(keys) != len(set(keys)):
        raise PublicAssetError(f"processed asset contains duplicate keys: {path}")
    groups: dict[str, Counter[str]] = defaultdict(Counter)
    for key in keys:
        tokens = re.split(r"([_|:\-/])", key)
        index, role = next(
            (index, token.casefold())
            for index, token in enumerate(tokens)
            if token.casefold() in {"ref", "mut"}
        )
        tokens[index] = "<ALLELE>"
        groups["".join(tokens)][role] += 1
    return (
        {
            "row_count": len(keys),
            "unique_key_count": len(keys),
            "normalized_group_count": len(groups),
            "syntactic_pair_count": sum(
                counts["ref"] == 1 and counts["mut"] == 1 for counts in groups.values()
            ),
            "ref_only_group_count": sum(
                counts["ref"] > 0 and counts["mut"] == 0 for counts in groups.values()
            ),
            "mut_only_group_count": sum(
                counts["mut"] > 0 and counts["ref"] == 0 for counts in groups.values()
            ),
            "role_multiplicity_group_count": sum(
                counts["ref"] > 1 or counts["mut"] > 1 for counts in groups.values()
            ),
            "measurement_missing_count": measurement_missing,
            "measurement_nonfinite_count": measurement_nonfinite,
        },
        set(keys),
    )


def inspect_official_public_assets(
    config: Mapping[str, Any], asset_dir: Path
) -> dict[str, Any]:
    """Read ordinary-public assets and return aggregate facts only.

    The caller must enforce production binding before invoking this function.
    It retains no member-level intermediate after return and emits no member
    identifier, sequence, row effect, row SE, or split assignment.
    """

    validate_protocol(config)
    contract = config["official_asset_contract"]
    supplement_path = asset_dir / contract["supplement_filename"]
    try:
        workbook = openpyxl.load_workbook(
            supplement_path, read_only=True, data_only=True
        )
    except (OSError, KeyError) as exc:
        raise PublicAssetError(
            f"cannot open official supplement: {supplement_path}"
        ) from exc
    sheet_name = contract["supplement_sheet"]
    if sheet_name not in workbook.sheetnames:
        raise PublicAssetError("official supplement sheet is absent")
    sheet = workbook[sheet_name]
    iterator = sheet.iter_rows(values_only=True)
    try:
        headers = list(next(iterator))
    except StopIteration as exc:
        raise PublicAssetError("official supplement is empty") from exc
    if headers != contract["supplement_headers"]:
        raise PublicAssetError("official supplement header closure differs")
    index = {name: position for position, name in enumerate(headers)}

    supplement_rows = 0
    region_counts: Counter[str] = Counter()
    versioned_accessions: set[str] = set()
    versioned_transcript_rows = 0
    mutant_ids: set[str] = set()
    variant_names: set[str] = set()
    single_base_rows = 0
    complete_sh = 0
    complete_hek = 0
    complete_both = 0
    complete_either = 0
    missing_both = 0
    for row in iterator:
        if len(row) != len(headers):
            raise PublicAssetError("official supplement row width differs")
        supplement_rows += 1
        region_counts[str(row[index["UTR_Group"]])] += 1
        variant_name = str(row[index["variant_name"]])
        match = re.match(r"([A-Z]+_\d+\.\d+)", variant_name)
        if match:
            versioned_transcript_rows += 1
            versioned_accessions.add(match.group(1))
        mutant_ids.add(str(row[index["Mutant"]]))
        variant_names.add(variant_name)
        if (
            len(str(row[index["ReferenceAllele"]])) == 1
            and len(str(row[index["AlternateAllele"]])) == 1
        ):
            single_base_rows += 1
        sh = all(
            row[index[name]] is not None
            for name in ("t05_WT_SH", "t05_mt_SH", "pval_SH")
        )
        hek = all(
            row[index[name]] is not None
            for name in ("t05_WT_HEK", "t05_mt_HEK", "pval_HEK")
        )
        complete_sh += int(sh)
        complete_hek += int(hek)
        complete_both += int(sh and hek)
        complete_either += int(sh or hek)
        missing_both += int(not sh and not hek)
    workbook.close()

    raw_geometry: dict[tuple[str, str], dict[str, Any]] = {}
    raw_keys: dict[tuple[str, str], set[str]] = {}
    for item in contract["processed_assets"]:
        geometry, keys = _raw_asset_geometry(asset_dir / item["filename"])
        role = (item["cell_line"], item["region"])
        raw_geometry[role] = geometry
        raw_keys[role] = keys

    same_region = all(
        raw_keys[("HEK293T", region)] == raw_keys[("SH-SY5Y", region)]
        for region in ("3UTR", "5UTR")
    )
    cross_region = len(raw_keys[("HEK293T", "3UTR")] & raw_keys[("HEK293T", "5UTR")])
    representative = [raw_geometry[("HEK293T", region)] for region in ("3UTR", "5UTR")]
    syntactic_pairs = sum(item["syntactic_pair_count"] for item in representative)
    allele_singletons = sum(
        item["ref_only_group_count"] + item["mut_only_group_count"]
        for item in representative
    )
    role_multiplicity = sum(
        item["role_multiplicity_group_count"] for item in representative
    )
    measurement_missing = sum(
        item["measurement_missing_count"] for item in raw_geometry.values()
    )
    measurement_nonfinite = sum(
        item["measurement_nonfinite_count"] for item in raw_geometry.values()
    )

    observation = json.loads(json.dumps(config["current_aggregate_observation"]))
    observation["source_mode"] = "LIVE_LOCAL_ORDINARY_PUBLIC_ASSET_AGGREGATION"
    observation["asset_geometry"].update(
        {
            "supplement_rows": supplement_rows,
            "supplement_columns": len(headers),
            "supplement_region_counts": dict(region_counts),
            "versioned_transcript_rows": versioned_transcript_rows,
            "unique_versioned_transcripts": len(versioned_accessions),
            "mutant_identifier_unique_count": len(mutant_ids),
            "variant_name_unique_count": len(variant_names),
            "single_base_ref_alt_rows": single_base_rows,
            "indel_or_complex_allele_rows": supplement_rows - single_base_rows,
            "complete_endpoint_rows_sh": complete_sh,
            "complete_endpoint_rows_hek": complete_hek,
            "complete_endpoint_rows_both": complete_both,
            "complete_endpoint_rows_either": complete_either,
            "missing_endpoint_rows_both": missing_both,
            "raw_rows_by_cell_region": {
                "HEK293T_3UTR": raw_geometry[("HEK293T", "3UTR")]["row_count"],
                "HEK293T_5UTR": raw_geometry[("HEK293T", "5UTR")]["row_count"],
                "SH_SY5Y_3UTR": raw_geometry[("SH-SY5Y", "3UTR")]["row_count"],
                "SH_SY5Y_5UTR": raw_geometry[("SH-SY5Y", "5UTR")]["row_count"],
            },
            "all_raw_rows_have_exactly_one_ref_or_mut_role_token": True,
            "same_region_key_sets_identical_across_cell_lines": same_region,
            "cross_region_key_intersection_count": cross_region,
            "syntactic_one_ref_one_mut_group_count": syntactic_pairs,
            "allele_singleton_group_count": allele_singletons,
            "groups_with_role_multiplicity_count": role_multiplicity,
            "raw_measurement_nonfinite_count": measurement_nonfinite,
            "raw_measurement_missing_count": measurement_missing,
        }
    )
    observation["crosswalk"]["raw_role_tokens_closed"] = role_multiplicity == 0
    observation["replicate_and_se"]["raw_three_by_three_geometry_closed"] = all(
        item["row_count"] > 0 for item in raw_geometry.values()
    )
    validate_observation(observation)
    return observation


def _require_production_bindings(config: Mapping[str, Any]) -> None:
    _validate_authority_binding(config["bindings"]["authority"])
    runtime_mode = _runtime_binding_mode(config["bindings"]["runtime"])
    implementation_mode = _implementation_binding_mode(
        config["bindings"]["implementation"], runtime_mode=runtime_mode
    )
    if runtime_mode != BOUND or implementation_mode != BOUND:
        raise BindingNotFrozen(
            "production bindings remain grouped UNKNOWN: "
            + ",".join(
                name
                for name, mode in (
                    ("runtime", runtime_mode),
                    ("implementation", implementation_mode),
                )
                if mode != BOUND
            )
        )


def _run_git(repo_root: Path, *arguments: str) -> str:
    try:
        result = subprocess.run(
            ["git", "-C", str(repo_root), *arguments],
            check=False,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
    except OSError as exc:
        raise CandidateContractError("git is unavailable for binding audit") from exc
    if result.returncode != 0:
        raise CandidateContractError("git repository binding audit failed")
    return result.stdout.strip()


def _git_blob(repo_root: Path, commit: str, path: str) -> bytes:
    try:
        result = subprocess.run(
            ["git", "-C", str(repo_root), "show", f"{commit}:{path}"],
            check=False,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
        )
    except OSError as exc:
        raise CandidateContractError("git is unavailable for blob audit") from exc
    if result.returncode != 0:
        raise CandidateContractError(f"cannot read bound Git blob: {path}")
    return result.stdout


def _changed_paths(repo_root: Path, commit: str) -> tuple[str, ...]:
    value = _run_git(
        repo_root,
        "diff-tree",
        "--no-commit-id",
        "--name-only",
        "-r",
        commit,
    )
    return tuple(sorted(line for line in value.splitlines() if line))


def _verify_frozen_commit(
    repo_root: Path,
    *,
    label: str,
    commit: str,
    expected_parent: str,
    expected_paths: Sequence[str],
    expected_blobs: Mapping[str, str] | None = None,
) -> None:
    ancestry = _run_git(repo_root, "rev-list", "--parents", "-n", "1", commit)
    fields = ancestry.split()
    if len(fields) != 2 or fields[0] != commit or fields[1] != expected_parent:
        raise CandidateContractError(f"{label} direct parent differs")
    if _changed_paths(repo_root, commit) != tuple(sorted(expected_paths)):
        raise CandidateContractError(f"{label} changed-path closure differs")
    for path, expected_sha in (expected_blobs or {}).items():
        observed_sha = hashlib.sha256(_git_blob(repo_root, commit, path)).hexdigest()
        if observed_sha != expected_sha:
            raise CandidateContractError(f"{label} blob identity differs: {path}")


def _live_origin_head(repo_root: Path, branch: str) -> str:
    ref = f"refs/heads/{branch}"
    value = _run_git(repo_root, "ls-remote", "--exit-code", "--heads", "origin", ref)
    lines = [line.split() for line in value.splitlines() if line.strip()]
    if len(lines) != 1 or len(lines[0]) != 2 or lines[0][1] != ref:
        raise CandidateContractError("live origin branch resolution differs")
    commit = lines[0][0]
    if not _is_hex(commit, 40):
        raise CandidateContractError("live origin head is not a commit id")
    return commit


def _normalise_own_binding(config: Mapping[str, Any]) -> dict[str, Any]:
    normalised = deepcopy(dict(config))
    own = normalised["bindings"]["implementation"]
    for field in OWN_BINDING_FIELDS:
        own[field] = UNKNOWN
    return normalised


def _audit_repository_bindings(
    config: Mapping[str, Any], config_path: Path, repo_root: Path
) -> dict[str, str]:
    """Verify A -> runtime I/B -> GSE217518 I1 -> I2/B2 before asset/output I/O."""

    _require_production_bindings(config)
    repository = config["repository_authority"]
    if repo_root.resolve() != Path(repository["production_repo_root"]).resolve():
        raise CandidateContractError("execution repository is not the frozen root")
    expected_config_path = (repo_root / CONFIG_REPO_PATH).resolve()
    if config_path.resolve() != expected_config_path:
        raise CandidateContractError("protocol path is outside the frozen repository")

    head = _run_git(repo_root, "rev-parse", "HEAD")
    upstream = _run_git(repo_root, "rev-parse", "@{upstream}")
    live_origin = _live_origin_head(repo_root, repository["branch"])
    if head != upstream or head != live_origin:
        raise CandidateContractError("HEAD, upstream, and live origin do not match")
    if _run_git(repo_root, "rev-parse", "--abbrev-ref", "HEAD") != repository["branch"]:
        raise CandidateContractError("production branch differs")
    if (
        _run_git(repo_root, "rev-parse", "--abbrev-ref", "@{upstream}")
        != repository["upstream_ref"]
    ):
        raise CandidateContractError("production upstream differs")
    if _run_git(repo_root, "status", "--porcelain=v1", "--untracked-files=all"):
        raise CandidateContractError("production worktree or index is dirty")

    authority = config["bindings"]["authority"]
    runtime = config["bindings"]["runtime"]
    own = config["bindings"]["implementation"]
    runtime_i = str(runtime["implementation_commit"])
    runtime_b = str(runtime["binding_commit"])
    implementation_i2 = str(own["implementation_commit"])

    _verify_frozen_commit(
        repo_root,
        label="DEC027 authority A",
        commit=AUTHORITY_COMMIT,
        expected_parent=AUTHORITY_PARENT,
        expected_paths=AUTHORITY_EXACT12,
        expected_blobs=authority["authority_blob_sha256_by_path"],
    )
    _verify_frozen_commit(
        repo_root,
        label="DEC027 runtime I",
        commit=runtime_i,
        expected_parent=AUTHORITY_COMMIT,
        expected_paths=RUNTIME_EXACT3,
        expected_blobs=runtime["implementation_blob_sha256_by_path"],
    )
    _verify_frozen_commit(
        repo_root,
        label="DEC027 runtime B",
        commit=runtime_b,
        expected_parent=runtime_i,
        expected_paths=(RUNTIME_CONFIG_PATH,),
        expected_blobs=runtime["binding_blob_sha256_by_path"],
    )
    _verify_frozen_commit(
        repo_root,
        label="GSE217518 frozen implementation I1",
        commit=I1_COMMIT,
        expected_parent=runtime_b,
        expected_paths=EXACT3,
        expected_blobs=I1_BLOBS,
    )
    _verify_frozen_commit(
        repo_root,
        label="GSE217518 dynamic implementation I2",
        commit=implementation_i2,
        expected_parent=I1_COMMIT,
        expected_paths=EXACT3,
        expected_blobs={
            SCRIPT_REPO_PATH: own["implementation_script_sha256"],
            TEST_REPO_PATH: own["implementation_test_sha256"],
        },
    )
    _verify_frozen_commit(
        repo_root,
        label="GSE217518 binding B2",
        commit=head,
        expected_parent=implementation_i2,
        expected_paths=(CONFIG_REPO_PATH,),
    )

    implementation_protocol = load_json_object_from_bytes(
        _git_blob(repo_root, implementation_i2, CONFIG_REPO_PATH),
        label="GSE217518 dynamic implementation I2 protocol",
    )
    if _normalise_own_binding(config) != implementation_protocol:
        raise CandidateContractError(
            "GSE217518 B2 changed fields outside its frozen four-scalar group"
        )
    head_config_blob = _git_blob(repo_root, head, CONFIG_REPO_PATH)
    if config_path.read_bytes() != head_config_blob:
        raise CandidateContractError("working protocol differs from GSE217518 B2")
    script_blob = _git_blob(repo_root, implementation_i2, SCRIPT_REPO_PATH)
    test_blob = _git_blob(repo_root, implementation_i2, TEST_REPO_PATH)
    executing_script = Path(__file__).resolve()
    if executing_script != (repo_root / SCRIPT_REPO_PATH).resolve():
        raise CandidateContractError("executing producer is a stale or copied script")
    if executing_script.read_bytes() != script_blob:
        raise CandidateContractError("executing producer differs from GSE217518 I2")
    if (repo_root / TEST_REPO_PATH).read_bytes() != test_blob:
        raise CandidateContractError("working focused test differs from GSE217518 I2")
    return {
        "status": PASS,
        "authority_commit": AUTHORITY_COMMIT,
        "runtime_binding_commit": runtime_b,
        "frozen_i1_commit": I1_COMMIT,
        "implementation_commit": implementation_i2,
        "binding_commit": head,
    }


def load_json_object_from_bytes(payload: bytes, *, label: str) -> dict[str, Any]:
    try:
        value = json.loads(
            payload.decode("utf-8"),
            object_pairs_hook=_reject_duplicate_keys,
            parse_constant=_reject_nonfinite,
        )
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise CandidateContractError(f"invalid strict UTF-8 {label}") from exc
    if not isinstance(value, dict):
        raise CandidateContractError(f"{label} root must be an object")
    return value


def write_report(output_dir: Path, report: Mapping[str, Any]) -> Path:
    _assert_aggregate_only(report)
    try:
        payload = (
            json.dumps(
                report,
                sort_keys=True,
                indent=2,
                ensure_ascii=False,
                allow_nan=False,
            )
            + "\n"
        ).encode("utf-8")
    except (TypeError, ValueError) as exc:
        raise OutputError("report is not finite aggregate JSON") from exc

    output_created = False
    directory_created = False
    temporary: Path | None = None
    path = output_dir / REPORT_FILENAME
    try:
        if output_dir.exists():
            if not output_dir.is_dir():
                raise OutputError("output path is not a directory")
            entries = list(output_dir.iterdir())
            if entries:
                if len(entries) == 1 and entries[0] == path:
                    if path.read_bytes() == payload:
                        return path
                    raise OutputError(
                        "different aggregate report already exists; replacement refused"
                    )
                raise OutputError("output directory contains an unexpected entry")
        else:
            output_dir.mkdir(parents=False, exist_ok=False)
            directory_created = True
            _fsync_directory(output_dir.parent)

        descriptor, temporary_name = tempfile.mkstemp(
            prefix=f".{REPORT_FILENAME}.", suffix=".tmp", dir=output_dir
        )
        temporary = Path(temporary_name)
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        try:
            os.link(temporary, path)
            output_created = True
        except FileExistsError as exc:
            if path.read_bytes() == payload:
                temporary.unlink()
                temporary = None
                return path
            raise OutputError(
                "different aggregate report appeared; replacement refused"
            ) from exc
        temporary.unlink()
        temporary = None
        _fsync_directory(output_dir)
        if list(output_dir.iterdir()) != [path]:
            raise OutputError("publication did not produce exactly one fixed report")
        return path
    except OutputError:
        if output_created:
            try:
                path.unlink()
            except OSError:
                pass
        raise
    except OSError as exc:
        if output_created:
            try:
                path.unlink()
            except OSError:
                pass
        raise OutputError("cannot atomically publish aggregate report") from exc
    finally:
        if temporary is not None:
            try:
                temporary.unlink()
            except FileNotFoundError:
                pass
        if directory_created and output_dir.exists():
            try:
                if not any(output_dir.iterdir()):
                    output_dir.rmdir()
            except OSError:
                pass


def _fsync_directory(path: Path) -> None:
    descriptor = os.open(path, os.O_RDONLY | getattr(os, "O_DIRECTORY", 0))
    try:
        os.fsync(descriptor)
    finally:
        os.close(descriptor)


def execute(config_path: Path, asset_dir: Path, output_dir: Path) -> Path:
    """Production entry point: binding check precedes asset and output access."""

    config = load_protocol(config_path)
    _require_production_bindings(config)
    _audit_repository_bindings(config, config_path, PRODUCTION_REPO_ROOT)
    observation = inspect_official_public_assets(config, asset_dir)
    report = evaluate_observation(config, observation)
    return write_report(output_dir, report)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", type=Path, default=CONFIG_PATH)
    parser.add_argument("--asset-dir", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    args = parser.parse_args()
    try:
        report_path = execute(args.config, args.asset_dir, args.output_dir)
    except (
        CandidateContractError,
        BindingNotFrozen,
        PublicAssetError,
        OutputError,
    ) as exc:
        parser.error(str(exc))
    print(report_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
