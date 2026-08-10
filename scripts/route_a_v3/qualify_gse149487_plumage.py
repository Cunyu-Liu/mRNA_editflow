#!/usr/bin/env python3
"""Fail-closed full-study qualification for public GSE149487/PLUMAGE.

The existing 293T V4 reconstruction is intentionally not imported as evidence
of qualification and is never modified.  This program binds a separate
21-asset trust root, reconstructs both public cell contexts from the eighteen
raw GEO count tables, keeps the paper-native barcode-distribution test separate
from the Route-A biological-replicate companion estimator, and materializes
CanonicalInterventionRecordV3 only when every pre-frozen gate passes.

The committed protocol deliberately leaves evidence that was not uniquely
established from the public materials as UNKNOWN_NOT_ASSERTED.  Running this
program with that protocol therefore produces an aggregate blocked bundle and
zero canonical records.  Qualification requires a new, hash-bound protocol
revision that supplies—not guesses—the missing mapping, paper-method, license,
and checkpoint-exposure evidence.  A staging revision whose implementation
commit is still UNKNOWN_NOT_ASSERTED is likewise blocked until the ordinary
two-stage implementation binding is completed.
"""
from __future__ import annotations

import argparse
import csv
import ctypes
import errno
import gzip
import hashlib
import importlib.util
import json
import math
import os
import random
import re
import stat
import statistics
import subprocess
import sys
import tempfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from statistics import NormalDist
from typing import Any, Callable, Iterable, Mapping, MutableMapping, Sequence

import pandas as pd
from openpyxl import load_workbook
from scipy.stats import mannwhitneyu


CONTRACT_ID = "mrna_xeditflow_route_a_v3"
DATASET_ID = "GSE149487"
DATASET_ALIAS = "PLUMAGE"
STUDY_GROUP_ID = "PLUMAGE_LIM_2021"
PROTOCOL_ID = "ROUTE_A_V3_GSE149487_PLUMAGE_FULL_A1_QUALIFICATION_V1"
PROTOCOL_STATUS = "PREFROZEN_FAIL_CLOSED_BEFORE_FULL_RAW_JOIN_RESULTS"
ASSET_MANIFEST_ID = "ROUTE_A_V3_GSE149487_ADDITIVE_ASSET_MANIFEST_V2"
QUALIFICATION_STATUS = "QUALIFIED_A1_ORDINARY"
BLOCKED_STATUS = "BLOCKED_PENDING_PUBLIC_EVIDENCE"
SCHEMA_VERSION = "3.0.0"

CONTEXTS: tuple[str, ...] = ("PC3", "293T")
ASSAYS: tuple[str, ...] = ("DNA", "TOTALRNA", "POLYSOME")
REPLICATES: tuple[int, ...] = (1, 2, 3)
ENDPOINTS: tuple[Mapping[str, str], ...] = (
    {
        "endpoint_key": "TRANSCRIPT",
        "endpoint_id": "transcript_log2_totalrna_over_dna",
        "endpoint_name": "TOTALRNA_OVER_DNA",
        "numerator_assay": "TOTALRNA",
        "denominator_assay": "DNA",
    },
    {
        "endpoint_key": "TRANSLATION_EFFICIENCY",
        "endpoint_id": "te_log2_polysome_over_totalrna",
        "endpoint_name": "POLYSOME_OVER_TOTALRNA",
        "numerator_assay": "POLYSOME",
        "denominator_assay": "TOTALRNA",
    },
)

FORBIDDEN_PATH_TOKENS: tuple[str, ...] = (
    "gse246381",
    "restricted",
    "sealed_external",
    "access_log",
)
SHA256_RE = re.compile(r"^[0-9a-f]{64}$")
COMMIT_RE = re.compile(r"^[0-9a-f]{40}$")
IDENTIFIER_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9_.:-]{0,159}$")
NONNEGATIVE_INTEGER_RE = re.compile(r"^(?:0|[1-9][0-9]*)$")
SEQUENCE_RE = re.compile(r"^[ACGT]+$")

ALWAYS_OUTPUT_FILES: tuple[str, ...] = (
    "ASSET_MANIFEST_EFFECTIVE.json",
    "CONSTRUCT_MAPPING_AUDIT.jsonl",
    "REJECTION_AUDIT.jsonl",
    "PAPER_NATIVE_REPRODUCTION.json",
    "replicate_effect_summaries.jsonl",
    "GROUP_LEAKAGE_AUDIT.json",
    "GROUP_POWER_AUDIT.json",
    "LICENSE_AND_EXPOSURE_AUDIT.json",
    "QUALIFICATION_REPORT.json",
    "SHA256SUMS",
)
CANONICAL_FILENAME = "canonical_intervention_records.jsonl"
PUBLICATION_COMMIT_FILENAME = "PUBLICATION_COMMIT.json"
PRIMARY_PUBLICATION_MODE = "KERNEL_ATOMIC_RENAME_NOREPLACE_V1"
FALLBACK_PUBLICATION_MODE = "ATOMIC_MKDIR_TERMINAL_COMMIT_MARKER_V1"
ATOMIC_NOREPLACE_UNSUPPORTED_ERRNOS = frozenset(
    {
        errno.ENOSYS,
        errno.EINVAL,
        getattr(errno, "ENOTSUP", errno.EINVAL),
        getattr(errno, "EOPNOTSUPP", errno.EINVAL),
    }
)
VERIFIED_JSON_BYTES_MODE = "PARSED_FROM_SINGLE_OPEN_VERIFIED_BYTES"
PRIVATE_READ_ONLY_SNAPSHOT_MODE = (
    "PRIVATE_READ_ONLY_SNAPSHOT_FROM_SINGLE_OPEN_VERIFIED_DESCRIPTOR"
)
RAW_KEY_CLASSIFICATION_HASH_DOMAIN = "GSE149487_RAW_KEY_CLASSIFICATION_V1"

# Test-only injection point.  Production execution leaves this as None.  The
# focused regression suite uses it to replace original paths after every
# parser input has been verified/snapshotted, proving that scientific parsing
# never reopens those mutable source paths.
_POST_VERIFIED_INPUT_SNAPSHOT_HOOK: Callable[[], None] | None = None


class QualificationError(RuntimeError):
    """An execution or evidence-integrity error that must fail closed."""


class ScopeViolation(QualificationError):
    """A forbidden or unsafe path was rejected before payload processing."""


class TransactionClaimContended(QualificationError):
    """Another execution owns the output/failure publication transaction."""


class PublicationContended(QualificationError):
    """A competing process already owns or materialized the final target."""


class AtomicNoReplaceUnsupported(QualificationError):
    """The kernel no-replace primitive returned an approved unsupported errno."""

    def __init__(self, error_number: int) -> None:
        self.error_number = int(error_number)
        super().__init__(
            f"atomic no-replace directory publication is unsupported (errno {error_number})"
        )


class PartialPublicationError(QualificationError):
    """Fallback final directory exists without a valid terminal commit marker."""

    def __init__(
        self,
        detail: str,
        *,
        unsupported_errno: int,
        commit_marker_present: bool,
    ) -> None:
        self.unsupported_errno = int(unsupported_errno)
        self.commit_marker_present = bool(commit_marker_present)
        super().__init__(detail)


class CommittedPublicationValidationError(QualificationError):
    """An atomically committed directory failed terminal-marker acceptance."""

    def __init__(
        self,
        detail: str,
        *,
        publication_mode: str,
        durability_warning_codes: Sequence[str],
    ) -> None:
        self.publication_mode = publication_mode
        self.durability_warning_codes = tuple(sorted(set(durability_warning_codes)))
        super().__init__(detail)


class ExclusiveWriteCommittedCloseError(QualificationError):
    """An exclusive file was fully written/fsynced before descriptor close failed."""


def _compact_json_bytes(value: Any) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=True,
        allow_nan=False,
        separators=(",", ":"),
        sort_keys=True,
    ).encode("utf-8")


def _pretty_json_bytes(value: Any) -> bytes:
    return (
        json.dumps(
            value,
            ensure_ascii=True,
            allow_nan=False,
            indent=2,
            sort_keys=True,
        )
        + "\n"
    ).encode("utf-8")


def _jsonl_bytes(rows: Sequence[Mapping[str, Any]]) -> bytes:
    return b"".join(_compact_json_bytes(dict(row)) + b"\n" for row in rows)


def _sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _sha256_file(path: Path) -> tuple[str, int]:
    digest = hashlib.sha256()
    size = 0
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1 << 20), b""):
            digest.update(block)
            size += len(block)
    return digest.hexdigest(), size


def _open_regular_readonly(
    path: Path,
    *,
    label: str,
    suffix: str | None = None,
) -> tuple[int, os.stat_result]:
    initial = _require_regular_file(path, label=label, suffix=suffix)
    flags = os.O_RDONLY | getattr(os, "O_CLOEXEC", 0) | getattr(os, "O_NOFOLLOW", 0)
    try:
        descriptor = os.open(path, flags)
    except OSError as exc:
        raise QualificationError(f"{label} could not be opened without following symlinks") from exc
    try:
        opened = os.fstat(descriptor)
        if not stat.S_ISREG(opened.st_mode):
            raise QualificationError(f"{label} opened descriptor is not a regular file")
        initial_identity = (
            initial.st_dev,
            initial.st_ino,
            initial.st_size,
            initial.st_mtime_ns,
        )
        opened_identity = (
            opened.st_dev,
            opened.st_ino,
            opened.st_size,
            opened.st_mtime_ns,
        )
        if initial_identity != opened_identity:
            raise QualificationError(f"{label} changed before its verified descriptor was opened")
    except Exception:
        os.close(descriptor)
        raise
    return descriptor, opened


def _verify_descriptor_final_state(
    descriptor: int,
    initial: os.stat_result,
    observed_bytes: int,
    *,
    label: str,
) -> None:
    final = os.fstat(descriptor)
    initial_identity = (
        initial.st_dev,
        initial.st_ino,
        initial.st_size,
        initial.st_mtime_ns,
    )
    final_identity = (
        final.st_dev,
        final.st_ino,
        final.st_size,
        final.st_mtime_ns,
    )
    if initial_identity != final_identity or observed_bytes != final.st_size:
        raise QualificationError(f"{label} changed while its verified bytes were captured")


def _validate_observed_file(
    *,
    observed_sha256: str,
    observed_bytes: int,
    expected_sha256: str,
    expected_bytes: int | None,
    label: str,
) -> None:
    if observed_sha256 != _require_sha256(
        expected_sha256, label=f"{label} expected hash"
    ):
        raise QualificationError(f"{label} SHA-256 mismatch")
    if expected_bytes is not None:
        if (
            isinstance(expected_bytes, bool)
            or not isinstance(expected_bytes, int)
            or expected_bytes < 0
        ):
            raise QualificationError(f"{label} expected bytes are invalid")
        if observed_bytes != expected_bytes:
            raise QualificationError(f"{label} byte count mismatch")


def _read_verified_bytes(
    path: Path,
    expected_sha256: str,
    *,
    label: str,
    expected_bytes: int | None = None,
    suffix: str | None = None,
) -> tuple[bytes, dict[str, Any]]:
    descriptor, initial = _open_regular_readonly(path, label=label, suffix=suffix)
    digest = hashlib.sha256()
    payload = bytearray()
    try:
        while True:
            block = os.read(descriptor, 1 << 20)
            if not block:
                break
            digest.update(block)
            payload.extend(block)
        _verify_descriptor_final_state(
            descriptor,
            initial,
            len(payload),
            label=label,
        )
    finally:
        os.close(descriptor)
    observed_sha256 = digest.hexdigest()
    observed_bytes = len(payload)
    _validate_observed_file(
        observed_sha256=observed_sha256,
        observed_bytes=observed_bytes,
        expected_sha256=expected_sha256,
        expected_bytes=expected_bytes,
        label=label,
    )
    return bytes(payload), {
        "sha256": observed_sha256,
        "bytes": observed_bytes,
        "filename": path.name,
        "parser_input_mode": VERIFIED_JSON_BYTES_MODE,
    }


def _snapshot_verified_file(
    source: Path,
    destination: Path,
    expected_sha256: str,
    *,
    label: str,
    expected_bytes: int | None = None,
    suffix: str | None = None,
) -> dict[str, Any]:
    source_descriptor, initial = _open_regular_readonly(
        source,
        label=label,
        suffix=suffix,
    )
    destination_descriptor: int | None = None
    digest = hashlib.sha256()
    observed_bytes = 0
    completed = False
    try:
        flags = os.O_WRONLY | os.O_CREAT | os.O_EXCL | getattr(os, "O_CLOEXEC", 0)
        destination_descriptor = os.open(destination, flags, 0o600)
        while True:
            block = os.read(source_descriptor, 1 << 20)
            if not block:
                break
            digest.update(block)
            observed_bytes += len(block)
            view = memoryview(block)
            while view:
                written = os.write(destination_descriptor, view)
                if written <= 0:
                    raise QualificationError(f"{label} snapshot write made no progress")
                view = view[written:]
        _verify_descriptor_final_state(
            source_descriptor,
            initial,
            observed_bytes,
            label=label,
        )
        observed_sha256 = digest.hexdigest()
        _validate_observed_file(
            observed_sha256=observed_sha256,
            observed_bytes=observed_bytes,
            expected_sha256=expected_sha256,
            expected_bytes=expected_bytes,
            label=label,
        )
        os.fsync(destination_descriptor)
        os.fchmod(destination_descriptor, stat.S_IRUSR)
        completed = True
    except OSError as exc:
        raise QualificationError(f"{label} private verified snapshot failed") from exc
    finally:
        if destination_descriptor is not None:
            os.close(destination_descriptor)
        os.close(source_descriptor)
        if not completed:
            try:
                destination.unlink()
            except FileNotFoundError:
                pass
    return {
        "sha256": digest.hexdigest(),
        "bytes": observed_bytes,
        "filename": source.name,
        "parser_input_mode": PRIVATE_READ_ONLY_SNAPSHOT_MODE,
    }


def _stable_id(prefix: str, value: Mapping[str, Any], *, length: int = 24) -> str:
    return prefix + _sha256_bytes(_compact_json_bytes(value))[:length]


def _require_sha256(value: Any, *, label: str) -> str:
    if not isinstance(value, str) or not SHA256_RE.fullmatch(value):
        raise QualificationError(f"{label} must be a lowercase SHA-256")
    return value


def _require_identifier(value: Any, *, label: str) -> str:
    if not isinstance(value, str) or not IDENTIFIER_RE.fullmatch(value):
        raise QualificationError(f"{label} is not a closed identifier")
    return value


def _require_exact_keys(value: Any, keys: set[str], *, label: str) -> Mapping[str, Any]:
    if not isinstance(value, Mapping):
        raise QualificationError(f"{label} must be an object")
    actual = set(value)
    if actual != keys:
        raise QualificationError(
            f"{label} keys are not closed; missing={sorted(keys - actual)}, "
            f"unexpected={sorted(actual - keys)}"
        )
    return value


def _require_exact_value(actual: Any, expected: Any, *, label: str) -> None:
    if type(actual) is not type(expected) or actual != expected:
        raise QualificationError(f"{label} does not match the pre-frozen value")


def _reject_forbidden_path(path: Path | str, *, label: str) -> None:
    text = os.fspath(path).casefold()
    matches = sorted(token for token in FORBIDDEN_PATH_TOKENS if token in text)
    if matches:
        raise ScopeViolation(
            f"{label} rejected before read; forbidden path token(s): "
            + ",".join(matches)
        )


def _absolute_without_resolving(path: Path | str) -> Path:
    return Path(os.path.abspath(os.path.expanduser(os.fspath(path))))


def _reject_symlink_components(path: Path, *, label: str) -> None:
    if not path.is_absolute():
        raise QualificationError(f"internal error: {label} path is not absolute")
    current = Path(path.anchor)
    for component in path.parts[1:]:
        current = current / component
        try:
            info = current.lstat()
        except FileNotFoundError:
            break
        if stat.S_ISLNK(info.st_mode):
            raise ScopeViolation(f"{label} must not contain a symlink component")


def _require_regular_file(path: Path, *, label: str, suffix: str | None = None) -> os.stat_result:
    _reject_forbidden_path(path, label=label)
    _reject_symlink_components(path, label=label)
    try:
        info = path.lstat()
    except FileNotFoundError as exc:
        raise QualificationError(f"{label} is missing") from exc
    if not stat.S_ISREG(info.st_mode):
        raise QualificationError(f"{label} must be a regular file")
    if suffix is not None and path.suffix.casefold() != suffix.casefold():
        raise QualificationError(f"{label} must have suffix {suffix}")
    return info


def _require_directory(path: Path, *, label: str) -> os.stat_result:
    _reject_forbidden_path(path, label=label)
    _reject_symlink_components(path, label=label)
    try:
        info = path.lstat()
    except FileNotFoundError as exc:
        raise QualificationError(f"{label} is missing") from exc
    if not stat.S_ISDIR(info.st_mode):
        raise QualificationError(f"{label} must be a directory")
    return info


def _require_absent_target(path: Path, *, label: str, suffix: str | None = None) -> None:
    _reject_forbidden_path(path, label=label)
    _reject_symlink_components(path, label=label)
    if suffix is not None and path.suffix.casefold() != suffix.casefold():
        raise QualificationError(f"{label} must have suffix {suffix}")
    if path.exists() or path.is_symlink():
        raise QualificationError(f"refusing to overwrite existing {label}")
    _require_directory(path.parent, label=f"{label} parent")


def _safe_relative_path(root: Path, relative: Any, *, label: str) -> Path:
    if not isinstance(relative, str) or not relative or Path(relative).is_absolute():
        raise QualificationError(f"{label} must be a nonempty relative path")
    if any(part in {"", ".", ".."} for part in Path(relative).parts):
        raise QualificationError(f"{label} may not contain dot or parent components")
    candidate = _absolute_without_resolving(root / relative)
    try:
        candidate.relative_to(root)
    except ValueError as exc:
        raise ScopeViolation(f"{label} escapes the repository root") from exc
    _reject_forbidden_path(candidate, label=label)
    return candidate


def _read_json(path: Path, *, label: str) -> dict[str, Any]:
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise QualificationError(f"{label} is not valid UTF-8 JSON") from exc
    if not isinstance(value, dict):
        raise QualificationError(f"{label} root must be an object")
    return value


def _read_json_bytes(payload: bytes, *, label: str) -> dict[str, Any]:
    try:
        value = json.loads(payload.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise QualificationError(f"{label} is not valid UTF-8 JSON") from exc
    if not isinstance(value, dict):
        raise QualificationError(f"{label} root must be an object")
    return value


def _verify_file_hash(
    path: Path,
    expected_sha256: str,
    *,
    label: str,
    expected_bytes: int | None = None,
) -> dict[str, Any]:
    initial = _require_regular_file(path, label=label)
    observed_sha256, observed_bytes = _sha256_file(path)
    final = _require_regular_file(path, label=label)
    initial_identity = (initial.st_dev, initial.st_ino, initial.st_size, initial.st_mtime_ns)
    final_identity = (final.st_dev, final.st_ino, final.st_size, final.st_mtime_ns)
    if initial_identity != final_identity or observed_bytes != final.st_size:
        raise QualificationError(f"{label} changed while being hashed")
    if observed_sha256 != _require_sha256(expected_sha256, label=f"{label} expected hash"):
        raise QualificationError(f"{label} SHA-256 mismatch")
    if expected_bytes is not None:
        if isinstance(expected_bytes, bool) or not isinstance(expected_bytes, int) or expected_bytes < 0:
            raise QualificationError(f"{label} expected bytes are invalid")
        if observed_bytes != expected_bytes:
            raise QualificationError(f"{label} byte count mismatch")
    return {"sha256": observed_sha256, "bytes": observed_bytes, "filename": path.name}


def _validate_protocol(document: Mapping[str, Any]) -> dict[str, Any]:
    expected_top_keys = {
        "contract_id",
        "schema_version",
        "protocol_id",
        "protocol_status",
        "dataset_id",
        "dataset_alias",
        "study_group_id",
        "independent_study_count",
        "authority",
        "scope",
        "input_contract",
        "mapping",
        "paper_faithful_measurement_transform",
        "route_a_companion_summary",
        "canonical_v3",
        "license_and_redistribution",
        "foundation_exposure",
        "split_and_leakage",
        "power_prefreeze",
        "qualification_gates",
        "output_contract",
        "known_external_evidence_blockers",
        "model_results_may_change_this_protocol",
    }
    _require_exact_keys(document, expected_top_keys, label="qualification protocol")
    for key, expected in (
        ("contract_id", CONTRACT_ID),
        ("schema_version", SCHEMA_VERSION),
        ("protocol_id", PROTOCOL_ID),
        ("protocol_status", PROTOCOL_STATUS),
        ("dataset_id", DATASET_ID),
        ("dataset_alias", DATASET_ALIAS),
        ("study_group_id", STUDY_GROUP_ID),
        ("independent_study_count", 1),
        ("model_results_may_change_this_protocol", False),
    ):
        _require_exact_value(document[key], expected, label=f"protocol.{key}")

    authority_keys = {
        "contract_path",
        "initial_contract_sha256",
        "contract_sha256",
        "accepted_a0_base_commit",
        "active_authority_commit",
        "active_amendment_decision_ids",
        "a1_qualification_path",
        "a1_qualification_sha256",
        "data_role_registry_path",
        "data_role_registry_sha256",
        "canonical_schema_path",
        "canonical_schema_sha256",
        "asset_manifest_path",
        "asset_manifest_sha256",
        "v4_helper_path",
        "v4_helper_sha256",
        "qualifier_path",
        "qualifier_sha256",
        "focused_test_path",
        "focused_test_sha256",
        "implementation_commit",
    }
    authority = _require_exact_keys(document["authority"], authority_keys, label="protocol.authority")
    for key in (
        "initial_contract_sha256",
        "contract_sha256",
        "a1_qualification_sha256",
        "data_role_registry_sha256",
        "canonical_schema_sha256",
        "asset_manifest_sha256",
        "v4_helper_sha256",
        "qualifier_sha256",
        "focused_test_sha256",
    ):
        _require_sha256(authority[key], label=f"protocol.authority.{key}")
    for key in (
        "contract_path",
        "a1_qualification_path",
        "data_role_registry_path",
        "canonical_schema_path",
        "asset_manifest_path",
        "v4_helper_path",
        "qualifier_path",
        "focused_test_path",
    ):
        if not isinstance(authority[key], str) or not authority[key]:
            raise QualificationError(f"protocol.authority.{key} must be a path")
    if not COMMIT_RE.fullmatch(str(authority["accepted_a0_base_commit"])):
        raise QualificationError("accepted A0 base commit must be a full commit hash")
    if not COMMIT_RE.fullmatch(str(authority["active_authority_commit"])):
        raise QualificationError("active authority commit must be a full commit hash")
    _require_exact_value(
        authority["active_amendment_decision_ids"],
        ["V3-DEC-017"],
        label="protocol.authority.active_amendment_decision_ids",
    )
    if authority["implementation_commit"] != "UNKNOWN_NOT_ASSERTED" and not COMMIT_RE.fullmatch(
        str(authority["implementation_commit"])
    ):
        raise QualificationError("implementation commit must be UNKNOWN_NOT_ASSERTED or a full commit hash")

    scope = document["scope"]
    for key, expected in (
        ("ordinary_public_data_only", True),
        ("included_contexts", list(CONTEXTS)),
        ("included_assays", list(ASSAYS)),
        ("included_endpoints", ["TRANSCRIPT", "TRANSLATION_EFFICIENCY"]),
        ("included_biological_replicates", list(REPLICATES)),
        ("region", "5UTR"),
        ("full_raw_geo_table_count", 18),
        ("supplement_count", 3),
        ("published_significant_only_tables_used_for_membership", False),
        ("endpoint_or_context_increases_independent_study_count", False),
        ("training_allowed", False),
        ("model_selection_allowed", False),
        ("authority_update_allowed_by_qualifier", False),
    ):
        if not isinstance(scope, Mapping) or key not in scope:
            raise QualificationError(f"protocol.scope.{key} is missing")
        _require_exact_value(scope[key], expected, label=f"protocol.scope.{key}")

    inputs = document["input_contract"]
    for key, expected in (
        ("asset_manifest_mode", "ADDITIVE_OVER_HASH_BOUND_P0_MANIFEST"),
        ("effective_asset_count", 21),
        ("raw_table_compression", "gzip"),
        ("raw_table_delimiter", "TAB"),
        ("raw_table_column_count", 2),
        ("raw_key_column_index", 0),
        ("raw_count_column_index", 1),
        ("raw_count_type", "NONNEGATIVE_INTEGER"),
        ("within_context_key_sets_must_be_exactly_equal", True),
        ("cross_context_missingness_allowed", True),
        ("cross_context_missing_is_zero", False),
        ("unexpected_regular_payload_action", "FAIL_CLOSED"),
        ("all_input_hashes_must_match_before_scientific_processing", True),
    ):
        if not isinstance(inputs, Mapping) or key not in inputs:
            raise QualificationError(f"protocol.input_contract.{key} is missing")
        _require_exact_value(inputs[key], expected, label=f"protocol.input_contract.{key}")

    mapping = document["mapping"]
    for key, expected in (
        ("mapping_source_asset_id", "GSE149487_LIM6C_293T"),
        ("mapping_sheet", "Sheet1"),
        ("description_column_index", 0),
        ("member_key_column_index", 1),
        ("sequence_source_asset_id", "GSE149487_MOESM8"),
        ("sequence_sheet", "6a 5' UTR sequences"),
        ("source_candidate_direction", "WT_TO_MUTANT"),
        ("edit_rule", "EXACTLY_ONE_SNV_WITH_REF_ALT_AND_POSITION_CONSISTENCY"),
        ("membership_may_depend_on_measured_effect_or_significance", False),
        ("all_unique_descriptions_must_be_classified", True),
        ("all_sequence_rows_must_be_classified", True),
        ("unadjudicated_class_action", "BLOCK_QUALIFICATION_AND_RETAIN_HASHED_REJECTION"),
    ):
        if not isinstance(mapping, Mapping) or key not in mapping:
            raise QualificationError(f"protocol.mapping.{key} is missing")
        _require_exact_value(mapping[key], expected, label=f"protocol.mapping.{key}")
    if mapping.get("outcome_blind_mapping_evidence_status") not in {
        "UNKNOWN_NOT_ASSERTED",
        "PASS",
    }:
        raise QualificationError("mapping evidence status is invalid")
    reconciliation = _require_exact_keys(
        mapping.get("raw_key_reconciliation"),
        {
            "classification_hash_domain",
            "mapped_class",
            "explicit_exclusion_classes",
            "prefrozen_excluded_key_sha256_by_class",
            "unclassified_raw_key_action",
            "mapping_key_absent_from_both_contexts_action",
            "cross_context_absence_imputed_as_zero",
            "raw_key_output_allowed",
        },
        label="protocol.mapping.raw_key_reconciliation",
    )
    for key, expected in (
        ("classification_hash_domain", RAW_KEY_CLASSIFICATION_HASH_DOMAIN),
        ("mapped_class", "MAPPED_TO_UNIQUE_DESCRIPTION"),
        ("explicit_exclusion_classes", ["EXCLUDED_CONTROL", "OUT_OF_SCOPE"]),
        ("unclassified_raw_key_action", "BLOCK_QUALIFICATION"),
        ("mapping_key_absent_from_both_contexts_action", "BLOCK_QUALIFICATION"),
        ("cross_context_absence_imputed_as_zero", False),
        ("raw_key_output_allowed", False),
    ):
        _require_exact_value(
            reconciliation[key],
            expected,
            label=f"protocol.mapping.raw_key_reconciliation.{key}",
        )
    exclusions = _require_exact_keys(
        reconciliation["prefrozen_excluded_key_sha256_by_class"],
        {"EXCLUDED_CONTROL", "OUT_OF_SCOPE"},
        label="protocol.mapping.raw_key_reconciliation.prefrozen exclusions",
    )
    seen_exclusion_hashes: set[str] = set()
    for classification in reconciliation["explicit_exclusion_classes"]:
        values = exclusions[classification]
        if not isinstance(values, list):
            raise QualificationError(
                f"protocol raw-key exclusion class {classification} must be a list"
            )
        for index, value in enumerate(values):
            digest = _require_sha256(
                value,
                label=f"protocol raw-key exclusion {classification}[{index}]",
            )
            if digest in seen_exclusion_hashes:
                raise QualificationError(
                    "protocol raw-key exclusion hashes must be unique across classes"
                )
            seen_exclusion_hashes.add(digest)

    transform = document["paper_faithful_measurement_transform"]
    fixed_transform_values = {
        "raw_count_to_cpm": "COUNT_DIVIDED_BY_SAMPLE_LIBRARY_SUM_TIMES_1E6",
        "original_cpm_minimum_inclusive": 0.5,
        "pseudocount": None,
        "transcript_ratio": "LOG2_TOTALRNA_CPM_MINUS_LOG2_DNA_CPM",
        "translation_efficiency_ratio": "LOG2_POLYSOME_CPM_MINUS_LOG2_TOTALRNA_CPM",
        "both_ratio_components_must_meet_threshold": True,
        "clipping_allowed": False,
        "missing_is_zero": False,
        "paper_test": "TWO_SIDED_MANN_WHITNEY_MUTANT_VS_WT",
        "paper_test_unit": "TECHNICAL_MEMBER_RATIO_DISTRIBUTION",
        "paper_test_pooling": "WITHIN_CONTEXT_AND_ENDPOINT_ACROSS_THREE_BIOLOGICAL_REPLICATES",
        "multiple_testing": "BENJAMINI_HOCHBERG",
        "multiple_testing_family": "WITHIN_CONTEXT_AND_ENDPOINT_ACROSS_ALL_OUTCOME_BLIND_STRICT_PAIRS",
        "fdr_threshold_exclusive": 0.1,
        "significance_may_change_membership": False,
    }
    if not isinstance(transform, Mapping):
        raise QualificationError("protocol paper transform must be an object")
    for key, expected in fixed_transform_values.items():
        if key not in transform:
            raise QualificationError(f"protocol paper transform is missing {key}")
        _require_exact_value(transform[key], expected, label=f"protocol paper transform.{key}")
    for key in (
        "method_source_status",
        "multiple_testing_family_status",
        "published_result_crosscheck_status",
    ):
        if transform.get(key) not in {"UNKNOWN_NOT_ASSERTED", "PASS"}:
            raise QualificationError(f"protocol paper transform.{key} is invalid")

    companion = document["route_a_companion_summary"]
    fixed_companion_values = {
        "classification": "ROUTE_A_COMPANION_SEPARATE_FROM_PAPER_NATIVE_TEST",
        "technical_member_role": "TECHNICAL_MEASUREMENT_UNIT_ONLY",
        "biological_replicate_role": "INFERENCE_UNIT",
        "within_arm_replicate_aggregation": "MEDIAN_OF_ELIGIBLE_TECHNICAL_MEMBER_LOG2_RATIOS",
        "within_replicate_effect": "CANDIDATE_MEDIAN_MINUS_SOURCE_MEDIAN",
        "across_replicate_effect": "EQUAL_WEIGHT_MEAN_OF_THREE_REPLICATE_EFFECTS",
        "standard_error": "SAMPLE_SD_OF_THREE_REPLICATE_EFFECTS_DIVIDED_BY_SQRT_3",
        "required_biological_replicates": 3,
        "minimum_eligible_members_per_arm_per_endpoint_per_replicate": 20,
        "technical_member_count_may_weight_replicates": False,
        "technical_member_count_may_increase_effective_n": False,
        "technical_member_count_may_define_standard_error": False,
    }
    if not isinstance(companion, Mapping):
        raise QualificationError("protocol companion summary must be an object")
    for key, expected in fixed_companion_values.items():
        if key not in companion:
            raise QualificationError(f"protocol companion summary is missing {key}")
        _require_exact_value(companion[key], expected, label=f"protocol companion.{key}")

    canonical = document["canonical_v3"]
    fixed_canonical_values = {
        "materialize_only_when_every_qualification_gate_passes": True,
        "data_role": "ORDINARY_A1_INTERVENTION",
        "evidence_status": "PASS",
        "claim_status": "NOT_ESTABLISHED",
        "sequence_alphabet": "DNA",
        "delta_direction": "CANDIDATE_MINUS_SOURCE",
        "region": "5UTR",
        "split_id": "ROUTE_A_V3_A1_GSE149487_QUALIFICATION_DEVELOPMENT_ONLY_V1",
        "split_partition": "DEVELOPMENT",
        "global_a2_benchmark_split_status": "NOT_RUN_DEFERRED_TO_A2",
        "qualified_record_reject_reason_code": None,
        "qualified_record_reject_reason_detail": None,
    }
    if not isinstance(canonical, Mapping):
        raise QualificationError("protocol canonical_v3 must be an object")
    for key, expected in fixed_canonical_values.items():
        if key not in canonical:
            raise QualificationError(f"protocol canonical_v3 is missing {key}")
        _require_exact_value(canonical[key], expected, label=f"protocol canonical_v3.{key}")

    license_protocol = _require_exact_keys(
        document["license_and_redistribution"],
        {
            "audit_status",
            "qualification_allowed_statuses",
            "unknown_status_blocks_qualification",
            "nonredistributable_release_mode",
            "license_id",
            "license_name",
            "license_uri",
            "verified_at",
        },
        label="protocol.license_and_redistribution",
    )
    if license_protocol["audit_status"] not in {"UNKNOWN_NOT_ASSERTED", "PASS"}:
        raise QualificationError("protocol license audit status is invalid")
    _require_exact_value(
        license_protocol["qualification_allowed_statuses"],
        ["VERIFIED_PUBLIC", "VERIFIED_NONREDISTRIBUTABLE"],
        label="protocol license qualification statuses",
    )
    _require_exact_value(
        license_protocol["unknown_status_blocks_qualification"],
        True,
        label="protocol license unknown blocker",
    )
    _require_exact_value(
        license_protocol["nonredistributable_release_mode"],
        "LOCATOR_HASH_AND_CODE_ONLY",
        label="protocol nonredistributable release mode",
    )
    for key in ("license_id", "verified_at"):
        if not isinstance(license_protocol[key], str) or not license_protocol[key]:
            raise QualificationError(f"protocol license {key} must be a string")
    for key in ("license_name", "license_uri"):
        if license_protocol[key] is not None and (
            not isinstance(license_protocol[key], str) or not license_protocol[key]
        ):
            raise QualificationError(f"protocol license {key} must be null or nonempty")

    exposure_protocol = _require_exact_keys(
        document["foundation_exposure"],
        {
            "audit_status",
            "checkpoint_id",
            "checkpoint_sha256",
            "stratum",
            "sequence_exposed",
            "label_exposed",
            "audit_id",
            "unknown_checkpoint_blocks_qualification",
        },
        label="protocol.foundation_exposure",
    )
    if exposure_protocol["audit_status"] not in {"UNKNOWN_NOT_ASSERTED", "PASS"}:
        raise QualificationError("protocol exposure audit status is invalid")
    if exposure_protocol["stratum"] not in {
        "UNTOUCHED",
        "SEQUENCE_EXPOSED",
        "LABEL_EXPOSED",
        "TRAINING_EXPOSED",
        "DEVELOPMENT_ONLY",
    }:
        raise QualificationError("protocol exposure stratum is invalid")
    for key in ("checkpoint_id", "checkpoint_sha256", "audit_id"):
        if not isinstance(exposure_protocol[key], str) or not exposure_protocol[key]:
            raise QualificationError(f"protocol exposure {key} must be a string")
    for key in ("sequence_exposed", "label_exposed", "unknown_checkpoint_blocks_qualification"):
        if type(exposure_protocol[key]) is not bool:
            raise QualificationError(f"protocol exposure {key} must be boolean")

    split = document["split_and_leakage"]
    fixed_split_values = {
        "qualification_partition": "DEVELOPMENT",
        "global_benchmark_split_may_be_claimed": False,
        "source_group_cross_partition_leakage_must_be_zero": True,
        "candidate_cross_partition_leakage_must_be_zero": True,
        "reverse_edge_cross_partition_leakage_must_be_zero": True,
        "gene_cross_partition_leakage_must_be_zero": True,
        "sequence_cluster_cross_partition_leakage_must_be_zero": True,
        "duplicate_directed_edge_count_must_be_zero": True,
    }
    if not isinstance(split, Mapping):
        raise QualificationError("protocol split_and_leakage must be an object")
    for key, expected in fixed_split_values.items():
        if key not in split:
            raise QualificationError(f"protocol split_and_leakage is missing {key}")
        _require_exact_value(split[key], expected, label=f"protocol split.{key}")

    power = document["power_prefreeze"]
    fixed_power_values = {
        "analysis_unit": "BIOLOGICAL_SOURCE_GROUP",
        "bootstrap_unit": "BIOLOGICAL_SOURCE_GROUP",
        "target_metric": "WITHIN_STUDY_SPEARMAN",
        "minimum_effect_at_alternative": 0.25,
        "alpha_two_sided": 0.05,
        "target_power": 0.8,
        "confidence_level": 0.95,
        "maximum_ci_full_width": 0.3,
        "simulation_seed": 20260810,
        "bootstrap_resamples": 2000,
        "simulation_trials": 1000,
        "model_results_may_change_this_rule": False,
    }
    if not isinstance(power, Mapping):
        raise QualificationError("protocol power_prefreeze must be an object")
    for key, expected in fixed_power_values.items():
        if key not in power:
            raise QualificationError(f"protocol power_prefreeze is missing {key}")
        _require_exact_value(power[key], expected, label=f"protocol power.{key}")

    output_contract = document["output_contract"]
    if not isinstance(output_contract, Mapping):
        raise QualificationError("protocol output_contract must be an object")
    for key, expected in (
        ("exclusive_new_output_directory_required", True),
        ("primary_publication_mode", PRIMARY_PUBLICATION_MODE),
        ("primary_atomic_sibling_staging_then_rename", True),
        ("atomic_no_replace_kernel_primitive_primary", True),
        (
            "primary_commit_marker_written_last_in_staging_before_atomic_rename",
            True,
        ),
        ("single_output_failure_transaction_claim_required", True),
        ("transaction_claim_loser_failure_record_allowed", False),
        (
            "directory_fsync_unsupported_errno_set",
            ["ENOSYS", "EINVAL", "ENOTSUP", "EOPNOTSUPP"],
        ),
        (
            "transaction_claim_parent_directory_fsync_unsupported_action",
            "CONTINUE_WITH_FIXED_CAPABILITY_WARNING",
        ),
        (
            "staging_directory_fsync_unsupported_action",
            "CONTINUE_WITH_FIXED_CAPABILITY_WARNING",
        ),
        (
            "staging_directory_fsync_other_error_action",
            "FAIL_CLOSED_BEFORE_PUBLICATION",
        ),
        (
            "failure_record_post_commit_directory_fsync_error_action",
            "RETURN_FAILURE_WITH_DURABILITY_WARNING",
        ),
        (
            "atomic_no_replace_unsupported_errno_fallback",
            ["ENOSYS", "EINVAL", "ENOTSUP", "EOPNOTSUPP"],
        ),
        ("fallback_publication_mode", FALLBACK_PUBLICATION_MODE),
        ("fallback_atomic_exclusive_final_mkdir_required", True),
        ("fallback_atomic_mkdir_loser_status", "CONTENDED_NO_FAILURE_RECORD"),
        ("fallback_bundle_file_write_mode", "O_EXCL_AND_FILE_FSYNC"),
        (
            "fallback_required_terminal_metadata_files",
            [PUBLICATION_COMMIT_FILENAME],
        ),
        (
            "required_terminal_metadata_files_all_publication_modes",
            [PUBLICATION_COMMIT_FILENAME],
        ),
        ("commit_marker_required_for_primary_and_fallback", True),
        (
            "unmarked_output_directory_acceptance",
            "REJECT_AS_PARTIAL_NOT_COMMITTED",
        ),
        ("post_commit_marker_validation_retry_count", 1),
        (
            "post_commit_marker_validation_failure_status",
            "COMMITTED_NOT_ACCEPTED",
        ),
        ("committed_not_accepted_failure_record_allowed", False),
        ("committed_not_accepted_canonical_accepted", False),
        ("fallback_commit_marker_schema_version", "1.0.0"),
        (
            "fallback_commit_marker_bindings",
            [
                "run_id",
                "execution_id",
                "SHA256SUMS_SHA256",
                "bundle_file_count",
                "bundle_filename_set_sha256",
                "final_output_directory_name_sha256",
                "final_output_target_sha256",
                "publication_mode",
                "committed_true",
            ],
        ),
        ("fallback_commit_marker_written_last", True),
        ("fallback_commit_marker_required_for_published_state", True),
        ("fallback_commit_marker_validation_required_before_published_return", True),
        ("fallback_directory_and_parent_fsync_after_commit_marker", True),
        ("fallback_partial_directory_status", "PARTIAL_NOT_COMMITTED"),
        ("fallback_partial_directory_must_be_preserved", True),
        ("fallback_partial_directory_canonical_accepted", False),
        ("fallback_retry_requires_new_run_id", True),
        ("success_stdout_includes_publication_mode", True),
        (
            "post_commit_durability_error_action",
            "RETURN_PUBLISHED_WITH_EXPLICIT_WARNING",
        ),
        (
            "scientific_parser_input_mode",
            "VERIFIED_BYTES_OR_PRIVATE_READ_ONLY_SNAPSHOT",
        ),
        ("retain_partial_staging_on_write_failure", True),
        ("raw_member_identifier_output_allowed", False),
        ("aggregate_only_stdout", True),
        ("always_materialized_files", list(ALWAYS_OUTPUT_FILES)),
        ("conditional_success_only_files", [CANONICAL_FILENAME]),
        ("failure_record_is_exclusive_sibling_json", True),
    ):
        if key not in output_contract:
            raise QualificationError(f"protocol output_contract is missing {key}")
        _require_exact_value(output_contract[key], expected, label=f"protocol output_contract.{key}")

    expected_gates = [
        "AUTHORITY_AND_CODE_TRUST_ROOTS",
        "EXACT_21_ASSET_MANIFEST_AND_PAYLOAD_INTEGRITY",
        "EXACT_18_TABLE_CONTEXT_ASSAY_REPLICATE_GRID",
        "WITHIN_CONTEXT_KEY_SET_ALIGNMENT_AND_MISSING_NOT_ZERO",
        "OUTCOME_BLIND_STRICT_SOURCE_CANDIDATE_MAPPING",
        "PAPER_NATIVE_TRANSFORM_TEST_AND_MULTIPLE_TESTING_REPRODUCTION",
        "THREE_BIOLOGICAL_REPLICATES_AND_ROUTE_A_SE",
        "CANONICAL_V3_SCHEMA_AND_HASH_LINEAGE",
        "LICENSE_AND_REDISTRIBUTION_AUDIT",
        "CHECKPOINT_SPECIFIC_FOUNDATION_EXPOSURE_AUDIT",
        "GROUP_AND_SEQUENCE_CLUSTER_LEAKAGE_AUDIT",
        "PREFROZEN_GROUP_POWER_SIMULATION",
    ]
    _require_exact_value(
        document["qualification_gates"], expected_gates, label="protocol qualification_gates"
    )
    if not isinstance(document["known_external_evidence_blockers"], list) or any(
        not isinstance(value, str) or not value
        for value in document["known_external_evidence_blockers"]
    ):
        raise QualificationError("known external blockers must be a list")
    if len(set(document["known_external_evidence_blockers"])) != len(
        document["known_external_evidence_blockers"]
    ):
        raise QualificationError("known external blockers must be unique")
    return dict(document)


def _validate_asset_manifest(document: Mapping[str, Any]) -> dict[str, Any]:
    _require_exact_keys(
        document,
        {
            "contract_id",
            "schema_version",
            "manifest_id",
            "manifest_status",
            "dataset_id",
            "dataset_alias",
            "expected_asset_count",
            "source_manifest",
            "inventory_policy",
            "assets",
        },
        label="asset manifest",
    )
    for key, expected in (
        ("contract_id", CONTRACT_ID),
        ("schema_version", "1.0.0"),
        ("manifest_id", ASSET_MANIFEST_ID),
        ("manifest_status", "PREFROZEN_ADDITIVE_TRUST_ROOT_BEFORE_FULL_QUALIFICATION"),
        ("dataset_id", DATASET_ID),
        ("dataset_alias", DATASET_ALIAS),
        ("expected_asset_count", 21),
    ):
        _require_exact_value(document[key], expected, label=f"asset manifest.{key}")
    source_manifest = _require_exact_keys(
        document["source_manifest"],
        {
            "filename",
            "sha256",
            "expected_declared_file_count",
            "role",
            "preserve_without_modification",
        },
        label="asset manifest.source_manifest",
    )
    for key, expected in (
        ("filename", "manifest.json"),
        ("expected_declared_file_count", 18),
        ("role", "HASH_BOUND_P0_GEO_RAW_COUNT_MANIFEST"),
        ("preserve_without_modification", True),
    ):
        _require_exact_value(source_manifest[key], expected, label=f"asset manifest.source_manifest.{key}")
    _require_sha256(source_manifest["sha256"], label="source manifest SHA-256")

    inventory = document["inventory_policy"]
    for key, expected in (
        ("mode", "ADDITIVE_OVER_HASH_BOUND_P0_MANIFEST"),
        ("geo_raw_assets_resolve_only_from_source_manifest", True),
        ("geo_raw_asset_sha256_and_bytes_inherited_only_after_unique_slot_resolution", True),
        ("supplement_assets_are_directly_sha256_bound", True),
        ("unexpected_regular_payload_action", "FAIL_CLOSED"),
        ("duplicate_logical_slot_action", "FAIL_CLOSED"),
        ("missing_logical_slot_action", "FAIL_CLOSED"),
        ("missing_is_zero", False),
        ("raw_identifier_output_allowed", False),
    ):
        if not isinstance(inventory, Mapping) or key not in inventory:
            raise QualificationError(f"asset manifest inventory policy is missing {key}")
        _require_exact_value(inventory[key], expected, label=f"asset manifest inventory.{key}")

    assets = document["assets"]
    if not isinstance(assets, list) or len(assets) != 21:
        raise QualificationError("asset manifest must contain exactly 21 assets")
    asset_ids: set[str] = set()
    raw_slots: set[tuple[str, str, int]] = set()
    raw_count = 0
    supplement_count = 0
    for index, asset in enumerate(assets):
        if not isinstance(asset, Mapping):
            raise QualificationError(f"asset manifest asset {index} must be an object")
        asset_id = _require_identifier(asset.get("asset_id"), label=f"asset {index} ID")
        if asset_id in asset_ids:
            raise QualificationError("asset manifest contains duplicate asset IDs")
        asset_ids.add(asset_id)
        kind = asset.get("asset_kind")
        license_block = asset.get("license")
        if not isinstance(license_block, Mapping):
            raise QualificationError(f"asset {asset_id} license must be an object")
        _require_exact_keys(
            license_block,
            {"status", "redistribution_allowed"},
            label=f"asset {asset_id} license",
        )
        if license_block.get("status") not in {
            "UNKNOWN_NOT_ASSERTED",
            "VERIFIED_PUBLIC",
            "VERIFIED_NONREDISTRIBUTABLE",
        }:
            raise QualificationError(f"asset {asset_id} license status is invalid")
        if license_block.get("redistribution_allowed") not in {None, True, False}:
            raise QualificationError(f"asset {asset_id} redistribution status is invalid")
        expected_redistribution = {
            "UNKNOWN_NOT_ASSERTED": None,
            "VERIFIED_PUBLIC": True,
            "VERIFIED_NONREDISTRIBUTABLE": False,
        }[license_block["status"]]
        if license_block["redistribution_allowed"] is not expected_redistribution:
            raise QualificationError(f"asset {asset_id} license/redistribution fields disagree")
        if kind == "GEO_RAW_COUNT":
            _require_exact_keys(
                asset,
                {
                    "asset_id",
                    "asset_kind",
                    "context",
                    "assay",
                    "biological_replicate",
                    "resolution",
                    "format",
                    "license",
                },
                label=f"raw asset {asset_id}",
            )
            raw_count += 1
            context = asset.get("context")
            assay = asset.get("assay")
            replicate = asset.get("biological_replicate")
            slot = (context, assay, replicate)
            if context not in CONTEXTS or assay not in ASSAYS or replicate not in REPLICATES:
                raise QualificationError(f"asset {asset_id} has an invalid raw grid slot")
            if slot in raw_slots:
                raise QualificationError("asset manifest contains a duplicate raw grid slot")
            raw_slots.add(slot)
            resolution = asset.get("resolution")
            if not isinstance(resolution, Mapping):
                raise QualificationError(f"asset {asset_id} resolution must be an object")
            _require_exact_keys(
                resolution,
                {"mode", "required_tokens", "replicate_regex"},
                label=f"asset {asset_id} resolution",
            )
            if resolution.get("mode") != "UNIQUE_CASEFOLD_FILENAME_TOKEN_MATCH_IN_HASH_BOUND_P0_MANIFEST":
                raise QualificationError(f"asset {asset_id} resolution mode is invalid")
            tokens = resolution.get("required_tokens")
            if not isinstance(tokens, list) or len(tokens) != 2 or any(
                not isinstance(token, str) or not token for token in tokens
            ):
                raise QualificationError(f"asset {asset_id} resolution tokens are invalid")
            pattern = resolution.get("replicate_regex")
            try:
                re.compile(str(pattern), re.IGNORECASE)
            except re.error as exc:
                raise QualificationError(f"asset {asset_id} replicate regex is invalid") from exc
            fmt = asset.get("format")
            expected_format = {
                "compression": "gzip",
                "delimiter": "TAB",
                "column_count": 2,
                "key_column_index": 0,
                "count_column_index": 1,
                "count_type": "NONNEGATIVE_INTEGER",
            }
            if not isinstance(fmt, Mapping) or dict(fmt) != expected_format:
                raise QualificationError(f"asset {asset_id} format is not pre-frozen")
        elif kind == "SUPPLEMENT_WORKBOOK":
            supplement_count += 1
            expected_supplement_keys = {
                "asset_id",
                "asset_kind",
                "filename",
                "bytes",
                "sha256",
                "source_uri",
                "role",
                "license",
            }
            if "workbook_contract" in asset:
                expected_supplement_keys.add("workbook_contract")
            if asset_id == "GSE149487_LIM6C_293T":
                expected_supplement_keys.update(
                    {
                        "source_commit",
                        "source_blob_git_sha1",
                        "source_commit_author_date",
                    }
                )
            _require_exact_keys(
                asset,
                expected_supplement_keys,
                label=f"supplement asset {asset_id}",
            )
            filename = asset.get("filename")
            if not isinstance(filename, str) or Path(filename).name != filename or not filename.endswith(".xlsx"):
                raise QualificationError(f"asset {asset_id} supplement filename is unsafe")
            if isinstance(asset.get("bytes"), bool) or not isinstance(asset.get("bytes"), int) or asset["bytes"] <= 0:
                raise QualificationError(f"asset {asset_id} supplement bytes are invalid")
            _require_sha256(asset.get("sha256"), label=f"asset {asset_id} SHA-256")
            if not isinstance(asset.get("source_uri"), str) or not asset["source_uri"]:
                raise QualificationError(f"asset {asset_id} source URI is missing")
            if asset_id == "GSE149487_LIM6C_293T":
                source_commit = str(asset.get("source_commit"))
                if source_commit != "d613b541d192d6c502a1ef8849c27e801a7fbfb9":
                    raise QualificationError("mapping workbook source commit is not pre-frozen")
                if source_commit not in asset["source_uri"]:
                    raise QualificationError("mapping workbook source URI is not commit-immutable")
                if asset.get("source_blob_git_sha1") != (
                    "2d4bae738074a1d1bffbafc5ec39da1dff679807"
                ):
                    raise QualificationError("mapping workbook Git blob SHA-1 is not pre-frozen")
                _require_exact_value(
                    asset.get("source_commit_author_date"),
                    "2024-06-17T19:25:39Z",
                    label="mapping workbook source commit author date",
                )
        else:
            raise QualificationError(f"asset {asset_id} kind is invalid")
    expected_slots = {
        (context, assay, replicate)
        for context in CONTEXTS
        for assay in ASSAYS
        for replicate in REPLICATES
    }
    if raw_count != 18 or supplement_count != 3 or raw_slots != expected_slots:
        raise QualificationError("asset manifest does not close the exact 18+3 inventory")
    return dict(document)


def _source_manifest_entries(document: Mapping[str, Any]) -> list[Mapping[str, Any]]:
    entries = document.get("files")
    if not isinstance(entries, list):
        raise QualificationError("P0 source manifest files must be a list")
    return entries


def _entry_bytes(entry: Mapping[str, Any], *, label: str) -> int:
    candidates = [entry.get("bytes"), entry.get("expected_bytes")]
    values = [value for value in candidates if isinstance(value, int) and not isinstance(value, bool)]
    if not values or any(value < 0 for value in values):
        raise QualificationError(f"{label} has no valid declared byte count")
    if len(set(values)) > 1:
        raise QualificationError(f"{label} declared byte counts disagree")
    return values[0]


def _resolve_and_verify_assets(
    *,
    data_root: Path,
    p0_manifest_path: Path,
    asset_manifest: Mapping[str, Any],
    snapshot_root: Path,
) -> tuple[list[dict[str, Any]], dict[str, Path], dict[str, Any]]:
    source_spec = asset_manifest["source_manifest"]
    if p0_manifest_path.name != source_spec["filename"] or p0_manifest_path.parent != data_root:
        raise QualificationError("P0 manifest must be the bound manifest inside the data root")
    p0_payload, p0_provenance = _read_verified_bytes(
        p0_manifest_path,
        source_spec["sha256"],
        label="P0 source manifest",
        suffix=".json",
    )
    p0_document = _read_json_bytes(p0_payload, label="P0 source manifest")
    source_entries = _source_manifest_entries(p0_document)
    if len(source_entries) != source_spec["expected_declared_file_count"]:
        raise QualificationError("P0 source manifest does not declare exactly 18 raw files")

    normalized_entries: list[dict[str, Any]] = []
    seen_names: set[str] = set()
    for index, entry in enumerate(source_entries):
        if not isinstance(entry, Mapping):
            raise QualificationError(f"P0 source manifest entry {index} must be an object")
        name = entry.get("name")
        if not isinstance(name, str) or not name or Path(name).name != name:
            raise QualificationError(f"P0 source manifest entry {index} has an unsafe name")
        if name in seen_names:
            raise QualificationError("P0 source manifest contains duplicate filenames")
        seen_names.add(name)
        if entry.get("downloaded") is not True:
            raise QualificationError(f"P0 source manifest entry {name} is not marked downloaded")
        normalized_entries.append(
            {
                "name": name,
                "sha256": _require_sha256(entry.get("sha256"), label=f"P0 entry {name} SHA-256"),
                "bytes": _entry_bytes(entry, label=f"P0 entry {name}"),
                "url": entry.get("url") if isinstance(entry.get("url"), str) else "UNKNOWN_NOT_ASSERTED",
            }
        )

    resolved_assets: list[dict[str, Any]] = []
    paths: dict[str, Path] = {}
    claimed_source_names: set[str] = set()
    asset_snapshot_root = snapshot_root / "assets"
    asset_snapshot_root.mkdir(mode=0o700)
    for asset_index, spec in enumerate(asset_manifest["assets"]):
        asset_id = spec["asset_id"]
        if spec["asset_kind"] == "GEO_RAW_COUNT":
            required_tokens = [token.casefold() for token in spec["resolution"]["required_tokens"]]
            replicate_pattern = re.compile(spec["resolution"]["replicate_regex"], re.IGNORECASE)
            matches = [
                entry
                for entry in normalized_entries
                if all(token in entry["name"].casefold() for token in required_tokens)
                and replicate_pattern.search(entry["name"]) is not None
            ]
            if len(matches) != 1:
                raise QualificationError(
                    f"raw logical slot {asset_id} resolved to {len(matches)} source entries"
                )
            source_entry = matches[0]
            if source_entry["name"] in claimed_source_names:
                raise QualificationError("one P0 source entry resolved to multiple raw logical slots")
            claimed_source_names.add(source_entry["name"])
            path = data_root / source_entry["name"]
            snapshot_path = asset_snapshot_root / f"{asset_index:02d}_{source_entry['name']}"
            observed = _snapshot_verified_file(
                path,
                snapshot_path,
                source_entry["sha256"],
                expected_bytes=source_entry["bytes"],
                label=f"raw asset {asset_id}",
                suffix=".gz",
            )
            effective = {
                "asset_id": asset_id,
                "asset_kind": "GEO_RAW_COUNT",
                "context": spec["context"],
                "assay": spec["assay"],
                "biological_replicate": spec["biological_replicate"],
                "filename": source_entry["name"],
                "bytes": observed["bytes"],
                "sha256": observed["sha256"],
                "source_uri": source_entry["url"],
                "license": dict(spec["license"]),
                "resolved_from_p0_manifest_sha256": p0_provenance["sha256"],
                "parser_input_mode": PRIVATE_READ_ONLY_SNAPSHOT_MODE,
            }
        else:
            path = data_root / spec["filename"]
            snapshot_path = asset_snapshot_root / f"{asset_index:02d}_{spec['filename']}"
            observed = _snapshot_verified_file(
                path,
                snapshot_path,
                spec["sha256"],
                expected_bytes=spec["bytes"],
                label=f"supplement asset {asset_id}",
                suffix=".xlsx",
            )
            effective = {
                "asset_id": asset_id,
                "asset_kind": "SUPPLEMENT_WORKBOOK",
                "filename": spec["filename"],
                "bytes": observed["bytes"],
                "sha256": observed["sha256"],
                "source_uri": spec["source_uri"],
                "role": spec["role"],
                "license": dict(spec["license"]),
                "parser_input_mode": PRIVATE_READ_ONLY_SNAPSHOT_MODE,
            }
            if "source_commit" in spec:
                effective["source_commit"] = spec["source_commit"]
                effective["source_blob_git_sha1"] = spec["source_blob_git_sha1"]
                effective["source_commit_author_date"] = spec[
                    "source_commit_author_date"
                ]
            if "workbook_contract" in spec:
                effective["workbook_contract"] = dict(spec["workbook_contract"])
        resolved_assets.append(effective)
        paths[asset_id] = snapshot_path

    if claimed_source_names != seen_names:
        raise QualificationError("not every P0 raw source entry resolved to exactly one logical slot")
    allowed_names = {p0_manifest_path.name} | {asset["filename"] for asset in resolved_assets}
    observed_names: set[str] = set()
    for child in data_root.iterdir():
        _reject_forbidden_path(child, label="data-root child")
        if child.is_symlink():
            raise ScopeViolation("data root must not contain symlink payloads")
        if child.is_file():
            observed_names.add(child.name)
        elif child.is_dir():
            raise QualificationError("data root contains an unexpected directory")
        else:
            raise QualificationError("data root contains an unsupported filesystem entry")
    if observed_names != allowed_names:
        raise QualificationError(
            "data-root inventory differs from the exact 21 assets plus source manifest"
        )
    resolved_assets.sort(key=lambda row: row["asset_id"])
    return resolved_assets, paths, p0_provenance


def _load_v4_helpers(path: Path) -> Any:
    spec = importlib.util.spec_from_file_location("route_a_v3_plumage_v4_helpers", path)
    if spec is None or spec.loader is None:
        raise QualificationError("could not load the hash-bound V4 helper module")
    module = importlib.util.module_from_spec(spec)
    try:
        spec.loader.exec_module(module)
    except Exception as exc:  # pragma: no cover - import failures are environment-specific
        raise QualificationError("hash-bound V4 helper module import failed") from exc
    required = (
        "_header_sha256",
        "_parse_6a_table",
        "_parse_description",
        "_matching_sequence_pair",
        "MOESM8_6A_COLUMNS",
    )
    if any(not hasattr(module, name) for name in required):
        raise QualificationError("hash-bound V4 helper module is missing required functions")
    return module


def _asset_by_id(assets: Sequence[Mapping[str, Any]], asset_id: str) -> Mapping[str, Any]:
    matches = [asset for asset in assets if asset["asset_id"] == asset_id]
    if len(matches) != 1:
        raise QualificationError(f"effective asset {asset_id} is not unique")
    return matches[0]


def _load_sequence_universe(
    *,
    path: Path,
    effective_asset: Mapping[str, Any],
    helper: Any,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], Counter[str]]:
    contract = effective_asset.get("workbook_contract")
    if not isinstance(contract, Mapping):
        raise QualificationError("MOESM8 effective asset is missing workbook contract")
    sheet_name = contract.get("sequence_sheet")
    expected_header = _require_sha256(
        contract.get("sequence_header_sha256"), label="MOESM8 sequence header SHA-256"
    )
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise QualificationError("MOESM8 workbook could not be opened") from exc
    try:
        if sheet_name not in workbook.sheetnames:
            raise QualificationError("MOESM8 sequence sheet is missing")
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration as exc:
            raise QualificationError("MOESM8 sequence sheet is empty") from exc
        if helper._header_sha256(header) != expected_header:
            raise QualificationError("MOESM8 sequence header SHA-256 mismatch")
        values = list(rows)
    finally:
        workbook.close()
    if len(header) != len(helper.MOESM8_6A_COLUMNS):
        raise QualificationError("MOESM8 sequence sheet column count mismatch")
    table = pd.DataFrame(values, columns=list(helper.MOESM8_6A_COLUMNS))
    try:
        entries, classifications = helper._parse_6a_table(
            table,
            moesm8_sha256=effective_asset["sha256"],
        )
    except Exception as exc:
        raise QualificationError("MOESM8 strict sequence-universe parsing failed") from exc
    counts = Counter(str(row["classification"]) for row in classifications)
    if sum(counts.values()) != len(values):
        raise QualificationError("MOESM8 sequence classification denominator mismatch")
    safe_audit = [
        {
            "audit_type": "SEQUENCE_UNIVERSE_CLASSIFICATION",
            "row_locator_sha256": row["row_locator_sha256"],
            "coordinate_sha256": row["coordinate_sha256"],
            "classification": row["classification"],
            "reason": row["classification_reason"],
            "raw_coordinate_emitted": False,
            "raw_sequence_emitted": False,
        }
        for row in classifications
    ]
    return [dict(entry) for entry in entries], safe_audit, counts


def _load_outcome_blind_member_mapping(
    *,
    path: Path,
    effective_asset: Mapping[str, Any],
    helper: Any,
    sequence_entries: Sequence[Mapping[str, Any]],
    mapping_evidence_status: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], Counter[str], set[str]]:
    contract = effective_asset.get("workbook_contract")
    if not isinstance(contract, Mapping):
        raise QualificationError("mapping workbook effective asset is missing workbook contract")
    sheet_name = contract.get("mapping_sheet")
    expected_header = _require_sha256(
        contract.get("mapping_header_sha256"), label="mapping header SHA-256"
    )
    description_column = contract.get("description_column_index")
    member_column = contract.get("member_key_column_index")
    if description_column != 0 or member_column != 1:
        raise QualificationError("mapping workbook columns are not pre-frozen")
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise QualificationError("mapping workbook could not be opened") from exc
    try:
        if sheet_name not in workbook.sheetnames:
            raise QualificationError("mapping workbook sheet is missing")
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration as exc:
            raise QualificationError("mapping workbook sheet is empty") from exc
        if helper._header_sha256(header) != expected_header:
            raise QualificationError("mapping workbook header SHA-256 mismatch")
        member_to_description: dict[str, str] = {}
        description_to_members: MutableMapping[str, set[str]] = defaultdict(set)
        duplicate_identical_rows = 0
        for physical_row, row in enumerate(rows, start=2):
            description_value = row[description_column] if len(row) > description_column else None
            member_value = row[member_column] if len(row) > member_column else None
            if description_value is None and member_value is None:
                continue
            if (
                not isinstance(description_value, str)
                or not description_value
                or description_value.strip() != description_value
                or not isinstance(member_value, str)
                or not member_value
                or member_value.strip() != member_value
            ):
                raise QualificationError("mapping workbook contains an invalid description/member row")
            existing = member_to_description.get(member_value)
            if existing is not None:
                if existing != description_value:
                    raise QualificationError("one member key maps to conflicting construct descriptions")
                duplicate_identical_rows += 1
                continue
            member_to_description[member_value] = description_value
            description_to_members[description_value].add(member_value)
    finally:
        workbook.close()
    if not member_to_description:
        raise QualificationError("mapping workbook contains no member-to-description rows")

    descriptions = sorted(description_to_members)
    parsed = {description: helper._parse_description(description) for description in descriptions}
    reasons: Counter[str] = Counter()
    audit_rows: list[dict[str, Any]] = []
    pairs: list[dict[str, Any]] = []
    used_wt: set[str] = set()
    decisions: dict[str, tuple[str, str, str | None]] = {}

    for description in descriptions:
        fields = parsed[description]
        description_digest = _sha256_bytes(
            f"{effective_asset['sha256']}\0{description}".encode("utf-8")
        )
        if fields is None:
            reason = "UNADJUDICATED_DESCRIPTION_CLASS"
            reasons[reason] += 1
            decisions[description] = ("REJECTED_UNADJUDICATED_DESCRIPTION", reason, None)
            continue
        if fields["type"] == "wt":
            decisions[description] = ("REFERENCE_ONLY_STRICT_WT", "REFERENCE_PENDING_PAIR_USE", None)
            continue
        wt_description = (
            f"{fields['gene']}_WT_{fields['chrom']}_{fields['start']}_{fields['end']}"
        )
        if wt_description not in description_to_members:
            reason = "MISSING_WT_MAPPING_DESCRIPTION"
            reasons[reason] += 1
            decisions[description] = ("REJECTED_STRICT_MUTANT", reason, None)
            continue
        try:
            sequence_pair, sequence_reason = helper._matching_sequence_pair(fields, sequence_entries)
        except Exception as exc:
            raise QualificationError("strict source-candidate sequence matching failed") from exc
        if sequence_pair is None:
            reason = str(sequence_reason or "UNKNOWN_SEQUENCE_PAIR_REJECTION")
            reasons[reason] += 1
            decisions[description] = ("REJECTED_STRICT_MUTANT", reason, None)
            continue
        source_sequence = str(sequence_pair["source_sequence"])
        candidate_sequence = str(sequence_pair["candidate_sequence"])
        if not SEQUENCE_RE.fullmatch(source_sequence) or not SEQUENCE_RE.fullmatch(candidate_sequence):
            raise QualificationError("strict sequence pair contains a non-DNA sequence")
        pair_id = _stable_id(
            "GSE149487_PAIR_",
            {
                "description_sha256": description_digest,
                "source_sequence_sha256": _sha256_bytes(source_sequence.encode("ascii")),
                "candidate_sequence_sha256": _sha256_bytes(candidate_sequence.encode("ascii")),
            },
        )
        group_id = _stable_id(
            "GSE149487_GROUP_",
            {
                "gene": fields["gene"],
                "chrom": fields["chrom"],
                "start": fields["start"],
                "end": fields["end"],
                "source_sequence": source_sequence,
            },
        )
        source_id = _stable_id(
            "GSE149487_SOURCE_",
            {"wt_description": wt_description, "source_sequence": source_sequence},
        )
        candidate_id = _stable_id(
            "GSE149487_CANDIDATE_",
            {"mutant_description": description, "candidate_sequence": candidate_sequence},
        )
        pairs.append(
            {
                "pair_id": pair_id,
                "biological_source_group_id": group_id,
                "source_id": source_id,
                "candidate_id": candidate_id,
                "gene_group_id": _stable_id("GSE149487_GENE_", {"gene": fields["gene"]}),
                "locus_id": _stable_id(
                    "GSE149487_LOCUS_",
                    {
                        "chrom": fields["chrom"],
                        "start": fields["start"],
                        "end": fields["end"],
                    },
                ),
                "source_sequence": source_sequence,
                "candidate_sequence": candidate_sequence,
                "source_sequence_sha256": _sha256_bytes(source_sequence.encode("ascii")),
                "candidate_sequence_sha256": _sha256_bytes(candidate_sequence.encode("ascii")),
                "sequence_index_0_based": int(sequence_pair["sequence_index_0_based"]),
                "ref": fields["ref"],
                "alt": fields["alt"],
                "source_members": set(description_to_members[wt_description]),
                "candidate_members": set(description_to_members[description]),
                "mutant_description_sha256": description_digest,
                "wt_description_sha256": _sha256_bytes(
                    f"{effective_asset['sha256']}\0{wt_description}".encode("utf-8")
                ),
            }
        )
        used_wt.add(wt_description)
        decisions[description] = ("INCLUDED_STRICT_SNV_MUTANT", "STRICT_OUTCOME_BLIND_PAIR", pair_id)

    for description in descriptions:
        classification, reason, pair_id = decisions[description]
        if classification == "REFERENCE_ONLY_STRICT_WT":
            reason = "WT_REFERENCE_FOR_INCLUDED_PAIR" if description in used_wt else "WT_REFERENCE_UNUSED"
        audit_rows.append(
            {
                "audit_type": "MAPPING_DESCRIPTION_CLASSIFICATION",
                "description_sha256": _sha256_bytes(
                    f"{effective_asset['sha256']}\0{description}".encode("utf-8")
                ),
                "classification": classification,
                "reason": reason,
                "pair_id": pair_id,
                "technical_member_count": len(description_to_members[description]),
                "raw_description_emitted": False,
                "raw_member_identifier_emitted": False,
            }
        )
    if duplicate_identical_rows:
        reasons["DUPLICATE_IDENTICAL_MEMBER_MAPPING_ROWS"] += duplicate_identical_rows
    pairs.sort(key=lambda row: row["pair_id"])
    audit_rows.sort(key=lambda row: (row["audit_type"], row["description_sha256"]))
    pair_audit = [
        {
            "audit_type": "STRICT_SOURCE_CANDIDATE_PAIR",
            "pair_id": pair["pair_id"],
            "biological_source_group_id": pair["biological_source_group_id"],
            "source_sequence_sha256": pair["source_sequence_sha256"],
            "candidate_sequence_sha256": pair["candidate_sequence_sha256"],
            "source_technical_member_count": len(pair["source_members"]),
            "candidate_technical_member_count": len(pair["candidate_members"]),
            "edit_count": 1,
            "outcome_blind_mapping_evidence_status": mapping_evidence_status,
            "raw_member_identifier_emitted": False,
            "raw_sequence_emitted": False,
        }
        for pair in pairs
    ]
    mapped_members = set(member_to_description)
    return pairs, audit_rows, pair_audit, reasons, mapped_members


def _key_set_sha256(keys: Iterable[str], *, asset_sha256: str) -> str:
    digest = hashlib.sha256()
    digest.update(asset_sha256.encode("ascii"))
    digest.update(b"\0SORTED_OPAQUE_KEYS\0")
    for key in sorted(keys):
        digest.update(key.encode("utf-8"))
        digest.update(b"\0")
    return digest.hexdigest()


def _raw_key_classification_sha256(key: str) -> str:
    return _sha256_bytes(
        f"{RAW_KEY_CLASSIFICATION_HASH_DOMAIN}\0{key}".encode("utf-8")
    )


def _classified_key_set_sha256(keys: Iterable[str], *, label: str) -> str:
    digest = hashlib.sha256()
    digest.update(RAW_KEY_CLASSIFICATION_HASH_DOMAIN.encode("ascii"))
    digest.update(b"\0SET\0")
    digest.update(label.encode("ascii"))
    digest.update(b"\0")
    for key_digest in sorted(_raw_key_classification_sha256(key) for key in keys):
        digest.update(key_digest.encode("ascii"))
        digest.update(b"\0")
    return digest.hexdigest()


def _read_raw_count_table(
    *,
    path: Path,
    asset: Mapping[str, Any],
    mapped_members: set[str],
) -> tuple[dict[str, int], set[str], dict[str, Any]]:
    counts_for_mapped_members: dict[str, int] = {}
    all_keys: set[str] = set()
    library_sum = 0
    row_count = 0
    header_sha256: str | None = None
    try:
        with gzip.open(path, "rt", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle, delimiter="\t", strict=True)
            try:
                header = next(reader)
            except StopIteration as exc:
                raise QualificationError(f"raw count asset {asset['asset_id']} is empty") from exc
            if len(header) != 2 or any(not isinstance(value, str) or not value for value in header):
                raise QualificationError(f"raw count asset {asset['asset_id']} header is not two nonempty fields")
            header_sha256 = _sha256_bytes(_compact_json_bytes(header))
            for line_number, row in enumerate(reader, start=2):
                if len(row) != 2:
                    raise QualificationError(f"raw count asset {asset['asset_id']} has a non-two-column row")
                key, count_text = row
                if not key or key.strip() != key:
                    raise QualificationError(f"raw count asset {asset['asset_id']} has an invalid opaque key")
                if key in all_keys:
                    raise QualificationError(f"raw count asset {asset['asset_id']} has a duplicate opaque key")
                if not NONNEGATIVE_INTEGER_RE.fullmatch(count_text):
                    raise QualificationError(f"raw count asset {asset['asset_id']} has a nonnegative-integer violation")
                count = int(count_text)
                all_keys.add(key)
                library_sum += count
                row_count += 1
                if key in mapped_members:
                    counts_for_mapped_members[key] = count
    except (OSError, UnicodeDecodeError, csv.Error, EOFError) as exc:
        raise QualificationError(f"raw count asset {asset['asset_id']} gzip/TSV validation failed") from exc
    if row_count == 0 or library_sum <= 0 or header_sha256 is None:
        raise QualificationError(f"raw count asset {asset['asset_id']} has no usable count universe")
    return counts_for_mapped_members, all_keys, {
        "asset_id": asset["asset_id"],
        "context": asset["context"],
        "assay": asset["assay"],
        "biological_replicate": asset["biological_replicate"],
        "row_count": row_count,
        "unique_key_count": len(all_keys),
        "mapped_key_count": len(counts_for_mapped_members),
        "unmapped_key_count": len(all_keys - mapped_members),
        "library_sum": library_sum,
        "header_sha256": header_sha256,
        "key_set_sha256": _key_set_sha256(all_keys, asset_sha256=asset["sha256"]),
        "raw_keys_emitted": False,
    }


def _load_geo_count_grid(
    *,
    resolved_assets: Sequence[Mapping[str, Any]],
    asset_paths: Mapping[str, Path],
    mapped_members: set[str],
    reconciliation_contract: Mapping[str, Any],
) -> tuple[
    dict[tuple[str, str, int], dict[str, int]],
    dict[tuple[str, str, int], int],
    list[dict[str, Any]],
    dict[str, Any],
    dict[str, Any],
]:
    raw_assets = [asset for asset in resolved_assets if asset["asset_kind"] == "GEO_RAW_COUNT"]
    raw_assets.sort(key=lambda row: (row["context"], row["assay"], row["biological_replicate"]))
    if len(raw_assets) != 18:
        raise QualificationError("effective raw grid must contain exactly 18 assets")
    counts: dict[tuple[str, str, int], dict[str, int]] = {}
    library_sums: dict[tuple[str, str, int], int] = {}
    table_audits: list[dict[str, Any]] = []
    reference_key_sets: dict[str, set[str]] = {}
    for asset in raw_assets:
        slot = (asset["context"], asset["assay"], int(asset["biological_replicate"]))
        if slot in counts:
            raise QualificationError("effective raw grid contains duplicate slots")
        selected_counts, all_keys, audit = _read_raw_count_table(
            path=asset_paths[asset["asset_id"]],
            asset=asset,
            mapped_members=mapped_members,
        )
        reference = reference_key_sets.get(asset["context"])
        if reference is None:
            reference_key_sets[asset["context"]] = all_keys
        elif all_keys != reference:
            raise QualificationError(
                f"raw key sets are not exactly aligned within context {asset['context']}"
            )
        counts[slot] = selected_counts
        library_sums[slot] = int(audit["library_sum"])
        table_audits.append(audit)
    expected_slots = {
        (context, assay, replicate)
        for context in CONTEXTS
        for assay in ASSAYS
        for replicate in REPLICATES
    }
    if set(counts) != expected_slots:
        raise QualificationError("loaded raw grid does not equal the exact 18 logical slots")
    pc3 = reference_key_sets["PC3"]
    t293 = reference_key_sets["293T"]
    cross_context = {
        "pc3_unique_key_count": len(pc3),
        "293t_unique_key_count": len(t293),
        "intersection_key_count": len(pc3 & t293),
        "pc3_only_key_count": len(pc3 - t293),
        "293t_only_key_count": len(t293 - pc3),
        "union_key_count": len(pc3 | t293),
        "missing_is_zero": False,
        "cross_context_missingness_allowed": True,
        "raw_keys_emitted": False,
    }
    exclusion_hashes = reconciliation_contract[
        "prefrozen_excluded_key_sha256_by_class"
    ]
    raw_union = pc3 | t293
    raw_key_digests = {
        key: _raw_key_classification_sha256(key) for key in raw_union
    }
    mapped_and_excluded = {
        key
        for key in raw_union & mapped_members
        if any(
            raw_key_digests[key] in set(exclusion_hashes[classification])
            for classification in reconciliation_contract["explicit_exclusion_classes"]
        )
    }
    per_context: dict[str, dict[str, Any]] = {}
    unclassified_union: set[str] = set()
    for context, universe in (("PC3", pc3), ("293T", t293)):
        mapped = universe & mapped_members
        classified_exclusions: dict[str, set[str]] = {}
        for classification in reconciliation_contract["explicit_exclusion_classes"]:
            allowed_hashes = set(exclusion_hashes[classification])
            classified_exclusions[classification] = {
                key
                for key in universe - mapped
                if raw_key_digests[key] in allowed_hashes
            }
        excluded = set().union(*classified_exclusions.values())
        unclassified = universe - mapped - excluded
        unclassified_union.update(unclassified)
        class_counts = {
            "MAPPED_TO_UNIQUE_DESCRIPTION": len(mapped),
            **{
                classification: len(classified_exclusions[classification])
                for classification in reconciliation_contract["explicit_exclusion_classes"]
            },
            "UNCLASSIFIED": len(unclassified),
        }
        per_context[context] = {
            "raw_key_count": len(universe),
            "classification_counts": class_counts,
            "raw_key_set_sha256": _classified_key_set_sha256(
                universe,
                label=f"{context}_RAW_UNIVERSE",
            ),
            "mapped_key_set_sha256": _classified_key_set_sha256(
                mapped,
                label=f"{context}_MAPPED",
            ),
            "excluded_key_set_sha256": _classified_key_set_sha256(
                excluded,
                label=f"{context}_EXCLUDED",
            ),
            "unclassified_key_set_sha256": _classified_key_set_sha256(
                unclassified,
                label=f"{context}_UNCLASSIFIED",
            ),
            "classification_complete_and_mutually_exclusive": (
                not unclassified and not (universe & mapped_and_excluded)
            ),
            "raw_keys_emitted": False,
        }
    mapping_absent_both = mapped_members - raw_union
    prefrozen_exclusion_hashes = set().union(
        *(set(exclusion_hashes[classification]) for classification in reconciliation_contract["explicit_exclusion_classes"])
    )
    observed_exclusion_hashes = {
        raw_key_digests[key]
        for key in raw_union - mapped_members
        if raw_key_digests[key] in prefrozen_exclusion_hashes
    }
    reconciliation = {
        "record_type": "GSE149487_OUTCOME_BLIND_RAW_KEY_RECONCILIATION",
        "classification_hash_domain": RAW_KEY_CLASSIFICATION_HASH_DOMAIN,
        "per_context": per_context,
        "raw_union_key_count": len(raw_union),
        "mapped_workbook_key_count": len(mapped_members),
        "unclassified_raw_key_count": len(unclassified_union),
        "unclassified_raw_key_set_sha256": _classified_key_set_sha256(
            unclassified_union,
            label="UNCLASSIFIED_RAW_UNION",
        ),
        "mapped_and_explicit_exclusion_overlap_count": len(mapped_and_excluded),
        "mapping_key_absent_from_both_contexts_count": len(mapping_absent_both),
        "mapping_key_absent_from_both_contexts_set_sha256": _classified_key_set_sha256(
            mapping_absent_both,
            label="MAPPING_ABSENT_BOTH_CONTEXTS",
        ),
        "prefrozen_exclusion_hash_count": len(prefrozen_exclusion_hashes),
        "observed_prefrozen_exclusion_hash_count": len(observed_exclusion_hashes),
        "cross_context_absence_imputed_as_zero": False,
        "raw_keys_emitted": False,
    }
    reconciliation["status"] = (
        "PASS"
        if reconciliation["unclassified_raw_key_count"] == 0
        and reconciliation["mapped_and_explicit_exclusion_overlap_count"] == 0
        and reconciliation["mapping_key_absent_from_both_contexts_count"] == 0
        else "BLOCKED_OUTCOME_BLIND_KEY_RECONCILIATION"
    )
    table_audits.sort(key=lambda row: row["asset_id"])
    return counts, library_sums, table_audits, cross_context, reconciliation


def _eligible_log2_ratios(
    *,
    members: Iterable[str],
    numerator_counts: Mapping[str, int],
    denominator_counts: Mapping[str, int],
    numerator_library_sum: int,
    denominator_library_sum: int,
    minimum_cpm: float,
) -> tuple[list[float], Counter[str]]:
    ratios: list[float] = []
    reasons: Counter[str] = Counter()
    for member in sorted(members):
        numerator = numerator_counts.get(member)
        denominator = denominator_counts.get(member)
        if numerator is None and denominator is None:
            reasons["BOTH_COMPONENT_KEYS_MISSING_NOT_ZERO"] += 1
            continue
        if numerator is None:
            reasons["NUMERATOR_KEY_MISSING_NOT_ZERO"] += 1
            continue
        if denominator is None:
            reasons["DENOMINATOR_KEY_MISSING_NOT_ZERO"] += 1
            continue
        numerator_cpm = numerator / numerator_library_sum * 1_000_000.0
        denominator_cpm = denominator / denominator_library_sum * 1_000_000.0
        if numerator_cpm < minimum_cpm and denominator_cpm < minimum_cpm:
            reasons["BOTH_COMPONENT_CPMS_BELOW_THRESHOLD"] += 1
            continue
        if numerator_cpm < minimum_cpm:
            reasons["NUMERATOR_CPM_BELOW_THRESHOLD"] += 1
            continue
        if denominator_cpm < minimum_cpm:
            reasons["DENOMINATOR_CPM_BELOW_THRESHOLD"] += 1
            continue
        ratios.append(math.log2(numerator_cpm) - math.log2(denominator_cpm))
    return ratios, reasons


def _benjamini_hochberg(pvalues: Sequence[float]) -> list[float]:
    if any(not math.isfinite(value) or value < 0 or value > 1 for value in pvalues):
        raise QualificationError("paper-native p-values must be finite probabilities")
    count = len(pvalues)
    order = sorted(range(count), key=lambda index: (pvalues[index], index))
    adjusted = [1.0] * count
    running = 1.0
    for reverse_rank, index in enumerate(reversed(order), start=1):
        rank = count - reverse_rank + 1
        candidate = min(1.0, pvalues[index] * count / rank)
        running = min(running, candidate)
        adjusted[index] = running
    return adjusted


def _build_paper_and_companion_results(
    *,
    pairs: Sequence[Mapping[str, Any]],
    counts: Mapping[tuple[str, str, int], Mapping[str, int]],
    library_sums: Mapping[tuple[str, str, int], int],
    protocol: Mapping[str, Any],
) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]], Counter[str]]:
    transform = protocol["paper_faithful_measurement_transform"]
    companion = protocol["route_a_companion_summary"]
    minimum_cpm = float(transform["original_cpm_minimum_inclusive"])
    minimum_members = int(companion["minimum_eligible_members_per_arm_per_endpoint_per_replicate"])
    replicate_rows: list[dict[str, Any]] = []
    effect_rows: list[dict[str, Any]] = []
    paper_rows: list[dict[str, Any]] = []
    rejections: Counter[str] = Counter()

    for pair in pairs:
        for context in CONTEXTS:
            for endpoint in ENDPOINTS:
                replicate_deltas: list[float] = []
                source_pooled: list[float] = []
                candidate_pooled: list[float] = []
                endpoint_reasons: Counter[str] = Counter()
                for replicate in REPLICATES:
                    numerator_slot = (context, endpoint["numerator_assay"], replicate)
                    denominator_slot = (context, endpoint["denominator_assay"], replicate)
                    source_ratios, source_reasons = _eligible_log2_ratios(
                        members=pair["source_members"],
                        numerator_counts=counts[numerator_slot],
                        denominator_counts=counts[denominator_slot],
                        numerator_library_sum=library_sums[numerator_slot],
                        denominator_library_sum=library_sums[denominator_slot],
                        minimum_cpm=minimum_cpm,
                    )
                    candidate_ratios, candidate_reasons = _eligible_log2_ratios(
                        members=pair["candidate_members"],
                        numerator_counts=counts[numerator_slot],
                        denominator_counts=counts[denominator_slot],
                        numerator_library_sum=library_sums[numerator_slot],
                        denominator_library_sum=library_sums[denominator_slot],
                        minimum_cpm=minimum_cpm,
                    )
                    source_pooled.extend(source_ratios)
                    candidate_pooled.extend(candidate_ratios)
                    endpoint_reasons.update(
                        {f"SOURCE_{key}": value for key, value in source_reasons.items()}
                    )
                    endpoint_reasons.update(
                        {f"CANDIDATE_{key}": value for key, value in candidate_reasons.items()}
                    )
                    eligible = (
                        len(source_ratios) >= minimum_members
                        and len(candidate_ratios) >= minimum_members
                    )
                    if eligible:
                        source_median = float(statistics.median(source_ratios))
                        candidate_median = float(statistics.median(candidate_ratios))
                        delta = candidate_median - source_median
                        replicate_deltas.append(delta)
                        rejection_reason = None
                    else:
                        source_median = None
                        candidate_median = None
                        delta = None
                        rejection_reason = "MINIMUM_TECHNICAL_MEMBER_COVERAGE_NOT_MET"
                        rejections[rejection_reason] += 1
                    replicate_rows.append(
                        {
                            "record_type": "ROUTE_A_COMPANION_REPLICATE_EFFECT",
                            "pair_id": pair["pair_id"],
                            "biological_source_group_id": pair["biological_source_group_id"],
                            "context": context,
                            "endpoint_id": endpoint["endpoint_id"],
                            "biological_replicate": replicate,
                            "source_technical_member_count_retained": len(source_ratios),
                            "candidate_technical_member_count_retained": len(candidate_ratios),
                            "source_ratio_median": source_median,
                            "candidate_ratio_median": candidate_median,
                            "candidate_minus_source_delta": delta,
                            "eligible": eligible,
                            "rejection_reason": rejection_reason,
                            "technical_members_are_independent_n": False,
                            "effective_n_unit": "BIOLOGICAL_REPLICATE",
                            "raw_member_identifier_emitted": False,
                        }
                    )

                if len(replicate_deltas) == 3:
                    effect = float(statistics.fmean(replicate_deltas))
                    standard_error = float(statistics.stdev(replicate_deltas) / math.sqrt(3.0))
                    effect_rejection = None
                else:
                    effect = None
                    standard_error = None
                    effect_rejection = "BIOLOGICAL_REPLICATE_COUNT_NOT_THREE"
                    rejections[effect_rejection] += 1

                if source_pooled and candidate_pooled:
                    try:
                        test_result = mannwhitneyu(
                            candidate_pooled,
                            source_pooled,
                            alternative="two-sided",
                            method="asymptotic",
                        )
                    except Exception as exc:
                        raise QualificationError("paper-native Mann-Whitney calculation failed") from exc
                    u_statistic = float(test_result.statistic)
                    p_value = float(test_result.pvalue)
                else:
                    u_statistic = None
                    p_value = None
                    rejections["PAPER_NATIVE_DISTRIBUTION_EMPTY"] += 1

                paper_rows.append(
                    {
                        "pair_id": pair["pair_id"],
                        "biological_source_group_id": pair["biological_source_group_id"],
                        "context": context,
                        "endpoint_id": endpoint["endpoint_id"],
                        "source_technical_ratio_count": len(source_pooled),
                        "candidate_technical_ratio_count": len(candidate_pooled),
                        "mann_whitney_u": u_statistic,
                        "two_sided_p_value": p_value,
                        "benjamini_hochberg_q_value": None,
                        "fdr_below_point_one": None,
                        "technical_members_are_independent_n": False,
                        "significance_used_for_membership": False,
                        "raw_member_identifier_emitted": False,
                    }
                )
                effect_rows.append(
                    {
                        "record_type": "ROUTE_A_COMPANION_AGGREGATE_EFFECT",
                        "pair_id": pair["pair_id"],
                        "biological_source_group_id": pair["biological_source_group_id"],
                        "context": context,
                        "endpoint_key": endpoint["endpoint_key"],
                        "endpoint_id": endpoint["endpoint_id"],
                        "endpoint_name": endpoint["endpoint_name"],
                        "candidate_minus_source_delta": effect,
                        "standard_error": standard_error,
                        "biological_replicate_count": len(replicate_deltas),
                        "biological_replicate_deltas": replicate_deltas,
                        "aggregate_effect_recomputable_from_replicate_deltas": True,
                        "standard_error_recomputable_from_replicate_deltas": True,
                        "inference_unit": "BIOLOGICAL_REPLICATE",
                        "paper_native_row": False,
                        "eligible": effect_rejection is None,
                        "rejection_reason": effect_rejection,
                        "technical_exclusion_reason_counts": dict(sorted(endpoint_reasons.items())),
                    }
                )

    family_counts: dict[str, int] = {}
    for context in CONTEXTS:
        for endpoint in ENDPOINTS:
            family = [
                row
                for row in paper_rows
                if row["context"] == context
                and row["endpoint_id"] == endpoint["endpoint_id"]
                and row["two_sided_p_value"] is not None
            ]
            q_values = _benjamini_hochberg(
                [float(row["two_sided_p_value"]) for row in family]
            )
            for row, q_value in zip(family, q_values):
                row["benjamini_hochberg_q_value"] = q_value
                row["fdr_below_point_one"] = q_value < 0.1
            family_counts[f"{context}:{endpoint['endpoint_id']}"] = len(family)

    statuses = {
        key: transform[key]
        for key in (
            "method_source_status",
            "multiple_testing_family_status",
            "published_result_crosscheck_status",
        )
    }
    paper_reproduced = (
        all(value == "PASS" for value in statuses.values())
        and bool(paper_rows)
        and all(row["benjamini_hochberg_q_value"] is not None for row in paper_rows)
    )
    paper_report = {
        "record_type": "GSE149487_PAPER_NATIVE_REPRODUCTION",
        "method": transform["paper_test"],
        "test_unit": transform["paper_test_unit"],
        "pooling": transform["paper_test_pooling"],
        "multiple_testing": transform["multiple_testing"],
        "multiple_testing_family": transform["multiple_testing_family"],
        "fdr_threshold_exclusive": transform["fdr_threshold_exclusive"],
        "method_evidence": statuses,
        "paper_method_reproduced": paper_reproduced,
        "computed_pair_context_endpoint_count": len(paper_rows),
        "family_counts": family_counts,
        "technical_members_are_independent_n": False,
        "significance_used_for_membership": False,
        "rows": sorted(
            paper_rows,
            key=lambda row: (row["pair_id"], row["context"], row["endpoint_id"]),
        ),
    }
    replicate_rows.sort(
        key=lambda row: (
            row["pair_id"],
            row["context"],
            row["endpoint_id"],
            row["biological_replicate"],
        )
    )
    effect_rows.sort(key=lambda row: (row["pair_id"], row["context"], row["endpoint_id"]))
    return paper_report, replicate_rows, effect_rows, rejections


def _fully_closed_pairs(
    pairs: Sequence[Mapping[str, Any]],
    effects: Sequence[Mapping[str, Any]],
) -> list[Mapping[str, Any]]:
    expected_slots = {
        (context, endpoint["endpoint_id"])
        for context in CONTEXTS
        for endpoint in ENDPOINTS
    }
    eligible: MutableMapping[str, set[tuple[str, str]]] = defaultdict(set)
    for effect in effects:
        if effect["eligible"]:
            eligible[str(effect["pair_id"])].add((str(effect["context"]), str(effect["endpoint_id"])))
    return [pair for pair in pairs if eligible.get(str(pair["pair_id"]), set()) == expected_slots]


def _cross_partition_duplicate_count(
    rows: Sequence[Mapping[str, Any]],
    value_key: str,
) -> int:
    partitions: MutableMapping[str, set[str]] = defaultdict(set)
    for row in rows:
        partitions[str(row[value_key])].add(str(row["partition"]))
    return sum(len(values) > 1 for values in partitions.values())


def _audit_group_leakage(pairs: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    rows = [
        {
            "pair_id": pair["pair_id"],
            "partition": "DEVELOPMENT",
            "source_group": pair["biological_source_group_id"],
            "candidate": pair["candidate_sequence_sha256"],
            "gene": pair["gene_group_id"],
            "sequence_cluster": pair["source_sequence_sha256"],
            "directed_edge": f"{pair['source_sequence_sha256']}->{pair['candidate_sequence_sha256']}",
            "reverse_edge": f"{pair['candidate_sequence_sha256']}->{pair['source_sequence_sha256']}",
        }
        for pair in pairs
    ]
    directed_counts = Counter(str(row["directed_edge"]) for row in rows)
    directed = set(directed_counts)
    reverse_duplicates = sum(
        1 for row in rows if str(row["reverse_edge"]) in directed
    ) // 2
    duplicate_directed = sum(count - 1 for count in directed_counts.values() if count > 1)
    counts = {
        "source_group_cross_partition": _cross_partition_duplicate_count(rows, "source_group"),
        "candidate_cross_partition": _cross_partition_duplicate_count(rows, "candidate"),
        "gene_cross_partition": _cross_partition_duplicate_count(rows, "gene"),
        "sequence_cluster_cross_partition": _cross_partition_duplicate_count(rows, "sequence_cluster"),
        "reverse_edge_duplicates": reverse_duplicates,
        "duplicate_directed_edges": duplicate_directed,
    }
    passed = all(value == 0 for value in counts.values())
    return {
        "record_type": "GSE149487_A1_GROUP_LEAKAGE_AUDIT",
        "split_id": "ROUTE_A_V3_A1_GSE149487_QUALIFICATION_DEVELOPMENT_ONLY_V1",
        "partition": "DEVELOPMENT",
        "qualification_pair_count": len(rows),
        "global_a2_benchmark_split_status": "NOT_RUN_DEFERRED_TO_A2",
        "global_benchmark_split_claimed": False,
        "counts": counts,
        "leakage_audit_status": "PASS" if passed else "FAIL_CURRENT_PROTOCOL",
    }


def _ranks(values: Sequence[float]) -> list[float]:
    order = sorted(range(len(values)), key=lambda index: (values[index], index))
    ranks = [0.0] * len(values)
    start = 0
    while start < len(order):
        end = start + 1
        while end < len(order) and values[order[end]] == values[order[start]]:
            end += 1
        average_rank = (start + 1 + end) / 2.0
        for position in range(start, end):
            ranks[order[position]] = average_rank
        start = end
    return ranks


def _pearson(left: Sequence[float], right: Sequence[float]) -> float:
    if len(left) != len(right) or len(left) < 2:
        return 0.0
    left_mean = statistics.fmean(left)
    right_mean = statistics.fmean(right)
    numerator = sum((x - left_mean) * (y - right_mean) for x, y in zip(left, right))
    left_ss = sum((x - left_mean) ** 2 for x in left)
    right_ss = sum((y - right_mean) ** 2 for y in right)
    denominator = math.sqrt(left_ss * right_ss)
    return numerator / denominator if denominator > 0 else 0.0


def _spearman(left: Sequence[float], right: Sequence[float]) -> float:
    return _pearson(_ranks(left), _ranks(right))


def _simulate_correlated_sample(
    rng: random.Random,
    sample_size: int,
    rho: float,
) -> tuple[list[float], list[float]]:
    residual_scale = math.sqrt(max(0.0, 1.0 - rho * rho))
    left: list[float] = []
    right: list[float] = []
    for _ in range(sample_size):
        x = rng.gauss(0.0, 1.0)
        y = rho * x + residual_scale * rng.gauss(0.0, 1.0)
        left.append(x)
        right.append(y)
    return left, right


def _run_group_power_simulation(group_count: int, protocol: Mapping[str, Any]) -> dict[str, Any]:
    power = protocol["power_prefreeze"]
    if group_count < 4:
        return {
            "record_type": "GSE149487_A1_GROUP_POWER_AUDIT",
            "status": "FAIL_CURRENT_PROTOCOL",
            "analysis_unit": power["analysis_unit"],
            "post_dedup_group_count": group_count,
            "estimated_power": 0.0,
            "bootstrap_ci_full_width": None,
            "simulation_trials": power["simulation_trials"],
            "bootstrap_resamples": power["bootstrap_resamples"],
            "seed": power["simulation_seed"],
            "reason": "FEWER_THAN_FOUR_BIOLOGICAL_SOURCE_GROUPS",
        }
    rho = float(power["minimum_effect_at_alternative"])
    alpha = float(power["alpha_two_sided"])
    trials = int(power["simulation_trials"])
    bootstrap_resamples = int(power["bootstrap_resamples"])
    seed = int(power["simulation_seed"])
    critical = NormalDist().inv_cdf(1.0 - alpha / 2.0)
    rng = random.Random(seed)
    rejected = 0
    for _ in range(trials):
        left, right = _simulate_correlated_sample(rng, group_count, rho)
        observed = max(-0.999999, min(0.999999, _spearman(left, right)))
        statistic = abs(math.atanh(observed) * math.sqrt(group_count - 3))
        if statistic > critical:
            rejected += 1
    estimated_power = rejected / trials

    baseline_rng = random.Random(seed ^ 0xA15EED)
    baseline_left, baseline_right = _simulate_correlated_sample(
        baseline_rng, group_count, rho
    )
    bootstrap_rng = random.Random(seed ^ 0xB0057)
    bootstrapped: list[float] = []
    for _ in range(bootstrap_resamples):
        indices = [bootstrap_rng.randrange(group_count) for _ in range(group_count)]
        bootstrapped.append(
            _spearman(
                [baseline_left[index] for index in indices],
                [baseline_right[index] for index in indices],
            )
        )
    bootstrapped.sort()
    tail = (1.0 - float(power["confidence_level"])) / 2.0
    lower_index = max(0, min(bootstrap_resamples - 1, int(math.floor(tail * bootstrap_resamples))))
    upper_index = max(
        0,
        min(
            bootstrap_resamples - 1,
            int(math.ceil((1.0 - tail) * bootstrap_resamples)) - 1,
        ),
    )
    ci_width = float(bootstrapped[upper_index] - bootstrapped[lower_index])
    passed = (
        estimated_power >= float(power["target_power"])
        and ci_width <= float(power["maximum_ci_full_width"])
    )
    return {
        "record_type": "GSE149487_A1_GROUP_POWER_AUDIT",
        "status": "PASS" if passed else "FAIL_CURRENT_PROTOCOL",
        "analysis_unit": power["analysis_unit"],
        "bootstrap_unit": power["bootstrap_unit"],
        "target_metric": power["target_metric"],
        "post_dedup_group_count": group_count,
        "alternative_effect": rho,
        "alpha_two_sided": alpha,
        "estimated_power": estimated_power,
        "target_power": power["target_power"],
        "bootstrap_ci_full_width": ci_width,
        "maximum_ci_full_width": power["maximum_ci_full_width"],
        "simulation_trials": trials,
        "bootstrap_resamples": bootstrap_resamples,
        "seed": seed,
        "reason": None if passed else "PREFROZEN_POWER_OR_CI_THRESHOLD_NOT_MET",
    }


def _parse_iso_datetime(value: Any, *, label: str) -> str:
    if not isinstance(value, str) or not value or value == "UNKNOWN_NOT_ASSERTED":
        raise QualificationError(f"{label} must be an asserted ISO-8601 timestamp")
    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError as exc:
        raise QualificationError(f"{label} must be an ISO-8601 timestamp") from exc
    if parsed.tzinfo is None or parsed.utcoffset() is None:
        raise QualificationError(f"{label} must include an explicit UTC offset")
    return value


def _audit_license_and_exposure(
    resolved_assets: Sequence[Mapping[str, Any]],
    protocol: Mapping[str, Any],
) -> dict[str, Any]:
    license_protocol = protocol["license_and_redistribution"]
    exposure_protocol = protocol["foundation_exposure"]
    allowed = set(license_protocol["qualification_allowed_statuses"])
    asset_status_counts = Counter(str(asset["license"]["status"]) for asset in resolved_assets)
    unknown_assets = sum(status not in allowed for status in asset_status_counts.elements())
    all_asset_licenses_closed = unknown_assets == 0
    license_audit_pass = (
        license_protocol["audit_status"] == "PASS"
        and all_asset_licenses_closed
        and license_protocol["license_id"] != "UNKNOWN_NOT_ASSERTED"
        and license_protocol["verified_at"] != "UNKNOWN_NOT_ASSERTED"
    )
    if license_audit_pass:
        _parse_iso_datetime(
            license_protocol["verified_at"], label="license verification timestamp"
        )
    verified_statuses = {
        str(asset["license"]["status"])
        for asset in resolved_assets
        if str(asset["license"]["status"]) in allowed
    }
    if "VERIFIED_NONREDISTRIBUTABLE" in verified_statuses:
        canonical_license_status = "VERIFIED_NONREDISTRIBUTABLE"
        redistribution_allowed = False
    elif verified_statuses == {"VERIFIED_PUBLIC"} and all_asset_licenses_closed:
        canonical_license_status = "VERIFIED_PUBLIC"
        redistribution_allowed = True
    else:
        canonical_license_status = "UNKNOWN_BLOCKED"
        redistribution_allowed = False

    checkpoint_hash = exposure_protocol["checkpoint_sha256"]
    exposure_audit_pass = (
        exposure_protocol["audit_status"] == "PASS"
        and exposure_protocol["checkpoint_id"] != "UNKNOWN_NOT_ASSERTED"
        and isinstance(checkpoint_hash, str)
        and SHA256_RE.fullmatch(checkpoint_hash) is not None
        and exposure_protocol["audit_id"] != "UNKNOWN_NOT_ASSERTED"
    )
    return {
        "record_type": "GSE149487_LICENSE_AND_EXPOSURE_AUDIT",
        "asset_license_status_counts": dict(sorted(asset_status_counts.items())),
        "all_21_asset_licenses_closed": all_asset_licenses_closed,
        "license_audit_status": "PASS" if license_audit_pass else "UNKNOWN_NOT_ASSERTED",
        "canonical_license_status": canonical_license_status,
        "redistribution_allowed": redistribution_allowed,
        "canonical_sequence_materialization_allowed": (
            license_audit_pass
            and canonical_license_status == "VERIFIED_PUBLIC"
            and redistribution_allowed is True
        ),
        "nonredistributable_release_mode": license_protocol["nonredistributable_release_mode"],
        "foundation_exposure_audit_status": "PASS" if exposure_audit_pass else "UNKNOWN_NOT_ASSERTED",
        "foundation_exposure_stratum": exposure_protocol["stratum"],
        "sequence_exposed": bool(exposure_protocol["sequence_exposed"]),
        "label_exposed": bool(exposure_protocol["label_exposed"]),
        "checkpoint_bound": exposure_audit_pass,
        "license_and_exposure_gate_status": (
            "PASS" if license_audit_pass and exposure_audit_pass else "BLOCKED_PENDING_PUBLIC_EVIDENCE"
        ),
    }


def _verify_git_binding(
    repo_root: Path,
    implementation_commit: str,
    *,
    accepted_a0_base_commit: str,
    active_authority_commit: str,
    authority_files: Sequence[tuple[str, str]],
    implementation_files: Sequence[tuple[str, str]],
) -> dict[str, Any]:
    if implementation_commit == "UNKNOWN_NOT_ASSERTED":
        return {
            "status": "UNKNOWN_NOT_ASSERTED",
            "accepted_a0_base_commit": accepted_a0_base_commit,
            "active_authority_commit": active_authority_commit,
            "implementation_commit": implementation_commit,
            "observed_head": None,
            "accepted_a0_is_ancestor_of_active_authority": None,
            "active_authority_is_ancestor_of_implementation": None,
            "implementation_commit_is_ancestor_of_head": None,
            "active_authority_file_hashes_match": None,
            "implementation_file_hashes_match": None,
            "worktree_clean": None,
        }

    def require_ancestor(older: str, newer: str, *, label: str) -> None:
        result = subprocess.run(
            ["git", "-C", str(repo_root), "merge-base", "--is-ancestor", older, newer],
            check=False,
            capture_output=True,
            text=True,
            timeout=15,
        )
        if result.returncode != 0:
            raise QualificationError(f"{label} ancestry is not satisfied")

    def require_blob_hashes(
        commit: str,
        files: Sequence[tuple[str, str]],
        *,
        label: str,
    ) -> None:
        for relative_path, expected_sha256 in files:
            blob = subprocess.run(
                ["git", "-C", str(repo_root), "show", f"{commit}:{relative_path}"],
                check=True,
                capture_output=True,
                timeout=30,
            ).stdout
            if _sha256_bytes(blob) != _require_sha256(
                expected_sha256,
                label=f"{label} blob expected hash",
            ):
                raise QualificationError(f"{label} file hash differs at its bound commit")

    try:
        observed_head = subprocess.run(
            ["git", "-C", str(repo_root), "rev-parse", "HEAD"],
            check=True,
            capture_output=True,
            text=True,
            timeout=15,
        ).stdout.strip()
        require_ancestor(
            accepted_a0_base_commit,
            active_authority_commit,
            label="accepted A0 to active authority",
        )
        require_ancestor(
            active_authority_commit,
            implementation_commit,
            label="active authority to implementation",
        )
        require_ancestor(
            implementation_commit,
            observed_head,
            label="implementation to executing HEAD",
        )
        require_blob_hashes(
            active_authority_commit,
            authority_files,
            label="active authority",
        )
        require_blob_hashes(
            implementation_commit,
            implementation_files,
            label="implementation",
        )
        porcelain = subprocess.run(
            ["git", "-C", str(repo_root), "status", "--porcelain", "--untracked-files=all"],
            check=True,
            capture_output=True,
            text=True,
            timeout=15,
        ).stdout
    except (OSError, subprocess.SubprocessError) as exc:
        raise QualificationError("repository binding could not be verified") from exc
    if porcelain:
        raise QualificationError("repository worktree is not clean during qualification")
    return {
        "status": "PASS",
        "accepted_a0_base_commit": accepted_a0_base_commit,
        "active_authority_commit": active_authority_commit,
        "implementation_commit": implementation_commit,
        "observed_head": observed_head,
        "accepted_a0_is_ancestor_of_active_authority": True,
        "active_authority_is_ancestor_of_implementation": True,
        "implementation_commit_is_ancestor_of_head": True,
        "active_authority_file_hashes_match": True,
        "implementation_file_hashes_match": True,
        "worktree_clean": True,
    }


def _validate_canonical_record(record: Mapping[str, Any], schema: Mapping[str, Any]) -> None:
    required = schema.get("required")
    properties = schema.get("properties")
    if not isinstance(required, list) or not isinstance(properties, Mapping):
        raise QualificationError("bound canonical schema has an invalid top-level contract")
    missing = set(required) - set(record)
    unexpected = set(record) - set(properties)
    if missing or unexpected:
        raise QualificationError(
            f"canonical record top-level keys invalid; missing={sorted(missing)}, unexpected={sorted(unexpected)}"
        )
    if record.get("contract_id") != CONTRACT_ID or record.get("schema_version") != SCHEMA_VERSION:
        raise QualificationError("canonical record authority fields are invalid")
    if record.get("evidence_status") != "PASS" or record.get("claim_status") != "NOT_ESTABLISHED":
        raise QualificationError("canonical record evidence/claim boundary is invalid")
    for key in ("record_id", "biological_source_group_id"):
        if not isinstance(record.get(key), str) or not record[key]:
            raise QualificationError(f"canonical record {key} is invalid")
    for sequence_key, hash_key in (
        ("source_sequence", "source_sequence_sha256"),
        ("candidate_sequence", "candidate_sequence_sha256"),
    ):
        sequence = record.get(sequence_key)
        if not isinstance(sequence, str) or not SEQUENCE_RE.fullmatch(sequence):
            raise QualificationError(f"canonical record {sequence_key} is invalid")
        if record.get(hash_key) != _sha256_bytes(sequence.encode("ascii")):
            raise QualificationError(f"canonical record {hash_key} does not bind the full sequence")
    if record.get("sequence_alphabet") != "DNA" or record.get("region") != "5UTR":
        raise QualificationError("canonical record sequence/region semantics are invalid")
    if record.get("edit_count") != 1 or not isinstance(record.get("edit_set"), list) or len(record["edit_set"]) != 1:
        raise QualificationError("canonical record must contain exactly one edit")
    edit = _require_exact_keys(
        record["edit_set"][0],
        {
            "edit_id",
            "position",
            "coordinate_system",
            "ref_base",
            "alt_base",
            "region",
            "distance_from_region_start",
        },
        label="canonical edit",
    )
    if (
        not isinstance(edit["position"], int)
        or isinstance(edit["position"], bool)
        or edit["position"] < 0
        or edit["coordinate_system"] != "ZERO_BASED_SOURCE"
        or edit["region"] != "5UTR"
        or edit["ref_base"] == edit["alt_base"]
        or not re.fullmatch(r"[ACGT]", str(edit["ref_base"]))
        or not re.fullmatch(r"[ACGT]", str(edit["alt_base"]))
    ):
        raise QualificationError("canonical edit semantics are invalid")

    nested_contracts: tuple[tuple[str, set[str], set[str]], ...] = (
        ("source", {"source_id", "sequence_id", "transcript_id", "gene_id", "locus_id", "design_family_id"}, {"source_id", "sequence_id"}),
        ("candidate", {"candidate_id", "sequence_id", "design_id"}, {"candidate_id", "sequence_id"}),
        ("study", {"study_id", "accession", "independent_study_group_id", "publication_doi"}, {"study_id", "accession", "independent_study_group_id"}),
        ("assay", {"assay_id", "assay_type", "protocol_version"}, {"assay_id", "assay_type"}),
        ("context", {"context_id", "observable_context", "cell_type", "condition"}, {"context_id", "observable_context"}),
        ("endpoint", {"endpoint_id", "endpoint_name", "beneficial_direction"}, {"endpoint_id", "endpoint_name", "beneficial_direction"}),
        ("raw_measurement", {"value", "unit", "scale", "source_column", "detection_limit"}, {"value", "unit", "scale", "source_column"}),
        ("paper_faithful_transform", {"transform_id", "description", "version", "direction_verified", "implementation_sha256"}, {"transform_id", "description", "version", "direction_verified", "implementation_sha256"}),
        ("replicate", {"replicate_id", "replicate_group_id", "replicate_count", "aggregation_rule"}, {"replicate_id", "replicate_group_id", "replicate_count"}),
        ("exposure", {"stratum", "label_exposed", "sequence_exposed", "audit_id"}, {"stratum", "label_exposed", "sequence_exposed", "audit_id"}),
        ("split", {"split_id", "partition", "leakage_audit_status"}, {"split_id", "partition", "leakage_audit_status"}),
        ("provenance", {"dataset_id", "asset_id", "source_uri", "source_file_sha256", "raw_record_locator", "acquired_at", "lineage"}, {"dataset_id", "asset_id", "source_uri", "source_file_sha256", "raw_record_locator", "lineage"}),
        ("license", {"license_id", "license_name", "license_uri", "status", "redistribution_allowed", "verified_at"}, {"license_id", "status", "redistribution_allowed", "verified_at"}),
        ("eligibility", {"status", "reject_reason_code", "reject_reason_detail"}, {"status", "reject_reason_code", "reject_reason_detail"}),
    )
    for key, allowed, required_nested in nested_contracts:
        value = record.get(key)
        if not isinstance(value, Mapping):
            raise QualificationError(f"canonical record {key} must be an object")
        if set(value) - allowed or required_nested - set(value):
            raise QualificationError(f"canonical record {key} has invalid closed keys")
    if record["data_role"] != "ORDINARY_A1_INTERVENTION":
        raise QualificationError("canonical data role is not A1 intervention")
    if record["split"] != {
        "split_id": "ROUTE_A_V3_A1_GSE149487_QUALIFICATION_DEVELOPMENT_ONLY_V1",
        "partition": "DEVELOPMENT",
        "leakage_audit_status": "PASS",
    }:
        raise QualificationError("canonical split boundary is invalid")
    if record["eligibility"] != {
        "status": "QUALIFIED",
        "reject_reason_code": None,
        "reject_reason_detail": None,
    }:
        raise QualificationError("canonical eligibility boundary is invalid")
    if record["replicate"]["replicate_count"] != 3:
        raise QualificationError("canonical replicate count must be three")
    if not isinstance(record["standard_error"], (int, float)) or record["standard_error"] < 0 or not math.isfinite(record["standard_error"]):
        raise QualificationError("canonical standard error is invalid")
    if not isinstance(record["delta"], (int, float)) or not math.isfinite(record["delta"]):
        raise QualificationError("canonical delta is invalid")
    if not SHA256_RE.fullmatch(str(record["paper_faithful_transform"]["implementation_sha256"])):
        raise QualificationError("canonical transform implementation hash is invalid")
    if not SHA256_RE.fullmatch(str(record["provenance"]["source_file_sha256"])):
        raise QualificationError("canonical provenance source hash is invalid")
    lineage = record["provenance"]["lineage"]
    if not isinstance(lineage, list) or not lineage:
        raise QualificationError("canonical provenance lineage is empty")
    for step in lineage:
        _require_exact_keys(
            step,
            {"step_id", "operation", "input_sha256", "output_sha256"},
            label="canonical lineage step",
        )
        _require_sha256(step["input_sha256"], label="canonical lineage input hash")
        _require_sha256(step["output_sha256"], label="canonical lineage output hash")
    if record["license"]["status"] != "VERIFIED_PUBLIC":
        raise QualificationError("canonical full-sequence license is not verified public")
    _parse_iso_datetime(record["license"]["verified_at"], label="canonical license timestamp")


def _build_canonical_records(
    *,
    closed_pairs: Sequence[Mapping[str, Any]],
    effects: Sequence[Mapping[str, Any]],
    effective_manifest_sha256: str,
    raw_bundle_sha256: str,
    protocol_sha256: str,
    script_sha256: str,
    protocol: Mapping[str, Any],
    schema: Mapping[str, Any],
    license_exposure: Mapping[str, Any],
) -> list[dict[str, Any]]:
    pair_by_id = {str(pair["pair_id"]): pair for pair in closed_pairs}
    license_protocol = protocol["license_and_redistribution"]
    exposure_protocol = protocol["foundation_exposure"]
    if (
        license_exposure.get("canonical_license_status") != "VERIFIED_PUBLIC"
        or license_exposure.get("redistribution_allowed") is not True
    ):
        raise QualificationError(
            "canonical full-sequence materialization requires verified public redistribution"
        )
    records: list[dict[str, Any]] = []
    for effect in effects:
        pair = pair_by_id.get(str(effect["pair_id"]))
        if pair is None or not effect["eligible"]:
            continue
        endpoint = next(item for item in ENDPOINTS if item["endpoint_id"] == effect["endpoint_id"])
        effect_digest = _sha256_bytes(_compact_json_bytes(effect))
        record_id = _stable_id(
            "GSE149487_CIRV3_",
            {
                "pair_id": pair["pair_id"],
                "context": effect["context"],
                "endpoint": effect["endpoint_id"],
                "effect_sha256": effect_digest,
            },
        )
        edit_id = _stable_id(
            "GSE149487_EDIT_",
            {
                "pair_id": pair["pair_id"],
                "position": pair["sequence_index_0_based"],
                "ref": pair["ref"],
                "alt": pair["alt"],
            },
        )
        record = {
            "contract_id": CONTRACT_ID,
            "schema_version": SCHEMA_VERSION,
            "record_id": record_id,
            "evidence_status": "PASS",
            "claim_status": "NOT_ESTABLISHED",
            "source": {
                "source_id": pair["source_id"],
                "sequence_id": pair["source_id"],
                "transcript_id": None,
                "gene_id": pair["gene_group_id"],
                "locus_id": pair["locus_id"],
                "design_family_id": "PLUMAGE_FULL_LENGTH_5UTR_SNV",
            },
            "candidate": {
                "candidate_id": pair["candidate_id"],
                "sequence_id": pair["candidate_id"],
                "design_id": pair["pair_id"],
            },
            "source_sequence": pair["source_sequence"],
            "candidate_sequence": pair["candidate_sequence"],
            "source_sequence_sha256": pair["source_sequence_sha256"],
            "candidate_sequence_sha256": pair["candidate_sequence_sha256"],
            "sequence_alphabet": "DNA",
            "edit_set": [
                {
                    "edit_id": edit_id,
                    "position": pair["sequence_index_0_based"],
                    "coordinate_system": "ZERO_BASED_SOURCE",
                    "ref_base": pair["ref"],
                    "alt_base": pair["alt"],
                    "region": "5UTR",
                    "distance_from_region_start": pair["sequence_index_0_based"],
                }
            ],
            "edit_count": 1,
            "region": "5UTR",
            "study": {
                "study_id": STUDY_GROUP_ID,
                "accession": DATASET_ID,
                "independent_study_group_id": STUDY_GROUP_ID,
                "publication_doi": None,
            },
            "assay": {
                "assay_id": "PLUMAGE_BARCODED_REPORTER",
                "assay_type": "BARCODED_FULL_LENGTH_5UTR_REPORTER",
                "protocol_version": "LIM_2021_PUBLIC",
            },
            "context": {
                "context_id": f"PLUMAGE_{effect['context']}",
                "observable_context": "CELL_LINE_REPORTER_ASSAY",
                "cell_type": effect["context"],
                "condition": None,
            },
            "endpoint": {
                "endpoint_id": endpoint["endpoint_id"],
                "endpoint_name": endpoint["endpoint_name"],
                "beneficial_direction": "HIGHER_IS_BETTER",
            },
            "raw_measurement": {
                "value": effect["candidate_minus_source_delta"],
                "unit": "LOG2_RATIO_DELTA",
                "scale": "LOG2",
                "source_column": "RAW_GEO_COUNTS_PAPER_ENDPOINT_DERIVED",
                "detection_limit": 0.5,
            },
            "paper_faithful_transform": {
                "transform_id": "GSE149487_RAW_COUNTS_TO_PAPER_ENDPOINT_V1",
                "description": "CPM floor, log2 assay ratio, and candidate-minus-source contrast; paper-native test retained separately",
                "version": "1.0.0",
                "direction_verified": True,
                "implementation_sha256": script_sha256,
            },
            "delta": effect["candidate_minus_source_delta"],
            "replicate": {
                "replicate_id": "BIOLOGICAL_REPLICATES_1_2_3_AGGREGATED",
                "replicate_group_id": f"{pair['pair_id']}:{effect['context']}:{effect['endpoint_id']}",
                "replicate_count": 3,
                "aggregation_rule": "EQUAL_WEIGHT_MEAN_WITH_SAMPLE_SD_DIV_SQRT_3_SE",
            },
            "standard_error": effect["standard_error"],
            "biological_source_group_id": pair["biological_source_group_id"],
            "gene_group_id": pair["gene_group_id"],
            "data_role": "ORDINARY_A1_INTERVENTION",
            "exposure": {
                "stratum": exposure_protocol["stratum"],
                "label_exposed": exposure_protocol["label_exposed"],
                "sequence_exposed": exposure_protocol["sequence_exposed"],
                "audit_id": exposure_protocol["audit_id"],
            },
            "split": {
                "split_id": "ROUTE_A_V3_A1_GSE149487_QUALIFICATION_DEVELOPMENT_ONLY_V1",
                "partition": "DEVELOPMENT",
                "leakage_audit_status": "PASS",
            },
            "provenance": {
                "dataset_id": DATASET_ID,
                "asset_id": "GSE149487_21_ASSET_EFFECTIVE_BUNDLE",
                "source_uri": "https://www.ncbi.nlm.nih.gov/geo/query/acc.cgi?acc=GSE149487",
                "source_file_sha256": effective_manifest_sha256,
                "raw_record_locator": "sha256:" + _sha256_bytes(
                    f"{effective_manifest_sha256}\0{pair['pair_id']}\0{effect['context']}\0{effect['endpoint_id']}".encode("utf-8")
                ),
                "acquired_at": None,
                "lineage": [
                    {
                        "step_id": "RESOLVE_21_ASSET_TRUST_ROOT",
                        "operation": "VERIFY_AND_RESOLVE_HASH_BOUND_ADDITIVE_MANIFEST",
                        "input_sha256": protocol_sha256,
                        "output_sha256": effective_manifest_sha256,
                    },
                    {
                        "step_id": "RAW_COUNTS_TO_REPLICATE_EFFECT",
                        "operation": "CPM_FILTER_LOG2_RATIO_MEDIAN_AND_THREE_REPLICATE_DELTA",
                        "input_sha256": raw_bundle_sha256,
                        "output_sha256": effect_digest,
                    },
                ],
            },
            "license": {
                "license_id": license_protocol["license_id"],
                "license_name": license_protocol["license_name"],
                "license_uri": license_protocol["license_uri"],
                "status": license_exposure["canonical_license_status"],
                "redistribution_allowed": license_exposure["redistribution_allowed"],
                "verified_at": license_protocol["verified_at"],
            },
            "eligibility": {
                "status": "QUALIFIED",
                "reject_reason_code": None,
                "reject_reason_detail": None,
            },
        }
        _validate_canonical_record(record, schema)
        records.append(record)
    records.sort(key=lambda row: row["record_id"])
    if not records:
        raise QualificationError("all-gates-pass state produced zero canonical records")
    return records


def _validate_run_metadata(
    *,
    run_id: str,
    execution_id: str,
    recorded_at: str,
) -> dict[str, str]:
    run_id = _require_identifier(run_id, label="run ID")
    execution_id = _require_identifier(execution_id, label="execution ID")
    recorded_at = _parse_iso_datetime(recorded_at, label="recorded_at")
    return {"run_id": run_id, "execution_id": execution_id, "recorded_at": recorded_at}


def _verify_trust_roots(
    *,
    repo_root: Path,
    protocol: Mapping[str, Any],
    explicit_asset_manifest_path: Path,
    snapshot_root: Path,
) -> tuple[
    dict[str, dict[str, Any]],
    dict[str, Path],
    dict[str, dict[str, Any]],
    dict[str, Any],
]:
    authority = protocol["authority"]
    path_fields = {
        "contract": "contract_path",
        "a1_qualification": "a1_qualification_path",
        "data_role_registry": "data_role_registry_path",
        "canonical_schema": "canonical_schema_path",
        "asset_manifest": "asset_manifest_path",
        "v4_helper": "v4_helper_path",
        "qualifier": "qualifier_path",
        "focused_test": "focused_test_path",
    }
    hash_fields = {
        "contract": "contract_sha256",
        "a1_qualification": "a1_qualification_sha256",
        "data_role_registry": "data_role_registry_sha256",
        "canonical_schema": "canonical_schema_sha256",
        "asset_manifest": "asset_manifest_sha256",
        "v4_helper": "v4_helper_sha256",
        "qualifier": "qualifier_sha256",
        "focused_test": "focused_test_sha256",
    }
    paths = {
        label: _safe_relative_path(repo_root, authority[field], label=f"authority {label} path")
        for label, field in path_fields.items()
    }
    if paths["asset_manifest"] != explicit_asset_manifest_path:
        raise QualificationError("explicit asset manifest path differs from the protocol trust root")
    provenance: dict[str, dict[str, Any]] = {}
    verified_documents: dict[str, dict[str, Any]] = {}
    parser_paths: dict[str, Path] = {}
    for label, path in paths.items():
        if label in {"asset_manifest", "canonical_schema"}:
            payload, provenance[label] = _read_verified_bytes(
                path,
                authority[hash_fields[label]],
                label=f"authority {label}",
            )
            verified_documents[label] = _read_json_bytes(
                payload,
                label=f"authority {label}",
            )
        elif label == "v4_helper":
            parser_path = snapshot_root / "v4_helper.py"
            provenance[label] = _snapshot_verified_file(
                path,
                parser_path,
                authority[hash_fields[label]],
                label="authority v4_helper",
                suffix=".py",
            )
            parser_paths[label] = parser_path
        else:
            provenance[label] = _verify_file_hash(
                path,
                authority[hash_fields[label]],
                label=f"authority {label}",
            )
    running_script = _absolute_without_resolving(Path(__file__))
    _reject_forbidden_path(running_script, label="running qualifier source")
    _require_regular_file(running_script, label="running qualifier source", suffix=".py")
    running_hash, _ = _sha256_file(running_script)
    if running_hash != authority["qualifier_sha256"]:
        raise QualificationError("executing qualifier source differs from the bound qualifier")
    git_binding = _verify_git_binding(
        repo_root,
        authority["implementation_commit"],
        accepted_a0_base_commit=authority["accepted_a0_base_commit"],
        active_authority_commit=authority["active_authority_commit"],
        authority_files=(
            (authority["contract_path"], authority["contract_sha256"]),
            (
                authority["data_role_registry_path"],
                authority["data_role_registry_sha256"],
            ),
        ),
        implementation_files=(
            (authority["qualifier_path"], authority["qualifier_sha256"]),
            (authority["focused_test_path"], authority["focused_test_sha256"]),
        ),
    )
    return provenance, parser_paths, verified_documents, git_binding


def _qualification_blockers(
    *,
    protocol: Mapping[str, Any],
    git_binding: Mapping[str, Any],
    sequence_class_counts: Mapping[str, int],
    mapping_reasons: Mapping[str, int],
    pair_count: int,
    closed_pair_count: int,
    paper_report: Mapping[str, Any],
    leakage_report: Mapping[str, Any],
    power_report: Mapping[str, Any],
    license_exposure: Mapping[str, Any],
    raw_key_reconciliation: Mapping[str, Any],
) -> list[str]:
    blockers: set[str] = set()
    if git_binding["status"] != "PASS":
        blockers.add("IMPLEMENTATION_COMMIT_UNKNOWN_NOT_ASSERTED")
    if protocol["mapping"]["outcome_blind_mapping_evidence_status"] != "PASS":
        blockers.add("OUTCOME_BLIND_LONG_READ_MAPPING_PROVENANCE_UNKNOWN_NOT_ASSERTED")
    if any(key.startswith("REJECTED_") and value for key, value in sequence_class_counts.items()):
        blockers.add("UNADJUDICATED_SEQUENCE_UNIVERSE_CLASSES_PRESENT")
    if any(int(value) > 0 for value in mapping_reasons.values()):
        blockers.add("UNADJUDICATED_OR_AMBIGUOUS_MAPPING_ROWS_PRESENT")
    if pair_count == 0:
        blockers.add("NO_STRICT_OUTCOME_BLIND_SOURCE_CANDIDATE_PAIRS")
    if closed_pair_count == 0:
        blockers.add("NO_PAIR_CLOSED_ACROSS_BOTH_CONTEXTS_AND_BOTH_ENDPOINTS")
    paper_evidence = paper_report["method_evidence"]
    if paper_evidence["method_source_status"] != "PASS":
        blockers.add("PAPER_NATIVE_METHOD_SOURCE_UNKNOWN_NOT_ASSERTED")
    if paper_evidence["multiple_testing_family_status"] != "PASS":
        blockers.add("PAPER_NATIVE_MULTIPLE_TESTING_FAMILY_UNKNOWN_NOT_ASSERTED")
    if paper_evidence["published_result_crosscheck_status"] != "PASS":
        blockers.add("PUBLISHED_RESULT_CROSSCHECK_UNKNOWN_NOT_ASSERTED")
    if not paper_report["paper_method_reproduced"]:
        blockers.add("PAPER_NATIVE_METHOD_NOT_REPRODUCED")
    if leakage_report["leakage_audit_status"] != "PASS":
        blockers.add("GROUP_OR_SEQUENCE_CLUSTER_LEAKAGE_NOT_ZERO")
    if power_report["status"] != "PASS":
        blockers.add("PREFROZEN_GROUP_POWER_OR_CI_GATE_FAILED")
    if license_exposure["license_audit_status"] != "PASS":
        blockers.add("LICENSE_AND_REDISTRIBUTION_UNKNOWN_NOT_ASSERTED")
    elif not license_exposure["canonical_sequence_materialization_allowed"]:
        blockers.add("CANONICAL_SEQUENCE_REDISTRIBUTION_NOT_ALLOWED")
    if license_exposure["foundation_exposure_audit_status"] != "PASS":
        blockers.add("CHECKPOINT_SPECIFIC_FOUNDATION_EXPOSURE_UNKNOWN_NOT_ASSERTED")
    if int(raw_key_reconciliation["unclassified_raw_key_count"]) != 0:
        blockers.add("RAW_KEY_UNCLASSIFIED_OUTCOME_BLIND_RECONCILIATION_NOT_ZERO")
    if int(raw_key_reconciliation["mapped_and_explicit_exclusion_overlap_count"]) != 0:
        blockers.add("RAW_KEY_CLASSIFICATION_NOT_MUTUALLY_EXCLUSIVE")
    if int(raw_key_reconciliation["mapping_key_absent_from_both_contexts_count"]) != 0:
        blockers.add("MAPPING_KEYS_ABSENT_FROM_BOTH_CONTEXTS_NOT_ZERO")
    return sorted(blockers)


def _raw_bundle_sha256(resolved_assets: Sequence[Mapping[str, Any]]) -> str:
    return _sha256_bytes(
        _compact_json_bytes(
            [
                {"asset_id": asset["asset_id"], "sha256": asset["sha256"], "bytes": asset["bytes"]}
                for asset in sorted(resolved_assets, key=lambda row: row["asset_id"])
                if asset["asset_kind"] == "GEO_RAW_COUNT"
            ]
        )
    )


def _safe_rejection_rows(
    *,
    sequence_class_counts: Mapping[str, int],
    mapping_reasons: Mapping[str, int],
    data_rejections: Mapping[str, int],
    pair_count: int,
    closed_pair_count: int,
    blockers: Sequence[str],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for namespace, values in (
        ("SEQUENCE_CLASSIFICATION", sequence_class_counts),
        ("MAPPING", mapping_reasons),
        ("ENDPOINT_OR_REPLICATE", data_rejections),
    ):
        for reason, count in sorted(values.items()):
            if int(count) > 0:
                rows.append(
                    {
                        "audit_type": namespace,
                        "reason": str(reason),
                        "count": int(count),
                        "raw_member_identifier_emitted": False,
                        "raw_sequence_emitted": False,
                    }
                )
    incomplete = pair_count - closed_pair_count
    if incomplete > 0:
        rows.append(
            {
                "audit_type": "PAIR_CLOSURE",
                "reason": "PAIR_NOT_CLOSED_ACROSS_BOTH_CONTEXTS_AND_BOTH_ENDPOINTS",
                "count": incomplete,
                "raw_member_identifier_emitted": False,
                "raw_sequence_emitted": False,
            }
        )
    for blocker in blockers:
        rows.append(
            {
                "audit_type": "QUALIFICATION_GATE",
                "reason": blocker,
                "count": 1,
                "raw_member_identifier_emitted": False,
                "raw_sequence_emitted": False,
            }
        )
    rows.sort(key=lambda row: (row["audit_type"], row["reason"]))
    return rows


def _gate_statuses(
    *,
    blockers: Sequence[str],
    paper_report: Mapping[str, Any],
    leakage_report: Mapping[str, Any],
    power_report: Mapping[str, Any],
    license_exposure: Mapping[str, Any],
    canonical_materialized: bool,
) -> dict[str, str]:
    blockers_set = set(blockers)
    return {
        "AUTHORITY_AND_CODE_TRUST_ROOTS": (
            "PASS" if "IMPLEMENTATION_COMMIT_UNKNOWN_NOT_ASSERTED" not in blockers_set else "UNKNOWN_NOT_ASSERTED"
        ),
        "EXACT_21_ASSET_MANIFEST_AND_PAYLOAD_INTEGRITY": "PASS",
        "EXACT_18_TABLE_CONTEXT_ASSAY_REPLICATE_GRID": "PASS",
        "WITHIN_CONTEXT_KEY_SET_ALIGNMENT_AND_MISSING_NOT_ZERO": "PASS",
        "OUTCOME_BLIND_STRICT_SOURCE_CANDIDATE_MAPPING": (
            "PASS"
            if not blockers_set
            & {
                "OUTCOME_BLIND_LONG_READ_MAPPING_PROVENANCE_UNKNOWN_NOT_ASSERTED",
                "UNADJUDICATED_SEQUENCE_UNIVERSE_CLASSES_PRESENT",
                "UNADJUDICATED_OR_AMBIGUOUS_MAPPING_ROWS_PRESENT",
                "NO_STRICT_OUTCOME_BLIND_SOURCE_CANDIDATE_PAIRS",
                "RAW_KEY_UNCLASSIFIED_OUTCOME_BLIND_RECONCILIATION_NOT_ZERO",
                "RAW_KEY_CLASSIFICATION_NOT_MUTUALLY_EXCLUSIVE",
                "MAPPING_KEYS_ABSENT_FROM_BOTH_CONTEXTS_NOT_ZERO",
            }
            else "BLOCKED_PENDING_PUBLIC_EVIDENCE"
        ),
        "PAPER_NATIVE_TRANSFORM_TEST_AND_MULTIPLE_TESTING_REPRODUCTION": (
            "PASS" if paper_report["paper_method_reproduced"] else "BLOCKED_PENDING_PUBLIC_EVIDENCE"
        ),
        "THREE_BIOLOGICAL_REPLICATES_AND_ROUTE_A_SE": (
            "PASS"
            if "NO_PAIR_CLOSED_ACROSS_BOTH_CONTEXTS_AND_BOTH_ENDPOINTS" not in blockers_set
            else "FAIL_CURRENT_PROTOCOL"
        ),
        "CANONICAL_V3_SCHEMA_AND_HASH_LINEAGE": (
            "PASS" if canonical_materialized else "NOT_RUN_BLOCKED_BY_UPSTREAM_GATES"
        ),
        "LICENSE_AND_REDISTRIBUTION_AUDIT": (
            "BLOCKED_CANONICAL_SEQUENCE_REDISTRIBUTION"
            if "CANONICAL_SEQUENCE_REDISTRIBUTION_NOT_ALLOWED" in blockers_set
            else license_exposure["license_audit_status"]
        ),
        "CHECKPOINT_SPECIFIC_FOUNDATION_EXPOSURE_AUDIT": license_exposure[
            "foundation_exposure_audit_status"
        ],
        "GROUP_AND_SEQUENCE_CLUSTER_LEAKAGE_AUDIT": leakage_report["leakage_audit_status"],
        "PREFROZEN_GROUP_POWER_SIMULATION": power_report["status"],
    }


def _build_qualification_payloads(
    *,
    repo_root: Path,
    data_root: Path,
    protocol_path: Path,
    asset_manifest_path: Path,
    expected_protocol_sha256: str,
    run_metadata: Mapping[str, str],
    snapshot_root: Path,
) -> tuple[dict[str, bytes], dict[str, Any]]:
    _require_directory(repo_root, label="repository root")
    _require_directory(data_root, label="public data root")
    _require_regular_file(protocol_path, label="qualification protocol", suffix=".json")
    _require_regular_file(asset_manifest_path, label="asset manifest", suffix=".json")

    protocol_payload, protocol_provenance = _read_verified_bytes(
        protocol_path,
        expected_protocol_sha256,
        label="qualification protocol",
        suffix=".json",
    )
    protocol = _validate_protocol(
        _read_json_bytes(protocol_payload, label="qualification protocol")
    )
    trust_provenance, parser_paths, trust_documents, git_binding = _verify_trust_roots(
        repo_root=repo_root,
        protocol=protocol,
        explicit_asset_manifest_path=asset_manifest_path,
        snapshot_root=snapshot_root,
    )
    asset_manifest = _validate_asset_manifest(trust_documents["asset_manifest"])
    p0_manifest_path = data_root / asset_manifest["source_manifest"]["filename"]
    resolved_assets, asset_paths, p0_provenance = _resolve_and_verify_assets(
        data_root=data_root,
        p0_manifest_path=p0_manifest_path,
        asset_manifest=asset_manifest,
        snapshot_root=snapshot_root,
    )
    if _POST_VERIFIED_INPUT_SNAPSHOT_HOOK is not None:
        _POST_VERIFIED_INPUT_SNAPSHOT_HOOK()
    snapshot_lineage = {
        "small_json_mode": VERIFIED_JSON_BYTES_MODE,
        "file_parser_mode": PRIVATE_READ_ONLY_SNAPSHOT_MODE,
        "verified_json_documents": [
            "qualification_protocol",
            "asset_manifest",
            "canonical_schema",
            "p0_source_manifest",
        ],
        "private_read_only_asset_snapshot_count": len(resolved_assets),
        "v4_helper_loaded_from_private_read_only_snapshot": True,
        "scientific_parser_original_path_reopen_after_verification": False,
        "snapshot_lifetime": "QUALIFICATION_PAYLOAD_BUILD_ONLY",
    }
    effective_manifest = {
        "contract_id": CONTRACT_ID,
        "schema_version": "1.0.0",
        "manifest_id": ASSET_MANIFEST_ID,
        "dataset_id": DATASET_ID,
        "asset_count": len(resolved_assets),
        "raw_geo_asset_count": sum(asset["asset_kind"] == "GEO_RAW_COUNT" for asset in resolved_assets),
        "supplement_asset_count": sum(asset["asset_kind"] == "SUPPLEMENT_WORKBOOK" for asset in resolved_assets),
        "source_manifest": p0_provenance,
        "source_manifest_preserved_without_modification": True,
        "assets": resolved_assets,
        "all_input_hashes_verified": True,
        "input_snapshot_lineage": snapshot_lineage,
        "raw_member_identifiers_emitted": False,
    }
    effective_manifest_sha256 = _sha256_bytes(_compact_json_bytes(effective_manifest))
    effective_manifest["effective_asset_set_sha256"] = effective_manifest_sha256
    raw_bundle_sha256 = _raw_bundle_sha256(resolved_assets)

    helper = _load_v4_helpers(parser_paths["v4_helper"])
    moesm8_asset = _asset_by_id(resolved_assets, "GSE149487_MOESM8")
    mapping_asset = _asset_by_id(resolved_assets, "GSE149487_LIM6C_293T")
    sequence_entries, sequence_audit, sequence_class_counts = _load_sequence_universe(
        path=asset_paths["GSE149487_MOESM8"],
        effective_asset=moesm8_asset,
        helper=helper,
    )
    pairs, mapping_audit, pair_audit, mapping_reasons, mapped_members = (
        _load_outcome_blind_member_mapping(
            path=asset_paths["GSE149487_LIM6C_293T"],
            effective_asset=mapping_asset,
            helper=helper,
            sequence_entries=sequence_entries,
            mapping_evidence_status=protocol["mapping"]["outcome_blind_mapping_evidence_status"],
        )
    )
    counts, library_sums, raw_table_audits, cross_context, raw_key_reconciliation = _load_geo_count_grid(
        resolved_assets=resolved_assets,
        asset_paths=asset_paths,
        mapped_members=mapped_members,
        reconciliation_contract=protocol["mapping"]["raw_key_reconciliation"],
    )
    paper_report, replicate_rows, effects, data_rejections = (
        _build_paper_and_companion_results(
            pairs=pairs,
            counts=counts,
            library_sums=library_sums,
            protocol=protocol,
        )
    )
    closed_pairs = _fully_closed_pairs(pairs, effects)
    leakage_report = _audit_group_leakage(closed_pairs)
    post_dedup_group_count = len(
        {str(pair["source_sequence_sha256"]) for pair in closed_pairs}
    )
    power_report = _run_group_power_simulation(post_dedup_group_count, protocol)
    license_exposure = _audit_license_and_exposure(resolved_assets, protocol)
    blockers = _qualification_blockers(
        protocol=protocol,
        git_binding=git_binding,
        sequence_class_counts=sequence_class_counts,
        mapping_reasons=mapping_reasons,
        pair_count=len(pairs),
        closed_pair_count=len(closed_pairs),
        paper_report=paper_report,
        leakage_report=leakage_report,
        power_report=power_report,
        license_exposure=license_exposure,
        raw_key_reconciliation=raw_key_reconciliation,
    )

    schema = trust_documents["canonical_schema"]
    canonical_records: list[dict[str, Any]] = []
    if not blockers:
        canonical_records = _build_canonical_records(
            closed_pairs=closed_pairs,
            effects=effects,
            effective_manifest_sha256=effective_manifest_sha256,
            raw_bundle_sha256=raw_bundle_sha256,
            protocol_sha256=protocol_provenance["sha256"],
            script_sha256=trust_provenance["qualifier"]["sha256"],
            protocol=protocol,
            schema=schema,
            license_exposure=license_exposure,
        )
    qualified = bool(canonical_records) and not blockers
    if not qualified:
        canonical_records = []

    beneficial = sum(
        1
        for effect in effects
        if effect["eligible"] and float(effect["candidate_minus_source_delta"]) > 0
    )
    adverse = sum(
        1
        for effect in effects
        if effect["eligible"] and float(effect["candidate_minus_source_delta"]) < 0
    )
    zero = sum(
        1
        for effect in effects
        if effect["eligible"] and float(effect["candidate_minus_source_delta"]) == 0
    )
    gate_statuses = _gate_statuses(
        blockers=blockers,
        paper_report=paper_report,
        leakage_report=leakage_report,
        power_report=power_report,
        license_exposure=license_exposure,
        canonical_materialized=qualified,
    )
    rejection_rows = _safe_rejection_rows(
        sequence_class_counts=sequence_class_counts,
        mapping_reasons=mapping_reasons,
        data_rejections=data_rejections,
        pair_count=len(pairs),
        closed_pair_count=len(closed_pairs),
        blockers=blockers,
    )
    mapping_output_rows = sequence_audit + mapping_audit + pair_audit
    mapping_output_rows.sort(
        key=lambda row: (
            str(row.get("audit_type")),
            str(row.get("pair_id", "")),
            str(row.get("description_sha256", row.get("row_locator_sha256", ""))),
        )
    )
    candidate_canonical_count = len(closed_pairs) * len(CONTEXTS) * len(ENDPOINTS)
    report = {
        "contract_id": CONTRACT_ID,
        "schema_version": "1.0.0",
        "record_type": "GSE149487_PLUMAGE_FULL_A1_QUALIFICATION_REPORT",
        "dataset_id": DATASET_ID,
        "dataset_alias": DATASET_ALIAS,
        "study_group_id": STUDY_GROUP_ID,
        "independent_study_count_if_qualified": 1,
        "endpoint_or_context_increases_independent_study_count": False,
        "run_id": run_metadata["run_id"],
        "execution_id": run_metadata["execution_id"],
        "recorded_at": run_metadata["recorded_at"],
        "qualification_status": QUALIFICATION_STATUS if qualified else BLOCKED_STATUS,
        "evidence_status": "PASS" if qualified else BLOCKED_STATUS,
        "scientific_claim_status": "NOT_ESTABLISHED",
        "qualified": qualified,
        "canonical_intervention_record_v3_materialized": qualified,
        "canonical_record_count": len(canonical_records),
        "candidate_canonical_record_count_before_gate": candidate_canonical_count,
        "qualified_independent_ordinary_study_count": 1 if qualified else 0,
        "qualified_a1_study_count": 1 if qualified else 0,
        "qualified_a2_dense_study_count": 0,
        "training_allowed": False,
        "model_selection_allowed": False,
        "authority_update_performed": False,
        "global_a2_benchmark_split_frozen": False,
        "blockers": blockers,
        "gate_statuses": gate_statuses,
        "authority": {
            "protocol_sha256": protocol_provenance["sha256"],
            "asset_manifest_sha256": trust_provenance["asset_manifest"]["sha256"],
            "canonical_schema_sha256": trust_provenance["canonical_schema"]["sha256"],
            "v4_helper_sha256": trust_provenance["v4_helper"]["sha256"],
            "qualifier_sha256": trust_provenance["qualifier"]["sha256"],
            "focused_test_sha256": trust_provenance["focused_test"]["sha256"],
            "implementation_commit_binding": git_binding,
            "input_snapshot_lineage": snapshot_lineage,
            "v4_artifacts_modified": False,
        },
        "asset_integrity": {
            "effective_asset_count": len(resolved_assets),
            "effective_asset_set_sha256": effective_manifest_sha256,
            "raw_bundle_sha256": raw_bundle_sha256,
            "all_hashes_verified": True,
            "scientific_parsers_used_only_verified_bytes_or_private_snapshots": True,
        },
        "raw_grid": {
            "table_count": len(raw_table_audits),
            "table_audits": raw_table_audits,
            "within_context_key_sets_exactly_equal": True,
            "cross_context": cross_context,
            "outcome_blind_key_reconciliation": raw_key_reconciliation,
        },
        "summary": {
            "nominal_mapping_member_rows": len(mapped_members),
            "strict_source_candidate_pairs": len(pairs),
            "fully_closed_source_candidate_pairs": len(closed_pairs),
            "distinct_candidates": len({pair["candidate_id"] for pair in closed_pairs}),
            "biological_source_groups": len({pair["biological_source_group_id"] for pair in closed_pairs}),
            "gene_groups": len({pair["gene_group_id"] for pair in closed_pairs}),
            "study_groups": 1 if closed_pairs else 0,
            "eligible_multi_candidate_pools": 0,
            "edit_count_strata": {"1": len(closed_pairs)},
            "replicate_and_se_coverage": {
                "eligible_context_endpoint_effects": sum(effect["eligible"] for effect in effects),
                "effects_with_three_biological_replicates_and_se": sum(
                    effect["eligible"]
                    and effect["biological_replicate_count"] == 3
                    and effect["standard_error"] is not None
                    for effect in effects
                ),
                "technical_members_are_independent_n": False,
            },
            "beneficial_and_noise_zone_balance": {
                "beneficial_direction": "HIGHER_IS_BETTER",
                "positive_effect_count": beneficial,
                "negative_effect_count": adverse,
                "exact_zero_effect_count": zero,
                "noise_equivalence_margin": "UNKNOWN_NOT_ASSERTED",
            },
            "post_dedup_effective_n": {
                "value": post_dedup_group_count,
                "unit": "UNIQUE_SOURCE_SEQUENCE_BIOLOGICAL_GROUP",
                "power_gate_status": power_report["status"],
            },
            "foundation_exposure": {
                "status": license_exposure["foundation_exposure_audit_status"],
                "stratum": license_exposure["foundation_exposure_stratum"],
                "checkpoint_bound": license_exposure["checkpoint_bound"],
            },
            "license_and_redistribution_status": {
                "status": license_exposure["license_audit_status"],
                "canonical_status": license_exposure["canonical_license_status"],
                "redistribution_allowed": license_exposure["redistribution_allowed"],
                "canonical_sequence_materialization_allowed": license_exposure[
                    "canonical_sequence_materialization_allowed"
                ],
            },
        },
        "paper_native_and_route_a_separate": True,
        "paper_native_artifact": "PAPER_NATIVE_REPRODUCTION.json",
        "route_a_companion_artifact": "replicate_effect_summaries.jsonl",
        "route_a_companion_artifact_record_type": "ROUTE_A_COMPANION_AGGREGATE_EFFECT",
        "route_a_companion_artifact_contains_replicate_deltas_mean_and_se": True,
        "route_a_companion_artifact_contains_paper_native_rows": False,
        "canonical_success_only_artifact": CANONICAL_FILENAME,
        "raw_member_identifiers_emitted": False,
        "raw_rows_emitted": False,
    }

    payloads: dict[str, bytes] = {
        "ASSET_MANIFEST_EFFECTIVE.json": _pretty_json_bytes(effective_manifest),
        "CONSTRUCT_MAPPING_AUDIT.jsonl": _jsonl_bytes(mapping_output_rows),
        "REJECTION_AUDIT.jsonl": _jsonl_bytes(rejection_rows),
        "PAPER_NATIVE_REPRODUCTION.json": _pretty_json_bytes(paper_report),
        "replicate_effect_summaries.jsonl": _jsonl_bytes(effects),
        "GROUP_LEAKAGE_AUDIT.json": _pretty_json_bytes(leakage_report),
        "GROUP_POWER_AUDIT.json": _pretty_json_bytes(power_report),
        "LICENSE_AND_EXPOSURE_AUDIT.json": _pretty_json_bytes(license_exposure),
        "QUALIFICATION_REPORT.json": _pretty_json_bytes(report),
    }
    if qualified:
        payloads[CANONICAL_FILENAME] = _jsonl_bytes(canonical_records)
    return payloads, report


def _write_exclusive(path: Path, payload: bytes) -> None:
    descriptor = os.open(path, os.O_WRONLY | os.O_CREAT | os.O_EXCL, 0o600)
    file_fsynced = False
    try:
        view = memoryview(payload)
        while view:
            written = os.write(descriptor, view)
            if written <= 0:
                raise QualificationError("exclusive output write made no progress")
            view = view[written:]
        os.fsync(descriptor)
        file_fsynced = True
    finally:
        try:
            os.close(descriptor)
        except OSError as exc:
            if file_fsynced:
                raise ExclusiveWriteCommittedCloseError(
                    "exclusive file descriptor close failed after file fsync"
                ) from exc
            raise


def _fsync_directory(path: Path) -> None:
    descriptor = os.open(path, os.O_RDONLY)
    try:
        os.fsync(descriptor)
    finally:
        os.close(descriptor)


def _directory_fsync_is_explicitly_unsupported(exc: OSError) -> bool:
    return exc.errno in ATOMIC_NOREPLACE_UNSUPPORTED_ERRNOS


def _rename_directory_noreplace(source: Path, destination: Path) -> list[str]:
    if source.parent != destination.parent:
        raise QualificationError("atomic no-replace publication requires one parent directory")
    parent_flags = (
        os.O_RDONLY
        | getattr(os, "O_CLOEXEC", 0)
        | getattr(os, "O_DIRECTORY", 0)
    )
    try:
        parent_descriptor = os.open(source.parent, parent_flags)
    except OSError as exc:
        raise QualificationError("publication parent directory could not be opened") from exc
    operation_error: BaseException | None = None
    operation_traceback: Any = None
    committed = False
    try:
        libc = ctypes.CDLL(None, use_errno=True)
        source_name = os.fsencode(source.name)
        destination_name = os.fsencode(destination.name)
        if sys.platform.startswith("linux"):
            rename_function = getattr(libc, "renameat2", None)
            if rename_function is None:
                raise AtomicNoReplaceUnsupported(errno.ENOSYS)
            rename_function.argtypes = [
                ctypes.c_int,
                ctypes.c_char_p,
                ctypes.c_int,
                ctypes.c_char_p,
                ctypes.c_uint,
            ]
            rename_function.restype = ctypes.c_int
            result = rename_function(
                parent_descriptor,
                source_name,
                parent_descriptor,
                destination_name,
                1,  # Linux RENAME_NOREPLACE
            )
        elif sys.platform == "darwin":
            rename_function = getattr(libc, "renameatx_np", None)
            if rename_function is None:
                raise AtomicNoReplaceUnsupported(errno.ENOSYS)
            rename_function.argtypes = [
                ctypes.c_int,
                ctypes.c_char_p,
                ctypes.c_int,
                ctypes.c_char_p,
                ctypes.c_uint,
            ]
            rename_function.restype = ctypes.c_int
            result = rename_function(
                parent_descriptor,
                source_name,
                parent_descriptor,
                destination_name,
                0x00000004,  # Darwin RENAME_EXCL
            )
        else:
            raise AtomicNoReplaceUnsupported(errno.ENOSYS)
        if result != 0:
            error_number = ctypes.get_errno()
            if error_number in {errno.EEXIST, errno.ENOTEMPTY}:
                raise PublicationContended(
                    "final output appeared during atomic no-replace publication"
                )
            if error_number in ATOMIC_NOREPLACE_UNSUPPORTED_ERRNOS:
                raise AtomicNoReplaceUnsupported(error_number)
            raise QualificationError(
                f"atomic no-replace directory publication failed with errno {error_number}"
            )
        committed = True
    except BaseException as exc:
        operation_error = exc
        operation_traceback = exc.__traceback__
    close_error: OSError | None = None
    try:
        os.close(parent_descriptor)
    except OSError as exc:
        close_error = exc
    if operation_error is not None:
        raise operation_error.with_traceback(operation_traceback)
    if not committed:
        raise QualificationError("atomic no-replace publication ended without commit")
    if close_error is not None:
        return ["POST_COMMIT_PARENT_DIRECTORY_DESCRIPTOR_CLOSE_FAILED"]
    return []


def _bundle_filename_set_sha256(names: Iterable[str]) -> str:
    normalized = sorted(names)
    if not normalized or len(normalized) != len(set(normalized)):
        raise QualificationError("publication bundle filename set is empty or duplicated")
    if any(Path(name).name != name or name in {"", PUBLICATION_COMMIT_FILENAME} for name in normalized):
        raise QualificationError("publication bundle filename set is unsafe")
    payload = (
        "GSE149487_PUBLICATION_BUNDLE_FILENAME_SET_V1\n"
        + "\n".join(normalized)
        + "\n"
    ).encode("utf-8")
    return _sha256_bytes(payload)


def _final_output_directory_name_sha256(name: str) -> str:
    if not isinstance(name, str) or not name or Path(name).name != name:
        raise QualificationError("final output directory name is unsafe")
    return _sha256_bytes(
        b"GSE149487_FINAL_OUTPUT_DIRECTORY_NAME_V1\n" + os.fsencode(name) + b"\n"
    )


def _final_output_target_sha256(path: Path) -> str:
    absolute = _absolute_without_resolving(path)
    return _sha256_bytes(
        b"GSE149487_FINAL_OUTPUT_TARGET_V1\n" + os.fsencode(absolute) + b"\n"
    )


def _publication_commit_document(
    *,
    complete_payloads: Mapping[str, bytes],
    run_metadata: Mapping[str, str],
    publication_mode: str,
    final_output_directory: Path,
) -> dict[str, Any]:
    if publication_mode not in {PRIMARY_PUBLICATION_MODE, FALLBACK_PUBLICATION_MODE}:
        raise QualificationError("publication commit mode is not allowed")
    metadata = _require_exact_keys(
        run_metadata,
        {"run_id", "execution_id", "recorded_at"},
        label="publication run metadata",
    )
    validated_metadata = _validate_run_metadata(
        run_id=metadata["run_id"],
        execution_id=metadata["execution_id"],
        recorded_at=metadata["recorded_at"],
    )
    names = set(complete_payloads)
    if "SHA256SUMS" not in names or PUBLICATION_COMMIT_FILENAME in names:
        raise QualificationError("publication bundle payload set cannot be committed")
    return {
        "schema_version": "1.0.0",
        "record_type": "GSE149487_PLUMAGE_PUBLICATION_COMMIT",
        "contract_id": CONTRACT_ID,
        "protocol_id": PROTOCOL_ID,
        "dataset_id": DATASET_ID,
        "publication_mode": publication_mode,
        "run_id": validated_metadata["run_id"],
        "execution_id": validated_metadata["execution_id"],
        "recorded_at": validated_metadata["recorded_at"],
        "sha256sums_filename": "SHA256SUMS",
        "sha256sums_sha256": _sha256_bytes(complete_payloads["SHA256SUMS"]),
        "bundle_file_count_excluding_commit_marker": len(names),
        "bundle_filename_set_sha256": _bundle_filename_set_sha256(names),
        "final_output_directory_name_sha256": (
            _final_output_directory_name_sha256(final_output_directory.name)
        ),
        "final_output_target_sha256": _final_output_target_sha256(
            final_output_directory
        ),
        "committed": True,
        "commit_marker_written_last": True,
        "canonical_acceptance_requires_valid_commit_marker": True,
    }


def _read_publication_regular_file(path: Path, *, label: str) -> bytes:
    flags = os.O_RDONLY | getattr(os, "O_CLOEXEC", 0) | getattr(os, "O_NOFOLLOW", 0)
    try:
        descriptor = os.open(path, flags)
    except OSError as exc:
        raise QualificationError(f"{label} could not be opened as a regular file") from exc
    try:
        if not stat.S_ISREG(os.fstat(descriptor).st_mode):
            raise QualificationError(f"{label} is not a regular file")
        chunks: list[bytes] = []
        while True:
            chunk = os.read(descriptor, 1 << 20)
            if not chunk:
                break
            chunks.append(chunk)
        return b"".join(chunks)
    except OSError as exc:
        raise QualificationError(f"{label} could not be read") from exc
    finally:
        os.close(descriptor)


def _validate_publication_commit(
    output_directory: Path,
    *,
    expected_run_metadata: Mapping[str, str] | None = None,
    expected_publication_mode: str | None = None,
    expected_final_output_target: Path | None = None,
) -> dict[str, Any]:
    try:
        output_stat = output_directory.lstat()
    except OSError as exc:
        raise QualificationError("fallback publication directory is unavailable") from exc
    if not stat.S_ISDIR(output_stat.st_mode):
        raise QualificationError("fallback publication target is not a directory")

    marker_path = output_directory / PUBLICATION_COMMIT_FILENAME
    marker = _read_json_bytes(
        _read_publication_regular_file(marker_path, label="publication commit marker"),
        label="publication commit marker",
    )
    marker_keys = {
        "schema_version",
        "record_type",
        "contract_id",
        "protocol_id",
        "dataset_id",
        "publication_mode",
        "run_id",
        "execution_id",
        "recorded_at",
        "sha256sums_filename",
        "sha256sums_sha256",
        "bundle_file_count_excluding_commit_marker",
        "bundle_filename_set_sha256",
        "final_output_directory_name_sha256",
        "final_output_target_sha256",
        "committed",
        "commit_marker_written_last",
        "canonical_acceptance_requires_valid_commit_marker",
    }
    marker = dict(
        _require_exact_keys(marker, marker_keys, label="publication commit marker")
    )
    for key, expected in (
        ("schema_version", "1.0.0"),
        ("record_type", "GSE149487_PLUMAGE_PUBLICATION_COMMIT"),
        ("contract_id", CONTRACT_ID),
        ("protocol_id", PROTOCOL_ID),
        ("dataset_id", DATASET_ID),
        ("sha256sums_filename", "SHA256SUMS"),
        ("committed", True),
        ("commit_marker_written_last", True),
        ("canonical_acceptance_requires_valid_commit_marker", True),
    ):
        _require_exact_value(marker[key], expected, label=f"publication commit marker.{key}")
    if marker["publication_mode"] not in {
        PRIMARY_PUBLICATION_MODE,
        FALLBACK_PUBLICATION_MODE,
    }:
        raise QualificationError("publication commit marker mode is not allowed")
    if expected_publication_mode is not None:
        if expected_publication_mode not in {
            PRIMARY_PUBLICATION_MODE,
            FALLBACK_PUBLICATION_MODE,
        }:
            raise QualificationError("expected publication mode is not allowed")
        _require_exact_value(
            marker["publication_mode"],
            expected_publication_mode,
            label="publication commit marker.publication_mode",
        )
    validated_metadata = _validate_run_metadata(
        run_id=marker["run_id"],
        execution_id=marker["execution_id"],
        recorded_at=marker["recorded_at"],
    )
    for key, value in validated_metadata.items():
        _require_exact_value(marker[key], value, label=f"publication commit marker.{key}")
    if expected_run_metadata is not None:
        expected = _require_exact_keys(
            expected_run_metadata,
            {"run_id", "execution_id", "recorded_at"},
            label="expected publication run metadata",
        )
        expected = _validate_run_metadata(
            run_id=expected["run_id"],
            execution_id=expected["execution_id"],
            recorded_at=expected["recorded_at"],
        )
        for key, value in expected.items():
            _require_exact_value(marker[key], value, label=f"publication commit marker.{key}")

    _require_sha256(marker["sha256sums_sha256"], label="publication SHA256SUMS SHA-256")
    _require_sha256(
        marker["bundle_filename_set_sha256"],
        label="publication bundle filename-set SHA-256",
    )
    _require_sha256(
        marker["final_output_directory_name_sha256"],
        label="publication final output directory-name SHA-256",
    )
    _require_sha256(
        marker["final_output_target_sha256"],
        label="publication final output target SHA-256",
    )
    final_output_target = (
        output_directory
        if expected_final_output_target is None
        else expected_final_output_target
    )
    if marker["final_output_directory_name_sha256"] != (
        _final_output_directory_name_sha256(final_output_target.name)
    ):
        raise QualificationError(
            "publication commit marker final output directory-name hash mismatch"
        )
    if marker["final_output_target_sha256"] != _final_output_target_sha256(
        final_output_target
    ):
        raise QualificationError(
            "publication commit marker final output target hash mismatch"
        )
    file_count = marker["bundle_file_count_excluding_commit_marker"]
    if type(file_count) is not int or file_count <= 0:
        raise QualificationError("publication bundle file count must be a positive integer")

    try:
        names = {entry.name for entry in output_directory.iterdir()}
    except OSError as exc:
        raise QualificationError("fallback publication directory could not be listed") from exc
    if any(Path(name).name != name or not name for name in names):
        raise QualificationError("fallback publication directory contains an unsafe filename")
    if PUBLICATION_COMMIT_FILENAME not in names:
        raise QualificationError("fallback publication commit marker is absent")
    bundle_names = names - {PUBLICATION_COMMIT_FILENAME}
    if len(bundle_names) != file_count:
        raise QualificationError("publication commit marker bundle file count mismatch")
    if _bundle_filename_set_sha256(bundle_names) != marker["bundle_filename_set_sha256"]:
        raise QualificationError("publication commit marker bundle filename-set hash mismatch")

    sums_payload = _read_publication_regular_file(
        output_directory / "SHA256SUMS",
        label="published SHA256SUMS",
    )
    if _sha256_bytes(sums_payload) != marker["sha256sums_sha256"]:
        raise QualificationError("publication commit marker SHA256SUMS hash mismatch")
    try:
        sums_text = sums_payload.decode("ascii")
    except UnicodeDecodeError as exc:
        raise QualificationError("published SHA256SUMS is not ASCII") from exc
    if not sums_text.endswith("\n") or "\r" in sums_text:
        raise QualificationError("published SHA256SUMS has non-canonical line endings")
    sum_lines = sums_text[:-1].split("\n")
    if not sum_lines or any(not line for line in sum_lines):
        raise QualificationError("published SHA256SUMS has an invalid line set")
    declared_hashes: dict[str, str] = {}
    for line in sum_lines:
        if len(line) < 67 or line[64:66] != "  ":
            raise QualificationError("published SHA256SUMS line is malformed")
        digest, name = line[:64], line[66:]
        _require_sha256(digest, label="published bundle member SHA-256")
        if Path(name).name != name or not name or name in declared_hashes:
            raise QualificationError("published SHA256SUMS filename is unsafe or duplicated")
        declared_hashes[name] = digest
    expected_member_names = bundle_names - {"SHA256SUMS"}
    if set(declared_hashes) != expected_member_names:
        raise QualificationError("published SHA256SUMS member set mismatch")
    for name, expected_digest in declared_hashes.items():
        payload = _read_publication_regular_file(
            output_directory / name,
            label="published bundle member",
        )
        if _sha256_bytes(payload) != expected_digest:
            raise QualificationError("published bundle member SHA-256 mismatch")
    try:
        final_names = {entry.name for entry in output_directory.iterdir()}
    except OSError as exc:
        raise QualificationError("fallback publication directory final listing failed") from exc
    if final_names != names:
        raise QualificationError("fallback publication directory changed during validation")
    return marker


def _publish_bundle_with_terminal_marker(
    *,
    output_directory: Path,
    complete_payloads: Mapping[str, bytes],
    run_metadata: Mapping[str, str],
    unsupported_errno: int,
    source_staging: Path,
    precommit_capability_warnings: Sequence[str],
) -> dict[str, Any]:
    if unsupported_errno not in ATOMIC_NOREPLACE_UNSUPPORTED_ERRNOS:
        raise QualificationError("fallback requested for a non-approved errno")
    commit_document = _publication_commit_document(
        complete_payloads=complete_payloads,
        run_metadata=run_metadata,
        publication_mode=FALLBACK_PUBLICATION_MODE,
        final_output_directory=output_directory,
    )
    try:
        os.mkdir(output_directory, 0o700)
    except FileExistsError as exc:
        raise PublicationContended(
            "final output appeared during atomic fallback directory creation"
        ) from exc
    except OSError as exc:
        raise QualificationError("atomic fallback directory creation failed") from exc

    marker_path = output_directory / PUBLICATION_COMMIT_FILENAME
    commit_marker: dict[str, Any] | None = None
    durability_warnings = list(precommit_capability_warnings)
    try:
        for name in sorted(complete_payloads):
            if Path(name).name != name or name == PUBLICATION_COMMIT_FILENAME:
                raise QualificationError("fallback output filename is unsafe")
            _write_exclusive(output_directory / name, complete_payloads[name])
        _write_exclusive(marker_path, _pretty_json_bytes(commit_document))
        commit_marker = _validate_publication_commit(
            output_directory,
            expected_run_metadata=run_metadata,
            expected_publication_mode=FALLBACK_PUBLICATION_MODE,
        )
    except Exception as exc:
        try:
            commit_marker = _validate_publication_commit(
                output_directory,
                expected_run_metadata=run_metadata,
                expected_publication_mode=FALLBACK_PUBLICATION_MODE,
            )
        except Exception as validation_exc:
            raise PartialPublicationError(
                "fallback publication is PARTIAL_NOT_COMMITTED; terminal commit "
                "marker is absent or invalid, the directory is preserved, and retry "
                "requires a new run ID",
                unsupported_errno=unsupported_errno,
                commit_marker_present=marker_path.exists() or marker_path.is_symlink(),
            ) from validation_exc
        durability_warnings.append(
            "POST_COMMIT_MARKER_WRITE_OR_VALIDATION_FINALIZATION_ERROR"
        )

    if commit_marker is None:
        raise PartialPublicationError(
            "fallback publication is PARTIAL_NOT_COMMITTED without a validated marker",
            unsupported_errno=unsupported_errno,
            commit_marker_present=marker_path.exists() or marker_path.is_symlink(),
        )
    try:
        _fsync_directory(output_directory)
    except OSError:
        durability_warnings.append("POST_COMMIT_OUTPUT_DIRECTORY_FSYNC_FAILED")
    try:
        _fsync_directory(output_directory.parent)
    except OSError:
        durability_warnings.append("POST_COMMIT_PARENT_DIRECTORY_FSYNC_FAILED")
    return {
        "status": (
            "PUBLISHED_WITH_POST_COMMIT_DURABILITY_WARNING"
            if durability_warnings
            else "PUBLISHED_DURABLE"
        ),
        "publication_mode": FALLBACK_PUBLICATION_MODE,
        "directory_committed": True,
        "atomic_no_replace": False,
        "atomic_exclusive_final_mkdir": True,
        "fallback_trigger_errno": unsupported_errno,
        "terminal_commit_marker_filename": PUBLICATION_COMMIT_FILENAME,
        "terminal_commit_marker_validated": True,
        "terminal_commit_marker_postcommit_revalidated": True,
        "bundle_file_count_excluding_commit_marker": commit_marker[
            "bundle_file_count_excluding_commit_marker"
        ],
        "sha256sums_sha256": commit_marker["sha256sums_sha256"],
        "partial_not_committed": False,
        "source_staging_preserved": source_staging.exists(),
        "durability_warning_codes": sorted(set(durability_warnings)),
        "failure_record_allowed": False,
    }


def _publish_bundle(
    output_directory: Path,
    payloads: Mapping[str, bytes],
    *,
    run_metadata: Mapping[str, str] | None = None,
) -> dict[str, Any]:
    if run_metadata is None:
        raise QualificationError(
            "publication requires validated run metadata for its terminal marker"
        )
    expected = set(ALWAYS_OUTPUT_FILES) - {"SHA256SUMS"}
    allowed = expected | {CANONICAL_FILENAME}
    payload_names = set(payloads)
    if payload_names != expected and payload_names != allowed:
        raise QualificationError("internal output payload set violates the closed contract")
    staging = Path(
        tempfile.mkdtemp(
            prefix=f".{output_directory.name}.partial-staging-",
            dir=output_directory.parent,
        )
    )
    precommit_capability_warnings: list[str] = []
    try:
        for name in sorted(payloads):
            if Path(name).name != name:
                raise QualificationError("internal output filename is unsafe")
            _write_exclusive(staging / name, payloads[name])
        sums = [
            f"{_sha256_bytes(payloads[name])}  {name}"
            for name in sorted(payloads)
        ]
        sums_payload = ("\n".join(sums) + "\n").encode("ascii")
        _write_exclusive(staging / "SHA256SUMS", sums_payload)
        complete_payloads = {**payloads, "SHA256SUMS": sums_payload}
        primary_commit_document = _publication_commit_document(
            complete_payloads=complete_payloads,
            run_metadata=run_metadata,
            publication_mode=PRIMARY_PUBLICATION_MODE,
            final_output_directory=output_directory,
        )
        _write_exclusive(
            staging / PUBLICATION_COMMIT_FILENAME,
            _pretty_json_bytes(primary_commit_document),
        )
        _validate_publication_commit(
            staging,
            expected_run_metadata=run_metadata,
            expected_publication_mode=PRIMARY_PUBLICATION_MODE,
            expected_final_output_target=output_directory,
        )
        try:
            _fsync_directory(staging)
        except OSError as exc:
            if not _directory_fsync_is_explicitly_unsupported(exc):
                raise
            precommit_capability_warnings.append(
                "PRECOMMIT_STAGING_DIRECTORY_FSYNC_UNSUPPORTED"
            )
    except Exception as exc:
        if isinstance(exc, QualificationError):
            raise
        raise QualificationError("bundle publication failed before atomic commit") from exc

    try:
        rename_warnings = _rename_directory_noreplace(staging, output_directory)
    except AtomicNoReplaceUnsupported as exc:
        return _publish_bundle_with_terminal_marker(
            output_directory=output_directory,
            complete_payloads=complete_payloads,
            run_metadata=run_metadata,
            unsupported_errno=exc.error_number,
            source_staging=staging,
            precommit_capability_warnings=precommit_capability_warnings,
        )

    durability_warnings = list(precommit_capability_warnings)
    durability_warnings.extend(rename_warnings or [])
    marker_postcommit_revalidated = False
    marker_validation_error: Exception | None = None
    for attempt in range(2):
        try:
            _validate_publication_commit(
                output_directory,
                expected_run_metadata=run_metadata,
                expected_publication_mode=PRIMARY_PUBLICATION_MODE,
            )
        except Exception as exc:
            marker_validation_error = exc
            continue
        if attempt == 1:
            durability_warnings.append(
                "POST_COMMIT_TERMINAL_MARKER_REVALIDATION_RETRY_REQUIRED"
            )
        marker_postcommit_revalidated = True
        marker_validation_error = None
        break
    try:
        _fsync_directory(output_directory.parent)
    except OSError:
        durability_warnings.append("POST_COMMIT_PARENT_DIRECTORY_FSYNC_FAILED")
    if marker_validation_error is not None:
        raise CommittedPublicationValidationError(
            "atomically committed output failed terminal commit-marker validation "
            "twice and is not accepted for qualification or canonical use",
            publication_mode=PRIMARY_PUBLICATION_MODE,
            durability_warning_codes=durability_warnings,
        ) from marker_validation_error
    return {
        "status": (
            "PUBLISHED_WITH_POST_COMMIT_DURABILITY_WARNING"
            if durability_warnings
            else "PUBLISHED_DURABLE"
        ),
        "publication_mode": PRIMARY_PUBLICATION_MODE,
        "directory_committed": True,
        "atomic_no_replace": True,
        "atomic_exclusive_final_mkdir": False,
        "terminal_commit_marker_filename": PUBLICATION_COMMIT_FILENAME,
        "terminal_commit_marker_validated": True,
        "terminal_commit_marker_postcommit_revalidated": marker_postcommit_revalidated,
        "partial_not_committed": False,
        "durability_warning_codes": sorted(set(durability_warnings)),
        "failure_record_allowed": False,
    }


def _transaction_claim_path(output_directory: Path) -> Path:
    return output_directory.parent / f".{output_directory.name}.qualification-transaction-claim"


def _acquire_transaction_claim(
    *,
    output_directory: Path,
    failure_record_path: Path,
) -> tuple[Path, int, list[str]]:
    if failure_record_path.parent != output_directory.parent:
        raise QualificationError(
            "failure record must be an exclusive sibling of the output directory"
        )
    if failure_record_path == output_directory:
        raise QualificationError("failure record and output directory must be distinct")
    claim = _transaction_claim_path(output_directory)
    _reject_forbidden_path(claim, label="qualification transaction claim")
    claim_created = False
    descriptor: int | None = None
    capability_warnings: list[str] = []
    try:
        descriptor = os.open(
            claim,
            os.O_WRONLY | os.O_CREAT | os.O_EXCL | getattr(os, "O_CLOEXEC", 0),
            0o600,
        )
        claim_created = True
        claim_payload = (
            "GSE149487_PLUMAGE_FULL_A1_OUTPUT_FAILURE_TRANSACTION\n"
            f"output={output_directory.name}\n"
            f"failure={failure_record_path.name}\n"
        ).encode("utf-8")
        view = memoryview(claim_payload)
        while view:
            written = os.write(descriptor, view)
            if written <= 0:
                raise QualificationError("transaction claim write made no progress")
            view = view[written:]
        os.fsync(descriptor)
        try:
            _fsync_directory(claim.parent)
        except OSError as exc:
            if not _directory_fsync_is_explicitly_unsupported(exc):
                raise
            capability_warnings.append(
                "TRANSACTION_CLAIM_PARENT_DIRECTORY_FSYNC_UNSUPPORTED"
            )
    except FileExistsError as exc:
        raise TransactionClaimContended(
            "qualification transaction claim is already owned"
        ) from exc
    except Exception as exc:
        if descriptor is not None:
            os.close(descriptor)
            descriptor = None
        if claim_created:
            try:
                claim.unlink()
            except FileNotFoundError:
                pass
        if isinstance(exc, QualificationError):
            raise
        raise QualificationError(
            "transaction claim could not be durably acquired"
        ) from exc
    if descriptor is None:
        raise QualificationError("transaction claim descriptor was not acquired")
    return claim, descriptor, capability_warnings


def _release_transaction_claim(claim: Path, descriptor: int) -> list[str]:
    warnings: list[str] = []
    try:
        os.close(descriptor)
    except OSError:
        warnings.append("TRANSACTION_CLAIM_DESCRIPTOR_CLOSE_FAILED")
    try:
        claim.unlink()
    except FileNotFoundError:
        pass
    except OSError:
        warnings.append("TRANSACTION_CLAIM_CLEANUP_FAILED")
    else:
        try:
            _fsync_directory(claim.parent)
        except OSError:
            warnings.append("TRANSACTION_CLAIM_CLEANUP_DIRECTORY_FSYNC_FAILED")
    return warnings


def _publish_failure_record(path: Path, payload: Mapping[str, Any]) -> list[str]:
    failure_bytes = _pretty_json_bytes(payload)
    warnings: list[str] = []
    try:
        _write_exclusive(path, failure_bytes)
    except ExclusiveWriteCommittedCloseError as exc:
        try:
            verified_bytes = _read_publication_regular_file(
                path,
                label="committed failure record",
            )
        except QualificationError as verification_exc:
            raise QualificationError(
                "failure record close failed and committed bytes could not be verified"
            ) from verification_exc
        if verified_bytes != failure_bytes:
            raise QualificationError(
                "failure record close failed and committed bytes do not match"
            ) from exc
        warnings.append("FAILURE_RECORD_DESCRIPTOR_CLOSE_FAILED")
    try:
        _fsync_directory(path.parent)
    except OSError:
        warnings.append("FAILURE_RECORD_PARENT_DIRECTORY_FSYNC_FAILED")
    return warnings


def qualify_gse149487_plumage(
    *,
    repo_root: Path,
    data_root: Path,
    protocol_path: Path,
    asset_manifest_path: Path,
    expected_protocol_sha256: str,
    output_directory: Path,
    run_id: str,
    execution_id: str,
    recorded_at: str,
) -> dict[str, Any]:
    raw_paths = {
        "repository root": Path(repo_root),
        "public data root": Path(data_root),
        "qualification protocol": Path(protocol_path),
        "asset manifest": Path(asset_manifest_path),
        "output directory": Path(output_directory),
    }
    for label, path in raw_paths.items():
        _reject_forbidden_path(path, label=label)
    paths = {label: _absolute_without_resolving(path) for label, path in raw_paths.items()}
    for label, path in paths.items():
        _reject_forbidden_path(path, label=label)
    _require_absent_target(paths["output directory"], label="output directory")
    run_metadata = _validate_run_metadata(
        run_id=run_id,
        execution_id=execution_id,
        recorded_at=recorded_at,
    )
    with tempfile.TemporaryDirectory(
        prefix="route-a-v3-gse149487-verified-inputs-"
    ) as snapshot_root_text:
        snapshot_root = Path(snapshot_root_text)
        os.chmod(snapshot_root, 0o700)
        payloads, report = _build_qualification_payloads(
            repo_root=paths["repository root"],
            data_root=paths["public data root"],
            protocol_path=paths["qualification protocol"],
            asset_manifest_path=paths["asset manifest"],
            expected_protocol_sha256=_require_sha256(
                expected_protocol_sha256, label="expected protocol SHA-256"
            ),
            run_metadata=run_metadata,
            snapshot_root=snapshot_root,
        )
    publication = _publish_bundle(
        paths["output directory"],
        payloads,
        run_metadata=run_metadata,
    )
    return {**report, "publication": publication}


def execute_qualification(
    *,
    failure_record_path: Path,
    **kwargs: Any,
) -> dict[str, Any]:
    failure_path = _absolute_without_resolving(Path(failure_record_path))
    if "output_directory" not in kwargs:
        raise QualificationError("output directory is required")
    output_path = _absolute_without_resolving(Path(kwargs["output_directory"]))
    _reject_forbidden_path(failure_path, label="failure record")
    _reject_forbidden_path(output_path, label="output directory")
    if failure_path.parent != output_path.parent:
        raise QualificationError(
            "failure record must be an exclusive sibling of the output directory"
        )
    if failure_path == output_path:
        raise QualificationError("failure record and output directory must be distinct")
    _require_absent_target(failure_path, label="failure record", suffix=".json")
    run_id = str(kwargs.get("run_id", "UNKNOWN_NOT_ASSERTED"))
    execution_id = str(kwargs.get("execution_id", "UNKNOWN_NOT_ASSERTED"))
    recorded_at = str(kwargs.get("recorded_at", "UNKNOWN_NOT_ASSERTED"))

    def contention(status: str, detail: str) -> dict[str, Any]:
        return {
            "kind": "CONTENDED",
            "contention": {
                "contract_id": CONTRACT_ID,
                "schema_version": "1.0.0",
                "record_type": "GSE149487_PLUMAGE_FULL_A1_TRANSACTION_CONTENTION",
                "dataset_id": DATASET_ID,
                "run_id": run_id,
                "execution_id": execution_id,
                "recorded_at": recorded_at,
                "status": status,
                "detail": detail,
                "qualified": False,
                "canonical_record_count": 0,
                "failure_record_materialized": False,
                "raw_member_identifiers_emitted": False,
                "raw_rows_emitted": False,
            },
        }

    def failure_payload(failure_type: str, failure_detail: str) -> dict[str, Any]:
        return {
            "contract_id": CONTRACT_ID,
            "schema_version": "1.0.0",
            "record_type": "GSE149487_PLUMAGE_FULL_A1_QUALIFICATION_FAILURE",
            "dataset_id": DATASET_ID,
            "run_id": run_id,
            "execution_id": execution_id,
            "recorded_at": recorded_at,
            "status": "FAIL_CLOSED_BEFORE_SUCCESS_BUNDLE_PUBLICATION",
            "failure_type": failure_type,
            "failure_detail": failure_detail,
            "qualified": False,
            "canonical_intervention_record_v3_materialized": False,
            "canonical_record_count": 0,
            "training_allowed": False,
            "model_selection_allowed": False,
            "raw_member_identifiers_emitted": False,
            "raw_rows_emitted": False,
        }

    try:
        claim, claim_descriptor, claim_acquisition_warnings = _acquire_transaction_claim(
            output_directory=output_path,
            failure_record_path=failure_path,
        )
    except TransactionClaimContended as exc:
        return contention("TRANSACTION_CLAIM_CONTENDED", str(exc))

    result: dict[str, Any]
    try:
        _require_absent_target(failure_path, label="failure record", suffix=".json")
        if output_path.exists() or output_path.is_symlink():
            result = contention(
                "OUTPUT_ALREADY_COMMITTED_OR_OCCUPIED",
                "output target exists; no failure record was published",
            )
        else:
            qualification_kwargs = dict(kwargs)
            qualification_kwargs["output_directory"] = output_path
            try:
                report = qualify_gse149487_plumage(**qualification_kwargs)
            except PublicationContended as exc:
                result = contention(
                    "OUTPUT_PUBLICATION_CONTENDED",
                    f"{exc}; no failure record was published",
                )
            except PartialPublicationError as exc:
                result = {
                    "kind": "PARTIAL_NOT_COMMITTED",
                    "partial": {
                        "contract_id": CONTRACT_ID,
                        "schema_version": "1.0.0",
                        "record_type": (
                            "GSE149487_PLUMAGE_FALLBACK_PARTIAL_NOT_COMMITTED"
                        ),
                        "dataset_id": DATASET_ID,
                        "run_id": run_id,
                        "execution_id": execution_id,
                        "recorded_at": recorded_at,
                        "status": "PARTIAL_NOT_COMMITTED",
                        "publication_mode": FALLBACK_PUBLICATION_MODE,
                        "fallback_trigger_errno": exc.unsupported_errno,
                        "terminal_commit_marker_present": (
                            exc.commit_marker_present
                        ),
                        "terminal_commit_marker_validated": False,
                        "output_directory_preserved": True,
                        "published": False,
                        "qualified": False,
                        "canonical_accepted": False,
                        "canonical_record_count": 0,
                        "failure_record_materialized": False,
                        "retry_requires_new_run_id": True,
                        "raw_member_identifiers_emitted": False,
                        "raw_rows_emitted": False,
                    },
                }
            except CommittedPublicationValidationError as exc:
                result = {
                    "kind": "COMMITTED_NOT_ACCEPTED",
                    "committed_not_accepted": {
                        "contract_id": CONTRACT_ID,
                        "schema_version": "1.0.0",
                        "record_type": (
                            "GSE149487_PLUMAGE_COMMITTED_NOT_ACCEPTED"
                        ),
                        "dataset_id": DATASET_ID,
                        "run_id": run_id,
                        "execution_id": execution_id,
                        "recorded_at": recorded_at,
                        "status": "COMMITTED_NOT_ACCEPTED",
                        "publication_mode": exc.publication_mode,
                        "directory_committed": True,
                        "output_directory_preserved": True,
                        "publication_accepted": False,
                        "terminal_commit_marker_validated": False,
                        "qualified": False,
                        "canonical_accepted": False,
                        "canonical_record_count": 0,
                        "failure_record_materialized": False,
                        "retry_requires_new_run_id": True,
                        "durability_warning_codes": list(
                            exc.durability_warning_codes
                        ),
                        "raw_member_identifiers_emitted": False,
                        "raw_rows_emitted": False,
                    },
                }
            except QualificationError as exc:
                if output_path.exists() or output_path.is_symlink():
                    result = contention(
                        "OUTPUT_COMMITTED_DURING_EXECUTION",
                        "output target appeared; no failure record was published",
                    )
                else:
                    failure = failure_payload(type(exc).__name__, str(exc))
                    failure_record_warnings = _publish_failure_record(
                        failure_path,
                        failure,
                    )
                    if failure_record_warnings:
                        failure["failure_record_durability_warning_codes"] = (
                            failure_record_warnings
                        )
                    result = {"kind": "FAILURE", "failure": failure}
            except Exception as exc:  # pragma: no cover - last-resort fail-closed boundary
                if output_path.exists() or output_path.is_symlink():
                    result = contention(
                        "OUTPUT_COMMITTED_DURING_UNEXPECTED_EXECUTION_ERROR",
                        "output target appeared; no failure record was published",
                    )
                else:
                    failure = failure_payload(
                        "UNEXPECTED_EXECUTION_ERROR",
                        type(exc).__name__,
                    )
                    failure_record_warnings = _publish_failure_record(
                        failure_path,
                        failure,
                    )
                    if failure_record_warnings:
                        failure["failure_record_durability_warning_codes"] = (
                            failure_record_warnings
                        )
                    result = {"kind": "FAILURE", "failure": failure}
            else:
                result = {"kind": "BUNDLE", "report": report}
    finally:
        claim_warnings = sorted(
            set(claim_acquisition_warnings)
            | set(_release_transaction_claim(claim, claim_descriptor))
        )

    if claim_warnings:
        if result["kind"] == "BUNDLE":
            publication = dict(result["report"]["publication"])
            publication["durability_warning_codes"] = sorted(
                set(publication["durability_warning_codes"]) | set(claim_warnings)
            )
            publication["status"] = "PUBLISHED_WITH_POST_COMMIT_DURABILITY_WARNING"
            result = {
                "kind": "BUNDLE",
                "report": {**result["report"], "publication": publication},
            }
        elif result["kind"] == "FAILURE":
            result["failure"]["transaction_claim_warning_codes"] = claim_warnings
        elif result["kind"] == "PARTIAL_NOT_COMMITTED":
            result["partial"]["transaction_claim_warning_codes"] = claim_warnings
        elif result["kind"] == "COMMITTED_NOT_ACCEPTED":
            result["committed_not_accepted"][
                "transaction_claim_warning_codes"
            ] = claim_warnings
        else:
            result["contention"]["transaction_claim_warning_codes"] = claim_warnings
    return result


def _parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--repo-root", required=True, type=Path)
    parser.add_argument("--data-root", required=True, type=Path)
    parser.add_argument("--protocol", required=True, type=Path)
    parser.add_argument("--asset-manifest", required=True, type=Path)
    parser.add_argument("--protocol-sha256", required=True)
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--failure-record", required=True, type=Path)
    parser.add_argument("--run-id", required=True)
    parser.add_argument("--execution-id", required=True)
    parser.add_argument("--recorded-at", required=True)
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = _parse_args(argv)
    result = execute_qualification(
        repo_root=args.repo_root,
        data_root=args.data_root,
        protocol_path=args.protocol,
        asset_manifest_path=args.asset_manifest,
        expected_protocol_sha256=args.protocol_sha256,
        output_directory=args.output_dir,
        failure_record_path=args.failure_record,
        run_id=args.run_id,
        execution_id=args.execution_id,
        recorded_at=args.recorded_at,
    )
    if result["kind"] == "FAILURE":
        failure = result["failure"]
        print(
            json.dumps(
                {
                    "dataset_id": DATASET_ID,
                    "status": failure["status"],
                    "failure_type": failure["failure_type"],
                    "qualified": False,
                    "canonical_record_count": 0,
                    "failure_record_durability_warning_count": len(
                        failure.get("failure_record_durability_warning_codes", [])
                    ),
                    "transaction_claim_warning_count": len(
                        failure.get("transaction_claim_warning_codes", [])
                    ),
                    "durability_warning_count": len(
                        failure.get("failure_record_durability_warning_codes", [])
                    )
                    + len(failure.get("transaction_claim_warning_codes", [])),
                },
                sort_keys=True,
            )
        )
        return 2
    if result["kind"] == "CONTENDED":
        contention = result["contention"]
        print(
            json.dumps(
                {
                    "dataset_id": DATASET_ID,
                    "status": contention["status"],
                    "qualified": False,
                    "canonical_record_count": 0,
                    "failure_record_materialized": False,
                },
                sort_keys=True,
            )
        )
        return 3
    if result["kind"] == "PARTIAL_NOT_COMMITTED":
        partial = result["partial"]
        print(
            json.dumps(
                {
                    "dataset_id": DATASET_ID,
                    "status": partial["status"],
                    "publication_mode": partial["publication_mode"],
                    "published": False,
                    "qualified": False,
                    "canonical_accepted": False,
                    "canonical_record_count": 0,
                    "failure_record_materialized": False,
                    "retry_requires_new_run_id": True,
                },
                sort_keys=True,
            )
        )
        return 4
    if result["kind"] == "COMMITTED_NOT_ACCEPTED":
        committed = result["committed_not_accepted"]
        print(
            json.dumps(
                {
                    "dataset_id": DATASET_ID,
                    "status": committed["status"],
                    "publication_mode": committed["publication_mode"],
                    "directory_committed": True,
                    "publication_accepted": False,
                    "qualified": False,
                    "canonical_accepted": False,
                    "canonical_record_count": 0,
                    "failure_record_materialized": False,
                    "durability_warning_count": len(
                        committed["durability_warning_codes"]
                    )
                    + len(committed.get("transaction_claim_warning_codes", [])),
                },
                sort_keys=True,
            )
        )
        return 5
    report = result["report"]
    print(
        json.dumps(
            {
                "dataset_id": DATASET_ID,
                "qualification_status": report["qualification_status"],
                "qualified": report["qualified"],
                "canonical_record_count": report["canonical_record_count"],
                "qualified_independent_ordinary_study_count": report[
                    "qualified_independent_ordinary_study_count"
                ],
                "publication_mode": report["publication"]["publication_mode"],
                "publication_status": report["publication"]["status"],
                "post_commit_durability_warning_count": len(
                    report["publication"]["durability_warning_codes"]
                ),
                "training_allowed": False,
                "model_selection_allowed": False,
            },
            sort_keys=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
