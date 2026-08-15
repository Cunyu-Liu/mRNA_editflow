#!/usr/bin/env python3
"""Convert the public GSE269595 MPRA to Route 2 Development listwise records."""

from __future__ import annotations

import argparse
import gzip
import json
import math
import os
import re
import shutil
import tempfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
CONFIG_PATH = REPO_ROOT / "configs/route_a_v3_route2_gse269595_converter_v1.json"
CANONICAL_SCHEMA_VERSION = "mrna_editflow_route2_canonical.v1"
BASES = set("ACGT")
LIBRARY_HEADER = (
    "gene_id", "pas_id", "type", "subtype", "experiment", "n_bc", "barcoded_seq_184bp",
)
MPRA_HEADER = (
    "sample", "replicate", "perturbation", "distal_site", "barcode", "gene_id", "pas_id",
    "aim", "subaim", "experiment", "n_bc", "total", "distal", "proximal", "log_odds",
)
EXPECTED_SAMPLE_FIELDS = {
    "CSTF3gA-rep1": ("rep1", "CSTF3"),
    "CSTF3gA-rep2": ("rep2", "CSTF3"),
    "CSTF3gB-rep1": ("rep1", "CSTF3"),
    "CSTF3gB-rep2": ("rep2", "CSTF3"),
    "NTgA-rep1": ("rep1", "NT"),
    "NTgA-rep2": ("rep2", "NT"),
    "NTgB-rep1": ("rep1", "NT"),
    "NTgB-rep2": ("rep2", "NT"),
    "NUDT21gA-rep1": ("rep1", "NUDT21"),
    "NUDT21gA-rep2": ("rep2", "NUDT21"),
    "NUDT21gB-rep1": ("rep1", "NUDT21"),
    "NUDT21gB-rep2": ("rep2", "NUDT21"),
}


class ConversionError(RuntimeError):
    pass


def _require(condition: bool, message: str) -> None:
    if not condition:
        raise ConversionError(message)


def load_config(path: Path = CONFIG_PATH) -> dict[str, Any]:
    config = json.loads(path.read_text(encoding="utf-8"))
    validate_config(config)
    return config


def validate_config(config: Mapping[str, Any]) -> None:
    _require(config["schema_version"] == "route_a_v3_route2_gse269595_converter.v1", "unexpected schema version")
    study = config["study"]
    _require(study["study_unit_id"] == "GSE269595", "unexpected study")
    _require(study["pool_assignment"] == "DEVELOPMENT", "study left Development")
    _require(study["qualification_class"] == "DEVELOPMENT_RELAXED", "qualification class changed")
    _require(study["study_role"] == "TRUE_A2_LISTWISE_DEVELOPMENT_EXPOSED_LIBRARY", "study role changed")
    _require(study["conversion_scope"] == "DENSE_SOURCE_FAMILY_PAIRED_REPLICATE_FINITE_SUB_TRAJECTORY", "conversion scope changed")
    _require(study["biological_context_perturbations"] == ["CSTF3", "NT", "NUDT21"], "perturbation set changed")
    _require(study["distal_reporter_contexts"] == ["CCT6A_moduleA", "CDK1_moduleB", "TMEM106C_moduleB", "TMEM237_moduleA", "bGH"], "distal context set changed")

    expected_input = config["input"]
    _require(expected_input["expected_library_member_count"] == 6113, "library count changed")
    _require(expected_input["expected_family_count"] == 373, "family count changed")
    _require(expected_input["expected_eligible_dense_family_count"] == 372, "dense family count changed")
    _require(expected_input["expected_pairwise_family_exclusion_count"] == 1, "pairwise exclusion count changed")
    _require(expected_input["expected_eligible_candidate_design_count"] == 3427, "candidate count changed")
    _require(expected_input["expected_distinct_eligible_sequence_pair_count"] == 3427, "sequence-pair count changed")
    _require(expected_input["expected_eligible_edit_count_histogram"] == {"11+": 771, "2": 8, "3-5": 551, "6-10": 2097}, "edit histogram changed")
    _require(expected_input["expected_processed_measurement_row_count"] == 366780, "measurement count changed")

    endpoint = config["endpoint_policy"]
    _require(endpoint["guide_arms_pooled_within_perturbation_and_biological_replicate"] is True, "guide pooling changed")
    _require(endpoint["biological_replicates"] == ["rep1", "rep2"], "replicate set changed")
    _require(endpoint["minimum_pooled_total_umi_per_replicate"] == 100, "minimum UMI changed")
    _require(endpoint["require_nonzero_proximal_and_distal_counts"] is True, "finite endpoint rule changed")
    _require(endpoint["source_and_candidate_must_pass_same_rule"] is True, "source endpoint rule changed")
    _require(endpoint["missing_is_zero"] is False, "missing values became zero")
    _require(endpoint["publisher_exact_censor_universe_claimed"] is False, "publisher censor universe overstated")

    expected = config["expected_conversion"]
    _require(expected["candidate_context_universe_count"] == 51405, "candidate-context universe changed")
    _require(expected["reject_candidate_endpoint_qc_count"] == 19518, "candidate QC reject count changed")
    _require(expected["reject_source_endpoint_qc_after_candidate_pass_count"] == 921, "source QC reject count changed")
    _require(expected["canonical_record_count"] == 30966, "canonical count changed")
    _require(expected["source_family_with_record_count"] == 363, "retained family count changed")
    _require(sum(expected["context_record_counts"].values()) == expected["canonical_record_count"], "context counts do not close")

    action = config["action_policy"]
    _require(action == {
        "allowed_candidate_action": "SUB",
        "multi_step_sub_trajectory_supported": True,
        "same_position_repeated_edit_required": False,
        "ins_supported": False,
        "del_supported": False,
    }, "action policy changed")
    development = config["development_policy"]
    _require(development["training_eligible"] is True and development["model_selection_eligible"] is True, "Development use disabled")
    _require(development["confirmatory_evaluation_eligible"] is False and development["unseen_evaluation_eligible"] is False, "Evaluation use enabled")
    _require(development["library_selection_exposure"] == "APARENT_AND_MEASURED_RESPONSE_GUIDED", "library exposure changed")
    _require(development["publisher_exact_censor_universe_status"] == "NOT_CLAIMED_DEVELOPMENT_RULE_EXPLICIT", "censor claim changed")
    credit = config["credit_policy"]
    _require(credit["measured_candidate"] is True and credit["generated_candidate"] is False, "candidate role changed")
    _require(not any(credit["qualified_credit_delta"].values()), "conversion increases qualified credit")
    _require(credit["qualified_counts_after_conversion"] == {"ordinary": 1, "a1": 1, "true_a2": 0, "canonical_records": 6547}, "qualified facts changed")
    output = config["output"]
    _require(output["overwrite_allowed"] is False and output["public_redistribution_allowed"] is False, "output policy changed")
    _require(output["directory"].startswith("/mnt/cunyuliu/mrna_xeditflow_routea_v3/route2/"), "output leaves Route 2 root")
    _require(config["scientific_claim_status"] == "NOT_ESTABLISHED", "scientific claim overstated")


def _positive_integer(value: Any, label: str) -> int:
    token = str(value)
    _require(re.fullmatch(r"[1-9][0-9]*", token) is not None, f"{label} is not a positive integer")
    return int(token)


def _nonnegative_integer(value: str, label: str) -> int:
    _require(re.fullmatch(r"(?:0|[1-9][0-9]*)", value) is not None, f"{label} is not a nonnegative integer")
    return int(value)


def _edit_bin(value: int) -> str:
    if value <= 2:
        return str(value)
    if value <= 5:
        return "3-5"
    if value <= 10:
        return "6-10"
    return "11+"


def _load_library(path: Path, sheet_name: str) -> tuple[dict[str, Any], Counter[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    _require(sheet_name in workbook.sheetnames, "publisher library sheet absent")
    rows = workbook[sheet_name].iter_rows(values_only=True)
    header = tuple(next(rows))
    _require(header == LIBRARY_HEADER, "publisher library header differs")
    barcode_to_design: dict[str, tuple[str, str, str]] = {}
    design_sequences: dict[tuple[str, str, str], set[str]] = defaultdict(set)
    design_metadata: dict[tuple[str, str, str], tuple[str, str, int]] = {}
    design_observed: Counter[tuple[str, str, str]] = Counter()
    family_designs: dict[tuple[str, str], set[tuple[str, str, str]]] = defaultdict(set)
    family_sources: dict[tuple[str, str], set[str]] = defaultdict(set)
    stats: Counter[str] = Counter()
    for values in rows:
        row = dict(zip(header, values))
        stats["library_member_count"] += 1
        full_sequence = str(row["barcoded_seq_184bp"]).upper()
        _require(len(full_sequence) == 184 and not (set(full_sequence) - BASES), "publisher construct sequence differs")
        barcode, sequence = full_sequence[:20], full_sequence[20:]
        family = (str(row["gene_id"]), str(row["pas_id"]))
        design = family + (str(row["experiment"]),)
        _require(barcode not in barcode_to_design, "publisher barcode is duplicated")
        barcode_to_design[barcode] = design
        design_sequences[design].add(sequence)
        metadata = (str(row["type"]), str(row["subtype"]), _positive_integer(row["n_bc"], "publisher n_bc"))
        previous = design_metadata.setdefault(design, metadata)
        _require(previous == metadata, "publisher design metadata differs within design")
        design_observed[design] += 1
        family_designs[family].add(design)
        if design[-1] == "wt":
            family_sources[family].add(sequence)
    for design, sequences in design_sequences.items():
        _require(len(sequences) == 1, f"publisher design has multiple constructs: {design}")
        if design_observed[design] != design_metadata[design][2]:
            stats["declared_multiplicity_mismatch_design_count"] += 1

    family_candidates = {
        family: {design for design in designs if design[-1] != "wt"}
        for family, designs in family_designs.items()
    }
    eligible_families = {
        family for family in family_designs
        if len(family_sources[family]) == 1 and len(family_candidates[family]) >= 3
    }
    pairwise_families = {
        family for family in family_designs
        if len(family_sources[family]) == 1 and len(family_candidates[family]) == 2
    }
    edit_histogram: Counter[str] = Counter()
    sequence_pairs: set[tuple[str, str]] = set()
    for family in eligible_families:
        source = next(iter(family_sources[family]))
        for design in family_candidates[family]:
            candidate = next(iter(design_sequences[design]))
            _require(len(source) == len(candidate) == 164, "source-candidate length differs")
            edit_count = sum(left != right for left, right in zip(source, candidate))
            _require(edit_count > 0, "zero-edit candidate appeared")
            edit_histogram[_edit_bin(edit_count)] += 1
            sequence_pairs.add((source, candidate))
    stats.update({
        "family_count": len(family_designs),
        "eligible_dense_family_count": len(eligible_families),
        "pairwise_family_exclusion_count": len(pairwise_families),
        "eligible_candidate_design_count": sum(len(family_candidates[family]) for family in eligible_families),
        "distinct_eligible_sequence_pair_count": len(sequence_pairs),
    })
    library = {
        "barcode_to_design": barcode_to_design,
        "design_sequences": {design: next(iter(sequences)) for design, sequences in design_sequences.items()},
        "design_metadata": design_metadata,
        "family_candidates": family_candidates,
        "family_sources": family_sources,
        "eligible_families": eligible_families,
        "edit_histogram": dict(sorted(edit_histogram.items())),
    }
    return library, stats


def _endpoint_formula_matches(total: int, distal: int, proximal: int, reported: str) -> bool:
    if total != distal + proximal:
        return False
    if distal == 0 and proximal == 0:
        return reported == "NA"
    if distal == 0:
        return reported == "Inf"
    if proximal == 0:
        return reported == "-Inf"
    try:
        observed = float(reported)
    except ValueError:
        return False
    return math.isfinite(observed) and math.isclose(observed, math.log2(proximal / distal), rel_tol=0.0, abs_tol=1e-12)


def _load_measurements(path: Path, library: Mapping[str, Any], config: Mapping[str, Any]) -> tuple[dict[tuple[str, ...], tuple[int, int, int]], Counter[str]]:
    barcodes = library["barcode_to_design"]
    metadata = library["design_metadata"]
    study = config["study"]
    contexts = {
        (sample, distal) for sample in EXPECTED_SAMPLE_FIELDS for distal in study["distal_reporter_contexts"]
    }
    seen_contexts: dict[str, set[tuple[str, str]]] = defaultdict(set)
    pooled: dict[tuple[str, ...], list[int]] = defaultdict(lambda: [0, 0, 0])
    stats: Counter[str] = Counter()
    with gzip.open(path, "rt", encoding="utf-8", newline="") as handle:
        header = tuple(handle.readline().split())
        _require(header == MPRA_HEADER, "processed MPRA header differs")
        for line in handle:
            values = line.split()
            _require(len(values) == len(MPRA_HEADER), "processed MPRA row width differs")
            row = dict(zip(MPRA_HEADER, values))
            stats["processed_measurement_row_count"] += 1
            design = barcodes.get(row["barcode"])
            _require(design is not None, "processed MPRA barcode is outside publisher library")
            _require(design[:2] == (row["gene_id"], row["pas_id"]) and design[-1] == row["experiment"], "processed MPRA design join differs")
            _require(metadata[design][0] == row["aim"], "processed MPRA design aim differs")
            if metadata[design][1] != row["subaim"]:
                stats["processed_subaim_label_mismatch_row_count"] += 1
            expected_sample = EXPECTED_SAMPLE_FIELDS.get(row["sample"])
            _require(expected_sample == (row["replicate"], row["perturbation"]), "processed MPRA sample context differs")
            context = (row["sample"], row["distal_site"])
            _require(context in contexts, "processed MPRA distal context differs")
            _require(context not in seen_contexts[row["barcode"]], "processed MPRA member context duplicated")
            seen_contexts[row["barcode"]].add(context)
            total = _nonnegative_integer(row["total"], "processed total")
            distal = _nonnegative_integer(row["distal"], "processed distal")
            proximal = _nonnegative_integer(row["proximal"], "processed proximal")
            if not _endpoint_formula_matches(total, distal, proximal, row["log_odds"]):
                stats["formula_mismatch_row_count"] += 1
            key = design + (row["perturbation"], row["distal_site"], row["replicate"])
            counts = pooled[key]
            counts[0] += total
            counts[1] += distal
            counts[2] += proximal
    stats["complete_context_member_count"] = sum(seen_contexts[barcode] == contexts for barcode in barcodes)
    stats["incomplete_context_member_count"] = len(barcodes) - stats["complete_context_member_count"]
    return {key: tuple(value) for key, value in pooled.items()}, stats


def _finite_endpoint(counts: tuple[int, int, int], minimum_total: int) -> float | None:
    total, distal, proximal = counts
    if total < minimum_total or distal == 0 or proximal == 0:
        return None
    _require(total == distal + proximal, "pooled count equation differs")
    return math.log2(proximal / distal)


def _eligible_units(config: Mapping[str, Any], library: Mapping[str, Any], pooled: Mapping[tuple[str, ...], tuple[int, int, int]]) -> tuple[list[dict[str, Any]], Counter[str]]:
    study = config["study"]
    endpoint = config["endpoint_policy"]
    minimum_total = endpoint["minimum_pooled_total_umi_per_replicate"]
    design_sequences = library["design_sequences"]
    design_metadata = library["design_metadata"]
    family_candidates = library["family_candidates"]
    family_sources = library["family_sources"]
    stats: Counter[str] = Counter()
    units: list[dict[str, Any]] = []
    for family in sorted(library["eligible_families"]):
        source_design = family + ("wt",)
        source = next(iter(family_sources[family]))
        for candidate_design in sorted(family_candidates[family]):
            candidate = design_sequences[candidate_design]
            changes = [index for index, (left, right) in enumerate(zip(source, candidate)) if left != right]
            _require(changes, "zero-edit candidate appeared during eligibility")
            for perturbation in study["biological_context_perturbations"]:
                for distal_context in study["distal_reporter_contexts"]:
                    stats["candidate_context_universe_count"] += 1
                    source_values: list[float] = []
                    candidate_values: list[float] = []
                    for replicate in endpoint["biological_replicates"]:
                        source_value = _finite_endpoint(pooled.get(source_design + (perturbation, distal_context, replicate), (0, 0, 0)), minimum_total)
                        candidate_value = _finite_endpoint(pooled.get(candidate_design + (perturbation, distal_context, replicate), (0, 0, 0)), minimum_total)
                        if source_value is not None:
                            source_values.append(source_value)
                        if candidate_value is not None:
                            candidate_values.append(candidate_value)
                    if len(candidate_values) != 2:
                        stats["reject_candidate_endpoint_qc_count"] += 1
                        continue
                    if len(source_values) != 2:
                        stats["reject_source_endpoint_qc_after_candidate_pass_count"] += 1
                        continue
                    replicate_deltas = [candidate_values[index] - source_values[index] for index in range(2)]
                    source_mean = sum(source_values) / 2.0
                    candidate_mean = sum(candidate_values) / 2.0
                    delta_mean = sum(replicate_deltas) / 2.0
                    standard_error = math.sqrt(sum((value - delta_mean) ** 2 for value in replicate_deltas)) / math.sqrt(2.0)
                    _require(all(math.isfinite(value) for value in (source_mean, candidate_mean, delta_mean, standard_error)), "nonfinite eligible endpoint appeared")
                    units.append({
                        "family": family,
                        "candidate_design": candidate_design,
                        "source_sequence": source,
                        "candidate_sequence": candidate,
                        "changes": changes,
                        "perturbation": perturbation,
                        "distal_context": distal_context,
                        "source_endpoint_value": source_mean,
                        "candidate_endpoint_value": candidate_mean,
                        "delta": delta_mean,
                        "standard_error": standard_error,
                        "type": design_metadata[candidate_design][0],
                        "subtype": design_metadata[candidate_design][1],
                    })
                    stats["canonical_record_count"] += 1
                    stats[f"context::{perturbation}|{distal_context}"] += 1
    stats["source_family_with_record_count"] = len({unit["family"] for unit in units})
    return units, stats


def _canonical_records(config: Mapping[str, Any], units: list[dict[str, Any]]) -> list[dict[str, Any]]:
    study = config["study"]
    development = config["development_policy"]
    credit = config["credit_policy"]
    records: list[dict[str, Any]] = []
    for index, unit in enumerate(units, start=1):
        gene_id, pas_id = unit["family"]
        experiment = unit["candidate_design"][-1]
        source_id = f"GSE269595:{gene_id}:{pas_id}:source"
        candidate_id = f"GSE269595:{gene_id}:{pas_id}:{experiment}"
        edit_operations = [
            {
                "type": "SUB",
                "position_zero_based": position,
                "ref": unit["source_sequence"][position],
                "alt": unit["candidate_sequence"][position],
            }
            for position in unit["changes"]
        ]
        records.append({
            "schema_version": CANONICAL_SCHEMA_VERSION,
            "canonical_record_id": f"GSE269595:record:{index:05d}",
            "study_unit_id": "GSE269595",
            "pool_assignment": study["pool_assignment"],
            "qualification_class": study["qualification_class"],
            "study_role": study["study_role"],
            "conversion_scope": study["conversion_scope"],
            "region": study["region"],
            "biological_context_id": f"PERTURBATION_{unit['perturbation']}__DISTAL_{unit['distal_context']}",
            "perturbation_id": unit["perturbation"],
            "distal_reporter_context_id": unit["distal_context"],
            "assay_id": study["assay_id"],
            "endpoint_id": study["endpoint_id"],
            "endpoint_direction": study["endpoint_direction"],
            "source_id": source_id,
            "source_sequence": unit["source_sequence"],
            "source_endpoint_value": unit["source_endpoint_value"],
            "candidate_id": candidate_id,
            "candidate_sequence": unit["candidate_sequence"],
            "candidate_endpoint_value": unit["candidate_endpoint_value"],
            "edit_operations": edit_operations,
            "edit_count": len(edit_operations),
            "direction_normalized_delta": unit["delta"],
            "biological_standard_error": unit["standard_error"],
            "standard_error_status": "DERIVED_FROM_TWO_PAIRED_BIOLOGICAL_REPLICATE_DELTAS_GUIDE_ARMS_POOLED",
            "published_biological_replicate_count": 2,
            "effective_n": 2,
            "group_id": source_id,
            "source_family_id": source_id,
            "gene_id": gene_id,
            "pas_id": pas_id,
            "candidate_experiment": experiment,
            "candidate_design_type": unit["type"],
            "candidate_design_subtype": unit["subtype"],
            "multi_step_sub_trajectory": True,
            "same_position_repeated_edit_required": False,
            "measured_candidate": credit["measured_candidate"],
            "generated_candidate": credit["generated_candidate"],
            "training_eligible": development["training_eligible"],
            "model_selection_eligible": development["model_selection_eligible"],
            "confirmatory_evaluation_eligible": development["confirmatory_evaluation_eligible"],
            "unseen_evaluation_eligible": development["unseen_evaluation_eligible"],
            "library_selection_exposure": development["library_selection_exposure"],
            "publisher_exact_censor_universe_status": development["publisher_exact_censor_universe_status"],
            "near_duplicate_split_status": development["near_duplicate_split_status"],
        })
    return records


def _write_json(path: Path, value: Mapping[str, Any]) -> None:
    path.write_text(json.dumps(value, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def execute(config: Mapping[str, Any], workbook_path: Path, mpra_path: Path, output_dir: Path) -> dict[str, Any]:
    _require(not output_dir.exists(), f"output directory already exists: {output_dir}")
    _require(workbook_path.is_file(), f"publisher library workbook absent: {workbook_path}")
    _require(mpra_path.is_file(), f"processed MPRA table absent: {mpra_path}")
    expected_input = config["input"]
    library, library_stats = _load_library(workbook_path, expected_input["required_library_sheet"])
    for key in (
        "library_member_count", "family_count", "eligible_dense_family_count",
        "pairwise_family_exclusion_count", "eligible_candidate_design_count",
        "distinct_eligible_sequence_pair_count",
    ):
        _require(library_stats[key] == expected_input[f"expected_{key}"], f"{key} differs")
    _require(library["edit_histogram"] == expected_input["expected_eligible_edit_count_histogram"], "eligible edit histogram differs")

    pooled, measurement_stats = _load_measurements(mpra_path, library, config)
    _require(measurement_stats["processed_measurement_row_count"] == expected_input["expected_processed_measurement_row_count"], "processed measurement count differs")
    _require(measurement_stats["complete_context_member_count"] == expected_input["expected_complete_context_member_count"], "complete member context count differs")
    _require(measurement_stats["incomplete_context_member_count"] == 0, "incomplete member context appeared")
    _require(measurement_stats["processed_subaim_label_mismatch_row_count"] == expected_input["expected_processed_subaim_label_mismatch_row_count"], "processed subaim label mismatch count differs")
    _require(measurement_stats["formula_mismatch_row_count"] == expected_input["expected_formula_mismatch_row_count"], "endpoint formula mismatch count differs")

    units, eligibility_stats = _eligible_units(config, library, pooled)
    expected = config["expected_conversion"]
    for key in (
        "candidate_context_universe_count", "reject_candidate_endpoint_qc_count",
        "reject_source_endpoint_qc_after_candidate_pass_count", "canonical_record_count",
        "source_family_with_record_count",
    ):
        _require(eligibility_stats[key] == expected[key], f"{key} differs")
    observed_context_counts = {
        key.removeprefix("context::"): value
        for key, value in eligibility_stats.items() if key.startswith("context::")
    }
    _require(observed_context_counts == expected["context_record_counts"], "context record counts differ")
    _require(expected["candidate_context_universe_count"] == expected["reject_candidate_endpoint_qc_count"] + expected["reject_source_endpoint_qc_after_candidate_pass_count"] + expected["canonical_record_count"], "candidate-context closure differs")
    records = _canonical_records(config, units)
    _require(len(records) == expected["canonical_record_count"], "canonical materialization count differs")
    _require(len({record["canonical_record_id"] for record in records}) == len(records), "canonical record IDs are not unique")

    output = config["output"]
    output_dir.parent.mkdir(parents=True, exist_ok=True)
    temporary = Path(tempfile.mkdtemp(prefix=f".{output_dir.name}.", dir=output_dir.parent))
    try:
        with (temporary / output["canonical_filename"]).open("w", encoding="utf-8") as handle:
            for record in records:
                handle.write(json.dumps(record, sort_keys=True) + "\n")
        summary = {
            "status": "PASS_GSE269595_DEVELOPMENT_RELAXED_TRUE_A2_LISTWISE_EXPOSED_LIBRARY",
            "study_unit_id": "GSE269595",
            "canonical_record_count": len(records),
            "source_family_with_record_count": eligibility_stats["source_family_with_record_count"],
            "eligible_dense_family_count": library_stats["eligible_dense_family_count"],
            "eligible_candidate_design_count": library_stats["eligible_candidate_design_count"],
            "context_record_counts": dict(sorted(observed_context_counts.items())),
            "edit_count_histogram": library["edit_histogram"],
            "library_selection_exposure": config["development_policy"]["library_selection_exposure"],
            "publisher_exact_censor_universe_status": config["development_policy"]["publisher_exact_censor_universe_status"],
            "qualified_credit_delta": config["credit_policy"]["qualified_credit_delta"],
            "qualified_counts_after_conversion": config["credit_policy"]["qualified_counts_after_conversion"],
            "scientific_claim_status": config["scientific_claim_status"],
            "limitations": config["limitations"],
        }
        rejects = {
            "library": dict(sorted(library_stats.items())),
            "measurement": dict(sorted(measurement_stats.items())),
            "eligibility": dict(sorted(eligibility_stats.items())),
            "candidate_context_universe_closed": True,
            "missing_is_zero": False,
            "non_sub_actions_retained": 0,
            "generated_candidates_retained": 0,
            "publisher_exact_censor_universe_claimed": False,
        }
        _write_json(temporary / output["conversion_summary_filename"], summary)
        _write_json(temporary / output["reject_summary_filename"], rejects)
        os.replace(temporary, output_dir)
    except Exception:
        shutil.rmtree(temporary, ignore_errors=True)
        raise
    return summary


def main() -> int:
    parser = argparse.ArgumentParser(description="Convert GSE269595 dense MPRA source families")
    parser.add_argument("--config", type=Path, default=CONFIG_PATH)
    parser.add_argument("--library-workbook", type=Path)
    parser.add_argument("--processed-mpra", type=Path)
    parser.add_argument("--output-dir", type=Path)
    args = parser.parse_args()
    config = load_config(args.config)
    summary = execute(
        config,
        args.library_workbook or Path(config["input"]["library_workbook_path"]),
        args.processed_mpra or Path(config["input"]["processed_mpra_path"]),
        args.output_dir or Path(config["output"]["directory"]),
    )
    print(json.dumps(summary, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
