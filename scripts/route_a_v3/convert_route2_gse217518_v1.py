#!/usr/bin/env python3
"""Reconstruct strict 115 bp GSE217518 SUB records for Route 2 Development."""

from __future__ import annotations

import argparse
import importlib.util
import json
import math
import os
import re
import shutil
import tempfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
CONFIG_PATH = REPO_ROOT / "configs/route_a_v3_route2_gse217518_converter_v1.json"
RECONSTRUCTION_MODULE_PATH = REPO_ROOT / "d1_staging/scripts/d1/reconstruct_gse217518_sequences.py"
CANONICAL_SCHEMA_VERSION = "mrna_editflow_route2_canonical.v1"
BASES = set("ACGT")


class ConversionError(RuntimeError):
    pass


def _require(condition: bool, message: str) -> None:
    if not condition:
        raise ConversionError(message)


def load_config(path: Path = CONFIG_PATH) -> dict[str, Any]:
    config = json.loads(path.read_text(encoding="utf-8"))
    validate_config(config)
    return config


def validate_config(config: Mapping[str, Any]) -> None:
    _require(config["schema_version"] == "route_a_v3_route2_gse217518_converter.v1", "unexpected schema version")
    study = config["study"]
    _require(study["study_unit_id"] == "GSE217518", "unexpected study")
    _require(study["pool_assignment"] == "DEVELOPMENT", "GSE217518 is not Development")
    _require(study["qualification_class"] == "DEVELOPMENT_RELAXED", "qualification class changed")
    _require(study["study_role"] == "A1_STRICT_115BP_RECONSTRUCTED_SUB_ONLY", "study role changed")
    _require(study["conversion_scope"] == "STRICT_PUBLIC_RECONSTRUCTION_SUB_ONLY_PARTIAL", "partial scope overstated")
    expected = config["input"]
    _require(expected["expected_supplement_row_count"] == 5072, "supplement row count changed")
    _require(expected["expected_strict_gc_valid_construct_row_count"] == 2761, "strict construct count changed")
    _require(expected["expected_sub_construct_row_count"] == 2759, "SUB construct count changed")
    _require(expected["expected_canonical_record_count"] == 4009, "canonical record count changed")
    _require(expected["construct_fragment_length"] == 115 and expected["variant_center_index_zero_based"] == 57, "construct geometry changed")
    primers = expected["primer_sequences"]
    _require(primers == {
        "5UTR": ["CGCTAGGGATCCTCTAGTCA", "ACCGGTCGCCACCATGGTGA"],
        "3UTR": ["GGACGAGCTGTACAAGTAAA", "GCGGCCGCGCAATAACTAGC"],
    }, "publisher primer sequences changed")
    _require(all(len(primer) == 20 and not (set(primer) - BASES) for pair in primers.values() for primer in pair), "invalid primer geometry")
    action = config["action_policy"]
    _require(action == {"allowed_candidate_action": "SUB", "ins_supported": False, "del_supported": False}, "action policy changed")
    output = config["output"]
    _require(output["overwrite_allowed"] is False, "successful output overwrite enabled")
    _require(output["public_redistribution_allowed"] is True, "closed public rights changed")
    _require(output["directory"].startswith("/mnt/cunyuliu/mrna_xeditflow_routea_v3/route2/"), "output leaves Route 2 root")
    development = config["development_policy"]
    _require(development["training_eligible"] is True and development["model_selection_eligible"] is True, "Development disabled")
    _require(development["near_duplicate_split_status"] == "NOT_RUN", "near-duplicate split overstated")
    _require(development["missing_standard_error_representation"] is None, "missing SE is not null")
    _require(development["missing_endpoint_is_zero"] is False, "missing endpoint changed to zero")
    _require(development["conflicting_exact_units_action"] == "REJECT_ALL_CONFLICTING_ROWS", "conflict action changed")
    credit = config["credit_policy"]
    _require(credit["measured_candidate"] is True and credit["generated_candidate"] is False, "candidate measurement role changed")
    _require(not any(credit["qualified_credit_delta"].values()), "Development conversion increases qualified credit")
    _require(credit["qualified_counts_after_conversion"] == {"ordinary": 1, "a1": 1, "true_a2": 0, "canonical_records": 6547}, "qualified facts changed")
    _require(config["scientific_claim_status"] == "NOT_ESTABLISHED", "scientific claim overstated")


def _load_reconstruction_module():
    _require(RECONSTRUCTION_MODULE_PATH.is_file(), f"reconstruction helper absent: {RECONSTRUCTION_MODULE_PATH}")
    spec = importlib.util.spec_from_file_location("route2_gse217518_reconstruction_helper", RECONSTRUCTION_MODULE_PATH)
    _require(spec is not None and spec.loader is not None, "cannot load reconstruction helper")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _finite(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(value)


def _load_transcripts(cache_path: Path, module) -> tuple[dict[str, dict[str, Any]], int]:
    content = cache_path.read_text(encoding="utf-8")
    genbank: dict[str, str] = {}
    for record in re.split(r"^//\s*$", content, flags=re.MULTILINE):
        match = re.match(r"^LOCUS\s+(\S+)", record.strip())
        if match:
            genbank[match.group(1)] = record.strip()
    transcripts: dict[str, dict[str, Any]] = {}
    for accession, text in genbank.items():
        cds = module.parse_cds_coordinates(text)
        sequence = module.parse_mrna_sequence(text)
        if cds and sequence and not (set(sequence) - BASES):
            cds_start, cds_end = cds
            transcripts[accession] = {
                "cds_start": cds_start,
                "cds_end": cds_end,
                "5utr": sequence[: cds_start - 1],
                "3utr": sequence[cds_end:],
            }
    return transcripts, len(genbank)


def _find_accession(variant_name: str, transcripts: Mapping[str, Any]) -> str | None:
    match = re.match(r"([A-Z]+_\d+\.\d+)", variant_name or "")
    if not match:
        return None
    accession = match.group(1)
    base = accession.split(".")[0]
    for candidate in transcripts:
        if candidate == accession or candidate.startswith(base):
            return candidate
    return None


def _crop_construct(source: str, candidate: str, fragment_length: int, center_index: int) -> tuple[str, str, int] | None:
    prefix = 0
    for left, right in zip(source, candidate):
        if left != right:
            break
        prefix += 1
    if prefix == min(len(source), len(candidate)):
        return None
    start = max(0, min(prefix - center_index, max(0, len(source) - fragment_length)))
    end = min(len(source), start + fragment_length)
    length_delta = len(candidate) - len(source)
    return source[start:end], candidate[start : end + length_delta], start


def _gc_construct_exact(fragment: str, fraction: Any, primers: Iterable[str]) -> tuple[bool, str]:
    if not _finite(fraction):
        return False, "DENOMINATOR"
    primers = tuple(primers)
    total_length = len(fragment) + sum(len(primer) for primer in primers)
    scaled = float(fraction) * total_length
    if abs(scaled - round(scaled)) > 1e-8:
        return False, "DENOMINATOR"
    expected_gc = sum(base in "GC" for base in fragment) + sum(base in "GC" for primer in primers for base in primer)
    return (int(round(scaled)) == expected_gc, "COMPOSITION")


def _read_supplement(path: Path, sheet: str) -> list[tuple[int, dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    _require(sheet in workbook.sheetnames, f"supplement sheet absent: {sheet}")
    worksheet = workbook[sheet]
    iterator = worksheet.iter_rows(values_only=True)
    header = list(next(iterator))
    required = {
        "Mutant", "variant_name", "GeneSymbol", "UTR_Group", "t05_WT_SH", "t05_mt_SH",
        "t05_WT_HEK", "t05_mt_HEK", "GCcontent_WT", "GCcontent_mt",
    }
    _require(required.issubset(header), f"supplement columns absent: {sorted(required - set(header))}")
    return [(row_number, dict(zip(header, values))) for row_number, values in enumerate(iterator, start=2)]


def _strict_sub_constructs(
    config: Mapping[str, Any], rows: list[tuple[int, dict[str, Any]]], transcripts: Mapping[str, Any], module
) -> tuple[list[dict[str, Any]], Counter[str], int]:
    input_config = config["input"]
    rejection_counts: Counter[str] = Counter()
    strict_gc_valid_count = 0
    accepted: list[dict[str, Any]] = []
    for row_number, row in rows:
        variant = module.parse_hgvs_c(row.get("variant_name") or "")
        if not variant or variant.get("position_type") not in ("5utr", "3utr"):
            rejection_counts["UNPARSEABLE_OR_NON_UTR"] += 1
            continue
        expected_position_type = "5utr" if row.get("UTR_Group") == "5'UTR" else "3utr" if row.get("UTR_Group") == "3'UTR" else None
        if variant["position_type"] != expected_position_type:
            rejection_counts["HGVS_SUPPLEMENT_REGION_MISMATCH"] += 1
            continue
        accession = _find_accession(row["variant_name"], transcripts)
        if accession is None:
            rejection_counts["TRANSCRIPT_OR_VARIANT_APPLICATION_FAILURE"] += 1
            continue
        transcript = transcripts[accession]
        applied = module.apply_variant_to_utr(
            transcript[expected_position_type], variant, transcript["cds_start"], transcript["cds_end"]
        )
        if not applied:
            rejection_counts["TRANSCRIPT_OR_VARIANT_APPLICATION_FAILURE"] += 1
            continue
        source, candidate, _ = applied
        cropped = _crop_construct(
            source, candidate, input_config["construct_fragment_length"], input_config["variant_center_index_zero_based"]
        )
        if not cropped:
            rejection_counts["TRANSCRIPT_OR_VARIANT_APPLICATION_FAILURE"] += 1
            continue
        source_fragment, candidate_fragment, window_start = cropped
        region = "5UTR" if expected_position_type == "5utr" else "3UTR"
        primers = input_config["primer_sequences"][region]
        source_gc, source_reason = _gc_construct_exact(source_fragment, row.get("GCcontent_WT"), primers)
        candidate_gc, candidate_reason = _gc_construct_exact(candidate_fragment, row.get("GCcontent_mt"), primers)
        if source_reason == "DENOMINATOR" or candidate_reason == "DENOMINATOR":
            rejection_counts["GC_FRACTION_DENOMINATOR_MISMATCH"] += 1
            continue
        if not source_gc or not candidate_gc:
            rejection_counts["CONSTRUCT_GC_COMPOSITION_MISMATCH"] += 1
            continue
        strict_gc_valid_count += 1
        if (
            variant.get("var_type") != "snv"
            or len(source_fragment) != len(candidate_fragment)
            or sum(left != right for left, right in zip(source_fragment, candidate_fragment)) != 1
        ):
            rejection_counts["UNSUPPORTED_ACTION_NOT_SUB"] += 1
            continue
        edit_position = next(index for index, (left, right) in enumerate(zip(source_fragment, candidate_fragment)) if left != right)
        accepted.append({
            "supplement_row_number": row_number,
            "mutant_identifier": row["Mutant"],
            "variant_name": row["variant_name"],
            "gene": row["GeneSymbol"],
            "transcript_accession": accession,
            "region": region,
            "source_sequence": source_fragment,
            "candidate_sequence": candidate_fragment,
            "window_start_zero_based": window_start,
            "edit_position_zero_based": edit_position,
            "edit_ref": source_fragment[edit_position],
            "edit_alt": candidate_fragment[edit_position],
            "t05_WT_SH": row.get("t05_WT_SH"),
            "t05_mt_SH": row.get("t05_mt_SH"),
            "t05_WT_HEK": row.get("t05_WT_HEK"),
            "t05_mt_HEK": row.get("t05_mt_HEK"),
        })
    return accepted, rejection_counts, strict_gc_valid_count


def _endpoint_units(constructs: list[dict[str, Any]]) -> tuple[dict[tuple[str, str, str], list[dict[str, Any]]], int]:
    units: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    endpoint_slot_count = 0
    for construct in constructs:
        for context, wt_field, mt_field in (
            ("SH_SY5Y", "t05_WT_SH", "t05_mt_SH"),
            ("HEK293T", "t05_WT_HEK", "t05_mt_HEK"),
        ):
            wt = construct[wt_field]
            candidate = construct[mt_field]
            if not (_finite(wt) and _finite(candidate)):
                continue
            endpoint_slot_count += 1
            value = dict(construct)
            value.update({
                "biological_context_id": context,
                "source_endpoint_value": float(wt),
                "candidate_endpoint_value": float(candidate),
                "direction_normalized_delta": float(candidate) - float(wt),
            })
            units[(construct["source_sequence"], construct["candidate_sequence"], context)].append(value)
    return units, endpoint_slot_count


def _resolve_units(units: Mapping[tuple[str, str, str], list[dict[str, Any]]]) -> tuple[list[dict[str, Any]], int, int, int]:
    resolved: list[dict[str, Any]] = []
    conflict_unit_count = 0
    conflict_row_count = 0
    exact_duplicate_unit_count = 0
    for key in sorted(units):
        values = units[key]
        deltas = {round(value["direction_normalized_delta"], 12) for value in values}
        if len(values) > 1 and len(deltas) > 1:
            conflict_unit_count += 1
            conflict_row_count += len(values)
            continue
        if len(values) > 1:
            exact_duplicate_unit_count += 1
        resolved.append(sorted(values, key=lambda value: value["supplement_row_number"])[0])
    return resolved, conflict_unit_count, conflict_row_count, exact_duplicate_unit_count


def _canonical_records(config: Mapping[str, Any], resolved: list[dict[str, Any]]) -> list[dict[str, Any]]:
    source_group_ids = {
        source: f"GSE217518_SOURCE_GROUP_{index:05d}"
        for index, source in enumerate(sorted({value["source_sequence"] for value in resolved}), start=1)
    }
    development = config["development_policy"]
    credit = config["credit_policy"]
    records = []
    for value in sorted(resolved, key=lambda item: (item["supplement_row_number"], item["biological_context_id"])):
        source_group_id = source_group_ids[value["source_sequence"]]
        records.append({
            "schema_version": CANONICAL_SCHEMA_VERSION,
            "canonical_record_id": f"GSE217518:row:{value['supplement_row_number']:05d}:context:{value['biological_context_id']}",
            "study_unit_id": "GSE217518",
            "pool_assignment": config["study"]["pool_assignment"],
            "qualification_class": config["study"]["qualification_class"],
            "study_role": config["study"]["study_role"],
            "conversion_scope": config["study"]["conversion_scope"],
            "region": value["region"],
            "biological_context_id": value["biological_context_id"],
            "assay_id": config["study"]["assay_id"],
            "endpoint_id": config["study"]["endpoint_id"],
            "endpoint_direction": config["study"]["endpoint_direction"],
            "source_id": source_group_id,
            "source_sequence": value["source_sequence"],
            "source_endpoint_value": value["source_endpoint_value"],
            "candidate_id": f"GSE217518:row:{value['supplement_row_number']:05d}:candidate",
            "candidate_sequence": value["candidate_sequence"],
            "candidate_endpoint_value": value["candidate_endpoint_value"],
            "edit_operations": [{
                "type": "SUB",
                "position_zero_based": value["edit_position_zero_based"],
                "ref": value["edit_ref"],
                "alt": value["edit_alt"],
            }],
            "direction_normalized_delta": value["direction_normalized_delta"],
            "biological_standard_error": development["missing_standard_error_representation"],
            "published_biological_experiment_count": 3,
            "effective_n": None,
            "group_id": source_group_id,
            "study_group_id": "GSE217518_SU_WANG_2025",
            "gene": value["gene"],
            "transcript_accession": value["transcript_accession"],
            "supplement_row_number": value["supplement_row_number"],
            "mutant_identifier": value["mutant_identifier"],
            "variant_name": value["variant_name"],
            "window_start_zero_based": value["window_start_zero_based"],
            "measured_candidate": credit["measured_candidate"],
            "generated_candidate": credit["generated_candidate"],
            "training_eligible": development["training_eligible"],
            "model_selection_eligible": development["model_selection_eligible"],
            "near_duplicate_split_status": development["near_duplicate_split_status"],
        })
    return records


def _write_json(path: Path, value: Mapping[str, Any]) -> None:
    path.write_text(json.dumps(value, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def execute(config: Mapping[str, Any], supplement_path: Path, cache_path: Path, output_dir: Path) -> dict[str, Any]:
    _require(not output_dir.exists(), f"output directory already exists: {output_dir}")
    _require(supplement_path.is_file(), f"supplement absent: {supplement_path}")
    _require(cache_path.is_file(), f"GenBank cache absent: {cache_path}")
    module = _load_reconstruction_module()
    rows = _read_supplement(supplement_path, config["input"]["supplement_sheet"])
    transcripts, genbank_record_count = _load_transcripts(cache_path, module)
    constructs, rejection_counts, strict_gc_valid_count = _strict_sub_constructs(config, rows, transcripts, module)
    units, endpoint_slot_count = _endpoint_units(constructs)
    resolved, conflict_unit_count, conflict_row_count, exact_duplicate_unit_count = _resolve_units(units)
    records = _canonical_records(config, resolved)
    context_counts = Counter(record["biological_context_id"] for record in records)
    region_counts = Counter(record["region"] for record in records)
    distinct_source_count = len({record["source_sequence"] for record in records})
    expected = config["input"]
    observed = {
        "supplement_row_count": len(rows),
        "genbank_record_count": genbank_record_count,
        "parsed_transcript_count": len(transcripts),
        "rejection_counts": dict(sorted(rejection_counts.items())),
        "strict_gc_valid_construct_row_count": strict_gc_valid_count,
        "sub_construct_row_count": len(constructs),
        "nonmissing_endpoint_slot_count": endpoint_slot_count,
        "conflicting_exact_unit_count": conflict_unit_count,
        "conflicting_endpoint_row_count": conflict_row_count,
        "exact_duplicate_unit_count": exact_duplicate_unit_count,
        "canonical_record_count": len(records),
        "context_record_counts": dict(sorted(context_counts.items())),
        "region_record_counts": dict(sorted(region_counts.items())),
        "distinct_source_group_count": distinct_source_count,
    }
    exact = (
        observed["supplement_row_count"] == expected["expected_supplement_row_count"]
        and observed["genbank_record_count"] == expected["expected_genbank_record_count"]
        and observed["parsed_transcript_count"] == expected["expected_parsed_transcript_count"]
        and observed["rejection_counts"] == expected["expected_rejection_counts"]
        and observed["strict_gc_valid_construct_row_count"] == expected["expected_strict_gc_valid_construct_row_count"]
        and observed["sub_construct_row_count"] == expected["expected_sub_construct_row_count"]
        and observed["nonmissing_endpoint_slot_count"] == expected["expected_nonmissing_endpoint_slot_count"]
        and observed["conflicting_exact_unit_count"] == expected["expected_conflicting_exact_unit_count"]
        and observed["conflicting_endpoint_row_count"] == expected["expected_conflicting_endpoint_row_count"]
        and observed["exact_duplicate_unit_count"] == expected["expected_exact_duplicate_unit_count"]
        and observed["canonical_record_count"] == expected["expected_canonical_record_count"]
        and observed["context_record_counts"] == expected["expected_context_record_counts"]
        and observed["region_record_counts"] == expected["expected_region_record_counts"]
        and observed["distinct_source_group_count"] == expected["expected_distinct_source_group_count"]
    )
    status = "CONVERTED_DEVELOPMENT_RELAXED_STRICT_115BP_SUB_ONLY_PARTIAL" if exact else "UNCONVERTIBLE_FOR_ROUTE2_V1_GEOMETRY_MISMATCH"
    output_dir.parent.mkdir(parents=True, exist_ok=True)
    temporary = Path(tempfile.mkdtemp(prefix=f".{output_dir.name}.", dir=output_dir.parent))
    try:
        canonical_path = temporary / config["output"]["canonical_filename"]
        with canonical_path.open("w", encoding="utf-8") as handle:
            if exact:
                for record in records:
                    handle.write(json.dumps(record, sort_keys=True) + "\n")
        summary = {
            "converter_id": config["converter_id"],
            "study_unit_id": "GSE217518",
            "pool_assignment": "DEVELOPMENT",
            "qualification_class": "DEVELOPMENT_RELAXED",
            "conversion_scope": config["study"]["conversion_scope"],
            "status": status,
            **observed,
            "measured_candidate_count": len(records) if exact else 0,
            "generated_candidate_count": 0,
            "biological_standard_error_count": 0,
            "qualified_credit_delta": config["credit_policy"]["qualified_credit_delta"],
            "qualified_counts_after_conversion": config["credit_policy"]["qualified_counts_after_conversion"],
            "public_redistribution_allowed": config["output"]["public_redistribution_allowed"],
            "near_duplicate_split_status": config["development_policy"]["near_duplicate_split_status"],
            "limitations": config["limitations"],
            "scientific_claim_status": config["scientific_claim_status"],
        }
        reject_summary = {
            "converter_id": config["converter_id"],
            "study_unit_id": "GSE217518",
            "construct_rejection_counts": observed["rejection_counts"],
            "missing_endpoint_slot_count_after_sub_construct_acceptance": len(constructs) * 2 - endpoint_slot_count,
            "conflicting_exact_unit_count": conflict_unit_count,
            "conflicting_endpoint_row_count": conflict_row_count,
            "conflicting_exact_units_action": config["development_policy"]["conflicting_exact_units_action"],
            "reject_payload_in_summary": False,
        }
        _write_json(temporary / config["output"]["conversion_summary_filename"], summary)
        _write_json(temporary / config["output"]["reject_summary_filename"], reject_summary)
        os.rename(temporary, output_dir)
        return summary
    except Exception:
        shutil.rmtree(temporary, ignore_errors=True)
        raise


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", type=Path, default=CONFIG_PATH)
    parser.add_argument("--supplement", type=Path)
    parser.add_argument("--genbank-cache", type=Path)
    parser.add_argument("--output-dir", type=Path)
    args = parser.parse_args()
    config = load_config(args.config)
    summary = execute(
        config,
        args.supplement or Path(config["input"]["supplement_path"]),
        args.genbank_cache or Path(config["input"]["genbank_cache_path"]),
        args.output_dir or Path(config["output"]["directory"]),
    )
    print(json.dumps(summary, sort_keys=True))
    return 0 if summary["status"].startswith("CONVERTED_") else 2


if __name__ == "__main__":
    raise SystemExit(main())
