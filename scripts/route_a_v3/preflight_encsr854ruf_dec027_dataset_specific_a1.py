#!/usr/bin/env python3
"""Fail-closed, aggregate-only ENCSR854RUF DEC027 A1 preflight.

The implementation candidate is deliberately inactive.  DEC027 authority, the
settled A1-EVT-059 runtime I1/I2/B2 history, and the GSE217518
I1/I2/B2/I3/B3 predecessor history are byte-bound.  This producer's own
four-scalar binding remains grouped UNKNOWN, so the production path stops
before Git, publisher-workbook, author-repository, or output-path I/O.  A future
bound run verifies the complete append-only commit chain and computes only
aggregate geometry; identifiers, sequences, row values, standard errors, and
split assignments are never serialized.
"""

from __future__ import annotations

import argparse
import copy
import csv
import hashlib
import json
import math
import os
import re
import subprocess
import sys
import tempfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence


SCHEMA_VERSION = "route_a_v3_encsr854ruf_dec027_dataset_specific_a1_preflight.v1"
REPORT_SCHEMA_VERSION = (
    "route_a_v3_encsr854ruf_dec027_dataset_specific_a1_preflight_record.v1"
)
PROTOCOL_ID = "ENCSR854RUF_DEC027_DATASET_SPECIFIC_A1_PREFLIGHT_V1"
DATASET_ID = "ENCSR854RUF"
DECISION_ID = "V3-DEC-027"
UNKNOWN = "UNKNOWN_NOT_ASSERTED"
BOUND = "BOUND"
REPORT_FILENAME = "ENCSR854RUF_DEC027_DATASET_SPECIFIC_A1_PREFLIGHT_RECORD_V1.json"

CONFIG_REPO_PATH = (
    "configs/route_a_v3_encsr854ruf_dec027_dataset_specific_a1_preflight_v1.json"
)
SCRIPT_REPO_PATH = (
    "scripts/route_a_v3/preflight_encsr854ruf_dec027_dataset_specific_a1.py"
)
TEST_REPO_PATH = (
    "tests/route_a_v3/test_preflight_encsr854ruf_dec027_dataset_specific_a1.py"
)
EXACT3 = (CONFIG_REPO_PATH, SCRIPT_REPO_PATH, TEST_REPO_PATH)

PRODUCTION_REPO_ROOT = Path(
    "/home/cunyuliu/mrna_editflow_goal/worktrees/routea_v3_a1_20260810"
)
PRODUCTION_BRANCH = "routea-v3-a1-20260810"
PRODUCTION_UPSTREAM = f"origin/{PRODUCTION_BRANCH}"
AUTHORITY_COMMIT = "3e0ad158a0b45b2f26ed82da3afe60667c712cd6"
AUTHORITY_PARENT = "b1ca33d852bad111ff31b4f60493d8c43c63d1a3"
AUTHORITY_EXACT12 = (
    "configs/route_a_v3.yaml",
    "configs/route_a_v3_a1_qualification.json",
    "docs/contracts/amendments/mrna_xeditflow_route_a_v3_dec027.yaml",
    "docs/contracts/supersession_mrna_xeditflow_v1_1_to_route_a_v3.yaml",
    "docs/execution/route_a_v3_a1_interim.yaml",
    "docs/execution/route_a_v3_a6_interim.yaml",
    "docs/execution/route_a_v3_data_role_registry.yaml",
    "docs/execution/route_a_v3_decision_log.yaml",
    "docs/execution/route_a_v3_registry_manifest.json",
    "docs/execution/route_a_v3_task_registry.yaml",
    "scripts/route_a_v3/validate_a0_bundle.py",
    "tests/route_a_v3/test_a0_integrity_guards.py",
)
AUTHORITY_BLOBS = {
    "configs/route_a_v3.yaml": "c5ec7d236443b506c09fd3f09e149ce5d082daff618887989af6e59472727a27",
    "configs/route_a_v3_a1_qualification.json": "261339c38f4b8bbd48bf8f63f6a588be57af9f6229119e84bf661d7ee8f855db",
    "docs/contracts/amendments/mrna_xeditflow_route_a_v3_dec027.yaml": "2a27c296539e8e665873363778d91cc223f56f933a815e9509b10a7267f6b5c4",
    "docs/contracts/supersession_mrna_xeditflow_v1_1_to_route_a_v3.yaml": "1c4b6e29c09eb24798207138047b68909d7bda8bacc3a2eab8e17a7ca789b44b",
    "docs/execution/route_a_v3_a1_interim.yaml": "fb50929ae2dfa0bdd1c50b003fc43c0e13b89baade203e8384c8b6ea3eba7b1e",
    "docs/execution/route_a_v3_a6_interim.yaml": "1d44bcfe8669a55dc42f619ed43178f0637e20a4297ca996ecab3f7165612769",
    "docs/execution/route_a_v3_data_role_registry.yaml": "80217a8114286f84960237819ac5b2d5828afbc23af118576541cc7cee64ae4e",
    "docs/execution/route_a_v3_decision_log.yaml": "e0dc2a7fb186c5c8d00c1c5604602b1b2f87b26241191d4da55a405a02387e05",
    "docs/execution/route_a_v3_registry_manifest.json": "73a39a566aa0310a80cc83f4eb17ddb95cabc87e5b070ef6484c05178ba32b75",
    "docs/execution/route_a_v3_task_registry.yaml": "a64d0b8bb5eb466b06daa46ed109bd19901ee775910bc5cc9221c39ead63a4bc",
    "scripts/route_a_v3/validate_a0_bundle.py": "81d1a8dc49375f53a1edd5c3f41625e734eacc31a4c328c8137340a041d77e65",
    "tests/route_a_v3/test_a0_integrity_guards.py": "106f847e957a40fffec4c1b57f8f572325ef96e8a3dc5729db68499e320f380b",
}
RUNTIME_CONFIG_PATH = "configs/route_a_v3_dec027_authority_runtime_sync_v1.json"
RUNTIME_SCRIPT_PATH = "scripts/route_a_v3/dec027_authority_runtime_sync.py"
RUNTIME_TEST_PATH = "tests/route_a_v3/test_dec027_authority_runtime_sync.py"
RUNTIME_EXACT3 = (RUNTIME_CONFIG_PATH, RUNTIME_SCRIPT_PATH, RUNTIME_TEST_PATH)
RUNTIME_I1_COMMIT = "de40c58ab81fc06196be3bb9ffb5aa35d39c9d03"
RUNTIME_I2_COMMIT = "5d66e8dc83eb9966f7698ac0fc677f1b06af8ea6"
RUNTIME_B2_COMMIT = "e60956cf59cbddc0406c5d116fb9714906db36e1"
GSE217_CONFIG_PATH = (
    "configs/route_a_v3_gse217518_corrected_a1_successor_candidate_v1.json"
)
GSE217_SCRIPT_PATH = (
    "scripts/route_a_v3/preflight_gse217518_corrected_a1_successor_candidate.py"
)
GSE217_TEST_PATH = (
    "tests/route_a_v3/test_preflight_gse217518_corrected_a1_successor_candidate.py"
)
GSE217_EXACT3 = (GSE217_CONFIG_PATH, GSE217_SCRIPT_PATH, GSE217_TEST_PATH)
GSE217_I1_COMMIT = "17a35f0f88cc988b938aaf25d94a8b32f0cacfc8"
GSE217_I2_COMMIT = "6fbd63be6d0edb9f73cf2f85e446917d3c3ff100"
GSE217_B2_COMMIT = "c3611b0f2e8baeb83422bb07f5446b42edce90ef"
GSE217_I3_COMMIT = "36b535f77b3f27bb872b182dcaf6c646d9781991"
GSE217_B3_COMMIT = "0a46400efee4ead95b1283df73d263f6f8033036"
GSE217_OWN_BINDING_FIELDS = (
    "status",
    "implementation_commit",
    "implementation_script_sha256",
    "implementation_test_sha256",
)
OWN_BINDING_FIELDS = (
    "status",
    "implementation_commit",
    "implementation_script_sha256",
    "implementation_test_sha256",
)

GATE_IDS = (
    "PUBLIC_SOURCE_ASSET_IDENTITY_AND_PRIMARY_ROUTE_CLOSED",
    "SOURCE_REFERENCE_TO_CANDIDATE_CROSSWALK_CLOSED",
    "FULL_REPORTER_AND_THREE_UTR_CONTEXT_CLOSED",
    "ENDPOINT_DIRECTION_SCALE_TRANSFORM_AND_SEMANTICS_CLOSED",
    "INDEPENDENT_BIOLOGICAL_REPLICATE_AND_VALID_STANDARD_ERROR_CLOSED",
    "MISSING_QC_AND_SELECTION_CLOSED",
    "LICENSE_AND_REUSE_RIGHTS_CLOSED",
    "HISTORICAL_ANALYTIC_OR_CHECKPOINT_EXPOSURE_CLOSED",
    "OUTCOME_BLIND_SOURCE_GROUP_NEAR_DUPLICATE_SPLIT_AND_ZERO_LEAKAGE_READINESS_CLOSED",
    "POST_DEDUP_INDEPENDENT_SOURCE_GROUP_EFFECTIVE_N_CLOSED",
    "PREFROZEN_SOURCE_GROUP_POWER_AND_FULL_CI_WIDTH_CLOSED",
)
NORMALIZED_CLASSES = {
    "PASS",
    "PARTIAL_OR_CONDITIONAL",
    "FAIL",
    "UNKNOWN_NOT_ASSERTED",
}
EXPECTED_REVIEW_COUNTS = {
    "pass": 3,
    "partial_or_conditional": 3,
    "fail": 1,
    "unknown_not_asserted": 4,
    "total": 11,
}
REQUIRED_SHEETS = (
    "README",
    "Variant MPRAu Results",
    "Oligo Variant Info",
    "Raw Counts HEK CMS",
    "Raw Counts HEK GWAS",
    "Raw Counts HEK Remaining",
)
CONTEXTS = ("HEK293FT", "HEPG2", "HMEC", "K562", "GM12878", "SKNSH")
HEX40 = re.compile(r"^[0-9a-f]{40}$")
HEX64 = re.compile(r"^[0-9a-f]{64}$")


class PreflightError(RuntimeError):
    """Base class for this fail-closed preflight."""


class ProtocolError(PreflightError):
    """The exact reviewed protocol surface is malformed or inconsistent."""


class ActivationBlocked(PreflightError):
    """Append-only authority/runtime/implementation binding is incomplete."""


class RepositoryError(PreflightError):
    """The active checkout does not match the future reviewed binding."""


class AssetAuditError(PreflightError):
    """A prepared ordinary-public asset differs from the frozen contract."""


class PublicationError(PreflightError):
    """The aggregate report cannot be published exactly once."""


def _exact_keys(value: Mapping[str, Any], expected: Iterable[str], *, label: str) -> None:
    if set(value) != set(expected):
        raise ProtocolError(f"{label} fields differ from the exact schema")


def _mapping(value: Any, *, label: str) -> Mapping[str, Any]:
    if not isinstance(value, dict):
        raise ProtocolError(f"{label} must be an object")
    return value


def _strict_json_bytes(payload: bytes, *, label: str) -> dict[str, Any]:
    def pairs_hook(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
        result: dict[str, Any] = {}
        for key, value in pairs:
            if key in result:
                raise ValueError(f"duplicate key {key!r}")
            result[key] = value
        return result

    def reject_constant(token: str) -> Any:
        raise ValueError(f"non-finite JSON token {token}")

    try:
        value = json.loads(
            payload.decode("utf-8"),
            object_pairs_hook=pairs_hook,
            parse_constant=reject_constant,
        )
    except (UnicodeDecodeError, ValueError, json.JSONDecodeError) as exc:
        raise ProtocolError(f"{label} is not strict UTF-8 JSON") from exc
    if not isinstance(value, dict):
        raise ProtocolError(f"{label} must be a JSON object")
    return value


def load_protocol(path: Path) -> dict[str, Any]:
    try:
        payload = path.read_bytes()
    except OSError as exc:
        raise ProtocolError(f"cannot read protocol: {path}") from exc
    protocol = _strict_json_bytes(payload, label="protocol")
    validate_protocol(protocol)
    return protocol


def _validate_sha_map(value: Any, paths: Sequence[str], *, label: str) -> None:
    if not isinstance(value, Mapping) or set(value) != set(paths):
        raise ProtocolError(f"{label} path closure differs")
    if any(not isinstance(item, str) or not HEX64.fullmatch(item) for item in value.values()):
        raise ProtocolError(f"{label} contains an invalid SHA-256")


def _validate_authority_group(group: Mapping[str, Any]) -> None:
    expected = {
        "status": BOUND,
        "authority_commit": AUTHORITY_COMMIT,
        "authority_expected_parent": AUTHORITY_PARENT,
        "authority_exact_changed_paths": list(AUTHORITY_EXACT12),
        "authority_blob_sha256_by_path": AUTHORITY_BLOBS,
    }
    if group != expected:
        raise ProtocolError("DEC027 authority A binding differs")


def _validate_runtime_group(group: Mapping[str, Any]) -> None:
    _exact_keys(
        group,
        (
            "status",
            "runtime_event_id",
            "i1_expected_parent",
            "i1_commit",
            "i1_exact_changed_paths",
            "i1_blob_sha256_by_path",
            "i2_expected_parent",
            "i2_commit",
            "i2_exact_changed_paths",
            "i2_blob_sha256_by_path",
            "b2_expected_parent",
            "b2_commit",
            "b2_exact_changed_paths",
            "b2_blob_sha256_by_path",
        ),
        label="runtime_group",
    )
    expected_scalars = {
        "status": BOUND,
        "runtime_event_id": "A1-EVT-059",
        "i1_expected_parent": AUTHORITY_COMMIT,
        "i1_commit": RUNTIME_I1_COMMIT,
        "i1_exact_changed_paths": list(RUNTIME_EXACT3),
        "i2_expected_parent": RUNTIME_I1_COMMIT,
        "i2_commit": RUNTIME_I2_COMMIT,
        "i2_exact_changed_paths": list(RUNTIME_EXACT3),
        "b2_expected_parent": RUNTIME_I2_COMMIT,
        "b2_commit": RUNTIME_B2_COMMIT,
        "b2_exact_changed_paths": [RUNTIME_CONFIG_PATH],
    }
    for field, expected in expected_scalars.items():
        if group.get(field) != expected:
            raise ProtocolError(f"runtime_group.{field} differs from EVT059")
    _validate_sha_map(
        group.get("i1_blob_sha256_by_path"),
        RUNTIME_EXACT3,
        label="runtime I1 blobs",
    )
    _validate_sha_map(
        group.get("i2_blob_sha256_by_path"),
        RUNTIME_EXACT3,
        label="runtime I2 blobs",
    )
    _validate_sha_map(
        group.get("b2_blob_sha256_by_path"),
        RUNTIME_EXACT3,
        label="runtime B2 blobs",
    )


def _predecessor_mode(group: Mapping[str, Any]) -> str:
    _exact_keys(
        group,
        (
            "status",
            "i1_expected_parent",
            "i1_commit",
            "i1_exact_changed_paths",
            "i1_blob_sha256_by_path",
            "i2_expected_parent",
            "i2_commit",
            "i2_exact_changed_paths",
            "i2_blob_sha256_by_path",
            "b2_expected_parent",
            "b2_commit",
            "b2_exact_changed_paths",
            "b2_blob_sha256_by_path",
            "i3_expected_parent",
            "i3_commit",
            "i3_exact_changed_paths",
            "i3_blob_sha256_by_path",
            "b3_expected_parent",
            "b3_commit",
            "b3_exact_changed_paths",
            "b3_blob_sha256_by_path",
        ),
        label="gse217518_predecessor_group",
    )
    expected_scalars = {
        "status": BOUND,
        "i1_expected_parent": RUNTIME_B2_COMMIT,
        "i1_commit": GSE217_I1_COMMIT,
        "i1_exact_changed_paths": list(GSE217_EXACT3),
        "i2_expected_parent": GSE217_I1_COMMIT,
        "i2_commit": GSE217_I2_COMMIT,
        "i2_exact_changed_paths": list(GSE217_EXACT3),
        "b2_expected_parent": GSE217_I2_COMMIT,
        "b2_commit": GSE217_B2_COMMIT,
        "b2_exact_changed_paths": [GSE217_CONFIG_PATH],
        "i3_expected_parent": GSE217_B2_COMMIT,
        "i3_commit": GSE217_I3_COMMIT,
        "i3_exact_changed_paths": list(GSE217_EXACT3),
        "b3_expected_parent": GSE217_I3_COMMIT,
        "b3_commit": GSE217_B3_COMMIT,
        "b3_exact_changed_paths": [GSE217_CONFIG_PATH],
    }
    for field, expected in expected_scalars.items():
        if group.get(field) != expected:
            raise ProtocolError(f"GSE217518 predecessor {field} differs")
    for step in ("i1", "i2", "b2", "i3", "b3"):
        _validate_sha_map(
            group.get(f"{step}_blob_sha256_by_path"),
            GSE217_EXACT3,
            label=f"GSE217518 {step.upper()} blobs",
        )
    return BOUND


def _own_mode(group: Mapping[str, Any]) -> str:
    _exact_keys(
        group,
        (
            "status",
            "implementation_commit",
            "implementation_script_path",
            "implementation_script_sha256",
            "implementation_test_path",
            "implementation_test_sha256",
            "implementation_exact_changed_paths",
            "binding_exact_changed_paths",
            "unknown_to_bound_scalar_paths",
        ),
        label="own_preflight_group",
    )
    if group.get("implementation_script_path") != SCRIPT_REPO_PATH:
        raise ProtocolError("own implementation script path differs")
    if group.get("implementation_test_path") != TEST_REPO_PATH:
        raise ProtocolError("own implementation test path differs")
    if tuple(group.get("implementation_exact_changed_paths", ())) != EXACT3:
        raise ProtocolError("own preflight implementation paths differ from exact3")
    if group.get("binding_exact_changed_paths") != [CONFIG_REPO_PATH]:
        raise ProtocolError("own preflight binding must be config-only")
    expected_scalar_paths = [
        f"implementation_binding.own_preflight_group.{field}"
        for field in OWN_BINDING_FIELDS
    ]
    if group.get("unknown_to_bound_scalar_paths") != expected_scalar_paths:
        raise ProtocolError("own four-scalar field list differs")
    dynamic = tuple(group.get(field) for field in OWN_BINDING_FIELDS)
    if group.get("status") == UNKNOWN:
        if dynamic != (UNKNOWN,) * len(OWN_BINDING_FIELDS):
            raise ProtocolError("own preflight group is partially bound")
        return UNKNOWN
    if group.get("status") != BOUND:
        raise ProtocolError("own preflight status is invalid")
    if not isinstance(group.get("implementation_commit"), str) or not HEX40.fullmatch(
        group["implementation_commit"]
    ):
        raise ProtocolError("own implementation commit is invalid")
    for field in ("implementation_script_sha256", "implementation_test_sha256"):
        if not isinstance(group.get(field), str) or not HEX64.fullmatch(group[field]):
            raise ProtocolError(f"own_preflight_group.{field} is not SHA-256")
    return BOUND


def _validate_binding(protocol: Mapping[str, Any]) -> tuple[str, str]:
    binding = _mapping(protocol["implementation_binding"], label="implementation_binding")
    _exact_keys(
        binding,
        (
            "binding_scheme",
            "authority_group",
            "runtime_group",
            "gse217518_predecessor_group",
            "own_preflight_group",
            "activation_rule",
        ),
        label="implementation_binding",
    )
    if binding.get("binding_scheme") != (
        "DEC027_A_TO_RUNTIME_I1_I2_B2_EVT059_TO_GSE217518_"
        "I1_I2_B2_I3_B3_TO_ENCSR854RUF_I_B"
    ):
        raise ProtocolError("append-only binding scheme differs")
    expected_activation = (
        "Authority, the complete A1-EVT-059 runtime A-to-I1-to-I2-to-B2 "
        "history, and the complete GSE217518 I1-to-I2-to-B2-to-I3-to-B3 "
        "history are frozen. ENCSR854RUF I must be the direct child of "
        "GSE217518 B3. This exact3 own four-scalar group remains grouped "
        "UNKNOWN_NOT_ASSERTED in I, so production must stop before Git, "
        "publisher workbook, author-repository asset, prepared-data path, or "
        "output-path inspection until a config-only ENCSR854RUF B changes only "
        "those four scalars. Generic registry requirements are never "
        "dataset-specific gate facts."
    )
    if binding.get("activation_rule") != expected_activation:
        raise ProtocolError("production activation rule differs")
    _validate_authority_group(
        _mapping(binding["authority_group"], label="authority_group")
    )
    _validate_runtime_group(_mapping(binding["runtime_group"], label="runtime_group"))
    predecessor_mode = _predecessor_mode(
        _mapping(binding["gse217518_predecessor_group"], label="gse217518_predecessor_group")
    )
    own_mode = _own_mode(
        _mapping(binding["own_preflight_group"], label="own_preflight_group")
    )
    return predecessor_mode, own_mode


def _normalized_counts(gates: Sequence[Mapping[str, Any]]) -> dict[str, int]:
    counter = Counter(str(item["normalized_status"]) for item in gates)
    return {
        "pass": counter["PASS"],
        "partial_or_conditional": counter["PARTIAL_OR_CONDITIONAL"],
        "fail": counter["FAIL"],
        "unknown_not_asserted": counter["UNKNOWN_NOT_ASSERTED"],
        "total": len(gates),
    }


def validate_protocol(protocol: Mapping[str, Any]) -> None:
    _exact_keys(
        protocol,
        (
            "schema_version",
            "protocol_id",
            "contract_id",
            "phase_id",
            "dataset_id",
            "decision_id",
            "protocol_status",
            "fresh_baseline",
            "repository_authority",
            "implementation_binding",
            "decision_boundary",
            "ordinary_public_asset_contract",
            "gate_contract",
            "public_research_snapshot",
            "aggregate_output_contract",
            "scientific_state",
        ),
        label="protocol",
    )
    expected = {
        "schema_version": SCHEMA_VERSION,
        "protocol_id": PROTOCOL_ID,
        "contract_id": "mrna_xeditflow_route_a_v3",
        "phase_id": "A1",
        "dataset_id": DATASET_ID,
        "decision_id": DECISION_ID,
    }
    for field, value in expected.items():
        if protocol.get(field) != value:
            raise ProtocolError(f"{field} differs from the frozen value")
    if protocol.get("protocol_status") != (
        "LOCAL_EXACT3_CANDIDATE_FULL_PREDECESSOR_HISTORY_BOUND_"
        "OWN_BINDING_UNKNOWN_NOT_ACTIVE"
    ):
        raise ProtocolError("protocol status differs")

    baseline = _mapping(protocol["fresh_baseline"], label="fresh_baseline")
    if baseline.get("dec027_authority_present") is not True:
        raise ProtocolError("DEC027 authority must be present in the observed baseline")
    if not HEX40.fullmatch(str(baseline.get("dec027_authority_commit", ""))):
        raise ProtocolError("fresh DEC027 authority commit is not full-length")
    if baseline.get("latest_settled_runtime_event_id") != "A1-EVT-059":
        raise ProtocolError("candidate must bind the settled EVT059 runtime")
    if baseline.get("dec027_runtime_binding_present") is not True:
        raise ProtocolError("candidate must preserve the settled runtime binding")
    if baseline.get("dec027_runtime_i1_commit") != RUNTIME_I1_COMMIT:
        raise ProtocolError("fresh runtime I1 differs")
    if baseline.get("dec027_runtime_i2_commit") != RUNTIME_I2_COMMIT:
        raise ProtocolError("fresh runtime I2 differs")
    if baseline.get("dec027_runtime_binding_commit") != RUNTIME_B2_COMMIT:
        raise ProtocolError("fresh runtime B2 differs")
    if baseline.get("gse217518_final_binding_commit") != GSE217_B3_COMMIT:
        raise ProtocolError("fresh GSE217518 B3 differs")
    for field in ("production_head", "upstream_head", "origin_head"):
        if baseline.get(field) != GSE217_B3_COMMIT:
            raise ProtocolError(f"fresh baseline {field} differs")
    if baseline.get("worktree_clean") is not True:
        raise ProtocolError("fresh baseline worktree must be clean")

    repository = _mapping(protocol["repository_authority"], label="repository_authority")
    if repository != {
        "production_repo_root": str(PRODUCTION_REPO_ROOT),
        "branch": PRODUCTION_BRANCH,
        "upstream_ref": PRODUCTION_UPSTREAM,
        "clean_head_equals_upstream_equals_live_origin_required": True,
    }:
        raise ProtocolError("production repository authority differs")

    _validate_binding(protocol)
    boundary = _mapping(protocol["decision_boundary"], label="decision_boundary")
    _exact_keys(
        boundary,
        (
            "public_research_role",
            "production_preflight_execution_allowed_after_all_bindings",
            "production_asset_read_allowed_after_all_bindings",
            "production_output_allowed_after_all_bindings",
            "network_download_allowed_by_candidate",
            "raw_fastq_body_read_allowed",
            "private_or_sealed_access_allowed",
            "qualification_allowed",
            "ordinary_credit_change_allowed",
            "a1_credit_change_allowed",
            "true_a2_credit_change_allowed",
            "canonical_mutation_allowed",
            "split_assignment_execution_allowed",
            "formal_power_gate_execution_allowed",
            "training_allowed",
            "gpu_work_allowed",
            "cuda_or_device_probe_allowed",
            "model_selection_allowed",
            "a7_allowed",
            "next_phase_authorized",
            "all_gates_passing_automatically_qualifies_dataset",
            "separate_promotion_authority_required",
        ),
        label="decision_boundary",
    )
    for field in (
        "production_preflight_execution_allowed_after_all_bindings",
        "production_asset_read_allowed_after_all_bindings",
        "production_output_allowed_after_all_bindings",
    ):
        if boundary.get(field) is not True:
            raise ProtocolError(f"decision_boundary.{field} must be true")
    for field in (
        "network_download_allowed_by_candidate",
        "raw_fastq_body_read_allowed",
        "private_or_sealed_access_allowed",
        "qualification_allowed",
        "ordinary_credit_change_allowed",
        "a1_credit_change_allowed",
        "true_a2_credit_change_allowed",
        "canonical_mutation_allowed",
        "split_assignment_execution_allowed",
        "formal_power_gate_execution_allowed",
        "training_allowed",
        "gpu_work_allowed",
        "cuda_or_device_probe_allowed",
        "model_selection_allowed",
        "a7_allowed",
        "next_phase_authorized",
        "all_gates_passing_automatically_qualifies_dataset",
    ):
        if boundary.get(field) is not False:
            raise ProtocolError(f"decision_boundary.{field} must be false")
    if boundary.get("separate_promotion_authority_required") is not True:
        raise ProtocolError("separate promotion authority must remain required")
    gate_contract = _mapping(protocol["gate_contract"], label="gate_contract")
    if tuple(gate_contract.get("gate_ids_exactly", ())) != GATE_IDS:
        raise ProtocolError("gate IDs differ from DEC027 exact 11")
    if gate_contract.get("independent_gate_axis_count") != len(GATE_IDS):
        raise ProtocolError("independent gate axis count must be 11")
    if gate_contract.get("registry_generic_requirements_may_be_reported_as_dataset_specific_results") is not False:
        raise ProtocolError("generic registry requirements cannot become dataset facts")

    snapshot = _mapping(protocol["public_research_snapshot"], label="public_research_snapshot")
    gates = snapshot.get("gate_statuses")
    if not isinstance(gates, list) or len(gates) != len(GATE_IDS):
        raise ProtocolError("public research snapshot must have exact 11 gates")
    if tuple(item.get("gate_id") for item in gates if isinstance(item, dict)) != GATE_IDS:
        raise ProtocolError("public research gate order differs from DEC027")
    for item in gates:
        _exact_keys(
            _mapping(item, label="gate status"),
            ("gate_id", "raw_status", "normalized_status", "fact_class", "reason_code", "aggregate_evidence"),
            label="gate status",
        )
        if item["normalized_status"] not in NORMALIZED_CLASSES:
            raise ProtocolError("unmapped normalized gate status")
    if _normalized_counts(gates) != EXPECTED_REVIEW_COUNTS:
        raise ProtocolError("public research gate counts differ from 3/3/1/4")
    if snapshot.get("normalized_gate_counts") != EXPECTED_REVIEW_COUNTS:
        raise ProtocolError("frozen normalized gate counts disagree with gate rows")
    if snapshot.get("aggregate_verdict") != "STOP_NOT_READY_FOR_A1_QUALIFICATION_OR_CREDIT":
        raise ProtocolError("aggregate verdict must remain STOP")

    output = _mapping(protocol["aggregate_output_contract"], label="aggregate_output_contract")
    _exact_keys(
        output,
        (
            "report_schema_version",
            "report_filename",
            "allowed_output",
            "same_directory_temporary_file_required",
            "file_fsync_before_publish_required",
            "atomic_no_replace_hard_link_required",
            "directory_fsync_after_publish_required",
            "identical_existing_report_is_idempotent",
            "different_existing_report_action",
            "failure_cleanup_required",
            "member_identifier_allowed",
            "source_or_candidate_sequence_allowed",
            "row_endpoint_or_effect_allowed",
            "row_standard_error_allowed",
            "replicate_identifier_allowed",
            "split_assignment_allowed",
            "final_file_count",
        ),
        label="aggregate_output_contract",
    )
    if output.get("report_schema_version") != REPORT_SCHEMA_VERSION:
        raise ProtocolError("report schema version differs")
    if output.get("report_filename") != REPORT_FILENAME or output.get("final_file_count") != 1:
        raise ProtocolError("report must be exactly one fixed JSON file")
    for field in (
        "same_directory_temporary_file_required",
        "file_fsync_before_publish_required",
        "atomic_no_replace_hard_link_required",
        "directory_fsync_after_publish_required",
        "identical_existing_report_is_idempotent",
        "failure_cleanup_required",
    ):
        if output.get(field) is not True:
            raise ProtocolError(f"aggregate_output_contract.{field} must be true")
    if output.get("different_existing_report_action") != "REJECT_WITHOUT_REPLACEMENT":
        raise ProtocolError("different existing report action differs")
    for field in (
        "member_identifier_allowed",
        "source_or_candidate_sequence_allowed",
        "row_endpoint_or_effect_allowed",
        "row_standard_error_allowed",
        "replicate_identifier_allowed",
        "split_assignment_allowed",
    ):
        if output.get(field) is not False:
            raise ProtocolError(f"{field} must remain false")

    state = _mapping(protocol["scientific_state"], label="scientific_state")
    if state.get("current_qualified_counts") != {
        "ordinary": 1,
        "a1": 1,
        "true_a2": 0,
        "canonical_records": 6547,
    }:
        raise ProtocolError("qualified-count baseline changed")
    if state.get("contribution_delta") != {
        "ordinary": 0,
        "a1": 0,
        "true_a2": 0,
        "canonical_records": 0,
    }:
        raise ProtocolError("preflight contribution delta must be zero")

def _require_activation(protocol: Mapping[str, Any]) -> None:
    """Stop before Git, asset, prepared-path, and output-path inspection."""

    predecessor_mode, own_mode = _validate_binding(protocol)
    if predecessor_mode != BOUND:
        raise ActivationBlocked(
            "gse217518_predecessor_group remains UNKNOWN_NOT_ASSERTED; "
            "stopped before Git/asset/output I/O"
        )
    if own_mode != BOUND:
        raise ActivationBlocked(
            "own_preflight_group remains UNKNOWN_NOT_ASSERTED; "
            "stopped before Git/asset/output I/O"
        )


def _run_git(repo_root: Path, *args: str) -> str:
    try:
        result = subprocess.run(
            ["git", "-C", os.fspath(repo_root), *args],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
    except (OSError, subprocess.CalledProcessError) as exc:
        raise RepositoryError(f"git {' '.join(args)} failed") from exc
    return result.stdout.strip()


def _changed_paths(repo_root: Path, commit: str) -> tuple[str, ...]:
    output = _run_git(repo_root, "diff-tree", "--no-commit-id", "--name-only", "-r", commit)
    return tuple(sorted(line for line in output.splitlines() if line))


def _git_blob(repo_root: Path, commit: str, path: str) -> bytes:
    try:
        result = subprocess.run(
            ["git", "-C", os.fspath(repo_root), "show", f"{commit}:{path}"],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
        )
    except (OSError, subprocess.CalledProcessError) as exc:
        raise RepositoryError(f"cannot read bound Git blob: {path}") from exc
    return result.stdout


def _verify_commit(
    repo_root: Path,
    *,
    label: str,
    commit: str,
    expected_parent: str,
    expected_paths: Sequence[str],
    expected_blobs: Mapping[str, str] | None = None,
) -> None:
    ancestry = _run_git(repo_root, "rev-list", "--parents", "-n", "1", commit).split()
    if ancestry != [commit, expected_parent]:
        raise RepositoryError(f"{label} direct parent differs")
    if _changed_paths(repo_root, commit) != tuple(sorted(expected_paths)):
        raise RepositoryError(f"{label} changed-path closure differs")
    for path, expected_sha in (expected_blobs or {}).items():
        observed_sha = hashlib.sha256(_git_blob(repo_root, commit, path)).hexdigest()
        if observed_sha != expected_sha:
            raise RepositoryError(f"{label} blob identity differs: {path}")


def _live_origin_head(repo_root: Path, branch: str) -> str:
    ref = f"refs/heads/{branch}"
    value = _run_git(repo_root, "ls-remote", "--exit-code", "--heads", "origin", ref)
    rows = [line.split() for line in value.splitlines() if line.strip()]
    if len(rows) != 1 or len(rows[0]) != 2 or rows[0][1] != ref:
        raise RepositoryError("live origin branch resolution differs")
    if not HEX40.fullmatch(rows[0][0]):
        raise RepositoryError("live origin head is not a commit")
    return rows[0][0]


def _normalise_own_binding(protocol: Mapping[str, Any]) -> dict[str, Any]:
    normalised = copy.deepcopy(dict(protocol))
    group = normalised["implementation_binding"]["own_preflight_group"]
    for field in OWN_BINDING_FIELDS:
        group[field] = UNKNOWN
    return normalised


def _normalise_gse217_binding(protocol: Mapping[str, Any]) -> dict[str, Any]:
    normalised = copy.deepcopy(dict(protocol))
    try:
        group = normalised["bindings"]["implementation"]
    except (KeyError, TypeError) as exc:
        raise RepositoryError("GSE217518 protocol binding structure differs") from exc
    for field in GSE217_OWN_BINDING_FIELDS:
        group[field] = UNKNOWN
    return normalised


def _read_repository_file(path: Path, *, label: str) -> bytes:
    try:
        return path.read_bytes()
    except OSError as exc:
        raise RepositoryError(f"cannot read {label}") from exc


def _audit_repository(
    protocol: Mapping[str, Any], config_path: Path, repo_root: Path
) -> dict[str, Any]:
    """Audit A -> runtime I1/I2/B2 -> GSE217518 I1/I2/B2/I3/B3 -> own I/B."""

    _require_activation(protocol)
    repository = protocol["repository_authority"]
    if repo_root.resolve() != Path(repository["production_repo_root"]).resolve():
        raise RepositoryError("execution repository is not the frozen production root")
    if config_path.resolve() != (repo_root / CONFIG_REPO_PATH).resolve():
        raise RepositoryError("executing config is outside the frozen repository path")

    head = _run_git(repo_root, "rev-parse", "HEAD")
    upstream = _run_git(repo_root, "rev-parse", "@{upstream}")
    live_origin = _live_origin_head(repo_root, repository["branch"])
    if head != upstream or head != live_origin:
        raise RepositoryError("HEAD, upstream, and live origin do not match")
    if _run_git(repo_root, "rev-parse", "--abbrev-ref", "HEAD") != repository["branch"]:
        raise RepositoryError("production branch differs")
    if _run_git(repo_root, "rev-parse", "--abbrev-ref", "@{upstream}") != repository["upstream_ref"]:
        raise RepositoryError("production upstream differs")
    if _run_git(repo_root, "status", "--porcelain=v1", "--untracked-files=all"):
        raise RepositoryError("production worktree or index is dirty")

    binding = protocol["implementation_binding"]
    authority = binding["authority_group"]
    runtime = binding["runtime_group"]
    predecessor = binding["gse217518_predecessor_group"]
    own = binding["own_preflight_group"]
    own_i = own["implementation_commit"]

    _verify_commit(
        repo_root,
        label="DEC027 authority A",
        commit=AUTHORITY_COMMIT,
        expected_parent=AUTHORITY_PARENT,
        expected_paths=AUTHORITY_EXACT12,
        expected_blobs=authority["authority_blob_sha256_by_path"],
    )
    _verify_commit(
        repo_root,
        label="DEC027 runtime I1",
        commit=RUNTIME_I1_COMMIT,
        expected_parent=AUTHORITY_COMMIT,
        expected_paths=RUNTIME_EXACT3,
        expected_blobs=runtime["i1_blob_sha256_by_path"],
    )
    _verify_commit(
        repo_root,
        label="DEC027 runtime I2",
        commit=RUNTIME_I2_COMMIT,
        expected_parent=RUNTIME_I1_COMMIT,
        expected_paths=RUNTIME_EXACT3,
        expected_blobs=runtime["i2_blob_sha256_by_path"],
    )
    _verify_commit(
        repo_root,
        label="DEC027 runtime B2",
        commit=RUNTIME_B2_COMMIT,
        expected_parent=RUNTIME_I2_COMMIT,
        expected_paths=(RUNTIME_CONFIG_PATH,),
        expected_blobs=runtime["b2_blob_sha256_by_path"],
    )
    _verify_commit(
        repo_root,
        label="GSE217518 I1",
        commit=GSE217_I1_COMMIT,
        expected_parent=RUNTIME_B2_COMMIT,
        expected_paths=GSE217_EXACT3,
        expected_blobs=predecessor["i1_blob_sha256_by_path"],
    )
    _verify_commit(
        repo_root,
        label="GSE217518 I2",
        commit=GSE217_I2_COMMIT,
        expected_parent=GSE217_I1_COMMIT,
        expected_paths=GSE217_EXACT3,
        expected_blobs=predecessor["i2_blob_sha256_by_path"],
    )
    _verify_commit(
        repo_root,
        label="GSE217518 B2",
        commit=GSE217_B2_COMMIT,
        expected_parent=GSE217_I2_COMMIT,
        expected_paths=(GSE217_CONFIG_PATH,),
        expected_blobs=predecessor["b2_blob_sha256_by_path"],
    )
    _verify_commit(
        repo_root,
        label="GSE217518 I3",
        commit=GSE217_I3_COMMIT,
        expected_parent=GSE217_B2_COMMIT,
        expected_paths=GSE217_EXACT3,
        expected_blobs=predecessor["i3_blob_sha256_by_path"],
    )
    _verify_commit(
        repo_root,
        label="GSE217518 B3",
        commit=GSE217_B3_COMMIT,
        expected_parent=GSE217_I3_COMMIT,
        expected_paths=(GSE217_CONFIG_PATH,),
        expected_blobs=predecessor["b3_blob_sha256_by_path"],
    )
    _verify_commit(
        repo_root,
        label="ENCSR854RUF implementation I",
        commit=own_i,
        expected_parent=GSE217_B3_COMMIT,
        expected_paths=EXACT3,
        expected_blobs={
            SCRIPT_REPO_PATH: own["implementation_script_sha256"],
            TEST_REPO_PATH: own["implementation_test_sha256"],
        },
    )
    _verify_commit(
        repo_root,
        label="ENCSR854RUF binding B",
        commit=head,
        expected_parent=own_i,
        expected_paths=(CONFIG_REPO_PATH,),
    )

    for implementation, binding_commit, label in (
        (GSE217_I2_COMMIT, GSE217_B2_COMMIT, "B2"),
        (GSE217_I3_COMMIT, GSE217_B3_COMMIT, "B3"),
    ):
        gse_i_protocol = _strict_json_bytes(
            _git_blob(repo_root, implementation, GSE217_CONFIG_PATH),
            label=f"GSE217518 {label} implementation protocol",
        )
        gse_b_protocol = _strict_json_bytes(
            _git_blob(repo_root, binding_commit, GSE217_CONFIG_PATH),
            label=f"GSE217518 {label} binding protocol",
        )
        if _normalise_gse217_binding(gse_b_protocol) != gse_i_protocol:
            raise RepositoryError(
                f"GSE217518 {label} changed fields outside its own four scalars"
            )

    own_i_protocol = _strict_json_bytes(
        _git_blob(repo_root, own_i, CONFIG_REPO_PATH),
        label="ENCSR854RUF I protocol",
    )
    if _normalise_own_binding(protocol) != own_i_protocol:
        raise RepositoryError("ENCSR854RUF B changed fields outside its own four scalars")
    if _read_repository_file(config_path, label="working ENCSR854RUF config") != _git_blob(
        repo_root, head, CONFIG_REPO_PATH
    ):
        raise RepositoryError("working config differs from ENCSR854RUF B")
    executing_script = Path(__file__).resolve()
    if executing_script != (repo_root / SCRIPT_REPO_PATH).resolve():
        raise RepositoryError("executing producer is a stale or copied script")
    if _read_repository_file(executing_script, label="executing producer") != _git_blob(
        repo_root, own_i, SCRIPT_REPO_PATH
    ):
        raise RepositoryError("executing producer differs from ENCSR854RUF I")
    focused_test = repo_root / TEST_REPO_PATH
    if _read_repository_file(focused_test, label="working focused test") != _git_blob(
        repo_root, own_i, TEST_REPO_PATH
    ):
        raise RepositoryError("working focused test differs from ENCSR854RUF I")
    return {
        "status": "BOUND_REPOSITORY_CHAIN_CLOSED",
        "head": head,
        "runtime_i1_commit": RUNTIME_I1_COMMIT,
        "runtime_i2_commit": RUNTIME_I2_COMMIT,
        "runtime_b2_commit": RUNTIME_B2_COMMIT,
        "gse217518_i1_commit": GSE217_I1_COMMIT,
        "gse217518_i2_commit": GSE217_I2_COMMIT,
        "gse217518_b2_commit": GSE217_B2_COMMIT,
        "gse217518_i3_commit": GSE217_I3_COMMIT,
        "gse217518_b3_commit": GSE217_B3_COMMIT,
        "implementation_commit": own["implementation_commit"],
        "binding_commit": head,
        "changed_path_counts": {
            "authority": 12,
            "runtime_i1": 3,
            "runtime_i2": 3,
            "runtime_b2": 1,
            "gse217518_i1": 3,
            "gse217518_i2": 3,
            "gse217518_b2": 1,
            "gse217518_i3": 3,
            "gse217518_b3": 1,
            "implementation": 3,
            "binding": 1,
        },
    }


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    try:
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
    except OSError as exc:
        raise AssetAuditError(f"cannot read required file: {path}") from exc
    return digest.hexdigest()


def _verify_file(path: Path, identity: Mapping[str, Any], *, label: str) -> None:
    try:
        size = path.stat().st_size
    except OSError as exc:
        raise AssetAuditError(f"missing {label}") from exc
    if size != identity.get("byte_count"):
        raise AssetAuditError(f"{label} byte count differs")
    if _sha256(path) != identity.get("sha256"):
        raise AssetAuditError(f"{label} SHA-256 differs")


def _header_map(values: Sequence[Any], *, sheet: str) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, value in enumerate(values):
        if value is None or str(value).strip() == "":
            continue
        key = str(value).strip()
        if key in result:
            raise AssetAuditError(f"duplicate header in {sheet}")
        result[key] = index
    return result


def _finite_number(value: Any) -> bool:
    if isinstance(value, bool) or value is None:
        return False
    try:
        number = float(value)
    except (TypeError, ValueError):
        return False
    return math.isfinite(number)


def _load_fasta_aliases(path: Path) -> tuple[dict[str, str], int]:
    aliases: dict[str, str] = {}
    header_count = 0
    header: str | None = None
    sequence_parts: list[str] = []

    def commit() -> None:
        nonlocal header_count
        if header is None:
            return
        sequence = "".join(sequence_parts).strip().upper()
        if not sequence or any(base not in "ACGTN" for base in sequence):
            raise AssetAuditError("author FASTA contains an invalid sequence alphabet")
        header_count += 1
        for token in header.split("/"):
            token = token.strip()
            if not token:
                raise AssetAuditError("author FASTA contains an empty slash alias")
            if token in aliases and aliases[token] != sequence:
                raise AssetAuditError("author FASTA slash alias maps to conflicting sequences")
            aliases[token] = sequence

    try:
        with path.open("r", encoding="utf-8") as handle:
            for raw in handle:
                line = raw.strip()
                if not line:
                    continue
                if line.startswith(">"):
                    commit()
                    header = line[1:].split()[0]
                    sequence_parts = []
                else:
                    if header is None:
                        raise AssetAuditError("author FASTA starts before its first header")
                    sequence_parts.append(line)
    except (OSError, UnicodeDecodeError) as exc:
        raise AssetAuditError("cannot parse author FASTA") from exc
    commit()
    return aliases, header_count


def _levenshtein(left: str, right: str) -> int:
    if len(left) < len(right):
        left, right = right, left
    previous = list(range(len(right) + 1))
    for left_index, left_base in enumerate(left, 1):
        current = [left_index]
        for right_index, right_base in enumerate(right, 1):
            current.append(
                min(
                    current[-1] + 1,
                    previous[right_index] + 1,
                    previous[right_index - 1] + (left_base != right_base),
                )
            )
        previous = current
    return previous[-1]


def _read_array_assignments(path: Path) -> list[dict[str, str]]:
    try:
        with path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle, delimiter="\t")
            expected = {"oligo_name", "tag", "dup", "ref_name", "CMS_array", "GWAS_array"}
            if set(reader.fieldnames or ()) != expected:
                raise AssetAuditError("author array-assignment schema differs")
            return [dict(row) for row in reader]
    except (OSError, UnicodeDecodeError, csv.Error) as exc:
        raise AssetAuditError("cannot parse author array assignments") from exc


def _rows_and_header(workbook: Any, sheet_name: str) -> tuple[dict[str, int], Any]:
    sheet = workbook[sheet_name]
    iterator = sheet.iter_rows(values_only=True)
    try:
        header = next(iterator)
    except StopIteration as exc:
        raise AssetAuditError(f"empty workbook sheet: {sheet_name}") from exc
    return _header_map(header, sheet=sheet_name), iterator


def _parse_workbook(workbook_path: Path, fasta_path: Path, array_path: Path) -> dict[str, int]:
    try:
        import openpyxl
    except ImportError as exc:
        raise AssetAuditError("openpyxl is required for the frozen XLSX route") from exc
    try:
        workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises multiple format-specific exceptions
        raise AssetAuditError("publisher workbook cannot be opened") from exc
    try:
        missing_sheets = sorted(set(REQUIRED_SHEETS) - set(workbook.sheetnames))
        if missing_sheets:
            raise AssetAuditError("publisher workbook required-sheet set differs")
        aliases, fasta_header_count = _load_fasta_aliases(fasta_path)
        array_rows = _read_array_assignments(array_path)

        oligo_header, oligo_rows = _rows_and_header(workbook, "Oligo Variant Info")
        required_oligo = {
            "mpra_variant_id",
            "tag",
            "oligo_id",
            "ref_allele",
            "alt_allele",
            "other_var_in_oligo_window",
        }
        if not required_oligo.issubset(oligo_header):
            raise AssetAuditError("publisher oligo-information schema differs")

        pair_members: dict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
        reporter_ids: set[str] = set()
        index_error_pairs: set[str] = set()
        for row in oligo_rows:
            pair_id = str(row[oligo_header["mpra_variant_id"]])
            role = str(row[oligo_header["tag"]]).strip().lower()
            reporter_id = str(row[oligo_header["oligo_id"]])
            if role not in {"ref", "alt"} or role in pair_members[pair_id]:
                raise AssetAuditError("source/candidate role closure differs")
            if reporter_id in reporter_ids:
                raise AssetAuditError("publisher reporter identifier is duplicated")
            reporter_ids.add(reporter_id)
            pair_members[pair_id][role] = {
                "reporter_id": reporter_id,
                "ref_allele": str(row[oligo_header["ref_allele"]]).upper(),
                "alt_allele": str(row[oligo_header["alt_allele"]]).upper(),
            }
            other = str(row[oligo_header["other_var_in_oligo_window"]] or "")
            if "index_error" in {part.strip().lower() for part in other.split(",")}:
                index_error_pairs.add(pair_id)

        malformed_pair_count = sum(set(members) != {"ref", "alt"} for members in pair_members.values())
        if malformed_pair_count:
            raise AssetAuditError("not every publisher pair has exactly one ref and one alt")
        crosswalk_missing = sum(
            member["reporter_id"] not in aliases
            for members in pair_members.values()
            for member in members.values()
        )
        if crosswalk_missing:
            raise AssetAuditError("publisher reporter to author FASTA crosswalk is incomplete")

        identical_pair_count = 0
        replay_mismatch_count = 0
        insert_lengths: Counter[int] = Counter()
        source_by_pair: dict[str, str] = {}
        candidate_by_pair: dict[str, str] = {}
        for pair_id, members in pair_members.items():
            source = aliases[members["ref"]["reporter_id"]]
            candidate = aliases[members["alt"]["reporter_id"]]
            source_by_pair[pair_id] = source
            candidate_by_pair[pair_id] = candidate
            insert_lengths[len(source)] += 1
            insert_lengths[len(candidate)] += 1
            identical_pair_count += source == candidate
            declared_ref = members["ref"]["ref_allele"]
            declared_alt = members["ref"]["alt_allele"]
            expected_distance = max(len(declared_ref), len(declared_alt))
            replay_mismatch_count += _levenshtein(source, candidate) != expected_distance

        result_header, result_rows = _rows_and_header(workbook, "Variant MPRAu Results")
        if "mpra_variant_id" not in result_header:
            raise AssetAuditError("publisher result identifier column is absent")
        effect_columns = [f"log2FoldChange_Skew_{context}" for context in CONTEXTS]
        se_columns = [f"lfcSE_Skew_{context}" for context in CONTEXTS]
        if not set(effect_columns + se_columns).issubset(result_header):
            raise AssetAuditError("six-context effect/standard-error schema differs")
        statistic_columns = [index for name, index in result_header.items() if name != "mpra_variant_id"]
        result_ids: set[str] = set()
        finite_all_pairs: set[str] = set()
        finite_any_pairs: set[str] = set()
        rows_with_nonnumeric = 0
        nonnumeric_cells = 0
        for row in result_rows:
            pair_id = str(row[result_header["mpra_variant_id"]])
            if pair_id in result_ids:
                raise AssetAuditError("publisher result pair identifier is duplicated")
            result_ids.add(pair_id)
            row_bad = sum(not _finite_number(row[index]) for index in statistic_columns)
            rows_with_nonnumeric += row_bad > 0
            nonnumeric_cells += row_bad
            finite_by_context = [
                _finite_number(row[result_header[effect]]) and _finite_number(row[result_header[se]])
                for effect, se in zip(effect_columns, se_columns)
            ]
            if all(finite_by_context):
                finite_all_pairs.add(pair_id)
            if any(finite_by_context):
                finite_any_pairs.add(pair_id)

        result_missing_pair_count = len(set(pair_members) - result_ids)
        unexpected_result_pair_count = len(result_ids - set(pair_members))
        if result_missing_pair_count or unexpected_result_pair_count:
            raise AssetAuditError("publisher oligo and result pair universes differ")
        eligible_all = finite_all_pairs - index_error_pairs
        eligible_any = finite_any_pairs - index_error_pairs

        raw_ids: dict[str, set[str]] = {}
        raw_numeric_cells = 0
        raw_invalid_cells = 0
        for sheet_name in (
            "Raw Counts HEK CMS",
            "Raw Counts HEK GWAS",
            "Raw Counts HEK Remaining",
        ):
            header, rows = _rows_and_header(workbook, sheet_name)
            if "oligo_id" not in header:
                raise AssetAuditError(f"{sheet_name} lacks oligo_id")
            numeric_columns = [index for name, index in header.items() if name != "oligo_id"]
            if not numeric_columns:
                raise AssetAuditError(f"{sheet_name} has no count columns")
            seen: set[str] = set()
            for row in rows:
                reporter_id = str(row[header["oligo_id"]])
                if reporter_id in seen:
                    raise AssetAuditError(f"{sheet_name} duplicates a reporter identifier")
                seen.add(reporter_id)
                for index in numeric_columns:
                    raw_numeric_cells += 1
                    value = row[index]
                    if not _finite_number(value):
                        raw_invalid_cells += 1
                        continue
                    number = float(value)
                    raw_invalid_cells += number < 0 or not number.is_integer()
            raw_ids[sheet_name] = seen

        expected_by_array: dict[str, set[str]] = {"CMS_array": set(), "GWAS_array": set()}
        for row in array_rows:
            aliases_for_row = {token.strip() for token in row["ref_name"].split("/") if token.strip()}
            for flag in expected_by_array:
                if str(row[flag]).strip() == "1":
                    expected_by_array[flag].update(aliases_for_row)
        expected_array_absence = len(
            (expected_by_array["CMS_array"] & reporter_ids) - raw_ids["Raw Counts HEK CMS"]
        ) + len(
            (expected_by_array["GWAS_array"] & reporter_ids) - raw_ids["Raw Counts HEK GWAS"]
        )
        remaining_absence = len(reporter_ids - raw_ids["Raw Counts HEK Remaining"])

        if len(insert_lengths) != 1:
            variable_insert_length = -1
        else:
            variable_insert_length = next(iter(insert_lengths))
        geometry = {
            "publisher_workbook_sheet_count": len(workbook.sheetnames),
            "published_pair_count": len(pair_members),
            "published_reporter_count": len(reporter_ids),
            "pair_size_two_count": len(pair_members),
            "reference_role_reporter_count": len(pair_members),
            "alternate_role_reporter_count": len(pair_members),
            "source_candidate_crosswalk_missing_count": crosswalk_missing,
            "author_fasta_header_count": fasta_header_count,
            "author_fasta_expanded_alias_token_count": len(aliases),
            "published_reporter_to_author_fasta_missing_after_documented_alias_expansion": crosswalk_missing,
            "variable_insert_length_bp": variable_insert_length,
            "ref_alt_sequence_identical_pair_count": identical_pair_count,
            "declared_allele_length_to_sequence_replay_mismatch_count": replay_mismatch_count,
            "index_error_affected_pair_count": len(index_error_pairs),
            "finite_effect_and_lfcse_all_six_context_pair_count_after_index_error_exclusion": len(eligible_all),
            "finite_effect_and_lfcse_at_least_one_context_pair_count_after_index_error_exclusion": len(eligible_any),
            "full_universe_distinct_exact_source_sequence_count": len(set(source_by_pair.values())),
            "finite_six_context_no_index_error_distinct_exact_source_sequence_count": len(
                {source_by_pair[pair_id] for pair_id in eligible_all}
            ),
            "finite_six_context_no_index_error_distinct_exact_source_candidate_pair_count": len(
                {(source_by_pair[pair_id], candidate_by_pair[pair_id]) for pair_id in eligible_all}
            ),
            "reported_result_rows_with_any_nonnumeric_statistic": rows_with_nonnumeric,
            "reported_nonnumeric_statistic_cell_count": nonnumeric_cells,
            "raw_count_numeric_cell_count": raw_numeric_cells,
            "raw_count_missing_nonfinite_negative_or_noninteger_cell_count": raw_invalid_cells,
            "expected_array_specific_hek_reporter_absence_count": expected_array_absence,
            "remaining_context_primary_reporter_absence_count": remaining_absence,
            "endpoint_context_count": len(CONTEXTS),
            "result_missing_pair_count": result_missing_pair_count,
            "unexpected_result_pair_count": unexpected_result_pair_count,
        }
        return geometry
    finally:
        workbook.close()


def _audit_prepared_assets(
    protocol: Mapping[str, Any], publisher_workbook: Path, author_repository: Path
) -> dict[str, int]:
    assets = protocol["ordinary_public_asset_contract"]
    publisher = assets["publisher_processed_asset"]
    _verify_file(publisher_workbook, publisher, label="publisher processed workbook")
    author = assets["author_repository"]
    verified_paths: dict[str, Path] = {}
    for relative, identity in author["required_assets"].items():
        path = author_repository / relative
        _verify_file(path, identity, label=f"author asset {relative}")
        verified_paths[relative] = path
    return _parse_workbook(
        publisher_workbook,
        verified_paths["data/GWASrewritepos_CMS_alignment_file.fasta"],
        verified_paths["data/GWASrewritepos_CMS_arrayassign"],
    )


def _replace_gate_failure(
    gates: list[dict[str, Any]], gate_id: str, reason_code: str, evidence: str
) -> None:
    for gate in gates:
        if gate["gate_id"] == gate_id:
            gate.update(
                {
                    "raw_status": f"FAIL_{reason_code}",
                    "normalized_status": "FAIL",
                    "fact_class": "CONFIRMED_EXECUTION_FACT",
                    "reason_code": reason_code,
                    "aggregate_evidence": evidence,
                }
            )
            return
    raise ProtocolError(f"cannot locate gate {gate_id}")


def evaluate_gate_statuses(
    protocol: Mapping[str, Any], geometry: Mapping[str, int]
) -> list[dict[str, Any]]:
    gates = copy.deepcopy(protocol["public_research_snapshot"]["gate_statuses"])
    if not (
        geometry["published_pair_count"] == geometry["pair_size_two_count"]
        and geometry["published_reporter_count"] == 2 * geometry["published_pair_count"]
        and geometry["source_candidate_crosswalk_missing_count"] == 0
        and geometry["result_missing_pair_count"] == 0
        and geometry["unexpected_result_pair_count"] == 0
    ):
        _replace_gate_failure(
            gates,
            GATE_IDS[1],
            "OBSERVED_SOURCE_CANDIDATE_CROSSWALK_NOT_CLOSED",
            "Aggregate execution found a role, crosswalk, or result-universe mismatch.",
        )
    if not (
        geometry["variable_insert_length_bp"] == 133
        and geometry["ref_alt_sequence_identical_pair_count"] == 0
        and geometry["declared_allele_length_to_sequence_replay_mismatch_count"] == 0
    ):
        _replace_gate_failure(
            gates,
            GATE_IDS[2],
            "OBSERVED_REPORTER_INSERT_OR_EDIT_REPLAY_DIFFERS",
            "Aggregate execution found a construct-length, identical-pair, or edit-replay mismatch.",
        )
    if geometry["endpoint_context_count"] != 6:
        _replace_gate_failure(
            gates,
            GATE_IDS[3],
            "OBSERVED_ENDPOINT_CONTEXT_SCHEMA_DIFFERS",
            "The six-context source-relative effect and standard-error schema was not closed.",
        )
    if geometry["finite_effect_and_lfcse_at_least_one_context_pair_count_after_index_error_exclusion"] == 0:
        _replace_gate_failure(
            gates,
            GATE_IDS[4],
            "NO_FINITE_SOURCE_RELATIVE_EFFECT_AND_STANDARD_ERROR_PAIR",
            "No pair retained a finite source-relative effect and standard error in any context.",
        )
    return gates


def build_aggregate_record(
    protocol: Mapping[str, Any],
    geometry: Mapping[str, int],
    repository_binding: Mapping[str, Any],
) -> dict[str, Any]:
    gates = evaluate_gate_statuses(protocol, geometry)
    counts = _normalized_counts(gates)
    record = {
        "schema_version": REPORT_SCHEMA_VERSION,
        "protocol_id": PROTOCOL_ID,
        "dataset_id": DATASET_ID,
        "decision_id": DECISION_ID,
        "source_snapshot_observed_at": protocol["public_research_snapshot"]["observed_at"],
        "status": "TERMINAL_AGGREGATE_PREFLIGHT_STOP_NOT_QUALIFIED",
        "repository_binding": copy.deepcopy(dict(repository_binding)),
        "aggregate_geometry": copy.deepcopy(dict(geometry)),
        "gate_statuses": gates,
        "normalized_gate_counts": counts,
        "aggregate_verdict": "STOP_NOT_READY_FOR_A1_QUALIFICATION_OR_CREDIT",
        "all_eleven_gates_pass": counts["pass"] == 11,
        "qualification_or_credit_change": False,
        "contribution_delta": {
            "ordinary": 0,
            "a1": 0,
            "true_a2": 0,
            "canonical_records": 0,
        },
        "retained_locks": {
            "training_allowed": False,
            "gpu_work_allowed": False,
            "model_selection_allowed": False,
            "a7_allowed": False,
            "next_phase_authorized": False,
            "scientific_claim_status": "NOT_ESTABLISHED_UNCHANGED",
        },
        "payload_output_counts": {
            "member_identifier_output_count": 0,
            "source_or_candidate_sequence_output_count": 0,
            "row_endpoint_or_effect_output_count": 0,
            "row_standard_error_output_count": 0,
            "replicate_identifier_output_count": 0,
            "split_assignment_output_count": 0,
        },
    }
    assert_aggregate_only(record)
    return record


FORBIDDEN_OUTPUT_KEYS = {
    "member_id",
    "source_sequence",
    "candidate_sequence",
    "row_endpoint",
    "row_effect",
    "row_standard_error",
    "replicate_id",
    "replicate_identifier_value",
    "split_assignment_value",
}


def assert_aggregate_only(value: Any) -> None:
    if isinstance(value, dict):
        forbidden = FORBIDDEN_OUTPUT_KEYS & set(value)
        if forbidden:
            raise PublicationError(f"member-level output key is forbidden: {sorted(forbidden)[0]}")
        for child in value.values():
            assert_aggregate_only(child)
    elif isinstance(value, list):
        for child in value:
            assert_aggregate_only(child)
    elif isinstance(value, float) and not math.isfinite(value):
        raise PublicationError("non-finite numeric output is forbidden")


def _write_exactly_one(record: Mapping[str, Any], output_dir: Path) -> Path:
    assert_aggregate_only(record)
    try:
        payload = (
            json.dumps(record, indent=2, sort_keys=True, allow_nan=False) + "\n"
        ).encode("utf-8")
    except (TypeError, ValueError) as exc:
        raise PublicationError("aggregate report is not finite JSON") from exc

    destination = output_dir / REPORT_FILENAME
    directory_created = False
    output_created = False
    temporary: Path | None = None
    try:
        if output_dir.exists():
            if not output_dir.is_dir():
                raise PublicationError("output path is not a directory")
            entries = list(output_dir.iterdir())
            if entries:
                if len(entries) == 1 and entries[0] == destination:
                    if destination.read_bytes() == payload:
                        return destination
                    raise PublicationError(
                        "different aggregate report already exists; replacement refused"
                    )
                raise PublicationError("output directory contains an unexpected entry")
        else:
            output_dir.mkdir(parents=False, exist_ok=False)
            directory_created = True
            _fsync_directory(output_dir.parent)

        descriptor, temporary_name = tempfile.mkstemp(
            prefix=f".{REPORT_FILENAME}.", suffix=".tmp", dir=output_dir
        )
        temporary = Path(temporary_name)
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        try:
            os.link(temporary, destination)
            output_created = True
        except FileExistsError as exc:
            if destination.read_bytes() == payload:
                temporary.unlink()
                temporary = None
                return destination
            raise PublicationError(
                "different aggregate report appeared; replacement refused"
            ) from exc
        temporary.unlink()
        temporary = None
        _fsync_directory(output_dir)
        if list(output_dir.iterdir()) != [destination]:
            raise PublicationError("publication did not produce exactly one fixed report")
        return destination
    except PublicationError:
        if output_created:
            try:
                destination.unlink()
            except OSError:
                pass
        raise
    except OSError as exc:
        if output_created:
            try:
                destination.unlink()
            except OSError:
                pass
        raise PublicationError("cannot atomically publish aggregate report") from exc
    finally:
        if temporary is not None:
            try:
                temporary.unlink()
            except FileNotFoundError:
                pass
        if directory_created and output_dir.exists():
            try:
                if not any(output_dir.iterdir()):
                    output_dir.rmdir()
            except OSError:
                pass


def _fsync_directory(path: Path) -> None:
    descriptor = os.open(path, os.O_RDONLY | getattr(os, "O_DIRECTORY", 0))
    try:
        os.fsync(descriptor)
    finally:
        os.close(descriptor)


def execute(
    protocol_path: Path,
    repository_root: Path,
    publisher_workbook: Path,
    author_repository: Path,
    output_dir: Path,
) -> Path:
    protocol = load_protocol(protocol_path)
    _require_activation(protocol)
    repository_binding = _audit_repository(protocol, protocol_path, repository_root)
    geometry = _audit_prepared_assets(protocol, publisher_workbook, author_repository)
    record = build_aggregate_record(
        protocol,
        geometry,
        repository_binding,
    )
    return _write_exactly_one(record, output_dir)


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", type=Path, required=True)
    parser.add_argument("--repository-root", type=Path, default=PRODUCTION_REPO_ROOT)
    parser.add_argument("--publisher-workbook", type=Path, required=True)
    parser.add_argument("--author-repository", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = _parser().parse_args(argv)
    try:
        destination = execute(
            args.config,
            args.repository_root,
            args.publisher_workbook,
            args.author_repository,
            args.output_dir,
        )
        print(destination)
        return 0
    except PreflightError as exc:
        print(f"STOP: {exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
