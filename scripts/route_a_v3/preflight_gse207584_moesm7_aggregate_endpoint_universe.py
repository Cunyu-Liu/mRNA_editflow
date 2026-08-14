#!/usr/bin/env python3
"""Aggregate-only GSE207584 publisher/GEO endpoint-universe preflight.

This local successor candidate establishes only the geometry of the 955-row
publisher-primary endpoint universe.  It verifies that the GEO processed table
is a within-family candidate-by-endpoint Cartesian expansion, without using row
order and without reading sequence.  It cannot identify which endpoint belongs
to which candidate and therefore cannot qualify the dataset or change A1/A2,
credit, canonical, training, GPU, model-selection, A7, or next-phase state.
"""

from __future__ import annotations

import argparse
import copy
import csv
import gzip
import hashlib
import json
import math
import os
import re
import subprocess
import tempfile
from collections import Counter, defaultdict
from contextlib import contextmanager
from pathlib import Path
from typing import Any, Callable, Iterator, Mapping, TextIO

from openpyxl import load_workbook


SCHEMA_VERSION = (
    "route_a_v3_gse207584_moesm7_aggregate_endpoint_universe_preflight.v1"
)
PROTOCOL_ID = "GSE207584_MOESM7_AGGREGATE_ENDPOINT_UNIVERSE_PREFLIGHT_V1"
DATASET_ID = "GSE207584"
BIOPROJECT_ID = "PRJNA856272"
DECISION_ID = "V3-DEC-023"
PROTOCOL_STATUS = "DRAFT_FOR_REVIEW_NOT_ACTIVE_PROTOCOL"
REPORT_FILENAME = "GSE207584_MOESM7_AGGREGATE_ENDPOINT_UNIVERSE_PREFLIGHT_V1.json"

UNKNOWN = "UNKNOWN_NOT_ASSERTED"
BOUND = "BOUND"
PROTOCOL_BASENAME = (
    "route_a_v3_gse207584_moesm7_aggregate_endpoint_universe_preflight_v1.json"
)
CONFIG_PATH = f"configs/{PROTOCOL_BASENAME}"
SCRIPT_PATH = (
    "scripts/route_a_v3/"
    "preflight_gse207584_moesm7_aggregate_endpoint_universe.py"
)
TEST_PATH = (
    "tests/route_a_v3/"
    "test_preflight_gse207584_moesm7_aggregate_endpoint_universe.py"
)
EXPECTED_EXACT3 = (CONFIG_PATH, SCRIPT_PATH, TEST_PATH)
UNKNOWN_BINDING_SCALARS = (
    "status",
    "implementation_commit",
    "implementation_script_sha256",
    "implementation_test_sha256",
)
UNKNOWN_TO_BOUND_PATHS = tuple(
    f"implementation_binding.{field}" for field in UNKNOWN_BINDING_SCALARS
)
CURRENT_PREDECESSOR_COMMIT = "0a6586814460b211cc730c463390e68f64aaa4f1"
INITIAL_IMPLEMENTATION_COMMIT = "374ea6166c74c898751c7a3d4d6951664ca1d524"
INITIAL_IMPLEMENTATION_FROZEN_BLOBS = {
    CONFIG_PATH: "308d9dda359a56d3d765c326b99de0bfdf5eff0434a7d3075ddc0c06152f547a",
    SCRIPT_PATH: "4473dacb00714e33a4fdd8e97920a297dae2b4e3ce33991d387b16d4d0d7ee28",
    TEST_PATH: "3c7dd85e415f18cfc93508bb0474b91aa04e8a68817db918cfca01cedb819626",
}
BINDING_SCHEME = (
    "CURRENT_PREDECESSOR_THEN_I1_EXACT3_THEN_I2_EXACT3_CONFIG_ONLY_B2_V2"
)
PRODUCTION_REPO_ROOT = Path(
    "/home/cunyuliu/mrna_editflow_goal/worktrees/routea_v3_a1_20260810"
)
PRODUCTION_BRANCH = "routea-v3-a1-20260810"
REMOTE_NAME = "origin"
EXECUTING_SCRIPT_PATH = Path(__file__).resolve()
HEX40_RE = re.compile(r"^[0-9a-f]{40}$")
HEX64_RE = re.compile(r"^[0-9a-f]{64}$")

PASS = "PASS_PREFLIGHT_ONLY"
FAIL = "FAIL_CLOSED"
UNKNOWN = "UNKNOWN_NOT_ASSERTED"
STATUS_CLOSED = (
    "AGGREGATE_ENDPOINT_UNIVERSE_CLOSED_"
    "SOURCE_CANDIDATE_MAPPING_UNKNOWN_NOT_QUALIFIED"
)
STATUS_STOP = "STOP_AGGREGATE_ENDPOINT_UNIVERSE_NOT_CLOSED_NOT_QUALIFIED"

PUBLISHER_HEADER = (
    "Protein_id",
    "Group",
    "2h_A",
    "2h_B",
    "2h_C",
    "5h_A",
    "5h_B",
    "5h_C",
    "8h_A",
    "8h_B",
    "8h_C",
    "Decay_rate",
    "Predicted_mRNA_stability",
)
GEO_HEADER = (
    "Name",
    "Protein_id",
    "Group",
    "zf_library_2h_1",
    "zf_library_2h_2",
    "zf_library_2h_3",
    "zf_library_5h_1",
    "zf_library_5h_2",
    "zf_library_5h_3",
    "zf_library_8h_1",
    "zf_library_8h_2",
    "zf_library_8h_3",
)
GATE_IDS = (
    "PUBLISHER_ENDPOINT_UNIVERSE_CLOSED",
    "GEO_AGGREGATE_COVERAGE_GEOMETRY_CLOSED",
    "PUBLISHER_TO_GEO_CARTESIAN_EXPANSION_SEMANTICS_CLOSED",
    "PRIVATE_ANALYSIS_AND_AGGREGATE_REPORTING_RIGHTS_CLOSED",
    "ROW_ORDER_AND_SEQUENCE_PAIRING_PROHIBITION_ENFORCED",
    "SOURCE_TO_CANDIDATE_MAPPING_CLOSED",
)
FALSE_CLAIM_KEYS = (
    "source_to_candidate_mapping_established",
    "source_to_candidate_edit_replay_closed",
    "a1_qualification_allowed_or_changed",
    "true_a2_status_allowed_or_changed",
    "study_credit_allowed_or_changed",
    "canonical_allowed_or_changed",
    "training_allowed_or_changed",
    "gpu_work_allowed_or_changed",
    "model_selection_allowed_or_changed",
    "a7_unlock_allowed_or_changed",
    "next_phase_allowed_or_changed",
)
EXPECTED_GEOMETRY = {
    "publisher_endpoint_row_count": 955,
    "publisher_unique_endpoint_signature_count": 955,
    "publisher_family_count": 100,
    "publisher_group_label_count": 7,
    "geo_body_row_count": 10227,
    "geo_unique_candidate_name_count": 955,
    "geo_family_count": 100,
    "endpoint_signature_numeric_decimal_places": 9,
}


class PreflightError(RuntimeError):
    """Base class for a fail-closed preflight error."""


class ProtocolError(PreflightError):
    """The local draft protocol differs from its frozen boundary."""


class BindingNotFrozen(ProtocolError):
    """The exact3-I/config-only-B implementation lifecycle is not closed."""


class AssetError(PreflightError):
    """An ordinary-public input cannot be interpreted under the protocol."""


class OutputError(PreflightError):
    """The sole aggregate report cannot be written safely."""


def _mapping(value: Any, *, label: str) -> Mapping[str, Any]:
    if not isinstance(value, Mapping):
        raise ProtocolError(f"{label} must be an object")
    return value


def _strict_json(payload: bytes) -> dict[str, Any]:
    def no_duplicates(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
        result: dict[str, Any] = {}
        for key, value in pairs:
            if key in result:
                raise ValueError(f"duplicate JSON key: {key}")
            result[key] = value
        return result

    try:
        value = json.loads(
            payload.decode("utf-8"),
            object_pairs_hook=no_duplicates,
            parse_constant=lambda token: (_ for _ in ()).throw(
                ValueError(f"non-finite JSON token: {token}")
            ),
        )
    except (UnicodeError, ValueError) as exc:
        raise ProtocolError("protocol is not strict UTF-8 JSON") from exc
    if not isinstance(value, dict):
        raise ProtocolError("protocol root must be an object")
    return value


def _validate_protocol(protocol: Mapping[str, Any]) -> None:
    exact_scalars = {
        "schema_version": SCHEMA_VERSION,
        "protocol_id": PROTOCOL_ID,
        "contract_id": "mrna_xeditflow_route_a_v3",
        "phase_id": "A1",
        "dataset_id": DATASET_ID,
        "bioproject_id": BIOPROJECT_ID,
        "decision_id": DECISION_ID,
        "protocol_status": PROTOCOL_STATUS,
    }
    for key, expected in exact_scalars.items():
        if protocol.get(key) != expected:
            raise ProtocolError(f"{key} differs from the draft boundary")

    binding = _mapping(
        protocol.get("implementation_binding"), label="implementation binding"
    )
    expected_binding_keys = {
        "binding_scheme",
        "status",
        "current_predecessor_commit",
        "initial_implementation_commit",
        "initial_implementation_exact_changed_paths",
        "initial_implementation_frozen_blobs",
        "production_repo_root",
        "production_branch",
        "remote_name",
        "implementation_commit",
        "implementation_script_path",
        "implementation_script_sha256",
        "implementation_test_path",
        "implementation_test_sha256",
        "unknown_to_bound_scalar_paths",
        "implementation_commit_exact_changed_paths",
        "binding_commit_exact_changed_paths",
        "activation_rule",
    }
    if set(binding) != expected_binding_keys:
        raise ProtocolError("implementation binding schema differs")
    expected_initial_blobs = [
        {"path": path, "sha256": INITIAL_IMPLEMENTATION_FROZEN_BLOBS[path]}
        for path in EXPECTED_EXACT3
    ]
    if (
        binding.get("binding_scheme") != BINDING_SCHEME
        or binding.get("current_predecessor_commit")
        != CURRENT_PREDECESSOR_COMMIT
        or binding.get("initial_implementation_commit")
        != INITIAL_IMPLEMENTATION_COMMIT
        or tuple(binding.get("initial_implementation_exact_changed_paths", ()))
        != EXPECTED_EXACT3
        or binding.get("initial_implementation_frozen_blobs")
        != expected_initial_blobs
        or binding.get("production_repo_root") != str(PRODUCTION_REPO_ROOT)
        or binding.get("production_branch") != PRODUCTION_BRANCH
        or binding.get("remote_name") != REMOTE_NAME
        or binding.get("implementation_script_path") != SCRIPT_PATH
        or binding.get("implementation_test_path") != TEST_PATH
        or tuple(binding.get("unknown_to_bound_scalar_paths", ()))
        != UNKNOWN_TO_BOUND_PATHS
        or tuple(binding.get("implementation_commit_exact_changed_paths", ()))
        != EXPECTED_EXACT3
        or binding.get("binding_commit_exact_changed_paths") != [CONFIG_PATH]
    ):
        raise ProtocolError("implementation binding boundary differs")
    status = binding.get("status")
    values = [binding.get(field) for field in UNKNOWN_BINDING_SCALARS]
    if status == UNKNOWN:
        if values != [UNKNOWN] * len(UNKNOWN_BINDING_SCALARS):
            raise ProtocolError("I implementation binding is partially known")
    elif status == BOUND:
        if not (
            isinstance(binding.get("implementation_commit"), str)
            and HEX40_RE.fullmatch(str(binding["implementation_commit"]))
        ):
            raise ProtocolError("BOUND implementation commit is invalid")
        for field in (
            "implementation_script_sha256",
            "implementation_test_sha256",
        ):
            if not (
                isinstance(binding.get(field), str)
                and HEX64_RE.fullmatch(str(binding[field]))
            ):
                raise ProtocolError(f"BOUND {field} is invalid")
    else:
        raise ProtocolError("implementation binding status is invalid")

    authority = _mapping(protocol.get("authority_assessment"), label="authority")
    if (
        authority.get(
            "current_authority_sufficient_for_this_candidate_and_aggregate_preflight"
        )
        is not True
        or authority.get("new_row_level_authority_granted") is not False
        or authority.get("sequence_field_classes_read_count") != 0
        or authority.get("allowed_output_class")
        != "AGGREGATE_COUNTS_HISTOGRAMS_AND_GATE_STATUSES_ONLY"
    ):
        raise ProtocolError("DEC023 authority boundary differs")

    assets = _mapping(
        protocol.get("ordinary_public_asset_contract"), label="asset contract"
    )
    publisher = _mapping(
        assets.get("publisher_endpoint_workbook"), label="publisher asset"
    )
    geo = _mapping(assets.get("geo_perfect_csv"), label="GEO asset")
    reference = _mapping(assets.get("reference_fasta"), label="reference policy")
    if (
        publisher.get("filename") != "41598_2022_15526_MOESM7_ESM.xlsx"
        or publisher.get("sheet_name") != "Supplemental-Table-6-library-pe"
        or tuple(publisher.get("required_header_exactly", ())) != PUBLISHER_HEADER
        or publisher.get("derived_effect_columns_values_read") is not False
        or geo.get("filename") != "GSE207584_Zebrafish-library-perfect.csv.gz"
        or tuple(geo.get("required_header_exactly", ())) != GEO_HEADER
        or reference.get("read_by_this_successor") is not False
    ):
        raise ProtocolError("ordinary-public asset boundary differs")

    geometry = _mapping(protocol.get("prefrozen_geometry"), label="geometry")
    for key, expected in EXPECTED_GEOMETRY.items():
        if geometry.get(key) != expected:
            raise ProtocolError(f"prefrozen geometry differs at {key}")
    if (
        geometry.get("row_order_pairing_permitted") is not False
        or geometry.get("sequence_pairing_permitted") is not False
    ):
        raise ProtocolError("row-order or sequence pairing was enabled")
    histogram_raw = _mapping(
        geometry.get("family_size_histogram"), label="family-size histogram"
    )
    try:
        histogram = {int(size): int(count) for size, count in histogram_raw.items()}
    except (TypeError, ValueError) as exc:
        raise ProtocolError("family-size histogram is malformed") from exc
    if any(size <= 0 or count <= 0 for size, count in histogram.items()):
        raise ProtocolError("family-size histogram contains nonpositive entries")
    if sum(histogram.values()) != EXPECTED_GEOMETRY["publisher_family_count"]:
        raise ProtocolError("family-size histogram family total differs")
    if (
        sum(size * count for size, count in histogram.items())
        != EXPECTED_GEOMETRY["publisher_endpoint_row_count"]
    ):
        raise ProtocolError("family-size histogram endpoint total differs")
    if (
        sum(size * size * count for size, count in histogram.items())
        != EXPECTED_GEOMETRY["geo_body_row_count"]
    ):
        raise ProtocolError("family-size histogram Cartesian total differs")

    rights = _mapping(protocol.get("rights_policy"), label="rights")
    if (
        rights.get("private_academic_analysis_allowed") is not True
        or rights.get("aggregate_derived_reporting_allowed") is not True
        or rights.get("member_payload_redistribution_allowed") is not False
        or rights.get("submitter_intellectual_property_transfer_asserted") is not False
    ):
        raise ProtocolError("rights scope differs")
    if not isinstance(rights.get("basis"), list) or len(rights["basis"]) != 3:
        raise ProtocolError("rights evidence basis differs")

    if tuple(protocol.get("gate_ids", ())) != GATE_IDS:
        raise ProtocolError("gate identifiers differ")
    claims = _mapping(protocol.get("claim_boundary"), label="claim boundary")
    if any(claims.get(key) is not False for key in FALSE_CLAIM_KEYS):
        raise ProtocolError("a forbidden scientific or execution claim was enabled")
    if claims.get("current_credit_delta") != {
        "ordinary": 0,
        "A1": 0,
        "true_A2": 0,
    }:
        raise ProtocolError("credit delta differs from zero")

    output = _mapping(protocol.get("output_contract"), label="output")
    if (
        output.get("filename") != REPORT_FILENAME
        or output.get("single_json_only") is not True
        or output.get("overwrite_allowed") is not False
        or output.get("existing_exact_payload_policy") != "ACCEPT_IDEMPOTENTLY"
        or output.get("existing_different_payload_policy") != "REJECT"
        or output.get("publication_strategy")
        != "SAME_DIRECTORY_UNIQUE_TEMP_WRITE_FLUSH_FSYNC_THEN_ATOMIC_LINK_NOREPLACE"
        or output.get("member_identifier_sequence_or_row_payload_output_count") != 0
        or output.get("row_or_member_material_persisted_count") != 0
    ):
        raise ProtocolError("aggregate-only output boundary differs")


def _run_git_text(repo_root: Path, *arguments: str) -> str:
    try:
        completed = subprocess.run(
            ["git", "-C", str(repo_root), *arguments],
            check=False,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
    except OSError as exc:
        raise ProtocolError("git is unavailable for implementation binding") from exc
    if completed.returncode != 0:
        raise ProtocolError("git implementation-binding check failed")
    return completed.stdout.strip()


def _git_blob(repo_root: Path, commit: str, relative_path: str) -> bytes:
    try:
        completed = subprocess.run(
            ["git", "-C", str(repo_root), "show", f"{commit}:{relative_path}"],
            check=False,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
        )
    except OSError as exc:
        raise ProtocolError("git is unavailable for implementation binding") from exc
    if completed.returncode != 0:
        raise ProtocolError("cannot read a bound implementation blob")
    return completed.stdout


def _changed_paths(repo_root: Path, commit: str) -> tuple[str, ...]:
    value = _run_git_text(
        repo_root,
        "diff-tree",
        "--no-commit-id",
        "--name-only",
        "-r",
        commit,
    )
    return tuple(sorted(line for line in value.splitlines() if line))


def _require_single_parent(
    repo_root: Path,
    commit: str,
    expected_parent: str,
    *,
    label: str,
) -> None:
    lineage = _run_git_text(
        repo_root, "rev-list", "--parents", "-n", "1", commit
    ).split()
    if lineage != [commit, expected_parent]:
        raise ProtocolError(f"{label} is not an exact single-parent commit")


def _semantic_diff_paths(left: Any, right: Any, prefix: str = "") -> set[str]:
    if type(left) is not type(right):
        return {prefix or "<root>"}
    if isinstance(left, dict):
        differences: set[str] = set()
        for key in sorted(set(left) | set(right)):
            path = f"{prefix}.{key}" if prefix else key
            if key not in left or key not in right:
                differences.add(path)
            else:
                differences.update(_semantic_diff_paths(left[key], right[key], path))
        return differences
    if isinstance(left, list):
        if len(left) != len(right):
            return {prefix}
        differences: set[str] = set()
        for index, (left_item, right_item) in enumerate(zip(left, right)):
            differences.update(
                _semantic_diff_paths(
                    left_item,
                    right_item,
                    f"{prefix}[{index}]",
                )
            )
        return differences
    return {prefix} if left != right else set()


def _normalise_binding_to_i(protocol: Mapping[str, Any]) -> dict[str, Any]:
    result = copy.deepcopy(dict(protocol))
    binding = result["implementation_binding"]
    for field in UNKNOWN_BINDING_SCALARS:
        binding[field] = UNKNOWN
    return result


def _truth_projection(protocol: Mapping[str, Any]) -> dict[str, Any]:
    """Return all scientific/geometry truth, excluding lifecycle bookkeeping."""

    result = copy.deepcopy(dict(protocol))
    result.pop("implementation_binding", None)
    return result


def _verify_live_origin_head(
    repo_root: Path,
    *,
    remote: str,
    branch: str,
    expected_head: str,
) -> None:
    reference = f"refs/heads/{branch}"
    rows = _run_git_text(
        repo_root,
        "ls-remote",
        "--heads",
        remote,
        reference,
    ).splitlines()
    if rows != [f"{expected_head}\t{reference}"]:
        raise ProtocolError("live origin branch is not current HEAD")


def _default_binding_auditor(
    protocol: Mapping[str, Any],
    protocol_path: Path,
    protocol_payload: bytes,
    repo_root: Path,
) -> dict[str, str]:
    """Close predecessor -> frozen I1 -> repair I2 -> config-only B2.

    This is called before any publisher/GEO asset or output operation.  A
    draft I file, a partially filled binding, a stale checkout, or a copied
    script therefore cannot reach the data readers.
    """

    binding = protocol["implementation_binding"]
    if binding.get("status") != BOUND:
        raise BindingNotFrozen("repair exact3-I2/config-only-B2 lifecycle is not BOUND")
    if any(binding.get(field) == UNKNOWN for field in UNKNOWN_BINDING_SCALARS):
        raise BindingNotFrozen("implementation binding remains UNKNOWN or partial")

    try:
        resolved_repo = repo_root.resolve(strict=True)
    except OSError as exc:
        raise ProtocolError("production repository root is unavailable") from exc
    if resolved_repo != PRODUCTION_REPO_ROOT:
        raise ProtocolError("production repository root differs")
    if protocol_path.resolve() != resolved_repo / CONFIG_PATH:
        raise ProtocolError("protocol path is outside the production config location")

    branch = str(binding["production_branch"])
    remote = str(binding["remote_name"])
    if _run_git_text(resolved_repo, "rev-parse", "--abbrev-ref", "HEAD") != branch:
        raise ProtocolError("production branch differs")
    if _run_git_text(
        resolved_repo, "status", "--porcelain=v1", "--untracked-files=all"
    ):
        raise ProtocolError("production worktree is not clean")

    head = _run_git_text(resolved_repo, "rev-parse", "HEAD")
    upstream_name = _run_git_text(
        resolved_repo,
        "rev-parse",
        "--abbrev-ref",
        "--symbolic-full-name",
        "@{u}",
    )
    if upstream_name != f"{remote}/{branch}":
        raise ProtocolError("configured upstream branch differs")
    upstream_head = _run_git_text(resolved_repo, "rev-parse", "@{u}")
    tracking_head = _run_git_text(
        resolved_repo, "rev-parse", f"refs/remotes/{remote}/{branch}"
    )
    if head != upstream_head or head != tracking_head:
        raise ProtocolError("HEAD, upstream, and origin tracking ref differ")
    _verify_live_origin_head(
        resolved_repo,
        remote=remote,
        branch=branch,
        expected_head=head,
    )

    initial_implementation_commit = str(binding["initial_implementation_commit"])
    implementation_commit = str(binding["implementation_commit"])
    predecessor = str(binding["current_predecessor_commit"])
    _require_single_parent(
        resolved_repo,
        initial_implementation_commit,
        predecessor,
        label="initial implementation I1",
    )
    _require_single_parent(
        resolved_repo,
        implementation_commit,
        initial_implementation_commit,
        label="repair implementation I2",
    )
    _require_single_parent(
        resolved_repo,
        head,
        implementation_commit,
        label="binding B2",
    )
    if _changed_paths(resolved_repo, initial_implementation_commit) != tuple(
        sorted(EXPECTED_EXACT3)
    ):
        raise ProtocolError("initial implementation I1 did not change exact3")
    if _changed_paths(resolved_repo, implementation_commit) != tuple(
        sorted(EXPECTED_EXACT3)
    ):
        raise ProtocolError("repair implementation I2 did not change exact3")
    if _changed_paths(resolved_repo, head) != (CONFIG_PATH,):
        raise ProtocolError("binding B2 did not change config-only")

    initial_blobs: dict[str, bytes] = {}
    for path in EXPECTED_EXACT3:
        blob = _git_blob(resolved_repo, initial_implementation_commit, path)
        if hashlib.sha256(blob).hexdigest() != INITIAL_IMPLEMENTATION_FROZEN_BLOBS[path]:
            raise ProtocolError(f"frozen initial I1 blob differs: {path}")
        initial_blobs[path] = blob
    initial_protocol = _strict_json(initial_blobs[CONFIG_PATH])
    i_config_blob = _git_blob(resolved_repo, implementation_commit, CONFIG_PATH)
    b_config_blob = _git_blob(resolved_repo, head, CONFIG_PATH)
    i_protocol = _strict_json(i_config_blob)
    b_protocol = _strict_json(b_config_blob)
    _validate_protocol(i_protocol)
    _validate_protocol(b_protocol)
    i_binding = i_protocol["implementation_binding"]
    if [i_binding.get(field) for field in UNKNOWN_BINDING_SCALARS] != [
        UNKNOWN
    ] * len(UNKNOWN_BINDING_SCALARS):
        raise ProtocolError("repair implementation I2 does not own exact four UNKNOWN scalars")
    truth = _truth_projection(initial_protocol)
    if _truth_projection(i_protocol) != truth or _truth_projection(b_protocol) != truth:
        raise ProtocolError("I1-to-I2/B2 normalized scientific truth changed")
    differences = _semantic_diff_paths(i_protocol, b_protocol)
    if differences != set(UNKNOWN_TO_BOUND_PATHS):
        raise ProtocolError("I2-to-B2 semantic change is not the exact four scalars")
    if _normalise_binding_to_i(b_protocol) != i_protocol:
        raise ProtocolError("B2 does not normalize exactly to repair I2")
    if b_protocol != dict(protocol) or b_config_blob != protocol_payload:
        raise ProtocolError("executed protocol is not the current B2 Git blob")

    script_blob = _git_blob(resolved_repo, implementation_commit, SCRIPT_PATH)
    test_blob = _git_blob(resolved_repo, implementation_commit, TEST_PATH)
    if _git_blob(resolved_repo, head, SCRIPT_PATH) != script_blob:
        raise ProtocolError("B2 script blob differs from repair I2")
    if _git_blob(resolved_repo, head, TEST_PATH) != test_blob:
        raise ProtocolError("B2 focused-test blob differs from repair I2")
    if hashlib.sha256(script_blob).hexdigest() != binding.get(
        "implementation_script_sha256"
    ):
        raise ProtocolError("bound implementation script hash differs")
    if hashlib.sha256(test_blob).hexdigest() != binding.get(
        "implementation_test_sha256"
    ):
        raise ProtocolError("bound focused-test hash differs")

    expected_script_path = (resolved_repo / SCRIPT_PATH).resolve()
    expected_test_path = (resolved_repo / TEST_PATH).resolve()
    if EXECUTING_SCRIPT_PATH.resolve() != expected_script_path:
        raise ProtocolError("executing __file__ is a stale or copied script")
    try:
        working_config = protocol_path.read_bytes()
        executing_script = EXECUTING_SCRIPT_PATH.read_bytes()
        working_test = expected_test_path.read_bytes()
    except OSError as exc:
        raise ProtocolError("bound working implementation is unreadable") from exc
    if working_config != b_config_blob:
        raise ProtocolError("working config differs from binding B2")
    if executing_script != script_blob:
        raise ProtocolError("executing __file__ bytes differ from repair I2")
    if working_test != test_blob:
        raise ProtocolError("working focused test differs from repair I2")

    return {
        "status": "BOUND_FROZEN_I1_REPAIR_EXACT3_I2_CONFIG_ONLY_B2_VERIFIED",
        "current_predecessor_commit": predecessor,
        "initial_implementation_commit": initial_implementation_commit,
        "implementation_commit": implementation_commit,
        "binding_commit": head,
        "upstream_head": upstream_head,
        "tracking_head": tracking_head,
        "live_origin_head": head,
    }


def _normalise_number(value: Any, decimal_places: int) -> float:
    if value is None or isinstance(value, bool):
        raise ValueError("missing or nonnumeric cell")
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError("nonnumeric cell") from exc
    if not math.isfinite(number):
        raise ValueError("nonfinite cell")
    result = round(number, decimal_places)
    return 0.0 if result == 0 else result


def _label(value: Any) -> str:
    if value is None:
        raise ValueError("missing label")
    result = str(value).strip()
    if not result:
        raise ValueError("empty label")
    return result


def _read_publisher(
    path: Path, contract: Mapping[str, Any], decimal_places: int
) -> dict[str, Any]:
    if path.name != contract.get("filename"):
        raise AssetError("publisher workbook filename differs")
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (OSError, ValueError) as exc:
        raise AssetError("publisher workbook is unreadable") from exc
    sheet_name = contract["sheet_name"]
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise AssetError("publisher endpoint sheet is absent")
    sheet = workbook[sheet_name]
    header = tuple(
        cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1, max_col=13))
    )
    if header != PUBLISHER_HEADER:
        workbook.close()
        raise AssetError("publisher endpoint header differs")

    family_signatures: dict[str, list[tuple[Any, ...]]] = defaultdict(list)
    groups: set[str] = set()
    body_rows = 0
    malformed_rows = 0
    invalid_numeric_rows = 0
    for row in sheet.iter_rows(min_row=2, max_col=11, values_only=True):
        body_rows += 1
        try:
            family = _label(row[0])
            group = _label(row[1])
            values = tuple(
                _normalise_number(value, decimal_places) for value in row[2:11]
            )
        except ValueError as exc:
            if "numeric" in str(exc) or "cell" in str(exc):
                invalid_numeric_rows += 1
            else:
                malformed_rows += 1
            continue
        signature = (group, *values)
        family_signatures[family].append(signature)
        groups.add(group)
    workbook.close()

    signature_counts = Counter(
        (family, signature)
        for family, signatures in family_signatures.items()
        for signature in signatures
    )
    family_sizes = Counter(len(rows) for rows in family_signatures.values())
    return {
        "body_row_count": body_rows,
        "valid_row_count": sum(len(rows) for rows in family_signatures.values()),
        "family_count": len(family_signatures),
        "group_label_count": len(groups),
        "unique_endpoint_signature_count": len(signature_counts),
        "duplicate_endpoint_signature_count": sum(
            count - 1 for count in signature_counts.values()
        ),
        "malformed_row_count": malformed_rows,
        "invalid_numeric_row_count": invalid_numeric_rows,
        "family_size_histogram": dict(sorted(family_sizes.items())),
        "_family_signatures": {
            family: tuple(signatures)
            for family, signatures in family_signatures.items()
        },
    }


@contextmanager
def _open_geo(path: Path) -> Iterator[TextIO]:
    try:
        if path.name.endswith(".gz"):
            with gzip.open(path, "rt", encoding="utf-8", newline="") as handle:
                yield handle
        else:
            with path.open("r", encoding="utf-8", newline="") as handle:
                yield handle
    except (OSError, UnicodeError) as exc:
        raise AssetError("GEO processed CSV is unreadable") from exc


def _read_geo(
    path: Path, contract: Mapping[str, Any], decimal_places: int
) -> dict[str, Any]:
    if path.name != contract.get("filename"):
        raise AssetError("GEO processed CSV filename differs")
    family_candidates: dict[str, set[str]] = defaultdict(set)
    family_signatures: dict[str, set[tuple[Any, ...]]] = defaultdict(set)
    family_row_counts: Counter[str] = Counter()
    pair_counts: Counter[tuple[str, str, tuple[Any, ...]]] = Counter()
    candidate_families: dict[str, set[str]] = defaultdict(set)
    groups: set[str] = set()
    body_rows = 0
    malformed_rows = 0
    invalid_numeric_rows = 0
    with _open_geo(path) as handle:
        reader = csv.reader(handle)
        try:
            header = tuple(next(reader))
        except StopIteration as exc:
            raise AssetError("GEO processed CSV is empty") from exc
        if header != GEO_HEADER:
            raise AssetError("GEO processed CSV header differs")
        for row in reader:
            body_rows += 1
            if len(row) != len(GEO_HEADER):
                malformed_rows += 1
                continue
            try:
                candidate = _label(row[0])
                family = _label(row[1])
                group = _label(row[2])
                values = tuple(
                    _normalise_number(value, decimal_places) for value in row[3:12]
                )
            except ValueError as exc:
                if "numeric" in str(exc) or "cell" in str(exc):
                    invalid_numeric_rows += 1
                else:
                    malformed_rows += 1
                continue
            signature = (group, *values)
            family_candidates[family].add(candidate)
            family_signatures[family].add(signature)
            family_row_counts[family] += 1
            pair_counts[(family, candidate, signature)] += 1
            candidate_families[candidate].add(family)
            groups.add(group)

    candidate_family_conflicts = sum(
        1 for families in candidate_families.values() if len(families) != 1
    )
    return {
        "body_row_count": body_rows,
        "valid_row_count": sum(family_row_counts.values()),
        "unique_candidate_name_count": len(candidate_families),
        "family_count": len(family_candidates),
        "group_label_count": len(groups),
        "unique_endpoint_signature_count": sum(
            len(signatures) for signatures in family_signatures.values()
        ),
        "unique_candidate_endpoint_pair_count": len(pair_counts),
        "duplicate_candidate_endpoint_pair_row_count": sum(
            count - 1 for count in pair_counts.values()
        ),
        "candidate_cross_family_conflict_count": candidate_family_conflicts,
        "malformed_row_count": malformed_rows,
        "invalid_numeric_row_count": invalid_numeric_rows,
        "candidate_count_by_family_histogram": dict(
            sorted(Counter(len(values) for values in family_candidates.values()).items())
        ),
        "_family_candidates": {
            family: frozenset(candidates)
            for family, candidates in family_candidates.items()
        },
        "_family_signatures": {
            family: frozenset(signatures)
            for family, signatures in family_signatures.items()
        },
        "_family_row_counts": dict(family_row_counts),
        "_pair_counts": pair_counts,
    }


def _gate(status: str, reason: str) -> dict[str, str]:
    return {"status": status, "reason": reason}


def aggregate(
    protocol: Mapping[str, Any],
    publisher_workbook: Path,
    geo_perfect_csv: Path,
) -> dict[str, Any]:
    assets = protocol["ordinary_public_asset_contract"]
    geometry = protocol["prefrozen_geometry"]
    decimal_places = geometry["endpoint_signature_numeric_decimal_places"]
    publisher = _read_publisher(
        publisher_workbook, assets["publisher_endpoint_workbook"], decimal_places
    )
    geo = _read_geo(geo_perfect_csv, assets["geo_perfect_csv"], decimal_places)

    publisher_families = publisher["_family_signatures"]
    geo_candidates = geo["_family_candidates"]
    geo_signatures = geo["_family_signatures"]
    geo_row_counts = geo["_family_row_counts"]
    pair_counts = geo["_pair_counts"]
    all_families = set(publisher_families) | set(geo_candidates) | set(geo_signatures)

    family_endpoint_candidate_count_match = 0
    family_signature_set_match = 0
    family_square_row_count_match = 0
    family_complete_cartesian = 0
    publisher_signatures_not_in_geo = 0
    geo_signatures_not_in_publisher = 0
    expected_geo_rows_from_publisher_family_squares = 0
    for family in all_families:
        publisher_rows = tuple(publisher_families.get(family, ()))
        publisher_signature_set = set(publisher_rows)
        candidates = set(geo_candidates.get(family, ()))
        observed_signature_set = set(geo_signatures.get(family, ()))
        expected_geo_rows_from_publisher_family_squares += len(publisher_rows) ** 2
        publisher_signatures_not_in_geo += len(
            publisher_signature_set - observed_signature_set
        )
        geo_signatures_not_in_publisher += len(
            observed_signature_set - publisher_signature_set
        )
        if len(publisher_rows) == len(candidates):
            family_endpoint_candidate_count_match += 1
        if publisher_signature_set == observed_signature_set:
            family_signature_set_match += 1
        if geo_row_counts.get(family, 0) == len(publisher_rows) * len(candidates):
            family_square_row_count_match += 1
        expected_pairs = {
            (family, candidate, signature)
            for candidate in candidates
            for signature in publisher_signature_set
        }
        observed_pairs = {
            pair for pair in pair_counts if pair[0] == family
        }
        if (
            expected_pairs == observed_pairs
            and all(pair_counts[pair] == 1 for pair in observed_pairs)
            and len(publisher_rows) == len(publisher_signature_set)
        ):
            family_complete_cartesian += 1

    expected_histogram = {
        int(size): count for size, count in geometry["family_size_histogram"].items()
    }
    expected_family_count = geometry["publisher_family_count"]
    publisher_pass = (
        publisher["body_row_count"] == geometry["publisher_endpoint_row_count"]
        and publisher["valid_row_count"] == geometry["publisher_endpoint_row_count"]
        and publisher["unique_endpoint_signature_count"]
        == geometry["publisher_unique_endpoint_signature_count"]
        and publisher["family_count"] == expected_family_count
        and publisher["group_label_count"] == geometry["publisher_group_label_count"]
        and publisher["family_size_histogram"] == expected_histogram
        and publisher["duplicate_endpoint_signature_count"] == 0
        and publisher["malformed_row_count"] == 0
        and publisher["invalid_numeric_row_count"] == 0
    )
    geo_pass = (
        geo["body_row_count"] == geometry["geo_body_row_count"]
        and geo["valid_row_count"] == geometry["geo_body_row_count"]
        and geo["unique_candidate_name_count"]
        == geometry["geo_unique_candidate_name_count"]
        and geo["family_count"] == geometry["geo_family_count"]
        and geo["group_label_count"] == geometry["publisher_group_label_count"]
        and geo["candidate_count_by_family_histogram"] == expected_histogram
        and geo["candidate_cross_family_conflict_count"] == 0
        and geo["malformed_row_count"] == 0
        and geo["invalid_numeric_row_count"] == 0
    )
    cartesian_pass = (
        len(all_families) == expected_family_count
        and family_endpoint_candidate_count_match == expected_family_count
        and family_signature_set_match == expected_family_count
        and family_square_row_count_match == expected_family_count
        and family_complete_cartesian == expected_family_count
        and publisher_signatures_not_in_geo == 0
        and geo_signatures_not_in_publisher == 0
        and geo["duplicate_candidate_endpoint_pair_row_count"] == 0
        and expected_geo_rows_from_publisher_family_squares
        == geometry["geo_body_row_count"]
    )
    rights = protocol["rights_policy"]
    rights_pass = (
        rights["private_academic_analysis_allowed"] is True
        and rights["aggregate_derived_reporting_allowed"] is True
        and rights["member_payload_redistribution_allowed"] is False
        and rights["submitter_intellectual_property_transfer_asserted"] is False
    )
    boundary_pass = (
        geometry["row_order_pairing_permitted"] is False
        and geometry["sequence_pairing_permitted"] is False
        and protocol["ordinary_public_asset_contract"]["reference_fasta"][
            "read_by_this_successor"
        ]
        is False
    )

    gates = {
        GATE_IDS[0]: _gate(
            PASS if publisher_pass else FAIL,
            "PUBLISHER_PRIMARY_955_ROW_ENDPOINT_UNIVERSE_AND_FAMILY_GEOMETRY_CLOSED"
            if publisher_pass
            else "PUBLISHER_ENDPOINT_COUNT_SCHEMA_OR_FAMILY_GEOMETRY_CONFLICT",
        ),
        GATE_IDS[1]: _gate(
            PASS if geo_pass else FAIL,
            "GEO_AGGREGATE_ROWS_CANDIDATE_CARDINALITY_AND_FAMILY_GEOMETRY_CLOSED"
            if geo_pass
            else "GEO_AGGREGATE_COUNT_OR_FAMILY_GEOMETRY_CONFLICT",
        ),
        GATE_IDS[2]: _gate(
            PASS if cartesian_pass else FAIL,
            "GEO_IS_WITHIN_FAMILY_CANDIDATE_BY_PUBLISHER_ENDPOINT_CARTESIAN_EXPANSION"
            if cartesian_pass
            else "PUBLISHER_GEO_ENDPOINT_COVERAGE_OR_CARTESIAN_SEMANTICS_CONFLICT",
        ),
        GATE_IDS[3]: _gate(
            PASS if rights_pass else UNKNOWN,
            "PRIVATE_ANALYSIS_AND_AGGREGATE_REPORTING_BOUND_REDISTRIBUTION_FORBIDDEN"
            if rights_pass
            else "PRIVATE_ANALYSIS_OR_AGGREGATE_REPORTING_RIGHTS_NOT_BOUND",
        ),
        GATE_IDS[4]: _gate(
            PASS if boundary_pass else FAIL,
            "ROW_ORDER_NOT_USED_SEQUENCE_NOT_READ_AND_PAIRING_INFERENCE_PROHIBITED"
            if boundary_pass
            else "ROW_ORDER_OR_SEQUENCE_PAIRING_BOUNDARY_CONFLICT",
        ),
        GATE_IDS[5]: _gate(
            UNKNOWN,
            "MOESM7_HAS_NO_CANDIDATE_IDENTIFIER_OR_SEQUENCE_AND_GEO_CARTESIAN_ROWS_DO_NOT_IDENTIFY_PAIRING",
        ),
    }
    closed_gates = GATE_IDS[:5]
    status = (
        STATUS_CLOSED
        if all(gates[gate_id]["status"] == PASS for gate_id in closed_gates)
        else STATUS_STOP
    )

    return {
        "schema_version": SCHEMA_VERSION,
        "protocol_id": PROTOCOL_ID,
        "dataset_id": DATASET_ID,
        "bioproject_id": BIOPROJECT_ID,
        "decision_id": DECISION_ID,
        "protocol_status": protocol["protocol_status"],
        "status": status,
        "aggregate_geometry": {
            "publisher_endpoint_universe": {
                "body_row_count": publisher["body_row_count"],
                "valid_row_count": publisher["valid_row_count"],
                "unique_endpoint_signature_count": publisher[
                    "unique_endpoint_signature_count"
                ],
                "duplicate_endpoint_signature_count": publisher[
                    "duplicate_endpoint_signature_count"
                ],
                "family_count": publisher["family_count"],
                "group_label_count": publisher["group_label_count"],
                "family_size_histogram": {
                    str(size): count
                    for size, count in publisher["family_size_histogram"].items()
                },
                "malformed_row_count": publisher["malformed_row_count"],
                "invalid_numeric_row_count": publisher["invalid_numeric_row_count"],
                "one_row_per_perfect_reporter_endpoint_universe": publisher_pass,
            },
            "geo_processed_geometry": {
                "body_row_count": geo["body_row_count"],
                "valid_row_count": geo["valid_row_count"],
                "unique_candidate_name_count": geo["unique_candidate_name_count"],
                "family_count": geo["family_count"],
                "group_label_count": geo["group_label_count"],
                "unique_endpoint_signature_count": geo[
                    "unique_endpoint_signature_count"
                ],
                "unique_candidate_endpoint_pair_count": geo[
                    "unique_candidate_endpoint_pair_count"
                ],
                "duplicate_candidate_endpoint_pair_row_count": geo[
                    "duplicate_candidate_endpoint_pair_row_count"
                ],
                "candidate_cross_family_conflict_count": geo[
                    "candidate_cross_family_conflict_count"
                ],
                "candidate_count_by_family_histogram": {
                    str(size): count
                    for size, count in geo[
                        "candidate_count_by_family_histogram"
                    ].items()
                },
                "malformed_row_count": geo["malformed_row_count"],
                "invalid_numeric_row_count": geo["invalid_numeric_row_count"],
            },
            "publisher_geo_comparison": {
                "family_union_count": len(all_families),
                "family_endpoint_candidate_count_match_count": (
                    family_endpoint_candidate_count_match
                ),
                "family_endpoint_signature_set_match_count": family_signature_set_match,
                "family_square_row_count_match_count": family_square_row_count_match,
                "family_complete_cartesian_count": family_complete_cartesian,
                "publisher_endpoint_signatures_not_in_geo_count": (
                    publisher_signatures_not_in_geo
                ),
                "geo_endpoint_signatures_not_in_publisher_count": (
                    geo_signatures_not_in_publisher
                ),
                "expected_geo_rows_from_publisher_family_squares": (
                    expected_geo_rows_from_publisher_family_squares
                ),
                "numeric_signature_decimal_places": decimal_places,
                "row_order_pairing_used": False,
                "sequence_pairing_used": False,
                "source_to_candidate_mapping_identifiable": False,
            },
        },
        "gates": {gate_id: gates[gate_id] for gate_id in GATE_IDS},
        "internal_access_attestation": {
            "ordinary_public_assets_read_count": 2,
            "sequence_asset_read_count": 0,
            "derived_effect_value_cell_read_count": 0,
            "private_or_restricted_asset_read_count": 0,
            "sealed_asset_contact_count": 0,
            "member_identifier_sequence_or_row_payload_output_count": 0,
            "row_or_member_material_persisted_count": 0,
            "split_assignment_output_count": 0,
            "training_run_count": 0,
            "gpu_run_count": 0,
            "model_selection_count": 0,
        },
        "claim_boundary": copy.deepcopy(protocol["claim_boundary"]),
    }


def _json_bytes(value: Any) -> bytes:
    try:
        return (
            json.dumps(
                value,
                ensure_ascii=True,
                sort_keys=True,
                indent=2,
                allow_nan=False,
            )
            + "\n"
        ).encode("utf-8")
    except (TypeError, ValueError) as exc:
        raise OutputError("aggregate report is not finite JSON") from exc


def _existing_payload_is_exact(output: Path, payload: bytes) -> bool:
    try:
        return output.read_bytes() == payload
    except OSError as exc:
        raise OutputError("existing aggregate output is unreadable") from exc


def _fsync_directory(path: Path) -> None:
    descriptor = os.open(
        path,
        os.O_RDONLY | getattr(os, "O_DIRECTORY", 0),
    )
    try:
        os.fsync(descriptor)
    finally:
        os.close(descriptor)


def _publish_no_replace(output: Path, payload: bytes) -> str:
    """Publish one fully durable payload without ever replacing a final path."""

    if output.exists():
        if _existing_payload_is_exact(output, payload):
            return "EXISTING_EXACT_PAYLOAD_ACCEPTED"
        raise OutputError("aggregate output exists with different payload")

    temporary_path: Path | None = None
    final_created = False
    try:
        output.parent.mkdir(parents=True, exist_ok=True)
        descriptor, temporary = tempfile.mkstemp(
            prefix=f".{output.name}.",
            suffix=".tmp",
            dir=output.parent,
        )
        temporary_path = Path(temporary)
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())

        try:
            # Same-directory hard-link publication is atomic and fails if the
            # final name already exists.  Unlike replace/rename, it cannot
            # overwrite operator material.
            os.link(temporary_path, output)
            final_created = True
        except FileExistsError:
            if not _existing_payload_is_exact(output, payload):
                raise OutputError("aggregate output appeared with different payload")
            temporary_path.unlink()
            temporary_path = None
            return "RACING_EXISTING_EXACT_PAYLOAD_ACCEPTED"

        temporary_path.unlink()
        temporary_path = None
        _fsync_directory(output.parent)
        return "NEW_ATOMIC_NOREPLACE_PUBLICATION"
    except OutputError as exc:
        error: OutputError = exc
        cause: OSError | None = None
    except OSError as exc:
        error = OutputError("cannot publish aggregate-only output")
        cause = exc

    cleanup_error: OSError | None = None
    if final_created:
        try:
            output.unlink(missing_ok=True)
        except OSError as exc:
            cleanup_error = exc
    if temporary_path is not None:
        try:
            temporary_path.unlink(missing_ok=True)
        except OSError as exc:
            cleanup_error = cleanup_error or exc
    if cleanup_error is not None:
        raise OutputError("publication failure cleanup did not complete") from cleanup_error
    if cause is not None:
        raise error from cause
    raise error


BindingAuditor = Callable[
    [Mapping[str, Any], Path, bytes, Path], Mapping[str, str]
]


def execute(
    protocol_path: Path,
    publisher_workbook: Path,
    geo_perfect_csv: Path,
    output: Path,
    *,
    repo_root: Path | None = None,
    binding_auditor: BindingAuditor = _default_binding_auditor,
) -> dict[str, Any]:
    try:
        protocol_payload = protocol_path.read_bytes()
    except OSError as exc:
        raise ProtocolError("protocol is unreadable") from exc
    protocol = _strict_json(protocol_payload)
    _validate_protocol(protocol)
    if output.name != REPORT_FILENAME:
        raise OutputError("output basename differs from the sole aggregate artifact")
    bound_repo_root = repo_root or protocol_path.parent.parent
    binding = binding_auditor(
        protocol,
        protocol_path,
        protocol_payload,
        bound_repo_root,
    )
    report = aggregate(protocol, publisher_workbook, geo_perfect_csv)
    report["implementation_binding"] = dict(binding)
    payload = _json_bytes(report)
    _publish_no_replace(output, payload)
    return report


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--protocol", required=True, type=Path)
    parser.add_argument("--publisher-endpoint-workbook", required=True, type=Path)
    parser.add_argument("--geo-perfect-csv", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    arguments = parser.parse_args()
    report = execute(
        arguments.protocol,
        arguments.publisher_endpoint_workbook,
        arguments.geo_perfect_csv,
        arguments.output,
    )
    print(
        json.dumps(
            {
                "status": report["status"],
                "protocol_status": report["protocol_status"],
                "output": str(arguments.output),
            },
            sort_keys=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
