#!/usr/bin/env python3
"""Convert E-MTAB-10902 publisher-qualified N-zip designs to Evaluation records."""

from __future__ import annotations

import argparse
import json
import math
import os
import tempfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
CONFIG_PATH = REPO_ROOT / "configs/route_a_v3_route2_emtab10902_converter_v1.json"
CANONICAL_SCHEMA_VERSION = "mrna_editflow_route2_canonical.v1"
BASES = set("ACGT")
PRIMARY_ENDPOINT = "Mean_log2ratio_NeuriteSoma_WT"


class ConversionError(RuntimeError):
    pass


def _require(condition: bool, message: str) -> None:
    if not condition:
        raise ConversionError(message)


def load_config(path: Path = CONFIG_PATH) -> dict[str, Any]:
    config = json.loads(path.read_text(encoding="utf-8"))
    validate_config(config)
    return config


def validate_config(config: Mapping[str, Any]) -> None:
    _require(config["schema_version"] == "route_a_v3_route2_emtab10902_converter.v1", "unexpected schema version")
    study = config["study"]
    _require(study["study_unit_id"] == "E-MTAB-10902", "unexpected study")
    _require(study["pool_assignment"] == "EVALUATION", "E-MTAB-10902 left Evaluation")
    _require(study["qualification_class"] == "EVALUATION_RESERVED", "qualification class changed")
    _require(study["study_role"] == "TRUE_A2_EXPLORATORY_ZERO_SHOT", "study role changed")
    _require(study["region"] == "3UTR", "region changed")
    _require(study["endpoint_direction"] == "HIGHER_IS_BETTER", "endpoint direction changed")
    input_spec = config["input"]
    _require(input_spec["expected_design_row_count"] == 6266, "publisher design count changed")
    _require(input_spec["expected_publisher_qc_passed_design_count"] == 5679, "publisher QC count changed")
    _require(input_spec["expected_family_count"] == 24, "publisher family count changed")
    _require(input_spec["expected_ambiguous_join_family_count"] == 2, "ambiguous family count changed")
    _require(input_spec["expected_sourceless_family_count"] == 4, "sourceless family count changed")
    _require(input_spec["expected_geometry_eligible_family_count"] == 18, "eligible family count changed")
    _require(input_spec["expected_geometry_eligible_candidate_row_count"] == 5737, "candidate geometry changed")
    _require(input_spec["expected_retained_duplicate_sequence_group_count"] == 5, "duplicate sequence geometry changed")
    expected = config["expected_conversion"]
    _require(expected["canonical_record_count"] > 0, "canonical count is not frozen")
    _require(expected["retained_source_family_count"] > 0, "retained source family count is not frozen")
    _require(sum(expected["edit_count_histogram"].values()) == expected["canonical_record_count"], "edit histogram does not close")
    policy = config["evaluation_policy"]
    for key in (
        "training_eligible",
        "model_selection_eligible",
        "hpo_eligible",
        "threshold_selection_eligible",
        "zero_shot_result_recorded",
    ):
        _require(policy[key] is False, f"Evaluation policy incorrectly enabled: {key}")
    _require(policy["predictor_generator_and_baselines_frozen"] is False, "Evaluation was prematurely unfrozen")
    action = config["action_policy"]
    _require(action == {
        "allowed_candidate_action": "SUB",
        "multi_step_sub_trajectory_supported": True,
        "same_position_repeated_edit_required": False,
        "ins_supported": False,
        "del_supported": False,
    }, "action policy changed")
    credit = config["credit_policy"]
    _require(credit["measured_candidate"] is True and credit["generated_candidate"] is False, "candidate role changed")
    _require(not any(credit["qualified_credit_delta"].values()), "Evaluation conversion increases qualified credit")
    _require(credit["qualified_counts_after_conversion"] == {"ordinary": 1, "a1": 1, "true_a2": 0, "canonical_records": 6547}, "qualified facts changed")
    output = config["output"]
    _require(output["overwrite_allowed"] is False, "successful output overwrite enabled")
    _require(output["public_redistribution_allowed"] is False, "private Evaluation output enabled for redistribution")
    _require(output["directory"].startswith("/mnt/cunyuliu/mrna_xeditflow_routea_v3/route2/"), "output leaves Route 2 root")
    _require(config["scientific_claim_status"] == "NOT_ESTABLISHED", "scientific claim overstated")


def _sheet_rows(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    _require(sheet_name in workbook.sheetnames, f"publisher sheet absent: {sheet_name}")
    rows = workbook[sheet_name].iter_rows(values_only=True)
    next(rows)
    next(rows)
    header = tuple(next(rows))
    return [dict(zip(header, row)) for row in rows if any(value is not None for value in row)]


def _key_a(row: Mapping[str, Any]) -> tuple[str, str, str, str]:
    return (
        str(row["Source gene id"]),
        str(row["Source gene name"]),
        str(row["Mutation type"]),
        str(row["Mutation position"]),
    )


def _key_b(row: Mapping[str, Any]) -> tuple[str, str, str, str]:
    return (
        str(row["source_gene_id"]),
        str(row["source_gene_name"]),
        str(row["mutation_type"]),
        str(row["mutation_position"]),
    )


def _family(row: Mapping[str, Any]) -> tuple[str, str, str]:
    return (
        str(row["Source gene id"]),
        str(row["Source gene name"]),
        str(row["Source tile id"]),
    )


def _finite(value: Any, label: str) -> float:
    _require(isinstance(value, (int, float)) and not isinstance(value, bool), f"{label} is not numeric")
    number = float(value)
    _require(math.isfinite(number), f"{label} is not finite")
    return number


def _load_qc(summary_path: Path, membership_path: Path, config: Mapping[str, Any]) -> dict[int, dict[str, Any]]:
    summary = json.loads(summary_path.read_text(encoding="utf-8"))
    expected_passed = config["input"]["expected_publisher_qc_passed_design_count"]
    _require(summary["status"] == "PUBLISHER_READ_QC_REPRODUCED", "publisher read QC was not reproduced")
    _require(summary["evaluation_outcome_sheet_read"] is False, "QC membership used Evaluation outcomes")
    _require(summary["passed_design_row_count"] == expected_passed, "publisher read QC count changed")
    rows = [json.loads(line) for line in membership_path.read_text(encoding="utf-8").splitlines()]
    _require(len(rows) == config["input"]["expected_design_row_count"], "QC membership row count changed")
    result = {int(row["design_row_number"]): row for row in rows}
    _require(len(result) == len(rows), "QC membership design row is duplicated")
    _require(sum(bool(row["passes_publisher_read_qc"]) for row in rows) == expected_passed, "QC membership pass count changed")
    return result


def _load_units(
    workbook_path: Path,
    qc: Mapping[int, Mapping[str, Any]],
    config: Mapping[str, Any],
) -> tuple[list[dict[str, Any]], Counter[str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    rows_a = _sheet_rows(workbook, "Supp_Table_2a")
    rows_b = _sheet_rows(workbook, "Supp_Table_2b")
    expected_rows = config["input"]["expected_design_row_count"]
    _require(len(rows_a) == len(rows_b) == expected_rows, "publisher table row count changed")

    by_key_a: dict[tuple[str, ...], list[dict[str, Any]]] = defaultdict(list)
    by_key_b: dict[tuple[str, ...], list[dict[str, Any]]] = defaultdict(list)
    for row in rows_a:
        by_key_a[_key_a(row)].append(row)
    for row in rows_b:
        by_key_b[_key_b(row)].append(row)
    _require(set(by_key_a) == set(by_key_b), "publisher metadata join key set changed")
    ambiguous_keys = {key for key in by_key_a if len(by_key_a[key]) != 1 or len(by_key_b[key]) != 1}
    ambiguous_gene_names = {key[1] for key in ambiguous_keys}
    _require(len(ambiguous_keys) == 7 and ambiguous_gene_names == {"Cflar_2"}, "publisher ambiguous join geometry changed")
    outcome_by_key = {key: values[0] for key, values in by_key_b.items() if key not in ambiguous_keys}

    families: dict[tuple[str, str, str], list[tuple[int, dict[str, Any], str]]] = defaultdict(list)
    for row_number, row in enumerate(rows_a, start=1):
        sequence = str(row["Sequence"]).upper()
        _require(len(sequence) == int(row["size"]) and len(sequence) in {85, 90, 100}, "publisher sequence length changed")
        _require(not (set(sequence) - BASES), "publisher sequence alphabet changed")
        _require(qc[row_number]["sequence_id"].startswith("NZSEQ"), "QC sequence id changed")
        families[_family(row)].append((row_number, row, sequence))
    _require(len(families) == config["input"]["expected_family_count"], "publisher family count changed")

    stats: Counter[str] = Counter()
    units: list[dict[str, Any]] = []
    for family, family_rows in sorted(families.items()):
        if family[1] in ambiguous_gene_names:
            stats["reject_ambiguous_join_family_count"] += 1
            stats["reject_ambiguous_join_design_row_count"] += len(family_rows)
            continue
        sources = [item for item in family_rows if item[1]["Mutation type"] == "WT"]
        if len(sources) != 1:
            stats["reject_sourceless_family_count"] += 1
            stats["reject_sourceless_design_row_count"] += len(family_rows)
            continue
        source_row_number, source_row, source_sequence = sources[0]
        source_qc_pass = bool(qc[source_row_number]["passes_publisher_read_qc"])
        source_value = _finite(outcome_by_key[_key_a(source_row)][PRIMARY_ENDPOINT], "source endpoint")

        candidate_groups: dict[str, list[tuple[int, dict[str, Any]]]] = defaultdict(list)
        for row_number, row, sequence in family_rows:
            if row["Mutation type"] != "WT":
                candidate_groups[sequence].append((row_number, row))
        for candidate_sequence, candidate_rows in sorted(candidate_groups.items()):
            qc_states = {bool(qc[row_number]["passes_publisher_read_qc"]) for row_number, _row in candidate_rows}
            _require(len(qc_states) == 1, "identical candidate sequences have different read QC membership")
            values = [_finite(outcome_by_key[_key_a(row)][PRIMARY_ENDPOINT], "candidate endpoint") for _row_number, row in candidate_rows]
            _require(len(set(values)) == 1, "identical retained candidate sequences have different endpoints")
            stats["geometry_eligible_candidate_sequence_count"] += 1
            stats["geometry_eligible_candidate_design_row_count"] += len(candidate_rows)
            if not next(iter(qc_states)):
                stats["reject_candidate_read_qc_sequence_count"] += 1
                stats["reject_candidate_read_qc_design_row_count"] += len(candidate_rows)
                continue
            if not source_qc_pass:
                stats["reject_source_read_qc_after_candidate_pass_sequence_count"] += 1
                stats["reject_source_read_qc_after_candidate_pass_design_row_count"] += len(candidate_rows)
                continue
            _require(len(candidate_sequence) == len(source_sequence), "candidate sequence length changed")
            changes = [index for index, pair in enumerate(zip(source_sequence, candidate_sequence)) if pair[0] != pair[1]]
            _require(changes, "zero-edit candidate appeared")
            units.append({
                "family": family,
                "source_design_row_number": source_row_number,
                "source_sequence_id": qc[source_row_number]["sequence_id"],
                "source_sequence": source_sequence,
                "source_endpoint_value": source_value,
                "candidate_design_row_numbers": [row_number for row_number, _row in candidate_rows],
                "candidate_sequence_id": qc[candidate_rows[0][0]]["sequence_id"],
                "candidate_sequence": candidate_sequence,
                "candidate_endpoint_value": values[0],
                "mutation_labels": [f"{row['Mutation type']}|{row['Mutation position']}" for _row_number, row in candidate_rows],
                "changes": changes,
            })
            stats["canonical_record_count"] += 1
            stats[f"edit_bin::{_edit_bin(len(changes))}"] += 1
    stats["retained_source_family_count"] = len({unit["family"] for unit in units})
    return units, stats


def _edit_bin(edit_count: int) -> str:
    if edit_count <= 2:
        return str(edit_count)
    if edit_count <= 5:
        return "3-5"
    if edit_count <= 10:
        return "6-10"
    return "11+"


def _canonical_records(config: Mapping[str, Any], units: list[dict[str, Any]]) -> list[dict[str, Any]]:
    study = config["study"]
    policy = config["evaluation_policy"]
    credit = config["credit_policy"]
    records: list[dict[str, Any]] = []
    for unit in units:
        gene_id, gene_name, tile_id = unit["family"]
        changes = unit["changes"]
        source = unit["source_sequence"]
        candidate = unit["candidate_sequence"]
        first_row = min(unit["candidate_design_row_numbers"])
        records.append({
            "schema_version": CANONICAL_SCHEMA_VERSION,
            "canonical_record_id": f"E-MTAB-10902:{gene_id}:tile{tile_id}:row{first_row}",
            "study_unit_id": study["study_unit_id"],
            "independent_study_group_id": study["independent_study_group_id"],
            "publication_doi": study["publication_doi"],
            "pool_assignment": study["pool_assignment"],
            "qualification_class": study["qualification_class"],
            "study_role": study["study_role"],
            "region": study["region"],
            "biological_context_id": study["biological_context_id"],
            "context": {
                "organism": "Mus musculus",
                "cell_type": "primary cortical neuron",
                "compartments": ["neurite", "soma"],
                "condition": "WT",
                "library": "secondary N-zip",
            },
            "assay_id": study["assay_id"],
            "assay_type": study["assay_type"],
            "endpoint_id": study["endpoint_id"],
            "endpoint_direction": study["endpoint_direction"],
            "source_id": f"E-MTAB-10902:{gene_id}:tile{tile_id}:WT",
            "gene_group_id": gene_id,
            "source_metadata": {
                "source_gene_name": gene_name,
                "source_tile_id": tile_id,
                "publisher_design_row_number": unit["source_design_row_number"],
                "publisher_sequence_id": unit["source_sequence_id"],
                "publisher_mean_log2ratio_neurite_soma_wt": unit["source_endpoint_value"],
            },
            "source_sequence": source,
            "candidate_id": f"E-MTAB-10902:{gene_id}:tile{tile_id}:row{first_row}",
            "candidate_metadata": {
                "publisher_design_row_numbers": unit["candidate_design_row_numbers"],
                "publisher_sequence_id": unit["candidate_sequence_id"],
                "mutation_labels": unit["mutation_labels"],
                "identical_design_row_count": len(unit["candidate_design_row_numbers"]),
                "publisher_mean_log2ratio_neurite_soma_wt": unit["candidate_endpoint_value"],
            },
            "candidate_sequence": candidate,
            "edit_operations": [
                {"type": "SUB", "position_zero_based": index, "ref": source[index], "alt": candidate[index]}
                for index in changes
            ],
            "multi_step_sub_trajectory": len(changes) > 1,
            "same_position_repeated_edit_required": False,
            "source_endpoint_value": unit["source_endpoint_value"],
            "candidate_endpoint_value": unit["candidate_endpoint_value"],
            "direction_normalized_delta": unit["candidate_endpoint_value"] - unit["source_endpoint_value"],
            "biological_standard_error": None,
            "biological_replicate_count": 3,
            "standard_error_status": "PUBLISHER_DID_NOT_REPORT_ROW_LEVEL_STANDARD_ERROR",
            "measured_candidate": credit["measured_candidate"],
            "generated_candidate": credit["generated_candidate"],
            "evaluation_outcome_access_stage": policy["outcome_access_stage"],
            "training_eligible": policy["training_eligible"],
            "model_selection_eligible": policy["model_selection_eligible"],
            "hpo_eligible": policy["hpo_eligible"],
            "threshold_selection_eligible": policy["threshold_selection_eligible"],
            "zero_shot_result_recorded": policy["zero_shot_result_recorded"],
            "provenance": {
                "study_accession": "E-MTAB-10902",
                "publisher_supplementary_table": "Supplementary Table 2",
                "publisher_read_qc_reconstructed_from_primary_wt_secondary_fastq": True,
                "publisher_read_qc_rule": "at least 20 reads in at least 3 of 6 samples",
            },
        })
    return records


def _write_json(path: Path, value: Mapping[str, Any]) -> None:
    path.write_text(json.dumps(value, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def execute(
    config: Mapping[str, Any],
    workbook_path: Path,
    qc_summary_path: Path,
    qc_membership_path: Path,
    output_dir: Path,
) -> dict[str, Any]:
    _require(not output_dir.exists(), f"output directory already exists: {output_dir}")
    qc = _load_qc(qc_summary_path, qc_membership_path, config)
    units, stats = _load_units(workbook_path, qc, config)
    records = _canonical_records(config, units)
    expected = config["expected_conversion"]
    observed_histogram = {
        key.removeprefix("edit_bin::"): value
        for key, value in sorted(stats.items())
        if key.startswith("edit_bin::")
    }
    _require(len(records) == expected["canonical_record_count"], "canonical count differs from frozen conversion")
    _require(stats["retained_source_family_count"] == expected["retained_source_family_count"], "retained family count differs")
    _require(observed_histogram == expected["edit_count_histogram"], "edit histogram differs")
    _require(stats["geometry_eligible_candidate_design_row_count"] == config["input"]["expected_geometry_eligible_candidate_row_count"], "candidate geometry differs")

    output_dir.parent.mkdir(parents=True, exist_ok=True)
    temporary = Path(tempfile.mkdtemp(prefix=f".{output_dir.name}.", dir=output_dir.parent))
    try:
        canonical_path = temporary / config["output"]["canonical_filename"]
        with canonical_path.open("w", encoding="utf-8") as handle:
            for record in records:
                handle.write(json.dumps(record, sort_keys=True) + "\n")
        summary = {
            "converter_id": config["converter_id"],
            "study_unit_id": config["study"]["study_unit_id"],
            "pool_assignment": config["study"]["pool_assignment"],
            "status": "CONVERTED_EVALUATION_RESERVED_TRUE_A2_EXPLORATORY",
            "canonical_record_count": len(records),
            "retained_source_family_count": stats["retained_source_family_count"],
            "edit_count_histogram": observed_histogram,
            "missing_standard_error_count": len(records),
            "evaluation_result_metric_computed": False,
            "zero_shot_result_recorded": False,
            "training_eligible_record_count": 0,
            "model_selection_eligible_record_count": 0,
            "qualified_credit_delta": config["credit_policy"]["qualified_credit_delta"],
            "qualified_counts_after_conversion": config["credit_policy"]["qualified_counts_after_conversion"],
            "scientific_claim_status": config["scientific_claim_status"],
        }
        reject_summary = {
            "converter_id": config["converter_id"],
            "study_unit_id": config["study"]["study_unit_id"],
            "reject_counts": {
                key: value for key, value in sorted(stats.items())
                if key.startswith("reject_")
            },
            "geometry_eligible_candidate_sequence_count": stats["geometry_eligible_candidate_sequence_count"],
            "geometry_eligible_candidate_design_row_count": stats["geometry_eligible_candidate_design_row_count"],
            "reject_payload_in_summary": False,
        }
        _write_json(temporary / config["output"]["conversion_summary_filename"], summary)
        _write_json(temporary / config["output"]["reject_summary_filename"], reject_summary)
        os.rename(temporary, output_dir)
        return summary
    finally:
        if temporary.exists():
            for child in temporary.iterdir():
                child.unlink()
            temporary.rmdir()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", type=Path, default=CONFIG_PATH)
    parser.add_argument("--workbook", type=Path)
    parser.add_argument("--qc-summary", type=Path)
    parser.add_argument("--qc-membership", type=Path)
    parser.add_argument("--output-dir", type=Path)
    args = parser.parse_args()
    config = load_config(args.config)
    input_spec = config["input"]
    summary = execute(
        config,
        args.workbook or Path(input_spec["publisher_workbook_path"]),
        args.qc_summary or Path(input_spec["qc_membership_summary_path"]),
        args.qc_membership or Path(input_spec["qc_membership_path"]),
        args.output_dir or Path(config["output"]["directory"]),
    )
    print(json.dumps(summary, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
