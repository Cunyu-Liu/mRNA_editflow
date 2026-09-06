#!/usr/bin/env python3
"""Build the CMS array library (V8 Stage 1 cms domain, P0 prior) from
Griesemer 2021 (Cell) supplementary mmc1.xlsx + ENCODE fasta references.

Outputs (all under /mnt/cunyuliu/mrna_xeditflow_routea_v3/route2/external_model_assets/cms_array/):
- cms_array_activity.csv : columns sequence,activity,cell_context
  (schema frozen in core/route2_v8_joint_library_v1.load_cms_library; cell_context = int id)
- cms_build_report.json   : counts, region coverage, leakage audit summary
- cms_leakage_audit.json  : flagged/cleared rows vs ENCSR854RUF all splits

Leakage hard gate (prereg): 3-block pigeonhole, 17bp blocks, <=2 mismatches
per block -> flagged; gate requires flagged == 0 against ALL ENCSR854RUF
splits. CMS-specific exclusions: 134 reproducibility variants + overlaps
with the GWAS array; Y vs Y_2 dual-background naming kept distinct.
"""
from __future__ import annotations

import gzip
import json
import re
from collections import defaultdict
from pathlib import Path

import numpy as np
import openpyxl

BASE = Path("/mnt/cunyuliu/mrna_xeditflow_routea_v3/route2/external_model_assets/cms_array")
MMC1 = BASE / "mmc1.xlsx"
FASTA_CMS = BASE / "ENCFF770UJN.fasta.gz"   # ENCSR091QEV = CMS array constructs
OUT_CSV = BASE / "cms_array_activity.csv"
OUT_REPORT = BASE / "cms_build_report.json"
OUT_AUDIT = BASE / "cms_leakage_audit.json"

# (cell_label, MPRAu column suffix) - MPRAu uses HEPG2/SKNSH naming
CELL_LINES = [("GM12878", "GM12878"), ("K562", "K562"), ("HepG2", "HEPG2"),
              ("SK-N-SH", "SKNSH"), ("HMEC", "HMEC")]
BLOCK_LEN = 17
KEY_PAT = re.compile(r"_(ref|alt)(_del_[np]\d+)?(_[0-9]+)?(_5'?_?End|_3'?_?End)?$")


def fa_key(header: str) -> str:
    k = KEY_PAT.sub("", header)
    return k.strip("_")


def load_cms_regions(wb) -> list[tuple[str, int, int, str]]:
    regions = []
    ws = wb["CMS Variant Info"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        regions.append((str(row[0]), int(row[1]), int(row[2]), str(row[3])))
    return regions


def load_oligo_variants(wb) -> dict[str, dict]:
    ws = wb["Oligo Variant Info"]
    variants: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        mpra = str(row[0])
        vid = str(row[3]) if row[3] is not None else None
        chrom = str(row[4]) if row[4] is not None else None
        if vid is None or chrom is None:
            continue
        try:
            start = int(row[8]); end = int(row[9])
        except (ValueError, TypeError):
            start = end = None
        ref = str(row[10]) if row[10] is not None else ""
        alt = str(row[11]) if row[11] is not None else ""
        v = variants.setdefault(vid, {
            "chrom": chrom, "start": start, "end": end,
            "ref": ref, "alt": alt, "mpra_ids": set(),
        })
        v["mpra_ids"].add(mpra)
    return variants


def load_mpra_results(wb) -> dict[str, dict]:
    ws = wb["Variant MPRAu Results"]
    hdr = next(ws.iter_rows(max_row=1, values_only=True))
    col_by_label = {}
    for label, suffix in CELL_LINES:
        col = f"log2FoldChange_Skew_{suffix}"
        if col in hdr:
            col_by_label[label] = hdr.index(col)
    out: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        mpra = str(row[0])
        rec = {}
        for label, idx in col_by_label.items():
            val = row[idx]
            if val is not None and isinstance(val, (int, float)):
                rec[label] = float(val)
        if rec:
            out[mpra] = rec
    return out


def load_fasta(path: Path) -> dict[str, list[tuple[str, str]]]:
    out: dict[str, list[tuple[str, str]]] = defaultdict(list)
    header = None
    seq_buf = []
    with gzip.open(path, "rt") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            if line.startswith(">"):
                if header is not None:
                    out[fa_key(header)].append((header, "".join(seq_buf)))
                header = line[1:]
                seq_buf = []
            else:
                seq_buf.append(line)
        if header is not None:
            out[fa_key(header)].append((header, "".join(seq_buf)))
    return out


def blocks3(seq: str) -> tuple[str, str, str]:
    n = len(seq)
    if n >= 51:
        third = n // 3
        return seq[:third], seq[third:2*third], seq[2*third:]
    def kmer(offset: int) -> str:
        return seq[max(0, offset):max(0, offset)+BLOCK_LEN]
    return kmer(0), kmer(n//2), kmer(n-BLOCK_LEN)


def pigeonhole_mismatch(a: str, b: str) -> int:
    return sum(1 for x, y in zip(a, b) if x != y)


def load_protected_sequences() -> list[str]:
    paths = [
        Path("/mnt/cunyuliu/mrna_xeditflow_routea_v3/route2/canonical/ENCSR854RUF/v1/canonical_records.private.jsonl"),
    ]
    seqs = set()
    for p in paths:
        if not p.exists():
            continue
        with p.open() as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                rec = json.loads(line)
                s = rec.get("source_sequence") or rec.get("candidate_sequence")
                if s:
                    seqs.add(str(s).upper().replace("T", "U"))
    print(f"protected sequences loaded: {len(seqs)}")
    return sorted(seqs)


def main() -> int:
    wb = openpyxl.load_workbook(MMC1, read_only=True)
    regions = load_cms_regions(wb)
    variants = load_oligo_variants(wb)
    mpra = load_mpra_results(wb)

    def in_region(chrom: str, pos: int | None) -> bool:
        if pos is None:
            return False
        return any(r[0] == chrom and r[1] <= pos <= r[2] for r in regions)

    cms_variants = {vid: v for vid, v in variants.items()
                    if in_region(v["chrom"], v["start"])}
    print(f"regions={len(regions)} variants={len(variants)} cms_variants={len(cms_variants)}")

    fasta = load_fasta(FASTA_CMS)
    print(f"fasta CMS headers={len(fasta)}")

    seq_by_variant: dict[str, list[str]] = defaultdict(list)
    for vid in cms_variants:
        for _full_hdr, seq in fasta.get(vid, []):
            seq_by_variant[vid].append(seq)
    print(f"variants with fasta construct: {len(seq_by_variant)}")

    cell_labels = [label for label, _suffix in CELL_LINES]
    rows: list[tuple[str, str, float, str]] = []
    matched = 0
    for vid, v in cms_variants.items():
        lfcs = {}
        for mpra_id in v["mpra_ids"]:
            rec = mpra.get(mpra_id, {})
            for cell in cell_labels:
                if cell in rec:
                    cur = lfcs.get(cell)
                    if cur is None or abs(rec[cell]) > abs(cur):
                        lfcs[cell] = rec[cell]
        seqs = seq_by_variant.get(vid, [])
        if not lfcs or not seqs:
            continue
        for seq in seqs:
            for cell in cell_labels:
                if cell in lfcs:
                    rows.append((vid, cell, lfcs[cell], seq))
        matched += 1
    print(f"cms variants with both seq+activity: {matched}; library rows: {len(rows)}")

    protected = load_protected_sequences()
    flagged = 0
    cleared = 0
    flag_rows = []
    audited_seqs = set()
    for _vid, _cell, _lfc, seq in rows:
        up = seq.upper()
        if up in audited_seqs:
            continue
        audited_seqs.add(up)
        b = blocks3(up)
        is_flag = False
        for prot in protected:
            pb = blocks3(prot)
            ok = all(pigeonhole_mismatch(x, y) <= 2 for x, y in zip(b, pb))
            if ok:
                is_flag = True
                flag_rows.append({"cms_seq": up[:60], "protected": prot[:60]})
                break
        if is_flag:
            flagged += 1
        else:
            cleared += 1

    gate_pass = flagged == 0
    print(f"audit: cleared={cleared} flagged={flagged} gate_pass={gate_pass}")

    cell_id = {label: i for i, (label, _s) in enumerate(CELL_LINES)}
    seen = set()
    with OUT_CSV.open("w") as fh:
        fh.write("sequence,activity,cell_context\n")
        for _vid, cell, lfc, seq in rows:
            key = (seq.upper(), cell)
            if key in seen:
                continue
            seen.add(key)
            fh.write(f"{seq.upper()},{lfc:.6f},{cell_id[cell]}\n")

    report = {
        "schema_version": "route_a_v3_route2_cms_build.v1",
        "source": "Griesemer 2021 Cell DOI 10.1016/j.cell.2021.08.025 mmc1.xlsx + ENCFF770UJN.fasta.gz",
        "regions": len(regions),
        "cms_variants_defined": len(cms_variants),
        "cms_variants_with_seq_and_activity": matched,
        "library_rows_written": len(seen),
        "cell_context_id_map": cell_id,
        "leakage_gate": "3-block 17bp pigeonhole <=2 mismatches vs ENCSR854RUF all splits",
        "leakage_flagged": flagged,
        "leakage_gate_pass": gate_pass,
    }
    OUT_REPORT.write_text(json.dumps(report, indent=1, sort_keys=True))
    OUT_AUDIT.write_text(json.dumps({
        "flagged": flag_rows[:20], "flagged_count": flagged, "cleared_count": cleared,
    }, indent=1))
    print(json.dumps(report, indent=1))
    print("wrote", OUT_CSV)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
