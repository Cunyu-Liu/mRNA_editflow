#!/usr/bin/env python3
"""Build the provenance-first NMI Benchmark v2 registry.

The canonical store contains two explicitly separated task families:

* ``local_delta`` records from the source-matched intervention tier (Layer C),
  which are the only records eligible as biological local-delta ground truth;
* ``absolute_property_*`` records from public absolute libraries (Layer B),
  which make context/assay axes observable without being relabelled as local
  interventions.

The builder never aliases one record into multiple final roles.  Final labels
remain physically present in the canonical store for reproducibility, but the
loader is fail-closed and refuses final manifests unless explicitly opened.
"""
from __future__ import annotations

import argparse
import csv
import difflib
import gzip
import hashlib
import json
import math
import os
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, Iterable, Iterator, List, Mapping, Optional, Sequence, Tuple


ROLES = ["train", "val", "test_id", "test_family", "test_context", "test_assay", "test_ood"]
FINAL_ROLES = set(ROLES[2:])
REQUIRED_SOURCE_MATCHED_FIELDS = [
    "source_id", "candidate_id", "source_sequence", "candidate_sequence",
    "edit_list", "edit_count", "measured_source", "measured_candidate",
    "measured_delta", "cargo", "cell_context", "assay", "batch", "replicate",
]


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def stable_fraction(value: str) -> float:
    return int(hashlib.sha256(value.encode("utf-8")).hexdigest()[:12], 16) / float(16**12)


def assign_local_role(old_role: str, source_id: str) -> str:
    """Re-use the frozen P3 source split while making all local axes distinct.

    Historical P3 ``test``/``ood`` and validation labels were exposed during
    prior development and are therefore never promoted to v2 final roles.
    They remain development/quarantine inputs only; new final roles come from
    the raw designed library below.
    """
    if old_role == "train":
        return "train"
    if old_role == "val":
        return "val"
    raise ValueError(f"unsupported P3 split role {old_role!r}")


def parse_float(value: object) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def normalize_sequence(value: object) -> str:
    return str(value or "").strip().upper().replace("T", "U")


def normalized_log2_delta(source: Optional[float], candidate: Optional[float]) -> Optional[float]:
    """Return the cross-assay scoring endpoint while retaining raw values."""
    if source is None or candidate is None or source <= -1.0 or candidate <= -1.0:
        return None
    return math.log2((candidate + 1.0) / (source + 1.0))


def normalize_p3_record(rec: Mapping[str, object], role: str) -> Dict[str, object]:
    confidence = str(rec.get("confidence", "unknown"))
    measured = confidence == "measured"
    source_id = str(rec.get("source_id", ""))
    record_id = str(rec.get("record_id", ""))
    source_sequence = normalize_sequence(rec.get("source_sequence"))
    candidate_sequence = normalize_sequence(rec.get("candidate_sequence"))
    internal = rec.get("internal_features") if isinstance(rec.get("internal_features"), dict) else {}
    measured_source = parse_float(rec.get("measured_or_proxy_source_value")) if measured else None
    measured_candidate = parse_float(rec.get("measured_or_proxy_candidate_value")) if measured else None
    measured_delta = parse_float(rec.get("delta")) if measured else None
    edit_list = rec.get("edit_list") if isinstance(rec.get("edit_list"), list) else []
    result = dict(rec)
    result.update({
        "benchmark_version": "nmi_benchmark_v2",
        "v2_source_role": role,
        "data_layer": "C_source_matched_intervention" if measured else "B_absolute_design_library",
        "task_kind": "local_delta" if measured else "derived_candidate_asset",
        "local_delta_eligible": measured and measured_delta is not None,
        "candidate_id": f"{record_id}:candidate",
        "mother_id": rec.get("mother_id") or source_id,
        "source_sequence": source_sequence,
        "candidate_sequence": candidate_sequence,
        "edit_list": edit_list,
        "edit_count": int(rec.get("edit_count") or len(edit_list)),
        "measured_source": measured_source,
        "measured_candidate": measured_candidate,
        "measured_delta": measured_delta,
        "scored_delta": normalized_log2_delta(measured_source, measured_candidate),
        "scoring_endpoint": "log2_plus_one_fold_change_candidate_over_source",
        "cargo": rec.get("cargo_id"),
        "protein_family_id": rec.get("protein_family_id") or (
            f"reporter_family:{str(rec.get('cargo_id') or 'unknown').casefold()}"
            if rec.get("cargo_id") else None
        ),
        "cell_context": rec.get("cell_context"),
        "assay": rec.get("assay_type"),
        "batch": rec.get("data_source"),
        "replicate": internal.get("replicate", internal.get("n_wt_replicates")),
        "source_sequence_sha256": sha256_text(source_sequence),
        "candidate_sequence_sha256": sha256_text(candidate_sequence),
        "label_visibility": "hidden_before_freeze" if role in FINAL_ROLES else "development_allowed",
        "label_semantics": "wet_lab_source_matched_delta" if measured else "not_ground_truth",
    })
    return result


def iter_fasta(path: Path) -> Iterator[Tuple[str, str]]:
    header: Optional[str] = None
    chunks: List[str] = []
    with path.open() as fh:
        for raw in fh:
            line = raw.strip()
            if not line:
                continue
            if line.startswith(">"):
                if header is not None:
                    yield header, normalize_sequence("".join(chunks))
                header = line[1:]
                chunks = []
            else:
                chunks.append(line)
    if header is not None:
        yield header, normalize_sequence("".join(chunks))


def header_field(header: str, key: str) -> Optional[str]:
    prefix = key + "="
    for part in header.split("|"):
        if part.startswith(prefix):
            return part[len(prefix):]
    return None


def make_absolute_record(
    *, record_id: str, source_id: str, sequence: str, value: float,
    cargo: str, context: str, assay: str, batch: str, replicate: Optional[object],
    family: str, role: str, task_kind: str, data_source: str,
) -> Dict[str, object]:
    return {
        "benchmark_version": "nmi_benchmark_v2",
        "record_id": record_id,
        "source_id": source_id,
        "mother_id": source_id,
        "candidate_id": f"{record_id}:candidate",
        "source_sequence": sequence,
        "candidate_sequence": sequence,
        "source_sequence_sha256": sha256_text(sequence),
        "candidate_sequence_sha256": sha256_text(sequence),
        "edit_list": [],
        "edit_count": 0,
        "measured_source": value,
        "measured_candidate": value,
        "measured_delta": None,
        "delta": None,
        "measured_or_proxy_source_value": value,
        "measured_or_proxy_candidate_value": value,
        "cargo": cargo,
        "cargo_id": cargo,
        "protein_family_id": f"reporter_family:{str(cargo).casefold()}",
        "cell_context": context,
        "assay": assay,
        "assay_type": assay,
        "batch": batch,
        "replicate": replicate,
        "family_cluster_id": family,
        "confidence": "measured",
        "data_source": data_source,
        "data_layer": "B_absolute_design_library",
        "task_kind": task_kind,
        "local_delta_eligible": False,
        "label_visibility": "hidden_before_freeze" if role in FINAL_ROLES else "development_allowed",
        "label_semantics": "absolute_property_only_not_local_delta_ground_truth",
        "split_role": role,
        "v2_source_role": role,
        "task_eligibility": "absolute_only",
        "value_qualifier": "wet-lab measured absolute property; not a source-matched intervention",
    }


def make_source_matched_axis_record(
    *, record_id: str, source_id: str, source_sequence: str,
    candidate_sequence: str, measured_source: float, measured_candidate: float,
    cargo: str, context: str, assay: str, batch: str,
    family: str, role: str, task_kind: str, data_source: str,
) -> Dict[str, object]:
    """Represent an assay/context intervention without calling it a sequence edit."""
    return {
        "benchmark_version": "nmi_benchmark_v2",
        "record_id": record_id,
        "source_id": source_id,
        "mother_id": source_id,
        "candidate_id": f"{record_id}:candidate",
        "source_sequence": source_sequence,
        "candidate_sequence": candidate_sequence,
        "source_sequence_sha256": sha256_text(source_sequence),
        "candidate_sequence_sha256": sha256_text(candidate_sequence),
        "edit_list": [],
        "edit_count": 0,
        "measured_source": measured_source,
        "measured_candidate": measured_candidate,
        "measured_delta": measured_candidate - measured_source,
        "delta": measured_candidate - measured_source,
        "measured_or_proxy_source_value": measured_source,
        "measured_or_proxy_candidate_value": measured_candidate,
        "cargo": cargo,
        "cargo_id": cargo,
        "cell_context": context,
        "assay": assay,
        "assay_type": assay,
        "batch": batch,
        "replicate": None,
        "family_cluster_id": family,
        "confidence": "measured",
        "data_source": data_source,
        "data_layer": "C_source_matched_intervention",
        "task_kind": task_kind,
        "local_delta_eligible": False,
        "label_visibility": "hidden_before_freeze" if role in FINAL_ROLES else "development_allowed",
        "label_semantics": "source_matched_assay_or_context_delta; not a nucleotide edit ground truth",
        "split_role": role,
        "v2_source_role": role,
        "task_eligibility": "axis_shift_only",
        "value_qualifier": "wet-lab measured source-matched assay/context shift",
    }


def derive_edit_operations(source: str, candidate: str) -> List[Dict[str, object]]:
    """Return auditable contiguous substitutions/insertions/deletions."""
    changes: List[Dict[str, object]] = []
    matcher = difflib.SequenceMatcher(a=source, b=candidate, autojunk=False)
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == "equal":
            continue
        changes.append({
            "pos": i1,
            "ref": source[i1:i2],
            "alt": candidate[j1:j2],
            "region": "five_utr",
            "op": tag,
        })
    return changes


def derive_single_nucleotide_edits(source: str, candidate: str) -> Optional[List[Dict[str, object]]]:
    if len(source) != len(candidate):
        return None
    changes = derive_edit_operations(source, candidate)
    if all(
        item["op"] == "replace"
        and len(str(item["ref"])) == 1
        and len(str(item["alt"])) == 1
        for item in changes
    ):
        return changes
    return None


def raw_ood_dimensions(source: str) -> List[str]:
    """Return declared OOD dimensions supported by the untouched raw library."""
    gc = (source.count("G") + source.count("C")) / max(1, len(source))
    dimensions: List[str] = []
    if gc <= 0.20 or gc >= 0.80:
        dimensions.append("gc_tail")
    # Sequences are normalized to RNA alphabet before this check. A uAUG is
    # a declared motif stratum, not a claim that the motif is causally
    # beneficial or deleterious.
    if "AUG" in source:
        dimensions.append("motif_uaug")
    return dimensions


def raw_source_role(source_id: str, family_id: str, source: str) -> Optional[str]:
    """Assign only new raw-library records to untouched final roles."""
    if raw_ood_dimensions(source):
        return "test_ood"
    family_fraction = stable_fraction("family:" + family_id)
    # Raw eGFP rows remain ID/OOD development controls; the final local-delta
    # family holdout is reserved for the independently sourced tdTomato
    # intervention panel below.
    source_fraction = stable_fraction("source:" + source_id)
    if 0.20 <= source_fraction < 0.40:
        return "test_id"
    return None


def iter_raw_untouched_records(
    path: Path, *, excluded_source_hashes: set[str], excluded_candidate_hashes: set[str],
) -> Iterator[Tuple[str, Dict[str, object]]]:
    """Read source-matched rows not present in the historical P3 measured tier."""
    anchors: Dict[str, float] = {}
    rows: List[Tuple[int, Dict[str, str]]] = []
    with gzip.open(path, "rt", newline="") as fh:
        reader = csv.DictReader(fh)
        for row_i, row in enumerate(reader):
            source = normalize_sequence(row.get("mother"))
            candidate = normalize_sequence(row.get("utr"))
            value = parse_float(row.get("rl"))
            if not source or not candidate or value is None:
                continue
            rows.append((row_i, row))
            if source == candidate:
                anchors.setdefault(source, value)
    seen_pairs = set()
    source_family_by_hash: Dict[str, str] = {}
    for row_i, row in rows:
        source = normalize_sequence(row.get("mother"))
        candidate = normalize_sequence(row.get("utr"))
        if source == candidate or source not in anchors:
            continue
        source_hash = sha256_text(source)
        candidate_hash = sha256_text(candidate)
        if source_hash in excluded_source_hashes or candidate_hash in excluded_candidate_hashes:
            continue
        edit_list = derive_single_nucleotide_edits(source, candidate)
        if not edit_list or len(edit_list) > 3:
            continue
        pair_key = (source, candidate)
        if pair_key in seen_pairs:
            continue
        seen_pairs.add(pair_key)
        family_raw = str(row.get("info1") or row.get("info2") or "unknown")
        family_raw = source_family_by_hash.setdefault(source_hash, family_raw)
        family_id = f"raw_library_family:{family_raw}"
        source_id = f"raw_sample2019:{source_hash[:20]}"
        role = raw_source_role(source_id, family_id, source)
        if role is None:
            continue
        candidate_value = parse_float(row.get("rl"))
        if candidate_value is None:
            continue
        record_id = f"v2untouched:sample2019:{row_i:07d}:{candidate_hash[:12]}"
        record = {
            "benchmark_version": "nmi_benchmark_v2",
            "record_id": record_id,
            "source_id": source_id,
            "mother_id": source_id,
            "candidate_id": f"{record_id}:candidate",
            "source_sequence": source,
            "candidate_sequence": candidate,
            "source_sequence_sha256": source_hash,
            "candidate_sequence_sha256": candidate_hash,
            "edit_list": edit_list,
            "edit_count": len(edit_list),
            "measured_source": anchors[source],
            "measured_candidate": candidate_value,
            "measured_delta": candidate_value - anchors[source],
            "scored_delta": normalized_log2_delta(anchors[source], candidate_value),
            "scoring_endpoint": "log2_plus_one_fold_change_candidate_over_source",
            "delta": candidate_value - anchors[source],
            "measured_or_proxy_source_value": anchors[source],
            "measured_or_proxy_candidate_value": candidate_value,
            "cargo": "eGFP",
            "cargo_id": "eGFP",
            "protein_family_id": "reporter_family:egfp",
            "cell_context": "HEK293T",
            "assay": "MPRA_2019_designed_library_untouched",
            "assay_type": "MPRA_2019_designed_library_untouched",
            "batch": path.name,
            "replicate": 1,
            "family_cluster_id": family_id,
            "confidence": "measured",
            "data_source": f"sample2019_mpra:{path.name}:raw_row_{row_i}",
            "data_layer": "C_source_matched_intervention",
            "task_kind": "local_delta",
            "local_delta_eligible": True,
            "label_semantics": "raw_source_matched_single_edit; excluded from historical P3 membership",
            "task_eligibility": "task_a_active",
            "value_qualifier": "wet-lab measured MPRA raw row not in historical P3 measured membership",
            "split_role": role,
            "v2_source_role": role,
            "ood_dimensions": raw_ood_dimensions(source),
        }
        yield role, record



def iter_xlsx_rows(path: Path, sheet_name: str) -> Iterator[Dict[str, object]]:
    """Read a small, provenance-registered XLSX sheet without pandas state."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "GSE246381 import requires openpyxl in the benchmark environment"
        ) from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header = [str(value) if value is not None else "" for value in next(rows)]
        for row in rows:
            yield dict(zip(header, row))
    finally:
        workbook.close()


def _gse_family_id(seq_id: str) -> str:
    match = re.search(r"(?:^|;)Family=([^;]+)", seq_id)
    return f"gse246381_family:{match.group(1) if match else 'unknown'}"


def _gse_variant_base(seq_id: str, allele: str) -> str:
    # Supplement 1 SeqID already carries the leading ``Variant;`` token.
    return f"{seq_id};{allele}"


def _gzip_header_is_valid(path: Path) -> bool:
    try:
        with gzip.open(path, "rb") as fh:
            fh.read(1)
    except (OSError, EOFError):
        return False
    return True


def _load_gse_umi_cpm(path: Path, barcode_map: Mapping[str, Tuple[str, str]]) -> Tuple[Dict[Tuple[str, str], float], int]:
    """Aggregate barcode UMI counts to mean per-sample CPM by allele."""
    with gzip.open(path, "rt", newline="") as fh:
        reader = csv.reader(fh)
        header = next(reader)
        n_samples = max(0, len(header) - 1)
        totals = [0.0] * n_samples
        rows: Dict[Tuple[str, str], List[float]] = defaultdict(lambda: [0.0] * n_samples)
        for row in reader:
            if len(row) < n_samples + 1:
                continue
            values = [parse_float(value) or 0.0 for value in row[1:n_samples + 1]]
            totals = [left + right for left, right in zip(totals, values)]
            mapped = barcode_map.get(row[0])
            if mapped is not None:
                target = rows[mapped]
                for i, value in enumerate(values):
                    target[i] += value
    measurements: Dict[Tuple[str, str], float] = {}
    for key, values in rows.items():
        cpm = [value / total * 1_000_000.0 for value, total in zip(values, totals) if total > 0]
        if cpm:
            measurements[key] = sum(cpm) / len(cpm)
    return measurements, n_samples


def _load_gse_label_map(path: Path, sheet_name: str) -> Dict[str, float]:
    result: Dict[str, float] = {}
    for row in iter_xlsx_rows(path, sheet_name):
        seq_id = str(row.get("SeqID") or "")
        value = parse_float(row.get("logFC"))
        if seq_id and value is not None:
            result[seq_id] = value
    return result


def iter_gse246381_records(data_root: Path) -> Iterator[Tuple[str, Dict[str, object]]]:
    """Emit exact paired tdTomato UTR interventions from GSE246381.

    HEK and mouse endpoints are split by the stable family hash so a paired
    sequence never appears in two final roles. Absolute values are derived
    only from the deposited UMI count matrices (mean sample-normalized CPM);
    the publication's reported allelic logFC is retained as an independent
    provenance field.
    """
    base = data_root / "data/raw/gse246381_utr_mutation"
    sequence_xlsx = base / "media-1.xlsx"
    hek_labels_xlsx = base / "media-6.xlsx"
    mouse_labels_xlsx = base / "media-3.xlsx"
    hek_counts = base / "GSE246381_hek_combined_umi_counts.csv.gz"
    mouse_counts = base / "GSE246381_vglut_combined_umi_counts.csv.gz"
    if not sequence_xlsx.exists():
        return

    variant_rows = list(iter_xlsx_rows(sequence_xlsx, "S1.2 MPRA Library Variants"))
    oligo_rows = list(iter_xlsx_rows(sequence_xlsx, "S1.3 Library Oligo Sequences"))

    def make_barcode_map() -> Dict[str, Tuple[str, str]]:
        result: Dict[str, Tuple[str, str]] = {}
        for row in oligo_rows:
            oligo_id = str(row.get("SeqID") or "")
            barcode = str(row.get("BC") or "")
            if not barcode or not oligo_id.endswith((";REF", ";ALT")):
                continue
            raw_seq_id, allele = oligo_id.rsplit(";", 1)
            seq_id = raw_seq_id.removeprefix("Variant;")
            result[f"{_gse_variant_base(raw_seq_id, allele)};{barcode}"] = (seq_id, allele)
        return result
    hek_measurements: Dict[Tuple[str, str], float] = {}
    mouse_measurements: Dict[Tuple[str, str], float] = {}
    hek_sample_count = 0
    mouse_sample_count = 0
    if hek_counts.exists() and _gzip_header_is_valid(hek_counts) and hek_labels_xlsx.exists():
        hek_measurements, hek_sample_count = _load_gse_umi_cpm(hek_counts, make_barcode_map())
        hek_reported = _load_gse_label_map(
            hek_labels_xlsx, "S2.4 Polysome-Total RNA Enrichm"
        )
    else:
        hek_reported = {}
    if mouse_counts.exists() and _gzip_header_is_valid(mouse_counts) and mouse_labels_xlsx.exists():
        mouse_measurements, mouse_sample_count = _load_gse_umi_cpm(mouse_counts, make_barcode_map())
        mouse_reported = _load_gse_label_map(
            mouse_labels_xlsx, "S5.4 CreON Polysome-Total RNA"
        )
    else:
        mouse_reported = {}

    for row in variant_rows:
        seq_id = str(row.get("SeqID") or "")
        source = normalize_sequence(row.get("RefSequence"))
        candidate = normalize_sequence(row.get("AltSequence"))
        if not seq_id or not source or not candidate:
            continue
        family_id = _gse_family_id(seq_id)
        family_fraction = stable_fraction("gse246381-family:" + family_id)
        edits = derive_edit_operations(source, candidate)
        if not edits:
            continue
        for endpoint in ("hek", "mouse"):
            if endpoint == "hek":
                if not hek_measurements or family_fraction >= 0.50:
                    continue
                measured_source = hek_measurements.get((seq_id, "REF"))
                measured_candidate = hek_measurements.get((seq_id, "ALT"))
                reported_logfc = hek_reported.get(seq_id)
                role = "test_family"
                context = "HEK293T"
                assay = "GSE246381_HEK_MPRA_combined_UMI"
                batch = "GSE246381_HEK_combined_umi_counts.csv.gz"
                sample_count = hek_sample_count
                ood_dimensions: List[str] = []
            else:
                if not mouse_measurements or family_fraction < 0.50:
                    continue
                measured_source = mouse_measurements.get((seq_id, "REF"))
                measured_candidate = mouse_measurements.get((seq_id, "ALT"))
                reported_logfc = mouse_reported.get(seq_id)
                role = "test_ood"
                context = "mouse_brain_Vglut2_CreON"
                assay = "GSE246381_mouse_Vglut_MPRA_combined_UMI"
                batch = "GSE246381_vglut_combined_umi_counts.csv.gz"
                sample_count = mouse_sample_count
                ood_dimensions = ["species_tail"]
                if len(source) <= 60 or len(source) >= 210:
                    ood_dimensions.append("length_tail")
            if measured_source is None or measured_candidate is None:
                continue
            rid = f"v2gse246381:{endpoint}:{sha256_text(seq_id)[:20]}"
            yield role, {
                "benchmark_version": "nmi_benchmark_v2",
                "record_id": rid,
                "source_id": f"{rid}:source",
                "mother_id": f"{rid}:source",
                "candidate_id": f"{rid}:candidate",
                "source_sequence": source,
                "candidate_sequence": candidate,
                "source_sequence_sha256": sha256_text(source),
                "candidate_sequence_sha256": sha256_text(candidate),
                "edit_list": edits,
                "edit_count": len(edits),
                "measured_source": measured_source,
                "measured_candidate": measured_candidate,
                "measured_delta": measured_candidate - measured_source,
                "scored_delta": normalized_log2_delta(measured_source, measured_candidate),
                "scoring_endpoint": "log2_plus_one_fold_change_candidate_over_source",
                "delta": measured_candidate - measured_source,
                "measured_or_proxy_source_value": measured_source,
                "measured_or_proxy_candidate_value": measured_candidate,
                "cargo": "tdTomato",
                "cargo_id": "tdTomato",
                "protein_family_id": "reporter_family:tdTomato",
                "cell_context": context,
                "assay": assay,
                "assay_type": assay,
                "batch": batch,
                "replicate": sample_count,
                "family_cluster_id": family_id,
                "confidence": "measured",
                "data_source": f"GSE246381:{endpoint}:GEO_processed_counts",
                "data_layer": "C_source_matched_intervention",
                "task_kind": "local_delta",
                "local_delta_eligible": True,
                "label_semantics": "wet_lab_source_matched_reporter_abundance_delta; reported_polysome_total_RNA_logFC_retained",
                "value_qualifier": "mean sample-normalized UMI CPM across deposited combined count matrix; not direct protein abundance; scored_delta is log2 plus-one fold change",
                "task_eligibility": "task_a_active_substitution" if all(
                    item["op"] == "replace"
                    and len(str(item["ref"])) == 1
                    and len(str(item["alt"])) == 1
                    for item in edits
                ) else "local_delta_only_indel_or_complex",
                "reported_logFC": reported_logfc,
                "species": "human" if endpoint == "hek" else "mouse",
                "gene": str(row.get("SYMBOL")) if row.get("SYMBOL") is not None else None,
                "phenotype": str(row.get("Pheno")) if row.get("Pheno") is not None else None,
                "transcript_id": str(row.get("Ensembl_TxID")) if row.get("Ensembl_TxID") is not None else None,
                "ood_dimensions": ood_dimensions,
            }


def iter_mpra_records(
    path: Path, *, role: str, max_records: int, task_kind: str, assay_label: str,
    cargo: str, context: str, batch: str, replicate: Optional[int],
) -> Iterator[Dict[str, object]]:
    """Read a bounded public MPRA absolute library without loading it in RAM."""
    with gzip.open(path, "rt", newline="") as fh:
        reader = csv.reader(fh)
        try:
            header = next(reader)
        except StopIteration:
            return
        try:
            seq_i = header.index("utr")
        except ValueError:
            seq_i = 1
        try:
            value_i = header.index("rl")
        except ValueError:
            value_i = len(header) - 1
        emitted = 0
        for row_i, row in enumerate(reader):
            if emitted >= max_records:
                break
            if len(row) <= max(seq_i, value_i):
                continue
            sequence = normalize_sequence(row[seq_i])
            value = parse_float(row[value_i])
            if not sequence or value is None:
                continue
            rid = f"v2abs:mpra:{path.stem}:{row_i:07d}"
            yield make_absolute_record(
                record_id=rid, source_id=f"{rid}:source", sequence=sequence,
                value=value, cargo=cargo, context=context, assay=assay_label,
                batch=batch, replicate=replicate, family=f"mpra:{cargo}", role=role,
                task_kind=task_kind, data_source=f"sample2019_mpra:{path.name}",
            )
            emitted += 1


def iter_codonbert_family_records(
    path: Path, *, data_root: Path, max_records: Optional[int] = None,
) -> Iterator[Tuple[str, Dict[str, object]]]:
    """Emit absolute CodonBERT records joined to exact CDS protein families."""
    cds_metadata: Dict[str, Dict[str, object]] = {}
    for stem in ("gencode_family", "refseq_family"):
        records_path = data_root / "data/reconstructed/p0_data_reconstruction_v1/combined" / f"{stem}.records.jsonl"
        metadata_path = data_root / "data/reconstructed/p0_data_reconstruction_v1/combined" / f"{stem}.metadata.jsonl"
        if not records_path.exists() or not metadata_path.exists():
            continue
        with records_path.open() as records_fh, metadata_path.open() as metadata_fh:
            for record_line, metadata_line in zip(records_fh, metadata_fh):
                record = json.loads(record_line)
                metadata = json.loads(metadata_line)
                cds = normalize_sequence(record.get("cds"))
                protein_sha = str(metadata.get("protein_sha256") or "")
                if cds and protein_sha:
                    cds_metadata.setdefault(sha256_text(cds), {
                        "protein_family_id": f"protein_family:{protein_sha}",
                        "protein_sha256": protein_sha,
                        "species": record.get("species"),
                        "transcript_id": record.get("transcript_id"),
                        "source": stem,
                    })
    emitted = 0
    with path.open(newline="") as fh:
        reader = csv.DictReader(fh)
        for row_i, row in enumerate(reader):
            if max_records is not None and emitted >= max_records:
                break
            sequence = normalize_sequence(row.get("Sequence"))
            value = parse_float(row.get("Value"))
            metadata = cds_metadata.get(sha256_text(sequence)) if sequence else None
            if not sequence or value is None or metadata is None:
                continue
            family_id = str(metadata["protein_family_id"])
            family_fraction = stable_fraction("codonbert-family:" + family_id)
            if family_fraction < 0.20:
                role = "test_family"
                task_kind = "absolute_property_family_shift"
            else:
                source_fraction = stable_fraction("codonbert-source:" + sha256_text(sequence))
                role = "val" if source_fraction < 0.20 else "train"
                task_kind = "absolute_property_regression"
            record_id = f"v2abs:codonbert:{row_i:07d}:{sha256_text(sequence)[:12]}"
            record = make_absolute_record(
                record_id=record_id, source_id=f"{record_id}:source", sequence=sequence,
                value=value, cargo=family_id, context=str(metadata.get("species") or "not_reported"),
                assay="CodonBERT_mRNA_Stability", batch=path.name, replicate=None,
                family=family_id, role=role, task_kind=task_kind,
                data_source=f"codonbert_stability:{path.name}:row_{row_i}",
            )
            record.update({
                "protein_family_id": family_id,
                "protein_sha256": metadata["protein_sha256"],
                "species": metadata.get("species"),
                "matched_transcript_id": metadata.get("transcript_id"),
                "family_join": "exact_cds_sha256_to_p0_gencode_or_refseq_metadata",
                "label_semantics": "absolute_property_protein_family_holdout_not_local_delta_ground_truth" if role == "test_family" else "absolute_property_only",
            })
            yield role, record
            emitted += 1


def iter_mpra_assay_pairs(
    source_path: Path, candidate_path: Path, *, role: str, max_pairs: int,
) -> Iterator[Dict[str, object]]:
    """Pair identical 50mers across chemistry conditions for assay-shift testing."""
    def values(path: Path) -> Dict[str, float]:
        result: Dict[str, float] = {}
        with gzip.open(path, "rt", newline="") as fh:
            reader = csv.reader(fh)
            header = next(reader)
            seq_i, value_i = header.index("utr"), header.index("rl")
            for row in reader:
                if len(row) <= max(seq_i, value_i):
                    continue
                sequence, value = normalize_sequence(row[seq_i]), parse_float(row[value_i])
                if sequence and value is not None:
                    result.setdefault(sequence, value)
        return result

    candidate_values = values(candidate_path)
    emitted = 0
    with gzip.open(source_path, "rt", newline="") as fh:
        reader = csv.reader(fh)
        header = next(reader)
        seq_i, value_i = header.index("utr"), header.index("rl")
        for row_i, row in enumerate(reader):
            if emitted >= max_pairs:
                break
            if len(row) <= max(seq_i, value_i):
                continue
            sequence, source_value = normalize_sequence(row[seq_i]), parse_float(row[value_i])
            candidate_value = candidate_values.get(sequence)
            if not sequence or source_value is None or candidate_value is None:
                continue
            rid = f"v2axis:assay:{source_path.stem}:{candidate_path.stem}:{row_i:07d}"
            yield make_source_matched_axis_record(
                record_id=rid, source_id=f"{rid}:source", source_sequence=sequence,
                candidate_sequence=sequence, measured_source=source_value,
                measured_candidate=candidate_value, cargo="eGFP", context="HEK293T",
                assay="MPRA_1methylpseudouridine_vs_unmodified",
                batch=f"{source_path.name}|{candidate_path.name}", family="mpra:eGFP",
                role=role, task_kind="assay_delta",
                data_source=f"sample2019_mpra:{source_path.name}+{candidate_path.name}",
            )
            emitted += 1


def iter_cao_records(path: Path, *, role: str, max_records: int) -> Iterator[Dict[str, object]]:
    emitted = 0
    context = "HEK293T" if path.name.startswith("hek_") else ("PC3" if path.name.startswith("pc3_") else "Muscle")
    task_kind = "absolute_property_context_shift" if context != "HEK293T" else "absolute_property_regression"
    for row_i, (header, sequence) in enumerate(iter_fasta(path)):
        if emitted >= max_records:
            break
        value = parse_float(header_field(header, "te"))
        if not sequence or value is None:
            continue
        accession = header.split("|", 1)[0]
        rid = f"v2abs:cao:{path.stem}:{row_i:06d}"
        yield make_absolute_record(
            record_id=rid, source_id=f"{rid}:source", sequence=sequence,
            value=value, cargo="endogenous_transcript", context=context,
            assay="Cao2021_TE", batch=path.name, replicate=None,
            family=f"cao:{accession}", role=role, task_kind=task_kind,
            data_source=f"cao2021_5utr:{path.name}",
        )
        emitted += 1


def iter_cao_context_pairs(
    hek_paths: Sequence[Path], context_path: Path, *, role: str, max_pairs: int,
) -> Iterator[Dict[str, object]]:
    """Pair matching transcript accessions across HEK293T and another context."""
    hek: Dict[str, Tuple[str, float, str]] = {}
    for path in hek_paths:
        for header, sequence in iter_fasta(path):
            value = parse_float(header_field(header, "te"))
            accession = header.split("|", 1)[0]
            if accession and sequence and value is not None:
                hek.setdefault(accession, (sequence, value, path.name))
    context = "PC3" if context_path.name.startswith("pc3_") else "Muscle"
    emitted = 0
    for row_i, (header, sequence) in enumerate(iter_fasta(context_path)):
        if emitted >= max_pairs:
            break
        value = parse_float(header_field(header, "te"))
        accession = header.split("|", 1)[0]
        source = hek.get(accession)
        if source is None or not sequence or value is None:
            continue
        source_sequence, source_value, source_batch = source
        rid = f"v2axis:context:{context_path.stem}:{row_i:06d}"
        yield make_source_matched_axis_record(
            record_id=rid, source_id=f"{rid}:source", source_sequence=source_sequence,
            candidate_sequence=sequence, measured_source=source_value,
            measured_candidate=value, cargo="endogenous_transcript", context=context,
            assay="Cao2021_TE_context_shift", batch=f"{source_batch}|{context_path.name}",
            family=f"cao:{accession}", role=role, task_kind="context_delta",
            data_source=f"cao2021_5utr:{source_batch}+{context_path.name}",
        )
        emitted += 1


def asset_entry(data_root: Path, relative: str, *, level: str, name: str,
                label_semantics: str, provenance: str, status: str = "available") -> Dict[str, object]:
    path = data_root / relative
    obj = {
        "name": name,
        "level": level,
        "relative_path": relative,
        "label_semantics": label_semantics,
        "provenance": provenance,
        "status": status if path.exists() else "not_present",
    }
    if path.exists() and path.is_file():
        obj["byte_size"] = path.stat().st_size
        obj["sha256"] = sha256_file(path)
    return obj


def build_asset_registry(data_root: Path) -> Dict[str, List[Dict[str, object]]]:
    """Register only existing assets, while making missing structure explicit."""
    assets = {
        "A_observational_pretraining": [
            asset_entry(data_root, "data/raw/gencode.v45.pc_transcripts.fa.gz", level="A", name="GENCODE v45 protein-coding transcripts", label_semantics="representation/observational_only", provenance="GENCODE v45 source asset"),
            asset_entry(data_root, "data/raw/gencode_mouse/gencode.vM36.pc_transcripts.fa.gz", level="A", name="GENCODE mouse vM36 protein-coding transcripts", label_semantics="representation/observational_only", provenance="GENCODE mouse vM36 source asset"),
            asset_entry(data_root, "data/raw/gencode_mouse/gencode_vM36_5utr.jsonl", level="A", name="GENCODE mouse vM36 species 5UTRs", label_semantics="representation/observational_only", provenance="Derived from GENCODE vM36 FASTA UTR5 header coordinates"),
            asset_entry(data_root, "data/reconstructed/p0_data_reconstruction_v1/sources/gencode_v45/canonical.records.jsonl", level="A", name="GENCODE canonical transcript records", label_semantics="representation/observational_only", provenance="P0 canonical reconstruction"),
            asset_entry(data_root, "data/raw/cao2021_5utr/final_endogenous_5utr.txt", level="A", name="endogenous 5UTR abundance/TE source table", label_semantics="representation/observational_only", provenance="Cao et al. 2021 source asset"),
            asset_entry(data_root, "data/raw/saluki_halflife/rna_hl_human.npz", level="A", name="human RNA half-life arrays", label_semantics="representation/auxiliary_only", provenance="Saluki RNA half-life asset"),
            asset_entry(data_root, "data/raw/saluki_halflife/rna_hl_mouse.npz", level="A", name="mouse RNA half-life arrays", label_semantics="representation/auxiliary_only", provenance="Saluki RNA half-life asset"),
            asset_entry(data_root, "data/raw/rna_structure/gencode_v45_5utr_structure.jsonl", level="A", name="GENCODE v45 5UTR RNA structure features", label_semantics="representation/auxiliary_only", provenance="ViennaRNA-derived observational feature asset from GENCODE canonical records"),
        ],
        "B_absolute_design_libraries": [
            asset_entry(data_root, "data/raw/sample2019_mpra/GSM3130435_egfp_unmod_1.csv.gz", level="B", name="Sample 2019 random 50mer absolute MPRA", label_semantics="absolute_property_only", provenance="Sample et al. 2019, GSE114002"),
            asset_entry(data_root, "data/raw/sample2019_mpra/GSM3130439_egfp_m1pseudo_1.csv.gz", level="B", name="Sample 2019 modified-RNA absolute MPRA", label_semantics="absolute_property_only", provenance="Sample et al. 2019, GSE114002"),
            asset_entry(data_root, "data/raw/sample2019_mpra/GSM3130441_mcherry_1.csv.gz", level="B", name="Sample 2019 mCherry absolute MPRA replicate 1", label_semantics="absolute_property_only; cargo-family holdout", provenance="Sample et al. 2019, GSE114002"),
            asset_entry(data_root, "data/raw/sample2019_mpra/GSM3130442_mcherry_2.csv.gz", level="B", name="Sample 2019 mCherry absolute MPRA replicate 2", label_semantics="absolute_property_only; cargo-family holdout", provenance="Sample et al. 2019, GSE114002"),
            asset_entry(data_root, "data/raw/sample2019_mpra/GSM4084997_varying_length_25to100.csv.gz", level="B", name="Sample 2019 varying-length absolute MPRA", label_semantics="absolute_property_only; length-shift auxiliary", provenance="Sample et al. 2019 varying-length MPRA asset"),
            asset_entry(data_root, "data/raw/cao2021_5utr/hek_top1000_high_TE.fasta", level="B", name="Cao HEK293T high-TE library", label_semantics="absolute_property_only", provenance="Cao et al. 2021 source asset"),
            asset_entry(data_root, "data/raw/cao2021_5utr/hek_bottom500_low_TE.fasta", level="B", name="Cao HEK293T low-TE library", label_semantics="absolute_property_only", provenance="Cao et al. 2021 source asset"),
            asset_entry(data_root, "data/raw/codonbert_stability/mRNA_Stability.csv", level="B", name="CodonBERT mRNA stability library", label_semantics="absolute_property_only", provenance="CodonBERT stability asset"),
            asset_entry(data_root, "data/raw/codonbert_stability/CoV_Vaccine_Degradation.csv", level="B", name="full-length/CDS-containing vaccine mRNA stability library", label_semantics="absolute_property_only", provenance="CodonBERT stability asset"),
        ],
        "C_source_matched_intervention": [
            {"name": "P3 measured source-matched intervention tier", "level": "C", "relative_path": "data/p3/benchmark/measured_tier.jsonl", "label_semantics": "wet_lab_local_delta_ground_truth", "provenance": "P3 source-matched measured tier", "status": "available"},
            {"name": "Sample 2019 chemistry-matched assay pairs", "level": "C", "relative_path": "data/nmi_benchmark_v2/records.jsonl", "label_semantics": "source_matched_assay_delta_not_sequence_edit", "provenance": "paired unmodified and 1-methylpseudouridine MPRA records", "status": "generated_by_builder"},
            {"name": "Cao context-matched pairs", "level": "C", "relative_path": "data/nmi_benchmark_v2/records.jsonl", "label_semantics": "source_matched_context_delta_not_sequence_edit", "provenance": "paired HEK293T and PC3/Muscle TE records by transcript accession", "status": "generated_by_builder"},
            asset_entry(data_root, "data/raw/gse246381_utr_mutation/media-1.xlsx", level="C", name="GSE246381 exact paired UTR sequences", label_semantics="source_matched_intervention_sequence_contract", provenance="GSE246381 Supplement 1; GEO GSE246381"),
            asset_entry(data_root, "data/raw/gse246381_utr_mutation/media-6.xlsx", level="C", name="GSE246381 HEK allelic-effect labels", label_semantics="source_matched_reporter_effect", provenance="GSE246381 Supplement 2; GEO GSE246381"),
            asset_entry(data_root, "data/raw/gse246381_utr_mutation/media-3.xlsx", level="C", name="GSE246381 mouse Vglut allelic-effect labels", label_semantics="source_matched_reporter_effect", provenance="GSE246381 Supplement 5; GEO GSE246381"),
            asset_entry(data_root, "data/raw/gse246381_utr_mutation/GSE246381_hek_combined_umi_counts.csv.gz", level="C", name="GSE246381 HEK combined UMI counts", label_semantics="measured_source_candidate_reporter_abundance", provenance="GEO GSE246381 processed supplementary counts"),
            asset_entry(data_root, "data/raw/gse246381_utr_mutation/GSE246381_vglut_combined_umi_counts.csv.gz", level="C", name="GSE246381 mouse Vglut combined UMI counts", label_semantics="measured_source_candidate_reporter_abundance", provenance="GEO GSE246381 processed supplementary counts"),
        ],
        "D_prospective_data": [
            {"name": "prospective post-freeze intake", "level": "D", "relative_path": "data/nmi_benchmark_v2/manifests/prospective.json", "label_semantics": "future_only_not_available_during_model_development", "provenance": "P1-01 freeze gate", "status": "empty_until_freeze", "frozen": False},
        ],
    }
    return assets


def build(input_paths: Iterable[Path], out_dir: Path, *, data_root: Path,
          mpra_train_limit: int = 1000, mpra_assay_limit: int = 1000,
          cao_train_limit: int = 1000, cao_context_limit: int = 1000) -> Dict:
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "manifests").mkdir(exist_ok=True)
    (out_dir / "indices").mkdir(exist_ok=True)
    records_path = out_dir / "records.jsonl"
    index_fhs = {role: (out_dir / "indices" / f"{role}.txt").open("w") for role in ROLES}
    counts = Counter()
    confidence = Counter()
    task_kinds = Counter()
    source_role: Dict[str, str] = {}
    family_by_role: Dict[str, set] = defaultdict(set)
    sources_by_role: Dict[str, set] = defaultdict(set)
    contexts_by_role: Dict[str, set] = defaultdict(set)
    assays_by_role: Dict[str, set] = defaultdict(set)
    legacy_excluded = Counter()
    historical_measured_source_hashes: set[str] = set()
    historical_measured_candidate_hashes: set[str] = set()

    def write_record(out, rec: Dict[str, object], role: str) -> None:
        rid = str(rec.get("record_id", ""))
        if not rid:
            raise ValueError("record missing record_id")
        if role not in ROLES:
            raise ValueError(f"invalid role {role!r}")
        rec["split_role"] = role
        rec["v2_source_role"] = role
        rec["label_visibility"] = "hidden_before_freeze" if role in FINAL_ROLES else "development_allowed"
        for field in REQUIRED_SOURCE_MATCHED_FIELDS:
            if field not in rec:
                raise ValueError(f"record {rid} missing required field {field}")
        out.write(json.dumps(rec, sort_keys=True, separators=(",", ":")) + "\n")
        index_fhs[role].write(rid + "\n")
        counts[role] += 1
        confidence[(role, str(rec.get("confidence")))] += 1
        task_kinds[(role, str(rec.get("task_kind")))] += 1
        sid = str(rec.get("source_id"))
        previous = source_role.get(sid)
        if previous is not None and previous != role:
            raise ValueError(f"source {sid} crosses roles: {previous} vs {role}")
        source_role[sid] = role
        sources_by_role[role].add(sid)
        family_by_role[role].add(str(rec.get("family_cluster_id")))
        contexts_by_role[role].add(str(rec.get("cell_context")))
        assays_by_role[role].add(str(rec.get("assay")))

    try:
        with records_path.open("w") as out:
            for input_path in input_paths:
                with input_path.open() as fh:
                    for line_no, line in enumerate(fh, 1):
                        if not line.strip():
                            continue
                        rec = json.loads(line)
                        old_role = str(rec.get("split_role", ""))
                        sid = str(rec.get("source_id", ""))
                        if old_role in {"test", "ood"}:
                            legacy_excluded[old_role] += 1
                            continue
                        role = assign_local_role(old_role, sid)
                        normalized = normalize_p3_record(rec, role)
                        write_record(out, normalized, role)
                        if str(rec.get("confidence")) == "measured":
                            historical_measured_source_hashes.add(str(normalized["source_sequence_sha256"]))
                            historical_measured_candidate_hashes.add(str(normalized["candidate_sequence_sha256"]))

            raw_library = data_root / "data/raw/sample2019_mpra/GSM3130443_designed_library.csv.gz"
            if raw_library.exists():
                for role, rec in iter_raw_untouched_records(
                    raw_library,
                    excluded_source_hashes=historical_measured_source_hashes,
                    excluded_candidate_hashes=historical_measured_candidate_hashes,
                ):
                    write_record(out, rec, role)

            # GSE246381 supplies exact paired sequences and independent
            # HEK/mouse reporter-abundance interventions. Its family split is
            # intentionally disjoint between test_family and species/length OOD.
            for role, rec in iter_gse246381_records(data_root):
                write_record(out, rec, role)

            # Exact CDS joins to P0 GENCODE/RefSeq protein metadata provide a
            # real protein-family absolute holdout. These records never enter
            # the local-delta metric stream.
            codonbert_path = data_root / "data/raw/codonbert_stability/mRNA_Stability.csv"
            if codonbert_path.exists():
                for role, rec in iter_codonbert_family_records(codonbert_path, data_root=data_root):
                    write_record(out, rec, role)

            # Absolute MPRA: unmodified HEK is development-only; m1pseudo is a
            # held-out assay/chemistry condition. These records never enter the
            # local-delta metric stream.
            mpra_dir = data_root / "data/raw/sample2019_mpra"
            unmodified = mpra_dir / "GSM3130435_egfp_unmod_1.csv.gz"
            if unmodified.exists():
                for i, rec in enumerate(iter_mpra_records(
                    unmodified, role="train", max_records=mpra_train_limit,
                    task_kind="absolute_property_regression", assay_label="MPRA_unmodified",
                    cargo="eGFP", context="HEK293T", batch=unmodified.name,
                    replicate=1,
                )):
                    write_record(out, rec, "train" if i % 5 else "val")
            assay_file = mpra_dir / "GSM3130439_egfp_m1pseudo_1.csv.gz"
            if assay_file.exists():
                for rec in iter_mpra_records(
                    assay_file, role="test_assay", max_records=mpra_assay_limit,
                    task_kind="absolute_property_assay_shift", assay_label="MPRA_1methylpseudouridine",
                    cargo="eGFP", context="HEK293T", batch=assay_file.name,
                    replicate=1,
                ):
                    write_record(out, rec, "test_assay")
            if unmodified.exists() and assay_file.exists():
                for rec in iter_mpra_assay_pairs(
                    mpra_dir / "GSM3130436_egfp_unmod_2.csv.gz", assay_file,
                    role="test_assay", max_pairs=1000,
                ):
                    write_record(out, rec, "test_assay")

            # mCherry is a measured cargo-family holdout relative to the
            # eGFP development libraries. These are absolute-property records
            # only; no sequence overlap is assumed or required.
            for replicate, name in enumerate(("GSM3130441_mcherry_1.csv.gz", "GSM3130442_mcherry_2.csv.gz"), 1):
                path = mpra_dir / name
                if not path.exists():
                    continue
                for rec in iter_mpra_records(
                    path, role="test_family", max_records=1000,
                    task_kind="absolute_property_family_shift", assay_label="MPRA_mCherry",
                    cargo="mCherry", context="HEK293T", batch=path.name,
                    replicate=replicate,
                ):
                    write_record(out, rec, "test_family")

            # Varying-length MPRA supplies a measured length-shift axis, but
            # has no source-matched nucleotide edit and therefore stays
            # absolute-only within test_ood.
            varying_length = mpra_dir / "GSM4084997_varying_length_25to100.csv.gz"
            if varying_length.exists():
                for rec in iter_mpra_records(
                    varying_length, role="test_ood", max_records=1000,
                    task_kind="absolute_property_length_shift", assay_label="MPRA_varying_length",
                    cargo="eGFP", context="HEK293T", batch=varying_length.name,
                    replicate=1,
                ):
                    rec["ood_dimensions"] = ["length_tail"]
                    rec["label_semantics"] = "absolute_property_length_shift_not_local_delta_ground_truth"
                    rec["value_qualifier"] = "wet-lab measured absolute MPRA property; varying-length auxiliary OOD axis"
                    write_record(out, rec, "test_ood")

            cao_dir = data_root / "data/raw/cao2021_5utr"
            for name in ("hek_top1000_high_TE.fasta", "hek_bottom500_low_TE.fasta"):
                path = cao_dir / name
                if not path.exists():
                    continue
                for i, rec in enumerate(iter_cao_records(path, role="train", max_records=cao_train_limit)):
                    write_record(out, rec, "train" if i % 5 else "val")
            for name in ("pc3_top1000_high_TE.fasta", "pc3_bottom500_low_TE.fasta", "muscle_all_5utr.fasta"):
                path = cao_dir / name
                if not path.exists():
                    continue
                for rec in iter_cao_records(path, role="test_context", max_records=cao_context_limit):
                    write_record(out, rec, "test_context")
                for rec in iter_cao_context_pairs(
                    (cao_dir / "hek_top1000_high_TE.fasta", cao_dir / "hek_bottom500_low_TE.fasta"),
                    path, role="test_context", max_pairs=cao_context_limit,
                ):
                    write_record(out, rec, "test_context")
    finally:
        for fh in index_fhs.values():
            fh.close()

    manifest_paths = {}
    for role in ROLES:
        idx = out_dir / "indices" / f"{role}.txt"
        task_role_counts = {k[1]: n for k, n in task_kinds.items() if k[0] == role}
        manifest = {
            "schema_version": "nmi_benchmark_v2",
            "role": role,
            "final_test": role in FINAL_ROLES,
            "records_path": "records.jsonl",
            "index_path": f"indices/{role}.txt",
            "record_count": int(counts[role]),
            "index_sha256": sha256_file(idx),
            "source_count": len(sources_by_role.get(role, set())),
            "task_kind_counts": dict(sorted(task_role_counts.items())),
            "label_policy": "hidden_by_default" if role in FINAL_ROLES else "development_allowed",
            "local_delta_ground_truth": role in {"train", "val", "test_id", "test_family", "test_ood"},
            "absolute_property_records_are_not_local_delta_ground_truth": True,
            "role_scope": {
                "test_context": "source_matched context_delta plus absolute_property_context_shift; no nucleotide edit context holdout",
                "test_assay": "source_matched assay_delta plus absolute_property_assay_shift; no nucleotide edit assay holdout",
                "test_family": "tdTomato GSE246381 source-matched local-delta cargo/protein-family holdout plus mCherry absolute holdout",
            }.get(role, "source_matched_or_distribution_split"),
        }
        p = out_dir / "manifests" / f"{role}.json"
        p.write_text(json.dumps(manifest, indent=2, sort_keys=True) + "\n")
        manifest_paths[role] = str(p.relative_to(out_dir))

    prospective = {
        "schema_version": "nmi_benchmark_v2_prospective_v1",
        "frozen": False,
        "label_access": "unavailable_until_post_development_freeze",
        "record_count": 0,
        "required_source_matched_fields": REQUIRED_SOURCE_MATCHED_FIELDS,
    }
    (out_dir / "manifests" / "prospective.json").write_text(json.dumps(prospective, indent=2, sort_keys=True) + "\n")

    registry = {
        "schema_version": "nmi_benchmark_v2",
        "source_inputs": [str(p) for p in input_paths],
        "additional_raw_inputs": [
            "data/raw/sample2019_mpra/GSM3130443_designed_library.csv.gz",
            "data/raw/sample2019_mpra/GSM3130441_mcherry_1.csv.gz",
            "data/raw/sample2019_mpra/GSM3130442_mcherry_2.csv.gz",
            "data/raw/sample2019_mpra/GSM4084997_varying_length_25to100.csv.gz",
            "data/raw/codonbert_stability/mRNA_Stability.csv",
            "data/reconstructed/p0_data_reconstruction_v1/combined/gencode_family.records.jsonl",
            "data/reconstructed/p0_data_reconstruction_v1/combined/gencode_family.metadata.jsonl",
            "data/reconstructed/p0_data_reconstruction_v1/combined/refseq_family.records.jsonl",
            "data/reconstructed/p0_data_reconstruction_v1/combined/refseq_family.metadata.jsonl",
            "data/raw/gse246381_utr_mutation/media-1.xlsx",
            "data/raw/gse246381_utr_mutation/media-3.xlsx",
            "data/raw/gse246381_utr_mutation/media-6.xlsx",
            "data/raw/gse246381_utr_mutation/GSE246381_hek_combined_umi_counts.csv.gz",
            "data/raw/gse246381_utr_mutation/GSE246381_vglut_combined_umi_counts.csv.gz",
        ],
        "records_path": str(records_path.relative_to(out_dir)),
        "records_sha256": sha256_file(records_path),
        "total_records": sum(counts.values()),
        "counts_by_role": dict(sorted(counts.items())),
        "confidence_by_role": {f"{r}:{c}": n for (r, c), n in sorted(confidence.items())},
        "task_kinds_by_role": {r: dict(sorted({k[1]: n for k, n in task_kinds.items() if k[0] == r}.items())) for r in ROLES},
        "source_counts": {r: len(v) for r, v in sorted(sources_by_role.items())},
        "family_counts": {r: len(v) for r, v in sorted(family_by_role.items())},
        "contexts_by_role": {r: sorted(v) for r, v in sorted(contexts_by_role.items())},
        "assays_by_role": {r: sorted(v) for r, v in sorted(assays_by_role.items())},
        "required_source_matched_fields": REQUIRED_SOURCE_MATCHED_FIELDS,
        "manifests": manifest_paths,
        "four_layers": build_asset_registry(data_root),
        "prospective_manifest": "manifests/prospective.json",
        "final_test_roles": sorted(FINAL_ROLES),
        "final_test_policy": "loader refuses final roles without explicit allow_final_labels",
        "proxy_is_biological_ground_truth": False,
        "legacy_excluded_split_counts": dict(sorted(legacy_excluded.items())),
        "untouched_test_policy": "historical P3 test/ood/val labels are never promoted to v2 final; final local-edit roles come from raw designed-library rows excluded by historical measured source/candidate membership",
        "family_axis_policy": "test_family contains an independently sourced tdTomato GSE246381 source-matched local-delta family holdout, with mCherry-vs-eGFP absolute holdout and exact CodonBERT protein-family absolute holdout; raw eGFP rows are not assigned to local-delta test_family",
        "protein_family_join_policy": "normalize RNA alphabet, SHA-256 exact CDS match, first provenance record wins for duplicate CDS; family split is stable_fraction(protein_family_id)",
        "ood_dimension_policy": {
            "local_delta_available": ["gc_tail", "motif_uaug", "species_tail", "length_tail"],
            "length": "source-matched GSE246381 length-tail subset is tagged on test_ood; Sample 2019 varying-length remains absolute-only",
            "species": "source-matched mouse GSE246381 Vglut labels are tagged on test_ood; GENCODE mouse 5UTRs remain observational-only",
            "record_field": "ood_dimensions",
        },
        "claim_policy": "absolute-property context/assay records do not authorize local-delta or SOTA claims",
    }
    (out_dir / "registry.json").write_text(json.dumps(registry, indent=2, sort_keys=True) + "\n")
    return registry


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", action="append", required=True, help="source P3 JSONL; repeatable")
    ap.add_argument("--out-dir", default="data/nmi_benchmark_v2")
    ap.add_argument("--data-root", default=".", help="repository root containing data/raw")
    ap.add_argument("--mpra-train-limit", type=int, default=1000)
    ap.add_argument("--mpra-assay-limit", type=int, default=1000)
    ap.add_argument("--cao-train-limit", type=int, default=1000)
    ap.add_argument("--cao-context-limit", type=int, default=1000)
    args = ap.parse_args()
    registry = build(
        [Path(p) for p in args.input], Path(args.out_dir), data_root=Path(args.data_root),
        mpra_train_limit=args.mpra_train_limit, mpra_assay_limit=args.mpra_assay_limit,
        cao_train_limit=args.cao_train_limit, cao_context_limit=args.cao_context_limit,
    )
    print(json.dumps(registry, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
