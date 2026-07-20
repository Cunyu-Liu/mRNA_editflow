#!/usr/bin/env python3
"""P1-05: Acquire Leppek 2022 PERSIST-Seq dataset (GSE173083).

Citation: Leppek K, Byeon GW, Kladwang W, et al. Combinatorial optimization
of mRNA structure, stability, and translation for RNA-based therapeutics.
Nat Commun. 2022;13(1):1536. PMID:35318324. GSE173083.

Note: The P1-05 design doc originally cited GSE151209 (PMID:34873328), but
that accession is actually "Thyroarytenoid muscle gene expression in Pink1
knockout rats" — unrelated. The correct Leppek 2022 PERSIST-Seq accession is
GSE173083 (PMID:35318324, Nat Commun).

PERSIST-Seq provides paired measurements on a pooled library of ~233 mRNAs:
    - In-cell mRNA stability (half-life)
    - Ribosome load (MRL proxy)
    - In-solution stability (degradation rate)
    - Protein output

Also acquires Table S5 (24 CDS designs) as supplementary data.

Output layout:
    data/raw/lepplek2022_persistseq/
        GSE173083_Table_S1.xlsx
        GSE173083_Table_S5.xlsx
        manifest.json
        lepplek2022_summary.json

Manifest fields (per data acquisition hard constraint):
    - source_url
    - sha256
    - license
    - record_count
    - split_stats (train/val/test counts after deterministic hashing)
    - schema (column names from xlsx)
"""
from __future__ import annotations

import hashlib
import json
import os
import ssl
import sys
import time
import urllib.request
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

DATA_ROOT = Path(_REPO_ROOT) / "data" / "raw" / "lepplek2022_persistseq"
DATA_ROOT.mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

GEO_ACC = "GSE173083"
PMID = "35318324"
CITATION = (
    "Leppek K, Byeon GW, Kladwang W, Wayment-Steele HK, Kerr CH, Xu AF, "
    "Kim DS, Topkar VV, Choe C, Rothschild D, Tiu GC, Wellington-Oguri R, "
    "Fujii K, Sharma E, Watkins AM, Nicol JJ, Romano J, Tunguz B, Diaz F, "
    "Cai H, Guo P, Wu J, Meng F, Shi S, Participants E, Dormitzer PR, "
    "Solorzano A, Barna M, Das R. Combinatorial optimization of mRNA "
    "structure, stability, and translation for RNA-based therapeutics. "
    "Nat Commun. 2022;13(1):1536. PMID:35318324. GSE173083."
)
LICENSE = (
    "NCBI GEO public data. Submitted to GEO per NIH data sharing policy. "
    "Cite: Leppek et al. 2022 Nat Commun (PMID:35318324). "
    "Use permitted for research purposes; redistribution subject to GEO terms."
)

# Supplementary files to download
FILES: List[Dict[str, str]] = [
    {
        "filename": "GSE173083_Table_S1.xlsx",
        "url": (
            "https://ftp.ncbi.nlm.nih.gov/geo/series/GSE173nnn/GSE173083/suppl/"
            "GSE173083_Table_S1_-_Attributes_for_pooled_233_sequences.xlsx"
        ),
        "description": "Attributes for pooled 233 sequences (PERSIST-seq main table)",
        "kind": "persist_seq_main",
    },
    {
        "filename": "GSE173083_Table_S5.xlsx",
        "url": (
            "https://ftp.ncbi.nlm.nih.gov/geo/series/GSE173nnn/GSE173083/suppl/"
            "GSE173083_Table_S5_-_Attributes_for_24_CDS_designs.xlsx"
        ),
        "description": "Attributes for 24 CDS designs",
        "kind": "cds_designs",
    },
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _http_get(url: str, dest: Path, timeout: int = 120, retries: int = 5) -> None:
    """Download a URL to dest with retries."""
    for attempt in range(1, retries + 1):
        try:
            req = urllib.request.Request(
                url, headers={"User-Agent": "mrna_editflow/1.0"}
            )
            ctx = ssl.create_default_context()
            ctx.check_hostname = False
            ctx.verify_mode = ssl.CERT_NONE
            with urllib.request.urlopen(req, timeout=timeout, context=ctx) as resp:
                if resp.status != 200:
                    raise RuntimeError(f"HTTP {resp.status} for {url}")
                tmp = dest.with_suffix(dest.suffix + ".part")
                with open(tmp, "wb") as f:
                    while True:
                        chunk = resp.read(1 << 16)
                        if not chunk:
                            break
                        f.write(chunk)
                tmp.replace(dest)
            return
        except Exception as e:
            if attempt == retries:
                raise
            wait = 5 * attempt
            print(f"  attempt {attempt} failed: {e}; retrying in {wait}s",
                  file=sys.stderr)
            time.sleep(wait)


def _sha256(path: Path, buf_size: int = 1 << 20) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        while True:
            chunk = f.read(buf_size)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def _split_bucket(seq: str) -> str:
    """Deterministic train/val/test bucket via SHA-256 of sequence.

    80% train / 10% val / 10% test (matches P1-02A/P1-03 convention).
    """
    h = hashlib.sha256(seq.upper().encode()).hexdigest()
    v = int(h[:8], 16) / 0xFFFFFFFF
    if v < 0.8:
        return "train"
    elif v < 0.9:
        return "val"
    return "test"


def _parse_xlsx(path: Path) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Parse an xlsx file into (column_names, list_of_row_dicts).

    Requires openpyxl. Falls back to pandas if available.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        try:
            import pandas as pd  # type: ignore
            df = pd.read_excel(path)
            cols = list(df.columns)
            rows = df.to_dict(orient="records")
            return cols, rows
        except ImportError as e:
            raise RuntimeError(
                "Neither openpyxl nor pandas is installed. "
                "Run: pip install openpyxl"
            ) from e

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows_iter))
    except StopIteration:
        wb.close()
        return [], []
    cols = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header)]
    rows: List[Dict[str, Any]] = []
    for row in rows_iter:
        if row is None or all(v is None for v in row):
            continue
        rows.append({cols[i]: row[i] for i in range(len(cols))})
    wb.close()
    return cols, rows


# ---------------------------------------------------------------------------
# Sequence / label extraction from Table S1
# ---------------------------------------------------------------------------

# Candidate column name patterns for each field (ordered by priority)
# Note: "Sequence ID" is the identifier, NOT the sequence itself.
# Prefer "RNA sequence" (full mRNA) and "Sequence 3 UTR" (3'UTR).
_SEQ_PATTERNS = [
    "rna sequence", "rna_seq", "full mrna", "mrna sequence",
    "sequence 3 utr", "3utr sequence", "three_prime_utr",
    "sequence cds", "cds sequence",
    "sequence 5 utr", "5utr sequence", "five_prime_utr",
]
_MRL_PATTERNS = ["ribosome load", "mrl", "rl", "translation"]
_HALFLIFE_PATTERNS = ["in-cell half-life", "half_life", "half-life", "halflife", "in_cell hl"]
_PROTEIN_PATTERNS = ["predicted expression", "protein", "expression", "output", "egfp", "gfp"]
_INCELL_STAB_PATTERNS = ["in-cell degradation", "in_cell degradation", "incell degradation"]
_INSOLN_STAB_PATTERNS = ["in-solution degradation", "in_soln degradation", "insoln degradation"]


def _find_column(cols: List[str], patterns: List[str]) -> Optional[str]:
    """Find the first column matching any of the patterns (case-insensitive)."""
    cols_lower = {c.lower(): c for c in cols}
    for pat in patterns:
        for cl, c in cols_lower.items():
            if pat in cl:
                return c
    return None


def _extract_sequence(row: Dict[str, Any], col: Optional[str]) -> str:
    if col is None:
        return ""
    v = row.get(col, "")
    if v is None:
        return ""
    s = str(v).strip().upper().replace("T", "U")
    # Filter to ACGU characters only
    return "".join(c for c in s if c in "ACGU")


def _extract_float(row: Dict[str, Any], col: Optional[str]) -> Optional[float]:
    if col is None:
        return None
    v = row.get(col, None)
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def extract_records(
    cols: List[str], rows: List[Dict[str, Any]]
) -> Tuple[List[Dict[str, Any]], Dict[str, str]]:
    """Extract records (sequence + labels) from parsed Table S1 rows.

    Returns (records, column_mapping).
    """
    seq_col = _find_column(cols, _SEQ_PATTERNS)
    mrl_col = _find_column(cols, _MRL_PATTERNS)
    hl_col = _find_column(cols, _HALFLIFE_PATTERNS)
    protein_col = _find_column(cols, _PROTEIN_PATTERNS)
    incell_col = _find_column(cols, _INCELL_STAB_PATTERNS)
    insoln_col = _find_column(cols, _INSOLN_STAB_PATTERNS)

    mapping = {
        "sequence": seq_col or "",
        "mrl": mrl_col or "",
        "half_life": hl_col or "",
        "protein_output": protein_col or "",
        "in_cell_stability": incell_col or "",
        "in_solution_stability": insoln_col or "",
    }
    print(f"  column mapping: {mapping}")

    records: List[Dict[str, Any]] = []
    skipped = 0
    for row in rows:
        seq = _extract_sequence(row, seq_col)
        if len(seq) < 20:  # need at least 20 nt to be useful
            skipped += 1
            continue
        rec: Dict[str, Any] = {
            "sequence": seq,
            "split": _split_bucket(seq),
            "dataset": "lepplek2022_persistseq",
            "cell_type": "HEK293T",  # PERSIST-seq main library
            "chemistry": "unmodified",  # may be pseudouridine for some
            "source_gse": "GSE173083",
            "source_table": "S1",
        }
        # Attach labels (any that are present)
        for label_name, col_name in [
            ("mrl", mrl_col),
            ("half_life", hl_col),
            ("protein_output", protein_col),
            ("in_cell_stability", incell_col),
            ("in_solution_stability", insoln_col),
        ]:
            val = _extract_float(row, col_name)
            if val is not None:
                rec[label_name] = val
        records.append(rec)

    print(f"  extracted {len(records)} records (skipped {skipped} short/empty)")
    return records, mapping


# ---------------------------------------------------------------------------
# Split computation
# ---------------------------------------------------------------------------

def compute_split_stats(records: List[Dict[str, Any]]) -> Dict[str, int]:
    counts = {"train": 0, "val": 0, "test": 0}
    for r in records:
        counts[r["split"]] += 1
    return counts


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    print(f"data_root = {DATA_ROOT}")
    print(f"GEO accession = {GEO_ACC}")
    print(f"PMID = {PMID}")
    print()

    # 1. Download supplementary files
    print("=== Downloading supplementary files ===")
    manifest_entries: List[Dict] = []
    for finfo in FILES:
        filename = finfo["filename"]
        url = finfo["url"]
        dest = DATA_ROOT / filename
        print(f"\n[{filename}]")
        print(f"  url: {url}")
        if dest.exists() and dest.stat().st_size > 0:
            print(f"  exists ({dest.stat().st_size} bytes), skipping download")
        else:
            try:
                _http_get(url, dest)
                print(f"  downloaded {dest.stat().st_size} bytes")
            except Exception as e:
                print(f"  ERROR: {e}", file=sys.stderr)
                manifest_entries.append({
                    "filename": filename,
                    "source_url": url,
                    "error": str(e),
                })
                continue

        sha = _sha256(dest)
        print(f"  sha256: {sha}")
        manifest_entries.append({
            "filename": filename,
            "source_url": url,
            "local_path": str(dest.relative_to(_REPO_ROOT)),
            "sha256": sha,
            "byte_size": dest.stat().st_size,
            "license": LICENSE,
            "citation": CITATION,
            "description": finfo["description"],
            "kind": finfo["kind"],
        })

    # 2. Parse Table S1 to extract records
    print("\n=== Parsing Table S1 ===")
    s1_path = DATA_ROOT / "GSE173083_Table_S1.xlsx"
    records: List[Dict[str, Any]] = []
    column_mapping: Dict[str, str] = {}
    schema: List[str] = []
    if s1_path.exists():
        try:
            cols, rows = _parse_xlsx(s1_path)
            schema = cols
            print(f"  columns ({len(cols)}): {cols}")
            print(f"  rows: {len(rows)}")
            records, column_mapping = extract_records(cols, rows)
        except Exception as e:
            print(f"  ERROR parsing: {e}", file=sys.stderr)
    else:
        print(f"  Table S1 not found at {s1_path}", file=sys.stderr)

    # 3. Compute split stats
    print("\n=== Computing split stats (deterministic 80/10/10) ===")
    split_stats = compute_split_stats(records)
    print(f"  {split_stats}")

    # 4. Write manifest
    print("\n=== Writing manifest ===")
    manifest = {
        "dataset_name": "lepplek2022_persistseq",
        "geo_accession": GEO_ACC,
        "pubmed_id": PMID,
        "citation": CITATION,
        "license": LICENSE,
        "acquisition_date_utc": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "n_files": len(FILES),
        "files": manifest_entries,
        "schema_s1": schema,
        "column_mapping": column_mapping,
        "record_count": len(records),
        "split_stats": split_stats,
        "labels_available": sorted({
            k for r in records for k in r.keys()
            if k in ("mrl", "half_life", "protein_output",
                     "in_cell_stability", "in_solution_stability")
        }),
    }
    manifest_path = DATA_ROOT / "manifest.json"
    with open(manifest_path, "w") as f:
        json.dump(manifest, f, indent=2, sort_keys=True, default=str)
    print(f"  manifest: {manifest_path}")

    # 5. Write summary
    summary = {
        "dataset_name": "lepplek2022_persistseq",
        "geo_accession": GEO_ACC,
        "total_records": len(records),
        "total_bytes": sum(e.get("byte_size", 0) for e in manifest_entries if "error" not in e),
        "split_stats": split_stats,
        "labels_available": manifest["labels_available"],
        "column_mapping": column_mapping,
    }
    summary_path = DATA_ROOT / "lepplek2022_summary.json"
    with open(summary_path, "w") as f:
        json.dump(summary, f, indent=2, sort_keys=True, default=str)
    print(f"  summary: {summary_path}")

    print("\n" + json.dumps(summary, indent=2, default=str))
    return 0


if __name__ == "__main__":
    sys.exit(main())
