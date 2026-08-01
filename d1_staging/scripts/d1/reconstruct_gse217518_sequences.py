#!/usr/bin/env python
"""D1-03: Reconstruct UTR sequences for GSE217518 from NCBI RefSeq.

GSE217518 (elife97682) contains 5072 paired UTR variants (Ref→Mut) with
RNA stability measurements, but NO sequences — only HGVS c. notation and
genomic coordinates. This script reconstructs the UTR sequences by:

1. Batch-fetching RefSeq mRNA GenBank records from NCBI E-utilities
2. Parsing CDS coordinates to identify 5'UTR and 3'UTR regions
3. Parsing HGVS c. notation to locate the variant position in the mRNA
4. Applying the variant (SNV/del/ins) to produce the Mut (candidate) sequence
5. Extracting the UTR subsequence (source = WT, candidate = Mut)

Output: JSONL with one record per (variant, cell_line_with_label) pair.

Contract: utr_editflow_contract_v2 (FROZEN)
Task: D1-03
"""

import argparse
import json
import os
import re
import sys
import time
import urllib.request
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl


# ---------------------------------------------------------------------------
# GenBank fetching
# ---------------------------------------------------------------------------

def fetch_genbank_batch(accessions: List[str], batch_size: int = 200) -> Dict[str, str]:
    """Fetch GenBank records from NCBI E-utilities in batches.

    Returns dict of accession -> GenBank flat file text.
    """
    result = {}
    n_batches = (len(accessions) + batch_size - 1) // batch_size
    for bi in range(n_batches):
        batch = accessions[bi * batch_size : (bi + 1) * batch_size]
        ids = ",".join(batch)
        url = (
            f"https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
            f"?db=nuccore&id={ids}&rettype=gb&retmode=text"
        )
        for attempt in range(3):
            try:
                req = urllib.request.Request(url, headers={"User-Agent": "mrna-editflow/1.0"})
                with urllib.request.urlopen(req, timeout=120) as r:
                    data = r.read().decode("utf-8", errors="replace")
                break
            except Exception as e:
                print(f"  batch {bi+1}/{n_batches} attempt {attempt+1} failed: {e}", file=sys.stderr)
                if attempt < 2:
                    time.sleep(5 * (attempt + 1))
                else:
                    print(f"  SKIP batch {bi+1}", file=sys.stderr)
                    data = ""
        # Split multi-record response into individual records
        records = re.split(r"^//\s*$", data, flags=re.MULTILINE)
        for rec in records:
            # Extract accession from LOCUS line
            m = re.match(r"^LOCUS\s+(\S+)", rec.strip())
            if m:
                acc = m.group(1)
                result[acc] = rec.strip()
        print(f"  fetched batch {bi+1}/{n_batches}: {len(records)-1} records", file=sys.stderr)
        if bi < n_batches - 1:
            time.sleep(0.5)  # rate limit: ~2 req/s without API key
    return result


def parse_cds_coordinates(genbank_text: str) -> Optional[Tuple[int, int]]:
    """Extract CDS start/end (1-indexed, inclusive) from GenBank flat file.

    Returns (cds_start, cds_end) or None if no CDS found.
    """
    # Match: "     CDS             join(34..1482,2000..2100,...)" or "     CDS             34..1482"
    # The GenBank flat file uses mRNA coordinates (already spliced)
    m = re.search(r"^\s{5}CDS\s+(.+)", genbank_text, re.MULTILINE)
    if not m:
        return None
    cds_loc = m.group(1).strip()
    # Handle join() for multi-exon CDS
    if cds_loc.startswith("join("):
        # Extract all ranges: join(34..1482,2000..2100)
        ranges = re.findall(r"(\d+)\.\.(\d+)", cds_loc)
        if not ranges:
            return None
        # CDS start = first start, CDS end = last end
        starts = [int(r[0]) for r in ranges]
        ends = [int(r[1]) for r in ranges]
        return (min(starts), max(ends))
    elif cds_loc.startswith("complement("):
        inner = cds_loc[len("complement("):-1]
        ranges = re.findall(r"(\d+)\.\.(\d+)", inner)
        if ranges:
            starts = [int(r[0]) for r in ranges]
            ends = [int(r[1]) for r in ranges]
            return (min(starts), max(ends))
    else:
        # Simple range: 34..1482
        m2 = re.match(r"(\d+)\.\.(\d+)", cds_loc)
        if m2:
            return (int(m2.group(1)), int(m2.group(2)))
        # Single position: 34
        m3 = re.match(r"(\d+)", cds_loc)
        if m3:
            return (int(m3.group(1)), int(m3.group(1)))
    return None


def parse_mrna_sequence(genbank_text: str) -> str:
    """Extract the mRNA sequence from GenBank ORIGIN section."""
    # Sequence is after ORIGIN line, format: "  1 agtttgcggc ..."
    idx = genbank_text.find("ORIGIN")
    if idx < 0:
        return ""
    seq_section = genbank_text[idx + len("ORIGIN"):]
    # Stop at the // record terminator
    end_idx = seq_section.find("//")
    if end_idx >= 0:
        seq_section = seq_section[:end_idx]
    # Remove line numbers and spaces
    seq = re.sub(r"[\d\s]", "", seq_section)
    return seq.upper().replace("U", "T")


# ---------------------------------------------------------------------------
# HGVS c. notation parsing
# ---------------------------------------------------------------------------

def parse_hgvs_c(variant_name: str) -> Optional[dict]:
    """Parse HGVS c. notation from variant_name.

    Returns dict with:
        - position_type: "5utr" / "3utr" / "cds" / "splice" / "unknown"
        - position: int (the main position, relative to CDS)
        - position2: int (for ranges, e.g. del)
        - var_type: "snv" / "del" / "ins" / "other"
        - ref: str (ref allele, if available)
        - alt: str (alt allele, if available)
        - raw: str (original notation)
    """
    # Extract the c. notation: NM_xxx(Gene):c.Notation
    m = re.search(r":c\.([^\s]+)", variant_name)
    if not m:
        return None
    notation = m.group(1)
    # Normalize: replace | with >
    notation = notation.replace("|", ">")

    result = {"raw": notation, "var_type": "other", "position_type": "unknown"}

    # SNV patterns: c.-NG>A, c.*NG>A, c.NG>A
    # 5'UTR: c.-NRef>Alt
    snv_m = re.match(r"-(\d+)([ACGT])>([ACGT])$", notation)
    if snv_m:
        result.update({
            "position_type": "5utr",
            "position": int(snv_m.group(1)),
            "var_type": "snv",
            "ref": snv_m.group(2),
            "alt": snv_m.group(3),
        })
        return result

    # 3'UTR: c.*NRef>Alt
    snv_m = re.match(r"\*(\d+)([ACGT])>([ACGT])$", notation)
    if snv_m:
        result.update({
            "position_type": "3utr",
            "position": int(snv_m.group(1)),
            "var_type": "snv",
            "ref": snv_m.group(2),
            "alt": snv_m.group(3),
        })
        return result

    # CDS SNV: c.NRef>Alt
    snv_m = re.match(r"(\d+)([ACGT])>([ACGT])$", notation)
    if snv_m:
        result.update({
            "position_type": "cds",
            "position": int(snv_m.group(1)),
            "var_type": "snv",
            "ref": snv_m.group(2),
            "alt": snv_m.group(3),
        })
        return result

    # Splice site: c.N-MRef>Alt (e.g., c.1394-1G>T)
    splice_m = re.match(r"(\d+)-(\d+)([ACGT])>([ACGT])$", notation)
    if splice_m:
        result.update({
            "position_type": "splice",
            "position": int(splice_m.group(1)),
            "var_type": "snv",
            "ref": splice_m.group(3),
            "alt": splice_m.group(4),
        })
        return result

    # Deletion in UTR: c.-N_-MdelSeq or c.*N_*MdelSeq
    del_m = re.match(r"(\*|-)(\d+)_(\*|-)(\d+)del([ACGT]*)$", notation)
    if del_m:
        sign1 = del_m.group(1)
        pos1 = int(del_m.group(2))
        sign2 = del_m.group(3)
        pos2 = int(del_m.group(4))
        del_seq = del_m.group(5)
        pt1 = "3utr" if sign1 == "*" else ("5utr" if sign1 == "-" else "cds")
        result.update({
            "position_type": pt1,
            "position": pos1 if sign1 == "*" else -pos1,
            "position2": pos2 if sign2 == "*" else -pos2,
            "var_type": "del",
            "ref": del_seq,
            "alt": "",
        })
        return result

    # Deletion: c.N_MdelSeq
    del_m = re.match(r"(\d+)_(\d+)del([ACGT]*)$", notation)
    if del_m:
        pos1 = int(del_m.group(1))
        pos2 = int(del_m.group(2))
        del_seq = del_m.group(3)
        pt = "cds" if pos1 > 0 else ("3utr" if notation.startswith("*") else "5utr")
        result.update({
            "position_type": pt,
            "position": pos1,
            "position2": pos2,
            "var_type": "del",
            "ref": del_seq,
            "alt": "",
        })
        return result

    # Insertion in UTR: c.*N_*MinsSeq or c.-N_-MinsSeq
    ins_m = re.match(r"(\*|-)(\d+)_(\*|-)(\d+)ins([ACGT]+)$", notation)
    if ins_m:
        sign1 = ins_m.group(1)
        pos1 = int(ins_m.group(2))
        sign2 = ins_m.group(3)
        pos2 = int(ins_m.group(4))
        ins_seq = ins_m.group(5)
        pt1 = "3utr" if sign1 == "*" else "5utr"
        result.update({
            "position_type": pt1,
            "position": pos1 if sign1 == "*" else -pos1,
            "position2": pos2 if sign2 == "*" else -pos2,
            "var_type": "ins",
            "ref": "",
            "alt": ins_seq,
        })
        return result

    return result


def hgvs_to_mrna_pos(position: int, position_type: str, cds_start_1idx: int, cds_end_1idx: int) -> Optional[int]:
    """Convert HGVS c. position to 0-indexed mRNA position.

    Args:
        position: HGVS position (positive for CDS, negative for 5'UTR, positive for 3'UTR with type)
        position_type: "5utr", "3utr", or "cds"
        cds_start_1idx: CDS start (1-indexed, inclusive)
        cds_end_1idx: CDS end (1-indexed, inclusive)

    Returns:
        0-indexed position in the mRNA, or None if out of range.
    """
    if position_type == "5utr":
        # c.-N → N bases before CDS start
        # c.-1 = position cds_start_1idx - 1 (0-indexed: cds_start_1idx - 2)
        # c.-N = position cds_start_1idx - N (0-indexed: cds_start_1idx - N - 1)
        pos_0idx = cds_start_1idx - position - 1
        if pos_0idx < 0:
            return None
        return pos_0idx
    elif position_type == "3utr":
        # c.*N → N bases after CDS end
        # c.*1 = position cds_end_1idx + 1 (0-indexed: cds_end_1idx)
        # c.*N = position cds_end_1idx + N (0-indexed: cds_end_1idx + N - 1)
        pos_0idx = cds_end_1idx + position - 1
        return pos_0idx
    elif position_type == "cds":
        # c.N → N-th base of CDS
        # c.1 = position cds_start_1idx (0-indexed: cds_start_1idx - 1)
        # c.N = position cds_start_1idx + N - 1 (0-indexed: cds_start_1idx + N - 2)
        pos_0idx = cds_start_1idx + position - 2
        return pos_0idx
    return None


# ---------------------------------------------------------------------------
# Variant application
# ---------------------------------------------------------------------------

def apply_variant_to_utr(
    utr_seq: str,
    var_info: dict,
    cds_start_1idx: int,
    cds_end_1idx: int,
) -> Optional[Tuple[str, str, str]]:
    """Apply a variant to a UTR sequence.

    Args:
        utr_seq: The UTR sequence (0-indexed string)
        var_info: Parsed HGVS variant info
        cds_start_1idx: CDS start (1-indexed) in the mRNA
        cds_end_1idx: CDS end (1-indexed) in the mRNA

    Returns:
        (source_seq, candidate_seq, status) or None if variant cannot be applied.
        source_seq = WT UTR, candidate_seq = Mut UTR
    """
    if var_info["position_type"] not in ("5utr", "3utr"):
        return None

    vt = var_info["var_type"]
    pos = var_info["position"]

    if vt == "snv":
        # Single position
        # For 5'UTR: position is negative (e.g., -134)
        # For 3'UTR: position is positive (e.g., 113)
        abs_pos = abs(pos)
        mrna_pos = hgvs_to_mrna_pos(abs_pos, var_info["position_type"], cds_start_1idx, cds_end_1idx)
        if mrna_pos is None:
            return None

        # The UTR sequence starts at a different offset than the mRNA
        # For 5'UTR: utr_seq = mRNA[0 : cds_start_1idx-1], so utr_pos = mrna_pos
        # For 3'UTR: utr_seq = mRNA[cds_end_1idx:], so utr_pos = mrna_pos - cds_end_1idx
        if var_info["position_type"] == "5utr":
            utr_pos = mrna_pos  # 0-indexed within 5'UTR
        else:
            utr_pos = mrna_pos - cds_end_1idx  # 0-indexed within 3'UTR

        if utr_pos < 0 or utr_pos >= len(utr_seq):
            return None

        # Verify ref allele
        ref_base = utr_seq[utr_pos]
        expected_ref = var_info.get("ref", "")
        if expected_ref and ref_base != expected_ref:
            return None  # Ref mismatch

        # Apply SNV
        candidate = utr_seq[:utr_pos] + var_info["alt"] + utr_seq[utr_pos + 1:]
        return (utr_seq, candidate, "ok")

    elif vt == "del":
        # Deletion range
        pos2 = var_info.get("position2")
        if pos2 is None:
            return None
        abs_pos1 = abs(pos)
        abs_pos2 = abs(pos2)
        mrna_pos1 = hgvs_to_mrna_pos(abs_pos1, var_info["position_type"], cds_start_1idx, cds_end_1idx)
        mrna_pos2 = hgvs_to_mrna_pos(abs_pos2, var_info["position_type"], cds_start_1idx, cds_end_1idx)
        if mrna_pos1 is None or mrna_pos2 is None:
            return None

        if var_info["position_type"] == "5utr":
            utr_pos1 = mrna_pos1
            utr_pos2 = mrna_pos2
        else:
            utr_pos1 = mrna_pos1 - cds_end_1idx
            utr_pos2 = mrna_pos2 - cds_end_1idx

        if utr_pos1 < 0 or utr_pos2 >= len(utr_seq) or utr_pos1 > utr_pos2:
            return None

        # Verify deleted sequence if provided
        del_seq_expected = var_info.get("ref", "")
        del_seq_actual = utr_seq[utr_pos1 : utr_pos2 + 1]
        if del_seq_expected and del_seq_actual != del_seq_expected:
            return None

        # Apply deletion
        candidate = utr_seq[:utr_pos1] + utr_seq[utr_pos2 + 1:]
        return (utr_seq, candidate, "ok")

    elif vt == "ins":
        # Insertion between positions
        pos2 = var_info.get("position2")
        if pos2 is None:
            return None
        abs_pos1 = abs(pos)
        abs_pos2 = abs(pos2)
        mrna_pos1 = hgvs_to_mrna_pos(abs_pos1, var_info["position_type"], cds_start_1idx, cds_end_1idx)
        mrna_pos2 = hgvs_to_mrna_pos(abs_pos2, var_info["position_type"], cds_start_1idx, cds_end_1idx)
        if mrna_pos1 is None or mrna_pos2 is None:
            return None

        if var_info["position_type"] == "5utr":
            utr_pos1 = mrna_pos1
            utr_pos2 = mrna_pos2
        else:
            utr_pos1 = mrna_pos1 - cds_end_1idx
            utr_pos2 = mrna_pos2 - cds_end_1idx

        # Insert after utr_pos2 (between pos1 and pos2, which are usually the same or adjacent)
        insert_after = max(utr_pos1, utr_pos2)
        if insert_after < 0 or insert_after >= len(utr_seq):
            return None

        ins_seq = var_info.get("alt", "")
        candidate = utr_seq[:insert_after + 1] + ins_seq + utr_seq[insert_after + 1:]
        return (utr_seq, candidate, "ok")

    return None


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Reconstruct GSE217518 UTR sequences from NCBI RefSeq")
    parser.add_argument("--excel", default="data/raw/elife97682_utr_stability/elife-97682-supp1-v1.xlsx",
                        help="Path to elife97682 Excel file")
    parser.add_argument("--output", default="data/raw/gse217518_utr_stability/reconstructed_variants.jsonl",
                        help="Output JSONL path")
    parser.add_argument("--cache-dir", default="data/raw/gse217518_utr_stability/genbank_cache",
                        help="Directory to cache GenBank records")
    args = parser.parse_args()

    print("=== D1-03: Reconstruct GSE217518 UTR sequences ===")

    # 1. Read Excel
    print(f"\n[1] Reading Excel: {args.excel}")
    wb = openpyxl.load_workbook(args.excel, read_only=True)
    ws = wb["Sup_T2_metaTable"]
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    print(f"  {len(rows)} variants, columns: {header}")

    # 2. Extract unique transcript accessions
    acc_set = set()
    for r in rows:
        vn = r[1] or ""
        m = re.match(r"([A-Z]+_\d+\.\d+)", vn)
        if m:
            acc_set.add(m.group(1))
    print(f"\n[2] Unique transcript accessions: {len(acc_set)}")

    # 3. Fetch GenBank records (with caching)
    cache_dir = Path(args.cache_dir)
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_file = cache_dir / "genbank_batch.txt"

    genbank_records = {}
    if cache_file.exists():
        print(f"\n[3] Loading cached GenBank records from {cache_file}")
        with open(cache_file) as f:
            content = f.read()
        records = re.split(r"^//\s*$", content, flags=re.MULTILINE)
        for rec in records:
            m = re.match(r"^LOCUS\s+(\S+)", rec.strip())
            if m:
                genbank_records[m.group(1)] = rec.strip()
        print(f"  Loaded {len(genbank_records)} cached records")
    else:
        print(f"\n[3] Fetching {len(acc_set)} GenBank records from NCBI...")
        acc_list = sorted(acc_set)
        genbank_records = fetch_genbank_batch(acc_list)
        # Save to cache
        with open(cache_file, "w") as f:
            for acc, rec in genbank_records.items():
                f.write(rec + "\n//\n")
        print(f"  Cached {len(genbank_records)} records to {cache_file}")

    # 4. Parse CDS coordinates and sequences
    print(f"\n[4] Parsing CDS coordinates and sequences")
    transcript_info = {}  # accession -> {cds_start, cds_end, sequence, 5utr, 3utr}
    for acc, gb_text in genbank_records.items():
        cds = parse_cds_coordinates(gb_text)
        seq = parse_mrna_sequence(gb_text)
        if cds and seq:
            cds_start, cds_end = cds
            five_utr = seq[:cds_start - 1]  # 0-indexed
            three_utr = seq[cds_end:]       # 0-indexed
            transcript_info[acc] = {
                "cds_start_1idx": cds_start,
                "cds_end_1idx": cds_end,
                "sequence": seq,
                "5utr": five_utr,
                "3utr": three_utr,
            }
    print(f"  Parsed {len(transcript_info)} transcripts with CDS info")

    # 5. Process variants
    print(f"\n[5] Processing variants")
    output_records = []
    stats = {"ok": 0, "skip_no_transcript": 0, "skip_no_cds": 0,
             "skip_position_out_of_range": 0, "skip_ref_mismatch": 0,
             "skip_unparseable": 0, "skip_cds_variant": 0}

    for r in rows:
        mutant_id = r[0]
        variant_name = r[1] or ""
        gene_symbol = r[2]
        utr_group = r[6]
        ref_allele = r[7]
        alt_allele = r[8]
        t05_wt_sh = r[9]
        t05_mt_sh = r[10]
        pval_sh = r[11]
        t05_wt_hek = r[13]
        t05_mt_hek = r[14]
        pval_hek = r[15]

        # Extract transcript accession
        m = re.match(r"([A-Z]+_\d+\.\d+)", variant_name)
        if not m:
            stats["skip_unparseable"] += 1
            continue
        acc = m.group(1)
        # Some accessions might not have the version in GenBank LOCUS
        # Try both with and without version
        acc_base = acc.split(".")[0]

        if acc not in transcript_info:
            # Try matching by base accession (LOCUS line might not include version)
            matched_acc = None
            for ga in genbank_records:
                if ga == acc or ga.startswith(acc_base):
                    matched_acc = ga
                    break
            if matched_acc is None:
                stats["skip_no_transcript"] += 1
                continue
            acc = matched_acc

        ti = transcript_info.get(acc)
        if ti is None:
            stats["skip_no_cds"] += 1
            continue

        # Parse HGVS notation
        var_info = parse_hgvs_c(variant_name)
        if var_info is None or var_info["position_type"] == "unknown":
            stats["skip_unparseable"] += 1
            continue
        if var_info["position_type"] in ("cds", "splice"):
            stats["skip_cds_variant"] += 1
            continue

        # Get UTR sequence
        utr_type = var_info["position_type"]  # "5utr" or "3utr"
        utr_seq = ti[utr_type]
        if not utr_seq:
            stats["skip_position_out_of_range"] += 1
            continue

        # Apply variant
        result = apply_variant_to_utr(
            utr_seq, var_info,
            ti["cds_start_1idx"], ti["cds_end_1idx"]
        )
        if result is None:
            stats["skip_ref_mismatch"] += 1
            continue

        source_seq, candidate_seq, status = result
        if status != "ok":
            stats["skip_" + status] = stats.get("skip_" + status, 0) + 1
            continue

        # Normalize sequences
        source_seq = source_seq.upper().replace("U", "T")
        candidate_seq = candidate_seq.upper().replace("U", "T")
        # Keep only ACGT
        source_seq = "".join(c for c in source_seq if c in "ACGT")
        candidate_seq = "".join(c for c in candidate_seq if c in "ACGT")

        if not source_seq or not candidate_seq or source_seq == candidate_seq:
            stats["skip_unparseable"] += 1
            continue

        stats["ok"] += 1

        # Build record
        region = "5'UTR" if utr_type == "5utr" else "3'UTR"
        record = {
            "record_id": f"GSE217518_{mutant_id}",
            "variant_name": variant_name,
            "gene_symbol": gene_symbol,
            "region": region,
            "source_sequence": source_seq,
            "candidate_sequence": candidate_seq,
            "variant_type": var_info["var_type"],
            "labels": {},
            "metadata": {
                "source_file": "elife-97682-supp1-v1.xlsx",
                "transcript_accession": acc,
                "cds_start": ti["cds_start_1idx"],
                "cds_end": ti["cds_end_1idx"],
                "utr_length": len(source_seq),
                "variant_position_type": utr_type,
                "variant_position": abs(var_info["position"]),
                "ref_allele_excel": ref_allele,
                "alt_allele_excel": alt_allele,
                "hgvs_raw": var_info["raw"],
            },
        }
        # Add labels (only non-None values)
        if t05_mt_hek is not None:
            record["labels"]["stability_hek"] = float(t05_mt_hek)
        if t05_mt_sh is not None:
            record["labels"]["stability_sh"] = float(t05_mt_sh)
        # Add WT stability to metadata
        if t05_wt_hek is not None:
            record["metadata"]["stability_wt_hek"] = float(t05_wt_hek)
        if t05_wt_sh is not None:
            record["metadata"]["stability_wt_sh"] = float(t05_wt_sh)
        if pval_hek is not None:
            record["metadata"]["pval_hek"] = float(pval_hek)
        if pval_sh is not None:
            record["metadata"]["pval_sh"] = float(pval_sh)

        output_records.append(record)

    # 6. Write output
    print(f"\n[6] Writing {len(output_records)} records to {args.output}")
    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    with open(args.output, "w") as f:
        for rec in output_records:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")

    # Print stats
    print(f"\n=== Statistics ===")
    print(f"  Total variants in Excel: {len(rows)}")
    print(f"  Successfully reconstructed: {stats['ok']}")
    print(f"  Skipped (no transcript): {stats['skip_no_transcript']}")
    print(f"  Skipped (no CDS): {stats['skip_no_cds']}")
    print(f"  Skipped (out of range): {stats['skip_position_out_of_range']}")
    print(f"  Skipped (ref mismatch): {stats['skip_ref_mismatch']}")
    print(f"  Skipped (CDS/splice variant): {stats['skip_cds_variant']}")
    print(f"  Skipped (unparseable): {stats['skip_unparseable']}")
    total_skip = sum(v for k, v in stats.items() if k != "ok")
    print(f"  Total skipped: {total_skip}")
    print(f"  Success rate: {stats['ok']/(stats['ok']+total_skip)*100:.1f}%")

    # Region breakdown
    by_region = {}
    for rec in output_records:
        by_region[rec["region"]] = by_region.get(rec["region"], 0) + 1
    print(f"\n  By region: {by_region}")

    by_type = {}
    for rec in output_records:
        by_type[rec["variant_type"]] = by_type.get(rec["variant_type"], 0) + 1
    print(f"  By variant type: {by_type}")

    print(f"\nDone. Output: {args.output}")


if __name__ == "__main__":
    main()
