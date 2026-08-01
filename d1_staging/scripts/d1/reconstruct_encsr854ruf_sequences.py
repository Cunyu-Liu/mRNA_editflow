#!/usr/bin/env python
"""D1-03: Reconstruct oligo sequences for ENCSR854RUF from hg19.

ENCSR854RUF (MPRAu) contains 15,266 3'UTR variants with MPRA activity data
across 6 cell types but NO actual oligo sequences. The xlsx "Oligo Variant
Info" sheet provides genomic coordinates for each oligo window (~100bp).
This script reconstructs WT (ref) and MUT (alt) oligo sequences by:

1. Reading oligo window coordinates from the xlsx
2. Fetching hg19 reference sequence for each window
3. Applying the variant (ref_allele -> alt_allele) on the + strand
4. Reverse-complementing if strand == '-'
5. Joining with activity labels from "Variant MPRAu Results" sheet

Output: JSONL with one record per mpra_variant_id.

Contract: utr_editflow_contract_v2 (FROZEN)
Task: D1-03
"""

import argparse
import json
import os
import subprocess
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl

# Make genome_fetcher importable
HERE = os.path.dirname(os.path.abspath(__file__))
if HERE not in sys.path:
    sys.path.insert(0, HERE)

from genome_fetcher import (  # noqa: E402
    reverse_complement,
    normalize_seq,
    ensure_fa,
    create_fetcher,
)


def apply_variant_to_plus_strand(
    plus_seq: str,
    var_offset: int,
    ref_allele: str,
    alt_allele: str,
) -> Optional[str]:
    """Apply variant on the + strand sequence.

    Args:
        plus_seq: + strand reference sequence
        var_offset: 0-indexed position of variant start in plus_seq
        ref_allele: reference allele (+ strand)
        alt_allele: alternate allele (+ strand)

    Returns:
        Sequence with alt_allele substituted for ref_allele, or None if
        ref_allele doesn't match.
    """
    ref_len = len(ref_allele)
    actual = plus_seq[var_offset : var_offset + ref_len]
    if actual.upper() != ref_allele.upper():
        return None
    return plus_seq[:var_offset] + alt_allele.upper() + plus_seq[var_offset + ref_len :]


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Reconstruct ENCSR854RUF oligo sequences from hg19"
    )
    parser.add_argument(
        "--excel",
        default="data/p0/ENCSR854RUF/processed/MPRAu_Supplementary_Table1.xlsx",
        help="Path to MPRAu Supplementary Table1 xlsx",
    )
    parser.add_argument(
        "--genome",
        default="ensembl",
        help="'ensembl' for REST API, or path to hg19 genome (.2bit or .fa)",
    )
    parser.add_argument(
        "--twobittofa",
        default="/mnt/cunyuliu/genomes/twoBitToFa",
        help="Path to twoBitToFa binary (for .2bit conversion)",
    )
    parser.add_argument(
        "--output",
        default="data/p0/ENCSR854RUF/reconstructed_oligos.jsonl",
        help="Output JSONL path",
    )
    args = parser.parse_args()

    print("=== D1-03: Reconstruct ENCSR854RUF oligo sequences ===")

    # 1. Create genome fetcher (Ensembl API or local FASTA)
    gf = create_fetcher(args.genome, args.twobittofa)

    # 2. Read xlsx — Oligo Variant Info
    print(f"\n[1] Reading Excel: {args.excel}")
    wb = openpyxl.load_workbook(args.excel, read_only=True)
    ws = wb["Oligo Variant Info"]
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    print(f"  columns: {header}")
    col = {name: i for i, name in enumerate(header)}

    # 3. Read xlsx — Variant MPRAu Results (labels)
    print(f"\n[2] Reading 'Variant MPRAu Results' sheet for labels")
    ws_results = wb["Variant MPRAu Results"]
    header_results = [c.value for c in next(ws_results.iter_rows(max_row=1))]
    print(f"  {len(header_results) - 1} label columns")

    # Build labels dict: mpra_variant_id -> {label_name: value}
    labels_by_vid = {}
    n_results = 0
    for row in ws_results.iter_rows(min_row=2, values_only=True):
        vid = row[0]
        if vid is None:
            continue
        labels = {}
        for i, val in enumerate(row[1:], start=1):
            if val is None:
                continue
            try:
                fv = float(val)
                if fv == fv:  # not NaN
                    labels[header_results[i]] = fv
            except (ValueError, TypeError):
                pass
        labels_by_vid[vid] = labels
        n_results += 1
    print(f"  loaded labels for {n_results} variants")

    # 4. Process oligo variant info — group by mpra_variant_id
    print(f"\n[3] Processing oligo variants")
    # Each variant has 2 rows (ref/alt). We only need the coordinates,
    # which are the same for both. Use the first row per variant.
    variants_by_vid = {}
    n_total_rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        vid = row[col["mpra_variant_id"]]
        if vid is None:
            continue
        n_total_rows += 1
        if vid not in variants_by_vid:
            variants_by_vid[vid] = {
                "mpra_variant_id": vid,
                "variant_id": row[col["variant_id"]],
                "chrom": row[col["chrom"]],
                "oligo_starts": row[col["oligo_starts"]],
                "oligo_ends": row[col["oligo_ends"]],
                "strand": row[col["strand"]],
                "var_start": row[col["var_start"]],
                "var_end": row[col["var_end"]],
                "ref_allele": row[col["ref_allele"]],
                "alt_allele": row[col["alt_allele"]],
                "genes": row[col["genes"]],
                "transcripts": row[col["transcripts"]],
                "gene_symbols": row[col["gene_symbols"]],
            }
    print(f"  {n_total_rows} rows -> {len(variants_by_vid)} unique variants")

    # 4b. Prefetch all oligo window sequences (batch API for Ensembl)
    prefetch_regions = []
    for vinfo in variants_by_vid.values():
        try:
            chrom_val = int(vinfo["chrom"])
            chrom = f"chr{chrom_val}"
            oligo_start = int(str(vinfo["oligo_starts"]).strip())
            oligo_end = int(str(vinfo["oligo_ends"]).strip())
            # 0-indexed half-open for prefetch
            prefetch_regions.append((chrom, oligo_start - 1, oligo_end))
        except (ValueError, TypeError):
            pass
    gf.prefetch(prefetch_regions)

    # 5. Reconstruct sequences
    print(f"\n[4] Reconstructing sequences from hg19")
    output_records = []
    stats = {
        "ok": 0,
        "skip_ref_mismatch": 0,
        "skip_fetch_error": 0,
        "skip_no_labels": 0,
    }

    for vid, vinfo in variants_by_vid.items():
        # Parse coordinates (stored as strings in xlsx)
        try:
            chrom_val = int(vinfo["chrom"])
            chrom = f"chr{chrom_val}"
            oligo_start = int(str(vinfo["oligo_starts"]).strip())
            oligo_end = int(str(vinfo["oligo_ends"]).strip())
            var_start = int(str(vinfo["var_start"]).strip())
            var_end = int(str(vinfo["var_end"]).strip())
        except (ValueError, TypeError) as e:
            print(f"  SKIP {vid}: coordinate parse error: {e}", file=sys.stderr)
            stats["skip_fetch_error"] += 1
            continue

        strand = str(vinfo["strand"]).strip()
        ref_allele = str(vinfo["ref_allele"]).strip().upper()
        alt_allele = str(vinfo["alt_allele"]).strip().upper()

        # Fetch + strand oligo sequence (1-indexed inclusive -> 0-indexed half-open)
        plus_seq = gf.fetch(chrom, oligo_start - 1, oligo_end)
        if not plus_seq or len(plus_seq) < 10:
            stats["skip_fetch_error"] += 1
            continue

        # Variant position within oligo (0-indexed)
        var_offset = var_start - oligo_start
        ref_len = var_end - var_start + 1  # 1-indexed inclusive

        # Verify ref_allele length matches coordinate span
        if ref_len != len(ref_allele):
            # Try alternative: var_end might be exclusive or 0-indexed
            ref_len = len(ref_allele)
            # Adjust offset if needed
            if var_end == var_start:
                ref_len = 1

        # Apply variant on + strand
        mut_plus = apply_variant_to_plus_strand(plus_seq, var_offset, ref_allele, alt_allele)
        if mut_plus is None:
            # Try offset adjustment: maybe oligo_starts is 0-indexed
            var_offset_alt = var_start - oligo_start - 1
            if 0 <= var_offset_alt < len(plus_seq):
                mut_plus = apply_variant_to_plus_strand(plus_seq, var_offset_alt, ref_allele, alt_allele)
                if mut_plus is not None:
                    var_offset = var_offset_alt
        if mut_plus is None:
            stats["skip_ref_mismatch"] += 1
            continue

        # RC if strand == '-'
        if strand == "-":
            wt_seq = reverse_complement(plus_seq)
            mut_seq = reverse_complement(mut_plus)
        else:
            wt_seq = plus_seq
            mut_seq = mut_plus

        wt_seq = normalize_seq(wt_seq)
        mut_seq = normalize_seq(mut_seq)

        if not wt_seq or not mut_seq or wt_seq == mut_seq:
            stats["skip_ref_mismatch"] += 1
            continue

        # Get labels
        labels = labels_by_vid.get(vid, {})
        if not labels:
            stats["skip_no_labels"] += 1
            # Still include the record (sequences are valid)

        stats["ok"] += 1

        record = {
            "record_id": f"ENCSR854RUF_{vid}",
            "mpra_variant_id": vid,
            "variant_id": vinfo["variant_id"],
            "source_sequence": wt_seq,
            "candidate_sequence": mut_seq,
            "region": "3'UTR",
            "variant_type": "snv" if len(ref_allele) == len(alt_allele) == 1 else "indel",
            "labels": labels,
            "metadata": {
                "chrom": chrom,
                "oligo_starts": oligo_start,
                "oligo_ends": oligo_end,
                "strand": strand,
                "var_start": var_start,
                "var_end": var_end,
                "ref_allele": ref_allele,
                "alt_allele": alt_allele,
                "genes": vinfo["genes"],
                "transcripts": vinfo["transcripts"],
                "gene_symbol": vinfo["gene_symbols"],
                "variant_position": var_offset,
                "oligo_length": len(wt_seq),
                "source_file": Path(args.excel).name,
            },
        }
        output_records.append(record)

    # 6. Write output
    print(f"\n[5] Writing {len(output_records)} records to {args.output}")
    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    with open(args.output, "w") as f:
        for rec in output_records:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")

    # Print stats
    print(f"\n=== Statistics ===")
    print(f"  Total variants: {len(variants_by_vid)}")
    print(f"  Successfully reconstructed: {stats['ok']}")
    print(f"  Skipped (ref mismatch): {stats['skip_ref_mismatch']}")
    print(f"  Skipped (fetch error): {stats['skip_fetch_error']}")
    print(f"  Skipped (no labels): {stats['skip_no_labels']}")
    total_skip = sum(v for k, v in stats.items() if k != "ok")
    if stats["ok"] + total_skip > 0:
        print(f"  Success rate: {stats['ok']/(stats['ok']+total_skip)*100:.1f}%")

    by_type = {}
    for rec in output_records:
        by_type[rec["variant_type"]] = by_type.get(rec["variant_type"], 0) + 1
    print(f"  By variant type: {by_type}")

    print(f"\nDone. Output: {args.output}")


if __name__ == "__main__":
    main()
