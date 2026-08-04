#!/usr/bin/env python3
"""Extract only deterministic sequence-label support from GSE176581.

The DE bin workbooks contain gene-level statistics without a bindable UTR
sequence and are therefore recorded as quarantine/support-only.  The natural
sequence supplementary table is emitted as F observations when its sequence
and numeric labels are valid.  synthetic/model tables remain AUX/reference
and are not promoted into E/F.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
from pathlib import Path

import openpyxl


SEQ_RE = re.compile(r"^[ACGTU]+$", re.IGNORECASE)


def text(value) -> str:
    return "" if value is None else str(value).strip()


def seq(value):
    raw = text(value)
    if not raw:
        return "", "EMPTY", []
    steps = ["STRIP"]
    upper = raw.upper()
    if upper != raw:
        steps.append("UPPERCASE")
    if "U" in upper:
        steps.append("U_TO_T")
    if not SEQ_RE.fullmatch(upper):
        return "", "INVALID_SYMBOL", steps
    return upper.replace("U", "T"), "PASS", steps


def num(value):
    try:
        if value is None:
            return None
        out = float(value)
        return out if math.isfinite(out) else None
    except (ValueError, TypeError):
        return None


def sha(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--raw-dir", type=Path, required=True)
    ap.add_argument("--output", type=Path, required=True)
    ap.add_argument("--dispositions", type=Path, required=True)
    args = ap.parse_args()
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.dispositions.parent.mkdir(parents=True, exist_ok=True)
    emitted = 0
    reasons: dict[str, int] = {}

    def reject(out, file_name, row_no, reason, **extra):
        reasons[reason] = reasons.get(reason, 0) + 1
        row = {
            "accession": "GSE176581",
            "source_file": file_name,
            "source_row_locator": f"{file_name}#row={row_no}",
            "disposition": "QUARANTINED",
            "reason": reason,
        }
        row.update(extra)
        out.write(json.dumps(row, sort_keys=True) + "\n")

    with args.output.open("w", encoding="utf-8") as out, args.dispositions.open(
        "w", encoding="utf-8"
    ) as rej:
        for path in sorted(args.raw_dir.glob("*.xlsx")):
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb["Naturally-occuring"] if "Naturally-occuring" in wb.sheetnames else None
            if ws is None:
                # The bin and synthetic tables remain evidence for this
                # dataset but are not sequence-label F rows in this adapter.
                support_ws = wb[wb.sheetnames[0]]
                for row_no, _ in enumerate(support_ws.iter_rows(), 1):
                    reject(rej, path.name, row_no, "SUPPORT_ONLY_NO_BINDABLE_UTR")
                wb.close()
                continue
            rows = ws.iter_rows(values_only=True)
            header = [text(x) for x in next(rows)]
            columns = {name: i for i, name in enumerate(header) if name}
            seq_i = columns.get("Sequence (with flanking regions)")
            te_i = columns.get("TE")
            rpkm_i = columns.get("RNAseq_rpkm")
            cell_i = columns.get("Cell tyes")
            if seq_i is None:
                reject(rej, path.name, 1, "MISSING_SEQUENCE_COLUMN")
                wb.close()
                continue
            for row_no, values in enumerate(rows, 2):
                raw = text(values[seq_i] if seq_i < len(values) else None)
                normalized, status, steps = seq(raw)
                if status != "PASS":
                    reject(rej, path.name, row_no, status, raw_sequence_sha256=sha(raw) if raw else None)
                    continue
                labels = {}
                if te_i is not None and te_i < len(values):
                    val = num(values[te_i])
                    if val is not None:
                        labels["te"] = val
                if rpkm_i is not None and rpkm_i < len(values):
                    val = num(values[rpkm_i])
                    if val is not None:
                        labels["rnaseq_rpkm"] = val
                if not labels:
                    reject(rej, path.name, row_no, "NO_NUMERIC_LABEL")
                    continue
                raw_cell = text(values[cell_i] if cell_i is not None and cell_i < len(values) else None)
                record_id = f"GSE176581_{path.stem}_{row_no}"
                record = {
                    "record_id": record_id,
                    "dataset": "gse176581",
                    "accession": "GSE176581",
                    "region": "5'UTR",
                    "source_sequence": None,
                    "candidate_sequence": normalized,
                    "edit_script": [],
                    "edit_script_verified": True,
                    "edit_distance": 0,
                    "n_ins": 0,
                    "n_del": 0,
                    "n_sub": 0,
                    "path_ambiguity": 1,
                    "labels": labels,
                    "metadata": {
                        "record_type": "observational",
                        "source_file": path.name,
                        "source_row_locator": f"{path.name}#row={row_no}",
                        "raw_candidate_sequence_sha256": sha(raw),
                        "candidate_normalization_steps": steps,
                        "candidate_alphabet_status": status,
                        "raw_context_label": raw_cell or None,
                        "adapter_id": "GSE176581_NATURAL_SEQUENCE_SUPPORT_V1",
                    },
                }
                out.write(json.dumps(record, ensure_ascii=False, sort_keys=True) + "\n")
                emitted += 1
            wb.close()

    summary = {
        "adapter_id": "GSE176581_NATURAL_SEQUENCE_SUPPORT_V1",
        "emitted": emitted,
        "dispositions": reasons,
        "output": str(args.output),
        "dispositions_path": str(args.dispositions),
    }
    args.output.with_suffix(".summary.json").write_text(
        json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
